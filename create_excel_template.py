import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def create_departments_users_template():
    """إنشاء ملف Excel نموذجي للأقسام والمستخدمين بتصميم جميل"""
    
    # بيانات الأقسام النموذجية
    departments_data = {
        'name': [
            'قسم تكنولوجيا المعلومات',
            'قسم الموارد البشرية', 
            'قسم المالية',
            'قسم المبيعات',
            'قسم التسويق'
        ],
        'manager_name': ['أحمد محمود', 'سارة علي', 'محمد عبدالله', 'علي حسن', 'فاطمة أحمد']
    }
    
    # بيانات المستخدمين النموذجية
    users_data = {
        'username': ['ahmed.it', 'sara.hr', 'mohamed.finance', 'ali.sales', 'fatima.marketing', 'layla.it'],
        'name': ['أحمد محمود', 'سارة علي', 'محمد عبدالله', 'علي حسن', 'فاطمة أحمد', 'ليلى كمال'],
        'password': ['123456', '654321', '112233', '445566', '778899', '999888'],
        'is_admin': ['نعم', 'لا', 'لا', 'لا', 'لا', 'لا'],
        'is_manager': ['نعم', 'نعم', 'نعم', 'نعم', 'نعم', 'لا'],
        'department_name': [
            'قسم تكنولوجيا المعلومات',
            'قسم الموارد البشرية',
            'قسم المالية', 
            'قسم المبيعات',
            'قسم التسويق',
            'قسم تكنولوجيا المعلومات'
        ]
    }
    
    # إنشاء ملف Excel جديد
    file_name = f"نموذج_الموظفين_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb = Workbook()
    
    # إزالة الورقة الافتراضية
    wb.remove(wb.active)
    
    # إنشاء ورقة الأقسام
    create_departments_sheet(wb, departments_data)
    
    # إنشاء ورقة المستخدمين
    create_users_sheet(wb, users_data)
    
    # إنشاء ورقة التعليمات
    create_instructions_sheet(wb)
    
    # حفظ الملف
    wb.save(file_name)
    print(f"✓ تم إنشاء الملف: {file_name}")
    print("📍 الموقع:", os.path.abspath(file_name))
    
    return file_name

def create_departments_sheet(wb, data):
    """إنشاء ورقة الأقسام"""
    ws = wb.create_sheet("الأقسام", 0)
    
    # العناوين العربية
    headers = ['اسم القسم', 'اسم المدير']
    
    # إضافة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    # إضافة البيانات
    for row_idx, (dept_name, manager_name) in enumerate(zip(data['name'], data['manager_name']), 2):
        ws.cell(row=row_idx, column=1, value=dept_name)
        ws.cell(row=row_idx, column=2, value=manager_name)
    
    # تنسيق الجدول
    format_table(ws, len(data['name']))
    
    # إضافة تعليمات ورقة الأقسام
    add_sheet_instructions(ws, [
        "• يجب إضافة الأقسام في هذه الورقة أولاً",
        "• اسم القسم يجب أن يكون فريداً",
        "• اسم المدير اختياري - يمكن تركه فارغاً",
        "• سيتم ربط المستخدمين بالأقسام باستخدام اسم القسم"
    ])

def create_users_sheet(wb, data):
    """إنشاء ورقة المستخدمين"""
    ws = wb.create_sheet("المستخدمين", 1)
    
    # العناوين العربية
    headers = [
        'اسم المستخدم', 
        'الاسم الكامل', 
        'كلمة المرور', 
        'مدير نظام', 
        'مدير قسم', 
        'اسم القسم'
    ]
    
    # إضافة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    # إضافة البيانات
    for row_idx in range(len(data['username'])):
        ws.cell(row=row_idx+2, column=1, value=data['username'][row_idx])
        ws.cell(row=row_idx+2, column=2, value=data['name'][row_idx])
        ws.cell(row=row_idx+2, column=3, value=data['password'][row_idx])
        ws.cell(row=row_idx+2, column=4, value=data['is_admin'][row_idx])
        ws.cell(row=row_idx+2, column=5, value=data['is_manager'][row_idx])
        ws.cell(row=row_idx+2, column=6, value=data['department_name'][row_idx])
    
    # تنسيق الجدول
    format_table(ws, len(data['username']))
    
    # إضافة تعليمات ورقة المستخدمين
    add_sheet_instructions(ws, [
        "• اسم المستخدم يجب أن يكون فريداً وغير مكرر",
        "• كلمة المرور يجب أن تكون 6 أحرف على الأقل",
        "• مدير نظام: 'نعم' أو 'لا'",
        "• مدير قسم: 'نعم' أو 'لا'", 
        "• اسم القسم: يجب أن يكون مطابقاً تماماً لورقة الأقسام",
        "• يمكن ترك اسم القسم فارغاً إذا لم يكن للمستخدم قسم محدد"
    ])

def create_instructions_sheet(wb):
    """إنشاء ورقة التعليمات العامة"""
    ws = wb.create_sheet("التعليمات", 2)
    
    # عنوان الورقة
    title_cell = ws.cell(row=1, column=1, value="تعليمات استيراد البيانات")
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center')
    
    # التعليمات العامة
    instructions = [
        "🟢 خطوات الاستيراد:",
        "1. املأ بيانات الأقسام في ورقة 'الأقسام'",
        "2. املأ بيانات المستخدمين في ورقة 'المستخدمين'", 
        "3. احفظ الملف",
        "4. استخدم سكريبت الاستيراد لتحميل البيانات للنظام",
        "",
        "🟡 ملاحظات هامة:",
        "• لا تحذف أو تعدل في العناوين (الصف الأول)",
        "• الحقول المطلوبة: اسم القسم، اسم المستخدم، الاسم الكامل، كلمة المرور",
        "• الحقول الاختيارية: اسم المدير، مدير نظام، مدير قسم، اسم القسم",
        "• يمكن إضافة أي عدد من الصفوف",
        "• يجب أن تتطابق أسماء الأقسام في الورقتين تماماً",
        "",
        "🔴 تجنب:",
        "• تكرار اسم المستخدم",
        "• تكرار اسم القسم", 
        "• كلمات مرور ضعيفة",
        "• تعديل تنسيق الملف"
    ]
    
    # إضافة التعليمات
    for row_idx, instruction in enumerate(instructions, 3):
        cell = ws.cell(row=row_idx, column=1, value=instruction)
        if "🟢" in instruction or "🟡" in instruction or "🔴" in instruction:
            cell.font = Font(bold=True, size=12)
        else:
            cell.font = Font(size=11)
    
    # ضبط عرض العمود
    ws.column_dimensions['A'].width = 60

def apply_header_style(cell):
    """تطبيق تنسيق رأس الجدول"""
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def format_table(ws, data_rows_count):
    """تنسيق الجدول"""
    # تحديد نطاق الجدول
    max_row = data_rows_count + 1
    max_col = ws.max_column
    
    # ضبط عرض الأعمدة
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        if col == 2:  # عمود الاسم الكامل
            ws.column_dimensions[column_letter].width = 25
        elif col == 6:  # عمود اسم القسم
            ws.column_dimensions[column_letter].width = 30
        else:
            ws.column_dimensions[column_letter].width = 15
    
    # تطبيق الحدود على جميع الخلايا
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # ليس عناوين
                cell.alignment = Alignment(horizontal='center', vertical='center')

def add_sheet_instructions(ws, instructions):
    """إضافة تعليمات أسفل الورقة"""
    start_row = ws.max_row + 3
    
    title_cell = ws.cell(row=start_row, column=1, value="ملاحظات هامة:")
    title_cell.font = Font(bold=True, color="FF0000", size=12)
    
    for idx, instruction in enumerate(instructions, 1):
        ws.cell(row=start_row + idx, column=1, value=instruction)
        ws.cell(row=start_row + idx, column=1).font = Font(size=10, color="555555")

def create_minimal_template():
    """إنشاء نموذج مبسط للاستخدام السريع"""
    wb = Workbook()
    
    # ورقة الأقسام
    ws_dept = wb.active
    ws_dept.title = "الأقسام"
    
    # عناوين الأقسام
    dept_headers = ['اسم القسم', 'اسم المدير']
    for col, header in enumerate(dept_headers, 1):
        ws_dept.cell(row=1, column=col, value=header)
    
    # بيانات نموذجية للأقسام
    sample_depts = [
        ['قسم تكنولوجيا المعلومات', 'أحمد محمود'],
        ['قسم الموارد البشرية', 'سارة علي'],
        ['قسم المالية', 'محمد عبدالله']
    ]
    
    for row, dept in enumerate(sample_depts, 2):
        ws_dept.cell(row=row, column=1, value=dept[0])
        ws_dept.cell(row=row, column=2, value=dept[1])
    
    # ورقة المستخدمين
    ws_users = wb.create_sheet("المستخدمين")
    
    # عناوين المستخدمين
    user_headers = ['اسم المستخدم', 'الاسم الكامل', 'كلمة المرور', 'مدير نظام', 'مدير قسم', 'اسم القسم']
    for col, header in enumerate(user_headers, 1):
        ws_users.cell(row=1, column=col, value=header)
    
    # بيانات نموذجية للمستخدمين
    sample_users = [
        ['ahmed.it', 'أحمد محمود', '123456', 'نعم', 'نعم', 'قسم تكنولوجيا المعلومات'],
        ['sara.hr', 'سارة علي', '654321', 'لا', 'نعم', 'قسم الموارد البشرية'],
        ['mohamed.finance', 'محمد عبدالله', '112233', 'لا', 'نعم', 'قسم المالية']
    ]
    
    for row, user in enumerate(sample_users, 2):
        for col, value in enumerate(user, 1):
            ws_users.cell(row=row, column=col, value=value)
    
    # حفظ الملف
    file_name = "نموذج_مبسط_للموظفين.xlsx"
    wb.save(file_name)
    print(f"✓ تم إنشاء النموذج المبسط: {file_name}")

if __name__ == "__main__":
    print("=" * 60)
    print("أداة إنشاء ملفات Excel لاستيراد الموظفين")
    print("=" * 60)
    
    while True:
        print("\nاختر نوع النموذج:")
        print("1 - نموذج مفصل (مصمم وجميل)")
        print("2 - نموذج مبسط (للاستخدام السريع)")
        print("3 - الخروج")
        
        choice = input("\nادخل رقم الخيار: ").strip()
        
        if choice == '1':
            print("جاري إنشاء النموذج المفصل...")
            file_name = create_departments_users_template()
            print("\n🎉 تم إنشاء الملف بنجاح!")
            print("📋 يحتوي الملف على 3 أوراق:")
            print("   - الأقسام: لإضافة أقسام الشركة")
            print("   - المستخدمين: لإضافة بيانات الموظفين") 
            print("   - التعليمات: لشرح طريقة الاستخدام")
            
        elif choice == '2':
            print("جاري إنشاء النموذج المبسط...")
            create_minimal_template()
            print("\n✓ تم إنشاء النموذج المبسط بنجاح!")
            print("📋 يحتوي الملف على ورقتين:")
            print("   - الأقسام")
            print("   - المستخدمين")
            
        elif choice == '3':
            print("مع السلامة!")
            break
            
        else:
            print("✗ خيار غير صحيح!")
        
        # سؤال إذا كان يريد إنشاء ملف آخر
        if choice in ['1', '2']:
            again = input("\nهل تريد إنشاء ملف آخر؟ (نعم/لا): ").strip().lower()
            if again not in ['نعم', 'yes', 'y', '']:
                break