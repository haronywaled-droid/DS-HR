import os
import json
import re
import warnings
import traceback
import smtplib
from io import BytesIO
from datetime import datetime, date, timedelta
from decimal import Decimal
from functools import wraps
from typing import Any, Dict, List, Union, Optional
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import date, datetime
import sys
import logging
from sqlalchemy import exc
import warnings
from sqlalchemy import func

import pandas as pd
from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import extract, or_
from sqlalchemy.exc import SAWarning
from sqlalchemy.orm import joinedload
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from database import init_db, db_session
from models import *
from sqlalchemy.orm import Session


app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here-change-this-in-production'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['EXPORT_FOLDER'] = 'exports'
app.config['SALARY_FOLDER'] = r'D:\Haron\hr_system\salary_slips'
logging.basicConfig(level=logging.DEBUG)

app.config['EXCEL_DATA_FOLDER'] = 'att'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'pdf'}
load_dotenv()


# Create folders if not exist
for folder in [app.config['UPLOAD_FOLDER'], app.config['EXPORT_FOLDER'], app.config['SALARY_FOLDER']]:
    os.makedirs(folder, exist_ok=True)

app.config.update(
    MAIL_SERVER='smtp.gmail.com',
    MAIL_PORT=587,
    MAIL_USE_TLS=True,
    MAIL_USERNAME=os.getenv('GMAIL_USERNAME', 'alharonywaled@gmail.com'),
    MAIL_PASSWORD=os.getenv('GMAIL_PASSWORD', 'H@r0n011**'),
    MAIL_DEFAULT_SENDER=os.getenv('GMAIL_USERNAME', 'alharonywaled@gmail.com')
)
from flask_mail import Mail
mail = Mail(app)
# Setup Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

warnings.filterwarnings('ignore', category=exc.SADeprecationWarning)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def calculate_age(birth_date):
    today = date.today()
    return today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))



def send_notification(user_id, title, message, notification_type='general', related_id=None):
    """إرسال إشعار للمستخدم"""
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        notification_type=notification_type,
        related_id=related_id
    )
    db_session.add(notification)
    db_session.commit()
    return notification



# ======== Messaging System Routes ========

@app.route('/messages')
@login_required
def messages_inbox():
    """Display user's inbox"""
    # Get messages where user is recipient
    recipient_messages = db_session.query(MessageRecipient)\
    .join(Message)\
    .filter(MessageRecipient.user_id == current_user.id, 
            MessageRecipient.is_archived == False, 
            MessageRecipient.is_deleted == False)\
    .order_by(Message.created_at.desc())\
    .all()
    
    messages = []
    for rm in recipient_messages:
        msg = rm.message
        msg.recipient_info = rm
        msg.recipient_type = rm.recipient_type
        msg.is_read = rm.is_read
        messages.append(msg)
    
    # Get unread count
    unread_count = db_session.query(MessageRecipient).filter_by(
        user_id=current_user.id,
        is_read=False,
        is_deleted=False
    ).count()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('messages/inbox.html',
                         messages=messages,
                         unread_count=unread_count,
                         folder='inbox',
                         notifications=notifications)

@app.route('/messages/sent')
@login_required
def messages_sent():
    """Display user's sent messages"""
    messages = db_session.query(Message)\
        .filter_by(sender_id=current_user.id, is_draft=False)\
        .order_by(Message.created_at.desc())\
        .all()
    
    # Load recipients for each message
    for msg in messages:
        msg.recipients_list = db_session.query(MessageRecipient)\
            .filter_by(message_id=msg.id)\
            .all()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('messages/sent.html',
                         messages=messages,
                         folder='sent',
                         notifications=notifications)

@app.route('/messages/drafts')
@login_required
def messages_drafts():
    """Display user's draft messages"""
    messages = db_session.query(Message)\
        .filter_by(sender_id=current_user.id, is_draft=True)\
        .order_by(Message.updated_at.desc())\
        .all()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('messages/drafts.html',
                         messages=messages,
                         folder='drafts',
                         notifications=notifications)

@app.route('/messages/starred')
@login_required
def messages_starred():
    """Display user's starred messages"""
    # Get starred messages where user is recipient
    recipient_starred = db_session.query(MessageRecipient)\
        .join(Message)\
        .filter(
            MessageRecipient.user_id == current_user.id,
            Message.is_starred == True,
            MessageRecipient.is_deleted == False
        )\
        .all()
    
    messages = [rm.message for rm in recipient_starred]
    
    # Get starred messages sent by user
    sent_starred = db_session.query(Message)\
        .filter(
            Message.sender_id == current_user.id,
            Message.is_starred == True
        )\
        .all()
    
    messages.extend(sent_starred)
    
    # Remove duplicates and sort
    unique_messages = {msg.id: msg for msg in messages}
    messages = sorted(unique_messages.values(), 
                     key=lambda x: x.created_at, 
                     reverse=True)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('messages/starred.html',
                         messages=messages,
                         folder='starred',
                         notifications=notifications)

@app.route('/messages/archived')
@login_required
def messages_archived():
    """Display user's archived messages"""
    archived = db_session.query(MessageRecipient)\
        .options(joinedload(MessageRecipient.message))\
        .filter(
            MessageRecipient.user_id == current_user.id,
            MessageRecipient.is_archived == True,
            MessageRecipient.is_deleted == False
        )\
        .all()
    
    messages = [a.message for a in archived]
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('messages/archived.html',
                         messages=messages,
                         folder='archived',
                         notifications=notifications)

@app.route('/messages/compose', methods=['GET', 'POST'])
@login_required
def messages_compose():
    """Compose a new message"""
    if request.method == 'POST':
        try:
            # Get form data
            to_users = request.form.getlist('to[]')
            cc_users = request.form.getlist('cc[]')
            bcc_users = request.form.getlist('bcc[]')
            subject = request.form.get('subject', '').strip()
            body = request.form.get('body', '').strip()
            priority = request.form.get('priority', 'normal')
            save_as_draft = 'save_draft' in request.form
            
            # Validate
            if not save_as_draft:
                if not to_users:
                    flash('يرجى تحديد مستلم واحد على الأقل', 'error')
                    return redirect(url_for('messages_compose'))
                
                if not subject:
                    flash('يرجى إدخال عنوان للرسالة', 'error')
                    return redirect(url_for('messages_compose'))
                
                if not body:
                    flash('يرجى إدخال نص الرسالة', 'error')
                    return redirect(url_for('messages_compose'))
            
            # Create message
            message = Message(
                sender_id=current_user.id,
                subject=subject,
                body=body,
                priority=priority,
                is_draft=save_as_draft
            )
            db_session.add(message)
            db_session.flush()  # Get message ID
            
            # Add recipients
            all_recipients = []
            
            # Add TO recipients
            for user_id in to_users:
                if user_id and user_id.isdigit():
                    recipient = MessageRecipient(
                        message_id=message.id,
                        user_id=int(user_id),
                        recipient_type='to'
                    )
                    db_session.add(recipient)
                    all_recipients.append(int(user_id))
            
            # Add CC recipients
            for user_id in cc_users:
                if user_id and user_id.isdigit() and int(user_id) not in all_recipients:
                    recipient = MessageRecipient(
                        message_id=message.id,
                        user_id=int(user_id),
                        recipient_type='cc'
                    )
                    db_session.add(recipient)
                    all_recipients.append(int(user_id))
            
            # Add BCC recipients
            for user_id in bcc_users:
                if user_id and user_id.isdigit() and int(user_id) not in all_recipients:
                    recipient = MessageRecipient(
                        message_id=message.id,
                        user_id=int(user_id),
                        recipient_type='bcc'
                    )
                    db_session.add(recipient)
                    all_recipients.append(int(user_id))
            
            # Handle file attachments
            files = request.files.getlist('attachments')
            for file in files:
                if file and file.filename:
                    if allowed_file(file.filename):
                        # Create attachments folder
                        attachments_folder = os.path.join(app.config['UPLOAD_FOLDER'], 
                                                         'messages', str(message.id))
                        os.makedirs(attachments_folder, exist_ok=True)
                        
                        # Save file
                        filename = secure_filename(file.filename)
                        # Add timestamp to avoid duplicates
                        name, ext = os.path.splitext(filename)
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        saved_filename = f"{name}_{timestamp}{ext}"
                        file_path = os.path.join(attachments_folder, saved_filename)
                        file.save(file_path)
                        
                        # Get file size
                        file_size = os.path.getsize(file_path)
                        
                        # Create attachment record
                        attachment = MessageAttachment(
                            message_id=message.id,
                            filename=saved_filename,
                            original_filename=filename,
                            file_path=file_path,
                            file_size=file_size,
                            mime_type=file.content_type
                        )
                        db_session.add(attachment)
                        
                        message.has_attachments = True
            
            db_session.commit()
            
            if save_as_draft:
                flash('تم حفظ المسودة بنجاح', 'success')
                return redirect(url_for('messages_drafts'))
            else:
                # Send notifications to recipients
                for user_id in all_recipients:
                    user = db_session.get(User, user_id)
                    if user:
                        # Determine recipient type for notification
                        if user_id in [int(uid) for uid in to_users]:
                            recipient_type = 'to'
                        elif user_id in [int(uid) for uid in cc_users]:
                            recipient_type = 'cc'
                        else:
                            recipient_type = 'bcc'
                        
                        # Create notification
                        notification_message = f'لديك رسالة جديدة من {current_user.name}'
                        if priority == 'high':
                            notification_message = f'🔴 [مهم] {notification_message}'
                        elif priority == 'urgent':
                            notification_message = f'⚠️ [عاجل] {notification_message}'
                        
                        create_notification(
                            user_id,
                            f'رسالة جديدة: {subject[:50]}{"..." if len(subject) > 50 else ""}',
                            notification_message,
                            'new_message',
                            related_id=message.id,
                            action_url=url_for('messages_view', message_id=message.id)
                        )
                
                flash('تم إرسال الرسالة بنجاح', 'success')
                return redirect(url_for('messages_sent'))
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'error')
            return redirect(url_for('messages_compose'))
    
    # GET request - show compose form
    # Get all users for recipient selection
    users = db_session.query(User).filter(
        User.is_active == True,
        User.id != current_user.id
    ).order_by(User.name).all()
    
    # Get departments for filtering
    departments = db_session.query(Department).all()
    
    # Get draft if replying/forwarding
    reply_to_id = request.args.get('reply_to')
    forward_id = request.args.get('forward')
    quote_id = request.args.get('quote')
    
    reply_to_msg = None
    forward_msg = None
    quote_msg = None
    
    if reply_to_id:
        reply_to_msg = db_session.get(Message, int(reply_to_id))
    elif forward_id:
        forward_msg = db_session.get(Message, int(forward_id))
    elif quote_id:
        quote_msg = db_session.get(Message, int(quote_id))
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('messages/compose.html',
                         users=users,
                         departments=departments,
                         reply_to_msg=reply_to_msg,
                         forward_msg=forward_msg,
                         quote_msg=quote_msg,
                         notifications=notifications)

@app.route('/messages/view/<int:message_id>')
@login_required
def messages_view(message_id):
    """View a single message"""
    message = db_session.get(Message, message_id)
    
    if not message:
        flash('الرسالة غير موجودة', 'error')
        return redirect(url_for('messages_inbox'))
    
    # Check if user has access to this message
    recipient = db_session.query(MessageRecipient).filter_by(
        message_id=message_id,
        user_id=current_user.id
    ).first()
    
    is_sender = message.sender_id == current_user.id
    
    if not recipient and not is_sender:
        flash('ليس لديك صلاحية لعرض هذه الرسالة', 'error')
        return redirect(url_for('messages_inbox'))
    
    # Mark as read if recipient
    if recipient and not recipient.is_read:
        recipient.is_read = True
        recipient.read_at = datetime.now()
        db_session.commit()
    
    # Load attachments
    attachments = db_session.query(MessageAttachment)\
        .filter_by(message_id=message_id)\
        .all()
    
    # Load all recipients
    recipients = db_session.query(MessageRecipient)\
        .filter_by(message_id=message_id)\
        .all()
    
    # Get thread (replies)
    thread_messages = []
    if message.thread_id:
        # This is a reply, get the whole thread
        thread_root_id = message.thread_id
        thread_messages = db_session.query(Message)\
            .filter(
                or_(
                    Message.id == thread_root_id,
                    Message.thread_id == thread_root_id,
                    Message.id == message_id
                )
            )\
            .order_by(Message.created_at)\
            .all()
    else:
        # This is the root message, get all replies
        thread_messages = db_session.query(Message)\
            .filter(
                or_(
                    Message.id == message_id,
                    Message.thread_id == message_id
                )
            )\
            .order_by(Message.created_at)\
            .all()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('messages/view.html',
                         message=message,
                         recipients=recipients,
                         attachments=attachments,
                         thread_messages=thread_messages,
                         is_sender=is_sender,
                         is_recipient=recipient is not None,
                         notifications=notifications)

@app.route('/messages/reply/<int:message_id>', methods=['POST'])
@login_required
def messages_reply(message_id):
    """Reply to a message"""
    original = db_session.get(Message, message_id)
    
    if not original:
        flash('الرسالة غير موجودة', 'error')
        return redirect(url_for('messages_inbox'))
    
    # Check if user can reply (must be recipient or sender)
    recipient = db_session.query(MessageRecipient).filter_by(
        message_id=message_id,
        user_id=current_user.id
    ).first()
    
    if not recipient and original.sender_id != current_user.id:
        flash('ليس لديك صلاحية للرد على هذه الرسالة', 'error')
        return redirect(url_for('messages_inbox'))
    
    try:
        body = request.form.get('body', '').strip()
        
        if not body:
            flash('يرجى إدخال نص الرد', 'error')
            return redirect(url_for('messages_view', message_id=message_id))
        
        # Determine recipients
        to_users = []
        
        if original.sender_id == current_user.id:
            # Sender replying - send to all original recipients (except BCC)
            original_recipients = db_session.query(MessageRecipient)\
                .filter_by(message_id=message_id)\
                .filter(MessageRecipient.recipient_type.in_(['to', 'cc']))\
                .all()
            to_users = [r.user_id for r in original_recipients if r.user_id != current_user.id]
        else:
            # Recipient replying - send to sender
            to_users = [original.sender_id]
        
        # Create reply
        reply = Message(
            sender_id=current_user.id,
            subject=f"Re: {original.subject}",
            body=body,
            thread_id=original.thread_id if original.thread_id else original.id,
            is_draft=False
        )
        db_session.add(reply)
        db_session.flush()
        
        # Add recipients
        for user_id in to_users:
            recipient = MessageRecipient(
                message_id=reply.id,
                user_id=user_id,
                recipient_type='to'
            )
            db_session.add(recipient)
            
            # Send notification
            create_notification(
                user_id,
                f'رد على: {original.subject[:50]}',
                f'قام {current_user.name} بالرد على رسالتك',
                'message_reply',
                related_id=reply.id,
                action_url=url_for('messages_view', message_id=reply.id)
            )
        
        db_session.commit()
        flash('تم إرسال الرد بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('messages_view', message_id=message_id))

@app.route('/messages/star/<int:message_id>', methods=['POST'])
@login_required
def messages_star(message_id):
    """Star or unstar a message"""
    message = db_session.get(Message, message_id)
    
    if not message:
        return jsonify({'success': False, 'message': 'الرسالة غير موجودة'})
    
    # Check access
    recipient = db_session.query(MessageRecipient).filter_by(
        message_id=message_id,
        user_id=current_user.id
    ).first()
    
    if not recipient and message.sender_id != current_user.id:
        return jsonify({'success': False, 'message': 'لا يمكنك الوصول إلى هذه الرسالة'})
    
    message.is_starred = not message.is_starred
    db_session.commit()
    
    return jsonify({
        'success': True,
        'is_starred': message.is_starred
    })

@app.route('/messages/archive/<int:message_id>', methods=['POST'])
@login_required
def messages_archive(message_id):
    """Archive a message"""
    recipient = db_session.query(MessageRecipient).filter_by(
        message_id=message_id,
        user_id=current_user.id
    ).first()
    
    if recipient:
        recipient.is_archived = not recipient.is_archived
        db_session.commit()
        
        return jsonify({
            'success': True,
            'is_archived': recipient.is_archived
        })
    
    return jsonify({'success': False, 'message': 'الرسالة غير موجودة'})

@app.route('/messages/delete/<int:message_id>', methods=['POST'])
@login_required
def messages_delete(message_id):
    """Delete a message (move to trash)"""
    recipient = db_session.query(MessageRecipient).filter_by(
        message_id=message_id,
        user_id=current_user.id
    ).first()
    
    if recipient:
        recipient.is_deleted = True
        db_session.commit()
        flash('تم نقل الرسالة إلى سلة المهملات', 'success')
    else:
        flash('الرسالة غير موجودة', 'error')
    
    return redirect(url_for('messages_inbox'))

@app.route('/messages/mark_read/<int:message_id>', methods=['POST'])
@login_required
def messages_mark_read(message_id):
    """Mark message as read/unread"""
    recipient = db_session.query(MessageRecipient).filter_by(
        message_id=message_id,
        user_id=current_user.id
    ).first()
    
    if recipient:
        data = request.get_json()
        is_read = data.get('is_read', True)
        
        recipient.is_read = is_read
        if is_read:
            recipient.read_at = datetime.now()
        else:
            recipient.read_at = None
        
        db_session.commit()
        
        return jsonify({'success': True, 'is_read': recipient.is_read})
    
    return jsonify({'success': False, 'message': 'الرسالة غير موجودة'})

@app.route('/messages/search')
@login_required
def messages_search():
    """Search messages"""
    query = request.args.get('q', '')
    
    if not query or len(query) < 2:
        flash('يرجى إدخال على الأقل 2 أحرف للبحث', 'error')
        return redirect(url_for('messages_inbox'))
    
    # Search in received messages
    received = db_session.query(Message)\
        .join(MessageRecipient)\
        .filter(
            MessageRecipient.user_id == current_user.id,
            MessageRecipient.is_deleted == False,
            or_(
                Message.subject.ilike(f'%{query}%'),
                Message.body.ilike(f'%{query}%')
            )
        )\
        .all()
    
    # Search in sent messages
    sent = db_session.query(Message)\
        .filter(
            Message.sender_id == current_user.id,
            or_(
                Message.subject.ilike(f'%{query}%'),
                Message.body.ilike(f'%{query}%')
            )
        )\
        .all()
    
    # Combine and remove duplicates
    all_messages = {msg.id: msg for msg in received + sent}
    messages = sorted(all_messages.values(), 
                     key=lambda x: x.created_at, 
                     reverse=True)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('messages/search_results.html',
                         messages=messages,
                         query=query,
                         notifications=notifications)

@app.route('/api/users/search')
@login_required
def api_users_search():
    """API endpoint for searching users (for compose form)"""
    query = request.args.get('q', '')
    department_id = request.args.get('department_id')
    
    users_query = db_session.query(User).filter(
        User.is_active == True,
        User.id != current_user.id
    )
    
    if query:
        users_query = users_query.filter(
            or_(
                User.name.ilike(f'%{query}%'),
                User.username.ilike(f'%{query}%')
            )
        )
    
    if department_id and department_id.isdigit():
        users_query = users_query.filter(User.department_id == int(department_id))
    
    users = users_query.limit(20).all()
    
    results = []
    for user in users:
        department = db_session.get(Department, user.department_id)
        results.append({
            'id': user.id,
            'name': user.name,
            'username': user.username,
            'department': department.name if department else '',
            'display': f"{user.name} ({user.username})"
        })
    
    return jsonify(results)

@app.route('/api/messages/unread_count')
@login_required
def api_messages_unread_count():
    """Get unread messages count"""
    count = db_session.query(MessageRecipient).filter_by(
        user_id=current_user.id,
        is_read=False,
        is_deleted=False
    ).count()
    
    return jsonify({'unread_count': count})

@app.route('/messages/download_attachment/<int:attachment_id>')
@login_required
def download_message_attachment(attachment_id):
    """Download a message attachment"""
    attachment = db_session.get(MessageAttachment, attachment_id)
    
    if not attachment:
        flash('الملف غير موجود', 'error')
        return redirect(request.referrer or url_for('messages_inbox'))
    
    # Check if user has access to this message
    message = db_session.get(Message, attachment.message_id)
    recipient = db_session.query(MessageRecipient).filter_by(
        message_id=attachment.message_id,
        user_id=current_user.id
    ).first()
    
    if not recipient and message.sender_id != current_user.id:
        flash('ليس لديك صلاحية لتحميل هذا الملف', 'error')
        return redirect(request.referrer or url_for('messages_inbox'))
    
    if os.path.exists(attachment.file_path):
        return send_file(
            attachment.file_path,
            as_attachment=True,
            download_name=attachment.original_filename
        )
    else:
        flash('الملف غير موجود على الخادم', 'error')
        return redirect(request.referrer or url_for('messages_inbox'))


def send_email(to_email, subject, html_content, text_content=None):
    """
    Send email using Gmail SMTP
    """
    try:
        # Get email credentials from config
        sender_email = app.config['MAIL_USERNAME']
        sender_password = app.config['MAIL_PASSWORD']
        
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = sender_email
        msg['To'] = to_email
        
        # Create the plain-text and HTML version of your message
        if text_content:
            part1 = MIMEText(text_content, 'plain')
            msg.attach(part1)
        
        part2 = MIMEText(html_content, 'html')
        msg.attach(part2)
        
        # Send email using Gmail SMTP
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
        
        print(f"✅ Email sent to {to_email}")
        return True
        
    except Exception as e:
        print(f"❌ Failed to send email to {to_email}: {str(e)}")
        return False


def send_notification_email(user_email, notification_title, notification_message, notification_type='general'):
    """
    Send formatted notification email to user
    """
    try:
        # Create HTML email content
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background-color: #4CAF50; color: white; padding: 20px; text-align: center; border-radius: 5px 5px 0 0; }}
                .content {{ background-color: #f9f9f9; padding: 30px; border-radius: 0 0 5px 5px; }}
                .notification-type {{ display: inline-block; padding: 5px 10px; background-color: #e7f3fe; color: #2196F3; border-radius: 3px; font-size: 12px; margin-bottom: 10px; }}
                .footer {{ text-align: center; margin-top: 30px; font-size: 12px; color: #666; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2>🔔 إشعار جديد</h2>
                </div>
                <div class="content">
                    <div class="notification-type">
                        {notification_type}
                    </div>
                    <h3>{notification_title}</h3>
                    <p>{notification_message}</p>
                    <p>📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
                </div>
                <div class="footer">
                    <p>هذا الإشعار تم إرساله تلقائياً من نظام إدارة الموظفين</p>
                    <p>© {datetime.now().year} جميع الحقوق محفوظة</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Create plain text version
        text_content = f"""
        إشعار جديد
        =========
        
        العنوان: {notification_title}
        النوع: {notification_type}
        الرسالة: {notification_message}
        
        التاريخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}
        
        هذا الإشعار تم إرساله تلقائياً من نظام إدارة الموظفين
        """
        
        # Send email
        subject = f"🔔 {notification_title}"
        return send_email(user_email, subject, html_content, text_content)
        
    except Exception as e:
        print(f"❌ Error creating notification email: {str(e)}")
        return False



def create_notification(user_id, title, message, notification_type='general', related_id=None, action_url=None):
    """
    Create notification and send email
    """
    # Create notification in database
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        notification_type=notification_type,
        related_id=related_id,
        action_url=action_url
    )
    db_session.add(notification)
    db_session.commit()
    
    # Send email notification
    try:
        # Get user information including email
        user = db_session.get(User, user_id)
        if user and hasattr(user, 'email') and user.email:
            # Send email in background (non-blocking)
            send_notification_email(user.email, title, message, notification_type)
        else:
            print(f"⚠️ User {user_id} has no email address or email field doesn't exist")
    except Exception as e:
        print(f"❌ Error sending email for notification: {str(e)}")
    
    return notification


def get_notification_email_template(notification_type, title, message):
    """
    Get different email templates based on notification type
    """
    templates = {
        'leave_approved': {
            'subject': '✅ موافقة على طلب الإجازة',
            'icon': '🎉',
            'color': '#4CAF50'
        },
        'leave_rejected': {
            'subject': '❌ رفض طلب الإجازة',
            'icon': '⚠️',
            'color': '#f44336'
        },
        'salary': {
            'subject': '💰 شيت مرتب جديد',
            'icon': '💰',
            'color': '#FF9800'
        },
        'permission_approved': {
            'subject': '✅ موافقة على طلب الإذن',
            'icon': '👍',
            'color': '#2196F3'
        },
        'schedule': {
            'subject': '📅 جدول عمل جديد',
            'icon': '📅',
            'color': '#9C27B0'
        },
        'default': {
            'subject': '🔔 إشعار جديد',
            'icon': '🔔',
            'color': '#607D8B'
        }
    }
    
    template = templates.get(notification_type, templates['default'])
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: 'Arial', sans-serif; direction: rtl; line-height: 1.6; }}
            .container {{ max-width: 600px; margin: 20px auto; background: #fff; border-radius: 10px; overflow: hidden; box-shadow: 0 0 20px rgba(0,0,0,0.1); }}
            .header {{ background: {template['color']}; padding: 30px; text-align: center; color: white; }}
            .header h1 {{ margin: 0; font-size: 24px; }}
            .icon {{ font-size: 40px; margin-bottom: 15px; }}
            .content {{ padding: 30px; }}
            .title {{ color: {template['color']}; font-size: 20px; margin-bottom: 20px; border-bottom: 2px solid {template['color']}; padding-bottom: 10px; }}
            .message {{ font-size: 16px; line-height: 1.8; color: #444; margin-bottom: 30px; }}
            .footer {{ background: #f5f5f5; padding: 20px; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #ddd; }}
            .button {{ display: inline-block; background: {template['color']}; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; margin-top: 20px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <div class="icon">{template['icon']}</div>
                <h1>{template['subject']}</h1>
            </div>
            <div class="content">
                <div class="title">{title}</div>
                <div class="message">{message}</div>
                <div style="text-align: center;">
                    <a href="{app.config.get('APP_URL', 'http://192.168.2.70:5551')}" class="button">الدخول إلى النظام</a>
                </div>
            </div>
            <div class="footer">
                <p>© {datetime.now().year} نظام إدارة الموظفين - جميع الحقوق محفوظة</p>
                <p>هذا إشعار تلقائي، يرجى عدم الرد على هذا البريد الإلكتروني</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return template['subject'], html_content


def send_notification_email(user_email, notification_title, notification_message, notification_type='general'):
    """
    Send formatted notification email using templates
    """
    try:
        # Get email template based on notification type
        subject, html_content = get_notification_email_template(
            notification_type, 
            notification_title, 
            notification_message
        )
        
        # Create plain text version
        text_content = f"""
        {notification_title}
        {'=' * len(notification_title)}
        
        {notification_message}
        
        نوع الإشعار: {notification_type}
        التاريخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}
        
        للدخول إلى النظام: {app.config.get('APP_URL', 'http://192.168.2.70:5551')}
        
        ---
        نظام إدارة الموظفين
        """
        
        # Send email
        return send_email(user_email, subject, html_content, text_content)
        
    except Exception as e:
        print(f"❌ Error sending notification email: {str(e)}")
        return False

def create_notification(user_id, title, message, notification_type='general', related_id=None, action_url=None):
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        notification_type=notification_type,
        related_id=related_id,
        action_url=action_url
    )
    db_session.add(notification)
    db_session.commit()
    return notification

def get_user_notifications(user_id):
    return db_session.query(Notification).filter_by(
        user_id=user_id, is_read=False
    ).order_by(Notification.created_at.desc()).all()

def export_to_excel(user_id):
    user = db_session.get(User, user_id)
    employee_data = db_session.query(EmployeeData).filter_by(user_id=user_id).first()
    
    if not employee_data:
        return None
    
    data_dict = {
        'الاسم بالعربية': [employee_data.arabic_name or ''],
        'الاسم بالإنجليزية': [employee_data.english_name or ''],
        'الرقم القومي': [employee_data.national_id or ''],
        'تاريخ إصدار البطاقة': [employee_data.id_issue_date or ''],
        'تاريخ الميلاد': [employee_data.birth_date or ''],
        'السن': [employee_data.age or ''],
        'واتساب': [employee_data.whatsapp or ''],
        'الهاتف': [employee_data.phone or ''],
        'العنوان': [employee_data.address or ''],
        'الموقف من التجنيد': [employee_data.military_status or ''],
        'الحالة الاجتماعية': [employee_data.marital_status or ''],
        'المؤهل الدراسي': [employee_data.qualification or ''],
        'سنة التخرج': [employee_data.graduation_year or ''],
        'التقدير': [employee_data.grade or ''],
        'يعمل حالياً': ['نعم' if employee_data.has_work else 'لا'],
        'جهة العمل': [employee_data.workplace or ''],
        'الوظيفة': [employee_data.job_title or ''],
        'الرقم التأميني': [employee_data.insurance_number or ''],
        'الرقم الضريبي': [employee_data.tax_number or ''],
        'ترخيص المهنة': [employee_data.profession_license or ''],
        'كارنية النقابة': [employee_data.union_card or ''],
        'جهة الاتصال 1 - الاسم': [employee_data.emergency1_name or ''],
        'جهة الاتصال 1 - التليفون': [employee_data.emergency1_phone or ''],
        'جهة الاتصال 1 - العنوان': [employee_data.emergency1_address or ''],
        'جهة الاتصال 1 - صلة القرابة': [employee_data.emergency1_relation or ''],
        'جهة الاتصال 2 - الاسم': [employee_data.emergency2_name or ''],
        'جهة الاتصال 2 - التليفون': [employee_data.emergency2_phone or ''],
        'جهة الاتصال 2 - العنوان': [employee_data.emergency2_address or ''],
        'جهة الاتصال 2 - صلة القرابة': [employee_data.emergency2_relation or ''],
        'نسبة الإكتمال': [f'{employee_data.completion_percentage}%'],
        'آخر تحديث': [employee_data.last_updated or ''],
        'تم التحديث بواسطة': [employee_data.updated_by or '']
    }
    
    df = pd.DataFrame(data_dict)
    
    user_export_folder = os.path.join(app.config['EXPORT_FOLDER'], str(user_id))
    os.makedirs(user_export_folder, exist_ok=True)
    
    filename = f"employee_data_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(user_export_folder, filename)
    df.to_excel(filepath, index=False, engine='openpyxl')
    
    return filepath



def update_database_schema():
    """تحديث هيكل قاعدة البيانات بإضافة الحقول المفقودة"""
    try:
        # استيراد سكريبت الإصلاح وتشغيله
        from database_fix import fix_database, check_specific_columns
        
        print("جاري تحديث هيكل قاعدة البيانات...")
        if fix_database():
            print("✓ تم تحديث هيكل قاعدة البيانات بنجاح")
            if check_specific_columns():
                print("✓ جميع الحقول المطلوبة موجودة")
                return True
            else:
                print("⚠️ بعض الحقول لا تزال مفقودة")
                return False
        else:
            print("❌ فشل تحديث قاعدة البيانات")
            return False
            
    except Exception as e:
        print(f"❌ خطأ في تحديث قاعدة البيانات: {e}")
        return False



def remove_weak_tables():
    """إزالة الجداول الضعيفة التي لم تعد مستخدمة"""
    try:
        from database_fix import remove_weak_tables as remove_tables
        print("جاري إزالة الجداول الضعيفة...")
        return remove_tables()
    except Exception as e:
        print(f"❌ خطأ في إزالة الجداول الضعيفة: {e}")
        return False

def update_database_schema():
    """تحديث هيكل قاعدة البيانات بإضافة الحقول المفقودة"""
    try:
        # استيراد سكريبت الإصلاح وتشغيله
        from database_fix import fix_database, check_specific_columns
        
        print("جاري تحديث هيكل قاعدة البيانات...")
        if fix_database():
            print("✓ تم تحديث هيكل قاعدة البيانات بنجاح")
            if check_specific_columns():
                print("✓ جميع الحقول المطلوبة موجودة")
                return True
            else:
                print("⚠️ بعض الحقول لا تزال مفقودة")
                return False
        else:
            print("❌ فشل تحديث قاعدة البيانات")
            return False
            
    except Exception as e:
        print(f"❌ خطأ في تحديث قاعدة البيانات: {e}")
        return False

def sync_schedule_with_department(schedule_id, db_session):
    """دالة مساعدة لمزامنة جدول مع هيكل القسم مع الحفاظ على الموظفين الحاليين"""
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            print(f"❌ الجدول {schedule_id} غير موجود")
            return False
        
        department = db_session.query(Department).get(schedule.department_id)
        if not department:
            print(f"❌ القسم {schedule.department_id} غير موجود")
            return False
        
        # الحصول على هيكل القسم
        structure_rows = db_session.query(ScheduleStructureRow).filter_by(
            department_id=schedule.department_id
        ).order_by('row_order').all()
        
        print(f"🔍 هيكل القسم {department.name} يحتوي على {len(structure_rows)} صف")
        
        if not structure_rows:
            print(f"⚠️ لا يوجد هيكل للقسم {department.name}")
            return False
        
        # الحصول على التفاصيل القديمة للجدول
        old_details = db_session.query(ScheduleDetail).filter_by(
            weekly_schedule_id=schedule.id
        ).all()
        
        # إنشاء قاموس للتفاصيل القديمة للحفاظ على الموظفين الحاليين
        old_details_map = {}
        for detail in old_details:
            key = f"{detail.day_date}-{detail.job_title}"
            old_details_map[key] = {
                'morning_shift': detail.morning_shift,
                'evening_shift': detail.evening_shift,
                'night_shift': detail.night_shift,
                'row_order': detail.row_order
            }
        
        print(f"🔍 تم حفظ {len(old_details_map)} سجل قديم للمحافظة عليها")
        
        # حذف التفاصيل القديمة التي ليس لها وظائف في الهيكل الجديد
        deleted_count = 0
        for detail in old_details:
            # التحقق إذا كانت الوظيفة لا تزال موجودة في الهيكل الجديد
            job_exists = any(row.job_title == detail.job_title for row in structure_rows)
            if not job_exists:
                # إذا كانت الوظيفة غير موجودة في الهيكل الجديد، احذف التفاصيل
                db_session.delete(detail)
                deleted_count += 1
        
        print(f"🗑️ تم حذف {deleted_count} سجل قديم (وظائف غير موجودة في الهيكل الجديد)")
        
        # الحصول على دالة get_arabic_day_name بشكل صحيح
        try:
            # محاولة استيراد الدالة من السياق الحالي
            if 'get_arabic_day_name_from_date' in globals():
                get_day_name = get_arabic_day_name_from_date
            else:
                # استخدام الإصدار البديل
                def get_day_name(date_obj):
                    arabic_days = {
                        0: 'الإثنين',    # Monday
                        1: 'الثلاثاء',   # Tuesday
                        2: 'الأربعاء',   # Wednesday
                        3: 'الخميس',     # Thursday
                        4: 'الجمعة',     # Friday
                        5: 'السبت',      # Saturday
                        6: 'الأحد'       # Sunday
                    }
                    return arabic_days.get(date_obj.weekday(), 'غير معروف')
        except:
            # دالة بسيطة كبديل
            def get_day_name(date_obj):
                arabic_days = {
                    0: 'الإثنين',    # Monday
                    1: 'الثلاثاء',   # Tuesday
                    2: 'الأربعاء',   # Wednesday
                    3: 'الخميس',     # Thursday
                    4: 'الجمعة',     # Friday
                    5: 'السبت',      # Saturday
                    6: 'الأحد'       # Sunday
                }
                return arabic_days.get(date_obj.weekday(), 'غير معروف')
        
        # إنشاء التفاصيل الجديدة والحفاظ على القيم القديمة
        added_count = 0
        for day_offset in range(7):
            current_date = schedule.week_start_date + timedelta(days=day_offset)
            day_name = get_day_name(current_date)
            
            row_order = 1
            for structure_row in structure_rows:
                key = f"{current_date}-{structure_row.job_title}"
                
                # استخدام القيم القديمة إذا كانت موجودة، وإلا استخدام القيم الافتراضية من الهيكل
                if key in old_details_map:
                    old_data = old_details_map[key]
                    morning = old_data['morning_shift']
                    evening = old_data['evening_shift']
                    night = old_data['night_shift']
                    row_order = old_data['row_order']
                    print(f"✅ المحافظة على الموظفين الحاليين لـ {structure_row.job_title} في {current_date}")
                else:
                    # استخدام القيم الافتراضية من الهيكل
                    morning = structure_row.morning_shift or ""
                    evening = structure_row.evening_shift or ""
                    night = structure_row.night_shift or ""
                
                # التحقق من وجود التفصيلة بالفعل قبل الإضافة
                existing_detail = db_session.query(ScheduleDetail).filter_by(
                    weekly_schedule_id=schedule.id,
                    day_date=current_date,
                    job_title=structure_row.job_title
                ).first()
                
                if existing_detail:
                    # تحديث التفصيلة الموجودة
                    existing_detail.morning_shift = morning
                    existing_detail.evening_shift = evening
                    existing_detail.night_shift = night
                    existing_detail.day_name = day_name
                    existing_detail.row_order = row_order
                    existing_detail.is_custom = False
                    print(f"📝 تحديث سجل موجود لـ {structure_row.job_title} في {current_date}")
                else:
                    # إنشاء تفصيلة جديدة
                    detail = ScheduleDetail(
                        weekly_schedule_id=schedule.id,
                        day_date=current_date,
                        day_name=day_name,
                        job_title=structure_row.job_title,
                        morning_shift=morning,
                        evening_shift=evening,
                        night_shift=night,
                        row_order=row_order,
                        is_custom=False
                    )
                    db_session.add(detail)
                    added_count += 1
                
                row_order += 1
        
        print(f"✅ تم إضافة/تحديث {added_count} سجل جديد للجدول {schedule.id}")
        print(f"📅 الفترة: {schedule.week_start_date} إلى {schedule.week_end_date}")
        print(f"📊 الحساب: {len(structure_rows)} وظيفة × 7 أيام = {len(structure_rows) * 7} سجل")
        
        db_session.commit()
        return True
        
    except Exception as e:
        print(f"❌ خطأ في مزامنة الجدول {schedule_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        db_session.rollback()
        return False

@app.route('/debug/schedule/<int:schedule_id>')
@login_required
def debug_schedule_details(schedule_id):
    """فحص تفاصيل جدول معين"""
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    details = db_session.query(ScheduleDetail).filter_by(
        weekly_schedule_id=schedule_id
    ).order_by('day_date', 'row_order').all()
    
    return f"""
    <h1>فحص الجدول {schedule_id}</h1>
    <p>الفترة: {schedule.week_start_date} إلى {schedule.week_end_date}</p>
    <p>عدد التفاصيل: {len(details)}</p>
    <table border="1">
        <tr>
            <th>التاريخ</th>
            <th>اليوم</th>
            <th>الوظيفة</th>
            <th>الصباحي</th>
            <th>المسائي</th>
            <th>السهر</th>
        </tr>
        {"".join([
            f'<tr><td>{detail.day_date}</td><td>{detail.day_name}</td><td>{detail.job_title}</td><td>{detail.morning_shift}</td><td>{detail.evening_shift}</td><td>{detail.night_shift}</td></tr>'
            for detail in details
        ])}
    </table>
    """

@app.route('/debug/schedule_info/<int:schedule_id>')
@login_required
def debug_schedule_info(schedule_id):
    """فحص معلومات جدول معين"""
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    
    if not schedule:
        return "الجدول غير موجود"
    
    department = db_session.query(Department).get(schedule.department_id)
    details_count = db_session.query(ScheduleDetail).filter_by(
        weekly_schedule_id=schedule_id
    ).count()
    
    return f"""
    <h1>معلومات الجدول {schedule_id}</h1>
    <table border="1">
        <tr><td>المعرف</td><td>{schedule.id}</td></tr>
        <tr><td>القسم</td><td>{department.name if department else 'غير معين'} (ID: {schedule.department_id})</td></tr>
        <tr><td>الفترة</td><td>{schedule.week_start_date} إلى {schedule.week_end_date}</td></tr>
        <tr><td>معتمد</td><td>{'نعم' if schedule.is_approved else 'لا'}</td></tr>
        <tr><td>مقفل</td><td>{'نعم' if schedule.is_locked else 'لا'}</td></tr>
        <tr><td>قالب</td><td>{'نعم' if schedule.is_template else 'لا'}</td></tr>
        <tr><td>عدد التفاصيل</td><td>{details_count}</td></tr>
        <tr><td>أنشئ بواسطة</td><td>{db_session.query(User).get(schedule.created_by).name if schedule.created_by else 'غير معروف'}</td></tr>
    </table>
    """

def create_default_admin():
    admin = db_session.query(User).filter_by(username='admin').first()
    if not admin:
        admin_user = User(
            username='admin',
            name='المسؤول العام',
            password_hash=generate_password_hash('admin123'),
            is_admin=True,
            is_manager=True
        )
        db_session.add(admin_user)
        
        # Create default department
        default_dept = Department(
            name='الإدارة العامة',
            created_by=1
        )
        db_session.add(default_dept)
        db_session.commit()
        
        # Update admin department
        admin_user.department_id = default_dept.id
        default_dept.primary_manager_id = admin_user.id
        db_session.commit()
        
        print('تم إنشاء المستخدم المسؤول الافتراضي: admin / admin123')

# ======== Basic Routes ========

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = db_session.query(User).filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            user.last_login = datetime.now()
            db_session.commit()
            
            if user.is_admin:
                return redirect(url_for('admin_dashboard'))
            elif user.is_manager:
                return redirect(url_for('manager_dashboard'))
            else:
                return redirect(url_for('user_dashboard'))
        else:
            flash('اسم المستخدم أو كلمة المرور غير صحيحة')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ======== Admin Routes ========

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    users_count = db_session.query(User).count()
    departments_count = db_session.query(Department).count()
    pending_leaves = db_session.query(LeaveRequest).filter_by(status='pending').count()
    pending_permissions = db_session.query(PermissionRequest).filter_by(status='pending').count()
    pending_schedules = db_session.query(WeeklySchedule).filter_by(is_approved=False).count()
    
    notifications = get_user_notifications(current_user.id)
    notifications_count = len(notifications)  # Add this line
    
    return render_template('admin_dashboard.html',
                         users_count=users_count,
                         departments_count=departments_count,
                         pending_leaves=pending_leaves,
                         pending_permissions=pending_permissions,
                         pending_schedules=pending_schedules,
                         notifications=notifications,
                         notifications_count=notifications_count)  # Add this
@app.route('/admin/users')
@login_required
def admin_users():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    users = db_session.query(User).all()
    departments = db_session.query(Department).all()
    
    for user in users:
        user.employee_data = db_session.query(EmployeeData).filter_by(user_id=user.id).first()
        if user.employee_data:
            user.employee_data.calculate_completion()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_users.html', 
                         users=users, 
                         departments=departments,
                         notifications=notifications)


@app.route('/admin/export_all_users')
@login_required
def export_all_users():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        # Get all non-admin users with department information
        users = db_session.query(User).filter_by(is_admin=False).all()
        
        if not users:
            flash('لا يوجد مستخدمين للتصدير')
            return redirect(url_for('admin_users'))
        
        # Preload departments for better performance
        departments = {dept.id: dept for dept in db_session.query(Department).all()}
        
        # Create a list to hold all user data
        all_users_data = []
        
        for user in users:
            employee_data = db_session.query(EmployeeData).filter_by(user_id=user.id).first()
            
            # Get department name safely
            department_name = 'غير معين'
            if user.department_id and user.department_id in departments:
                department_name = departments[user.department_id].name
            
            if employee_data:
                user_dict = {
                    'اسم المستخدم': user.username,
                    'الاسم الكامل': user.name,
                    'الصفة': 'مدير' if user.is_manager else 'موظف',
                    'القسم': department_name,
                    'الاسم بالعربية': employee_data.arabic_name or '',
                    'الاسم بالإنجليزية': employee_data.english_name or '',
                    'الرقم القومي': employee_data.national_id or '',
                    'تاريخ إصدار البطاقة': employee_data.id_issue_date.strftime('%Y-%m-%d') if employee_data.id_issue_date else '',
                    'تاريخ الميلاد': employee_data.birth_date.strftime('%Y-%m-%d') if employee_data.birth_date else '',
                    'السن': employee_data.age or '',
                    'واتساب': employee_data.whatsapp or '',
                    'الهاتف': employee_data.phone or '',
                    'العنوان': employee_data.address or '',
                    'الموقف من التجنيد': employee_data.military_status or '',
                    'الحالة الاجتماعية': employee_data.marital_status or '',
                    'المؤهل الدراسي': employee_data.qualification or '',
                    'سنة التخرج': employee_data.graduation_year or '',
                    'التقدير': employee_data.grade or '',
                    'يعمل حالياً': 'نعم' if employee_data.has_work else 'لا',
                    'جهة العمل': employee_data.workplace or '',
                    'الوظيفة': employee_data.job_title or '',
                    'الرقم التأميني': employee_data.insurance_number or '',
                    'الرقم الضريبي': employee_data.tax_number or '',
                    'ترخيص المهنة': employee_data.profession_license or '',
                    'كارنية النقابة': employee_data.union_card or '',
                    'جهة الاتصال 1 - الاسم': employee_data.emergency1_name or '',
                    'جهة الاتصال 1 - التليفون': employee_data.emergency1_phone or '',
                    'جهة الاتصال 1 - العنوان': employee_data.emergency1_address or '',
                    'جهة الاتصال 1 - صلة القرابة': employee_data.emergency1_relation or '',
                    'جهة الاتصال 2 - الاسم': employee_data.emergency2_name or '',
                    'جهة الاتصال 2 - التليفون': employee_data.emergency2_phone or '',
                    'جهة الاتصال 2 - العنوان': employee_data.emergency2_address or '',
                    'جهة الاتصال 2 - صلة القرابة': employee_data.emergency2_relation or '',
                    'نسبة الإكتمال': f'{employee_data.completion_percentage}%',
                    'آخر تحديث': employee_data.last_updated.strftime('%Y-%m-%d %H:%M') if employee_data.last_updated else '',
                    'تم التحديث بواسطة': employee_data.updated_by or ''
                }
            else:
                user_dict = {
                    'اسم المستخدم': user.username,
                    'الاسم الكامل': user.name,
                    'الصفة': 'مدير' if user.is_manager else 'موظف',
                    'القسم': department_name,
                    'الاسم بالعربية': '',
                    'الاسم بالإنجليزية': '',
                    'الرقم القومي': '',
                    'تاريخ إصدار البطاقة': '',
                    'تاريخ الميلاد': '',
                    'السن': '',
                    'واتساب': '',
                    'الهاتف': '',
                    'العنوان': '',
                    'الموقف من التجنيد': '',
                    'الحالة الاجتماعية': '',
                    'المؤهل الدراسي': '',
                    'سنة التخرج': '',
                    'التقدير': '',
                    'يعمل حالياً': 'لا',
                    'جهة العمل': '',
                    'الوظيفة': '',
                    'الرقم التأميني': '',
                    'الرقم الضريبي': '',
                    'ترخيص المهنة': '',
                    'كارنية النقابة': '',
                    'جهة الاتصال 1 - الاسم': '',
                    'جهة الاتصال 1 - التليفون': '',
                    'جهة الاتصال 1 - العنوان': '',
                    'جهة الاتصال 1 - صلة القرابة': '',
                    'جهة الاتصال 2 - الاسم': '',
                    'جهة الاتصال 2 - التليفون': '',
                    'جهة الاتصال 2 - العنوان': '',
                    'جهة الاتصال 2 - صلة القرابة': '',
                    'نسبة الإكتمال': '0%',
                    'آخر تحديث': '',
                    'تم التحديث بواسطة': ''
                }
            
            all_users_data.append(user_dict)
        
        # Create DataFrame
        df = pd.DataFrame(all_users_data)
        
        # Create export folder if not exists
        export_folder = os.path.join(app.config['EXPORT_FOLDER'], 'all_users')
        os.makedirs(export_folder, exist_ok=True)
        
        # Generate filename
        filename = f"all_users_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(export_folder, filename)
        
        # Save to Excel
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        # Log the action
        print(f'تم تصدير بيانات {len(users)} مستخدم بواسطة {current_user.name}')
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        flash(f'حدث خطأ أثناء التصدير: {str(e)}')
        return redirect(url_for('admin_users'))
    
@app.route('/admin/create_user', methods=['GET', 'POST'])
@login_required
def create_user():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    departments = db_session.query(Department).all()
    
    if request.method == 'POST':
        username = request.form['username']
        name = request.form['name']
        password = request.form['password']
        is_manager = 'is_manager' in request.form
        department_id = request.form.get('department_id')
        
        if db_session.query(User).filter_by(username=username).first():
            flash('اسم المستخدم موجود بالفعل')
            return redirect(url_for('create_user'))
        
        user = User(
            username=username,
            name=name,
            password_hash=generate_password_hash(password),
            is_admin=False,
            is_manager=is_manager,
            department_id=department_id
        )
        
        db_session.add(user)
        db_session.commit()
        
        # Create employee balance
        balance = EmployeeBalance(user_id=user.id)
        db_session.add(balance)
        db_session.commit()
        
        create_notification(
            current_user.id,
            'تم إنشاء مستخدم جديد',
            f'تم إنشاء المستخدم {name} بنجاح',
            'user_created'
        )
        
        flash('تم إنشاء المستخدم بنجاح')
        return redirect(url_for('admin_users'))
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_create_user.html', 
                         departments=departments,
                         notifications=notifications)






@app.route('/admin/users/<int:user_id>/edit_complete', methods=['GET', 'POST'])
@login_required
def admin_edit_user_complete(user_id):
    """
    Complete user data editing - for admin only
    Allows editing all user fields including personal information
    """
    # Get the user to edit
    user = db_session.query(User).filter_by(id=user_id).first()
    if not user:
        flash('المستخدم غير موجود', 'error')
        return redirect(url_for('admin_users'))
    
    # Get employee data or create new if doesn't exist
    employee_data = db_session.query(EmployeeData).filter_by(user_id=user_id).first()
    if not employee_data:
        employee_data = EmployeeData(user_id=user_id)
        db_session.add(employee_data)
        db_session.commit()
    
    # Get departments for dropdown
    departments = db_session.query(Department).all()
    
    if request.method == 'POST':
        try:
            # Update basic user information
            user.username = request.form.get('username', user.username)
            user.name = request.form.get('name', user.name)
            user.department_id = request.form.get('department_id') or None
            user.is_manager = 'is_manager' in request.form
            user.is_active = 'is_active' in request.form
            
            # Update password only if provided
            new_password = request.form.get('password', '').strip()
            if new_password:
                if len(new_password) < 6:
                    flash('كلمة المرور يجب أن تكون至少 6 أحرف', 'error')
                    return redirect(url_for('admin_edit_user_complete', user_id=user_id))
                user.password_hash = generate_password_hash(new_password)
            
            # Update employee data
            employee_data.arabic_name = request.form.get('arabic_name') or None
            employee_data.english_name = request.form.get('english_name') or None
            employee_data.national_id = request.form.get('national_id') or None
            
            # Handle dates
            if request.form.get('id_issue_date'):
                employee_data.id_issue_date = datetime.strptime(request.form['id_issue_date'], '%Y-%m-%d').date()
            else:
                employee_data.id_issue_date = None
            
            if request.form.get('birth_date'):
                employee_data.birth_date = datetime.strptime(request.form['birth_date'], '%Y-%m-%d').date()
                employee_data.age = calculate_age(employee_data.birth_date)
            else:
                employee_data.birth_date = None
                employee_data.age = None
            
            # Contact information
            employee_data.whatsapp = request.form.get('whatsapp') or None
            employee_data.phone = request.form.get('phone') or None
            employee_data.address = request.form.get('address') or None
            
            # Personal information
            employee_data.military_status = request.form.get('military_status') or None
            employee_data.marital_status = request.form.get('marital_status') or None
            employee_data.qualification = request.form.get('qualification') or None
            
            if request.form.get('graduation_year'):
                employee_data.graduation_year = int(request.form['graduation_year'])
            else:
                employee_data.graduation_year = None
            
            employee_data.grade = request.form.get('grade') or None
            employee_data.has_work = 'has_work' in request.form
            employee_data.workplace = request.form.get('workplace') or None
            employee_data.job_title = request.form.get('job_title') or None
            
            # Professional information
            employee_data.insurance_number = request.form.get('insurance_number') or None
            employee_data.tax_number = request.form.get('tax_number') or None
            employee_data.profession_license = request.form.get('profession_license') or None
            employee_data.union_card = request.form.get('union_card') or None
            
            # Emergency contacts
            employee_data.emergency1_name = request.form.get('emergency1_name') or None
            employee_data.emergency1_phone = request.form.get('emergency1_phone') or None
            employee_data.emergency1_address = request.form.get('emergency1_address') or None
            employee_data.emergency1_relation = request.form.get('emergency1_relation') or None
            employee_data.emergency2_name = request.form.get('emergency2_name') or None
            employee_data.emergency2_phone = request.form.get('emergency2_phone') or None
            employee_data.emergency2_address = request.form.get('emergency2_address') or None
            employee_data.emergency2_relation = request.form.get('emergency2_relation') or None
            
            # Handle file uploads
            user_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], str(user_id))
            os.makedirs(user_upload_folder, exist_ok=True)
            
            file_fields = [
                ('national_id_image', 'national_id_image'),
                ('military_status_image', 'military_status_image'),
                ('qualification_image', 'qualification_image'),
                ('salary_details', 'salary_details'),
                ('employment_status', 'employment_status')
            ]
            
            for form_field, db_field in file_fields:
                file = request.files.get(form_field)
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    file_path = os.path.join(user_upload_folder, filename)
                    file.save(file_path)
                    setattr(employee_data, db_field, filename)
            
            # Update timestamps
            employee_data.last_updated = datetime.now()
            employee_data.updated_by = current_user.name
            
            # Calculate completion percentage
            employee_data.calculate_completion()
            
            db_session.commit()
            
            # Create notification
            create_notification(
                current_user.id,
                'تم تحديث بيانات المستخدم',
                f'تم تحديث البيانات الكاملة للمستخدم {user.name}',
                'user_data_updated'
            )
            
            flash('تم تحديث بيانات المستخدم بنجاح', 'success')
            return redirect(url_for('admin_users'))
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ أثناء تحديث البيانات: {str(e)}', 'error')
            print(f"Error updating user data: {str(e)}")
            return redirect(url_for('admin_edit_user_complete', user_id=user_id))
    
    # GET request - display the form
    notifications = get_user_notifications(current_user.id)
    
    # Calculate completion percentage for display
    completion_percentage = employee_data.calculate_completion()
    missing_fields = employee_data.get_missing_fields()
    
    return render_template('admin/admin_edit_user_complete.html', 
                         user=user,
                         employee_data=employee_data,
                         departments=departments,
                         completion_percentage=completion_percentage,
                         missing_fields=missing_fields,
                         notifications=notifications)






@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    """
    Edit user information - only allows password updates
    All other fields are read-only as per the HTML template
    """
    
    # Get the user to edit
    user = db_session.query(User).filter_by(id=user_id).first()
    if not user:
        flash('المستخدم غير موجود', 'error')
        return redirect(url_for('admin_users'))
    
    # Get departments for display (read-only in template)
    departments = db_session.query(Department).all()
    
    if request.method == 'POST':
        try:
            # Only process password update if provided
            new_password = request.form.get('password', '').strip()
            
            # Update password only if not empty
            if new_password:
                # Validate password strength (optional)
                if len(new_password) < 6:
                    flash('كلمة المرور يجب أن تكون至少 6 أحرف', 'error')
                    return redirect(url_for('edit_user', user_id=user_id))
                
                user.password_hash = generate_password_hash(new_password)
                db_session.commit()
                
                # Create notification for password change
                create_notification(
                    current_user.id,
                    'تم تحديث كلمة المرور',
                    f'تم تحديث كلمة مرور المستخدم {user.name} بنجاح',
                    'password_updated'
                )
                
                flash('تم تحديث كلمة المرور بنجاح', 'success')
            else:
                flash('لم يتم إجراء أي تغييرات', 'info')
            
            return redirect(url_for('admin_users'))
            
        except Exception as e:
            db_session.rollback()
            flash('حدث خطأ أثناء تحديث البيانات', 'error')
            # Log the error for debugging
            print(f"Error updating user: {str(e)}")
            return redirect(url_for('edit_user', user_id=user_id))
    
    # GET request - display the form
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_edit_user.html', 
                         user=user,
                         departments=departments,
                         notifications=notifications)


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
def delete_user(user_id):
    """Delete a user and all related records from the database"""
    
    # Authorization checks
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    # Get user
    user = db_session.query(User).filter_by(id=user_id).first()
    if not user:
        flash('المستخدم غير موجود', 'error')
        return redirect(url_for('admin_users'))
    
    # Prevent self-deletion
    if user.id == current_user.id:
        flash('لا يمكنك ح delete حسابك الخاص', 'warning')
        return redirect(url_for('admin_users'))
    
    # Store user info for logging
    user_name = user.name
    user_email = user.email
    
    try:
        # Begin transaction
        db_session.begin_nested()
        
        # 1. Handle message_recipients - FIXED: Use MessageRecipient (not MessageRecipients)
        try:
            message_recipients = db_session.query(MessageRecipient).filter_by(user_id=user_id).all()
            for recipient in message_recipients:
                db_session.delete(recipient)
        except Exception as e:
            app.logger.warning(f"Error deleting message recipients: {e}")
        
        # 2. Handle messages sent by the user
        try:
            user_messages = db_session.query(Message).filter_by(sender_id=user_id).all()
            for message in user_messages:
                db_session.delete(message)
        except Exception as e:
            app.logger.warning(f"Error deleting messages: {e}")
        
        # 3. Handle notifications
        try:
            notifications = db_session.query(Notification).filter_by(user_id=user_id).all()
            for notification in notifications:
                db_session.delete(notification)
        except Exception as e:
            app.logger.warning(f"Error deleting notifications: {e}")
        
        # 4. Handle employee data
        try:
            employee_data = db_session.query(EmployeeData).filter_by(user_id=user_id).all()
            for data in employee_data:
                db_session.delete(data)
        except Exception as e:
            app.logger.warning(f"Error deleting employee data: {e}")
        
        # 5. Handle employee balance
        try:
            employee_balances = db_session.query(EmployeeBalance).filter_by(user_id=user_id).all()
            for balance in employee_balances:
                db_session.delete(balance)
        except Exception as e:
            app.logger.warning(f"Error deleting employee balance: {e}")
        
        # 6. Handle salary slips
        try:
            salary_slips = db_session.query(SalarySlip).filter_by(user_id=user_id).all()
            for slip in salary_slips:
                # Delete physical file if it exists
                if slip.file_path and os.path.exists(slip.file_path):
                    try:
                        os.remove(slip.file_path)
                    except:
                        pass
                db_session.delete(slip)
        except Exception as e:
            app.logger.warning(f"Error deleting salary slips: {e}")
        
        # 7. Handle attendance records - FIXED: Check if Attendance model exists
        try:
            # Check if Attendance model is defined in the current scope
            if 'Attendance' in globals() or 'Attendance' in locals():
                attendance_records = db_session.query(Attendance).filter_by(user_id=user_id).all()
                for record in attendance_records:
                    db_session.delete(record)
        except Exception as e:
            app.logger.warning(f"Error deleting attendance records: {e}")
        
        # 8. Handle leave requests
        try:
            leave_requests = db_session.query(LeaveRequest).filter_by(user_id=user_id).all()
            for leave in leave_requests:
                db_session.delete(leave)
        except Exception as e:
            app.logger.warning(f"Error deleting leave requests: {e}")
        
        # 9. Handle permission requests
        try:
            permission_requests = db_session.query(PermissionRequest).filter_by(user_id=user_id).all()
            for permission in permission_requests:
                db_session.delete(permission)
        except Exception as e:
            app.logger.warning(f"Error deleting permission requests: {e}")
        
        # 10. Handle advance requests
        try:
            advance_requests = db_session.query(AdvanceRequest).filter_by(user_id=user_id).all()
            for advance in advance_requests:
                db_session.delete(advance)
        except Exception as e:
            app.logger.warning(f"Error deleting advance requests: {e}")
        
        # 11. Handle reward penalties
        try:
            rewards_penalties = db_session.query(RewardPenalty).filter_by(user_id=user_id).all()
            for rp in rewards_penalties:
                db_session.delete(rp)
        except Exception as e:
            app.logger.warning(f"Error deleting reward penalties: {e}")
        
        # 12. Handle schedule details (if user appears in schedules)
        try:
            schedule_details = db_session.query(ScheduleDetail).filter(
                (ScheduleDetail.morning_shift.like(f'%{user_name}%')) |
                (ScheduleDetail.evening_shift.like(f'%{user_name}%')) |
                (ScheduleDetail.night_shift.like(f'%{user_name}%'))
            ).all()
            
            for detail in schedule_details:
                # Clear the user from shifts instead of deleting the whole detail
                if detail.morning_shift and user_name in detail.morning_shift:
                    detail.morning_shift = detail.morning_shift.replace(user_name, '').strip()
                if detail.evening_shift and user_name in detail.evening_shift:
                    detail.evening_shift = detail.evening_shift.replace(user_name, '').strip()
                if detail.night_shift and user_name in detail.night_shift:
                    detail.night_shift = detail.night_shift.replace(user_name, '').strip()
        except Exception as e:
            app.logger.warning(f"Error updating schedule details: {e}")
        
        # 13. Handle department manager records
        try:
            dept_managers = db_session.query(DepartmentManager).filter_by(user_id=user_id).all()
            for dm in dept_managers:
                db_session.delete(dm)
        except Exception as e:
            app.logger.warning(f"Error deleting department manager records: {e}")
        
        # 14. Finally, delete the user
        db_session.delete(user)
        
        # Commit all changes
        db_session.commit()
        
        # Log the deletion
        app.logger.info(f"User deleted - ID: {user_id}, Name: {user_name}, Email: {user_email}, Deleted by: {current_user.id}")
        
        # Create notification for admin
        try:
            create_notification(
                current_user.id,
                'تم حذف المستخدم',
                f'تم حذف المستخدم {user_name} بنجاح',
                'user_deleted'
            )
        except Exception as notif_error:
            app.logger.warning(f"Could not create notification: {notif_error}")
        
        flash(f'تم حذف المستخدم {user_name} بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        app.logger.error(f"Error deleting user {user_id}: {str(e)}")
        flash('حدث خطأ أثناء حذف المستخدم. يرجى المحاولة مرة أخرى.', 'error')
        print(f"Error deleting user: {e}")
        # Print full traceback for debugging
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('admin_users'))

def get_all_user_related_records(user_id):
    """Get all database records related to a user"""
    related_models = {
        'message_recipients': MessageRecipients,
        'messages_sent': Message,
        'notifications': Notification,
        'employee_data': EmployeeData,
        'employee_balances': EmployeeBalance,
        'salary_slips': SalarySlip,
        'attendance': Attendance,
        'leave_requests': LeaveRequest,
        # Add any other related models here
    }
    
    records = {}
    for name, model in related_models.items():
        if hasattr(model, 'user_id'):
            records[name] = db_session.query(model).filter_by(user_id=user_id).count()
        elif name == 'messages_sent' and hasattr(model, 'sender_id'):
            records[name] = db_session.query(model).filter_by(sender_id=user_id).count()
    
    return records

# You can call this before deletion to show summary
@app.route('/admin/users/<int:user_id>/delete_preview', methods=['GET'])
@login_required
def delete_user_preview(user_id):
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    user = db_session.query(User).filter_by(id=user_id).first()
    if not user:
        flash('المستخدم غير موجود', 'error')
        return redirect(url_for('admin_users'))
    
    related_records = get_all_user_related_records(user_id)
    total_records = sum(related_records.values())
    
    return render_template('admin/delete_user_preview.html', 
                         user=user, 
                         related_records=related_records,
                         total_records=total_records)


@app.route('/admin/users/<int:user_id>/toggle_status', methods=['POST'])
@login_required
def toggle_user_status(user_id):
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    user = db_session.query(User).filter_by(id=user_id).first()
    if not user:
        flash('المستخدم غير موجود')
        return redirect(url_for('admin_users'))
    
    # Prevent deactivating yourself
    if user.id == current_user.id:
        flash('لا يمكنك تعطيل حسابك الخاص')
        return redirect(url_for('admin_users'))
    
    user.is_active = not user.is_active
    db_session.commit()
    
    status_text = "مفعل" if user.is_active else "معطل"
    create_notification(
        current_user.id,
        'تم تغيير حالة المستخدم',
        f'تم تغيير حالة المستخدم {user.name} إلى {status_text}',
        'user_status_changed'
    )
    
    flash(f'تم تغيير حالة المستخدم إلى {status_text}')
    return redirect(url_for('admin_users'))

@app.route('/admin/assign_manager', methods=['POST'])
@login_required
def assign_department_manager():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    department_id = request.form['department_id']
    manager_id = request.form['manager_id']
    
    department = db_session.query(Department).get(department_id)
    new_manager = db_session.query(User).get(manager_id)
    
    if department and new_manager:
        # Remove manager role from previous manager if exists
        if department.manager_id:
            old_manager = db_session.query(User).get(department.manager_id)
            if old_manager:
                # Check if old manager is manager of any other department
                other_departments = db_session.query(Department).filter(
                    Department.manager_id == old_manager.id,
                    Department.id != department.id
                ).count()
                if other_departments == 0:
                    old_manager.is_manager = False
        
        # Assign new manager
        department.manager_id = manager_id
        new_manager.is_manager = True
        new_manager.department_id = department_id
        
        db_session.commit()
        
        create_notification(
            new_manager.id,
            'تم تعيينك كمدير قسم',
            f'تم تعيينك كمدير للقسم {department.name}',
            'manager_assigned'
        )
        
        flash(f'تم تعيين {new_manager.name} كمدير للقسم {department.name} بنجاح')
    else:
        flash('حدث خطأ في تعيين المدير')
    
    return redirect(url_for('admin_departments'))


def get_arabic_day_name(day_index):
    """الحصول على اسم اليوم بالعربية"""
    # day_index يجب أن يكون من weekday() في Python
    # weekday(): Monday=0, Tuesday=1, Wednesday=2, Thursday=3, Friday=4, Saturday=5, Sunday=6
    
    arabic_days = {
        0: 'الإثنين',    # Monday
        1: 'الثلاثاء',   # Tuesday
        2: 'الأربعاء',   # Wednesday
        3: 'الخميس',     # Thursday
        4: 'الجمعة',     # Friday
        0: 'السبت',      # Saturday
        1: 'الأحد'       # Sunday
    }
    
    return arabic_days.get(day_index, f'يوم {day_index}')



@app.route('/api/department_employeess')
@login_required
def api_department_employeess():
    """الحصول على موظفي القسم الحالي"""
    try:
        dept_id = request.args.get('dept_id')
        if not dept_id:
            # إذا لم يتم تمرير dept_id، استخدم قسم المستخدم الحالي
            dept_id = current_user.department_id
        
        if not dept_id:
            return jsonify({
                'success': False,
                'message': 'لا يوجد قسم محدد'
            })
        
        # جلب جميع موظفي القسم (غير المسؤولين)
        employees = db_session.query(User).filter(
            User.department_id == dept_id,
            User.is_admin == False,
            User.is_active == True
        ).all()
        
        # تحضير البيانات
        employees_data = []
        for emp in employees:
            # الحصول على بيانات الموظف
            employee_data = db_session.query(EmployeeData).filter_by(user_id=emp.id).first()
            
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'username': emp.username,
                'position': employee_data.job_title if employee_data else 'موظف',
                'full_name': employee_data.arabic_name if employee_data else emp.name
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        print(f"خطأ في جلب موظفي القسم: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'خطأ في جلب البيانات: {str(e)}'
        })


@app.route('/admin/departments', methods=['GET', 'POST'])
@login_required
def admin_departments():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    # Handle POST request for creating department
    if request.method == 'POST':
        try:
            name = request.form['name']
            primary_manager_id = request.form.get('primary_manager_id') or None
            
            # Handle empty values for numeric fields
            max_advance_amount_str = request.form.get('max_advance_amount', '0')
            max_installments_str = request.form.get('max_installments', '1')
            
            max_advance_amount = float(max_advance_amount_str) if max_advance_amount_str else 0.0
            max_installments = int(max_installments_str) if max_installments_str else 1
            
            department = Department(
                name=name,
                primary_manager_id=primary_manager_id,
                advance_policy_max_amount=max_advance_amount,
                advance_policy_max_installments=max_installments,
                created_by=current_user.id
            )
            
            db_session.add(department)
            db_session.commit()
            
            flash('تم إنشاء القسم بنجاح', 'success')
            return redirect(url_for('admin_departments'))
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'error')
            return redirect(url_for('admin_departments'))
    
    # For GET requests, show the page
    try:
        # Get all departments
        departments = db_session.query(Department).all()
        
        # Get all users
        users = db_session.query(User).all()
        
        # Get all managers (users with is_manager=True)
        managers = db_session.query(User).filter_by(is_manager=True).all()
        
        # Prepare department data with managers
        departments_data = []
        for department in departments:
            # Create a dictionary with department info
            dept_data = {
                'id': department.id,
                'name': department.name,
                'primary_manager_id': department.primary_manager_id,
                'employee_count': db_session.query(User).filter_by(
                    department_id=department.id,
                    is_admin=False
                ).count(),
                'managers': []
            }
            
            # Get managers for this department
            dept_managers = db_session.query(DepartmentManager).filter_by(
                department_id=department.id
            ).all()
            
            # Add user info for each manager
            for dept_manager in dept_managers:
                user = db_session.get(User, dept_manager.user_id)  # Use get() instead of query().get()
                if user:
                    dept_data['managers'].append({
                        'manager_id': dept_manager.id,
                        'user_id': dept_manager.user_id,
                        'user_name': user.name,
                        'user_username': user.username,
                        'can_manage_schedules': dept_manager.can_manage_schedules,
                        'can_manage_leaves': dept_manager.can_manage_leaves,
                        'can_manage_permissions': dept_manager.can_manage_permissions,
                        'can_manage_advances': dept_manager.can_manage_advances,
                        'can_manage_rewards': dept_manager.can_manage_rewards,
                        'can_view_reports': dept_manager.can_view_reports,
                    })
            
            departments_data.append(dept_data)
        
        notifications = get_user_notifications(current_user.id)
        
        return render_template('admin/admin_departments.html',
                             departments=departments_data,  # Pass prepared data instead of ORM objects
                             users=users,
                             managers=managers,
                             notifications=notifications)
                             
    except Exception as e:
        flash(f'حدث خطأ في تحميل البيانات: {str(e)}', 'error')
        import traceback
        print(f"Error details: {traceback.format_exc()}")
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/add_manager_to_department', methods=['POST'])
@login_required
def admin_add_manager_to_department():
    """Add a manager to a department with specific permissions"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        department_id = request.form['department_id']
        user_id = request.form['user_id']
        
        # Check if user is already a manager for this department
        existing_manager = db_session.query(DepartmentManager).filter_by(
            department_id=department_id,
            user_id=user_id
        ).first()
        
        if existing_manager:
            flash('هذا المستخدم مدير بالفعل لهذا القسم', 'warning')
            return redirect(url_for('admin_departments'))
        
        # Get permissions from form
        can_manage_schedules = 'can_manage_schedules' in request.form
        can_manage_leaves = 'can_manage_leaves' in request.form
        can_manage_permissions = 'can_manage_permissions' in request.form
        can_manage_advances = 'can_manage_advances' in request.form
        can_manage_rewards = 'can_manage_rewards' in request.form
        can_view_reports = 'can_view_reports' in request.form
        
        # Check if should be primary manager
        if 'is_primary_manager' in request.form:
            department = db_session.get(Department, department_id)  # Use get() instead of query().get()
            if department:
                department.primary_manager_id = user_id
        
        # Create department manager record
        department_manager = DepartmentManager(
            department_id=department_id,
            user_id=user_id,
            can_manage_schedules=can_manage_schedules,
            can_manage_leaves=can_manage_leaves,
            can_manage_permissions=can_manage_permissions,
            can_manage_advances=can_manage_advances,
            can_manage_rewards=can_manage_rewards,
            can_view_reports=can_view_reports,
            created_by=current_user.id
        )
        
        db_session.add(department_manager)
        
        # Update user to be a manager if not already
        user = db_session.get(User, user_id)  # Use get() instead of query().get()
        if user and not user.is_manager:
            user.is_manager = True
        
        db_session.commit()
        
        flash('تم إضافة المدير بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'error')
        import traceback
        print(f"Error in add_manager_to_department: {traceback.format_exc()}")
    
    return redirect(url_for('admin_departments'))

@app.route('/admin/remove_department_manager/<int:manager_id>')
@login_required
def admin_remove_department_manager(manager_id):
    """Remove a manager from a department"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        manager = db_session.get(DepartmentManager, manager_id)  # Use get() instead of query().get()
        if not manager:
            flash('المدير غير موجود', 'error')
            return redirect(url_for('admin_departments'))
        
        # Check if this manager is the primary manager
        department = db_session.get(Department, manager.department_id)
        if department and department.primary_manager_id == manager.user_id:
            department.primary_manager_id = None
        
        # Remove the manager record
        db_session.delete(manager)
        
        # Check if user is manager of any other department
        other_managements = db_session.query(DepartmentManager).filter_by(
            user_id=manager.user_id
        ).count()
        
        # If not managing any other department, remove manager role
        if other_managements == 0:
            user = db_session.get(User, manager.user_id)
            if user:
                user.is_manager = False
        
        db_session.commit()
        flash('تم حذف المدير بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('admin_departments'))

@app.route('/admin/set_primary_manager/<int:department_id>', methods=['POST'])
@login_required
def admin_set_primary_manager(department_id):
    """Set or remove primary manager for a department"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        manager_id = request.form.get('manager_id')
        department = db_session.get(Department, department_id)  # Use get() instead of query().get()
        
        if not department:
            flash('القسم غير موجود', 'error')
            return redirect(url_for('admin_departments'))
        
        if manager_id:
            # Verify manager exists in this department
            manager = db_session.query(DepartmentManager).filter_by(
                department_id=department_id,
                user_id=manager_id
            ).first()
            
            if not manager:
                flash('المدير غير موجود في هذا القسم', 'error')
                return redirect(url_for('admin_departments'))
            
            department.primary_manager_id = manager_id
        else:
            department.primary_manager_id = None
        
        db_session.commit()
        flash('تم تحديث المدير الرئيسي بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('admin_departments'))

def get_department_managers_with_users(department_id):
    """Get all managers for a department with their user info"""
    return db_session.query(
        DepartmentManager, User
    ).join(
        User, DepartmentManager.user_id == User.id
    ).filter(
        DepartmentManager.department_id == department_id
    ).all()

@app.route('/admin/employee_progress')
@login_required
def admin_employee_progress():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    users = db_session.query(User).filter_by(is_admin=False).all()
    employee_data_list = []
    
    for user in users:
        data = db_session.query(EmployeeData).filter_by(user_id=user.id).first()
        if data:
            # استخدم الدوال من النموذج مباشرة
            data.calculate_completion()
            employee_data_list.append({
                'user': user,
                'data': data,
                'missing_fields': data.get_missing_fields()  # استدعاء الدالة من النموذج
            })
        else:
            # إذا لم يكن هناك بيانات للموظف، أضفه كـ "لم يبدأ"
            employee_data_list.append({
                'user': user,
                'data': None,
                'missing_fields': []
            })
    
    db_session.commit()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_employee_progress.html',
                         employee_data_list=employee_data_list,
                         notifications=notifications)

@app.route('/admin/request_data_update/<int:user_id>')
@login_required
def request_data_update(user_id):
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    employee_data = db_session.query(EmployeeData).filter_by(user_id=user_id).first()
    if employee_data:
        employee_data.needs_update = True
        db_session.commit()
        
        create_notification(
            user_id,
            'مطلوب تحديث البيانات',
            'يرجى تحديث بياناتك الشخصية',
            'data_update_requested',
            action_url=url_for('user_form')
        )
        
        flash('تم إرسال طلب تحديث البيانات')
    
    return redirect(url_for('admin_employee_progress'))



from sqlalchemy.orm import joinedload


@app.teardown_appcontext
def shutdown_session(exception=None):
    db_session.remove()

@app.before_request
def create_session():
    # التأكد من وجود جلسة قاعدة بيانات
    pass

from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from io import BytesIO


@app.route('/user/dashboard')
@login_required
def user_dashboard():
    if current_user.is_admin:
        return redirect(url_for('admin_dashboard'))
    elif current_user.is_manager:
        return redirect(url_for('manager_dashboard'))
    
    employee_data = db_session.query(EmployeeData).filter_by(user_id=current_user.id).first()
    
    if employee_data:
        completion_percentage = employee_data.calculate_completion()
        missing_fields = employee_data.get_missing_fields()
        db_session.commit()
    else:
        completion_percentage = 0
        missing_fields = []
    
    # Check for new salary slips
    new_salary_slips = db_session.query(SalarySlip).filter_by(
        user_id=current_user.id, is_viewed=False
    ).count()
    
    # Check schedule notifications
    current_week_start = date.today() - timedelta(days=date.today().weekday())
    has_current_schedule = db_session.query(WeeklySchedule).filter_by(
        department_id=current_user.department_id,
        week_start_date=current_week_start,
        is_approved=True
    ).first()
    
    # Get employee balance
    employee_balance = db_session.query(EmployeeBalance).filter_by(user_id=current_user.id).first()
    if not employee_balance:
        # Create balance if it doesn't exist
        employee_balance = EmployeeBalance(user_id=current_user.id)
        db_session.add(employee_balance)
        db_session.commit()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_dashboard.html', 
                         employee_data=employee_data,
                         completion_percentage=completion_percentage,
                         missing_fields=missing_fields,
                         new_salary_slips=new_salary_slips,
                         has_current_schedule=has_current_schedule,
                         employee_balance=employee_balance,
                         notifications=notifications)

@app.route('/user/form', methods=['GET', 'POST'])
@login_required
def user_form():

    employee_data = db_session.query(EmployeeData).filter_by(user_id=current_user.id).first()
    
    if request.method == 'POST':
        if not current_user.is_admin or not current_user.is_manager:
            if not employee_data:
                employee_data = EmployeeData(user_id=current_user.id)
                db_session.add(employee_data)
        
        # Update all fields from form - none are required
        employee_data.arabic_name = request.form.get('arabic_name') or None
        employee_data.english_name = request.form.get('english_name') or None
        employee_data.national_id = request.form.get('national_id') or None
        
        if request.form.get('id_issue_date'):
            employee_data.id_issue_date = datetime.strptime(request.form['id_issue_date'], '%Y-%m-%d').date()
        else:
            employee_data.id_issue_date = None
        
        if request.form.get('birth_date'):
            employee_data.birth_date = datetime.strptime(request.form['birth_date'], '%Y-%m-%d').date()
            employee_data.age = calculate_age(employee_data.birth_date)
        else:
            employee_data.birth_date = None
            employee_data.age = None
        
        employee_data.whatsapp = request.form.get('whatsapp') or None
        employee_data.phone = request.form.get('phone') or None
        employee_data.address = request.form.get('address') or None
        employee_data.military_status = request.form.get('military_status') or None
        employee_data.marital_status = request.form.get('marital_status') or None
        employee_data.qualification = request.form.get('qualification') or None
        
        if request.form.get('graduation_year'):
            employee_data.graduation_year = int(request.form['graduation_year'])
        else:
            employee_data.graduation_year = None
        
        employee_data.grade = request.form.get('grade') or None
        employee_data.has_work = 'has_work' in request.form
        employee_data.workplace = request.form.get('workplace') or None
        employee_data.job_title = request.form.get('job_title') or None
        employee_data.insurance_number = request.form.get('insurance_number') or None
        employee_data.tax_number = request.form.get('tax_number') or None
        
        # Emergency contacts
        employee_data.emergency1_name = request.form.get('emergency1_name') or None
        employee_data.emergency1_phone = request.form.get('emergency1_phone') or None
        employee_data.emergency1_address = request.form.get('emergency1_address') or None
        employee_data.emergency1_relation = request.form.get('emergency1_relation') or None
        employee_data.emergency2_name = request.form.get('emergency2_name') or None
        employee_data.emergency2_phone = request.form.get('emergency2_phone') or None
        employee_data.emergency2_address = request.form.get('emergency2_address') or None
        employee_data.emergency2_relation = request.form.get('emergency2_relation') or None
        
        employee_data.profession_license = request.form.get('profession_license') or None
        employee_data.union_card = request.form.get('union_card') or None
        
        # Handle file uploads
        user_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], str(current_user.name))
        os.makedirs(user_upload_folder, exist_ok=True)
        
        file_fields = [
            ('national_id_image', 'national_id_image'),
            ('military_status_image', 'military_status_image'),
            ('qualification_image', 'qualification_image'),
            ('salary_details', 'salary_details'),
            ('employment_status', 'employment_status')
        ]
        
        for form_field, db_field in file_fields:
            file = request.files.get(form_field)
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file_path = os.path.join(user_upload_folder, filename)
                file.save(file_path)
                setattr(employee_data, db_field, filename)
        
        employee_data.last_updated = datetime.now()
        employee_data.updated_by = current_user.name
        
        # حساب نسبة الإكتمال باستخدام الدالة المساعدة
        completion_percentage = employee_data.calculate_completion()

        # Check if this is a final update or temporary save
        if 'save_complete' in request.form:
            employee_data.needs_update = False
            flash('تم التحديث النهائي للبيانات بنجاح')
            
            # Export to Excel
            export_to_excel(current_user.id)
            
            # Notify admin
            admins = db_session.query(User).filter_by(is_admin=True).all()
            for admin in admins:
                create_notification(
                    admin.id,
                    'تم تحديث البيانات نهائياً',
                    f'قام {current_user.name} بتحديث بياناته نهائياً',
                    'data_updated'
                )
        else:
            # Temporary save
            flash('تم الحفظ المؤقت للبيانات بنجاح')
        
        db_session.commit()
        
        return redirect(url_for('user_dashboard'))
    
    missing_fields = []
    completion_percentage = 0
    
    if employee_data:
        completion_percentage = employee_data.calculate_completion()
        missing_fields = employee_data.get_missing_fields()
        db_session.commit()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_form.html', 
                         employee_data=employee_data,
                         completion_percentage=completion_percentage,
                         missing_fields=missing_fields,
                         notifications=notifications)


@app.route('/user/leave', methods=['GET', 'POST'])
@login_required
def user_leave():

    balance = db_session.query(EmployeeBalance).filter_by(user_id=current_user.id).first()
    
    # الحصول على نطاق الشهر المالي
    financial_month_start, financial_month_end = get_financial_month_range()
    
    # الحصول على قائمة الوظائف من هيكل القسم
    department_jobs = get_department_jobs(current_user.department_id)
    
    if request.method == 'POST':
        # التحقق من وجود الحقول المطلوبة
        required_fields = ['leave_type', 'leave_date', 'reason']
        for field in required_fields:
            if field not in request.form:
                flash(f'حقل {field} مطلوب', 'error')
                return redirect(url_for('user_leave'))
        
        leave_type = request.form['leave_type']
        leave_date = datetime.strptime(request.form['leave_date'], '%Y-%m-%d').date()
        reason = request.form['reason']
        
        # جلب معلومات الشيفتات المختارة
        shifts = request.form.getlist('shifts[]')
        
        if not shifts:
            flash('يجب اختيار شيفت واحد على الأقل', 'error')
            return redirect(url_for('user_leave'))
        
        # التحقق من أن التاريخ في نطاق الشهر المالي
        if leave_date < financial_month_start or leave_date > financial_month_end:
            flash(f'يمكن اختيار التواريخ فقط من {financial_month_start.strftime("%Y-%m-%d")} إلى {financial_month_end.strftime("%Y-%m-%d")}', 'error')
            return redirect(url_for('user_leave'))
        
        # إنشاء معرف فريد لهذه المجموعة من الشيفتات
        parent_request_id = int(datetime.now().timestamp())
        created_requests = 0
        
        try:
            for i, shift in enumerate(shifts):
                # الحصول على الوظيفة المختارة لهذا الشيفت
                job = request.form.get(f'job_{shift}', '')
                
                # التحقق من اختيار وظيفة لكل شيفت (خاص بالإجازات من الرصيد)
                if leave_type == 'من رصيد الإجازات' and not job:
                    flash(f'يجب اختيار وظيفة للشيفت {shift}', 'error')
                    return redirect(url_for('user_leave'))
                
                # التحقق من الرصيد فقط للشيفت الأول من النوع "من رصيد الإجازات"
                if leave_type == 'من رصيد الإجازات'  and i == 0:
                    if balance.leave_balance < len(shifts):
                        flash(f'رصيد الإجازات غير كافي. تحتاج {len(shifts)} يوم، رصيدك الحالي: {balance.leave_balance} يوم', 'error')
                        return redirect(url_for('user_leave'))
                
                # إنشاء طلب إجازة منفصل لكل شيفت
                leave_request = LeaveRequest(
                    user_id=current_user.id,
                    department_id=current_user.department_id,
                    leave_type=leave_type,
                    start_date=leave_date,  # استخدام leave_date كـ start_date
                    end_date=leave_date,    # إذا كان نفس اليوم لبداية ونهاية الإجازة
                    leave_date=leave_date,  # أيضًا حفظه كـ leave_date إذا كان الحقل موجودًا
                    shift_name=shift,
                    shift_job=job,
                    total_days=1,  # كل شيفت = يوم واحد
                    reason=reason,
                    parent_request_id=parent_request_id,
                    shift_order=i+1,
                    status='pending'
                )
                
                db_session.add(leave_request)
                created_requests += 1
            
            # خصم من الرصيد إذا كانت الإجازة من رصيد الإجازات
            if leave_type == 'من رصيد الإجازات':
                balance.leave_balance -= len(shifts)
                balance.last_updated = datetime.now()
            
            db_session.commit()
            
            # إرسال إشعار للمدير
            department = db_session.query(Department).get(current_user.department_id)
            if department and department.primary_manager_id:
                shift_names = "، ".join(shifts)
                
                notification_message = f'طلب {current_user.name} إجازة من نوع {leave_type} '
                notification_message += f'بعدد {len(shifts)} شيفت ({shift_names}) '
                notification_message += f'بتاريخ {leave_date.strftime("%Y-%m-%d")}'
                
                create_notification(
                    department.primary_manager_id,
                    'طلب إجازة جديد',
                    notification_message,
                    'leave_request',
                    related_id=parent_request_id,
                    action_url=url_for('manager_leave_requests')
                )
            
            flash(f'تم تقديم {created_requests} طلب إجازة منفصل بنجاح', 'success')
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ أثناء تقديم الطلب: {str(e)}', 'error')
        
        return redirect(url_for('user_leave_requests'))
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_leave.html', 
                         balance=balance,
                         financial_month_start=financial_month_start,
                         financial_month_end=financial_month_end,
                         department_jobs=department_jobs,
                         notifications=notifications)

@app.template_filter('from_json')
def from_json_filter(value):
    """تحويل نص JSON إلى كائن Python في القوالب"""
    if not value or not isinstance(value, str):
        return {}
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return {}
    
# تأكد من أن Jinja2 يعرف الفلتر الجديد
app.jinja_env.filters['from_json'] = from_json_filter

def is_date_in_financial_month(check_date):
    """Check if date is within current financial month (26th to 25th)"""
    today = date.today()
    
    # Determine financial month range
    if today.day >= 26:
        # From 26th of current month to 25th of next month
        financial_start = date(today.year, today.month, 26)
        if today.month == 12:
            financial_end = date(today.year + 1, 1, 25)
        else:
            financial_end = date(today.year, today.month + 1, 25)
    else:
        # From 26th of previous month to 25th of current month
        if today.month == 1:
            financial_start = date(today.year - 1, 12, 26)
            financial_end = date(today.year, 1, 25)
        else:
            financial_start = date(today.year, today.month - 1, 26)
            financial_end = date(today.year, today.month, 25)
    
    return financial_start <= check_date <= financial_end


from datetime import date  # تأكد من وجود هذا الاستيراد في الأعلى
def get_financial_month_range():
    """الحصول على نطاق الشهر المالي (26 إلى 25 من الشهر التالي)"""
    today = date.today()
    
    if today.day >= 26:
        # الفترة: 26 من الشهر الحالي إلى 25 من الشهر التالي
        start_date = date(today.year, today.month, 26)
        if today.month == 12:
            end_date = date(today.year + 1, 1, 25)
        else:
            end_date = date(today.year, today.month + 1, 25)
    else:
        # الفترة: 26 من الشهر السابق إلى 25 من الشهر الحالي
        if today.month == 1:
            start_date = date(today.year - 1, 12, 26)
            end_date = date(today.year, 1, 25)
        else:
            start_date = date(today.year, today.month - 1, 26)
            end_date = date(today.year, today.month, 25)
    
    return start_date, end_date


def get_department_jobs(department_id):
    """الحصول على قائمة الوظائف من هيكل الجدول الخاص بالقسم"""
    department = db_session.query(Department).get(department_id)
    
    if not department:
        return []
    
    jobs = []
    
    # محاولة استخراج الوظائف من هيكل الجدول
    if department.schedule_structure:
        try:
            structure_data = json.loads(department.schedule_structure)
            
            # استخراج الوظائف من الهيكل
            if isinstance(structure_data, dict):
                # النوع الجديد: يحتوي على schedule كمفتاح
                if 'schedule' in structure_data:
                    schedule_list = structure_data['schedule']
                    for day_schedule in schedule_list:
                        if isinstance(day_schedule, dict):
                            job = day_schedule.get('job')
                            if job and job not in jobs:
                                jobs.append(job)
                
                # محاولة استخراج الوظائف من المفاتيح الأخرى
                elif 'jobs' in structure_data:
                    jobs = structure_data['jobs']
                
                # النوع القديم: قد يحتوي على حقل الوظائف مباشرة
                elif 'structure' in structure_data:
                    for item in structure_data['structure']:
                        if isinstance(item, dict):
                            job = item.get('job') or item.get('الوظيفة')
                            if job and job not in jobs:
                                jobs.append(job)
            
            elif isinstance(structure_data, list):
                # إذا كانت البيانات مباشرة كقائمة
                for item in structure_data:
                    if isinstance(item, dict):
                        job = item.get('job') or item.get('الوظيفة')
                        if job and job not in jobs:
                            jobs.append(job)
        
        except Exception as e:
            print(f"خطأ في استخراج الوظائف من هيكل القسم: {str(e)}")
    
    # إذا لم توجد وظائف في الهيكل، استخدم القائمة الافتراضية
    if not jobs:
        jobs = [
            'طبيب',
            'ممرض',
            'فني مختبر',
            'سكرتير طبي',
            'موظف استقبال',
            'أخصائي أشعة',
            'صيدلي',
            'إداري',
            'موظف خدمة'
        ]
    
    # إزالة القيم الفارغة وتصفية القيم
    jobs = [job for job in jobs if job and job.strip()]
    
    # إضافة "موظف" كخيار افتراضي إذا لم يكن موجوداً
    if 'موظف' not in jobs:
        jobs.insert(0, 'موظف')
    
    return jobs


@app.route('/approve_permission/<int:request_id>')
@login_required
def approve_permission(request_id):
    """Alias for manager_approve_permission to maintain template compatibility"""
    return redirect(url_for('manager_approve_permission', request_id=request_id))

@app.route('/user/leave_requests')
@login_required
def user_leave_requests():
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    # Get period filter parameter
    period_filter = request.args.get('period', 'all')
    
    # Get all leave requests (individual shifts)
    query = db_session.query(LeaveRequest)\
        .filter_by(user_id=current_user.id)\
        .order_by(LeaveRequest.created_at.desc())
    
    # Apply period filter
    if period_filter != 'all':
        today = date.today()
        
        if period_filter == 'week':
            week_ago = today - timedelta(days=7)
            query = query.filter(LeaveRequest.created_at >= week_ago)
        elif period_filter == 'month':
            month_ago = today - timedelta(days=30)
            query = query.filter(LeaveRequest.created_at >= month_ago)
        elif period_filter == '3months':
            three_months_ago = today - timedelta(days=90)
            query = query.filter(LeaveRequest.created_at >= three_months_ago)
    
    # Get all individual leave requests
    leave_requests = query.all()
    
    # Get permission requests with proper handling
    permission_query = db_session.query(PermissionRequest).filter_by(
        user_id=current_user.id
    ).order_by(PermissionRequest.created_at.desc())
    
    if period_filter != 'all':
        today = date.today()
        if period_filter == 'week':
            week_ago = today - timedelta(days=7)
            permission_query = permission_query.filter(PermissionRequest.created_at >= week_ago)
        elif period_filter == 'month':
            month_ago = today - timedelta(days=30)
            permission_query = permission_query.filter(PermissionRequest.created_at >= month_ago)
        elif period_filter == '3months':
            three_months_ago = today - timedelta(days=90)
            permission_query = permission_query.filter(PermissionRequest.created_at >= three_months_ago)
    
    permission_requests = permission_query.all()
    
    # Debug: Print permission request data
    print(f"Found {len(permission_requests)} permission requests")
    for i, req in enumerate(permission_requests[:5]):  # Show first 5
        print(f"Permission {i+1}: ID={req.id}, Type={req.permission_type}, Extra Data={req.extra_data}")
    
    # Calculate completion percentage
    employee_data = db_session.query(EmployeeData).filter_by(user_id=current_user.id).first()
    new_salary_slips = 0
    if employee_data:
        employee_data.calculate_completion()
        new_salary_slips = db_session.query(SalarySlip).filter_by(
            user_id=current_user.id, is_viewed=False
        ).count()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_leave_requests.html',
                         leave_requests=leave_requests,
                         permission_requests=permission_requests,
                         period_filter=period_filter,
                         new_salary_slips=new_salary_slips,
                         notifications=notifications)



@app.route('/debug_permissions')
@login_required
def debug_permissions():
    """Debug route to check permission data structure"""
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    permission_requests = db_session.query(PermissionRequest).filter_by(
        user_id=current_user.id
    ).order_by(PermissionRequest.created_at.desc()).all()
    
    debug_data = []
    for req in permission_requests:
        debug_data.append({
            'id': req.id,
            'permission_type': req.permission_type,
            'date': req.date.strftime('%Y-%m-%d') if req.date else None,
            'time': req.time,
            'reason': req.reason,
            'extra_data': req.extra_data,
            'extra_data_dict': req.extra_data_dict if hasattr(req, 'extra_data_dict') else {},
            'shift_name': req.shift_name if hasattr(req, 'shift_name') else 'N/A',
            'shift_job': req.shift_job if hasattr(req, 'shift_job') else 'N/A',
            'exchange_employee_name': req.exchange_employee_name if hasattr(req, 'exchange_employee_name') else 'N/A',
            'attendance_type': req.attendance_type if hasattr(req, 'attendance_type') else 'N/A',
            'overtime_hours': req.overtime_hours if hasattr(req, 'overtime_hours') else 'N/A',
            'status': req.status
        })
    
    return jsonify({
        'user_id': current_user.id,
        'total_requests': len(permission_requests),
        'requests': debug_data
    })

# In your app.py, add these template filters:

@app.template_filter('get_extra_data')
def get_extra_data_filter(extra_data_json):
    """Parse extra_data JSON in template"""
    if not extra_data_json:
        return {}
    try:
        return json.loads(extra_data_json)
    except:
        return {}

@app.route('/api/manager/permission_statistics')
@login_required
def api_permission_statistics():
    """API للحصول على إحصائيات الإذونات"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        department_ids = [dept.id for dept in managed_departments]
        
        if not department_ids:
            return jsonify({'success': True, 'statistics': {}})
        
        # Get filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        department_id = request.args.get('department_id')
        permission_type = request.args.get('permission_type')
        
        # Build query
        query = db_session.query(PermissionRequest).filter(
            PermissionRequest.department_id.in_(department_ids)
        )
        
        if start_date:
            query = query.filter(PermissionRequest.date >= start_date)
        if end_date:
            query = query.filter(PermissionRequest.date <= end_date)
        if department_id and department_id != 'all':
            query = query.filter(PermissionRequest.department_id == department_id)
        if permission_type and permission_type != 'all':
            query = query.filter(PermissionRequest.permission_type == permission_type)
        
        all_requests = query.all()
        
        # Calculate statistics
        stats = get_permission_statistics(department_ids, all_requests)
        
        return jsonify({
            'success': True,
            'statistics': stats
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })


@app.template_filter('get_shift_name')
def get_shift_name_filter(extra_data_json):
    """Get shift name from extra_data in template"""
    data = get_extra_data_filter(extra_data_json)
    return data.get('shift', '')

@app.template_filter('get_shift_job')
def get_shift_job_filter(extra_data_json):
    """Get job from extra_data in template"""
    data = get_extra_data_filter(extra_data_json)
    return data.get('job', '')

@app.template_filter('get_employee_name')
def get_employee_name_filter(extra_data_json):
    """Get employee name from extra_data in template"""
    data = get_extra_data_filter(extra_data_json)
    return data.get('employee_name', '')

@app.template_filter('get_attendance_type')
def get_attendance_type_filter(extra_data_json):
    """Get attendance type from extra_data in template"""
    data = get_extra_data_filter(extra_data_json)
    return data.get('attendance_type', '')

@app.template_filter('get_overtime_hours')
def get_overtime_hours_filter(extra_data_json):
    """Get overtime hours from extra_data in template"""
    data = get_extra_data_filter(extra_data_json)
    return data.get('hours') or data.get('overtime_hours') or 0




def reset_monthly_permission_balances():
    """Reset permission balance to 2 for all employees at the beginning of each month"""
    try:
        today = date.today()
        
        # Check if it's the first day of the month
        if today.day == 26:
            print(f"=== Resetting permission balances for {today.strftime('%Y-%m')} ===")
            
            # Get all employee balances
            employee_balances = db_session.query(EmployeeBalance).all()
            reset_count = 0
            
            for balance in employee_balances:
                # Only reset if the balance is different from default (2)
                if balance.permission_balance != 2:
                    old_balance = balance.permission_balance
                    balance.permission_balance = 2
                    balance.last_updated = datetime.now()
                    reset_count += 1
                    
                    print(f"Reset permission balance for user {balance.user_id}: {old_balance} -> 2")
                    
                    # Notify employee about balance reset
                    create_notification(
                        balance.user_id,
                        'تم تجديد رصيد الإذونات',
                        f'تم تجديد رصيد الإذونات الخاص بك إلى 2 إذن لشهر {today.strftime("%Y-%m")}',
                        'balance_reset',
                        action_url=url_for('user_dashboard')
                    )
            
                
            return reset_count
        else:
            print(f"Not the first day of month ({today.day}), skipping permission balance reset")
            return 0
            
    except Exception as e:
        print(f"Error resetting permission balances: {str(e)}")
        db_session.rollback()
        return 0
    

def check_and_reset_balances():
    """Check if balances need to be reset (runs daily)"""
    try:
        # Check if we've already reset balances this month
        today = date.today()
        current_month = today.strftime('%Y-%m')
        
        if not hasattr(app, 'last_balance_reset_month'):
            app.last_balance_reset_month = None
        
        # Reset if it's a new month and we haven't reset yet
        if app.last_balance_reset_month != current_month and today.day == 26:
            reset_count = reset_monthly_permission_balances()
            app.last_balance_reset_month = current_month
            return reset_count
        
        return 0
        
    except Exception as e:
        print(f"Error in balance reset check: {str(e)}")
        return 0    


@app.route('/user/permission', methods=['GET', 'POST'])
@login_required
def user_permission():

    balance = db_session.query(EmployeeBalance).filter_by(user_id=current_user.id).first()
    department_jobs = get_department_jobs(current_user.department_id)
    
    # Get department employees for auto-complete
    department_employees = db_session.query(User).filter(
        User.department_id == current_user.department_id,
        User.is_active == True,
        User.is_admin == False
    ).all()
    

    if request.method == 'POST':
        permission_type = request.form['permission_type']
        permission_date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        reason = request.form['reason']
        

        
        created_count = 0
        
        try:
            # معالجة أنواع الإذن المختلفة
            if permission_type == 'اذن تبديل وردية':
                # جمع بيانات تبديل الشيفتات
                exchange_shifts = request.form.getlist('exchange_shifts[]')
                
                if not exchange_shifts:
                    flash('يرجى اختيار شيفت واحد على الأقل للتبديل')
                    return redirect(url_for('user_permission'))
                
                # التحقق من أن جميع الشيفتات المختارة لها وظائف محددة
                for shift in exchange_shifts:
                    job_key = f'exchange_job_{shift}'
                    job = request.form.get(job_key)
                    
                    if not job:
                        flash(f'يرجى اختيار وظيفة للشيفت {shift}')
                        return redirect(url_for('user_permission'))
                
                # إنشاء سجل منفصل لكل شيفت
                for shift in exchange_shifts:
                    job_key = f'exchange_job_{shift}'
                    employee_key = f'exchange_employee_{shift}'
                    employee_id_key = f'exchange_employee_{shift}_id'
                    
                    job = request.form.get(job_key)
                    employee_name = request.form.get(employee_key)
                    employee_id = request.form.get(employee_id_key)
                    
                    # بيانات إضافية لهذا الشيفت المحدد
                    extra_data = {
                        'type': 'shift_exchange',
                        'shift': shift,
                        'job': job,
                        'employee_name': employee_name,
                        'employee_id': employee_id
                    }
                    
                    # حساب الوقت
                    time_minutes = 0
                    time_input = request.form.get('time')
                    if time_input:
                        try:
                            time_obj = datetime.strptime(time_input, '%H:%M').time()
                            time_minutes = time_obj.hour * 60 + time_obj.minute
                        except:
                            pass
                    
                    # إنشاء طلب إذن منفصل لهذا الشيفت
                    permission_request = PermissionRequest(
                        user_id=current_user.id,
                        department_id=current_user.department_id,
                        permission_type=f'{permission_type} - {shift}',
                        date=permission_date,
                        time=time_minutes,
                        reason=reason,
                        extra_data=json.dumps(extra_data, ensure_ascii=False),
                        status='pending'
                    )
                    
                    db_session.add(permission_request)
                    created_count += 1
                                
            elif permission_type == 'اذن طلب ساعات اضافي':
                # جمع بيانات الساعات الإضافية
                overtime_shifts = request.form.getlist('overtime_shifts[]')
                overtime_hours = int(request.form.get('overtime_hours', 1))
                
                if not overtime_shifts:
                    flash('يرجى اختيار شيفت واحد على الأقل للعمل الإضافي')
                    return redirect(url_for('user_permission'))
                
                # التحقق من عدد الساعات
                if overtime_hours < 1 or overtime_hours > 12:
                    flash('عدد الساعات الإضافية يجب أن يكون بين 1 و 12 ساعة')
                    return redirect(url_for('user_permission'))
                
                # التحقق من أن جميع الشيفتات المختارة لها وظائف محددة
                for shift in overtime_shifts:
                    job_key = f'overtime_job_{shift}'
                    job = request.form.get(job_key)
                    
                    if not job:
                        flash(f'يرجى اختيار وظيفة للشيفت {shift}')
                        return redirect(url_for('user_permission'))
                    
                    # التحقق من اختيار توقيت
                    if shift == 'صباحي':
                        attendance_type = request.form.get('morning_attendance_type')
                        if not attendance_type:
                            flash(f'يرجى اختيار توقيت للشيفت الصباحي')
                            return redirect(url_for('user_permission'))
                    else:
                        attendance_type = request.form.get('evening_attendance_type')
                        if not attendance_type:
                            flash(f'يرجى اختيار توقيت للشيفت المسائي')
                            return redirect(url_for('user_permission'))
                
                # إنشاء سجل منفصل لكل شيفت
                for shift in overtime_shifts:
                    job_key = f'overtime_job_{shift}'
                    job = request.form.get(job_key)
                    
                    # تحديد نوع الحضور/الانصراف
                    if shift == 'صباحي':
                        attendance_type = request.form.get('morning_attendance_type')
                    else:
                        attendance_type = request.form.get('evening_attendance_type')
                    
                    # بيانات إضافية لهذا الشيفت المحدد
                    extra_data = {
                        'type': 'overtime',
                        'shift': shift,
                        'job': job,
                        'hours': overtime_hours,
                        'attendance_type': attendance_type
                    }
                    
                    # حساب الوقت
                    time_minutes = 0
                    time_input = request.form.get('time')
                    if time_input:
                        try:
                            time_obj = datetime.strptime(time_input, '%H:%M').time()
                            time_minutes = time_obj.hour * 60 + time_obj.minute
                        except:
                            pass
                    
                    # إنشاء طلب إذن منفصل لهذا الشيفت
                    permission_request = PermissionRequest(
                        user_id=current_user.id,
                        department_id=current_user.department_id,
                        permission_type=f'{permission_type} - {shift}',
                        date=permission_date,
                        time=time_minutes,
                        reason=reason,
                        extra_data=json.dumps(extra_data, ensure_ascii=False),
                        status='pending'
                    )
                    
                    db_session.add(permission_request)
                    created_count += 1
                            
            else:
                # الإذونات العادية (حضور/انصراف/نسيان توقيع)
                if balance.permission_balance <= 0:
                    flash('رصيد الإذونات غير كافي')
                    return redirect(url_for('user_permission'))

                extra_data = {}
                time_minutes = 0
                time_input = request.form.get('time')
                
                if time_input:
                    try:
                        time_obj = datetime.strptime(time_input, '%H:%M').time()
                        time_minutes = time_obj.hour * 60 + time_obj.minute
                    except:
                        pass
                
                # إنشاء طلب إذن واحد
                permission_request = PermissionRequest(
                    user_id=current_user.id,
                    department_id=current_user.department_id,
                    permission_type=permission_type,
                    date=permission_date,
                    time=time_minutes,
                    reason=reason,
                    extra_data=json.dumps(extra_data) if extra_data else None,
                    status='pending'
                )
                
                db_session.add(permission_request)
                created_count += 1
                
                # خصم من رصيد الإذونات
                balance.permission_balance -= 1
            
            # تحديث رصيد الموظف
            balance.last_updated = datetime.now()
            
            # حفظ جميع التغييرات
            db_session.commit()
            
            # إرسال إشعار للمدير
            department = db_session.query(Department).get(current_user.department_id)
            if department and department.primary_manager_id:
                notification_message = f'طلب {current_user.name} {created_count} إذن من نوع {permission_type}'
                
                create_notification(
                    department.primary_manager_id,
                    'طلب إذن جديد',
                    notification_message,
                    'permission_request',
                    related_id=permission_request.id if created_count == 1 else None,
                    action_url=url_for('manager_permission_requests')
                )
            
            if created_count > 1:
                flash(f'تم تقديم {created_count} طلب إذن منفصل بنجاح')
            else:
                flash('تم تقديم طلب الإذن بنجاح وفي انتظار الموافقة')
            
            return redirect(url_for('user_leave_requests'))
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ أثناء تقديم الطلب: {str(e)}', 'error')
            print(f"Error in permission submission: {str(e)}")
            import traceback
            traceback.print_exc()
            return redirect(url_for('user_permission'))
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_permission.html',
                         balance=balance,
                         department_jobs=department_jobs,
                         department_employees=department_employees,
                         notifications=notifications)


@app.route('/admin/create_department', methods=['POST'])
@login_required
def admin_create_department():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        name = request.form['name']
        primary_manager_id = request.form.get('primary_manager_id') or None
        max_advance_amount = float(request.form.get('max_advance_amount', 0))
        max_installments = int(request.form.get('max_installments', 1))
        auto_generate = 'auto_generate_schedule' in request.form
        
        department = Department(
            name=name,
            primary_manager_id=primary_manager_id,
            advance_policy_max_amount=max_advance_amount,
            advance_policy_max_installments=max_installments,
            auto_generate_schedule=auto_generate,
            created_by=current_user.id
        )
        
        db_session.add(department)
        db_session.commit()
        
        flash('تم إنشاء القسم بنجاح', 'success')
        
    except Exception as e:
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('admin_departments'))

@app.route('/admin/add_department_manager', methods=['POST'])
@login_required
def admin_add_department_manager():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        department_id = request.form['department_id']
        user_id = request.form['user_id']
        
        # Get permissions from form
        permissions = {
            'can_manage_schedules': 'can_manage_schedules' in request.form,
            'can_manage_leaves': 'can_manage_leaves' in request.form,
            'can_manage_permissions': 'can_manage_permissions' in request.form,
            'can_manage_advances': 'can_manage_advances' in request.form,
            'can_manage_rewards': 'can_manage_rewards' in request.form,
            'can_view_reports': 'can_view_reports' in request.form
        }
        
        # Create department manager record
        department_manager = DepartmentManager(
            department_id=department_id,
            user_id=user_id,
            created_by=current_user.id,
            **permissions
        )
        
        db_session.add(department_manager)
        
        # Update user to be a manager if not already
        user = db_session.get(User, user_id)
        if not user.is_manager:
            user.is_manager = True
        
        db_session.commit()
        
        flash('تم إضافة المدير بنجاح', 'success')
        
    except Exception as e:
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('admin_departments'))

@app.route('/admin/update_manager_permissions/<int:manager_id>', methods=['POST'])
@login_required
def admin_update_manager_permissions(manager_id):
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        manager = db_session.query(DepartmentManager).get(manager_id)
        if not manager:
            flash('المدير غير موجود', 'error')
            return redirect(url_for('admin_departments'))
        
        # Update permissions
        manager.can_manage_schedules = 'can_manage_schedules' in request.form
        manager.can_manage_leaves = 'can_manage_leaves' in request.form
        manager.can_manage_permissions = 'can_manage_permissions' in request.form
        manager.can_manage_advances = 'can_manage_advances' in request.form
        manager.can_manage_rewards = 'can_manage_rewards' in request.form
        manager.can_view_reports = 'can_view_reports' in request.form
        
        db_session.commit()
        flash('تم تحديث الصلاحيات بنجاح', 'success')
        
    except Exception as e:
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('admin_departments'))


# Add this helper function
def get_manager_permissions_display(manager):
    """Return a string representation of manager permissions"""
    if not manager:
        return "لا توجد صلاحيات"
    
    permissions = []
    if manager.can_manage_schedules:
        permissions.append('إدارة الجداول')
    if manager.can_manage_leaves:
        permissions.append('إدارة الإجازات')
    if manager.can_manage_permissions:
        permissions.append('إدارة الإذونات')
    if manager.can_manage_advances:
        permissions.append('إدارة السلف')
    if manager.can_manage_rewards:
        permissions.append('إدارة المكافآت')
    if manager.can_view_reports:
        permissions.append('عرض التقارير')
    
    return '، '.join(permissions) if permissions else 'لا توجد صلاحيات'

# Update the context processor
@app.context_processor
def utility_processor():
    def get_manager_permissions(manager):
        return get_manager_permissions_display(manager)
    
    def get_approver_name(approver_id):
        if not approver_id:
            return "غير معين"
        approver = db_session.query(User).get(approver_id)
        return approver.name if approver else "غير معين"
    
    return dict(
        get_manager_permissions_display=get_manager_permissions,
        get_approver_name=get_approver_name
    )


@app.route('/user/cancel_leave/<int:request_id>')
@login_required
def cancel_leave_request(request_id):
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    leave_request = db_session.query(LeaveRequest).filter_by(id=request_id).first()
    
    if leave_request and leave_request.user_id == current_user.id and leave_request.status == 'pending':
        # Check if this is a leave from balance and restore it
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=current_user.id
            ).first()
            
            if balance:
                
                balance.leave_balance += 1  # Restore 1 day for each shift
                balance.last_updated = datetime.now()
        
        db_session.delete(leave_request)
        db_session.commit()
        flash('تم إلغاء الشيفت بنجاح')
    
    return redirect(url_for('user_leave_requests'))


@app.route('/user/cancel_permission/<int:request_id>')
@login_required
def cancel_permission_request(request_id):
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    permission_request = db_session.get(PermissionRequest, request_id)
    


    if permission_request and permission_request.user_id == current_user.id and permission_request.status == 'pending':
        
        balance = db_session.query(EmployeeBalance).filter_by(user_id=current_user.id).first()
        balance.permission_balance += 1
        balance.last_updated = datetime.now()
        
        
        db_session.delete(permission_request)
        db_session.commit()
        flash('تم إلغاء طلب الإذن بنجاح')
    



    return redirect(url_for('user_leave_requests'))

@app.route('/user/permission_requests')
@login_required
def user_permission_requests():
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    permission_requests = db_session.query(PermissionRequest).filter_by(
        user_id=current_user.id
    ).order_by(PermissionRequest.created_at.desc()).all()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_permission_requests.html',
                         permission_requests=permission_requests,
                         notifications=notifications)


import glob
import re
def detect_salary_slips():
    """اكتشاف شيتات المرتب الجديدة تلقائياً"""
    try:
        salary_folder = app.config['SALARY_FOLDER']
        if not os.path.exists(salary_folder):
            os.makedirs(salary_folder)
        
        detected_count = 0
        for filename in os.listdir(salary_folder):
            if filename.lower().endswith('.pdf'):
                file_path = os.path.join(salary_folder, filename)
                
                # استخراج معلومات الملف
                file_info = extract_file_info_by_username(filename)
                
                if file_info:
                    username = file_info['username']
                    month = file_info['month']
                    arabic_month = file_info['arabic_month']
                    
                    # البحث عن المستخدم باستخدام username فقط
                    user = db_session.query(User).filter(
                        User.username == username
                    ).first()
                    
                    if user:
                        # التحقق من عدم وجود شيت مكرر
                        existing_slip = db_session.query(SalarySlip).filter_by(
                            user_id=user.id,
                            month=month
                        ).first()
                        
                        if not existing_slip:
                            # إنشاء سجل جديد
                            salary_slip = SalarySlip(
                                user_id=user.id,
                                month=month,
                                arabic_month=arabic_month,
                                file_path=file_path,
                                file_name=filename,
                                uploaded_by=1,  # النظام
                                is_auto_detected=True,
                                is_viewed=False
                            )
                            db_session.add(salary_slip)
                            
                            # إنشاء إشعار للمستخدم
                            create_notification(
                                user.id,
                                'شيت مرتب جديد',
                                f'تم اكتشاف شيت مرتب جديد لشهر {arabic_month}',
                                'salary',
                                salary_slip.id
                            )
                            
                            detected_count += 1
                            print(f"✅ تم اكتشاف شيت مرتب جديد: {filename} للمستخدم: {username} لشهر: {arabic_month}")
                    else:
                        print(f"⚠️  لم يتم العثور على مستخدم باسم: {username}")
        
        db_session.commit()
        return detected_count
        
    except Exception as e:
        print(f'❌ خطأ في اكتشاف شيتات المرتب: {str(e)}')
        db_session.rollback()
        return 0

import os
import re
from datetime import datetime

def extract_month_from_filename(filename):
    """استخراج الشهر من اسم الملف مع دعم العربية والإنجليزية"""
    try:
        name_without_ext = os.path.splitext(filename)[0]
        
        # البحث عن أنماط التاريخ المختلفة
        # النمط الجديد: 2025_9---510.PDF
        year_month_match = re.search(r'(\d{4})_(\d{1,2})', name_without_ext)
        if year_month_match:
            year = int(year_month_match.group(1))
            month = int(year_month_match.group(2))
            return f"{year}-{month:02d}"
        
        # الأنماط القديمة (للتوافق مع الإصدارات السابقة)
        month_year_match = re.search(r'(\d{4})-(\d{2})', name_without_ext)  # 2024-01
        if month_year_match:
            return month_year_match.group(0)
        
        # الأشهر العربية والإنجليزية
        month_names = {
            # العربية الكاملة
            'يناير': '01', 'فبراير': '02', 'مارس': '03', 'أبريل': '04',
            'مايو': '05', 'يونيو': '06', 'يوليو': '07', 'أغسطس': '08',
            'سبتمبر': '09', 'أكتوبر': '10', 'نوفمبر': '11', 'ديسمبر': '12',
            # العربية المختصرة
            'ينا': '01', 'فبر': '02', 'مار': '03', 'أبر': '04',
            'ماي': '05', 'يون': '06', 'يول': '07', 'أغس': '08',
            'سبت': '09', 'أكت': '10', 'نوف': '11', 'ديس': '12',
            # الإنجليزية الكاملة
            'january': '01', 'february': '02', 'march': '03', 'april': '04',
            'may': '05', 'june': '06', 'july': '07', 'august': '08',
            'september': '09', 'october': '10', 'november': '11', 'december': '12',
            # الإنجليزية المختصرة
            'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04', 'may': '05', 'jun': '06',
            'jul': '07', 'aug': '08', 'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'
        }
        
        for month_name, month_num in month_names.items():
            if re.search(month_name, name_without_ext, re.IGNORECASE):
                current_year = datetime.now().year
                return f"{current_year}-{month_num}"
        
        # إذا لم يتم العثور على شهر، نستخدم الشهر الحالي
        return datetime.now().strftime('%Y-%m')
        
    except Exception as e:
        print(f'خطأ في استخراج الشهر من {filename}: {str(e)}')
        return datetime.now().strftime('%Y-%m')

def get_arabic_month_name(month_num):
    """الحصول على اسم الشهر بالعربية"""
    arabic_months = {
        '01': 'يناير', '02': 'فبراير', '03': 'مارس', '04': 'أبريل',
        '05': 'مايو', '06': 'يونيو', '07': 'يوليو', '08': 'أغسطس',
        '09': 'سبتمبر', '10': 'أكتوبر', '11': 'نوفمبر', '12': 'ديسمبر'
    }
    return arabic_months.get(month_num, 'غير معروف')

def extract_file_info_by_username(filename):
    """استخراج معلومات المستخدم والشهر من اسم الملف مع دعم العربية"""
    try:
        # إزالة الامتداد
        name_without_ext = os.path.splitext(filename)[0]
        
        print(f"معالجة الملف: {filename}")
        print(f"الاسم بدون امتداد: {name_without_ext}")
        
        # محاولة استخراج الشهر من اسم الملف
        month = extract_month_from_filename(filename)
        print(f"الشهر المستخرج: {month}")
        
        # استخراج اسم الشهر بالعربية
        month_num = month.split('-')[1] if '-' in month else datetime.now().strftime('%m')
        arabic_month = get_arabic_month_name(month_num)
        print(f"اسم الشهر بالعربية: {arabic_month}")
        
        # استخراج اسم المستخدم
        username = None
        
        # النمط 1: تاريخ_شهر---اسم مستخدم (2025_9---510)
        pattern1 = r'\d{4}_\d{1,2}---(\w+)'
        match1 = re.search(pattern1, name_without_ext)
        if match1:
            username = match1.group(1)
            print(f"تم التعرف على النمط 1 - اسم المستخدم: {username}")
        
        # النمط 2: اسم مستخدم_تاريخ (username_2024-01)
        if not username:
            pattern2 = r'^([^_]+)_\d'
            match2 = re.search(pattern2, name_without_ext)
            if match2:
                username = match2.group(1)
                print(f"تم التعرف على النمط 2 - اسم المستخدم: {username}")
        
        # النمط 3: اسم مستخدم فقط (username.pdf)
        if not username:
            # إزالة أنماط التاريخ المعروفة
            temp_username = re.sub(r'_\d{4}-\d{2}', '', name_without_ext)
            temp_username = re.sub(r'_\d{4}_\d{1,2}', '', temp_username)
            temp_username = re.sub(r'---\d+$', '', temp_username)
            temp_username = re.sub(r'_\d{6}', '', temp_username)
            
            # إزالة أسماء الأشهر العربية والإنجليزية
            month_patterns = [
                # العربية الكاملة
                r'_يناير', r'_فبراير', r'_مارس', r'_أبريل', r'_مايو', r'_يونيو',
                r'_يوليو', r'_أغسطس', r'_سبتمبر', r'_أكتوبر', r'_نوفمبر', r'_ديسمبر',
                # العربية المختصرة
                r'_ينا', r'_فبر', r'_مار', r'_أبر', r'_ماي', r'_يون',
                r'_يول', r'_أغس', r'_سبت', r'_أكت', r'_نوف', r'_ديس',
                # الإنجليزية
                r'_january', r'_february', r'_march', r'_april', r'_may', r'_june',
                r'_july', r'_august', r'_september', r'_october', r'_november', r'_december',
                r'_jan', r'_feb', r'_mar', r'_apr', r'_may', r'_jun',
                r'_jul', r'_aug', r'_sep', r'_oct', r'_nov', r'_dec'
            ]
            
            for pattern in month_patterns:
                temp_username = re.sub(pattern, '', temp_username, flags=re.IGNORECASE)
            
            username = temp_username.strip(' _-')
            print(f"تم التعرف على النمط 3 - اسم المستخدم: {username}")
        
        if username:
            return {
                'username': username, 
                'month': month,  # التنسيق الرقمي: 2025-09
                'arabic_month': arabic_month,  # اسم الشهر العربي: سبتمبر
                'display_month': arabic_month  # اسم العرض العربي
            }
        else:
            print(f"لم يتم العثور على اسم مستخدم في: {filename}")
            return None
            
    except Exception as e:
        print(f'خطأ في تحليل اسم الملف {filename}: {str(e)}')
        return None

@app.route('/user/salary_slips')
@login_required
def user_salary_slips():
    if current_user.is_admin:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    # اكتشاف شيتات المرتب الجديدة تلقائياً قبل العرض
    detect_salary_slips()
    
    # الحصول على جميع شيتات المرتب للمستخدم
    all_salary_slips = db_session.query(SalarySlip).filter_by(
        user_id=current_user.id
    ).order_by(SalarySlip.month.desc()).all()
    
    # تصفية الشيتات التي لا يوجد لها ملف فعلي
    valid_salary_slips = []
    for slip in all_salary_slips:
        if slip.file_path and os.path.exists(slip.file_path):
            valid_salary_slips.append(slip)
        else:
            # حذف السجل من قاعدة البيانات إذا لم يكن الملف موجوداً
            print(f"تحذير: ملف شيت المرتب غير موجود - {slip.file_path}")
            db_session.delete(slip)
    
    # تحديث حالة المشاهدة للشيتات الصالحة فقط
    for slip in valid_salary_slips:
        if not slip.is_viewed:
            slip.is_viewed = True
    
    db_session.commit()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_salary_slips.html',
                         salary_slips=valid_salary_slips,  # إرسال القائمة المصفاة فقط
                         notifications=notifications,
                         can_extract_images=True)

@app.route('/admin/upload_salary_slip', methods=['GET', 'POST'])
@login_required
def admin_upload_salary_slip():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    if request.method == 'POST':
        user_id = request.form['user_id']
        month = request.form['month']  # 2025-09
        file = request.files['salary_file']
        
        if file and allowed_file(file.filename):
            user = db_session.get(User, user_id)
            if not user:
                flash('المستخدم غير موجود')
                return redirect(url_for('admin_upload_salary_slip'))
            
            # استخراج اسم الشهر العربي
            month_num = month.split('-')[1] if '-' in month else datetime.now().strftime('%m')
            arabic_month = get_arabic_month_name(month_num)
            
            existing_slip = db_session.query(SalarySlip).filter_by(
                user_id=user_id,
                month=month
            ).first()
            
            if existing_slip:
                flash('يوجد already شيت مرتب لنفس الشهر')
                return redirect(url_for('admin_upload_salary_slip'))
            
            filename = secure_filename(f"{user_id}_{month}_{file.filename}")
            file_path = os.path.join(app.config['SALARY_FOLDER'], filename)
            file.save(file_path)
            
            salary_slip = SalarySlip(
                user_id=user_id,
                month=month,
                arabic_month=arabic_month,  # حفظ اسم الشهر العربي
                file_path=file_path,
                file_name=filename,
                uploaded_by=current_user.id
            )
            db_session.add(salary_slip)
            
            create_notification(
                user_id,
                'شيت مرتب جديد',
                f'تم رفع شيت المرتب لشهر {arabic_month}',  # استخدام الاسم العربي
                'salary',
                salary_slip.id
            )
            
            db_session.commit()
            flash('تم رفع شيت المرتب بنجاح')
            return redirect(url_for('admin_salary_slips'))
        else:
            flash('الملف غير مسموح به')
    
    users = db_session.query(User).filter_by(is_admin=False).all()
    notifications_list = db_session.query(Notification).filter_by(
        user_id=current_user.id, is_read=False
    ).order_by(Notification.created_at.desc()).limit(5).all()
    
    return render_template('admin_upload_salary_slip.html', 
                         users=users, 
                         notifications=notifications_list)

@app.route('/admin/salary_slips', methods=['GET', 'POST'])
@login_required
def admin_salary_slips():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    if request.method == 'POST':
        # Manual upload functionality
        user_id = request.form['user_id']
        month = request.form['month']
        file = request.files['file']
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['SALARY_FOLDER'], filename)
            file.save(file_path)
            
            salary_slip = SalarySlip(
                user_id=user_id,
                month=month,
                file_path=file_path,
                file_name=filename,
                uploaded_by=current_user.id,
                is_auto_detected=False
            )
            db_session.add(salary_slip)
            
            # Notify user
            user = db_session.get(User, user_id)
            create_notification(
                user_id,
                'مستحق راتب جديد',
                f'تم إضافة مستحق راتب جديد لشهر {month}',
                'salary_slip',
                action_url=url_for('user_salary_slips')
            )
            
            db_session.commit()
            flash('تم رفع مستحق الراتب بنجاح')
    
    # Auto-detect on page load
    detected_count = detect_salary_slips()
    if detected_count > 0:
        flash(f'تم اكتشاف {detected_count} مستحق راتب جديد تلقائياً')
    
    users = db_session.query(User).filter_by(is_admin=False).all()
    salary_slips = db_session.query(SalarySlip).order_by(SalarySlip.month.desc()).all()
    
    # Add user info to salary slips for display
    for slip in salary_slips:
        slip.user = db_session.query(User).get(slip.user_id)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_salary_slips.html',
                         users=users,
                         salary_slips=salary_slips,
                         notifications=notifications)

@app.route('/download_salary_slip/<int:slip_id>')
@login_required
def download_salary_slip(slip_id):
    salary_slip = db_session.query(SalarySlip).get(slip_id)
    
    # Check permissions

    if salary_slip and os.path.exists(salary_slip.file_path):
        return send_file(salary_slip.file_path, as_attachment=True)
    else:
        flash('الملف غير موجود')
        return redirect(request.referrer or url_for('user_dashboard'))




@app.route('/user/rewards_penalties')
@login_required
def user_rewards_penalties():
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    rewards_penalties = db_session.query(RewardPenalty).filter_by(
        user_id=current_user.id
    ).order_by(RewardPenalty.effective_date.desc()).all()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_rewards_penalties.html',
                         rewards_penalties=rewards_penalties,
                         notifications=notifications)




@app.route('/admin/all_leaves')
@login_required
def admin_all_leaves():
    """عرض جميع الإجازات في كل قسم حسب الشهر المالي"""

    # الحصول على معاملات التصفية
    month_filter = request.args.get('month', '')
    department_filter = request.args.get('department', '')
    status_filter = request.args.get('status', '')
    
    # بناء الاستعلام الأساسي بدون joinedload
    query = db_session.query(LeaveRequest).options(joinedload(LeaveRequest.approver))

    
    # تطبيق الفلاتر
    if department_filter and department_filter.isdigit():
        query = query.filter(LeaveRequest.department_id == int(department_filter))
    
    if status_filter:
        query = query.filter(LeaveRequest.status == status_filter)
    
    # الحصول على جميع الإجازات
    all_leaves = query.order_by(LeaveRequest.created_at.desc()).all()
    
    # تحميل البيانات المرتبطة يدوياً
    for leave in all_leaves:
        leave.user = db_session.query(User).get(leave.user_id)
        leave.department = db_session.query(Department).get(leave.department_id)
    
    # باقي الكود يبقى كما هو...
    # تصفية حسب الشهر المالي إذا تم تحديده
    filtered_leaves = []
    if month_filter:
        try:
            # تحويل month_filter إلى تاريخ (تنسيق: YYYY-MM)
            year, month = map(int, month_filter.split('-'))
            
            # حساب بداية ونهاية الشهر المالي (26 إلى 25)
            if month == 12:
                financial_month_start = date(year, 12, 26)
                financial_month_end = date(year + 1, 1, 25)
            else:
                financial_month_start = date(year, month, 26)
                financial_month_end = date(year, month + 1, 25)
            
            # تصفية الإجازات التي تتداخل مع الشهر المالي
            for leave in all_leaves:
                if (leave.start_date <= financial_month_end and 
                    leave.end_date >= financial_month_start):
                    filtered_leaves.append(leave)
        except ValueError:
            filtered_leaves = all_leaves
    else:
        filtered_leaves = all_leaves
    
    # إحصائيات
    stats = {
        'total': len(filtered_leaves),
        'approved': len([l for l in filtered_leaves if l.status == 'approved']),
        'pending': len([l for l in filtered_leaves if l.status == 'pending']),
        'rejected': len([l for l in filtered_leaves if l.status == 'rejected']),
        'current_month': len([l for l in filtered_leaves if is_in_current_financial_month(l.start_date)])
    }
    
    # تجميع البيانات للعرض
    leaves_by_department = {}
    for leave in filtered_leaves:
        dept_name = leave.department.name if leave.department else 'غير معين'
        if dept_name not in leaves_by_department:
            leaves_by_department[dept_name] = []
        leaves_by_department[dept_name].append(leave)
    
    # الحصول على الأقسام للفلتر
    departments = db_session.query(Department).all()
    
    # توليد قائمة الأشهر المالية للفلتر
    financial_months = generate_financial_months()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_all_leaves.html',
                         leaves=filtered_leaves,
                         leaves_by_department=leaves_by_department,
                         departments=departments,
                         financial_months=financial_months,
                         selected_month=month_filter,
                         selected_department=department_filter,
                         selected_status=status_filter,
                         stats=stats,
                         notifications=notifications)


def get_approver_name(approver_id):
    """الحصول على اسم الموافق على الإجازة"""
    if not approver_id:
        return "غير معين"
    
    approver = db_session.query(User).get(approver_id)
    return approver.name if approver else "غير معين"

@app.context_processor
def utility_processor():
    def get_approver_name(approver_id):
        if not approver_id:
            return "غير معين"
        approver = db_session.query(User).get(approver_id)
        return approver.name if approver else "غير معين"
    
    return dict(get_approver_name=get_approver_name)

def generate_financial_months():
    """توليد قائمة بالأشهر المالية للسنوات السابقة والحالية والمستقبلية"""
    months = []
    current_date = date.today()
    current_year = current_date.year
    
    # إضافة الأشهر للعام الحالي
    for month_num in range(1, 13):
        # تنسيقين: YYYY-MM للفلتر و MM للعرض البسيط
        year_month = f"{current_year}-{month_num:02d}"
        month_only = str(month_num)
        
        month_name = get_month_name_arabic(month_num)
        prev_month_name = get_month_name_arabic(month_num-1 if month_num>1 else 12)
        prev_year = current_year if month_num > 1 else current_year - 1
        
        display_name = f"{month_name} {current_year} (26 {prev_month_name} {prev_year} - 25 {month_name} {current_year})"
        
        months.append({
            'value': year_month,  # للفلتر
            'simple_value': month_only,  # للتوافق مع القيم القديمة
            'display': display_name
        })
    
    # تحديد الشهر الحالي
    if current_date.day >= 26:
        current_financial_month = current_date.month
    else:
        current_financial_month = current_date.month - 1 if current_date.month > 1 else 12
    
    # وضع علامة على الشهر الحالي
    for month in months:
        if month['simple_value'] == str(current_financial_month):
            month['display'] += " - الحالي"
            break
    
    # إضافة أشهر من العام الماضي
    for month_num in range(1, 13):
        year_month = f"{current_year-1}-{month_num:02d}"
        month_only = str(month_num)
        
        month_name = get_month_name_arabic(month_num)
        prev_month_name = get_month_name_arabic(month_num-1 if month_num>1 else 12)
        prev_year = current_year-1 if month_num > 1 else current_year - 2
        
        display_name = f"{month_name} {current_year-1} (26 {prev_month_name} {prev_year} - 25 {month_name} {current_year-1})"
        
        months.insert(0, {
            'value': year_month,
            'simple_value': month_only,
            'display': display_name
        })
    
    return months

def get_month_name_arabic(month):
    """إرجاع اسم الشهر بالعربية"""
    months = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    return months.get(month, '')

def is_in_current_financial_month(leave_date):
    """التحقق مما إذا كان تاريخ الإجارة يقع في الشهر المالي الحالي"""
    today = date.today()
    
    # حساب بداية ونهاية الشهر المالي الحالي
    if today.day >= 26:
        # الشهر المالي يبدأ من 26 من الشهر الحالي إلى 25 من الشهر القادم
        financial_month_start = date(today.year, today.month, 26)
        if today.month == 12:
            financial_month_end = date(today.year + 1, 1, 25)
        else:
            financial_month_end = date(today.year, today.month + 1, 25)
    else:
        # الشهر المالي يبدأ من 26 من الشهر الماضي إلى 25 من الشهر الحالي
        if today.month == 1:
            financial_month_start = date(today.year - 1, 12, 26)
            financial_month_end = date(today.year, 1, 25)
        else:
            financial_month_start = date(today.year, today.month - 1, 26)
            financial_month_end = date(today.year, today.month, 25)
    
    return financial_month_start <= leave_date <= financial_month_end


@app.route('/admin/export_leaves')
@login_required
def admin_export_leaves():
    """تصدير بيانات الإجازات إلى Excel - بدون استخدام العلاقات"""

    try:
        # الحصول على معاملات التصفية
        month_filter = request.args.get('month', '')
        department_filter = request.args.get('department', '')
        status_filter = request.args.get('status', '')
        
        # بناء الاستعلام الأساسي
        query = db_session.query(LeaveRequest)
        
        if department_filter and department_filter.isdigit():
            query = query.filter(LeaveRequest.department_id == int(department_filter))
        
        if status_filter:
            query = query.filter(LeaveRequest.status == status_filter)
        
        all_leaves = query.order_by(LeaveRequest.created_at.desc()).all()
        
        # الحصول على بيانات المستخدمين والأقسام مسبقاً
        user_ids = [leave.user_id for leave in all_leaves]
        department_ids = [leave.department_id for leave in all_leaves]
        
        users = {user.id: user for user in db_session.query(User).filter(User.id.in_(user_ids)).all()}
        departments = {dept.id: dept for dept in db_session.query(Department).filter(Department.id.in_(department_ids)).all()}
        
        # تطبيق فلتر الشهر المالي
        filtered_leaves = []
        if month_filter:
            try:
                year, month = map(int, month_filter.split('-'))
                if month == 12:
                    financial_month_start = date(year, 12, 26)
                    financial_month_end = date(year + 1, 1, 25)
                else:
                    financial_month_start = date(year, month, 26)
                    financial_month_end = date(year, month + 1, 25)
                
                for leave in all_leaves:
                    if (leave.start_date <= financial_month_end and 
                        leave.end_date >= financial_month_start):
                        filtered_leaves.append(leave)
            except ValueError:
                filtered_leaves = all_leaves
        else:
            filtered_leaves = all_leaves
        
        # باقي كود التصدير يبقى كما هو...
        # إنشاء ملف Excel
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "تقرير الإجازات"
        
        # إضافة العنوان
        title = "تقرير الإجازات"
        if month_filter:
            month_name = get_month_name_arabic(int(month_filter.split('-')[1]))
            year = month_filter.split('-')[0]
            title += f" - {month_name} {year}"
        
        if department_filter:
            department = db_session.query(Department).get(int(department_filter))
            if department:
                title += f" - قسم {department.name}"
        
        if status_filter:
            status_names = {'approved': 'المقبولة', 'pending': 'المعلقة', 'rejected': 'المرفوضة'}
            title += f" - {status_names.get(status_filter, status_filter)}"
        
        worksheet.merge_cells('A1:H1')
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # رؤوس الأعمدة
        headers = [
            'اسم الموظف', 'القسم', 'نوع الإجازة', 'تاريخ البداية', 
            'تاريخ النهاية', 'عدد الأيام', 'الحالة', 'ملاحظات'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # بيانات الإجازات
        row = 4
        for leave in filtered_leaves:
            user = users.get(leave.user_id)
            department = departments.get(leave.department_id)
            
            worksheet.cell(row=row, column=1, value=user.name if user else 'غير معين')
            worksheet.cell(row=row, column=2, value=department.name if department else 'غير معين')
            worksheet.cell(row=row, column=3, value=leave.leave_type)
            worksheet.cell(row=row, column=4, value=leave.start_date.strftime('%Y-%m-%d'))
            worksheet.cell(row=row, column=5, value=leave.end_date.strftime('%Y-%m-%d'))
            worksheet.cell(row=row, column=6, value=leave.total_days)
            
            # تحويل الحالة للعربية
            status_arabic = {
                'approved': 'مقبولة',
                'pending': 'معلقة', 
                'rejected': 'مرفوضة'
            }.get(leave.status, leave.status)
            
            worksheet.cell(row=row, column=7, value=status_arabic)
            worksheet.cell(row=row, column=8, value=leave.reason or '')
            
            row += 1
        
        # ضبط عرض الأعمدة
        column_widths = [25, 20, 15, 15, 15, 12, 12, 30]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width
        
        workbook.save(output)
        output.seek(0)
        
        filename = f"تقرير_الإجازات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'خطأ في تصدير الملف: {str(e)}', 'error')
        return redirect(url_for('admin_all_leaves'))

# في قسم inject_manager_stats (البحث عن هذا القسم في app.py):
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404



@app.context_processor
def inject_manager_stats():
    """Inject manager stats into all templates"""
    try:
        if current_user.is_authenticated and hasattr(current_user, 'is_manager') and current_user.is_manager:
            # استخدام primary_manager_id بدلاً من manager_id
            managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
            department_ids = [dept.id for dept in managed_departments]
            
            stats = {
                'pending_leaves': db_session.query(LeaveRequest).filter(
                    LeaveRequest.department_id.in_(department_ids),
                    LeaveRequest.status == 'pending'
                ).count(),
                'pending_permissions': db_session.query(PermissionRequest).filter(
                    PermissionRequest.department_id.in_(department_ids),
                    PermissionRequest.status == 'pending'
                ).count(),
                'pending_advances': db_session.query(AdvanceRequest).filter(
                    AdvanceRequest.department_id.in_(department_ids),
                    AdvanceRequest.status == 'pending'
                ).count(),
                'department_employees': db_session.query(User).filter(
                    User.department_id.in_(department_ids),
                    User.is_admin == False
                ).count()
            }
            return {'stats': stats}
    except Exception as e:
        print(f"Error in inject_manager_stats: {e}")
    
    return {'stats': {}}


@app.route('/admin/all_permissions')
@login_required
def admin_all_permissions():
    """عرض جميع طلبات الإذن في كل قسم حسب الشهر المالي"""
    # الحصول على معاملات التصفية
    month_filter = request.args.get('month', '')
    department_filter = request.args.get('department', '')
    status_filter = request.args.get('status', '')
    
    # بناء الاستعلام الأساسي
    query = db_session.query(PermissionRequest)
    
    # تطبيق الفلاتر
    if department_filter and department_filter.isdigit():
        query = query.filter(PermissionRequest.department_id == int(department_filter))
    
    if status_filter:
        query = query.filter(PermissionRequest.status == status_filter)
    
    # الحصول على جميع طلبات الإذن
    all_permissions = query.order_by(PermissionRequest.created_at.desc()).all()
    
    # تحميل البيانات المرتبطة يدوياً
    for permission in all_permissions:
        permission.user = db_session.query(User).get(permission.user_id)
        permission.department = db_session.query(Department).get(permission.department_id)
    
    # تصفية حسب الشهر المالي إذا تم تحديده
    filtered_permissions = []
    if month_filter:
        try:
            # تحويل month_filter إلى تاريخ (تنسيق: YYYY-MM)
            year, month = map(int, month_filter.split('-'))
            
            # حساب بداية ونهاية الشهر المالي (26 إلى 25)
            if month == 12:
                financial_month_start = date(year, 12, 26)
                financial_month_end = date(year + 1, 1, 25)
            else:
                financial_month_start = date(year, month, 26)
                financial_month_end = date(year, month + 1, 25)
            
            # تصفية طلبات الإذن التي تتداخل مع الشهر المالي
            for permission in all_permissions:
                if (permission.date <= financial_month_end and 
                    permission.date >= financial_month_start):
                    filtered_permissions.append(permission)
        except ValueError:
            filtered_permissions = all_permissions
    else:
        filtered_permissions = all_permissions
    
    # إحصائيات
    stats = {
        'total': len(filtered_permissions),
        'approved': len([p for p in filtered_permissions if p.status == 'approved']),
        'pending': len([p for p in filtered_permissions if p.status == 'pending']),
        'rejected': len([p for p in filtered_permissions if p.status == 'rejected']),
        'current_month': len([p for p in filtered_permissions if is_in_current_financial_month(p.date)])
    }
    
    # تجميع البيانات للعرض
    permissions_by_department = {}
    for permission in filtered_permissions:
        dept_name = permission.department.name if permission.department else 'غير معين'
        if dept_name not in permissions_by_department:
            permissions_by_department[dept_name] = []
        permissions_by_department[dept_name].append(permission)
    
    # الحصول على الأقسام للفلتر
    departments = db_session.query(Department).all()
    
    # توليد قائمة الأشهر المالية للفلتر
    financial_months = generate_financial_months()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_all_permissions.html',
                         permissions=filtered_permissions,
                         permissions_by_department=permissions_by_department,
                         departments=departments,
                         financial_months=financial_months,
                         selected_month=month_filter,
                         selected_department=department_filter,
                         selected_status=status_filter,
                         stats=stats,
                         notifications=notifications)

@app.route('/admin/export_permissions')
@login_required
def admin_export_permissions():
    """تصدير بيانات طلبات الإذن إلى Excel"""

    try:
        # الحصول على معاملات التصفية
        month_filter = request.args.get('month', '')
        department_filter = request.args.get('department', '')
        status_filter = request.args.get('status', '')
        
        # بناء الاستعلام الأساسي
        query = db_session.query(PermissionRequest)
        
        if department_filter and department_filter.isdigit():
            query = query.filter(PermissionRequest.department_id == int(department_filter))
        
        if status_filter:
            query = query.filter(PermissionRequest.status == status_filter)
        
        all_permissions = query.order_by(PermissionRequest.created_at.desc()).all()
        
        # الحصول على بيانات المستخدمين والأقسام مسبقاً
        user_ids = [permission.user_id for permission in all_permissions]
        department_ids = [permission.department_id for permission in all_permissions]
        
        users = {user.id: user for user in db_session.query(User).filter(User.id.in_(user_ids)).all()}
        departments = {dept.id: dept for dept in db_session.query(Department).filter(Department.id.in_(department_ids)).all()}
        
        # تطبيق فلتر الشهر المالي
        filtered_permissions = []
        if month_filter:
            try:
                year, month = map(int, month_filter.split('-'))
                if month == 12:
                    financial_month_start = date(year, 12, 26)
                    financial_month_end = date(year + 1, 1, 25)
                else:
                    financial_month_start = date(year, month, 26)
                    financial_month_end = date(year, month + 1, 25)
                
                for permission in all_permissions:
                    if (permission.date <= financial_month_end and 
                        permission.date >= financial_month_start):
                        filtered_permissions.append(permission)
            except ValueError:
                filtered_permissions = all_permissions
        else:
            filtered_permissions = all_permissions
        
        # إنشاء ملف Excel
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "تقرير الإذونات"
        
        # إضافة العنوان
        title = "تقرير طلبات الإذن"
        if month_filter:
            month_name = get_month_name_arabic(int(month_filter.split('-')[1]))
            year = month_filter.split('-')[0]
            title += f" - {month_name} {year}"
        
        if department_filter:
            department = db_session.query(Department).get(int(department_filter))
            if department:
                title += f" - قسم {department.name}"
        
        if status_filter:
            status_names = {'approved': 'المقبولة', 'pending': 'المعلقة', 'rejected': 'المرفوضة'}
            title += f" - {status_names.get(status_filter, status_filter)}"
        
        worksheet.merge_cells('A1:H1')
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # رؤوس الأعمدة
        headers = [
            'اسم الموظف', 'القسم', 'نوع الإذن', 'التاريخ', 
            'المدة (دقيقة)', 'الحالة', 'السبب'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # بيانات طلبات الإذن
        row = 4
        for permission in filtered_permissions:
            user = users.get(permission.user_id)
            department = departments.get(permission.department_id)
            
            worksheet.cell(row=row, column=1, value=user.name if user else 'غير معين')
            worksheet.cell(row=row, column=2, value=department.name if department else 'غير معين')
            worksheet.cell(row=row, column=3, value=permission.permission_type)
            worksheet.cell(row=row, column=4, value=permission.date.strftime('%Y-%m-%d'))
            worksheet.cell(row=row, column=5, value=permission.time or '-')
            
            # تحويل الحالة للعربية
            status_arabic = {
                'approved': 'مقبولة',
                'pending': 'معلقة', 
                'rejected': 'مرفوضة'
            }.get(permission.status, permission.status)
            
            worksheet.cell(row=row, column=6, value=status_arabic)
            worksheet.cell(row=row, column=7, value=permission.reason or '')
            
            row += 1
        
        # ضبط عرض الأعمدة
        column_widths = {
            'A': 25,  # اسم الموظف
            'B': 20,  # القسم
            'C': 15,  # نوع الإذن
            'D': 12,  # التاريخ
            'E': 10,  # من
            'F': 10,  # إلى
            'G': 15,  # المدة
            'H': 12,  # الحالة
            'I': 25   # السبب
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        workbook.save(output)
        output.seek(0)
        
        filename = f"تقرير_الإذونات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'خطأ في تصدير الملف: {str(e)}', 'error')
        return redirect(url_for('admin_all_permissions'))



def parse_shift_info(shift_info_json):
    """تحليل معلومات الشيفتات من JSON"""
    if not shift_info_json:
        return []
    
    try:
        shift_data = json.loads(shift_info_json)
        return shift_data.get('shifts', [])
    except:
        return []

def get_shift_jobs(shift_info_json):
    """الحصول على الوظائف المحددة لكل شيفت"""
    if not shift_info_json:
        return {}
    
    try:
        shift_data = json.loads(shift_info_json)
        return shift_data.get('jobs', {})
    except:
        return {}



@app.route('/manager/approve_leave_partial/<int:request_id>', methods=['POST'])
@login_required
def approve_leave_partial(request_id):
    """موافقة جزئية على طلب إجازة (قبول شيفتات معينة)"""

    leave_request = db_session.query(LeaveRequest).filter_by(id=request_id).first()
    if not leave_request:
        flash('طلب الإجازة غير موجود', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_request.department_id)

    
    # الحصول على الشيفتات المقبولة والمرفوضة
    approved_shifts_json = request.form.get('approved_shifts', '[]')
    rejected_shifts_json = request.form.get('rejected_shifts', '[]')
    approval_notes = request.form.get('approval_notes', '')
    
    try:
        approved_shifts = json.loads(approved_shifts_json)
        rejected_shifts = json.loads(rejected_shifts_json)
        
        if not approved_shifts:
            flash('لم يتم تحديد أي شيفتات للموافقة', 'error')
            return redirect(url_for('manager_leave_requests'))
        
        # حساب عدد الشيفتات المقبولة والمرفوضة
        approved_count = len(approved_shifts)
        rejected_count = len(rejected_shifts)
        
        # إذا تم رفض جميع الشيفتات، اعتبر الطلب مرفوضاً بالكامل
        if approved_count == 0:
            leave_request.status = 'rejected'
            leave_request.rejection_reason = f'تم رفض جميع الشيفتات. {approval_notes}'
        # إذا تم قبول جميع الشيفتات، اعتبر الطلب مقبولاً بالكامل
        elif rejected_count == 0:
            leave_request.status = 'approved'
        # إذا كان هناك قبول ورفض جزئي، اعتبر الطلب مقبولاً جزئياً
        else:
            leave_request.status = 'partially_approved'
        
        # حفظ بيانات القبول/الرفض الجزئي
        leave_request.partial_leave_data = json.dumps({
            'approved_shifts': approved_shifts,
            'rejected_shifts': rejected_shifts,
            'approval_notes': approval_notes,
            'approved_count': approved_count,
            'rejected_count': rejected_count
        })
        
        leave_request.approved_shifts = json.dumps(approved_shifts)
        leave_request.rejected_shifts = json.dumps(rejected_shifts)
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        
        # إذا كانت الإجازة من رصيد الإجازات، خصم فقط الشيفتات المقبولة
        if leave_request.leave_type == 'من رصيد الإجازات' and approved_count > 0:
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=leave_request.user_id
            ).first()
            
            if balance:
                # خصم فقط الشيفتات المقبولة
                balance.leave_balance -= approved_count
                balance.last_updated = datetime.now()
        
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'تمت معالجة طلب إجازتك جزئياً. '
        notification_message += f'تم قبول {approved_count} شيفت ورفض {rejected_count} شيفت.'
        if approval_notes:
            notification_message += f' ملاحظات: {approval_notes}'
        
        create_notification(
            leave_request.user_id,
            'معالجة جزئية لطلب الإجازة',
            notification_message,
            'leave_partially_processed',
            related_id=leave_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        flash(f'تمت الموافقة على {approved_count} شيفت ورفض {rejected_count} شيفت بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء معالجة الطلب: {str(e)}', 'error')
    
    return redirect(url_for('manager_leave_requests'))



@app.route('/manager/reject_leave_partial/<int:request_id>', methods=['POST'])
@login_required
def reject_leave_partial(request_id):
    """رفض جزئي لطلب إجازة مع زيادة رصيد الموظف إذا كانت من رصيد الإجازات"""

    
    leave_request = db_session.query(LeaveRequest).filter_by(id=request_id).first()
    if not leave_request:
        flash('طلب الإجازة غير موجود', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.get(Department, leave_request.department_id)

    # التحقق من التاريخ (لا يمكن الرفض إلا في نفس الفترة الزمنية)
    leave_date = leave_request.leave_date if hasattr(leave_request, 'leave_date') and leave_request.leave_date else leave_request.start_date
    
    if not can_reject_leave_by_date(leave_date):
        # الحصول على معلومات الفترة الزمنية للعرض
        period_start, period_end = get_rejection_period_dates(leave_date)
        flash(
            f'لا يمكن رفض الإجازة إلا خلال الفترة من {period_start.strftime("%Y-%m-%d")} '
            f'إلى {period_end.strftime("%Y-%m-%d")}',
            'error'
        )
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_request.department_id)

    
    # التحقق من التاريخ (لا يمكن الرفض بعد أو في يوم 26)
    if not can_reject_leave_by_date(leave_request.leave_date or leave_request.start_date):
        flash('لا يمكن رفض الإجازة بعد أو في يوم 26 من الشهر الحالي', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # الحصول على بيانات الرفض
    rejected_shifts_json = request.form.get('rejected_shifts', '[]')
    approved_shifts_json = request.form.get('approved_shifts', '[]')
    rejection_reason = request.form.get('rejection_reason', '').strip()
    
    if not rejection_reason:
        flash('يرجى كتابة سبب الرفض', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    try:
        rejected_shifts = json.loads(rejected_shifts_json)
        approved_shifts = json.loads(approved_shifts_json)
        
        if not rejected_shifts:
            flash('لم يتم تحديد أي شيفتات للرفض', 'error')
            return redirect(url_for('manager_leave_requests'))
        
        rejected_count = len(rejected_shifts)
        approved_count = len(approved_shifts)
        
        # تحديد حالة الطلب النهائية
        if approved_count == 0:
            leave_request.status = 'rejected'
        elif rejected_count == 0:
            leave_request.status = 'approved'
        else:
            leave_request.status = 'partially_rejected'
        
        # حفظ بيانات الرفض الجزئي
        leave_request.partial_leave_data = json.dumps({
            'rejected_shifts': rejected_shifts,
            'approved_shifts': approved_shifts,
            'rejection_reason': rejection_reason,
            'rejected_count': rejected_count,
            'approved_count': approved_count
        })
        
        leave_request.rejected_shifts = json.dumps(rejected_shifts)
        leave_request.approved_shifts = json.dumps(approved_shifts)
        leave_request.rejection_reason = rejection_reason
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        
        # زيادة رصيد الموظف إذا كانت الإجازة من رصيد الإجازات
        balance_increased = False
        new_balance = None
        
        if leave_request.leave_type == 'من رصيد الإجازات' and rejected_count > 0:
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=leave_request.user_id
            ).first()
            
            if balance:
                old_balance = balance.leave_balance
                balance.leave_balance += rejected_count
                balance.last_updated = datetime.now()
                new_balance = balance.leave_balance
                balance_increased = True
        
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'تم رفض {rejected_count} شيفت من طلب إجازتك. السبب: {rejection_reason}'
        if balance_increased:
            notification_message += f'\nتم زيادة رصيدك بمقدار {rejected_count} يوم (الرصيد الجديد: {new_balance} يوم)'
        
        create_notification(
            leave_request.user_id,
            'رفض جزئي لطلب الإجازة',
            notification_message,
            'leave_partially_rejected',
            related_id=leave_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        success_message = f'تم رفض {rejected_count} شيفت بنجاح'
        if balance_increased:
            success_message += f' وتم زيادة رصيد الموظف بمقدار {rejected_count} يوم'
        
        flash(success_message, 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء رفض الطلب: {str(e)}', 'error')
    
    return redirect(url_for('manager_leave_requests'))



def update_employee_balance_on_rejection(leave_request, rejected_shifts_count):
    """زيادة رصيد الموظف عند رفض إجازة من رصيد الإجازات"""
    if leave_request.leave_type == 'من رصيد الإجازات' and rejected_shifts_count > 0:
        balance = db_session.query(EmployeeBalance).filter_by(
            user_id=leave_request.user_id
        ).first()
        
        if balance:
            # زيادة الرصيد بعدد الشيفتات المرفوضة
            balance.leave_balance += rejected_shifts_count
            balance.last_updated = datetime.now()
            db_session.commit()
            
            return True, balance.leave_balance
    
    return False, None


@app.route('/manager/approve_leave/<int:request_id>')
@login_required
def approve_leave(request_id):
    """موافقة على طلب إجازة فردي"""
    
    leave_request = db_session.get(LeaveRequest, request_id)
    
    if not leave_request:
        flash('طلب الإجازة غير موجود', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_request.department_id)
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    
    if leave_request.status != 'pending':
        flash('تم معالجة هذا الطلب مسبقاً', 'warning')
        return redirect(url_for('manager_leave_requests'))
    
    try:
        # موافقة على الشيفت الفردي
        leave_request.status = 'approved'
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        
        # إذا كانت الإجازة من رصيد الإجازات، خصم من الرصيد
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=leave_request.user_id
            ).first()
            if balance:
                balance.leave_balance -= 1  # خصم يوم واحد لكل شيفت
                balance.last_updated = datetime.now()
        
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'تم الموافقة على شيفت إجازتك ({leave_request.shift_name})'
        if leave_request.shift_job:
            notification_message += f' للوظيفة {leave_request.shift_job}'
        
        create_notification(
            leave_request.user_id,
            'موافقة على طلب الإجازة',
            notification_message,
            'leave_approved',
            related_id=leave_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        flash('تم الموافقة على الشيفت بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء الموافقة على الطلب: {str(e)}', 'error')
    
    return redirect(url_for('manager_leave_requests'))




@app.context_processor
def utility_processor():
    def can_reject_leave(leave_request):
        """دالة مساعدة للقالب للتحقق من إمكانية رفض الإجازة"""
        leave_date = getattr(leave_request, 'leave_date', None) or leave_request.start_date
        return can_reject_leave_by_date(leave_date)
    
    def get_month_name_arabic(month):
        """إرجاع اسم الشهر بالعربية"""
        months = {
            1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
            5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
            9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
        }
        return months.get(month, '')
    
    return dict(
        can_reject_leave=can_reject_leave,
        get_month_name_arabic=get_month_name_arabic
    )


@app.route('/api/reject_leave/<int:request_id>', methods=['POST'])
@login_required
def api_reject_leave(request_id):
    """API لرفض طلب إجازة (للاستخدام مع AJAX)"""
    
    leave_request = db_session.query(LeaveRequest).filter_by(id=request_id).first()
    if not leave_request:
        return jsonify({'success': False, 'message': 'طلب الإجازة غير موجود'})
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_request.department_id)
    
    # التحقق من أن الطلب معلق
    if leave_request.status != 'pending':
        return jsonify({'success': False, 'message': 'تم معالجة هذا الطلب مسبقاً'})
    
    # الحصول على البيانات من طلب AJAX
    data = request.get_json()
    if not data or 'rejection_reason' not in data:
        return jsonify({'success': False, 'message': 'بيانات غير كاملة'})
    
    rejection_reason = data['rejection_reason'].strip()
    if len(rejection_reason) < 10:
        return jsonify({'success': False, 'message': 'سبب الرفض يجب أن يكون على الأقل 10 أحرف'})
    
    try:
        # تحديث حالة الطلب
        leave_request.status = 'rejected'
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        leave_request.rejection_reason = rejection_reason
        
        # زيادة الرصيد إذا كانت الإجازة من الرصيد
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=leave_request.user_id
            ).first()
            
            if balance:
                # زيادة الرصيد
                balance.leave_balance += leave_request.total_days
                balance.last_updated = datetime.now()
                
                # يمكنك إضافة سجل للعملية
                print(f"تم زيادة رصيد الموظف {leave_request.user_id} بمقدار {leave_request.total_days} يوم")
        
        db_session.commit()
        
        # إرسال إشعار
        create_notification(
            leave_request.user_id,
            'تم رفض طلب الإجازة',
            f'تم رفض طلب إجازتك. السبب: {rejection_reason[:50]}...',
            'leave_rejected',
            related_id=leave_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        return jsonify({
            'success': True,
            'message': f'تم رفض طلب الإجازة بنجاح{" وتم زيادة رصيد الموظف" if leave_request.leave_type == "من رصيد الإجازات" else ""}',
            'new_balance': balance.leave_balance if balance and leave_request.leave_type == 'من رصيد الإجازات' else None
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})
    


@app.route('/manager/approve_permission/<int:request_id>')
@login_required
def manager_approve_permission(request_id):
    """موافقة المدير على طلب إذن مع معالجة الأنواع الجديدة"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    permission_request = db_session.get(PermissionRequest, request_id)
    
    # Check if manager manages this department
    department = db_session.get(Department, permission_request.department_id) if permission_request else None

    
    if permission_request.status == 'pending':
        permission_request.status = 'approved'
        permission_request.approved_by = current_user.id
        permission_request.approved_at = datetime.now()
        
        # معالجة خاصة لأنواع الإذن الجديدة
        extra_data = {}
        if permission_request.extra_data:
            try:
                extra_data = json.loads(permission_request.extra_data)
            except:
                pass
        
        # إرسال إشعار خاص للموظف بناءً على نوع الإذن
        notification_message = 'تم الموافقة على طلب الإذن الخاص بك'
        
        if permission_request.permission_type == 'اذن طلب ساعات اضافي':
            hours = extra_data.get('hours', 0)
            shifts = extra_data.get('shifts', [])
            notification_message = f'تم الموافقة على طلب ساعات إضافية لمدة {hours} ساعات للشيفتات: {", ".join(shifts)}'
        
        elif permission_request.permission_type == 'اذن تبديل وردية':
            shifts = extra_data.get('shifts', [])
            notification_message = f'تم الموافقة على طلب تبديل وردية للشيفتات: {", ".join(shifts)}'
            
            # يمكنك هنا إضافة منطق إضافي لتحديث جداول العمل
        
        # Notify employee
        create_notification(
            permission_request.user_id,
            'موافقة على طلب الإذن',
            notification_message,
            'permission_approved',
            action_url=url_for('user_leave_requests')
        )
        
        db_session.commit()
        flash('تم الموافقة على طلب الإذن')
    
    return redirect(url_for('manager_permission_requests'))



@app.route('/export/permission_report')
@login_required
def export_permission_report():
    """تصدير تقرير الإذونات إلى Excel"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        department_ids = [dept.id for dept in managed_departments]
    
        
        # الحصول على معاملات التصفية
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        # بناء الاستعلام
        query = db_session.query(PermissionRequest)\
            .options(
                joinedload(PermissionRequest.user),
                joinedload(PermissionRequest.department),
                joinedload(PermissionRequest.approver)
            )\
            .filter(PermissionRequest.department_id.in_(department_ids))
        
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                query = query.filter(PermissionRequest.created_at >= start_date)
            except ValueError:
                pass
        
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                query = query.filter(PermissionRequest.created_at <= end_date)
            except ValueError:
                pass
        
        permissions = query.order_by(PermissionRequest.created_at.desc()).all()
        
        # إنشاء ملف Excel
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "تقرير الإذونات"
        
        # إضافة العنوان
        title = f"تقرير الإذونات - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        if start_date_str and end_date_str:
            title += f" ({start_date_str} إلى {end_date_str})"
        
        worksheet.merge_cells('A1:H1')
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # إضافة رؤوس الأعمدة
        headers = ['اسم الموظف', 'القسم', 'نوع الإذن', 'التاريخ', 
                  'المدة (دقيقة)', 'السبب', 'الحالة', 'تاريخ الطلب']
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # إضافة البيانات
        row = 4
        for perm in permissions:
            worksheet.cell(row=row, column=1, value=perm.user.name if perm.user else 'غير معروف')
            worksheet.cell(row=row, column=2, value=perm.department.name if perm.department else 'غير معين')
            worksheet.cell(row=row, column=3, value=perm.permission_type)
            worksheet.cell(row=row, column=4, value=perm.date.strftime('%Y-%m-%d') if perm.date else '')
            worksheet.cell(row=row, column=5, value=perm.time or '')
            worksheet.cell(row=row, column=6, value=perm.reason or '')
            
            # تحويل الحالة للعربية
            status_arabic = {
                'approved': 'مقبولة',
                'pending': 'معلقة',
                'rejected': 'مرفوضة'
            }.get(perm.status, perm.status)
            worksheet.cell(row=row, column=7, value=status_arabic)
            
            worksheet.cell(row=row, column=8, value=perm.created_at.strftime('%Y-%m-%d %H:%M') if perm.created_at else '')
            row += 1
        
        # ضبط عرض الأعمدة
        column_widths = [25, 20, 20, 15, 15, 30, 15, 20]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width
        
        workbook.save(output)
        output.seek(0)
        
        # إنشاء اسم الملف
        filename = f"تقرير_الإذونات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'خطأ في تصدير التقرير: {str(e)}', 'error')
        return redirect(url_for('manager_permission_requests'))

@app.template_filter('parse_permission_extra_data')
def parse_permission_extra_data_filter(extra_data_json):
    """فلتر لعرض البيانات الإضافية لطلب الإذن"""
    if not extra_data_json:
        return {}
    try:
        return json.loads(extra_data_json)
    except:
        return {}


@app.context_processor
def utility_processor():
    def parse_permission_extra_data(extra_data_json):
        if not extra_data_json:
            return {}
        try:
            return json.loads(extra_data_json)
        except:
            return {}
    
    return dict(
        parse_permission_extra_data=parse_permission_extra_data,
        get_approver_name=get_approver_name
    )

@app.route('/manager/reject_permission/<int:request_id>', methods=['POST'])
@login_required
def reject_permission(request_id):
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    permission_request = db_session.get(PermissionRequest, request_id)
    
    # Check if manager manages this department
    department = db_session.get(Department, permission_request.department_id) if permission_request else None
    
    if permission_request.status == 'pending':
        permission_request.status = 'rejected'
        permission_request.approved_by = current_user.id
        permission_request.approved_at = datetime.now()
        permission_request.rejection_reason = request.form.get('rejection_reason')
        
        # Notify employee
        create_notification(
            permission_request.user_id,
            'رفض طلب الإذن',
            'تم رفض طلب الإذن الخاص بك',
            'permission_rejected',
            action_url=url_for('user_leave_requests')
        )
        
        db_session.commit()
        flash('تم رفض طلب الإذن')
    
    return redirect(url_for('manager_permission_requests'))

@app.route('/manager/approve_advance/<int:request_id>')
@login_required
def approve_advance(request_id):
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    advance_request = db_session.query(AdvanceRequest).get(request_id)
    
    # Check if manager manages this department
    department = db_session.query(Department).get(advance_request.department_id)
    
    if advance_request.status == 'pending':
        advance_request.status = 'approved'
        advance_request.approved_by = current_user.id
        advance_request.approved_at = datetime.now()
        
        # Update balance
        balance = db_session.query(EmployeeBalance).filter_by(user_id=advance_request.user_id).first()
        if balance:
            balance.advance_balance += advance_request.amount
            balance.last_updated = datetime.now()
        
        # Notify employee
        create_notification(
            advance_request.user_id,
            'موافقة على طلب السلفة',
            f'تم الموافقة على طلب سلفتك بقيمة {advance_request.amount} جنيه',
            'advance_approved',
            action_url=url_for('user_advance_requests')
        )
        
        db_session.commit()
        flash('تم الموافقة على طلب السلفة')
    
    return redirect(url_for('manager_advance_requests'))

@app.route('/manager/reject_advance/<int:request_id>', methods=['POST'])
@login_required
def reject_advance(request_id):
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    advance_request = db_session.query(AdvanceRequest).get(request_id)
    
    # Check if manager manages this department
    department = db_session.query(Department).get(advance_request.department_id)
    
    if advance_request.status == 'pending':
        advance_request.status = 'rejected'
        advance_request.approved_by = current_user.id
        advance_request.approved_at = datetime.now()
        advance_request.rejection_reason = request.form.get('rejection_reason')
        
        # Notify employee
        create_notification(
            advance_request.user_id,
            'رفض طلب السلفة',
            'تم رفض طلب السلفة الخاص بك',
            'advance_rejected',
            action_url=url_for('user_advance_requests')
        )
        
        db_session.commit()
        flash('تم رفض طلب السلفة')
    
    return redirect(url_for('manager_advance_requests'))

# ======== Notification Routes ========

@app.route('/notifications')
@login_required
def notifications():
    notifications_list = db_session.query(Notification).filter_by(
        user_id=current_user.id
    ).order_by(Notification.created_at.desc()).all()
    
    # Get only unread notifications
    unread_notifications = [n for n in notifications_list if not n.is_read]
    unread_count = len(unread_notifications)
    
    return render_template('notifications.html', 
                         notifications=notifications_list,
                         unread_notifications=unread_notifications,
                         unread_count=unread_count)

@app.route('/mark_notification_read/<int:notification_id>')
@login_required
def mark_notification_read(notification_id):
    notification = db_session.query(Notification).get(notification_id)
    if notification and notification.user_id == current_user.id:
        notification.is_read = True
        db_session.commit()
    
    return redirect(request.referrer or url_for('notifications'))

@app.route('/mark_all_read')
@login_required
def mark_all_read():
    notifications_list = db_session.query(Notification).filter_by(
        user_id=current_user.id, is_read=False
    ).all()
    
    for notification in notifications_list:
        notification.is_read = True
    
    db_session.commit()
    flash('تم تعيين جميع الإشعارات كمقروءة')
    return redirect(url_for('notifications'))

# ======== API Routes ========

@app.route('/api/employee_progress')
@login_required
def api_employee_progress():
    if current_user.is_admin:
        return jsonify({'success': False})
    
    employee_data = db_session.query(EmployeeData).filter_by(user_id=current_user.id).first()
    
    if employee_data:
        completion_percentage = employee_data.calculate_completion()
        missing_fields = employee_data.get_missing_fields()
        db_session.commit()
    else:
        completion_percentage = 0
        missing_fields = []
    
    return jsonify({
        'completion_percentage': completion_percentage,
        'missing_fields': missing_fields
    })


@app.route('/department/employees')
@login_required
def department_employees():
    """عرض جميع موظفي القسم"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    # الحصول على معرف القسم من المعاملات
    dept_id = request.args.get('dept_id', current_user.department_id)
    
    # جلب القسم
    department = db_session.get(Department, dept_id)
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('manager_dashboard' if current_user.is_manager else 'admin_dashboard'))
    
    # جلب موظفي القسم
    employees = db_session.query(User).filter(
        User.department_id == dept_id,
        User.is_admin == False,
        User.is_active == True
    ).order_by(User.name).all()
    
    # جلب بيانات الموظفين الإضافية
    for emp in employees:
        emp.employee_data = db_session.query(EmployeeData).filter_by(user_id=emp.id).first()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('department_employees.html',
                         department=department,
                         employees=employees,
                         notifications=notifications)


# ======== Admin User Data View ========
@app.route('/admin/view_user/<int:user_id>')
@login_required
def view_user_data(user_id):
    if not current_user.is_admin and not current_user.is_manager:
        return redirect(url_for('user_dashboard'))
    
    user = db_session.get(User, user_id)
    employee_data = db_session.query(EmployeeData).filter_by(user_id=user_id).first()
    
    if employee_data:
        # Use the method from EmployeeData model if it exists
        missing_fields = employee_data.get_missing_fields()
    else:
        missing_fields = []
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_view_user.html', 
                         user=user, 
                         employee_data=employee_data,
                         missing_fields=missing_fields,
                         notifications=notifications)

@app.route('/admin/export_user/<int:user_id>')
@login_required
def export_user_data(user_id):
    if not current_user.is_admin and not current_user.is_manager:
        return redirect(url_for('user_dashboard'))
    
    filepath = export_to_excel(user_id)
    if filepath:
        return send_file(filepath, as_attachment=True)
    else:
        flash('لا توجد بيانات للتصدير')
        return redirect(url_for('admin_users'))


@login_manager.user_loader
def load_user(user_id):
    try:
        return db_session.get(User, int(user_id))
    except Exception as e:
        print(f"Error loading user {user_id}: {e}")
        return None


@app.context_processor
def inject_timedelta():
    return dict(timedelta=timedelta)


def get_model_by_id(model, model_id):
    """دالة مساعدة للحصول على نموذج بواسطة المعرف مع التوافق مع SQLAlchemy 2.0"""
    try:
        return db_session.get(model, model_id)
    except Exception as e:
        print(f"Error getting {model.__name__} with id {model_id}: {e}")
        return None

# ======== Manager Routes ========

@app.route('/manager/dashboard')
@login_required
def manager_dashboard():
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('manager_dashboard'))
    
    managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
    department_ids = [dept.id for dept in managed_departments]
    
    # Get all users for department counts
    users = db_session.query(User).filter(User.department_id.in_(department_ids)).all()
    
    stats = {
        'pending_leaves': db_session.query(LeaveRequest).filter(
            LeaveRequest.department_id.in_(department_ids),
            LeaveRequest.status == 'pending'
        ).count(),
        'pending_permissions': db_session.query(PermissionRequest).filter(
            PermissionRequest.department_id.in_(department_ids),
            PermissionRequest.status == 'pending'
        ).count(),
        'pending_advances': db_session.query(AdvanceRequest).filter(
            AdvanceRequest.department_id.in_(department_ids),
            AdvanceRequest.status == 'pending'
        ).count(),
        'department_employees': db_session.query(User).filter(
            User.department_id.in_(department_ids),
            User.is_admin == False
        ).count()
    }
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager_dashboard.html',
                         managed_departments=managed_departments,
                         stats=stats,
                         users=users,
                         notifications=notifications)

@app.route('/manager/employees')
@login_required
def manager_employees():
    """إدارة موظفي القسم"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
    department_ids = [dept.id for dept in managed_departments]
    
    employees = db_session.query(User).filter(
        User.department_id.in_(department_ids),
        User.is_admin == False
    ).all()
    
    # Load employee data
    for employee in employees:
        employee.employee_data = db_session.query(EmployeeData).filter_by(user_id=employee.id).first()
        if employee.employee_data:
            employee.employee_data.calculate_completion()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager/manager_employees.html',
                         employees=employees,
                         notifications=notifications)

@app.context_processor
def utility_processor():
    def get_month_name_arabic(month):
        """إرجاع اسم الشهر بالعربية"""
        months = {
            1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
            5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
            9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
        }
        return months.get(month, '')
    
    return dict(get_month_name_arabic=get_month_name_arabic)


@app.route('/manager/export_leave_requests')
@login_required
def export_leave_requests():
    """تصدير طلبات الإجازات إلى Excel"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
    department_ids = [dept.id for dept in managed_departments]
    
    # الحصول على معاملات التصفية
    filter_status = request.args.get('status', 'all')
    filter_type = request.args.get('type', 'all')
    filter_month = request.args.get('month', 'all')
    
    # بناء الاستعلام الأساسي (نفس منطق manager_leave_requests)
    query = db_session.query(LeaveRequest).filter(
        LeaveRequest.department_id.in_(department_ids)
    )
    
    # تطبيق الفلاتر (نفس منطق manager_leave_requests)
    if filter_status != 'all':
        query = query.filter(LeaveRequest.status == filter_status)
    
    if filter_type != 'all':
        query = query.filter(LeaveRequest.leave_type == filter_type)
    
    if filter_month != 'all':
        try:
            if '-' not in filter_month:
                current_year = date.today().year
                month_num = int(filter_month)
                year_month = f"{current_year}-{month_num:02d}"
            else:
                year_month = filter_month
            
            year, month = map(int, year_month.split('-'))
            
            if month == 12:
                financial_month_start = date(year, 12, 26)
                financial_month_end = date(year + 1, 1, 25)
            else:
                financial_month_start = date(year, month, 26)
                financial_month_end = date(year, month + 1, 25)
            
            query = query.filter(
                LeaveRequest.start_date <= financial_month_end,
                LeaveRequest.end_date >= financial_month_start
            )
        except (ValueError, IndexError):
            pass
    
    leave_requests = query.order_by(LeaveRequest.created_at.desc()).all()
    
    # تحضير البيانات للتصدير
    data = []
    for req in leave_requests:
        user = db_session.get(User, req.user_id)
        department = db_session.get(Department, req.department_id)
        
        data.append({
            'اسم الموظف': user.name if user else 'غير معين',
            'القسم': department.name if department else 'غير معين',
            'نوع الإجازة': req.leave_type,
            'تاريخ البداية': req.start_date.strftime('%Y-%m-%d'),
            'تاريخ النهاية': req.end_date.strftime('%Y-%m-%d'),
            'عدد الأيام': req.total_days,
            'السبب': req.reason or '',
            'حالة الطلب': get_status_arabic(req.status),
            'تاريخ الطلب': req.created_at.strftime('%Y-%m-%d %H:%M'),
            'تاريخ القرار': req.approved_at.strftime('%Y-%m-%d %H:%M') if req.approved_at else '',
            'سبب الرفض': req.rejection_reason or ''
        })
    
    # إنشاء DataFrame
    df = pd.DataFrame(data)
    
    # إنشاء ملف Excel في الذاكرة
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='طلبات الإجازات', index=False)
        
        # تحسين تنسيق الأعمدة
        worksheet = writer.sheets['طلبات الإجازات']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # إنشاء اسم الملف
    filename = f"طلبات_الإجازات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def get_status_arabic(status):
    """تحويل حالة الطلب إلى العربية"""
    status_map = {
        'pending': 'معلقة',
        'approved': 'مقبولة',
        'rejected': 'مرفوضة'
    }
    return status_map.get(status, status)


from sqlalchemy.orm import joinedload, contains_eager


def can_reject_leave_by_date(leave_date):
    print(leave_date)
    """
    التحقق مما إذا كان يمكن رفض الإجازة بناءً على التاريخ
    القاعدة: يمكن رفض الإجازة فقط في نفس الفترة الزمنية (الشهر المالي)
    
    أمثلة:
    - إجازة 15 يناير 2025 ← تتبع لنطاق يناير (26 ديسمبر 2024 إلى 25 يناير 2025)
    - اليوم 20 يناير 2025 ← يقع في نفس النطاق ← يمكن الرفض
    - اليوم 26 يناير 2025 ← يقع في نطاق فبراير ← لا يمكن الرفض
    """
    if not leave_date:
        return False
    
    # تحويل التاريخ إذا كان نصياً
    if isinstance(leave_date, str):
        try:
            leave_date = datetime.strptime(leave_date, '%Y-%m-%d').date()
        except:
            return False
    
    today = date.today()
    
    # تحديد الفترة الزمنية (الشهر المالي) لتاريخ الإجازة
    leave_period_start, leave_period_end = get_financial_period_for_date(leave_date)
    
    # تحديد الفترة الزمنية (الشهر المالي) لليوم الحالي
    today_period_start, today_period_end = get_financial_period_for_date(today)
    
    # يمكن الرفض فقط إذا كان اليوم في نفس الفترة الزمنية لتاريخ الإجازة
    return (leave_period_start == today_period_start and 
            leave_period_end == today_period_end)



def get_manager_department_stats(manager_id):
    """الحصول على إحصائيات قسم المدير"""
    # الحصول على الأقسام التي يديرها المدير
    managed_departments = db_session.query(Department).filter_by(
        primary_manager_id=manager_id
    ).all()
    
    if not managed_departments:
        return {}
    
    department_ids = [dept.id for dept in managed_departments]
    
    # حساب عدد الموظفين في القسم
    department_employees = db_session.query(User).filter(
        User.department_id.in_(department_ids),
        User.is_admin == False,
        User.is_active == True
    ).all()
    
    # حساب عدد الإذونات اليوم
    today = date.today()
    today_permissions = db_session.query(PermissionRequest).filter(
        PermissionRequest.department_id.in_(department_ids),
        PermissionRequest.date == today,
        PermissionRequest.status == 'approved'
    ).count()
    
    return {
        'department_employees': department_employees,
        'today_permissions': today_permissions,
        'managed_departments': managed_departments
    }

def get_department_employees_without_managers(department_id):
    """الحصول على موظفي القسم (غير المديرين)"""
    employees = db_session.query(User).filter(
        User.department_id == department_id,
        User.is_admin == False,
        User.is_active == True
    ).all()
    
    return employees



@app.route('/manager/create_permission', methods=['GET', 'POST'])
@login_required
def manager_create_permission():
    """صفحة إنشاء إذن للموظفين بواسطة المدير"""
    
    try:
        # الحصول على إحصائيات قسم المدير
        stats = get_manager_department_stats(current_user.id)
    
        print(1)
        # الحصول على موظفي القسم الأول الذي يديره (لتبسيط الأمثلة)
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        
        if managed_departments:
            # الحصول على إحصائيات
            department = managed_departments[0]
        print(6)
        department_employees = get_department_employees_without_managers(department.id)
        print(2)
        # معالجة طلب POST (إنشاء إذن)
        if request.method == 'POST':
            return handle_manager_permission_creation(request, current_user, department)
        
        # تحضير البيانات للعرض
        notifications = get_user_notifications(current_user.id)
        print(3)
        today = date.today()
        today_permissions = db_session.query(PermissionRequest).filter(
                PermissionRequest.department_id == department.id,
                PermissionRequest.date == today,
                PermissionRequest.status == 'approved'
            ).count()
        return render_template('manager_create_permission.html',
                             department_employees=department_employees,
                             today_permissions=today_permissions,
                             current_user=current_user,
                             notifications=notifications)
    
    except Exception as e:
        print(f"Error in manager_create_permission: {str(e)}")
        flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('manager_dashboard'))


def handle_manager_permission_creation(request, manager, department):
    """معالجة إنشاء إذن من قبل المدير"""
    try:
        # التحقق من البيانات المطلوبة
        required_fields = ['employee_id', 'permission_type', 'date', 'time', 'reason']
        for field in required_fields:
            if field not in request.form or not request.form[field].strip():
                flash(f'حقل {field} مطلوب', 'error')
                return redirect(url_for('manager_create_permission'))
        
        employee_id = request.form['employee_id']
        permission_type = request.form['permission_type']
        permission_date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        time_str = request.form['time']
        reason = request.form['reason']
        
        # التحقق من أن الموظف ينتمي لنفس قسم المدير
        employee = db_session.query(User).get(employee_id)
        if not employee or employee.department_id != department.id:
            flash('الموظف غير موجود أو لا ينتمي لقسمك', 'error')
            return redirect(url_for('manager_create_permission'))
        
        # التحقق من أن التاريخ ليس في الماضي
        if permission_date < date.today():
            flash('لا يمكن إنشاء إذن لتاريخ ماضي', 'error')
            return redirect(url_for('manager_create_permission'))
        
        # تحويل الوقت من تنسيق HH:MM إلى دقائق
        try:
            time_obj = datetime.strptime(time_str, '%H:%M').time()
            time_minutes = time_obj.hour * 60 + time_obj.minute
        except ValueError:
            flash('تنسيق الوقت غير صحيح', 'error')
            return redirect(url_for('manager_create_permission'))
        
        # إنشاء طلب الإذن مع الموافقة المباشرة
        permission_request = PermissionRequest(
            user_id=employee_id,
            department_id=department.id,
            permission_type=permission_type,
            date=permission_date,
            time=time_minutes,
            reason=reason,
            extra_data=json.dumps({
                'created_by_manager': True,
                'manager_id': manager.id,
                'manager_name': manager.name,
                'source': 'manager_direct'
            }, ensure_ascii=False),
            status='approved',  # موافقة مباشرة من المدير
            approved_by=manager.id,
            approved_at=datetime.now(),
            created_at=datetime.now()
        )
        
        db_session.add(permission_request)
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'قام مديرك {manager.name} بمنحك إذن من نوع {permission_type}'
        notification_message += f' بتاريخ {permission_date.strftime("%Y-%m-%d")}'
        notification_message += f' بسبب: {reason}'
        
        create_notification(
            employee_id,
            'تم منحك إذن جديد',
            notification_message,
            'permission_granted',
            related_id=permission_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        # إرسال إشعار للمدير نفسه
        create_notification(
            manager.id,
            'تم إنشاء إذن للموظف',
            f'تم بنجاح منح إذن للموظف {employee.name}',
            'manager_permission_created',
            related_id=permission_request.id,
            action_url=url_for('manager_permission_requests')
        )
        
        flash(f'تم منح الإذن للموظف {employee.name} بنجاح', 'success')
        return redirect(url_for('manager_create_permission'))
    
    except Exception as e:
        db_session.rollback()
        print(f"Error handling manager permission creation: {str(e)}")
        flash(f'حدث خطأ أثناء إنشاء الإذن: {str(e)}', 'error')
        return redirect(url_for('manager_create_permission'))



@app.route('/api/manager/employee_suggestions')
@login_required
def api_manager_employee_suggestions():
    """API للحصول على اقتراحات الموظفين للبحث"""
    
    try:
        query = request.args.get('q', '')
        
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        
        if not managed_departments:
            return jsonify({'success': True, 'employees': []})
        
        department_ids = [dept.id for dept in managed_departments]
        
        # البحث عن الموظفين
        employees_query = db_session.query(User).filter(
            User.department_id.in_(department_ids),
            User.is_admin == False,
            User.is_active == True
        )
        
        if query:
            employees_query = employees_query.filter(
                User.name.ilike(f'%{query}%') |
                User.username.ilike(f'%{query}%')
            )
        
        employees = employees_query.order_by(User.name).limit(10).all()
        
        # تحضير البيانات
        employees_data = []
        for emp in employees:
            employee_data = db_session.query(EmployeeData).filter_by(
                user_id=emp.id
            ).first()
            
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'username': emp.username,
                'department_id': emp.department_id,
                'job_title': employee_data.job_title if employee_data else 'موظف',
                'display_text': f"{emp.name} ({emp.username})"
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        })



@app.context_processor
def inject_manager_context():
    """حقن سياق المدير في جميع القوالب"""
    if current_user.is_authenticated and current_user.is_manager:
        # الحصول على قسم المدير
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        
        if managed_departments:
            # الحصول على إحصائيات
            department = managed_departments[0]
            employee_count = db_session.query(User).filter(
                User.department_id == department.id,
                User.is_admin == False
            ).count()
            
            today = date.today()
            today_permissions = db_session.query(PermissionRequest).filter(
                PermissionRequest.department_id == department.id,
                PermissionRequest.date == today,
                PermissionRequest.status == 'approved'
            ).count()
            
            return {
                'managed_department': department,
                'department_employee_count': employee_count,
                'today_permissions': today_permissions
            }
    
    return {}

def get_financial_period_for_date(check_date):
    """
    الحصول على بداية ونهاية الفترة المالية (الشهر المالي) لتاريخ معين
    
    القاعدة:
    - يناير: 26 ديسمبر السابق إلى 25 يناير الحالي
    - فبراير: 26 يناير إلى 25 فبراير
    - مارس: 26 فبراير إلى 25 مارس
    - ... وهكذا
    """
    # حساب الشهر المالي بناءً على تاريخ الإجازة
    if check_date.day >= 26:
        # يقع في الشهر المالي الذي يبدأ من يوم 26 من الشهر الحالي
        period_start = date(check_date.year, check_date.month, 26)
        if check_date.month == 12:
            period_end = date(check_date.year + 1, 1, 25)
        else:
            period_end = date(check_date.year, check_date.month + 1, 25)
    else:
        # يقع في الشهر المالي الذي يبدأ من يوم 26 من الشهر السابق
        if check_date.month == 1:
            period_start = date(check_date.year - 1, 12, 26)
            period_end = date(check_date.year, 1, 25)
        else:
            period_start = date(check_date.year, check_date.month - 1, 26)
            period_end = date(check_date.year, check_date.month, 25)
    
    return period_start, period_end


@app.route('/manager/reject_leave/<int:request_id>', methods=['POST'])
@login_required
def reject_leave(request_id):
    
    leave_request = db_session.query(LeaveRequest).filter_by(id=request_id).first()
    if not leave_request:
        flash('طلب الإجازة غير موجود', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # الحصول على تاريخ الإجازة الصحيح
    # استخدم leave_date إذا كان موجوداً، وإلا استخدم start_date
    leave_date = leave_request.leave_date if hasattr(leave_request, 'leave_date') and leave_request.leave_date else leave_request.start_date
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.get(Department, leave_request.department_id)
    
    # التحقق من أن الطلب معلق
    if leave_request.status != 'pending':
        flash('تم معالجة هذا الطلب مسبقاً', 'warning')
        return redirect(url_for('manager_leave_requests'))
    
    rejection_reason = request.form.get('rejection_reason', '').strip()
    
    if not rejection_reason:
        flash('يرجى كتابة سبب الرفض', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    try:
        # رفض الشيفت الفردي
        leave_request.status = 'rejected'
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        leave_request.rejection_reason = rejection_reason
        
        # زيادة الرصيد إذا كانت الإجازة من رصيد الإجازات
        balance_increased = False
        new_balance = None
        
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=leave_request.user_id
            ).first()
            
            if balance:
                old_balance = balance.leave_balance
                balance.leave_balance += 1  # إرجاع يوم واحد
                balance.last_updated = datetime.now()
                new_balance = balance.leave_balance
                balance_increased = True
        
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'تم رفض شيفت إجازتك ({leave_request.shift_name}). السبب: {rejection_reason}'
        if balance_increased:
            notification_message += f'\nتم زيادة رصيدك بمقدار 1 يوم (الرصيد الجديد: {new_balance} يوم)'
        
        create_notification(
            leave_request.user_id,
            'رفض طلب الإجازة',
            notification_message,
            'leave_rejected',
            related_id=leave_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        success_message = 'تم رفض الشيفت بنجاح'
        if balance_increased:
            success_message += ' وتم زيادة رصيد الموظف'
        
        flash(success_message, 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء رفض الطلب: {str(e)}', 'error')
    
    return redirect(url_for('manager_leave_requests'))


@app.context_processor
def inject_rejection_period():
    """تمرير معلومات الفترة الزمنية الحالية للرفض لجميع القوالب"""
    if current_user.is_authenticated and (current_user.is_manager or current_user.is_admin):
        period_start, period_end = get_financial_period_for_date(date.today())
        return {
            'rejection_period_start': period_start,
            'rejection_period_end': period_end,
            'can_reject_leave_by_date': can_reject_leave_by_date,
            'get_financial_period_for_date': get_financial_period_for_date
        }
    return {}



def get_financial_month_range(check_date=None):
    """الحصول على نطاق الشهر المالي (26 الشهر السابق إلى 25 الشهر الحالي)"""
    if not check_date:
        check_date = date.today()
    
    if check_date.day >= 26:
        # الفترة: 26 من الشهر الحالي إلى 25 من الشهر التالي
        start_date = date(check_date.year, check_date.month, 26)
        if check_date.month == 12:
            end_date = date(check_date.year + 1, 1, 25)
        else:
            end_date = date(check_date.year, check_date.month + 1, 25)
    else:
        # الفترة: 26 من الشهر السابق إلى 25 من الشهر الحالي
        if check_date.month == 1:
            start_date = date(check_date.year - 1, 12, 26)
            end_date = date(check_date.year, 1, 25)
        else:
            start_date = date(check_date.year, check_date.month - 1, 26)
            end_date = date(check_date.year, check_date.month, 25)
    
    return start_date, end_date

@app.context_processor
def utility_processor():
    def get_month_name_arabic_filter(month_num):
        months = {
            1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
            5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
            9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
        }
        return months.get(month_num, f'شهر {month_num}')
    
    return dict(get_month_name_arabic=get_month_name_arabic_filter)


def get_month_name_arabic(month_num):
    """Get Arabic month name - make sure this function exists"""
    months = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    return months.get(month_num, '')

@app.template_filter('get_month_name_arabic')
def get_month_name_arabic_filter(month_num):
    """فلتر للقالب للحصول على اسم الشهر بالعربية"""
    months = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    return months.get(month_num, f'شهر {month_num}')

@app.route('/manager/approve_leave_group/<int:parent_request_id>')
@login_required
def approve_leave_group(parent_request_id):
    """موافقة على مجموعة كاملة من الشيفتات"""
    
    # الحصول على جميع الشيفتات المرتبطة بهذا الطلب
    leave_requests = db_session.query(LeaveRequest).filter_by(
        parent_request_id=parent_request_id,
        status='pending'
    ).all()
    
    if not leave_requests:
        flash('لم يتم العثور على طلبات إجازة معلقة', 'warning')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_requests[0].department_id)
    
    try:
        approved_count = 0
        user_id = leave_requests[0].user_id
        
        for request in leave_requests:
            if request.status == 'pending':
                request.status = 'approved'
                request.approved_by = current_user.id
                request.approved_at = datetime.now()
                approved_count += 1
        
        # خصم من الرصيد فقط إذا كانت من رصيد الإجازات
        if leave_requests[0].leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(user_id=user_id).first()
            if balance:
                balance.leave_balance -= approved_count
                balance.last_updated = datetime.now()
        
        db_session.commit()
        
        # إرسال إشعار للموظف
        create_notification(
            user_id,
            'موافقة على طلب الإجازة',
            f'تم الموافقة على جميع شيفتات طلب إجازتك البالغ عددها {approved_count} شيفت',
            'leave_approved',
            related_id=parent_request_id,
            action_url=url_for('user_leave_requests')
        )
        
        flash(f'تم الموافقة على {approved_count} شيفت بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء الموافقة على الطلب: {str(e)}', 'error')
    
    return redirect(url_for('manager_leave_requests'))


@app.route('/manager/approve_single_shift/<int:request_id>')
@login_required
def approve_single_shift(request_id):
    """موافقة على شيفت فردي"""
    
    leave_request = db_session.query(LeaveRequest).filter_by(id=request_id).first()
    
    if not leave_request:
        flash('طلب الإجازة غير موجود')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_request.department_id)
    
    if leave_request.status == 'pending':
        leave_request.status = 'approved'
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        
        # خصم من الرصيد فقط إذا كانت من رصيد الإجازات
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(user_id=leave_request.user_id).first()
            if balance:
                balance.leave_balance -= 1  # شيفت واحد = يوم واحد
                balance.last_updated = datetime.now()
        
        db_session.commit()
        
        flash('تم الموافقة على الشيفت بنجاح')
    
    return redirect(url_for('manager_leave_requests'))


@app.template_filter('parse_shift_info')
def parse_shift_info_filter(shift_info_json):
    """فلتر للقالب لتحليل معلومات الشيفتات"""
    if not shift_info_json:
        return []
    try:
        shift_data = json.loads(shift_info_json)
        return shift_data.get('shifts', [])
    except:
        return []

@app.template_filter('get_shift_jobs')
def get_shift_jobs_filter(shift_info_json):
    """فلتر للقالب للحصول على الوظائف المحددة لكل شيفت"""
    if not shift_info_json:
        return {}
    try:
        shift_data = json.loads(shift_info_json)
        return shift_data.get('jobs', {})
    except:
        return {}



@app.route('/api/manager/permission_requests/statistics')
@login_required
def api_permission_requests_statistics():
    """API للحصول على إحصائيات الإذونات للقسم"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        department_ids = [dept.id for dept in managed_departments]
        
        if not department_ids:
            return jsonify({'success': True, 'statistics': {}})
        
        # الحصول على جميع طلبات الإذن
        all_permissions = db_session.query(PermissionRequest)\
            .filter(PermissionRequest.department_id.in_(department_ids))\
            .all()
        
        # حساب الإحصائيات
        stats = calculate_detailed_permission_statistics(all_permissions)
        
        return jsonify({
            'success': True,
            'statistics': stats
        })
        
    except Exception as e:
        print(f"Error in permission statistics API: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

def calculate_permission_statistics(permissions):
    """حساب إحصائيات الإذونات"""
    total = len(permissions)
    pending = len([p for p in permissions if p.status == 'pending'])
    approved = len([p for p in permissions if p.status == 'approved'])
    rejected = len([p for p in permissions if p.status == 'rejected'])
    
    # حساب الإحصائيات حسب النوع
    type_stats = {}
    for permission in permissions:
        perm_type = permission.permission_type or 'غير محدد'
        if perm_type not in type_stats:
            type_stats[perm_type] = {
                'total': 0,
                'pending': 0,
                'approved': 0,
                'rejected': 0
            }
        
        type_stats[perm_type]['total'] += 1
        if permission.status == 'pending':
            type_stats[perm_type]['pending'] += 1
        elif permission.status == 'approved':
            type_stats[perm_type]['approved'] += 1
        elif permission.status == 'rejected':
            type_stats[perm_type]['rejected'] += 1
    
    # حساب الإحصائيات حسب الشهر
    monthly_stats = {}
    for permission in permissions:
        if permission.created_at:
            month_key = permission.created_at.strftime('%Y-%m')
        else:
            month_key = 'غير محدد'
        
        if month_key not in monthly_stats:
            monthly_stats[month_key] = {
                'total': 0,
                'pending': 0,
                'approved': 0,
                'rejected': 0
            }
        
        monthly_stats[month_key]['total'] += 1
        if permission.status == 'pending':
            monthly_stats[month_key]['pending'] += 1
        elif permission.status == 'approved':
            monthly_stats[month_key]['approved'] += 1
        elif permission.status == 'rejected':
            monthly_stats[month_key]['rejected'] += 1
    
    # حساب الإحصائيات حسب الموظف
    employee_stats = {}
    for permission in permissions:
        if permission.user:
            user_key = f"{permission.user.name} ({permission.user.username})"
            if user_key not in employee_stats:
                employee_stats[user_key] = {
                    'department': permission.department.name if permission.department else 'غير معين',
                    'total': 0,
                    'pending': 0,
                    'approved': 0,
                    'rejected': 0
                }
            
            employee_stats[user_key]['total'] += 1
            if permission.status == 'pending':
                employee_stats[user_key]['pending'] += 1
            elif permission.status == 'approved':
                employee_stats[user_key]['approved'] += 1
            elif permission.status == 'rejected':
                employee_stats[user_key]['rejected'] += 1
    
    # حساب معدل القبول
    approval_rate = 0
    if total > 0:
        approval_rate = round((approved / total) * 100, 1)
    
    # طلبات الشهر الحالي
    current_month = datetime.now().strftime('%Y-%m')
    current_month_requests = [
        p for p in permissions 
        if p.created_at and p.created_at.strftime('%Y-%m') == current_month
    ]
    
    return {
        'total_permissions': total,
        'pending_permissions': pending,
        'approved_permissions': approved,
        'rejected_permissions': rejected,
        'approval_rate': approval_rate,
        'current_month_requests': len(current_month_requests),
        'permission_types': type_stats,
        'monthly_stats': monthly_stats,
        'employee_stats': employee_stats
    }

# ======== Helper Functions for Rewards/Penalties ========

def get_managed_department_ids(user_id):
    """الحصول على IDs جميع الأقسام التي يديرها المستخدم"""
    if not user_id:
        return []
    
    user = db_session.query(User).get(user_id)
    if user and user.is_admin:
        departments = db_session.query(Department).all()
        return [dept.id for dept in departments]
    
    # If user is manager, get their departments from DepartmentManager
    managed_departments = db_session.query(DepartmentManager).filter_by(
        user_id=user_id
    ).all()
    
    return [dm.department_id for dm in managed_departments]

def get_department_employees(department_id):
    """الحصول على موظفي قسم معين"""
    return db_session.query(User).filter(
        User.department_id == department_id,
        User.is_admin == False,
        User.is_active == True
    ).order_by(User.name).all()

def can_manage_rewards(user_id, department_id=None):
    """التحقق مما إذا كان المستخدم يمكنه إدارة المكافآت"""
    user = db_session.query(User).get(user_id)
    
    if user.is_admin:
        return True
    
    # التحقق من أن المستخدم مدير للقسم المحدد
    if department_id:
        manager = db_session.query(DepartmentManager).filter_by(
            user_id=user_id,
            department_id=department_id
        ).first()
        
        if manager and manager.can_manage_rewards:
            return True
    
    return False

# Change the name of the helper function to avoid conflict
def get_department_employees_list(department_id):
    """الحصول على موظفي قسم معين (قائمة) - لتمييزها عن API"""
    return db_session.query(User).filter(
        User.department_id == department_id,
        User.is_admin == False,
        User.is_active == True
    ).order_by(User.name).all()


@app.route('/manager/rewards_penalties', methods=['GET', 'POST'])
@login_required
def manager_rewards_penalties():
    """إدارة المكافآت والخصومات"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_dept_ids = get_managed_department_ids(current_user.id)
        
        # معالجة POST لإنشاء مكافأة/خصم
        if request.method == 'POST':
            return handle_create_reward_penalty(request, current_user, managed_dept_ids)
        
        # GET - عرض الصفحة
        departments = db_session.query(Department).filter(
            Department.id.in_(managed_dept_ids)
        ).all()
        
        # الحصول على المكافآت والخصومات السابقة
        rewards_penalties = db_session.query(RewardPenalty).filter(
            RewardPenalty.department_id.in_(managed_dept_ids)
        ).order_by(RewardPenalty.effective_date.desc()).all()
        
        # تحميل البيانات المرتبطة
        for rp in rewards_penalties:
            rp.user = db_session.query(User).get(rp.user_id)
            rp.department = db_session.query(Department).get(rp.department_id)
            rp.creator = db_session.query(User).get(rp.created_by)
        
        # إحصائيات
        stats = {
            'total': len(rewards_penalties),
            'rewards': len([rp for rp in rewards_penalties if rp.type == 'reward']),
            'penalties': len([rp for rp in rewards_penalties if rp.type == 'penalty']),
            'total_amount': sum([rp.amount for rp in rewards_penalties if rp.type == 'reward']) - 
                          sum([rp.amount for rp in rewards_penalties if rp.type == 'penalty'])
        }
        
        # التحضير لقائمة موظفي الأقسام (لتجنب مشاكل AJAX)
        department_employees_dict = {}
        for dept in departments:
            # استخدام الدالة الجديدة
            employees = get_department_employees_list(dept.id)
            department_employees_dict[str(dept.id)] = []
            
            for emp in employees:
                employee_data = db_session.query(EmployeeData).filter_by(user_id=emp.id).first()
                employee_info = {
                    'id': emp.id,
                    'name': emp.name,
                    'arabic_name': employee_data.arabic_name if employee_data else emp.name,
                    'job_title': employee_data.job_title if employee_data else 'موظف'
                }
                department_employees_dict[str(dept.id)].append(employee_info)
        
        # تحويل إلى JSON لتجنب مشاكل الترميز
        import json
        department_employees_json = json.dumps(department_employees_dict, ensure_ascii=False)
        
        notifications = get_user_notifications(current_user.id)
        
        return render_template('manager/manager_rewards_penalties.html',
                             departments=departments,
                             department_employees=department_employees_dict,  # للقالب
                             department_employees_json=department_employees_json,  # لجافاسكريبت
                             rewards_penalties=rewards_penalties,
                             stats=stats,
                             notifications=notifications)
    
    except Exception as e:
        print(f"Error in manager_rewards_penalties: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('manager_dashboard'))

def handle_create_reward_penalty(request, current_user, managed_dept_ids):
    """معالجة إنشاء مكافأة/خصم"""
    try:
        # التحقق من البيانات المطلوبة
        required_fields = ['employee_id', 'type', 'amount', 'effective_date', 'reason']
        for field in required_fields:
            if field not in request.form or not request.form[field].strip():
                flash(f'حقل {field} مطلوب', 'error')
                return redirect(url_for('manager_rewards_penalties'))
        
        employee_id = int(request.form['employee_id'])
        type_ = request.form['type']
        amount = float(request.form['amount'])
        effective_date = datetime.strptime(request.form['effective_date'], '%Y-%m-%d').date()
        reason = request.form['reason'].strip()
        
        # التحقق من أن الموظف ينتمي لقسم يديره المدير
        employee = db_session.query(User).get(employee_id)
        if not employee:
            flash('الموظف غير موجود', 'error')
            return redirect(url_for('manager_rewards_penalties'))
        
        if employee.department_id not in managed_dept_ids:
            flash('الموظف لا ينتمي إلى قسم تديره', 'error')
            return redirect(url_for('manager_rewards_penalties'))
        
        # التحقق من المبلغ
        if amount <= 0:
            flash('المبلغ يجب أن يكون أكبر من صفر', 'error')
            return redirect(url_for('manager_rewards_penalties'))
        
        # التحقق من التاريخ
        if effective_date > date.today():
            flash('لا يمكن إضافة مكافأة/خصم لتاريخ مستقبلي', 'error')
            return redirect(url_for('manager_rewards_penalties'))
        
        # إنشاء سجل المكافأة/الخصم
        reward_penalty = RewardPenalty(
            user_id=employee_id,
            department_id=employee.department_id,
            type=type_,
            amount=amount,
            reason=reason,
            effective_date=effective_date,
            created_by=current_user.id,
            created_at=datetime.now()
        )
        
        db_session.add(reward_penalty)
        
        # إرسال إشعار للموظف
        type_arabic = 'مكافأة' if type_ == 'reward' else 'خصم'
        notification_message = f'تم تطبيق {type_arabic} عليك بقيمة {amount:.2f} جنيه'
        notification_message += f'\nالتاريخ: {effective_date.strftime("%Y-%m-%d")}'
        notification_message += f'\nالسبب: {reason}'
        
        create_notification(
            employee_id,
            f'{type_arabic} جديد',
            notification_message,
            'reward_penalty',
            related_id=reward_penalty.id,
            action_url=url_for('user_rewards_penalties')
        )
        
        # إرسال إشعار للمدير
        create_notification(
            current_user.id,
            f'تم إضافة {type_arabic}',
            f'تم إضافة {type_arabic} للموظف {employee.name} بنجاح',
            'reward_penalty_created',
            related_id=reward_penalty.id,
            action_url=url_for('manager_rewards_penalties')
        )
        
        db_session.commit()
        
        flash(f'تم إضافة {type_arabic} بنجاح', 'success')
        return redirect(url_for('manager_rewards_penalties'))
    
    except ValueError as e:
        flash('البيانات المدخلة غير صحيحة', 'error')
        return redirect(url_for('manager_rewards_penalties'))
    except Exception as e:
        db_session.rollback()
        print(f"Error creating reward/penalty: {str(e)}")
        flash(f'حدث خطأ أثناء الإضافة: {str(e)}', 'error')
        return redirect(url_for('manager_rewards_penalties'))
    




@app.route('/api/department_employees/<int:department_id>')
@login_required
def api_department_employees(department_id):
    """API للحصول على موظفي قسم معين"""
    
    try:
        # التحقق من أن المدير يدير هذا القسم
        managed_dept_ids = get_managed_department_ids(current_user.id)
        
        employees = get_department_employees(department_id)
        
        employees_data = []
        for emp in employees:
            employee_data = db_session.query(EmployeeData).filter_by(user_id=emp.id).first()
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'username': emp.username,
                'arabic_name': employee_data.arabic_name if employee_data else emp.name,
                'job_title': employee_data.job_title if employee_data else 'موظف'
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

def calculate_detailed_permission_statistics(permissions):
    """حساب إحصائيات مفصلة للإذونات"""
    try:
        total_permissions = len(permissions)
        pending_permissions = len([p for p in permissions if p.status == 'pending'])
        approved_permissions = len([p for p in permissions if p.status == 'approved'])
        rejected_permissions = len([p for p in permissions if p.status == 'rejected'])
        
        # حساب الإحصائيات حسب النوع
        permission_types = {}
        for permission in permissions:
            perm_type = permission.permission_type or 'غير محدد'
            if perm_type not in permission_types:
                permission_types[perm_type] = {
                    'total': 0,
                    'pending': 0,
                    'approved': 0,
                    'rejected': 0
                }
            
            permission_types[perm_type]['total'] += 1
            if permission.status == 'pending':
                permission_types[perm_type]['pending'] += 1
            elif permission.status == 'approved':
                permission_types[perm_type]['approved'] += 1
            elif permission.status == 'rejected':
                permission_types[perm_type]['rejected'] += 1
        
        # حساب الإحصائيات حسب الشهر
        monthly_stats = {}
        for permission in permissions:
            if permission.created_at:
                month_key = permission.created_at.strftime('%Y-%m')
            else:
                month_key = 'غير محدد'
            
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {
                    'total': 0,
                    'pending': 0,
                    'approved': 0,
                    'rejected': 0
                }
            
            monthly_stats[month_key]['total'] += 1
            if permission.status == 'pending':
                monthly_stats[month_key]['pending'] += 1
            elif permission.status == 'approved':
                monthly_stats[month_key]['approved'] += 1
            elif permission.status == 'rejected':
                monthly_stats[month_key]['rejected'] += 1
        
        # حساب الإحصائيات حسب الموظف
        employee_stats = {}
        for permission in permissions:
            user = db_session.query(User).get(permission.user_id)
            department = db_session.query(Department).get(permission.department_id)
            
            if user:
                user_key = f"{user.name} ({user.username})"
                if user_key not in employee_stats:
                    employee_stats[user_key] = {
                        'department': department.name if department else 'غير معين',
                        'total': 0,
                        'pending': 0,
                        'approved': 0,
                        'rejected': 0,
                        'total_hours': 0
                    }
                
                employee_stats[user_key]['total'] += 1
                if permission.status == 'pending':
                    employee_stats[user_key]['pending'] += 1
                elif permission.status == 'approved':
                    employee_stats[user_key]['approved'] += 1
                elif permission.status == 'rejected':
                    employee_stats[user_key]['rejected'] += 1
                
                # حساب ساعات الإضافي إذا كانت موجودة
                if permission.extra_data:
                    try:
                        extra_data = json.loads(permission.extra_data)
                        if 'hours' in extra_data:
                            hours = int(extra_data['hours'])
                            employee_stats[user_key]['total_hours'] += hours
                    except:
                        pass
        
        # حساب معدل القبول
        approval_rate = 0
        if total_permissions > 0:
            approval_rate = round((approved_permissions / total_permissions) * 100, 1)
        
        # طلبات الشهر الحالي
        current_month = datetime.now().strftime('%Y-%m')
        current_month_requests = [
            p for p in permissions 
            if p.created_at and p.created_at.strftime('%Y-%m') == current_month
        ]
        
        return {
            'total_permissions': total_permissions,
            'pending_permissions': pending_permissions,
            'approved_permissions': approved_permissions,
            'rejected_permissions': rejected_permissions,
            'approval_rate': approval_rate,
            'current_month_requests': len(current_month_requests),
            'permission_types': permission_types,
            'monthly_stats': monthly_stats,
            'employee_stats': employee_stats,
            'avg_processing_time': 1.2,  # يمكن حسابها لاحقاً
            'peak_permission_type': max(permission_types, key=lambda k: permission_types[k]['total']) if permission_types else 'لا توجد بيانات'
        }
        
    except Exception as e:
        print(f"Error calculating detailed statistics: {str(e)}")
        return {}

def calculate_detailed_statistics(department_ids):
    """Calculate detailed statistics for all request types"""
    stats = {
        'leaves': {
            'total': db_session.query(LeaveRequest).filter(
                LeaveRequest.department_id.in_(department_ids)
            ).count(),
            'pending': db_session.query(LeaveRequest).filter(
                LeaveRequest.department_id.in_(department_ids),
                LeaveRequest.status == 'pending'
            ).count(),
            'approved': db_session.query(LeaveRequest).filter(
                LeaveRequest.department_id.in_(department_ids),
                LeaveRequest.status == 'approved'
            ).count(),
            'rejected': db_session.query(LeaveRequest).filter(
                LeaveRequest.department_id.in_(department_ids),
                LeaveRequest.status == 'rejected'
            ).count()
        },
        'permissions': {
            'total': db_session.query(PermissionRequest).filter(
                PermissionRequest.department_id.in_(department_ids)
            ).count(),
            'pending': db_session.query(PermissionRequest).filter(
                PermissionRequest.department_id.in_(department_ids),
                PermissionRequest.status == 'pending'
            ).count(),
            'approved': db_session.query(PermissionRequest).filter(
                PermissionRequest.department_id.in_(department_ids),
                PermissionRequest.status == 'approved'
            ).count(),
            'rejected': db_session.query(PermissionRequest).filter(
                PermissionRequest.department_id.in_(department_ids),
                PermissionRequest.status == 'rejected'
            ).count()
        },
        'advances': {
            'total': db_session.query(AdvanceRequest).filter(
                AdvanceRequest.department_id.in_(department_ids)
            ).count(),
            'pending': db_session.query(AdvanceRequest).filter(
                AdvanceRequest.department_id.in_(department_ids),
                AdvanceRequest.status == 'pending'
            ).count()
        },
        'employees': {
            'total': db_session.query(User).filter(
                User.department_id.in_(department_ids),
                User.is_admin == False
            ).count(),
            'active': db_session.query(User).filter(
                User.department_id.in_(department_ids),
                User.is_admin == False,
                User.is_active == True
            ).count()
        }
    }
    
    # Calculate rates
    for category in ['leaves', 'permissions']:
        if stats[category]['total'] > 0:
            stats[category]['approval_rate'] = round(
                (stats[category]['approved'] / stats[category]['total']) * 100, 1
            )
            stats[category]['pending_rate'] = round(
                (stats[category]['pending'] / stats[category]['total']) * 100, 1
            )
        else:
            stats[category]['approval_rate'] = 0
            stats[category]['pending_rate'] = 0
    
    return stats


@app.route('/api/manager/permission_chart_data')
@login_required
def api_permission_chart_data():
    """API لتوليد بيانات الرسوم البيانية للإذونات"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        department_ids = [dept.id for dept in managed_departments]
        
        if not department_ids:
            return jsonify({'success': True, 'data': {}})
        
        # الحصول على الفترة المحددة
        period = request.args.get('period', 'month')
        
        # حساب تاريخ البدء بناءً على الفترة
        today = datetime.now()
        if period == 'week':
            start_date = today - timedelta(days=7)
        elif period == 'month':
            start_date = today - timedelta(days=30)
        elif period == 'year':
            start_date = today - timedelta(days=365)
        else:
            start_date = today - timedelta(days=30)  # افتراضي شهر
        
        # الحصول على البيانات
        query = db_session.query(PermissionRequest).filter(
            PermissionRequest.department_id.in_(department_ids),
            PermissionRequest.created_at >= start_date
        )
        
        all_requests = query.all()
        
        # ====== توزيع أنواع الإذونات ======
        type_distribution = {}
        for request in all_requests:
            # تنظيف نوع الإذن
            permission_type = "غير محدد"
            if request.permission_type:
                if "اذن حضور" in request.permission_type:
                    permission_type = "اذن حضور"
                elif "اذن انصراف" in request.permission_type:
                    permission_type = "اذن انصراف"
                elif "اذن نسيان" in request.permission_type:
                    permission_type = "اذن نسيان توقيع"
                elif "اذن تبديل" in request.permission_type:
                    permission_type = "اذن تبديل وردية"
                elif "اذن طلب ساعات" in request.permission_type:
                    permission_type = "اذن طلب ساعات اضافي"
                else:
                    permission_type = request.permission_type
            
            type_distribution[permission_type] = type_distribution.get(permission_type, 0) + 1
        
        # تحويل إلى تنسيق مناسب للرسم البياني
        chart_data_types = {
            'labels': list(type_distribution.keys()),
            'values': list(type_distribution.values()),
            'colors': [
                '#3b82f6',  # أزرق
                '#10b981',  # أخضر
                '#f59e0b',  # أصفر
                '#8b5cf6',  # بنفسجي
                '#ec4899',  # وردي
                '#ef4444'   # أحمر
            ]
        }
        
        # ====== الاتجاه الشهري ======
        monthly_trend = {}
        for request in all_requests:
            if request.created_at:
                month_key = request.created_at.strftime('%Y-%m')
                if month_key not in monthly_trend:
                    monthly_trend[month_key] = {'approved': 0, 'rejected': 0}
                
                if request.status == 'approved':
                    monthly_trend[month_key]['approved'] += 1
                elif request.status == 'rejected':
                    monthly_trend[month_key]['rejected'] += 1
        
        # فرز الأشهر
        sorted_months = sorted(monthly_trend.keys())
        chart_data_trend = {
            'months': sorted_months,
            'approved': [monthly_trend[month].get('approved', 0) for month in sorted_months],
            'rejected': [monthly_trend[month].get('rejected', 0) for month in sorted_months]
        }
        
        return jsonify({
            'success': True,
            'period': period,
            'type_distribution': chart_data_types,
            'monthly_trend': chart_data_trend,
            'total_requests': len(all_requests)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })



@app.route('/api/manager/permission_employee_report')
@login_required
def api_permission_employee_report():
    """API لتقرير الموظفين والإذونات"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        department_ids = [dept.id for dept in managed_departments]
        
        if not department_ids:
            return jsonify({'success': True, 'employees': []})
        
        # الحصول على تاريخ البدء والنهاية
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date') or datetime.now().strftime('%Y-%m-%d')
        
        # بناء الاستعلام
        query = db_session.query(PermissionRequest).filter(
            PermissionRequest.department_id.in_(department_ids)
        )
        
        if start_date:
            query = query.filter(PermissionRequest.date >= start_date)
        if end_date:
            query = query.filter(PermissionRequest.date <= end_date)
        
        all_requests = query.all()
        
        # تجميع بيانات الموظفين
        employee_data = {}
        for request in all_requests:
            user = db_session.query(User).get(request.user_id)
            department = db_session.query(Department).get(request.department_id)
            
            if user:
                user_id = user.id
                if user_id not in employee_data:
                    employee_data[user_id] = {
                        'employee_id': user_id,
                        'employee_name': user.name,
                        'department': department.name if department else 'غير معين',
                        'total': 0,
                        'approved': 0,
                        'rejected': 0,
                        'pending': 0,
                        'total_hours': 0,
                        'approval_rate': 0
                    }
                
                employee_data[user_id]['total'] += 1
                
                if request.status == 'approved':
                    employee_data[user_id]['approved'] += 1
                    # حساب الساعات الإضافية
                    if request.extra_data:
                        try:
                            extra_data = json.loads(request.extra_data)
                            if 'hours' in extra_data:
                                employee_data[user_id]['total_hours'] += int(extra_data['hours'])
                        except:
                            pass
                elif request.status == 'rejected':
                    employee_data[user_id]['rejected'] += 1
                elif request.status == 'pending':
                    employee_data[user_id]['pending'] += 1
        
        # حساب معدل القبول لكل موظف
        for emp_id, data in employee_data.items():
            total_processed = data['approved'] + data['rejected']
            if total_processed > 0:
                data['approval_rate'] = round((data['approved'] / total_processed) * 100, 1)
            else:
                data['approval_rate'] = 0
            
            # حساب متوسط الساعات الإضافية
            if data['total'] > 0:
                data['avg_hours'] = round(data['total_hours'] / data['total'], 1)
            else:
                data['avg_hours'] = 0
        
        # تحويل إلى قائمة وترتيب حسب عدد الطلبات
        employees_list = list(employee_data.values())
        employees_list.sort(key=lambda x: x['total'], reverse=True)
        
        # ====== إحصائيات إضافية ======
        # النوع الأكثر طلباً
        permission_types = {}
        for request in all_requests:
            perm_type = request.permission_type or 'غير محدد'
            permission_types[perm_type] = permission_types.get(perm_type, 0) + 1
        
        peak_permission_type = max(permission_types, key=permission_types.get) if permission_types else 'لا توجد بيانات'
        
        # متوسط وقت المعالجة (بسيط - يمكن تطويره)
        avg_processing_time = 1.2
        
        return jsonify({
            'success': True,
            'employees': employees_list,
            'summary': {
                'total_employees': len(employees_list),
                'total_requests': len(all_requests),
                'peak_permission_type': peak_permission_type,
                'avg_processing_time': avg_processing_time
            }
        })
        
    except Exception as e:
        print(f"Error in employee report API: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })


# ======== Manager Permission API Routes ========

@app.route('/api/manager/pending_permissions')
@login_required
def api_manager_pending_permissions():
    """API للحصول على طلبات الإذونات المعلقة"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        department_ids = [dept.id for dept in managed_departments]
        
        if not department_ids:
            return jsonify({'success': True, 'requests': []})
        
        # الحصول على الطلبات المعلقة
        pending_requests = db_session.query(PermissionRequest)\
            .options(
                joinedload(PermissionRequest.user),
                joinedload(PermissionRequest.department)
            )\
            .filter(
                PermissionRequest.department_id.in_(department_ids),
                PermissionRequest.status == 'pending'
            )\
            .order_by(PermissionRequest.created_at.desc())\
            .all()
        
        # تحضير البيانات للاستجابة
        requests_data = []
        for request in pending_requests:
            request_data = {
                'id': request.id,
                'user': {
                    'id': request.user.id if request.user else None,
                    'name': request.user.name if request.user else 'غير معروف',
                    'username': request.user.username if request.user else 'غير معروف'
                } if request.user else None,
                'department': {
                    'id': request.department.id if request.department else None,
                    'name': request.department.name if request.department else 'غير معين'
                } if request.department else None,
                'permission_type': request.permission_type,
                'date': request.date.strftime('%Y-%m-%d') if request.date else None,
                'time': request.time,
                'reason': request.reason,
                'extra_data': request.extra_data,
                'status': request.status,
                'created_at': request.created_at.strftime('%Y-%m-%d %H:%M') if request.created_at else None,
                'balance': 5  # يمكنك استبدال هذا بقيمة حقيقية من قاعدة البيانات
            }
            requests_data.append(request_data)
        
        return jsonify({
            'success': True,
            'requests': requests_data,
            'count': len(requests_data)
        })
        
    except Exception as e:
        print(f"Error in pending permissions API: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })


@app.route('/api/manager/permission_stats_summary')
@login_required
def api_manager_permission_stats_summary():
    """API للحصول على ملخص إحصائيات الإذونات"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        department_ids = [dept.id for dept in managed_departments]
        
        if not department_ids:
            return jsonify({'success': True, 'statistics': {}})
        
        # الحصول على جميع طلبات الإذن
        all_permissions = db_session.query(PermissionRequest)\
            .filter(PermissionRequest.department_id.in_(department_ids))\
            .all()
        
        # حساب الإحصائيات الأساسية
        total_permissions = len(all_permissions)
        pending_permissions = len([p for p in all_permissions if p.status == 'pending'])
        approved_permissions = len([p for p in all_permissions if p.status == 'approved'])
        rejected_permissions = len([p for p in all_permissions if p.status == 'rejected'])
        
        # حساب معدل القبول
        approval_rate = 0
        if total_permissions > 0:
            approval_rate = round((approved_permissions / total_permissions) * 100, 1)
        
        # طلبات الشهر الحالي
        current_month = datetime.now().strftime('%Y-%m')
        current_month_requests = [
            p for p in all_permissions 
            if p.created_at and p.created_at.strftime('%Y-%m') == current_month
        ]
        
        return jsonify({
            'success': True,
            'statistics': {
                'total_permissions': total_permissions,
                'pending_permissions': pending_permissions,
                'approved_permissions': approved_permissions,
                'rejected_permissions': rejected_permissions,
                'approval_rate': approval_rate,
                'current_month_requests': len(current_month_requests)
            }
        })
        
    except Exception as e:
        print(f"Error in permission stats summary API: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })


def get_manager_navigation_stats(manager_id):
    """الحصول على إحصائيات التنقل للمدير"""
    try:
        # الحصول على الأقسام التي يديرها المدير
        managed_departments = db_session.query(Department).filter_by(
            primary_manager_id=manager_id
        ).all()
        
        if not managed_departments:
            return {}
        
        department_ids = [dept.id for dept in managed_departments]
        
        stats = {
            'pending_leaves': db_session.query(LeaveRequest).filter(
                LeaveRequest.department_id.in_(department_ids),
                LeaveRequest.status == 'pending'
            ).count(),
            'pending_permissions': db_session.query(PermissionRequest).filter(
                PermissionRequest.department_id.in_(department_ids),
                PermissionRequest.status == 'pending'
            ).count(),
            'pending_advances': db_session.query(AdvanceRequest).filter(
                AdvanceRequest.department_id.in_(department_ids),
                AdvanceRequest.status == 'pending'
            ).count(),
            'department_employees': db_session.query(User).filter(
                User.department_id.in_(department_ids),
                User.is_admin == False
            ).count()
        }
        return stats
    except Exception as e:
        print(f"Error getting navigation stats: {e}")
        return {}


@app.context_processor
def inject_manager_stats():
    """Inject manager stats into all templates"""
    if current_user.is_authenticated and hasattr(current_user, 'is_manager') and current_user.is_manager:
        stats = get_manager_navigation_stats(current_user.id)
        return {'stats': stats}
    return {'stats': {}}

def get_permission_statistics(department_ids, all_requests=None):
    """الحصول على إحصائيات متقدمة للإذونات بشكل صحيح"""
    try:
        if not department_ids:
            return get_empty_permission_stats()
        
        # If all_requests is not provided, fetch with proper joins
        if all_requests is None:
            all_requests = db_session.query(PermissionRequest)\
                .options(
                    joinedload(PermissionRequest.user),  # This should be a relationship
                    joinedload(PermissionRequest.department),  # This should be a relationship
                    joinedload(PermissionRequest.approver)  # Use the relationship name, not the column
                )\
                .filter(PermissionRequest.department_id.in_(department_ids))\
                .all()
        
        # ====== إحصائيات الأساسية ======
        total_permissions = len(all_requests)
        pending_permissions = len([r for r in all_requests if r.status == 'pending'])
        approved_permissions = len([r for r in all_requests if r.status == 'approved'])
        rejected_permissions = len([r for r in all_requests if r.status == 'rejected'])
        
        # ====== إحصائيات حسب النوع ======
        permission_types_summary = {}
        for request in all_requests:
            # تنظيف نوع الإذن
            permission_type = "غير محدد"
            if request.permission_type:
                if "اذن حضور" in request.permission_type:
                    permission_type = "اذن حضور"
                elif "اذن انصراف" in request.permission_type:
                    permission_type = "اذن انصراف"
                elif "اذن نسيان" in request.permission_type:
                    permission_type = "اذن نسيان توقيع"
                elif "اذن تبديل" in request.permission_type:
                    permission_type = "اذن تبديل وردية"
                elif "اذن طلب ساعات" in request.permission_type:
                    permission_type = "اذن طلب ساعات اضافي"
                else:
                    permission_type = request.permission_type
            
            if permission_type not in permission_types_summary:
                permission_types_summary[permission_type] = {
                    'total': 0, 
                    'approved': 0, 
                    'pending': 0, 
                    'rejected': 0
                }
            
            permission_types_summary[permission_type]['total'] += 1
            if request.status == 'approved':
                permission_types_summary[permission_type]['approved'] += 1
            elif request.status == 'pending':
                permission_types_summary[permission_type]['pending'] += 1
            elif request.status == 'rejected':
                permission_types_summary[permission_type]['rejected'] += 1
        
        # ====== إحصائيات حسب الشهر ======
        monthly_stats = {}
        for request in all_requests:
            if request.date:  # استخدام تاريخ الإذن وليس تاريخ الإنشاء
                month_key = request.date.strftime('%Y-%m')
            elif request.created_at:
                month_key = request.created_at.strftime('%Y-%m')
            else:
                month_key = 'غير محدد'
            
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {
                    'total': 0, 
                    'approved': 0, 
                    'pending': 0, 
                    'rejected': 0
                }
            
            monthly_stats[month_key]['total'] += 1
            if request.status == 'approved':
                monthly_stats[month_key]['approved'] += 1
            elif request.status == 'pending':
                monthly_stats[month_key]['pending'] += 1
            elif request.status == 'rejected':
                monthly_stats[month_key]['rejected'] += 1
        
        # ====== حساب معدل القبول ======
        approval_rate = 0
        if total_permissions > 0:
            approval_rate = round((approved_permissions / total_permissions) * 100, 1)
        
        # ====== طلبات الشهر الحالي ======
        current_month = datetime.now().strftime('%Y-%m')
        current_month_requests = [
            r for r in all_requests 
            if (r.date and r.date.strftime('%Y-%m') == current_month) or 
               (r.created_at and r.created_at.strftime('%Y-%m') == current_month)
        ]
        
        # ====== إحصائيات حسب الموظف ======
        employee_stats = {}
        for request in all_requests:
            if request.user:
                employee_name = request.user.name
                if employee_name not in employee_stats:
                    employee_stats[employee_name] = {
                        'total': 0,
                        'approved': 0,
                        'rejected': 0,
                        'pending': 0,
                        'department': request.department.name if request.department else 'غير معين'
                    }
                
                employee_stats[employee_name]['total'] += 1
                if request.status == 'approved':
                    employee_stats[employee_name]['approved'] += 1
                elif request.status == 'rejected':
                    employee_stats[employee_name]['rejected'] += 1
                elif request.status == 'pending':
                    employee_stats[employee_name]['pending'] += 1
        
        # ====== تجهيز النتائج ======
        result = {
            # إحصائيات العرض في التبويب
            'total_permissions': total_permissions,
            'pending_permissions': pending_permissions,
            'approved_permissions': approved_permissions,
            'rejected_permissions': rejected_permissions,
            'approval_rate': approval_rate,
            'current_month_requests': len(current_month_requests),
            
            # إحصائيات للتقارير
            'permission_types': permission_types_summary,
            'monthly_stats': monthly_stats,
            'employee_stats': employee_stats,
            
            # إحصائيات التنقل (للتوافق)
            'pending_leaves': db_session.query(LeaveRequest).filter(
                LeaveRequest.department_id.in_(department_ids),
                LeaveRequest.status == 'pending'
            ).count(),
            'pending_advances': db_session.query(AdvanceRequest).filter(
                AdvanceRequest.department_id.in_(department_ids),
                AdvanceRequest.status == 'pending'
            ).count(),
        }
        
        return result
        
    except Exception as e:
        print(f"Error calculating permission statistics: {str(e)}")
        import traceback
        traceback.print_exc()
        return get_empty_permission_stats()

def get_empty_permission_stats():
    """إرجاع إحصائيات فارغة"""
    return {
        'total_permissions': 0,
        'pending_permissions': 0,
        'approved_permissions': 0,
        'rejected_permissions': 0,
        'approval_rate': 0,
        'current_month_requests': 0,
        'permission_types': {},
        'monthly_stats': {},
        'employee_stats': {},
        'pending_leaves': 0,
        'pending_advances': 0
    }

@app.context_processor
def inject_manager_stats():
    """Inject manager stats into all templates"""
    try:
        if current_user.is_authenticated and hasattr(current_user, 'is_manager') and current_user.is_manager:
            # استخدام primary_manager_id بدلاً من manager_id
            managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
            department_ids = [dept.id for dept in managed_departments]
            
            stats = {
                'pending_leaves': db_session.query(LeaveRequest).filter(
                    LeaveRequest.department_id.in_(department_ids),
                    LeaveRequest.status == 'pending'
                ).count(),
                'pending_permissions': db_session.query(PermissionRequest).filter(
                    PermissionRequest.department_id.in_(department_ids),
                    PermissionRequest.status == 'pending'
                ).count(),
                'pending_advances': db_session.query(AdvanceRequest).filter(
                    AdvanceRequest.department_id.in_(department_ids),
                    AdvanceRequest.status == 'pending'
                ).count(),
                'department_employees': db_session.query(User).filter(
                    User.department_id.in_(department_ids),
                    User.is_admin == False
                ).count()
            }
            return {'stats': stats}
    except Exception as e:
        print(f"Error in inject_manager_stats: {e}")
    
    return {'stats': {}}


@app.route('/manager/advance_requests')
@login_required
def manager_advance_requests():
    """طلبات السلف"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
    department_ids = [dept.id for dept in managed_departments]
    
    advance_requests = db_session.query(AdvanceRequest).filter(
        AdvanceRequest.department_id.in_(department_ids)
    ).order_by(AdvanceRequest.created_at.desc()).all()
    
    # Load related data
    for request in advance_requests:
        request.user = db_session.query(User).get(request.user_id)
        request.department = db_session.query(Department).get(request.department_id)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager/manager_advance_requests.html',
                         advance_requests=advance_requests,
                         notifications=notifications)



def is_within_rejection_period(submit_date, reject_date=None):
    """
    التحقق مما إذا كان تاريخ الرفض يقع في نفس الفترة الزمنية لتقديم الطلب
    الفترات: من 26 من الشهر إلى 25 من الشهر التالي
    """
    if not reject_date:
        reject_date = date.today()
    
    # تحويل التواريخ إذا كانت نصوصاً
    if isinstance(submit_date, str):
        submit_date = datetime.strptime(submit_date, '%Y-%m-%d').date()
    if isinstance(reject_date, str):
        reject_date = datetime.strptime(reject_date, '%Y-%m-%d').date()
    
    # تحديد الفترة الزمنية لتاريخ تقديم الطلب
    if submit_date.day >= 26:
        # الفترة: من 26 من الشهر الحالي إلى 25 من الشهر التالي
        period_start = date(submit_date.year, submit_date.month, 26)
        if submit_date.month == 12:
            period_end = date(submit_date.year + 1, 1, 25)
        else:
            period_end = date(submit_date.year, submit_date.month + 1, 25)
    else:
        # الفترة: من 26 من الشهر السابق إلى 25 من الشهر الحالي
        if submit_date.month == 1:
            period_start = date(submit_date.year - 1, 12, 26)
            period_end = date(submit_date.year, 1, 25)
        else:
            period_start = date(submit_date.year, submit_date.month - 1, 26)
            period_end = date(submit_date.year, submit_date.month, 25)
    
    # التحقق مما إذا كان تاريخ الرفض يقع ضمن هذه الفترة
    return period_start <= reject_date <= period_end


def get_rejection_period_dates(check_date=None):
    """
    الحصول على تواريخ بداية ونهاية الفترة الزمنية للرفض
    """
    if not check_date:
        check_date = date.today()
    
    if check_date.day >= 26:
        # الفترة: من 26 من الشهر الحالي إلى 25 من الشهر التالي
        period_start = date(check_date.year, check_date.month, 26)
        if check_date.month == 12:
            period_end = date(check_date.year + 1, 1, 25)
        else:
            period_end = date(check_date.year, check_date.month + 1, 25)
    else:
        # الفترة: من 26 من الشهر السابق إلى 25 من الشهر الحالي
        if check_date.month == 1:
            period_start = date(check_date.year - 1, 12, 26)
            period_end = date(check_date.year, 1, 25)
        else:
            period_start = date(check_date.year, check_date.month - 1, 26)
            period_end = date(check_date.year, check_date.month, 25)
    
    return period_start, period_end


def can_reject_leave_by_financial_month(leave_date):
    """
    التحقق من رفض الإجازة بناءً على نطاق الشهر المالي
    القاعدة: لا يمكن رفض الإجازة بعد الوصول إلى يوم 26 من نفس الشهر المالي
    """
    if not leave_date:
        return True
    
    # تحويل التاريخ إذا كان نصياً
    if isinstance(leave_date, str):
        try:
            leave_date = datetime.strptime(leave_date, '%Y-%m-%d').date()
        except:
            return True
    
    today = date.today()
    
    # إذا كان تاريخ الإجازة في الماضي، لا يمكن رفضه
    if leave_date < today:
        return False
    
    # تحديد الشهر المالي لتاريخ الإجازة
    # الشهر المالي يبدأ من 26 الشهر السابق إلى 25 الشهر الحالي
    if leave_date.day >= 26:
        # يقع في الشهر المالي التالي
        financial_month_26th = date(leave_date.year, leave_date.month, 26)
    else:
        # يقع في الشهر المالي الحالي
        if leave_date.month == 1:
            financial_month_26th = date(leave_date.year - 1, 12, 26)
        else:
            financial_month_26th = date(leave_date.year, leave_date.month - 1, 26)
    
    # إذا وصلنا إلى يوم 26 من الشهر المالي، لا يمكن الرفض
    if today >= financial_month_26th:
        return False
    
    return True



@app.route('/manager/approve_advance/<int:request_id>')
@login_required
def manager_approve_advance(request_id):
    """موافقة المدير على طلب سلفة"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    advance_request = db_session.query(AdvanceRequest).get(request_id)
    
    # Check if manager manages this department
    department = db_session.query(Department).get(advance_request.department_id)
    
    if advance_request.status == 'pending':
        advance_request.status = 'approved'
        advance_request.approved_by = current_user.id
        advance_request.approved_at = datetime.now()
        
        # Update balance
        balance = db_session.query(EmployeeBalance).filter_by(user_id=advance_request.user_id).first()
        if balance:
            balance.advance_balance += advance_request.amount
            balance.last_updated = datetime.now()
        
        # Notify employee
        create_notification(
            advance_request.user_id,
            'موافقة على طلب السلفة',
            f'تم الموافقة على طلب سلفتك بقيمة {advance_request.amount} جنيه',
            'advance_approved',
            action_url=url_for('user_advance_requests')
        )
        
        db_session.commit()
        flash('تم الموافقة على طلب السلفة')
    
    return redirect(url_for('manager_advance_requests'))


@app.route('/admin/schedules')
@login_required
def admin_schedules():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    # Get all departments
    departments = db_session.query(Department).all()
    
    # Get all schedules
    schedules = db_session.query(WeeklySchedule).order_by(
        WeeklySchedule.week_start_date.desc()
    ).all()
    
    # Load related data
    for schedule in schedules:
        schedule.department = db_session.query(Department).get(schedule.department_id)
        schedule.creator = db_session.query(User).get(schedule.created_by)
        if schedule.approved_by:
            schedule.approver = db_session.query(User).get(schedule.approved_by)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_schedules.html',
                         schedules=schedules,
                         departments=departments,
                         notifications=notifications)




# ======== Schedule Structure Routes ========
@app.route('/admin/schedule_structure/<int:dept_id>', methods=['GET', 'POST'])
@login_required
def admin_schedule_structure(dept_id):
    """إدارة هيكل الجدول للقسم"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    department = db_session.get(Department, dept_id)
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('admin_departments'))
    
    # GET request - عرض هيكل الجدول
    if request.method == 'GET':
        # الحصول على صفوف هيكل الجدول
        structure_rows = db_session.query(ScheduleStructureRow).filter_by(
            department_id=dept_id
        ).order_by('row_order').all()
        
        notifications = get_user_notifications(current_user.id)
        
        return render_template('admin_schedule_structure.html',
                             department=department,
                             structure_rows=structure_rows,
                             notifications=notifications)
    
    # POST request - معالجة تحديثات الهيكل
    try:
        # إضافة صف جديد
        if 'add_row' in request.form:
            job_title = request.form.get('job_title', '').strip()
            
            if job_title:
                # العثور على أعلى ترتيب
                max_order = db_session.query(func.max(ScheduleStructureRow.row_order)).filter_by(
                    department_id=dept_id
                ).scalar() or 0
                
                new_row = ScheduleStructureRow(
                    department_id=dept_id,
                    job_title=job_title,
                    morning_shift="",
                    evening_shift="",
                    night_shift="",
                    row_order=max_order + 1,
                    created_by=current_user.id
                )
                db_session.add(new_row)
                db_session.commit()
                flash('تم إضافة الصف بنجاح', 'success')
            else:
                flash('يرجى إدخال اسم الوظيفة', 'error')
        
        # حفظ التغييرات على الصفوف الحالية
        elif 'save_changes' in request.form:
            # الحصول على جميع الصفوف الحالية
            rows = db_session.query(ScheduleStructureRow).filter_by(
                department_id=dept_id
            ).all()
            
            for row in rows:
                row_id = str(row.id)
                
                # تحديث اسم الوظيفة
                job_title = request.form.get(f'job_title_{row_id}')
                if job_title:
                    row.job_title = job_title.strip()
                
                # تحديث أسماء الموظفين في الشيفتات
                row.morning_shift = request.form.get(f'morning_shift_{row_id}', '').strip()
                row.evening_shift = request.form.get(f'evening_shift_{row_id}', '').strip()
                row.night_shift = request.form.get(f'night_shift_{row_id}', '').strip()
                
                # تحديث الترتيب
                order = request.form.get(f'order_{row_id}')
                if order and order.isdigit():
                    row.row_order = int(order)
            
            # تحديث إصدار الهيكل
            department.schedule_structure_version = (department.schedule_structure_version or 0) + 1
            department.structure_last_modified = datetime.now()
            
            db_session.commit()
            flash('تم حفظ التغييرات بنجاح', 'success')
        
        # حذف صف
        elif 'delete_row' in request.form:
            row_id = request.form.get('delete_row')
            if row_id and row_id.isdigit():
                row = db_session.query(ScheduleStructureRow).get(int(row_id))
                if row and row.department_id == dept_id:
                    db_session.delete(row)
                    
                    # تحديث ترتيب الصفوف المتبقية
                    remaining_rows = db_session.query(ScheduleStructureRow).filter_by(
                        department_id=dept_id
                    ).order_by('row_order').all()
                    
                    for i, remaining_row in enumerate(remaining_rows, 1):
                        remaining_row.row_order = i
                    
                    # تحديث إصدار الهيكل
                    department.schedule_structure_version = (department.schedule_structure_version or 0) + 1
                    department.structure_last_modified = datetime.now()
                    
                    db_session.commit()
                    flash('تم حذف الصف بنجاح', 'success')
                else:
                    flash('الصف غير موجود', 'error')
            else:
                flash('معرف الصف غير صالح', 'error')
        
        # رفع ملف Excel
        elif 'excel_file' in request.files:
            file = request.files['excel_file']
            if file and file.filename.endswith(('.xlsx', '.xls')):
                try:
                    # قراءة ملف Excel
                    import pandas as pd
                    df = pd.read_excel(file)
                    
                    # التحقق من وجود الأعمدة المطلوبة
                    required_columns = ['الوظيفة', 'الشيفت الصباحي', 'الشيفت المسائي', 'شيفت السهر']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    
                    if missing_columns:
                        flash(f'الأعمدة المفقودة: {", ".join(missing_columns)}', 'error')
                        return redirect(url_for('admin_schedule_structure', dept_id=dept_id))
                    
                    # حذف جميع الصفوف الحالية
                    db_session.query(ScheduleStructureRow).filter_by(
                        department_id=dept_id
                    ).delete()
                    
                    # إضافة الصفوف الجديدة من Excel
                    row_order = 1
                    for _, row_data in df.iterrows():
                        job_title = str(row_data.get('الوظيفة', '')).strip()
                        if not job_title:
                            continue
                        
                        new_row = ScheduleStructureRow(
                            department_id=dept_id,
                            job_title=job_title,
                            morning_shift=str(row_data.get('الشيفت الصباحي', '')).strip(),
                            evening_shift=str(row_data.get('الشيفت المسائي', '')).strip(),
                            night_shift=str(row_data.get('شيفت السهر', '')).strip(),
                            row_order=row_order,
                            created_by=current_user.id
                        )
                        db_session.add(new_row)
                        row_order += 1
                    
                    # تحديث إصدار الهيكل
                    department.schedule_structure_version = (department.schedule_structure_version or 0) + 1
                    department.structure_last_modified = datetime.now()
                    
                    db_session.commit()
                    flash(f'تم استيراد {row_order-1} صف بنجاح', 'success')
                    
                except Exception as e:
                    db_session.rollback()
                    flash(f'خطأ في معالجة ملف Excel: {str(e)}', 'error')
            else:
                flash('يرجى رفع ملف Excel صالح', 'error')
        
        # إعادة ترتيب الصفوف (يمكن إضافتها إذا أردت)
        elif 'reorder_rows' in request.form:
            order_data = request.form.get('order_data')
            if order_data:
                try:
                    order_list = json.loads(order_data)
                    for item in order_list:
                        row_id = item.get('id')
                        new_order = item.get('order')
                        if row_id and new_order:
                            row = db_session.query(ScheduleStructureRow).get(int(row_id))
                            if row and row.department_id == dept_id:
                                row.row_order = int(new_order)
                    
                    db_session.commit()
                    flash('تم إعادة ترتيب الصفوف بنجاح', 'success')
                except Exception as e:
                    flash(f'خطأ في إعادة الترتيب: {str(e)}', 'error')
        
        return redirect(url_for('admin_schedule_structure', dept_id=dept_id))
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('admin_schedule_structure', dept_id=dept_id))


@app.route('/admin/export_schedule_structure/<int:dept_id>', methods=['POST'])
@login_required
def export_schedule_structure(dept_id):
    """تصدير هيكل الجدول إلى Excel"""
    
    try:
        department = db_session.get(Department, dept_id)
        if not department:
            flash('القسم غير موجود', 'error')
            return redirect(url_for('admin_departments'))
        
        # الحصول على صفوف الهيكل
        structure_rows = db_session.query(ScheduleStructureRow).filter_by(
            department_id=dept_id
        ).order_by('row_order').all()
        
        # إنشاء DataFrame بنفس تنسيق ملفك الأصلي
        data = []
        days = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        
        for day in days:
            for row in structure_rows:
                data.append({
                    'شيفت السهر': row.night_shift or '',
                    'الشيفت المسائي': row.evening_shift or '',
                    'الشيفت الصباحي': row.morning_shift or '',
                    'الوظيفة': row.job_title,
                    'القسم': department.name,
                    'التاريخ': '',  # يمكنك إضافة تاريخ إذا أردت
                    'اليوم': day
                })
        
        df = pd.DataFrame(data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='هيكل الجدول', index=False)
            
            # تحسين التنسيق
            worksheet = writer.sheets['هيكل الجدول']
            column_widths = {
                'A': 20,  # شيفت السهر
                'B': 20,  # الشيفت المسائي
                'C': 20,  # الشيفت الصباحي
                'D': 30,  # الوظيفة
                'E': 20,  # القسم
                'F': 15,  # التاريخ
                'G': 15   # اليوم
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
        
        output.seek(0)
        
        # اسم الملف
        filename = f"هيكل_الجدول_{department.name.replace(' ', '_')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'خطأ في التصدير: {str(e)}', 'error')
        return redirect(url_for('admin_schedule_structure', dept_id=dept_id))

@app.template_filter('unique')
def unique_filter(sequence, attribute=None):
    """فلتر للقالب للحصول على قيم فريدة"""
    if attribute:
        seen = set()
        result = []
        for item in sequence:
            value = getattr(item, attribute)
            if value not in seen:
                seen.add(value)
                result.append(item)
        return result
    else:
        return list(set(sequence))

@app.route('/admin/schedule_structure/update/<int:dept_id>', methods=['POST'])
@login_required
def admin_schedule_structure_update(dept_id):
    """تحديث هيكل الجدول"""
    
    try:
        department = db_session.get(Department, dept_id)
        if not department:
            flash('القسم غير موجود', 'error')
            return redirect(url_for('admin_departments'))
        
        # إضافة صف جديد
        if 'add_row' in request.form:
            job_title = request.form.get('job_title', '').strip()
            
            if job_title:
                # العثور على أعلى ترتيب
                max_order = db_session.query(func.max(ScheduleStructureRow.row_order)).filter_by(
                    department_id=dept_id
                ).scalar() or 0
                
                new_row = ScheduleStructureRow(
                    department_id=dept_id,
                    job_title=job_title,
                    job_code=None,  # إضافة job_code
                    morning_shift="",
                    evening_shift="",
                    night_shift="",
                    row_order=max_order + 1,
                    created_by=current_user.id
                )
                db_session.add(new_row)
                db_session.commit()
                flash('تم إضافة الصف بنجاح', 'success')
            else:
                flash('يرجى إدخال اسم الوظيفة', 'error')
        
        # حفظ التغييرات على الصفوف الحالية
        elif 'save_changes' in request.form:
            # الحصول على جميع الصفوف الحالية
            rows = db_session.query(ScheduleStructureRow).filter_by(
                department_id=dept_id
            ).all()
            
            for row in rows:
                row_id = str(row.id)
                
                # تحديث اسم الوظيفة
                job_title = request.form.get(f'job_title_{row_id}')
                if job_title:
                    row.job_title = job_title.strip()
                
                # تحديث كود الوظيفة - هذا الجزء الجديد
                job_code = request.form.get(f'job_code_{row_id}')
                if job_code is not None:  # حتى لو كان فارغاً
                    row.job_code = job_code.strip() if job_code.strip() else None
                
                # تحديث أسماء الموظفين في الشيفتات
                row.morning_shift = request.form.get(f'morning_shift_{row_id}', '').strip()
                row.evening_shift = request.form.get(f'evening_shift_{row_id}', '').strip()
                row.night_shift = request.form.get(f'night_shift_{row_id}', '').strip()
                
                # تحديث الترتيب
                order = request.form.get(f'order_{row_id}')
                if order and order.isdigit():
                    row.row_order = int(order)
            
            # تحديث إصدار الهيكل
            department.schedule_structure_version = (department.schedule_structure_version or 0) + 1
            department.structure_last_modified = datetime.now()
            
            db_session.commit()
            flash('تم حفظ التغييرات بنجاح', 'success')
        
        # حذف صف
        elif 'delete_row' in request.form:
            row_id = request.form.get('delete_row')
            if row_id and row_id.isdigit():
                row = db_session.query(ScheduleStructureRow).get(int(row_id))
                if row and row.department_id == dept_id:
                    db_session.delete(row)
                    
                    # تحديث ترتيب الصفوف المتبقية
                    remaining_rows = db_session.query(ScheduleStructureRow).filter_by(
                        department_id=dept_id
                    ).order_by('row_order').all()
                    
                    for i, remaining_row in enumerate(remaining_rows, 1):
                        remaining_row.row_order = i
                    
                    # تحديث إصدار الهيكل
                    department.schedule_structure_version = (department.schedule_structure_version or 0) + 1
                    department.structure_last_modified = datetime.now()
                    
                    db_session.commit()
                    flash('تم حذف الصف بنجاح', 'success')
                else:
                    flash('الصف غير موجود', 'error')
            else:
                flash('معرف الصف غير صالح', 'error')
        
        # رفع ملف Excel
        elif 'excel_file' in request.files:
            file = request.files['excel_file']
            if file and file.filename.endswith(('.xlsx', '.xls')):
                try:
                    # قراءة ملف Excel
                    import pandas as pd
                    df = pd.read_excel(file)
                    
                    # التحقق من وجود الأعمدة المطلوبة (مع إضافة كود الوظيفة)
                    required_columns = ['الوظيفة', 'كود الوظيفة', 'الشيفت الصباحي', 'الشيفت المسائي', 'شيفت السهر']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    
                    if missing_columns:
                        flash(f'الأعمدة المفقودة: {", ".join(missing_columns)}', 'error')
                        return redirect(url_for('admin_schedule_structure', dept_id=dept_id))
                    
                    # حذف جميع الصفوف الحالية
                    db_session.query(ScheduleStructureRow).filter_by(
                        department_id=dept_id
                    ).delete()
                    
                    # إضافة الصفوف الجديدة من Excel
                    row_order = 1
                    for _, row_data in df.iterrows():
                        job_title = str(row_data.get('الوظيفة', '')).strip()
                        if not job_title:
                            continue
                        
                        # قراءة كود الوظيفة
                        job_code = row_data.get('كود الوظيفة', '')
                        if pd.notna(job_code):
                            job_code = str(job_code).strip()
                        else:
                            job_code = None
                        
                        new_row = ScheduleStructureRow(
                            department_id=dept_id,
                            job_title=job_title,
                            job_code=job_code,  # إضافة job_code
                            morning_shift=str(row_data.get('الشيفت الصباحي', '')).strip(),
                            evening_shift=str(row_data.get('الشيفت المسائي', '')).strip(),
                            night_shift=str(row_data.get('شيفت السهر', '')).strip(),
                            row_order=row_order,
                            created_by=current_user.id
                        )
                        db_session.add(new_row)
                        row_order += 1
                    
                    # تحديث إصدار الهيكل
                    department.schedule_structure_version = (department.schedule_structure_version or 0) + 1
                    department.structure_last_modified = datetime.now()
                    
                    db_session.commit()
                    flash(f'تم استيراد {row_order-1} صف بنجاح', 'success')
                    
                except Exception as e:
                    db_session.rollback()
                    flash(f'خطأ في معالجة ملف Excel: {str(e)}', 'error')
            else:
                flash('يرجى رفع ملف Excel صالح', 'error')
        
        return redirect(url_for('admin_schedule_structure', dept_id=dept_id))
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('admin_schedule_structure', dept_id=dept_id))


@app.route('/admin/update_all_schedules', methods=['POST'])
@login_required
def update_all_schedules():
    """تحديث جميع الجداول بهيكل القسم الحالي"""
    
    try:
        department_id = request.form.get('department_id')
        if not department_id:
            return jsonify({'success': False, 'message': 'معرف القسم مطلوب'})
        
        department = db_session.query(Department).get(department_id)
        if not department:
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
        
        # الحصول على جميع الجداول غير المعتمدة للقسم
        schedules = db_session.query(WeeklySchedule).filter_by(
            department_id=department_id,
            is_approved=False
        ).all()
        
        updated_count = 0
        
        for schedule in schedules:
            try:
                # استخدام الدالة المحسنة للمزامنة
                if sync_schedule_with_department(schedule.id, db_session):
                    updated_count += 1
                    print(f"✓ تم تحديث جدول {schedule.id} للفترة {schedule.week_start_date}")
                else:
                    print(f"✗ فشل تحديث جدول {schedule.id}")
            except Exception as e:
                print(f"❌ خطأ في تحديث جدول {schedule.id}: {str(e)}")
                continue
        
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم تحديث {updated_count} جدول',
            'updated_count': updated_count
        })
        
    except Exception as e:
        db_session.rollback()
        print(f"❌ خطأ عام في تحديث الجداول: {str(e)}")
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@app.route('/admin/regenerate_all_from_structure/<int:dept_id>')
@login_required
def regenerate_all_from_structure(dept_id):
    """إعادة توليد جميع الجداول من الهيكل الحالي"""
    
    try:
        department = db_session.get(Department, dept_id)
        if not department:
            flash('القسم غير موجود', 'error')
            return redirect(url_for('admin_departments'))
        
        # الحصول على جميع الجداول للقسم
        schedules = db_session.query(WeeklySchedule).filter_by(
            department_id=dept_id
        ).all()
        
        regenerated_count = 0
        
        for schedule in schedules:
            if schedule.sync_with_department_structure(db_session, force=True):
                regenerated_count += 1
        
        db_session.commit()
        
        flash(f'تم إعادة توليد {regenerated_count} جدول بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('admin_schedule_structure', dept_id=dept_id))


def get_week_start_date_saturday(date_obj):
    """الحصول على تاريخ بداية الأسبوع (السبت) بشكل صحيح"""
    # date_obj.weekday(): Monday=0, Tuesday=1, Wednesday=2, Thursday=3, Friday=4, Saturday=5, Sunday=6
    # نريد: Saturday=0, Sunday=1, ..., Friday=6
    
    # Calculate days to subtract to get to Saturday
    # If today is Monday (0), we need to go back 2 days to get to Saturday (5)
    days_since_saturday = (date_obj.weekday() - 5) % 7
    if days_since_saturday == 0:
        # Already Saturday, return the same date
        return date_obj
    else:
        # Go back to the most recent Saturday
        return date_obj - timedelta(days=days_since_saturday)
    
@app.route('/admin/delete_schedule_confirm/<int:schedule_id>', methods=['GET', 'POST'])
@login_required
def delete_schedule_confirm(schedule_id):
    """صفحة تأكيد حذف جدول"""
    
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('admin_weekly_schedules'))
    
    department = db_session.query(Department).get(schedule.department_id)
    
    if request.method == 'POST':
        try:
            # الحصول على إحصاءات قبل الحذف
            detail_count = db_session.query(ScheduleDetail).filter_by(
                weekly_schedule_id=schedule_id
            ).count()
            
            # حذف الجدول باستخدام الدالة المساعدة
            success, message = delete_schedule_with_related_data(schedule_id)
            
            if success:
                flash(message, 'success')
            else:
                flash(message, 'error')
                
            return redirect(url_for('admin_weekly_schedules'))
        
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'error')
            return redirect(url_for('admin_weekly_schedules'))
    
    # GET request - عرض صفحة التأكيد
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/confirm_delete_schedule.html',
                         schedule=schedule,
                         department=department,
                         notifications=notifications)


def delete_schedule_with_related_data(schedule_id):
    """دالة مساعدة لحذف الجدول مع البيانات المرتبطة"""
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return False, 'الجدول غير موجود'
        
        # إحصاءات قبل الحذف
        detail_count = db_session.query(ScheduleDetail).filter_by(
            weekly_schedule_id=schedule_id
        ).count()
        
        history_count = db_session.query(ScheduleApprovalHistory).filter_by(
            schedule_id=schedule_id
        ).count()
        
        # الحذف
        db_session.query(ScheduleDetail).filter_by(
            weekly_schedule_id=schedule_id
        ).delete()
        
        db_session.query(ScheduleApprovalHistory).filter_by(
            schedule_id=schedule_id
        ).delete()
        
        db_session.delete(schedule)
        db_session.commit()
        
        # تسجيل في سجل النظام
        log_deletion(current_user.id, 'schedule', schedule_id, {
            'detail_count': detail_count,
            'history_count': history_count,
            'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
            'department_id': schedule.department_id
        })
        
        message = f'تم حذف الجدول بنجاح (تم حذف {detail_count} سجل تفصيلي)'
        return True, message
        
    except Exception as e:
        db_session.rollback()
        return False, f'حدث خطأ أثناء الحذف: {str(e)}'


def log_deletion(user_id, entity_type, entity_id, details):
    """تسجيل عمليات الحذف"""
    try:
        log_entry = DeletionLog(
            user_id=user_id,
            entity_type=entity_type,
            entity_id=entity_id,
            details=json.dumps(details, ensure_ascii=False),
            deleted_at=datetime.now()
        )
        db_session.add(log_entry)
        db_session.commit()
    except Exception as e:
        print(f"خطأ في تسجيل عملية الحذف: {e}")


@app.route('/admin/delete_multiple_schedules', methods=['POST'])
@login_required
def delete_multiple_schedules():
    """حذف عدة جداول مرة واحدة"""
    
    try:
        data = request.get_json()
        schedule_ids = data.get('schedule_ids', [])
        
        if not schedule_ids:
            return jsonify({'success': False, 'message': 'لم يتم تحديد جداول للحذف'})
        
        deleted_count = 0
        failed_count = 0
        
        for schedule_id in schedule_ids:
            try:
                schedule = db_session.query(WeeklySchedule).get(schedule_id)
                if not schedule:
                    failed_count += 1
                    continue
                
                # حذف البيانات المرتبطة
                db_session.query(ScheduleDetail).filter_by(
                    weekly_schedule_id=schedule_id
                ).delete()
                
                db_session.query(ScheduleApprovalHistory).filter_by(
                    schedule_id=schedule_id
                ).delete()
                
                # حذف الجدول
                db_session.delete(schedule)
                deleted_count += 1
                
            except Exception as e:
                failed_count += 1
                print(f"خطأ في حذف الجدول {schedule_id}: {str(e)}")
        
        db_session.commit()
        
        message = f'تم حذف {deleted_count} جدول'
        if failed_count > 0:
            message += f'، فشل في حذف {failed_count} جدول'
        
        return jsonify({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'failed_count': failed_count
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})



@app.route('/admin/create_schedules_from_structure', methods=['POST'])
@login_required
def create_schedules_from_structure():
    """إنشاء جداول جديدة من الهيكل الحالي"""
    
    try:
        data = request.get_json()
        department_id = data.get('department_id')
        weeks_count = data.get('weeks_count', 4)
        
        if not department_id:
            return jsonify({'success': False, 'message': 'معرف القسم مطلوب'})
        
        department = db_session.query(Department).get(department_id)
        if not department:
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
        
        created_schedules = []
        
        # التاريخ المحدد لبداية الجداول
        base_start_date = date(2025, 12, 20)  # 20 ديسمبر 2025
        
        for i in range(weeks_count):
            # حساب تاريخ بداية الأسبوع
            week_start_date = base_start_date + timedelta(days=(i * 7))
            week_end_date = week_start_date + timedelta(days=6)
            
            # التحقق من اليوم بشكل صحيح
            day_name = get_arabic_day_name(week_start_date)
            print(f"الأسبوع {i+1}: يبدأ من {week_start_date} ({day_name}, weekday={week_start_date.weekday()})")
            
            # تحقق إذا كان تاريخ البداية هو السبت (اختياري)
            if week_start_date.weekday() != 5:
                print(f"ملاحظة: تاريخ البداية {week_start_date} ليس يوم السبت، بل هو {day_name}")
            
            # التحقق من عدم وجود جدول لهذا الأسبوع
            existing_schedule = db_session.query(WeeklySchedule).filter_by(
                department_id=department_id,
                week_start_date=week_start_date
            ).first()
            
            if existing_schedule:
                print(f"يوجد already جدول للأسبوع {week_start_date}")
                continue
            
            # إنشاء جدول جديد
            week_number = week_start_date.isocalendar()[1]
            month = week_start_date.month
            year = week_start_date.year
            
            new_schedule = WeeklySchedule(
                department_id=department_id,
                week_start_date=week_start_date,
                week_end_date=week_end_date,
                week_number=week_number,
                month=month,
                year=year,
                created_by=current_user.id,
                is_template=False
            )
            
            db_session.add(new_schedule)
            db_session.flush()  # للحصول على ID
            
            # مزامنة الجدول مع هيكل القسم
            if new_schedule.sync_with_department_structure(db_session):
                created_schedules.append({
                    'id': new_schedule.id,
                    'start_date': week_start_date.strftime('%Y-%m-%d'),
                    'end_date': week_end_date.strftime('%Y-%m-%d'),
                    'day_name': day_name,
                    'day_weekday': week_start_date.weekday(),
                    'year': year,
                    'month': month
                })
        
        db_session.commit()
        
        # عرض تفاصيل الجداول المنشأة
        print(f"\n=== تم إنشاء {len(created_schedules)} جدول جديد ===")
        for schedule in created_schedules:
            print(f"✓ جدول {schedule['id']}: {schedule['start_date']} ({schedule['day_name']}) إلى {schedule['end_date']}")
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء {len(created_schedules)} جدول جديد',
            'created_count': len(created_schedules),
            'schedules': created_schedules,
            'start_date': '2025-12-20',
            'total_weeks': weeks_count
        })
        
    except Exception as e:
        db_session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@app.route('/debug/dates/<target_date>')
def debug_dates(target_date):
    """Debug route to check date calculations"""
    try:
        target = datetime.strptime(target_date, '%Y-%m-%d').date()
        
        results = {
            'target_date': target.strftime('%Y-%m-%d'),
            'target_weekday': target.weekday(),
            'arabic_day_name': get_arabic_day_name(target),
            'next_saturday': get_next_saturday(target).strftime('%Y-%m-%d'),
            'week_start_saturday': get_week_start_date_saturday(target).strftime('%Y-%m-%d'),
            'week_dates': get_week_dates_with_correct_days(get_week_start_date_saturday(target))
        }
        
        return jsonify(results)
    except Exception as e:
        return jsonify({'error': str(e)})

def get_correct_week_start(date_obj, start_day='saturday'):
    """الحصول على تاريخ بداية الأسبوع الصحيح"""
    # start_day: 'saturday' أو 'monday' أو 'sunday'
    
    weekday_map = {
        'saturday': 5,  # Saturday=5 في weekday()
        'sunday': 6,    # Sunday=6
        'monday': 0     # Monday=0
    }
    
    if start_day not in weekday_map:
        start_day = 'saturday'  # الافتراضي
    
    target_weekday = weekday_map[start_day]
    current_weekday = date_obj.weekday()
    
    # حساب عدد الأيام حتى يوم البداية المطلوب
    days_until_target = (target_weekday - current_weekday) % 7
    
    # إذا كان اليوم هو يوم البداية نفسه، نذهب للأسبوع القادم
    if days_until_target == 0:
        days_until_target = 7
    
    return date_obj + timedelta(days=days_until_target)


def get_next_saturday(start_date=None):
    """الحصول على تاريخ يوم السبت القادم بشكل صحيح"""
    if start_date is None:
        start_date = date.today()
    
    # date.weekday(): Monday=0, Tuesday=1, ..., Sunday=6
    # Saturday = 5
    
    # Calculate days until next Saturday
    days_until_saturday = (5 - start_date.weekday()) % 7
    
    # If today is Saturday, return next Saturday (add 7 days)
    if days_until_saturday == 0:
        days_until_saturday = 7
    
    next_saturday = start_date + timedelta(days=days_until_saturday)
    
    # Verify it's actually Saturday
    assert next_saturday.weekday() == 5, f"Date {next_saturday} is not Saturday!"
    
    return next_saturday

# ======== Weekly Schedule Routes ========

@app.route('/admin/weekly_schedules')
@login_required
def admin_weekly_schedules():
    """إدارة الجداول الأسبوعية"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    # الحصول على معاملات التصفية
    department_id = request.args.get('department', 'all')
    year = request.args.get('year', date.today().year)
    month = request.args.get('month', 'all')
    status = request.args.get('status', 'all')
    
    # بناء الاستعلام
    query = db_session.query(WeeklySchedule)
    
    if department_id != 'all':
        query = query.filter_by(department_id=department_id)
    
    # Handle year parameter properly - convert to int safely
    year_value = None
    if year:
        try:
            # Convert to int if it's a string
            if isinstance(year, str):
                if year.isdigit():
                    year_value = int(year)
            elif isinstance(year, int):
                year_value = year
                
            if year_value and 2000 <= year_value <= 2100:
                query = query.filter_by(year=year_value)
        except (ValueError, TypeError, AttributeError):
            # If conversion fails, ignore the filter
            pass
    
    if month != 'all' and month.isdigit():
        query = query.filter_by(month=int(month))
    
    if status != 'all':
        if status == 'approved':
            query = query.filter_by(is_approved=True)
        elif status == 'pending':
            query = query.filter_by(is_approved=False, is_locked=False)
        elif status == 'locked':
            query = query.filter_by(is_locked=True)
    
    schedules = query.order_by(
        WeeklySchedule.year.desc(),
        WeeklySchedule.week_number.desc()
    ).all()
    
    # تحميل البيانات المرتبطة
    for schedule in schedules:
        schedule.department = db_session.query(Department).get(schedule.department_id)
        schedule.creator = db_session.query(User).get(schedule.created_by)
        if schedule.approved_by:
            schedule.approver = db_session.query(User).get(schedule.approved_by)
    
    departments = db_session.query(Department).all()
    
    notifications = get_user_notifications(current_user.id)
    
    # FIXED: Safely determine current_year for template
    current_year = date.today().year
    if year:
        try:
            if isinstance(year, str) and year.isdigit():
                current_year = int(year)
            elif isinstance(year, int):
                current_year = year
        except (ValueError, TypeError, AttributeError):
            pass
    
    return render_template('admin/admin_weekly_schedules.html',
                         schedules=schedules,
                         departments=departments,
                         notifications=notifications,
                         current_year=current_year,
                         selected_department=department_id,
                         selected_month=month,
                         selected_status=status)

@app.route('/admin/weekly_schedule/<int:schedule_id>')
@login_required
def admin_weekly_schedule_detail(schedule_id):
    """تفاصيل الجدول الأسبوعي"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('admin_weekly_schedules'))
    
    # تحميل التفاصيل
    schedule_details = db_session.query(ScheduleDetail).filter_by(
        weekly_schedule_id=schedule_id
    ).order_by('day_date', 'row_order').all()
    
    # تحميل البيانات المرتبطة
    schedule.department = db_session.query(Department).get(schedule.department_id)
    schedule.creator = db_session.query(User).get(schedule.created_by)
    if schedule.approved_by:
        schedule.approver = db_session.query(User).get(schedule.approved_by)
    
    # تجميع التفاصيل حسب اليوم
    details_by_day = {}
    for detail in schedule_details:
        day_key = detail.day_date.strftime('%Y-%m-%d')
        if day_key not in details_by_day:
            details_by_day[day_key] = {
                'date': detail.day_date,
                'day_name': detail.day_name,
                'details': []
            }
        details_by_day[day_key]['details'].append(detail)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_weekly_schedule_detail.html',
                         schedule=schedule,
                         details_by_day=details_by_day,
                         notifications=notifications)


@app.route('/admin/create_weekly_schedule', methods=['GET', 'POST'])
@login_required
def create_weekly_schedule():
    """إنشاء جدول أسبوعي جديد"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    if request.method == 'POST':
        try:
            department_id = request.form['department_id']
            week_start_date = datetime.strptime(request.form['week_start_date'], '%Y-%m-%d').date()
            
            # التحقق من عدم وجود جدول لهذا الأسبوع
            existing_schedule = db_session.query(WeeklySchedule).filter_by(
                department_id=department_id,
                week_start_date=week_start_date
            ).first()
            
            if existing_schedule:
                flash('يوجد already جدول لنفس الأسبوع', 'error')
                return redirect(url_for('create_weekly_schedule'))
            
            # حساب التواريخ
            week_end_date = week_start_date + timedelta(days=6)
            week_number = week_start_date.isocalendar()[1]
            month = week_start_date.month
            year = week_start_date.year
            
            # إنشاء الجدول
            new_schedule = WeeklySchedule(
                department_id=department_id,
                week_start_date=week_start_date,
                week_end_date=week_end_date,
                week_number=week_number,
                month=month,
                year=year,
                created_by=current_user.id,
                is_template='is_template' in request.form
            )
            
            db_session.add(new_schedule)
            db_session.flush()  # للحصول على ID
            
            # مزامنة مع هيكل القسم
            new_schedule.sync_with_department_structure(db_session)
            
            db_session.commit()
            
            flash('تم إنشاء الجدول بنجاح', 'success')
            return redirect(url_for('admin_weekly_schedule_detail', schedule_id=new_schedule.id))
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'error')
    
    departments = db_session.query(Department).all()
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/create_weekly_schedule.html',
                         departments=departments,
                         notifications=notifications)


@app.route('/admin/approve_schedule/<int:schedule_id>')
@login_required
def approve_schedule(schedule_id):
    """اعتماد الجدول الأسبوعي"""
    
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('admin_weekly_schedules'))
    
    if schedule.is_approved:
        flash('الجدول معتمد already', 'warning')
        return redirect(url_for('admin_weekly_schedule_detail', schedule_id=schedule_id))
    
    schedule.is_approved = True
    schedule.approved_by = current_user.id
    schedule.approved_at = datetime.now()
    
    # إنشاء سجل في تاريخ الاعتماد
    approval_history = ScheduleApprovalHistory(
        schedule_id=schedule_id,
        action='approved',
        comments='تم اعتماد الجدول',
        performed_by=current_user.id
    )
    db_session.add(approval_history)
    
    db_session.commit()
    
    flash('تم اعتماد الجدول بنجاح', 'success')
    return redirect(url_for('admin_weekly_schedule_detail', schedule_id=schedule_id))


@app.route('/admin/lock_schedule/<int:schedule_id>')
@login_required
def lock_schedule(schedule_id):
    """قفل الجدول الأسبوعي"""
    
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('admin_weekly_schedules'))
    
    if schedule.is_locked:
        flash('الجدول مقفل already', 'warning')
        return redirect(url_for('admin_weekly_schedule_detail', schedule_id=schedule_id))
    
    schedule.is_locked = True
    schedule.updated_by = current_user.id
    schedule.updated_at = datetime.now()
    
    db_session.commit()
    
    flash('تم قفل الجدول بنجاح', 'success')
    return redirect(url_for('admin_weekly_schedule_detail', schedule_id=schedule_id))

@app.route('/manager/unlock_schedule/<int:schedule_id>')
@login_required
def manager_unlock_schedule(schedule_id):
    """فتح جدول من قبل مدير (يسمح فقط للمستخدمين الخاصين)"""
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('manager_schedules'))
    
    department = db_session.query(Department).get(schedule.department_id)
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('manager_schedules'))
    
    # التحقق من أن المستخدم مدير لهذا القسم أو من المستخدمين الخاصين
    managed_dept_ids = get_managed_department_ids(current_user.id)
    if schedule.department_id not in managed_dept_ids and current_user.username not in ['98', '510', '100']:
        flash('ليس لديك صلاحية لفتح هذا الجدول', 'error')
        return redirect(url_for('manager_schedules'))
    
    schedule.is_locked = False
    schedule.is_approved = False
    schedule.updated_by = current_user.id
    schedule.updated_at = datetime.now()
    db_session.commit()
    
    flash('تم فتح قفل الجدول بنجاح', 'success')
    return redirect(url_for('manager_edit_schedule', schedule_id=schedule.id))

@app.route('/admin/unlock_schedule/<int:schedule_id>')
@login_required
def unlock_schedule(schedule_id):
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('admin_weekly_schedules'))
    
    # السماح للمسؤول أو المستخدمين الخاصين
    if not (current_user.is_admin or current_user.username in SPECIAL_USERS):
        flash('غير مصرح لك بفتح الجدول', 'error')
        return redirect(url_for('admin_weekly_schedules'))
    
    schedule.is_locked = False
    schedule.updated_by = current_user.id
    schedule.updated_at = datetime.now()
    db_session.commit()
    
    flash('تم فتح قفل الجدول بنجاح', 'success')
    return redirect(url_for('admin_weekly_schedule_detail', schedule_id=schedule_id))

@app.route('/admin/delete_schedule/<int:schedule_id>')
@login_required
def delete_schedule(schedule_id):
    """حذف الجدول الأسبوعي"""

    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('admin_weekly_schedules'))
    
    if schedule.is_approved and not current_user.is_admin:
        flash('لا يمكن حذف جدول معتمد', 'error')
        return redirect(url_for('admin_weekly_schedule_detail', schedule_id=schedule_id))
    
    try:
        # حذف التفاصيل المرتبطة أولاً
        db_session.query(ScheduleDetail).filter_by(
            weekly_schedule_id=schedule_id
        ).delete()
        
        # حذف سجل الاعتماد
        db_session.query(ScheduleApprovalHistory).filter_by(
            schedule_id=schedule_id
        ).delete()
        
        # حذف الجدول
        db_session.delete(schedule)
        db_session.commit()
        
        flash('تم حذف الجدول بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء الحذف: {str(e)}', 'error')
    
    return redirect(url_for('admin_weekly_schedules'))


@app.route('/admin/update_schedule_detail/<int:detail_id>', methods=['POST'])
@login_required
def update_schedule_detail(detail_id):
    """تحديث تفصيلة في الجدول"""

    
    try:
        detail = db_session.query(ScheduleDetail).get(detail_id)
        if not detail:
            return jsonify({'success': False, 'message': 'التفصيلة غير موجودة'})
        
        # التحقق من أن الجدول غير مقفل
        schedule = db_session.query(WeeklySchedule).get(detail.weekly_schedule_id)
        if schedule.is_locked:
            return jsonify({'success': False, 'message': 'الجدول مقفل ولا يمكن التعديل'})
        
        # تحديث البيانات
        data = request.get_json()
        if 'morning_shift' in data:
            detail.morning_shift = data['morning_shift']
        if 'evening_shift' in data:
            detail.evening_shift = data['evening_shift']
        if 'night_shift' in data:
            detail.night_shift = data['night_shift']
        if 'notes' in data:
            detail.notes = data['notes']
        
        detail.modified_by = current_user.id
        detail.modified_at = datetime.now()
        
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم التحديث بنجاح',
            'detail': detail.to_dict()
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})


@app.route('/admin/copy_schedule_row/<int:schedule_id>', methods=['POST'])
@login_required
def copy_schedule_row(schedule_id):
    """نسخ صف في الجدول"""
    
    try:
        data = request.get_json()
        source_row_id = data.get('source_row_id')
        new_job_title = data.get('new_job_title')
        
        if not source_row_id or not new_job_title:
            return jsonify({'success': False, 'message': 'بيانات غير كاملة'})
        
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return jsonify({'success': False, 'message': 'الجدول غير موجود'})
        
        # نسخ الصف
        new_row_id = schedule.copy_row(db_session, source_row_id, new_job_title)
        
        if new_row_id:
            db_session.commit()
            return jsonify({
                'success': True,
                'message': 'تم نسخ الصف بنجاح',
                'new_row_id': new_row_id
            })
        else:
            return jsonify({'success': False, 'message': 'فشل في نسخ الصف'})
        
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})


@app.route('/admin/export_schedule/<int:schedule_id>')
@login_required
def export_schedule(schedule_id):
    """تصدير الجدول إلى Excel"""
    
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            flash('الجدول غير موجود', 'error')
            return redirect(url_for('admin_weekly_schedules'))
        
        # الحصول على التفاصيل
        details = db_session.query(ScheduleDetail).filter_by(
            weekly_schedule_id=schedule_id
        ).order_by('day_date', 'row_order').all()
        
        # إنشاء DataFrame
        data = []
        for detail in details:
            data.append({
                'اليوم': detail.day_name,
                'التاريخ': detail.day_date.strftime('%Y-%m-%d'),
                'الوظيفة': detail.job_title,
                'الشيفت الصباحي': detail.morning_shift or '',
                'الشيفت المسائي': detail.evening_shift or '',
                'شيفت السهر': detail.night_shift or '',
                'ملاحظات': detail.notes or ''
            })
        
        df = pd.DataFrame(data)
        
        # إنشاء ملف Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='الجدول الأسبوعي', index=False)
            
            # تحسين التنسيق
            worksheet = writer.sheets['الجدول الأسبوعي']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # اسم الملف
        dept_name = schedule.department.name if schedule.department else 'جدول'
        filename = f"{dept_name}_الجدول_الأسبوعي_{schedule.week_start_date.strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'خطأ في التصدير: {str(e)}', 'error')
        return redirect(url_for('admin_weekly_schedule_detail', schedule_id=schedule_id))


# ======== Manager Schedule Routes ========


@app.route('/manager/schedule/<int:schedule_id>')
@login_required
def manager_schedule_detail(schedule_id):
    """تفاصيل الجدول للمدير"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('manager_schedules'))
    
    # التحقق من أن المدير مسؤول عن القسم
    
    # تحميل التفاصيل
    schedule_details = db_session.query(ScheduleDetail).filter_by(
        weekly_schedule_id=schedule_id
    ).order_by('day_date', 'row_order').all()
    
    # تحميل البيانات المرتبطة
    schedule.department = db_session.query(Department).get(schedule.department_id)
    schedule.creator = db_session.query(User).get(schedule.created_by)
    
    # تجميع التفاصيل حسب اليوم
    details_by_day = {}
    for detail in schedule_details:
        day_key = detail.day_date.strftime('%Y-%m-%d')
        if day_key not in details_by_day:
            details_by_day[day_key] = {
                'date': detail.day_date,
                'day_name': detail.day_name,
                'details': []
            }
        details_by_day[day_key]['details'].append(detail)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager/manager_schedule.html',
                         schedule=schedule,
                         details_by_day=details_by_day,
                         today=date.today(),
                         notifications=notifications)




@app.route('/manager/schedules')
@login_required
def manager_schedules():
    """عرض الجداول الأسبوعية مع الفلترة المتقدمة - النسخة المحسنة"""
    
    try:
        # ====== 1. الحصول على الأقسام التي يديرها المستخدم ======
        department_managers = db_session.query(DepartmentManager).filter_by(
            user_id=current_user.id
        ).all()

        
        # الحصول على أقسام المدير
        department_ids = [dm.department_id for dm in department_managers]
        departments = db_session.query(Department).filter(
            Department.id.in_(department_ids)
        ).all()
        
        # ====== 2. تحديد القسم المحدد (الاسم الصحيح هو 'dept_id') ======
        selected_dept_id = request.args.get('dept_id', '')
        
        print(f"DEBUG: Received dept_id parameter: '{selected_dept_id}'")
        print(f"DEBUG: Department IDs managed: {department_ids}")
        print(f"DEBUG: Available departments: {[d.name for d in departments]}")
        
        # تحويل selected_dept_id إلى int إذا كان رقم
        if selected_dept_id and selected_dept_id.isdigit():
            selected_dept_id = int(selected_dept_id)
            print(f"DEBUG: Converted dept_id to int: {selected_dept_id}")
            
            # التحقق أن القسم المحدد من ضمن أقسام المدير
            if selected_dept_id not in department_ids:
                print(f"DEBUG: {selected_dept_id} not in department_ids, using first department")
                selected_dept_id = department_ids[0] if department_ids else None
        else:
            # إذا لم يتم تحديد قسم، استخدم القسم الأول
            selected_dept_id = department_ids[0] if department_ids else None
            print(f"DEBUG: No department selected, using first: {selected_dept_id}")
        
        # الحصول على كائن القسم المحدد
        selected_department = None
        if selected_dept_id:
            selected_department = db_session.get(Department, selected_dept_id)
            print(f"DEBUG: Selected department: {selected_department.name if selected_department else 'None'}")
        
        # ====== 3. الحصول على معاملات الفلترة الأخرى ======
        selected_status = request.args.get('status', '')
        selected_month = request.args.get('month', '')
        selected_year = request.args.get('year', '')
        
        print(f"DEBUG: Filters - status: {selected_status}, month: {selected_month}, year: {selected_year}")
        
        # ====== 4. بناء الاستعلام الأساسي ======
        query = db_session.query(WeeklySchedule)
        
        # تطبيق فلتر القسم (هام جداً)
        if selected_department:
            query = query.filter_by(department_id=selected_department.id)
            print(f"DEBUG: Filtering by department: {selected_department.name} (ID: {selected_department.id})")
        else:
            # إذا لم يتم تحديد قسم، استخدام جميع أقسام المدير
            query = query.filter(WeeklySchedule.department_id.in_(department_ids))
            print(f"DEBUG: Filtering by all manager departments: {department_ids}")
        
        # تطبيق فلتر الحالة
        if selected_status:
            if selected_status == 'approved':
                query = query.filter_by(is_approved=True)
            elif selected_status == 'pending':
                query = query.filter_by(is_approved=False, is_locked=False)
            elif selected_status == 'draft':
                query = query.filter_by(is_approved=False)
            elif selected_status == 'locked':
                query = query.filter_by(is_locked=True)
        
        # تطبيق فلتر الشهر
        if selected_month and selected_month.isdigit():
            month_value = int(selected_month)
            if 1 <= month_value <= 12:
                query = query.filter_by(month=month_value)
        
        # تطبيق فلتر السنة
        if selected_year and selected_year.isdigit():
            year_value = int(selected_year)
            if 2000 <= year_value <= 2100:
                query = query.filter_by(year=year_value)
        
        # ====== 5. الحصول على الجداول ======
        schedules = query.order_by(
            WeeklySchedule.year.desc(),
            WeeklySchedule.month.desc(),
            WeeklySchedule.week_number.desc()
        ).all()
        
        print(f"DEBUG: Found {len(schedules)} schedules for department {selected_dept_id}")
        if schedules:
            print(f"DEBUG: First schedule department ID: {schedules[0].department_id}")
            print(f"DEBUG: First schedule dates: {schedules[0].week_start_date} to {schedules[0].week_end_date}")
        
        # تحميل البيانات المرتبطة
        for schedule in schedules:
            schedule.department = db_session.get(Department, schedule.department_id)
            schedule.creator = db_session.get(User, schedule.created_by)
            if schedule.approved_by:
                schedule.approver = db_session.get(User, schedule.approved_by)
            else:
                schedule.approver = None
        
        # ====== 6. تجميع الجداول حسب الفترات ======
        schedules_by_period = {}
        today = date.today()
        
        for schedule in schedules:
            # معالجة قيم None في year و month
            year_value = schedule.year or today.year
            month_value = schedule.month or today.month
            
            # التأكد من أن القيم ضمن المدى الصحيح
            if not (1 <= month_value <= 12):
                month_value = today.month
            if not (2000 <= year_value <= 2100):
                year_value = today.year
            
            # إنشاء مفتاح الفترة
            period_key = f"{year_value}-{month_value:02d}"
            
            # الحصول على اسم الشهر العربي
            arabic_month = get_month_name_arabic(month_value) if month_value else "غير محدد"
            display_name = f"{arabic_month} {year_value}"
            
            # إضافة الجدول إلى الفترة المناسبة
            if period_key not in schedules_by_period:
                schedules_by_period[period_key] = {
                    'display_name': display_name,
                    'year': year_value,
                    'month': month_value,
                    'schedules': []
                }
            
            schedules_by_period[period_key]['schedules'].append(schedule)
        
        # فرز الفترات من الأحدث إلى الأقدم
        sorted_periods = dict(sorted(schedules_by_period.items(), reverse=True))
        
        # ====== 7. الحصول على السنوات المتاحة للفلتر ======
        # استخدام نفس فلتر القسم للسنوات المتاحة
        year_query = db_session.query(
            func.distinct(WeeklySchedule.year)
        )
        
        if selected_department:
            year_query = year_query.filter_by(department_id=selected_department.id)
        else:
            year_query = year_query.filter(WeeklySchedule.department_id.in_(department_ids))
        
        year_query = year_query.order_by(WeeklySchedule.year.desc()).all()
        
        # تصفية القيم الفارغة والغير صالحة
        available_years = []
        for year_tuple in year_query:
            if year_tuple and year_tuple[0]:
                year_value = year_tuple[0]
                if 2000 <= year_value <= 2100:
                    available_years.append(year_value)
        
        print(f"DEBUG: Available years for department {selected_dept_id}: {available_years}")
        
        # ====== 8. حساب الأسبوع الحالي ======
        current_week_start = get_week_start_date(today)
        current_week_end = current_week_start + timedelta(days=6)
        
        # ====== 9. الحصول على الإشعارات ======
        notifications = get_user_notifications(current_user.id)
        
        # ====== 10. إحصائيات القسم ======
        stats = get_manager_navigation_stats(current_user.id)
        
        # ====== 11. شيتات المرتب الجديدة ======
        new_salary_slips = db_session.query(SalarySlip).filter_by(
            user_id=current_user.id, is_viewed=False
        ).count()
        
        # ====== 12. إضافة بيانات القسم الحالي للحصول على الموظفين ======
        if selected_department:
            # الحصول على عدد موظفي القسم
            employee_count = db_session.query(User).filter(
                User.department_id == selected_department.id,
                User.is_admin == False,
                User.is_active == True
            ).count()
            
            selected_department.employees_count = employee_count
            
            # الحصول على عدد الجداول النشطة
            active_schedules = db_session.query(WeeklySchedule).filter(
                WeeklySchedule.department_id == selected_department.id,
                WeeklySchedule.is_approved == True,
                WeeklySchedule.week_start_date <= today,
                WeeklySchedule.week_end_date >= today
            ).count()
            
            selected_department.active_schedules = active_schedules
            
            # الحصول على عدد طلبات الإجازة المعلقة
            pending_leaves = db_session.query(LeaveRequest).filter(
                LeaveRequest.department_id == selected_department.id,
                LeaveRequest.status == 'pending'
            ).count()
            
            selected_department.pending_leaves = pending_leaves
        
        return render_template('manager/manager_schedules.html',
                             departments=departments,  # جميع الأقسام
                             selected_department=selected_department,  # القسم المحدد
                             schedules=schedules,
                             schedules_by_period=sorted_periods,
                             today=today,
                             current_week_start=current_week_start,
                             current_week_end=current_week_end,
                             selected_status=selected_status,
                             selected_month=selected_month,
                             selected_year=selected_year,
                             available_years=available_years,
                             get_month_name_arabic=get_month_name_arabic,
                             notifications=notifications,
                             stats=stats,
                             new_salary_slips=new_salary_slips,
                             current_user=current_user)
                             
    except Exception as e:
        # تسجيل الخطأ للتصحيح
        print(f"❌ خطأ في manager_schedules: {str(e)}")
        import traceback
        traceback.print_exc()
        
        flash(f'حدث خطأ في تحميل الجداول: {str(e)}', 'error')
        
        return render_template('manager/manager_schedules.html',
                             departments=[],
                             selected_department=None,
                             schedules=[],
                             schedules_by_period={},
                             today=date.today(),
                             current_week_start=None,
                             current_week_end=None,
                             selected_status='',
                             selected_month='',
                             selected_year='',
                             available_years=[],
                             notifications=[],
                             stats={},
                             new_salary_slips=0,
                             current_user=current_user)

import sys
import traceback

# Add this to your app.py before your routes
import json
from datetime import datetime, date
from decimal import Decimal
import numpy as np
import pandas as pd

class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        elif isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, (np.integer, np.int64, np.int32, np.int16, np.int8)):
            return int(obj)
        elif isinstance(obj, (np.floating, np.float64, np.float32, np.float16)):
            return float(obj)
        elif isinstance(obj, (np.ndarray,)):
            return obj.tolist()
        elif isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        elif isinstance(obj, pd.Series):
            return obj.to_dict()
        elif pd.isna(obj):
            return None
        elif hasattr(obj, 'item'):  # numpy/pandas scalar
            try:
                return obj.item()
            except:
                return str(obj)
        return super().default(obj)

# Set the custom encoder
app.json_encoder = CustomJSONEncoder



def get_manager_navigation_stats(user_id):
    """الحصول على إحصائيات المدير لجميع الأقسام"""
    try:
        # الحصول على جميع أقسام المدير
        department_ids = get_manager_department_ids(user_id)
        
        if not department_ids:
            return {
                'pending_leaves': 0,
                'pending_permissions': 0,
                'pending_advances': 0
            }
        
        # حساب الإحصائيات لجميع الأقسام
        pending_leaves = db_session.query(LeaveRequest).filter(
            LeaveRequest.department_id.in_(department_ids),
            LeaveRequest.status == 'pending'
        ).count()
        
        pending_permissions = db_session.query(PermissionRequest).filter(
            PermissionRequest.department_id.in_(department_ids),
            PermissionRequest.status == 'pending'
        ).count()
        
        pending_advances = db_session.query(AdvanceRequest).filter(
            AdvanceRequest.department_id.in_(department_ids),
            AdvanceRequest.status == 'pending'
        ).count()
        
        return {
            'pending_leaves': pending_leaves,
            'pending_permissions': pending_permissions,
            'pending_advances': pending_advances
        }
        
    except Exception as e:
        print(f"Error in get_manager_navigation_stats: {str(e)}")
        return {
            'pending_leaves': 0,
            'pending_permissions': 0,
            'pending_advances': 0
        }

def get_manager_department_ids(user_id):
    """الحصول على IDs جميع الأقسام التي يديرها المستخدم"""
    try:
        dept_managers = db_session.query(DepartmentManager).filter_by(
            user_id=user_id
        ).all()
        return [dm.department_id for dm in dept_managers]
    except:
        return []


@app.errorhandler(Exception)
def handle_error(e):
    # تسجيل الخطأ
    print(f"Error: {e}", file=sys.stderr)
    print(traceback.format_exc(), file=sys.stderr)
    
    # محاولة تحديد إذا كان خطأ تنسيق
    if "unsupported format string passed to NoneType.__format__" in str(e):
        print("DEBUG: Format error on None value detected")
        # أضف مزيد من المعلومات للتصحيح
        import inspect
        for frame_info in inspect.stack():
            if 'app.py' in frame_info.filename:
                print(f"  File: {frame_info.filename}, Line: {frame_info.lineno}")
    
    return "An error occurred", 500

@app.route('/api/schedule/<int:schedule_id>/submit_for_approval', methods=['POST'])
@login_required
def submit_schedule_for_approval(schedule_id):
    """تقديم الجدول للاعتماد"""
    
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return jsonify({'success': False, 'message': 'الجدول غير موجود'})
        
        # التحقق من أن المدير مسؤول عن القسم
        department = db_session.query(Department).get(schedule.department_id)
        if not department:
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
        
        
        # التحقق من أن الجدول غير مقفل
        if schedule.is_locked:
            return jsonify({'success': False, 'message': 'الجدول مقفل ولا يمكن التعديل'})
        
        # تحديث حالة الجدول
        schedule.status = 'pending'
        schedule.updated_at = datetime.now()
        schedule.updated_by = current_user.id
        schedule.is_locked = True
        schedule.is_approved = True
        schedule.status = 'approved'
        schedule.approved_by = current_user.id
        schedule.approved_at = datetime.now()
        
        # إنشاء سجل في تاريخ الاعتماد
        approval_history = ScheduleApprovalHistory(
            schedule_id=schedule_id,
            action='submitted',
            comments='تم إرسال الجدول للاعتماد',
            performed_by=current_user.id
        )
        db_session.add(approval_history)
        
        # إرسال إشعار للمسؤولين
        admins = db_session.query(User).filter_by(is_admin=True).all()
        for admin in admins:
            create_notification(
                admin.id,
                'جدول جديد يحتاج اعتماد',
                f'قام المدير {current_user.name} بإرسال جدول قسم {department.name} للاعتماد',
                'schedule_pending',
                related_id=schedule_id,
                action_url=url_for('admin_weekly_schedule_detail', schedule_id=schedule_id)
            )
        
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم إرسال الجدول للاعتماد بنجاح',
            'is_locked': True,
            'redirect_url': url_for('manager_schedules')
        })
        
    except Exception as e:
        db_session.rollback()
        print(f"Error submitting schedule: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})
    

def get_week_start_date(date_obj):
    """الحصول على تاريخ بداية الأسبوع (السبت)"""
    # weekday() returns: Monday=0, Tuesday=1, ..., Sunday=6
    # نحتاج: Saturday=0, Sunday=1, ..., Friday=6
    days_since_saturday = (date_obj.weekday() + 2) % 7
    return date_obj - timedelta(days=days_since_saturday)


@app.route('/manager/schedule/submit/<int:schedule_id>', methods=['POST'])
@login_required
def manager_submit_schedule(schedule_id):
    """إرسال الجدول للاعتماد"""
    
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return jsonify({'success': False, 'message': 'الجدول غير موجود'})
        
        # التحقق من أن المدير مسؤول عن القسم
        department = db_session.query(Department).get(schedule.department_id)
        
        # تحديث حالة الجدول
        schedule.status = 'pending'
        schedule.updated_at = datetime.now()
        schedule.updated_by = current_user.id
        
        # إنشاء سجل في تاريخ الاعتماد
        approval_history = ScheduleApprovalHistory(
            schedule_id=schedule_id,
            action='submitted',
            comments='تم إرسال الجدول للاعتماد',
            performed_by=current_user.id
        )
        db_session.add(approval_history)
        
        # إرسال إشعار للمسؤولين
        admins = db_session.query(User).filter_by(is_admin=True).all()
        for admin in admins:
            create_notification(
                admin.id,
                'جدول جديد يحتاج اعتماد',
                f'قام المدير {current_user.name} بإرسال جدول قسم {department.name} للاعتماد',
                'schedule_pending',
                related_id=schedule_id,
                action_url=url_for('admin_weekly_schedule_detail', schedule_id=schedule_id)
            )
        
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم إرسال الجدول للاعتماد بنجاح'
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })


@app.route('/manager/create_future_schedules', methods=['POST'])
@login_required
def manager_create_future_schedules():
    """إنشاء جداول مستقبلية بناءً على آخر جدول لكل قسم يديره المدير"""
    
    try:
        # الحصول على جميع الأقسام التي يديرها المدير
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
    
        
        total_created = 0
        department_details = []
        
        for department in managed_departments:
            # الحصول على آخر جدول لكل قسم
            last_schedule = db_session.query(WeeklySchedule).filter_by(
                department_id=department.id
            ).order_by(
                WeeklySchedule.week_start_date.desc()
            ).first()
            
            if not last_schedule:
                department_details.append({
                    'name': department.name,
                    'status': 'لا يوجد جدول سابق للنسخ منه',
                    'created': 0
                })
                continue
            
            created_count = 0
            
            # إنشاء جداول للأسابيع القادمة (52 أسبوع = سنة كاملة)
            for i in range(1, 53):  # 52 أسبوع
                # حساب تاريخ بداية الأسبوع
                next_week_start = last_schedule.week_start_date + timedelta(days=(i * 7))
                next_week_end = next_week_start + timedelta(days=6)
                
                # التحقق من عدم وجود جدول لهذا الأسبوع
                existing_schedule = db_session.query(WeeklySchedule).filter_by(
                    department_id=department.id,
                    week_start_date=next_week_start
                ).first()
                
                if existing_schedule:
                    continue
                
                # إنشاء جدول جديد
                week_number = next_week_start.isocalendar()[1]
                month = next_week_start.month
                year = next_week_start.year
                
                new_schedule = WeeklySchedule(
                    department_id=department.id,
                    week_start_date=next_week_start,
                    week_end_date=next_week_end,
                    week_number=week_number,
                    month=month,
                    year=year,
                    created_by=current_user.id,
                    is_approved=False,
                    is_locked=False,
                    is_template=False
                )
                
                db_session.add(new_schedule)
                db_session.flush()  # للحصول على ID
                
                # نسخ تفاصيل آخر جدول
                last_details = db_session.query(ScheduleDetail).filter_by(
                    weekly_schedule_id=last_schedule.id
                ).all()
                
                for detail in last_details:
                    new_detail = ScheduleDetail(
                        weekly_schedule_id=new_schedule.id,
                        day_date=next_week_start + (detail.day_date - last_schedule.week_start_date),
                        day_name=WeeklySchedule.get_arabic_day_name(
                            next_week_start + (detail.day_date - last_schedule.week_start_date)
                        ),
                        job_title=detail.job_title,
                        morning_shift=detail.morning_shift,
                        evening_shift=detail.evening_shift,
                        night_shift=detail.night_shift,
                        row_order=detail.row_order,
                        is_custom=detail.is_custom,
                        notes=detail.notes
                    )
                    db_session.add(new_detail)
                
                created_count += 1
                total_created += 1
            
            department_details.append({
                'name': department.name,
                'status': 'تم إنشاء جدول جديد' if created_count > 0 else 'لا يوجد حاجة لإنشاء جداول جديدة',
                'created': created_count
            })
        
        db_session.commit()
        
        # إنشاء رسالة تفصيلية
        flash_message = f'تم إنشاء {total_created} جدول جديد في {len(managed_departments)} قسم<br>'
        for detail in department_details:
            flash_message += f'• {detail["name"]}: {detail["created"]} جدول ({detail["status"]})<br>'
        
        flash(flash_message, 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء إنشاء الجداول: {str(e)}', 'error')
    
    return redirect(url_for('manager_schedules'))


@app.route('/view_schedule/<int:schedule_id>')
@login_required
def view_schedule(schedule_id):
    """عرض تفاصيل الجدول"""
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('manager_schedules' if current_user.is_manager else 'user_schedule'))
    
    # التحقق من الصلاحيات
    if current_user.is_manager:
        department = db_session.query(Department).get(schedule.department_id)
    
    # تحميل التفاصيل
    schedule_details = db_session.query(ScheduleDetail).filter_by(
        weekly_schedule_id=schedule_id
    ).order_by('day_date', 'row_order').all()
    
    # تحميل البيانات المرتبطة
    schedule.department = db_session.query(Department).get(schedule.department_id)
    schedule.creator = db_session.query(User).get(schedule.created_by)
    if schedule.approved_by:
        schedule.approver = db_session.query(User).get(schedule.approved_by)
    
    # تجميع التفاصيل حسب اليوم
    details_by_day = {}
    for detail in schedule_details:
        day_key = detail.day_date.strftime('%Y-%m-%d')
        if day_key not in details_by_day:
            details_by_day[day_key] = {
                'date': detail.day_date,
                'day_name': detail.day_name,
                'details': []
            }
        details_by_day[day_key]['details'].append(detail)
    
    notifications = get_user_notifications(current_user.id)
    
    template_name = 'manager/manager_schedule.html' if current_user.is_manager else 'user/user_schedule.html'
    
    return render_template(template_name,
                         schedule=schedule,
                         today=date.today(),
                         details_by_day=details_by_day,
                         notifications=notifications)


@app.route('/user/schedule')
@login_required
def user_schedule():
    """عرض الجدول للموظف - للموظفين العاديين فقط"""
    
    
    # الحصول على التاريخ الحالي
    today = date.today()
    
    print(f"DEBUG - User Schedule Page Accessed")
    print(f"DEBUG - User ID: {current_user.id}, Name: {current_user.name}")
    print(f"DEBUG - User Department ID: {current_user.department_id}")
    print(f"DEBUG - Today: {today}")
    
    # الحصول على نطاق الشهر المالي أولاً
    financial_month_start, financial_month_end = get_financial_month_range()
    
    print(f"DEBUG - Financial Month Range: {financial_month_start} to {financial_month_end}")
    
    # استخدام نطاق الشهر المالي كبداية ونهاية للاستعلام
    financial_month = financial_month_start.month
    financial_year = financial_month_start.year
    
    print(f"DEBUG - Financial Month/Year: {financial_month}/{financial_year}")
    
    # الحصول على الجداول المعتمدة للقسم ضمن نطاق الشهر المالي فقط
    approved_schedules = db_session.query(WeeklySchedule).filter(
        WeeklySchedule.department_id == current_user.department_id,
        WeeklySchedule.week_start_date >= financial_month_start,
        WeeklySchedule.week_end_date <= financial_month_end,
    ).order_by(WeeklySchedule.week_start_date).all()  # ترتيب حسب تاريخ بداية الأسبوع
    
    print(f"DEBUG - Found {len(approved_schedules)} approved schedules for financial month")
    
    schedule_data_list = []
    
    # تحديد الأسبوع الحالي
    current_week_schedule = None
    
    for schedule in approved_schedules:
        print(f"DEBUG - Processing schedule: Week {schedule.week_number}, "
              f"Start: {schedule.week_start_date}, End: {schedule.week_end_date}")
        
        # تحميل تفاصيل الجدول
        schedule_details = db_session.query(ScheduleDetail).filter_by(
            weekly_schedule_id=schedule.id
        ).order_by('day_date', 'row_order').all()
        
        # تحويل التفاصيل إلى تنسيق مناسب للقالب
        days_data = []
        for detail in schedule_details:
            days_data.append({
                'day': detail.day_name or '',
                'date': detail.day_date.strftime('%Y-%m-%d') if detail.day_date else '',
                'morning_shift': detail.morning_shift or '',
                'evening_shift': detail.evening_shift or '',
                'night_shift': detail.night_shift or '',
                'job': detail.job_title or ''
            })
        
        # التحقق إذا كان هذا هو الأسبوع الحالي
        is_current_week = (schedule.week_start_date <= today <= schedule.week_end_date)
        
        # تحديد جدول الأسبوع الحالي
        if is_current_week:
            current_week_schedule = schedule
            print(f"DEBUG - Current week found: Week {schedule.week_number}")
        
        schedule_data_list.append({
            'schedule': schedule,
            'data': {
                'schedule': days_data
            },
            'is_current_week': is_current_week,
            'month_name': get_month_name_arabic(schedule.month),
            'year': schedule.year,
            'week_number': schedule.week_number
        })
    
    # تسجيل مشاهدة الجدول الحالي إن وجد
    if current_week_schedule:
        print(f"DEBUG - Current week schedule ID: {current_week_schedule.id}")
        if current_user not in current_week_schedule.viewed_by_users:
            current_week_schedule.viewed_by_users.append(current_user)
            db_session.commit()
            print(f"DEBUG - User marked as viewed schedule {current_week_schedule.id}")
    
    # معلومات الشهر المالي للقالب
    month_name = get_month_name_arabic(financial_month)
    year = financial_year
    
    # الحصول على الإشعارات
    notifications = get_user_notifications(current_user.id)
    
    # شيتات المرتب الجديدة للقالب
    new_salary_slips = db_session.query(SalarySlip).filter_by(
        user_id=current_user.id, is_viewed=False
    ).count()
    
    # إحصاءات
    stats = {
        'total_schedules': len(approved_schedules),
        'current_week_schedule': current_week_schedule is not None,
        'financial_month_start': financial_month_start,
        'financial_month_end': financial_month_end
    }
    
    return render_template('user_schedule.html',
                         schedule_data_list=schedule_data_list,
                         financial_month_start=financial_month_start,
                         financial_month_end=financial_month_end,
                         month_name=month_name,
                         year=year,
                         today=today,
                         notifications=notifications,
                         new_salary_slips=new_salary_slips,
                         stats=stats)



@app.route('/api/user_schedule_details')
@login_required
def api_user_schedule_details():
    """API للحصول على تفاصيل الجدول للموظف"""
    
    try:
        today = date.today()
        
        # البحث عن جدول اليوم
        schedule = db_session.query(WeeklySchedule).filter(
            WeeklySchedule.department_id == current_user.department_id,
            WeeklySchedule.is_approved == True,
            WeeklySchedule.week_start_date <= today,
            WeeklySchedule.week_end_date >= today
        ).first()
        
        if not schedule:
            return jsonify({
                'success': False,
                'message': 'لا يوجد جدول معتمد لهذا الأسبوع'
            })
        
        # الحصول على تفاصيل اليوم الحالي
        details = db_session.query(ScheduleDetail).filter_by(
            weekly_schedule_id=schedule.id,
            day_date=today
        ).order_by('row_order').all()
        
        # البحث عن تفاصيل الموظف
        user_details = []
        employee_data = db_session.query(EmployeeData).filter_by(
            user_id=current_user.id
        ).first()
        
        user_job_title = employee_data.job_title if employee_data else 'موظف'
        
        for detail in details:
            if detail.job_title == user_job_title:
                user_details.append({
                    'job_title': detail.job_title,
                    'morning_shift': detail.morning_shift,
                    'evening_shift': detail.evening_shift,
                    'night_shift': detail.night_shift,
                    'notes': detail.notes or ''
                })
        
        return jsonify({
            'success': True,
            'schedule': {
                'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
                'week_end_date': schedule.week_end_date.strftime('%Y-%m-%d'),
                'week_number': schedule.week_number
            },
            'today': today.strftime('%Y-%m-%d'),
            'user_job_title': user_job_title,
            'details': user_details
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })



@app.route('/api/schedule/update/<int:schedule_id>', methods=['POST'])
@login_required
def update_schedule(schedule_id):
    """تحديث الجدول كمسودة فقط"""
    
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return jsonify({'success': False, 'message': 'الجدول غير موجود'})
        
        # التحقق من أن المدير مسؤول عن القسم
        department = db_session.query(Department).get(schedule.department_id)
        if not department:
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
    
        
        # التحقق من أن الجدول غير مقفل
        if schedule.is_locked:
            return jsonify({'success': False, 'message': 'الجدول مقفل ولا يمكن التعديل'})
        
        data = request.get_json()
        
        print(f"Received data for schedule {schedule_id}: {data}")
        
        updated_count = 0
        new_rows_count = 0
        
        # Handle existing rows
        if 'existing_rows' in data:
            print(f"Processing {len(data['existing_rows'])} existing rows")
            for row_data in data['existing_rows']:
                detail_id = row_data.get('detail_id')
                if detail_id:
                    detail = db_session.query(ScheduleDetail).get(detail_id)
                    if detail and detail.weekly_schedule_id == schedule.id:
                        detail.morning_shift = row_data.get('morning_shift', '').strip()
                        detail.evening_shift = row_data.get('evening_shift', '').strip()
                        detail.night_shift = row_data.get('night_shift', '').strip()
                        
                        # Update job title if provided
                        if 'job_title' in row_data and row_data['job_title']:
                            detail.job_title = row_data['job_title'].strip()
                        
                        detail.modified_by = current_user.id
                        detail.modified_at = datetime.now()
                        updated_count += 1
        
        # Handle new rows
        if 'new_rows' in data:
            print(f"Processing {len(data['new_rows'])} new rows")
            for row_data in data['new_rows']:
                job_title = row_data.get('job_title', '').strip()
                date_str = row_data.get('date', '')
                day_name = row_data.get('day', '')
                
                print(f"New row data: job={job_title}, date={date_str}, day={day_name}")
                
                if job_title and date_str and day_name:
                    try:
                        day_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        
                        # Verify date is within schedule range
                        if schedule.week_start_date <= day_date <= schedule.week_end_date:
                            # Get max row order for this day
                            max_order = db_session.query(func.max(ScheduleDetail.row_order)).filter_by(
                                weekly_schedule_id=schedule.id,
                                day_date=day_date
                            ).scalar() or 0
                            
                            # Create new detail
                            new_detail = ScheduleDetail(
                                weekly_schedule_id=schedule.id,
                                day_date=day_date,
                                day_name=day_name,
                                job_title=job_title,
                                morning_shift=row_data.get('morning_shift', '').strip(),
                                evening_shift=row_data.get('evening_shift', '').strip(),
                                night_shift=row_data.get('night_shift', '').strip(),
                                row_order=max_order + 1,
                                is_custom=True
                            )
                            
                            db_session.add(new_detail)
                            new_rows_count += 1
                            print(f"✓ Added new row: {job_title} on {date_str}")
                            
                    except Exception as e:
                        print(f"Error creating new row: {str(e)}")
                        import traceback
                        traceback.print_exc()
                        continue
                else:
                    print(f"Missing required fields in row: {row_data}")
        
        # Update schedule timestamps (مسودة فقط)
        schedule.updated_at = datetime.now()
        schedule.updated_by = current_user.id
        
        db_session.commit()
        
        print(f"Successfully updated: {updated_count} rows, created: {new_rows_count} rows")
        
        return jsonify({
            'success': True,
            'message': f'تم حفظ المسودة ({updated_count} صف محدث، {new_rows_count} صف جديد)',
            'updated_count': updated_count,
            'new_rows_count': new_rows_count,
            'has_new_rows': new_rows_count > 0
        })
        
    except Exception as e:
        db_session.rollback()
        print(f"Error updating schedule: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})
    


@app.route('/admin/approve_schedule_final/<int:schedule_id>')
@login_required
def approve_schedule_final(schedule_id):
    """اعتماد الجدول النهائي من قبل الإدمن"""
    
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('admin_weekly_schedules'))
    
    if schedule.is_approved:
        flash('الجدول معتمد already', 'warning')
        return redirect(url_for('admin_weekly_schedule_detail', schedule_id=schedule_id))
    
    schedule.is_approved = True
    schedule.status = 'approved'
    schedule.approved_by = current_user.id
    schedule.approved_at = datetime.now()
    schedule.is_locked = True  # التأكد من أنه مقفل
    
    # إنشاء سجل في تاريخ الاعتماد
    approval_history = ScheduleApprovalHistory(
        schedule_id=schedule_id,
        action='approved',
        comments='تم اعتماد الجدول نهائياً',
        performed_by=current_user.id
    )
    db_session.add(approval_history)
    
    # إرسال إشعار للمدير
    department = db_session.query(Department).get(schedule.department_id)
    if department and department.primary_manager_id:
        create_notification(
            department.primary_manager_id,
            'تم اعتماد الجدول',
            f'تم اعتماد جدول قسم {department.name} للفترة {schedule.week_start_date.strftime("%Y-%m-%d")}',
            'schedule_approved',
            related_id=schedule_id,
            action_url=url_for('view_schedule', schedule_id=schedule_id)
        )
    
    db_session.commit()
    
    flash('تم اعتماد الجدول بنجاح', 'success')
    return redirect(url_for('admin_weekly_schedule_detail', schedule_id=schedule_id))


def get_arabic_day_name(date_obj):
    """الحصول على اسم اليوم بالعربية بشكل صحيح"""
    # weekday() returns: Monday=0, Tuesday=1, Wednesday=2, Thursday=3, Friday=4, Saturday=5, Sunday=6
    
    arabic_days = {
        0: 'الإثنين',    # Monday
        1: 'الثلاثاء',   # Tuesday
        2: 'الأربعاء',   # Wednesday
        3: 'الخميس',     # Thursday
        4: 'الجمعة',     # Friday
        5: 'السبت',      # Saturday
        6: 'الأحد'       # Sunday
    }
    
    return arabic_days.get(date_obj.weekday(), 'غير معروف')

def get_arabic_day_name_english(date_obj):
    """الحصول على اسم اليوم بالعربية من تاريخ إنجليزي"""
    # English weekday to Arabic mapping
    english_to_arabic = {
        0: 'الإثنين',    # Monday
        1: 'الثلاثاء',   # Tuesday
        2: 'الأربعاء',   # Wednesday
        3: 'الخميس',     # Thursday
        4: 'الجمعة',     # Friday
        5: 'السبت',      # Saturday
        6: 'الأحد'       # Sunday
    }
    
    # date_obj.weekday() returns: Monday=0, Sunday=6
    return english_to_arabic.get(date_obj.weekday(), 'غير معروف')

def get_week_dates_with_correct_days(start_date):
    """الحصول على جميع أيام الأسبوع مع الأسماء الصحيحة"""
    week_dates = []
    
    # Saturday is the start of the week
    for i in range(7):
        current_date = start_date + timedelta(days=i)
        day_name = get_arabic_day_name(current_date)
        
        week_dates.append({
            'date': current_date,
            'day_name': day_name,
            'date_str': current_date.strftime('%Y-%m-%d'),
            'day_index': i,
            'weekday': current_date.weekday()  # Add weekday for debugging
        })
    
    # Log for debugging
    print(f"Week starting from {start_date}:")
    for wd in week_dates:
        print(f"  {wd['date_str']}: {wd['day_name']} (weekday: {wd['weekday']})")
    
    return week_dates

@app.route('/api/schedule/<int:schedule_id>/delete_detail/<int:detail_id>', methods=['DELETE'])
@login_required
def delete_schedule_detail(schedule_id, detail_id):
    """حذف صف من الجدول"""
    
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return jsonify({'success': False, 'message': 'الجدول غير موجود'})
        
        # التحقق من أن المدير مسؤول عن القسم
        department = db_session.query(Department).get(schedule.department_id)
        
        # التحقق من أن الجدول غير مقفل
        if schedule.is_locked:
            return jsonify({'success': False, 'message': 'الجدول مقفل ولا يمكن التعديل'})
        
        detail = db_session.query(ScheduleDetail).get(detail_id)
        if not detail or detail.weekly_schedule_id != schedule_id:
            return jsonify({'success': False, 'message': 'الصف غير موجود'})
        
        db_session.delete(detail)
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم حذف الصف بنجاح'
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})
    

@app.route('/api/schedule/<int:schedule_id>/add_row', methods=['POST'])
@login_required
def add_schedule_row(schedule_id):
    """API لإضافة صف جديد للجدول"""
    
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return jsonify({'success': False, 'message': 'الجدول غير موجود'})
        
        # التحقق من أن المدير مسؤول عن القسم
        department = db_session.query(Department).get(schedule.department_id)
        
        # التحقق من أن الجدول غير مقفل
        if schedule.is_locked:
            return jsonify({'success': False, 'message': 'الجدول مقفل ولا يمكن التعديل'})
        
        data = request.get_json()
        
        # التحقق من البيانات المطلوبة
        required_fields = ['day_date', 'job_title']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'success': False, 'message': f'حقل {field} مطلوب'})
        
        # تحويل التاريخ
        day_date = datetime.strptime(data['day_date'], '%Y-%m-%d').date()
        
        # التحقق من أن التاريخ ضمن نطاق الجدول
        if not (schedule.week_start_date <= day_date <= schedule.week_end_date):
            return jsonify({
                'success': False,
                'message': f'التاريخ خارج نطاق الجدول ({schedule.week_start_date} إلى {schedule.week_end_date})'
            })
        
        # إضافة الصف الجديد
        new_detail = schedule.add_custom_row(
            db_session=db_session,
            day_date=day_date,
            job_title=data['job_title'],
            day_name=data.get('day_name'),
            current_user_id=current_user.id
        )
        
        # حفظ أي بيانات إضافية
        if 'morning_shift' in data:
            new_detail.morning_shift = data['morning_shift']
        if 'evening_shift' in data:
            new_detail.evening_shift = data['evening_shift']
        if 'night_shift' in data:
            new_detail.night_shift = data['night_shift']
        
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم إضافة الصف بنجاح',
            'detail_id': new_detail.id,
            'detail': new_detail.to_dict()
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

# ======== Schedule Editing Routes ========

@app.route('/api/department/<int:department_id>/employees')
@login_required
def get_department_employees(department_id):
    """API للحصول على موظفي القسم"""
    
    try:
        # التحقق من أن المدير مسؤول عن القسم
        department = db_session.query(Department).get(department_id)
        if not department:
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
        
        
        # الحصول على موظفي القسم
        employees = db_session.query(User).filter(
            User.department_id == department_id,
            User.is_active == True,
            User.is_admin == False
        ).all()
        
        # تحضير بيانات الموظفين
        employees_data = []
        for employee in employees:
            # الحصول على بيانات الموظف الإضافية
            employee_data = db_session.query(EmployeeData).filter_by(
                user_id=employee.id
            ).first()
            
            employee_info = {
                'id': employee.id,
                'username': employee.username,
                'name': employee.name,
                'job_title': employee_data.job_title if employee_data else '',
                'arabic_name': employee_data.arabic_name if employee_data else '',
                'english_name': employee_data.english_name if employee_data else '',
                'display_name': employee_data.arabic_name if employee_data and employee_data.arabic_name else employee.name
            }
            employees_data.append(employee_info)
        
        return jsonify({
            'success': True,
            'department': {
                'id': department.id,
                'name': department.name,
                'manager_id': department.primary_manager_id
            },
            'employees': employees_data,
            'count': len(employees_data)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })


@app.route('/api/schedule/<int:schedule_id>/department_employees')
@login_required
def get_schedule_department_employees(schedule_id):
    """API للحصول على موظفي قسم جدول محدد"""
    
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return jsonify({'success': False, 'message': 'الجدول غير موجود'})
        
        # التحقق من الصلاحيات
        department = db_session.query(Department).get(schedule.department_id)
        
        # الحصول على موظفي القسم
        employees = db_session.query(User).filter(
            User.department_id == schedule.department_id,
            User.is_active == True,
            User.is_admin == False
        ).all()
        
        # تحضير البيانات
        employees_data = []
        for employee in employees:
            employee_data = db_session.query(EmployeeData).filter_by(
                user_id=employee.id
            ).first()
            
            employees_data.append({
                'id': employee.id,
                'username': employee.username,
                'name': employee.name,
                'job_title': employee_data.job_title if employee_data else '',
                'arabic_name': employee_data.arabic_name if employee_data else '',
                'display_name': employee_data.arabic_name if employee_data and employee_data.arabic_name else employee.name
            })
        
        return jsonify({
            'success': True,
            'schedule': {
                'id': schedule.id,
                'department_id': schedule.department_id,
                'department_name': department.name,
                'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
                'week_end_date': schedule.week_end_date.strftime('%Y-%m-%d')
            },
            'employees': employees_data,
            'count': len(employees_data)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })


@app.route('/api/employee/search')
@login_required
def search_employees():
    """API للبحث عن الموظفين"""
    
    try:
        query = request.args.get('q', '')
        department_id = request.args.get('department_id', '')
        
        if not department_id:
            # إذا لم يتم تحديد قسم، استخدام قسم المدير
            if current_user.is_manager:
                department = db_session.query(Department).filter_by(
                    primary_manager_id=current_user.id
                ).first()
                if department:
                    department_id = department.id
                else:
                    return jsonify({'success': True, 'employees': []})
        
        # بناء الاستعلام
        employees_query = db_session.query(User).filter(
            User.department_id == department_id,
            User.is_active == True,
            User.is_admin == False
        )
        
        if query:
            employees_query = employees_query.filter(
                or_(
                    User.username.ilike(f'%{query}%'),
                    User.name.ilike(f'%{query}%')
                )
            )
        
        employees = employees_query.limit(20).all()
        
        # تحضير البيانات
        employees_data = []
        for employee in employees:
            employee_data = db_session.query(EmployeeData).filter_by(
                user_id=employee.id
            ).first()
            
            display_name = employee.name
            if employee_data and employee_data.arabic_name:
                display_name = employee_data.arabic_name
            
            employees_data.append({
                'id': employee.id,
                'username': employee.username,
                'name': employee.name,
                'display_name': display_name,
                'job_title': employee_data.job_title if employee_data else '',
                'department_id': employee.department_id
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })


@app.route('/api/schedule/<int:schedule_id>/get_department_info')
@login_required
def get_schedule_department_info(schedule_id):
    """الحصول على معلومات قسم جدول معين"""
    
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return jsonify({'success': False, 'message': 'الجدول غير موجود'})
        
        department = db_session.query(Department).get(schedule.department_id)
        if not department:
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
        
        # التحقق من الصلاحيات
        
        return jsonify({
            'success': True,
            'department': {
                'id': department.id,
                'name': department.name,
                'manager_id': department.primary_manager_id,
                'manager_name': db_session.query(User).get(department.primary_manager_id).name if department.primary_manager_id else '',
                'employee_count': db_session.query(User).filter(
                    User.department_id == department.id,
                    User.is_active == True,
                    User.is_admin == False
                ).count()
            },
            'schedule': {
                'id': schedule.id,
                'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
                'week_end_date': schedule.week_end_date.strftime('%Y-%m-%d')
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })



@app.route('/manager/edit_schedule/<int:schedule_id>')
@login_required
def manager_edit_schedule(schedule_id):
    """تعديل الجدول - النسخة المحسنة مع Select2"""
    
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('manager_schedules'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(schedule.department_id)
    
    # التحقق من أن الجدول غير مقفل وغير معتمد
    if schedule.is_locked or schedule.is_approved:
        flash('لا يمكن تعديل جدول مقفل أو معتمد', 'error')
        return redirect(url_for('view_schedule', schedule_id=schedule_id))
    
    # تحميل التفاصيل
    schedule_details = db_session.query(ScheduleDetail).filter_by(
        weekly_schedule_id=schedule_id
    ).order_by('day_date', 'row_order').all()
    
    # الحصول على موظفي القسم
    department_employees = db_session.query(User).filter(
        User.department_id == department.id,
        User.is_active == True,
        User.is_admin == False
    ).all()
    
    # تحضير بيانات الموظفين
    employees_list = []
    for emp in department_employees:
        emp_data = db_session.query(EmployeeData).filter_by(user_id=emp.id).first()
        employees_list.append({
            'id': emp.id,
            'name': emp.name,
            'username': emp.username,
            'arabic_name': emp_data.arabic_name if emp_data else '',
            'job_title': emp_data.job_title if emp_data else ''
        })
    
    # تجميع التفاصيل حسب اليوم
    details_by_day = {}
    for detail in schedule_details:
        if detail.day_date:  # Make sure day_date is not None
            day_key = detail.day_date.strftime('%Y-%m-%d')
            if day_key not in details_by_day:
                details_by_day[day_key] = {
                    'date': detail.day_date,
                    'day_name': detail.day_name,
                    'details': []
                }
            details_by_day[day_key]['details'].append(detail)
    
    # Sort by date
    details_by_day = dict(sorted(details_by_day.items()))
    
    # Get all days in the week for reference
    week_dates = []
    current_date = schedule.week_start_date
    day_counter = 0
    
    while current_date <= schedule.week_end_date:
        # Use the get_arabic_day_name function for consistency
        day_name = get_arabic_day_name(current_date)
        
        week_dates.append({
            'date': current_date,
            'day_name': day_name,
            'date_str': current_date.strftime('%Y-%m-%d'),
            'day_index': day_counter
        })
        
        current_date += timedelta(days=1)
        day_counter += 1
    
    # Debug information
    print(f"Week dates from {schedule.week_start_date} to {schedule.week_end_date}:")
    for wd in week_dates:
        print(f"  {wd['date_str']}: {wd['day_name']}")
    
    unique_jobs = set()
    for detail in schedule_details:
        if detail.job_title:
            unique_jobs.add(detail.job_title)
    
    notifications = get_user_notifications(current_user.id)
    today = date.today()
    
    return render_template('manager/edit_schedule.html',
                         schedule=schedule,
                         details_by_day=details_by_day,
                         week_dates=week_dates,  # Pass week dates to template
                         department=department,
                         department_employees=employees_list,
                         unique_jobs=list(unique_jobs),
                         notifications=notifications,
                         today=today)


@property
def created_by_user(self):
    if self.created_by:
        return db_session.get(User, self.created_by)
    return None


@app.route('/api/schedule/<int:schedule_id>/employee_stats')
@login_required
def get_schedule_employee_stats(schedule_id):
    """الحصول على إحصائيات الموظفين في الجدول"""
    
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return jsonify({'success': False, 'message': 'الجدول غير موجود'})
        
        department = db_session.query(Department).get(schedule.department_id)
        
        # الحصول على جميع الموظفين المذكورين في الجدول
        details = db_session.query(ScheduleDetail).filter_by(
            weekly_schedule_id=schedule_id
        ).all()
        
        # إحصاء الموظفين في كل شيفت
        employee_counts = {
            'morning': {},
            'evening': {},
            'night': {}
        }
        
        for detail in details:
            # الشيفت الصباحي
            if detail.morning_shift:
                employee_counts['morning'][detail.morning_shift] = \
                    employee_counts['morning'].get(detail.morning_shift, 0) + 1
            
            # الشيفت المسائي
            if detail.evening_shift:
                employee_counts['evening'][detail.evening_shift] = \
                    employee_counts['evening'].get(detail.evening_shift, 0) + 1
            
            # شيفت السهر
            if detail.night_shift:
                employee_counts['night'][detail.night_shift] = \
                    employee_counts['night'].get(detail.night_shift, 0) + 1
        
        # إحصاء الوظائف
        job_counts = {}
        for detail in details:
            if detail.job_title:
                job_counts[detail.job_title] = job_counts.get(detail.job_title, 0) + 1
        
        # عدد الموظفين الفريدين
        all_employees = set()
        for shift_type, employees in employee_counts.items():
            all_employees.update(employees.keys())
        
        return jsonify({
            'success': True,
            'stats': {
                'total_employees': len(all_employees),
                'total_shifts': len(details),
                'job_count': len(job_counts),
                'employee_counts': employee_counts,
                'job_counts': job_counts,
                'department_employee_count': db_session.query(User).filter(
                    User.department_id == department.id,
                    User.is_active == True,
                    User.is_admin == False
                ).count()
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })



# ======== Manager Leave Requests Routes ========
@app.route('/manager/leave_requests')
@login_required
def manager_leave_requests():
    """طلبات الإجازات مع إحصائيات متقدمة وتقارير"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()

        
        
        department_ids = [dept.id for dept in managed_departments]
        print(department_ids)
        # Get ALL leave requests
        all_requests = db_session.query(LeaveRequest)\
            .options(
                joinedload(LeaveRequest.user),
                joinedload(LeaveRequest.department),
                joinedload(LeaveRequest.approver)
            )\
            .filter(LeaveRequest.department_id.in_(department_ids))\
            .order_by(LeaveRequest.created_at.desc())\
            .all()
        
        # Separate pending and processed requests
        pending_requests = [req for req in all_requests if req.status == 'pending']
        processed_requests = [req for req in all_requests if req.status != 'pending']
        
        # الحصول على الإحصائيات المتقدمة
        stats = get_advanced_leave_statistics(department_ids, all_requests)
        
        # إضافة الإحصائيات الأساسية المطلوبة للقالب
        stats.update({
            'pending_leaves': len(pending_requests),
            'pending_permissions': db_session.query(PermissionRequest).filter(
                PermissionRequest.department_id.in_(department_ids),
                PermissionRequest.status == 'pending'
            ).count(),
            'pending_advances': db_session.query(AdvanceRequest).filter(
                AdvanceRequest.department_id.in_(department_ids),
                AdvanceRequest.status == 'pending'
            ).count(),
            'department_employees': db_session.query(User).filter(
                User.department_id.in_(department_ids),
                User.is_admin == False
            ).count()
        })
        
        # الحصول على شيتات المرتب الجديدة
        new_salary_slips = db_session.query(SalarySlip).filter_by(
            user_id=current_user.id, is_viewed=False
        ).count()
        
        notifications = get_user_notifications(current_user.id)
        
        return render_template('manager/manager_leave_requests.html',
                            pending_requests=pending_requests,
                            processed_requests=processed_requests,
                            departments=managed_departments,
                            stats=stats,
                            notifications=notifications,
                            new_salary_slips=new_salary_slips,
                            current_user=current_user)
            
    except Exception as e:
        print(f"Error in manager_leave_requests: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'حدث خطأ في تحميل البيانات: {str(e)}', 'error')
        return render_template('manager/manager_leave_requests.html',
                            pending_requests=[],
                            processed_requests=[],
                            departments=[],
                            stats={},
                            notifications=[])


@app.context_processor
def inject_datetime():
    """Inject date and timedelta into all templates"""
    return {
        'date': date,
        'datetime': datetime,
        'timedelta': timedelta
    }



def get_advanced_leave_statistics(department_ids, all_requests):
    """الحصول على إحصائيات متقدمة للإجازات"""
    try:
        # إحصائيات أساسية
        total_leaves = len(all_requests)
        pending_leaves = len([r for r in all_requests if r.status == 'pending'])
        approved_leaves = len([r for r in all_requests if r.status == 'approved'])
        rejected_leaves = len([r for r in all_requests if r.status == 'rejected'])
        
        all_requests = db_session.query(LeaveRequest)\
            .options(
                joinedload(LeaveRequest.user),
                joinedload(LeaveRequest.department),
                joinedload(LeaveRequest.approver)
            )\
            .filter(LeaveRequest.department_id.in_(department_ids))\
            .all()
        # إحصائيات حسب النوع
        leave_types = {}
        for request in all_requests:
            leave_type = request.leave_type or 'غير محدد'
            if leave_type not in leave_types:
                leave_types[leave_type] = {'total': 0, 'approved': 0, 'pending': 0, 'rejected': 0}
            leave_types[leave_type]['total'] += 1
            if request.status == 'approved':
                leave_types[leave_type]['approved'] += 1
            elif request.status == 'pending':
                leave_types[leave_type]['pending'] += 1
            elif request.status == 'rejected':
                leave_types[leave_type]['rejected'] += 1
        
        # إحصائيات حسب الشهر
        monthly_stats = {}
        for request in all_requests:
            month_key = request.created_at.strftime('%Y-%m') if request.created_at else 'غير محدد'
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {'total': 0, 'approved': 0, 'pending': 0, 'rejected': 0}
            monthly_stats[month_key]['total'] += 1
            if request.status == 'approved':
                monthly_stats[month_key]['approved'] += 1
            elif request.status == 'pending':
                monthly_stats[month_key]['pending'] += 1
            elif request.status == 'rejected':
                monthly_stats[month_key]['rejected'] += 1
        
        # إحصائيات حسب القسم
        department_stats = {}
        departments = db_session.query(Department).filter(Department.id.in_(department_ids)).all()
        for dept in departments:
            dept_requests = [r for r in all_requests if r.department_id == dept.id]
            department_stats[dept.name] = {
                'total': len(dept_requests),
                'approved': len([r for r in dept_requests if r.status == 'approved']),
                'pending': len([r for r in dept_requests if r.status == 'pending']),
                'rejected': len([r for r in dept_requests if r.status == 'rejected']),
                'approval_rate': calculate_approval_rate(dept_requests)
            }
        
        # إحصائيات الموظفين
        employee_stats = {}
        for request in all_requests:
            if request.user:
                user_key = f"{request.user.name} ({request.user.username})"
                if user_key not in employee_stats:
                    employee_stats[user_key] = {
                        'department': request.department.name if request.department else 'غير معين',
                        'total': 0,
                        'approved': 0,
                        'pending': 0,
                        'rejected': 0,
                        'total_days': 0
                    }
                employee_stats[user_key]['total'] += 1
                employee_stats[user_key]['total_days'] += request.total_days if request.total_days else 1
                if request.status == 'approved':
                    employee_stats[user_key]['approved'] += 1
                elif request.status == 'pending':
                    employee_stats[user_key]['pending'] += 1
                elif request.status == 'rejected':
                    employee_stats[user_key]['rejected'] += 1
        
        # حساب معدل القبول الشهري
        current_month = datetime.now().strftime('%Y-%m')
        current_month_requests = [r for r in all_requests 
                                  if r.created_at and r.created_at.strftime('%Y-%m') == current_month]
        
        return {
            'total_leaves': total_leaves,
            'pending_leaves': pending_leaves,
            'approved_leaves': approved_leaves,
            'rejected_leaves': rejected_leaves,
            'leave_types': leave_types,
            'monthly_stats': monthly_stats,
            'department_stats': department_stats,
            'employee_stats': employee_stats,
            'approval_rate': calculate_approval_rate(all_requests),
            'avg_processing_time': calculate_avg_processing_time(all_requests),
            'current_month_requests': len(current_month_requests),
            'current_month_approval_rate': calculate_approval_rate(current_month_requests),
            'leave_types_summary': {
                'total': len(leave_types),
                'types': list(leave_types.keys())[:5]  # أول 5 أنواع فقط للعرض
            }
        }
        
    except Exception as e:
        print(f"Error calculating statistics: {str(e)}")
        # إرجاع إحصائيات افتراضية في حالة الخطأ
        return {
            'total_leaves': 0,
            'pending_leaves': 0,
            'approved_leaves': 0,
            'rejected_leaves': 0,
            'leave_types': {},
            'monthly_stats': {},
            'department_stats': {},
            'employee_stats': {},
            'approval_rate': 0,
            'avg_processing_time': 0,
            'current_month_requests': 0,
            'current_month_approval_rate': 0,
            'leave_types_summary': {'total': 0, 'types': []}
        }



def calculate_approval_rate(requests):
    """حساب معدل القبول"""
    approved = len([r for r in requests if r.status == 'approved'])
    total_processed = len([r for r in requests if r.status != 'pending'])
    
    if total_processed == 0:
        return 0
    return round((approved / total_processed) * 100, 1)

def calculate_avg_processing_time(requests):
    """حساب متوسط وقت المعالجة"""
    processed_requests = [r for r in requests if r.approved_at and r.created_at]
    if not processed_requests:
        return 0
    
    total_hours = sum([
        (r.approved_at - r.created_at).total_seconds() / 3600
        for r in processed_requests
    ])
    
    return round(total_hours / len(processed_requests), 1)

@app.route('/api/manager/leave_statistics')
@login_required
def api_leave_statistics():
    """API للحصول على إحصائيات الإجازات"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        department_ids = [dept.id for dept in managed_departments]
        
        # Get filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        department_id = request.args.get('department_id')
        leave_type = request.args.get('leave_type')
        
        # Build query
        query = db_session.query(LeaveRequest).filter(LeaveRequest.department_id.in_(department_ids))
        
        if start_date:
            query = query.filter(LeaveRequest.created_at >= start_date)
        if end_date:
            query = query.filter(LeaveRequest.created_at <= end_date)
        if department_id and department_id != 'all':
            query = query.filter(LeaveRequest.department_id == department_id)
        if leave_type and leave_type != 'all':
            query = query.filter(LeaveRequest.leave_type == leave_type)
        
        all_requests = query.all()
        
        # Calculate statistics
        stats = get_advanced_leave_statistics(department_ids, all_requests)
        
        return jsonify({
            'success': True,
            'statistics': stats
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@app.route('/api/manager/leave_reports')
@login_required
def api_leave_reports():
    """API لتوليد تقارير الإجازات"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        department_ids = [dept.id for dept in managed_departments]
        
        # Get filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        report_type = request.args.get('report_type', 'summary')
        
        # Build query
        query = db_session.query(LeaveRequest)\
            .options(
                joinedload(LeaveRequest.user),
                joinedload(LeaveRequest.department),
                joinedload(LeaveRequest.approver)
            )\
            .filter(LeaveRequest.department_id.in_(department_ids))
        
        if start_date:
            query = query.filter(LeaveRequest.created_at >= start_date)
        if end_date:
            query = query.filter(LeaveRequest.created_at <= end_date)
        
        all_requests = query.all()
        
        # Generate report based on type
        if report_type == 'detailed':
            report_data = generate_detailed_report(all_requests)
        elif report_type == 'employee':
            report_data = generate_employee_report(all_requests)
        elif report_type == 'department':
            report_data = generate_department_report(all_requests)
        else:  # summary
            report_data = generate_summary_report(all_requests)
        
        return jsonify({
            'success': True,
            'report_type': report_type,
            'report_data': report_data,
            'total_records': len(all_requests),
            'generated_at': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

def generate_summary_report(requests):
    """توليد تقرير ملخص"""
    summary = {
        'total_requests': len(requests),
        'pending_requests': len([r for r in requests if r.status == 'pending']),
        'approved_requests': len([r for r in requests if r.status == 'approved']),
        'rejected_requests': len([r for r in requests if r.status == 'rejected']),
        'total_days': sum([r.total_days or 1 for r in requests]),
        'approval_rate': calculate_approval_rate(requests),
        'avg_days_per_request': calculate_average_days(requests)
    }
    
    return summary

def generate_detailed_report(requests):
    """توليد تقرير مفصل"""
    detailed_data = []
    
    for request in requests:
        detailed_data.append({
            'employee_name': request.user.name if request.user else 'غير معروف',
            'department': request.department.name if request.department else 'غير معين',
            'leave_type': request.leave_type,
            'start_date': request.start_date.strftime('%Y-%m-%d') if request.start_date else '',
            'end_date': request.end_date.strftime('%Y-%m-%d') if request.end_date else '',
            'total_days': request.total_days or 1,
            'status': request.status,
            'reason': request.reason or '',
            'created_at': request.created_at.strftime('%Y-%m-%d %H:%M') if request.created_at else '',
            'processed_by': request.approver.name if request.approver else '',
            'processed_at': request.approved_at.strftime('%Y-%m-%d %H:%M') if request.approved_at else '',
            'rejection_reason': request.rejection_reason or ''
        })
    
    return detailed_data

def generate_employee_report(requests):
    """توليد تقرير الموظفين"""
    employee_data = {}
    
    for request in requests:
        if request.user:
            user_key = request.user_id
            if user_key not in employee_data:
                employee_data[user_key] = {
                    'employee_name': request.user.name,
                    'employee_username': request.user.username,
                    'department': request.department.name if request.department else 'غير معين',
                    'total_requests': 0,
                    'total_days': 0,
                    'approved_requests': 0,
                    'rejected_requests': 0,
                    'pending_requests': 0,
                    'approval_rate': 0
                }
            
            employee_data[user_key]['total_requests'] += 1
            employee_data[user_key]['total_days'] += request.total_days or 1
            
            if request.status == 'approved':
                employee_data[user_key]['approved_requests'] += 1
            elif request.status == 'rejected':
                employee_data[user_key]['rejected_requests'] += 1
            elif request.status == 'pending':
                employee_data[user_key]['pending_requests'] += 1
    
    # Calculate approval rates
    for employee in employee_data.values():
        total_processed = employee['approved_requests'] + employee['rejected_requests']
        if total_processed > 0:
            employee['approval_rate'] = round((employee['approved_requests'] / total_processed) * 100, 1)
    
    return list(employee_data.values())

def generate_department_report(requests):
    """توليد تقرير الأقسام"""
    department_data = {}
    
    for request in requests:
        if request.department:
            dept_key = request.department_id
            if dept_key not in department_data:
                department_data[dept_key] = {
                    'department_name': request.department.name,
                    'total_requests': 0,
                    'total_days': 0,
                    'approved_requests': 0,
                    'rejected_requests': 0,
                    'pending_requests': 0,
                    'unique_employees': set()
                }
            
            department_data[dept_key]['total_requests'] += 1
            department_data[dept_key]['total_days'] += request.total_days or 1
            
            if request.status == 'approved':
                department_data[dept_key]['approved_requests'] += 1
            elif request.status == 'rejected':
                department_data[dept_key]['rejected_requests'] += 1
            elif request.status == 'pending':
                department_data[dept_key]['pending_requests'] += 1
            
            if request.user:
                department_data[dept_key]['unique_employees'].add(request.user_id)
    
    # Convert sets to counts and calculate rates
    result = []
    for dept_id, data in department_data.items():
        data['unique_employees_count'] = len(data['unique_employees'])
        del data['unique_employees']
        
        total_processed = data['approved_requests'] + data['rejected_requests']
        if total_processed > 0:
            data['approval_rate'] = round((data['approved_requests'] / total_processed) * 100, 1)
        else:
            data['approval_rate'] = 0
        
        if data['total_requests'] > 0:
            data['avg_days_per_request'] = round(data['total_days'] / data['total_requests'], 1)
        else:
            data['avg_days_per_request'] = 0
        
        result.append(data)
    
    return result

def calculate_average_days(requests):
    """حساب متوسط أيام الإجازة"""
    if not requests:
        return 0
    
    total_days = sum([r.total_days or 1 for r in requests])
    return round(total_days / len(requests), 1)



@app.route('/manager/permission_requests')
@login_required
def manager_permission_requests():
    """طلبات الإذونات مع الفلاتر والإحصائيات المتقدمة"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    # Initialize variables
    managed_departments = []
    department_ids = []
    all_permissions = []
    pending_requests = []
    history_requests = []
    departments = []
    permission_stats = get_empty_permission_stats()
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        
        if not managed_departments:
            empty_stats = get_empty_permission_stats()
            return render_template('manager/manager_permission_requests.html',
                                 pending_requests=[],
                                 history_requests=[],
                                 departments=[],
                                 stats=empty_stats,
                                 notifications=[],
                                 new_salary_slips=0)
        
        department_ids = [dept.id for dept in managed_departments]
        
        # ====== الحصول على جميع طلبات الإذن مع JOINs ======
        all_permissions = db_session.query(PermissionRequest)\
            .options(
                joinedload(PermissionRequest.user),
                joinedload(PermissionRequest.department)
            )\
            .filter(PermissionRequest.department_id.in_(department_ids))\
            .order_by(PermissionRequest.created_at.desc())\
            .all()
        
        # ====== فصل الطلبات المعلقة عن السابقة ======
        for request in all_permissions:
            # Load approver separately if needed
            if request.approved_by:
                request.approved_by_user = db_session.query(User).get(request.approved_by)
            else:
                request.approved_by_user = None
                    
            if request.status == 'pending':
                pending_requests.append(request)
            else:
                history_requests.append(request)
        
        # ====== الحصول على الإحصائيات ======
        permission_stats = get_permission_statistics(department_ids, all_permissions)
        
        print(f"DEBUG: Total permissions found: {len(all_permissions)}")
        print(f"DEBUG: Pending permissions: {len(pending_requests)}")
        print(f"DEBUG: Processed permissions: {len(history_requests)}")
        
        # الحصول على الأقسام للفلتر
        departments = db_session.query(Department).filter(
            Department.id.in_(department_ids)
        ).all()
        
        # ====== شيتات المرتب الجديدة ======
        new_salary_slips = db_session.query(SalarySlip).filter_by(
            user_id=current_user.id, is_viewed=False
        ).count()
        
        # ====== الإشعارات ======
        notifications = get_user_notifications(current_user.id)
        
        return render_template('manager/manager_permission_requests.html',
                             pending_requests=pending_requests,
                             history_requests=history_requests,
                             departments=departments,
                             stats=permission_stats,
                             notifications=notifications,
                             new_salary_slips=new_salary_slips)
        
    except Exception as e:
        print(f"Error in manager_permission_requests: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'حدث خطأ في تحميل البيانات: {str(e)}', 'error')
        # Return empty data instead of crashing
        return render_template('manager/manager_permission_requests.html',
                             pending_requests=[],
                             history_requests=[],
                             departments=[],
                             stats=get_empty_permission_stats(),
                             notifications=[])
    
@app.route('/api/manager/history_permissions')
@login_required
def api_manager_history_permissions():
    """API للحصول على سجل طلبات الإذونات"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        department_ids = [dept.id for dept in managed_departments]
        
        if not department_ids:
            return jsonify({'success': True, 'requests': []})
        
        # الحصول على الطلبات السابقة (غير معلقة)
        history_requests = db_session.query(PermissionRequest)\
            .options(
                joinedload(PermissionRequest.user),
                joinedload(PermissionRequest.department),
                joinedload(PermissionRequest.approver)
            )\
            .filter(
                PermissionRequest.department_id.in_(department_ids),
                PermissionRequest.status != 'pending'
            )\
            .order_by(PermissionRequest.created_at.desc())\
            .all()
        
        # تحضير البيانات للاستجابة
        requests_data = []
        for req in history_requests:
            request_data = {
                'id': req.id,
                'user': {
                    'id': req.user.id if req.user else None,
                    'name': req.user.name if req.user else 'غير معروف'
                } if req.user else None,
                'department': {
                    'id': req.department.id if req.department else None,
                    'name': req.department.name if req.department else 'غير معين'
                } if req.department else None,
                'permission_type': req.permission_type,
                'date': req.date.strftime('%Y-%m-%d') if req.date else None,
                'time': req.time,
                'reason': req.reason,
                'rejection_reason': req.rejection_reason,
                'status': req.status,
                'created_at': req.created_at.strftime('%Y-%m-%d %H:%M') if req.created_at else None,
                'approved_at': req.approved_at.strftime('%Y-%m-%d %H:%M') if req.approved_at else None,
                'approved_by': {
                    'id': req.approver.id if req.approver else None,
                    'name': req.approver.name if req.approver else 'غير معروف'
                } if req.approver else None
            }
            requests_data.append(request_data)
        
        return jsonify({
            'success': True,
            'requests': requests_data,
            'count': len(requests_data)
        })
        
    except Exception as e:
        print(f"Error in history permissions API: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@app.route('/api/manager/permission_reports')
@login_required
def api_manager_permission_reports():
    """API للحصول على تقارير الإذونات"""
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
        department_ids = [dept.id for dept in managed_departments]
        
        if not department_ids:
            return jsonify({'success': True, 'employees': [], 'summary': {}})
        
        # الحصول على معاملات التقرير
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # بناء الاستعلام
        query = db_session.query(PermissionRequest)\
            .filter(PermissionRequest.department_id.in_(department_ids))
        
        if start_date:
            query = query.filter(PermissionRequest.created_at >= start_date)
        if end_date:
            query = query.filter(PermissionRequest.created_at <= end_date)
        
        all_permissions = query.all()
        
        # تجميع بيانات الموظفين
        employee_data = {}
        for perm in all_permissions:
            user = db_session.query(User).get(perm.user_id)
            department = db_session.query(Department).get(perm.department_id)
            
            if user:
                user_id = user.id
                if user_id not in employee_data:
                    employee_data[user_id] = {
                        'employee_id': user_id,
                        'employee_name': user.name,
                        'department': department.name if department else 'غير معين',
                        'total': 0,
                        'approved': 0,
                        'rejected': 0,
                        'pending': 0,
                        'total_hours': 0,
                        'approval_rate': 0
                    }
                
                employee_data[user_id]['total'] += 1
                
                if perm.status == 'approved':
                    employee_data[user_id]['approved'] += 1
                    # حساب الساعات الإضافية
                    if perm.extra_data:
                        try:
                            extra_data = json.loads(perm.extra_data)
                            if 'hours' in extra_data:
                                employee_data[user_id]['total_hours'] += int(extra_data['hours'])
                        except:
                            pass
                elif perm.status == 'rejected':
                    employee_data[user_id]['rejected'] += 1
                elif perm.status == 'pending':
                    employee_data[user_id]['pending'] += 1
        
        # حساب معدل القبول لكل موظف
        for emp_id, data in employee_data.items():
            total_processed = data['approved'] + data['rejected']
            if total_processed > 0:
                data['approval_rate'] = round((data['approved'] / total_processed) * 100, 1)
            else:
                data['approval_rate'] = 0
            
            # حساب متوسط الساعات الإضافية
            if data['total'] > 0:
                data['avg_hours'] = round(data['total_hours'] / data['total'], 1)
            else:
                data['avg_hours'] = 0
        
        # تحويل إلى قائمة وترتيب حسب عدد الطلبات
        employees_list = list(employee_data.values())
        employees_list.sort(key=lambda x: x['total'], reverse=True)
        
        # إحصائيات إضافية
        permission_types = {}
        for perm in all_permissions:
            perm_type = perm.permission_type or 'غير محدد'
            permission_types[perm_type] = permission_types.get(perm_type, 0) + 1
        
        peak_permission_type = max(permission_types, key=permission_types.get) if permission_types else 'لا توجد بيانات'
        
        return jsonify({
            'success': True,
            'employees': employees_list,
            'summary': {
                'total_employees': len(employees_list),
                'total_requests': len(all_permissions),
                'peak_permission_type': peak_permission_type,
                'avg_processing_time': 1.2
            }
        })
        
    except Exception as e:
        print(f"Error in permission reports API: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        })

@app.route('/admin/schedule_history/<int:dept_id>')
@login_required
def admin_schedule_history(dept_id):
    """عرض التاريخ الكامل للجداول لقسم معين"""
    
    department = db_session.query(Department).get(dept_id)
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('admin_schedules'))
    
    # الحصول على جميع الجداول للقسم مرتبة من الأحدث إلى الأقدم
    schedules = db_session.query(WeeklySchedule).filter_by(
        department_id=dept_id
    ).order_by(WeeklySchedule.week_start_date.desc()).all()
    
    # تجميع الجداول حسب الشهر والسنة
    schedules_by_period = {}
    for schedule in schedules:
        period_key = schedule.week_start_date.strftime('%Y-%m')  # السنة-الشهر
        if period_key not in schedules_by_period:
            schedules_by_period[period_key] = {
                'display_name': schedule.week_start_date.strftime('%B %Y'),  # اسم الشهر والسنة
                'schedules': []
            }
        schedules_by_period[period_key]['schedules'].append(schedule)
    
    # إحصائيات
    stats = {
        'total_schedules': len(schedules),
        'approved_schedules': len([s for s in schedules if s.is_approved]),
        'pending_schedules': len([s for s in schedules if s.status == 'pending']),
        'draft_schedules': len([s for s in schedules if s.status == 'draft']),
        'locked_schedules': len([s for s in schedules if s.is_locked]),
        'oldest_schedule': min(schedules, key=lambda x: x.week_start_date).week_start_date if schedules else None,
        'newest_schedule': max(schedules, key=lambda x: x.week_start_date).week_start_date if schedules else None
    }
    
    # حساب التواريخ الحالية
    today = date.today()
    days_since_saturday = (today.weekday() - 5) % 7
    current_week_start = today - timedelta(days=days_since_saturday)
    current_week_end = current_week_start + timedelta(days=6)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_schedule_history.html',
                         department=department,
                         schedules_by_period=schedules_by_period,
                         stats=stats,
                         today=today,
                         current_week_start=current_week_start,
                         current_week_end=current_week_end,
                         notifications=notifications)


@app.route('/admin/generate_schedules')
@login_required
def admin_generate_schedules():
    """إنشاء الجداول الأسبوعية يدوياً"""
    
    try:
        generated_count = auto_generate_weekly_schedules()
        return jsonify({
            'success': True, 
            'message': f'تم إنشاء {generated_count} جدول أسبوعي بنجاح',
            'count': generated_count
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في إنشاء الجداول: {str(e)}'
        })


@app.route('/admin/regenerate_all_future')
@login_required
def admin_regenerate_all_future():
    """إنشاء الجداول المستقبلية لجميع الأقسام"""
    
    try:
        departments = db_session.query(Department).filter_by(auto_generate_schedule=True).all()
        total_generated = 0
        
        for department in departments:
            generated_count = generate_future_schedules(department.id)
            total_generated += generated_count
            if generated_count > 0:
                print(f"تم إنشاء {generated_count} جدول للقسم {department.name}")
        
        flash(f'تم إنشاء {total_generated} جدول مستقبلي جديد لجميع الأقسام')
        
    except Exception as e:
        flash(f'خطأ في إنشاء الجداول: {str(e)}', 'error')
    
    return redirect(url_for('admin_schedules'))


@app.route('/admin/schedule/export/<int:schedule_id>')
@login_required
def admin_export_schedule(schedule_id):
    """تصدير الجدول بنفس هيكل العرض"""
    
    schedule = db_session.get(WeeklySchedule, schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('admin_schedules'))
    
    try:
        # الحصول على اسم القسم
        department_name = "غير معين"
        department = db_session.query(Department).get(schedule.department_id)
        if department:
            department_name = department.name
        
        print(f"تصدير جدول: {schedule_id}")
        print(f"بيانات الجدول: {schedule.schedule_data}")
        
        # إنشاء ملف Excel
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "جدول العمل"
        
        # إضافة العنوان
        worksheet.merge_cells('A1:G1')
        worksheet['A1'] = f'جدول العمل الأسبوعي - {department_name}'
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        worksheet.merge_cells('A2:G2')
        worksheet['A2'] = f'الفترة: {schedule.week_start_date} إلى {schedule.week_end_date}'
        worksheet['A2'].font = Font(size=12, bold=True)
        worksheet['A2'].alignment = Alignment(horizontal='center')
        
        # إضافة رؤوس الأعمدة
        headers = ['اليوم', 'التاريخ', 'الشيفت الصباحي', 'الشيفت المسائي', 'شيفت السهر', 'الوظيفة', 'القسم']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # معالجة البيانات بناءً على الهيكل الجديد
        row = 5
        entry_count = 0
        
        if schedule.schedule_data:
            try:
                schedule_data = json.loads(schedule.schedule_data)
                print(f"نوع البيانات: {type(schedule_data)}")
                print(f"مفاتيح البيانات: {schedule_data.keys() if isinstance(schedule_data, dict) else 'Not a dict'}")
                
                # الهيكل الجديد: يحتوي على schedule كمفتاح رئيسي
                if isinstance(schedule_data, dict) and 'schedule' in schedule_data:
                    schedule_entries = schedule_data['schedule']
                    print(f"عدد الإدخالات في الجدول: {len(schedule_entries)}")
                    
                    for entry in schedule_entries:
                        if isinstance(entry, dict):
                            # إضافة بيانات كل صف
                            worksheet.cell(row=row, column=1, value=entry.get('day', ''))
                            worksheet.cell(row=row, column=2, value=entry.get('date', ''))
                            worksheet.cell(row=row, column=3, value=entry.get('morning_shift', ''))
                            worksheet.cell(row=row, column=4, value=entry.get('evening_shift', ''))
                            worksheet.cell(row=row, column=5, value=entry.get('night_shift', ''))
                            worksheet.cell(row=row, column=6, value=entry.get('job', ''))
                            worksheet.cell(row=row, column=7, value=entry.get('department', department_name))
                            
                            row += 1
                            entry_count += 1
                            
                    print(f"تم إضافة {entry_count} إدخال إلى Excel")
                    
                else:
                    # محاولة الهيكل القديم
                    print("الهيكل غير متوقع، محاولة الهيكل القديم")
                    for key, value in schedule_data.items():
                        if isinstance(value, dict):
                            employee = db_session.query(User).get(int(key)) if key.isdigit() else None
                            if employee:
                                worksheet.cell(row=row, column=1, value=employee.name)
                                worksheet.cell(row=row, column=2, value=value.get('saturday', ''))
                                worksheet.cell(row=row, column=3, value=value.get('sunday', ''))
                                worksheet.cell(row=row, column=4, value=value.get('monday', ''))
                                worksheet.cell(row=row, column=5, value=value.get('tuesday', ''))
                                worksheet.cell(row=row, column=6, value=value.get('wednesday', ''))
                                worksheet.cell(row=row, column=7, value=value.get('thursday', ''))
                                row += 1
                                entry_count += 1
                                
            except Exception as e:
                print(f"خطأ في معالجة بيانات الجدول: {str(e)}")
                import traceback
                print(f"Traceback: {traceback.format_exc()}")
        
        # إذا لم توجد بيانات، إضافة رسالة
        if entry_count == 0:
            worksheet.merge_cells('A5:G5')
            worksheet.cell(row=5, column=1, value="لا توجد بيانات في الجدول")
            worksheet.cell(row=5, column=1).alignment = Alignment(horizontal='center')
            worksheet.cell(row=5, column=1).font = Font(color="FF0000", bold=True)
            print("لم يتم العثور على بيانات للتصدير")
        
        # ضبط عرض الأعمدة
        column_widths = {
            'A': 15,  # اليوم
            'B': 12,  # التاريخ
            'C': 20,  # الشيفت الصباحي
            'D': 20,  # الشيفت المسائي
            'E': 20,  # شيفت السهر
            'F': 30,  # الوظيفة
            'G': 20   # القسم
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # إضافة حدود للجدول
        from openpyxl.styles import Border, Side
        
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        for row_num in range(4, worksheet.max_row + 1):
            for col in range(1, 8):
                worksheet.cell(row=row_num, column=col).border = thin_border
        
        workbook.save(output)
        output.seek(0)
        
        filename = f"جدول_{department_name}_{schedule.week_start_date}_to_{schedule.week_end_date}.xlsx".replace(" ", "_")
        
        # تحديث حالة الجدول إذا كان معتمداً
        if schedule.is_approved and not schedule.is_locked:
            schedule.is_locked = True
            schedule.exported_at = datetime.now()
            schedule.exported_by = current_user.id
            db_session.commit()
        
        flash(f'تم تصدير الجدول بنجاح ({entry_count} إدخال)', 'success')
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"خطأ تفصيلي في التصدير: {e}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        flash(f'خطأ في تصدير الملف: {str(e)}', 'error')
        return redirect(url_for('admin_schedules'))


@app.route('/api/schedule/<int:schedule_id>/repopulate_from/<int:source_schedule_id>', methods=['POST'])
@login_required
def repopulate_schedule_from_another(schedule_id, source_schedule_id):
    """إعادة ملء الجدول الحالي من جدول آخر (نسخ الموظفين)"""

    
    try:
        # الحصول على الجدول الهدف
        target_schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not target_schedule:
            return jsonify({'success': False, 'message': 'الجدول الهدف غير موجود'})
        
        # الحصول على الجدول المصدر
        source_schedule = db_session.query(WeeklySchedule).get(source_schedule_id)
        if not source_schedule:
            return jsonify({'success': False, 'message': 'الجدول المصدر غير موجود'})
        
        # التحقق من أن الجداول لنفس القسم
        if target_schedule.department_id != source_schedule.department_id:
            return jsonify({
                'success': False, 
                'message': 'لا يمكن نسخ الجداول من أقسام مختلفة'
            })
        
        # التحقق من صلاحيات المستخدم
        department = db_session.query(Department).get(target_schedule.department_id)
        
        # التحقق من أن الجدول الهدف غير مقفل
        if target_schedule.is_locked:
            return jsonify({
                'success': False, 
                'message': 'الجدول الهدف مقفل ولا يمكن التعديل'
            })
        
        # ====== 1. الحصول على تفاصيل الجدول المصدر ======
        source_details = db_session.query(ScheduleDetail).filter_by(
            weekly_schedule_id=source_schedule_id
        ).order_by('day_date', 'row_order').all()
        
        print(f"تم العثور على {len(source_details)} تفصيلة في الجدول المصدر")
        
        # ====== 2. حذف جميع تفاصيل الجدول الهدف القديمة ======
        deleted_count = db_session.query(ScheduleDetail).filter(
            ScheduleDetail.weekly_schedule_id == schedule_id
        ).delete(synchronize_session=False)
        
        print(f"تم حذف {deleted_count} سجل من الجدول الهدف")
        
        # ====== 3. نسخ تفاصيل الجدول المصدر ======
        copied_count = 0
        
        # خريطة لتتبع التواريخ بين الجدولين
        date_mapping = {}
        
        # إنشاء تعيين للتواريخ (نفس الأيام من الأسبوع)
        for i in range(7):
            source_date = source_schedule.week_start_date + timedelta(days=i)
            target_date = target_schedule.week_start_date + timedelta(days=i)
            date_mapping[source_date] = target_date
        
        for source_detail in source_details:
            if source_detail.day_date in date_mapping:
                # إنشاء تفصيلة جديدة في الجدول الهدف
                new_detail = ScheduleDetail(
                    weekly_schedule_id=schedule_id,
                    day_date=date_mapping[source_detail.day_date],
                    day_name=get_arabic_day_name(date_mapping[source_detail.day_date]),
                    job_title=source_detail.job_title,
                    morning_shift=source_detail.morning_shift,
                    evening_shift=source_detail.evening_shift,
                    night_shift=source_detail.night_shift,
                    row_order=source_detail.row_order,
                    is_custom=False,  # تم إنشاؤها تلقائياً من نسخ
                    notes=source_detail.notes
                )
                
                db_session.add(new_detail)
                copied_count += 1
        
        # ====== 4. حفظ التغييرات ======
        target_schedule.updated_at = datetime.now()
        target_schedule.updated_by = current_user.id
        
        db_session.commit()
        
        # ====== 5. تسجيل العملية ======
        log_entry = ScheduleRepopulationLog(
            target_schedule_id=schedule_id,
            source_schedule_id=source_schedule_id,
            user_id=current_user.id,
            details=json.dumps({
                'source_week': source_schedule.week_start_date.strftime('%Y-%m-%d'),
                'target_week': target_schedule.week_start_date.strftime('%Y-%m-%d'),
                'copied_details': copied_count,
                'deleted_details': deleted_count,
                'operation': 'repopulate_from_another'
            }, ensure_ascii=False),
            performed_at=datetime.now()
        )
        db_session.add(log_entry)
        db_session.commit()
        
        # ====== 6. إرسال إشعار ======
        notification_message = f'تمت إعادة ملء جدول الأسبوع {target_schedule.week_start_date.strftime("%Y-%m-%d")} '
        notification_message += f'ببيانات من جدول الأسبوع {source_schedule.week_start_date.strftime("%Y-%m-%d")}'
        
        create_notification(
            current_user.id,
            'إعادة ملء جدول',
            notification_message,
            'schedule_repopulated',
            related_id=schedule_id,
            action_url=url_for('manager_edit_schedule', schedule_id=schedule_id)
        )
        
        return jsonify({
            'success': True,
            'message': f'تم نسخ {copied_count} تفصيلة وحذف {deleted_count} تفصيلة قديمة',
            'details': {
                'copied': copied_count,
                'deleted': deleted_count,
                'source_schedule': source_schedule.week_start_date.strftime('%Y-%m-%d'),
                'target_schedule': target_schedule.week_start_date.strftime('%Y-%m-%d')
            },
            'redirect_url': url_for('manager_edit_schedule', schedule_id=schedule_id)
        })
        
    except Exception as e:
        db_session.rollback()
        print(f"خطأ في إعادة ملء الجدول: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False, 
            'message': f'حدث خطأ: {str(e)}'
        })


from datetime import datetime, date

@app.route('/api/schedule/<int:schedule_id>/available_source_schedules')
@login_required
def get_available_source_schedules(schedule_id):
    """الحصول على الجداول المتاحة لاستخدامها كمصدر للإعادة الملء"""
    
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return jsonify({'success': False, 'message': 'الجدول غير موجود'})
        
        department = db_session.query(Department).get(schedule.department_id)
        
        # حساب بداية ونهاية السنة الحالية
        current_year = datetime.now().year
        start_of_year = date(current_year, 1, 1)
        end_of_year = date(current_year, 12, 31)
        
        # الحصول على جميع الجداول لنفس القسم في السنة الحالية، باستثناء الجدول الحالي
        source_schedules = db_session.query(WeeklySchedule).filter(
            WeeklySchedule.department_id == schedule.department_id,
            WeeklySchedule.id != schedule_id,  # استبعاد الجدول الحالي
            WeeklySchedule.week_start_date >= start_of_year,
            WeeklySchedule.week_start_date <= end_of_year
            # يمكن إضافة شرط للجداول المعتمدة إذا رغبت: ,WeeklySchedule.is_approved == True
        ).order_by(WeeklySchedule.week_start_date.desc()).all()
        
        schedules_data = []
        for src_schedule in source_schedules:
            # حساب عدد التفاصيل
            detail_count = db_session.query(ScheduleDetail).filter_by(
                weekly_schedule_id=src_schedule.id
            ).count()
            
            schedules_data.append({
                'id': src_schedule.id,
                'week_start_date': src_schedule.week_start_date.strftime('%Y-%m-%d'),
                'week_end_date': src_schedule.week_end_date.strftime('%Y-%m-%d'),
                'week_number': src_schedule.week_number,
                'detail_count': detail_count,
                'is_approved': src_schedule.is_approved,
                'created_at': src_schedule.created_at.strftime('%Y-%m-%d') if src_schedule.created_at else '',
                'display_name': f'أسبوع {src_schedule.week_number} ({src_schedule.week_start_date.strftime("%Y-%m-%d")}) - {detail_count} تفصيلة'
            })
        
        return jsonify({
            'success': True,
            'schedules': schedules_data,
            'current_schedule': {
                'id': schedule.id,
                'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
                'week_end_date': schedule.week_end_date.strftime('%Y-%m-%d'),
                'department_name': department.name if department else 'غير معين'
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': f'حدث خطأ: {str(e)}'
        })

@app.route('/api/schedule/<int:target_schedule_id>/duplicate/<int:source_schedule_id>', methods=['POST'])
@login_required
def duplicate_schedule(target_schedule_id, source_schedule_id):
    """نسخ تفاصيل جدول إلى جدول آخر"""
    
    try:
        target_schedule = db_session.query(WeeklySchedule).get(target_schedule_id)
        source_schedule = db_session.query(WeeklySchedule).get(source_schedule_id)
        
        if not target_schedule or not source_schedule:
            return jsonify({'success': False, 'message': 'الجدول غير موجود'})
        
        if target_schedule.department_id != source_schedule.department_id:
            return jsonify({'success': False, 'message': 'لا يمكن نسخ جدول من قسم آخر'})
        
        # حذف التفاصيل الحالية للجدول الهدف
        db_session.query(ScheduleDetail).filter_by(weekly_schedule_id=target_schedule_id).delete()
        
        # نسخ التفاصيل من الجدول المصدر
        source_details = db_session.query(ScheduleDetail).filter_by(
            weekly_schedule_id=source_schedule_id
        ).all()
        
        for detail in source_details:
            new_detail = ScheduleDetail(
                weekly_schedule_id=target_schedule_id,
                employee_id=detail.employee_id,
                shift_type=detail.shift_type,
                monday=detail.monday,
                tuesday=detail.tuesday,
                wednesday=detail.wednesday,
                thursday=detail.thursday,
                friday=detail.friday,
                saturday=detail.saturday,
                sunday=detail.sunday,
                notes=detail.notes,
                created_by=current_user.id,
                created_at=datetime.now()
            )
            db_session.add(new_detail)
        
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم نسخ {len(source_details)} تفصيلة بنجاح',
            'copied_count': len(source_details)
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({
            'success': False, 
            'message': f'حدث خطأ: {str(e)}'
        })



import os
import pandas as pd
import glob
from werkzeug.utils import secure_filename


def get_user_excel_files(user_id=None, username=None, department_id=None):
    """
    Get Excel files for a user or department
    """
    excel_data_folder = app.config['EXCEL_DATA_FOLDER']
    
    if not os.path.exists(excel_data_folder):
        os.makedirs(excel_data_folder)
        return []
    
    files_data = []
    
    if username:
        # Handle username with .0 suffix for folder lookup
        possible_usernames = [username]
        
        # If username doesn't have .0 suffix, also check with .0
        if not username.endswith('.0'):
            possible_usernames.append(username + '.0')
        # If username has .0 suffix, also check without it
        else:
            base_username = username.rstrip('.0')
            possible_usernames.append(base_username)
        
        for possible_username in possible_usernames:
            user_folder = os.path.join(excel_data_folder, possible_username)
            if os.path.exists(user_folder):
                excel_files = glob.glob(os.path.join(user_folder, "*.xlsx")) + \
                             glob.glob(os.path.join(user_folder, "*.xls"))
                
                for file_path in excel_files:
                    try:
                        file_info = extract_excel_data(file_path, username)  # Store original username
                        if file_info:
                            files_data.append(file_info)
                    except Exception as e:
                        print(f"Error reading Excel file {file_path}: {str(e)}")
                break  # Found folder, no need to check other variations
    
    elif department_id:
        # Get files for all users in department
        users = db_session.query(User).filter_by(department_id=department_id, is_admin=False).all()
        for user in users:
            # Check both username formats
            possible_usernames = [user.username]
            
            # Check with .0 suffix
            username_with_suffix = user.username + '.0'
            possible_usernames.append(username_with_suffix)
            
            for possible_username in possible_usernames:
                user_folder = os.path.join(excel_data_folder, possible_username)
                if os.path.exists(user_folder):
                    excel_files = glob.glob(os.path.join(user_folder, "*.xlsx")) + \
                                 glob.glob(os.path.join(user_folder, "*.xls"))
                    
                    for file_path in excel_files:
                        try:
                            file_info = extract_excel_data(file_path, user.username)  # Store database username
                            if file_info:
                                file_info['user_name'] = user.name
                                file_info['user_id'] = user.id
                                # Store the actual folder username for download
                                file_info['folder_username'] = possible_username
                                files_data.append(file_info)
                        except Exception as e:
                            print(f"Error reading Excel file {file_path}: {str(e)}")
                    break  # Found folder, no need to check other variations
    
    return files_data


def extract_excel_data(file_path, username):
    """
    Extract data from Excel file with specified structure
    """
    try:
        # Read Excel file
        df = pd.read_excel(file_path)
        
        # Keep original column names as they are in Excel
        original_columns = list(df.columns)
        
        # Map expected column names (for filtering/search)
        column_mapping = {
            'رقم المكينة': 'machine_number',
            'اسم الموظف': 'employee_name',  # Updated from 'اسم المستخدم'
            'اسم المستخدم': 'username',  # This is the username/code
            'كود المستخدم': 'user_code',  # Updated mapping
            'كود المستخدم الالي': 'auto_user_code',
            'التاريخ': 'date',
            # English column names
            'machine_number': 'machine_number',
            'employee_name': 'employee_name',
            'username': 'username',
            'user_code': 'user_code',
            'auto_user_code': 'auto_user_code',
            'date': 'date'
        }
        
        # Create a reverse mapping for display
        reverse_mapping = {v: k for k, v in column_mapping.items()}
        
        # Create a normalized copy for filtering, keep original for display
        df_normalized = df.copy()
        
        # Rename columns in normalized copy for consistent filtering
        for original_col in original_columns:
            # Check Arabic column names
            if original_col in column_mapping:
                df_normalized.rename(columns={original_col: column_mapping[original_col]}, inplace=True)
            # Check case-insensitive
            else:
                for arabic_col, english_col in column_mapping.items():
                    if isinstance(original_col, str) and original_col.lower() == arabic_col.lower():
                        df_normalized.rename(columns={original_col: english_col}, inplace=True)
                        break
        
        # Convert date column to datetime if exists
        if 'date' in df_normalized.columns:
            try:
                # Keep original date string for display
                df['date_display'] = df_normalized['date'].astype(str)
                df_normalized['date'] = pd.to_datetime(df_normalized['date'], errors='coerce')
            except:
                df['date_display'] = ''
        
        # Add username from parameter for filtering
        df_normalized['folder_username'] = username
        
        # Limit to 1000 rows for performance
        if len(df) > 1000:
            df = df.head(1000)
            df_normalized = df_normalized.head(1000)
        
        # Convert to records while preserving original column names
        records = []
        for idx in range(len(df)):
            record = {}
            # Add original Excel columns
            for col in original_columns:
                value = df.iloc[idx][col]
                # Keep original format for display
                record[col] = str(value) if pd.notna(value) else ''
            
            # Add normalized values for filtering
            for col in df_normalized.columns:
                if col not in original_columns:  # Only add normalized columns not in original
                    value = df_normalized.iloc[idx][col]
                    record[f'_{col}'] = value if pd.notna(value) else None
            
            records.append(record)
        
        return {
            'file_path': file_path,
            'file_name': os.path.basename(file_path),
            'original_columns': original_columns,  # Keep original column names
            'data': records,
            'row_count': len(df),
            'username': username,
            'last_modified': os.path.getmtime(file_path),
            'has_employee_name': any('اسم الموظف' in col or 'employee_name' in col.lower() for col in original_columns),
            'has_username': any('اسم المستخدم' in col or 'username' in col.lower() for col in original_columns),
            'has_user_code': any('كود المستخدم' in col or 'user_code' in col.lower() for col in original_columns)
        }
        
    except Exception as e:
        print(f"Error processing Excel file {file_path}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def filter_excel_data(files_data, start_date=None, end_date=None, username=None):
    """
    Filter Excel data by date and username
    """
    filtered_data = []
    
    for file_info in files_data:
        filtered_records = []
        
        for record in file_info['data']:
            # Filter by date if provided
            date_match = True
            if start_date and 'date' in record and record['date']:
                try:
                    record_date = pd.to_datetime(record['date']).date()
                    if record_date < pd.to_datetime(start_date).date():
                        date_match = False
                except:
                    pass
            
            if end_date and 'date' in record and record['date']:
                try:
                    record_date = pd.to_datetime(record['date']).date()
                    if record_date > pd.to_datetime(end_date).date():
                        date_match = False
                except:
                    pass
            
            # Filter by username if provided
            username_match = True
            if username and 'username' in record:
                if isinstance(username, list):
                    if record['username'] not in username:
                        username_match = False
                elif record['username'] != username:
                    username_match = False
            
            if date_match and username_match:
                filtered_records.append(record)
        
        if filtered_records:
            filtered_file_info = file_info.copy()
            filtered_file_info['data'] = filtered_records
            filtered_file_info['row_count'] = len(filtered_records)
            filtered_data.append(filtered_file_info)
    
    return filtered_data


def filter_excel_data(files_data, start_date=None, end_date=None, username=None, employee_name=None):
    """
    Filter Excel data by date and username
    """
    filtered_data = []
    
    for file_info in files_data:
        filtered_records = []
        
        for record in file_info['data']:
            # Filter by date if provided
            date_match = True
            if start_date:
                date_value = None
                # Try to get date from normalized field
                if '_date' in record and record['_date']:
                    date_value = record['_date']
                # Or look for date in original columns
                elif any('تاريخ' in key or 'date' in key.lower() for key in record.keys()):
                    for key in record.keys():
                        if 'تاريخ' in key or 'date' in key.lower():
                            try:
                                date_value = pd.to_datetime(record[key], errors='coerce')
                                break
                            except:
                                pass
                
                if date_value:
                    try:
                        if pd.to_datetime(date_value).date() < pd.to_datetime(start_date).date():
                            date_match = False
                    except:
                        pass
            
            if end_date and date_value:
                try:
                    if pd.to_datetime(date_value).date() > pd.to_datetime(end_date).date():
                        date_match = False
                except:
                    pass
            
            # Filter by username if provided (using normalized field)
            username_match = True
            if username and '_folder_username' in record:
                if record['_folder_username'] != username:
                    username_match = False
            
            # Filter by employee name if provided
            employee_match = True
            if employee_name:
                found_employee = False
                for key in record.keys():
                    if 'اسم الموظف' in key or 'employee_name' in key.lower():
                        if employee_name.lower() in str(record[key]).lower():
                            found_employee = True
                            break
                if not found_employee:
                    employee_match = False
            
            if date_match and username_match and employee_match:
                filtered_records.append(record)
        
        if filtered_records:
            filtered_file_info = file_info.copy()
            filtered_file_info['data'] = filtered_records
            filtered_file_info['row_count'] = len(filtered_records)
            filtered_data.append(filtered_file_info)
    
    return filtered_data

@app.route('/user/excel_data')
@login_required
def user_excel_data():
    """Display Excel data exactly as in Excel files for the logged-in user"""
    
    # Check if user is admin or manager
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('manager_excel_data'))
    
    # Get filter parameters
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Get Excel files for this user
    files_data = get_user_excel_files(username=current_user.username)
    
    # Apply filters
    filtered_data = []
    for file_info in files_data:
        filtered_records = []
        for record in file_info['data']:
            # Check date filter
            date_match = True
            if start_date:
                date_value = None
                # Try to get date from normalized field
                if '_date' in record and record['_date']:
                    date_value = record['_date']
                # Or look for date in original columns
                else:
                    for key in record.keys():
                        if 'تاريخ' in key or 'date' in key.lower():
                            try:
                                date_value = pd.to_datetime(record[key], errors='coerce')
                                break
                            except:
                                pass
                
                if date_value:
                    try:
                        if pd.to_datetime(date_value).date() < pd.to_datetime(start_date).date():
                            date_match = False
                    except:
                        pass
            
            if end_date and date_value:
                try:
                    if pd.to_datetime(date_value).date() > pd.to_datetime(end_date).date():
                        date_match = False
                except:
                    pass
            
            if date_match:
                filtered_records.append(record)
        
        if filtered_records:
            filtered_file = file_info.copy()
            filtered_file['data'] = filtered_records
            filtered_file['row_count'] = len(filtered_records)
            filtered_data.append(filtered_file)
    
    files_data = filtered_data
    
    # Calculate statistics
    total_records = sum([file_info['row_count'] for file_info in files_data])
    total_files = len(files_data)
    
    # Prepare data for template - show exactly as in Excel
    all_records = []
    for file_info in files_data:
        for record in file_info['data']:
            # Create a copy of the record with original Excel data
            excel_record = {
                'file_name': file_info['file_name'],
                'file_path': file_info['file_path'],
                'excel_data': record  # Keep all original Excel columns
            }
            
            # Extract common fields for easy access in template
            for key, value in record.items():
                if not key.startswith('_'):  # Skip normalized fields
                    if 'رقم المكينة' in key or 'machine_number' in key.lower():
                        excel_record['machine_number'] = value
                    elif 'اسم الموظف' in key or 'employee_name' in key.lower():
                        excel_record['employee_name'] = value
                    elif 'اسم المستخدم' in key or 'username' in key.lower():
                        excel_record['username'] = value
                    elif 'كود المستخدم' in key or 'user_code' in key.lower():
                        excel_record['user_code'] = value
                    elif 'كود المستخدم الالي' in key or 'auto_user_code' in key.lower():
                        excel_record['auto_user_code'] = value
                    elif 'تاريخ' in key or 'date' in key.lower():
                        excel_record['date'] = value
                        # Try to format date for display
                        try:
                            if pd.notna(value):
                                excel_record['formatted_date'] = pd.to_datetime(value).strftime('%Y-%m-%d')
                            else:
                                excel_record['formatted_date'] = ''
                        except:
                            excel_record['formatted_date'] = str(value) if pd.notna(value) else ''
            
            all_records.append(excel_record)
    
    # Sort by date if available
    try:
        all_records.sort(key=lambda x: pd.to_datetime(x.get('date', ''), errors='coerce') or pd.Timestamp.min, reverse=True)
    except:
        pass
    
    # Get all unique column names from all Excel files for dynamic table headers
    all_columns = set()
    for file_info in files_data:
        if 'original_columns' in file_info:
            all_columns.update(file_info['original_columns'])
    
    # Common column order (Arabic first, then English)
    common_columns_order = [
        'رقم المكينة',
        'اسم الموظف',
        'اسم المستخدم',
        'كود المستخدم',
        'كود المستخدم الالي',
        'التاريخ',
        'machine_number',
        'employee_name',
        'username',
        'user_code',
        'auto_user_code',
        'date'
    ]
    
    # Sort columns: common ones first, then others
    sorted_columns = []
    for col in common_columns_order:
        if col in all_columns:
            sorted_columns.append(col)
            all_columns.remove(col)
    
    # Add remaining columns
    sorted_columns.extend(sorted(all_columns))
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_excel_data.html',
                         records=all_records,
                         total_records=total_records,
                         total_files=total_files,
                         columns=sorted_columns,
                         start_date=start_date,
                         end_date=end_date,
                         notifications=notifications)


def format_date_for_display(date_value):
    """Format date for display"""
    if pd.isna(date_value) or not date_value:
        return ''
    
    try:
        if isinstance(date_value, str):
            date_obj = pd.to_datetime(date_value)
        else:
            date_obj = pd.to_datetime(date_value)
        
        return date_obj.strftime('%Y-%m-%d')
    except:
        return str(date_value)

def parse_date(date_value):
    """Parse date for sorting"""
    if pd.isna(date_value) or not date_value:
        return pd.Timestamp.min
    
    try:
        return pd.to_datetime(date_value)
    except:
        return pd.Timestamp.min
    

@app.route('/manager/excel_data')
@login_required
def manager_excel_data():
    """Display Excel data exactly as in Excel files for all users in manager's department"""
    
    # Get manager's departments
    managed_departments = db_session.query(Department).filter((Department.primary_manager_id == current_user.id) | (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(DepartmentManager.user_id == current_user.id)))).all()
    

    
    department_ids = [dept.id for dept in managed_departments]
    
    # Get filter parameters
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    selected_department = request.args.get('department', '')
    selected_username = request.args.get('username', '')
    selected_employee = request.args.get('employee', '')
    
    # Get users in selected department or all managed departments
    if selected_department and selected_department != 'all':
        department_id = int(selected_department)
        # Verify manager has access to this department
        if department_id in department_ids:
            target_department_id = department_id
        else:
            target_department_id = department_ids[0]
    else:
        # Use all managed departments
        target_department_id = None
    
    # Get all Excel data for the department(s)
    if target_department_id:
        files_data = get_user_excel_files(department_id=target_department_id)
    else:
        files_data = []
        for dept_id in department_ids:
            files_data.extend(get_user_excel_files(department_id=dept_id))
    
    # Apply filters
    filtered_files_data = []
    for file_info in files_data:
        filtered_records = []
        
        for record in file_info['data']:
            # Check date filter
            date_match = True
            if start_date:
                date_value = None
                if '_date' in record and record['_date']:
                    date_value = record['_date']
                else:
                    for key in record.keys():
                        if 'تاريخ' in key or 'date' in key.lower():
                            try:
                                date_value = pd.to_datetime(record[key], errors='coerce')
                                break
                            except:
                                pass
                
                if date_value:
                    try:
                        if pd.to_datetime(date_value).date() < pd.to_datetime(start_date).date():
                            date_match = False
                    except:
                        pass
            
            if end_date and date_value:
                try:
                    if pd.to_datetime(date_value).date() > pd.to_datetime(end_date).date():
                        date_match = False
                except:
                    pass
            
            # Check employee name filter
            employee_match = True
            if selected_employee and selected_employee != 'all':
                found = False
                for key in record.keys():
                    if 'اسم الموظف' in key or 'employee_name' in key.lower():
                        if selected_employee.lower() in str(record[key]).lower():
                            found = True
                            break
                if not found:
                    employee_match = False
            
            if date_match and employee_match:
                filtered_records.append(record)
        
        if filtered_records:
            filtered_file = file_info.copy()
            filtered_file['data'] = filtered_records
            filtered_file['row_count'] = len(filtered_records)
            filtered_files_data.append(filtered_file)
    
    files_data = filtered_files_data
    
    # Calculate statistics
    total_records = sum([file_info['row_count'] for file_info in files_data])
    total_files = len(files_data)
    
    # Get unique usernames and employee names for filter dropdowns
    unique_employees = set()
    
    for file_info in files_data:
        for record in file_info['data']:
            # Extract employee name
            for key in record.keys():
                if not key.startswith('_'):
                    if 'اسم الموظف' in key or 'employee_name' in key.lower():
                        if record[key]:
                            unique_employees.add(str(record[key]))
    
    # Prepare data for template - CONVERT TO JSON-SERIALIZABLE FORMAT
    all_records = []
    for file_info in files_data:
        user_name = file_info.get('user_name', file_info.get('username', ''))
        user_id = file_info.get('user_id', '')
        department_name = ''
        
        # Get department name from user
        if user_id:
            user_obj = db_session.query(User).get(user_id)
            if user_obj and user_obj.department_id:
                dept = db_session.query(Department).get(user_obj.department_id)
                department_name = dept.name if dept else ''
        
        for record in file_info['data']:
            # Convert all values to string for JSON serialization
            excel_data_serializable = {}
            for key, value in record.items():
                if not key.startswith('_'):
                    # Convert pandas types to Python types
                    if pd.isna(value):
                        excel_data_serializable[key] = ''
                    elif hasattr(value, 'item'):  # numpy/pandas scalar
                        try:
                            excel_data_serializable[key] = value.item()
                        except:
                            excel_data_serializable[key] = str(value)
                    elif isinstance(value, (pd.Timestamp, pd.DatetimeIndex)):
                        excel_data_serializable[key] = str(value)
                    else:
                        excel_data_serializable[key] = str(value)
            
            # Create record with serializable data
            excel_record = {
                'file_name': file_info['file_name'],
                'file_path': file_info['file_path'],
                'user_name': str(user_name) if user_name else 'غير محدد',
                'department_name': str(department_name) if department_name else 'غير محدد',
                'excel_data': excel_data_serializable,
                'user_id': str(user_id) if user_id else '',
                'row_index': len(all_records) + 1
            }
            
            # Extract common fields for easy access
            for key, value in excel_data_serializable.items():
                if 'رقم المكينة' in key or 'machine_number' in key.lower():
                    excel_record['machine_number'] = value
                elif 'اسم الموظف' in key or 'employee_name' in key.lower():
                    excel_record['employee_name'] = value
                elif 'اسم المستخدم' in key or 'username' in key.lower():
                    excel_record['username'] = value
                elif 'كود المستخدم' in key or 'user_code' in key.lower():
                    excel_record['user_code'] = value
                elif 'كود المستخدم الالي' in key or 'auto_user_code' in key.lower():
                    excel_record['auto_user_code'] = value
                elif 'تاريخ' in key or 'date' in key.lower():
                    excel_record['date'] = value
                    # Try to format date for display
                    try:
                        if value and str(value).strip():
                            excel_record['formatted_date'] = pd.to_datetime(value).strftime('%Y-%m-%d')
                        else:
                            excel_record['formatted_date'] = ''
                    except:
                        excel_record['formatted_date'] = str(value) if value else ''
            
            all_records.append(excel_record)
    
    # Sort by date if available
    try:
        all_records.sort(key=lambda x: (
            pd.to_datetime(x.get('date', ''), errors='coerce') 
            or pd.Timestamp.min
        ), reverse=True)
    except:
        pass
    
    # Get departments for filter dropdown
    departments = db_session.query(Department).filter(
        Department.id.in_(department_ids)
    ).all()
    
    # Calculate date ranges for quick actions
    today = date.today().isoformat()
    
    # Get week start (Saturday) and end (Friday)
    today_obj = date.today()
    days_since_saturday = (today_obj.weekday() + 2) % 7
    week_start = (today_obj - timedelta(days=days_since_saturday)).isoformat()
    week_end = (today_obj + timedelta(days=(6 - days_since_saturday))).isoformat()
    
    # Get month start and end
    month_start = date(today_obj.year, today_obj.month, 1).isoformat()
    if today_obj.month == 12:
        month_end = date(today_obj.year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(today_obj.year, today_obj.month + 1, 1) - timedelta(days=1)
    month_end = month_end.isoformat()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager_excel_data.html',
                         records=all_records,
                         total_records=total_records,
                         total_files=total_files,
                         departments=departments,
                         employees=sorted(unique_employees),
                         selected_department=selected_department,
                         selected_employee=selected_employee,
                         start_date=start_date,
                         end_date=end_date,
                         today=today,
                         week_start=week_start,
                         week_end=week_end,
                         month_start=month_start,
                         month_end=month_end,
                         notifications=notifications)

@app.route('/download_excel/<path:file_path>')
@login_required
def download_excel(file_path):
    """Download an Excel file"""
    
    # Security check
    if '..' in file_path or file_path.startswith('/'):
        flash('مسار الملف غير صالح', 'error')
        return redirect(request.referrer or url_for('user_dashboard'))
    
    # Check if file exists
    if not os.path.exists(file_path):
        flash('الملف غير موجود', 'error')
        return redirect(request.referrer or url_for('user_dashboard'))
    
    # For regular users, check if file belongs to them
    if not current_user.is_admin and not current_user.is_manager:
        user_folder = os.path.join(app.config['EXCEL_DATA_FOLDER'], current_user.username)
        if not file_path.startswith(os.path.abspath(user_folder)):
            flash('ليس لديك صلاحية للوصول إلى هذا الملف', 'error')
            return redirect(url_for('user_dashboard'))
    
    # For managers, check if file belongs to their department
    elif current_user.is_manager:
        # Extract username from file path
        path_parts = file_path.split(os.sep)
        if len(path_parts) >= 2:
            username_from_path = path_parts[-2]  # Second last part should be username
            
            # Check if this user is in manager's department
            user = db_session.query(User).filter_by(username=username_from_path).first()
            if user:
                managed_departments = db_session.query(Department).filter_by(
                    primary_manager_id=current_user.id
                ).all()
                department_ids = [dept.id for dept in managed_departments]
                
                if user.department_id not in department_ids:
                    flash('ليس لديك صلاحية للوصول إلى ملفات هذا المستخدم', 'error')
                    return redirect(url_for('manager_dashboard'))
    
    # Download the file
    return send_file(
        file_path,
        as_attachment=True,
        download_name=os.path.basename(file_path)
    )




@app.route('/api/excel_details')
@login_required
def api_excel_details():
    """API to get Excel file details"""
    try:
        file_path = request.args.get('file_path')
        
        if not file_path:
            return jsonify({'success': False, 'message': 'مسار الملف مطلوب'})
        
        # Check file exists
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'الملف غير موجود'})
        
        # Extract file info
        file_info = extract_excel_data(file_path, '')
        
        if not file_info:
            return jsonify({'success': False, 'message': 'تعذر قراءة الملف'})
        
        return jsonify({
            'success': True,
            'details': {
                'file_name': os.path.basename(file_path),
                'row_count': file_info.get('row_count', 0),
                'last_modified': os.path.getmtime(file_path),
                'original_columns': file_info.get('original_columns', []),
                'sample_data': file_info.get('data', [])[:5]  # First 5 rows
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/export/excel')
@login_required
def export_excel_data():
    """Export filtered data to Excel"""
    try:
        # Get filter parameters
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        department = request.args.get('department', '')
        employee = request.args.get('employee', '')
        
        # Get data based on filters (reuse your existing logic)
        # ... (implement based on your existing manager_excel_data logic)
        
        # Create Excel file
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "بيانات البصمات"
        
        # Set Arabic headers
        headers = ['الموظف', 'القسم', 'التاريخ والوقت', 'رقم الماكينة', 'كود المستخدم']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
        
        # Add data
        row = 2
        # ... (add your data here)
        
        # Save to buffer
        workbook.save(output)
        output.seek(0)
        
        filename = f"بيانات_البصمات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'خطأ في التصدير: {str(e)}', 'error')
        return redirect(url_for('manager_excel_data'))

def download_excel(file_path):
    """Download Excel file with proper encoding"""
    try:
        # Decode URL
        decoded_path = unquote(file_path)
        
        # Security check
        if '..' in decoded_path or decoded_path.startswith('/'):
            return "مسار غير صالح", 400
        
        # Check if file exists
        if not os.path.exists(decoded_path):
            return "الملف غير موجود", 404
        
        # Get filename
        filename = os.path.basename(decoded_path)
        
        return send_file(
            decoded_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return str(e), 500



@app.route('/admin/fix_schedule_status')
@login_required
def fix_schedule_status():
    """Fix schedule status - make is_approved match is_locked"""
    if not current_user.is_admin:
        flash('غير مصرح', 'error')
        return redirect(url_for('admin_dashboard'))
    
    try:
        # Get all schedules
        schedules = db_session.query(WeeklySchedule).all()
        updated_count = 0
        
        print(f"Found {len(schedules)} schedules total")
        
        for schedule in schedules:
            if schedule.is_locked != schedule.is_approved:
                print(f"Schedule {schedule.id}: is_locked={schedule.is_locked}, is_approved={schedule.is_approved} -> setting to {schedule.is_locked}")
                schedule.is_approved = schedule.is_locked
                updated_count += 1
        
        db_session.commit()
        
        flash(f'✅ تم تحديث {updated_count} جدول: تم جعل is_approved مطابقاً لـ is_locked', 'success')
        print(f"Successfully updated {updated_count} schedules")
        
    except Exception as e:
        db_session.rollback()
        flash(f'❌ حدث خطأ: {str(e)}', 'error')
        print(f"Error details: {str(e)}")
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('admin_weekly_schedules'))

from sqlalchemy import func


@app.route('/export/weekly_schedule_excel', methods=['GET', 'POST'])
@login_required
def export_weekly_schedule_excel():
    """تصدير جدول العمل الأسبوعي بنفس الهيكل المطلوب"""
    
    try:
        # Get filter parameters from request
        if request.method == 'POST':
            # Handle form submission
            department_ids = request.form.getlist('department_ids[]')
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
        else:
            # Handle GET request with query parameters
            department_ids = request.args.getlist('department_ids[]')
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
        
        # Validate inputs
        if not department_ids:
            if request.method == 'POST':
                flash('يرجى اختيار قسم واحد على الأقل', 'error')
                return redirect(request.referrer or url_for('admin_weekly_schedules'))
            else:
                return jsonify({'success': False, 'message': 'يرجى اختيار قسم واحد على الأقل'})
        
        if not start_date or not end_date:
            if request.method == 'POST':
                flash('يرجى تحديد الفترة الزمنية', 'error')
                return redirect(request.referrer or url_for('admin_weekly_schedules'))
            else:
                return jsonify({'success': False, 'message': 'يرجى تحديد الفترة الزمنية'})
        
        # Convert date strings to date objects
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Get departments
        departments = db_session.query(Department).filter(Department.id.in_(department_ids)).all()
        
        if not departments:
            if request.method == 'POST':
                flash('لم يتم العثور على الأقسام المحددة', 'error')
                return redirect(request.referrer or url_for('admin_weekly_schedules'))
            else:
                return jsonify({'success': False, 'message': 'لم يتم العثور على الأقسام المحددة'})
        
        # ====== تجهيز قاموس بأكواد الوظائف من قاعدة البيانات ======
        job_codes_dict = {}
        for department in departments:
            # جلب جميع أكواد الوظائف لهذا القسم
            structure_rows = db_session.query(ScheduleStructureRow).filter_by(
                department_id=department.id
            ).all()
            
            for row in structure_rows:
                if row.job_title and row.job_code:
                    job_codes_dict[row.job_title] = row.job_code
        
        print(f"✅ تم تحميل {len(job_codes_dict)} كود وظيفة من قاعدة البيانات")
        
        # Create Excel file
        output = BytesIO()
        workbook = Workbook()
        
        # Remove default sheet if empty
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Create a worksheet for each department
        for department in departments:
            # Get schedules for this department within the date range
            schedules = db_session.query(WeeklySchedule).filter(
                WeeklySchedule.department_id == department.id,
                WeeklySchedule.week_start_date >= start_date_obj,
                WeeklySchedule.week_end_date <= end_date_obj,
            ).order_by(WeeklySchedule.week_start_date).all()
            
            if not schedules:
                continue
            
            # Create worksheet for this department
            sheet_name = department.name[:20]  # Excel sheet name max 31 chars
            worksheet = workbook.create_sheet(title=sheet_name)
            
            # Get schedule details for all weeks
            all_details = []
            for schedule in schedules:
                details = db_session.query(ScheduleDetail).filter_by(
                    weekly_schedule_id=schedule.id
                ).order_by('day_date', 'row_order').all()
                all_details.extend(details)
            
            if not all_details:
                continue
            
            # ====== HEADER SECTION ======
            # Title
            worksheet.merge_cells('A1:O1')
            title_cell = worksheet['A1']
            title_cell.value = f'جدول العمل الأسبوعي - {department.name}'
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Period
            worksheet.merge_cells('A2:O2')
            period_cell = worksheet['A2']
            period_cell.value = f'الفترة: {start_date} إلى {end_date}'
            period_cell.font = Font(size=12, bold=True)
            period_cell.alignment = Alignment(horizontal='center')
            
            # ====== COLUMN HEADERS ======
            headers = [
                'اليوم', 'التاريخ', 
                'كود موظف شفت A', 'الشيفت الصباحي',
                'كود موظف شفتB', 'الشيفت المسائي',
                'كود موظف شفتC', 'شيفت السهر',
                'كود الوظيفة', 'الوظيفة', 'القسم', 'البيان',
                'لا يحسب شفتA', 'لا يحسب شفتB', 'لا يحسب شفتC'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # ====== ADD DATA ======
            row = 5
            
            # Group details by date
            details_by_date = {}
            for detail in all_details:
                date_key = detail.day_date.strftime('%Y-%m-%d')
                if date_key not in details_by_date:
                    details_by_date[date_key] = []
                details_by_date[date_key].append(detail)
            
            # Sort by date
            sorted_dates = sorted(details_by_date.keys())
            
            for date_key in sorted_dates:
                day_details = details_by_date[date_key]
                
                # Sort by row_order
                day_details.sort(key=lambda x: x.row_order)
                
                for detail in day_details:
                    # Extract employee usernames and names from shift texts
                    morning_username = get_employee_username_from_shift(detail.morning_shift)
                    morning_name = extract_employee_name_from_shift(detail.morning_shift)
                    
                    evening_username = get_employee_username_from_shift(detail.evening_shift)
                    evening_name = extract_employee_name_from_shift(detail.evening_shift)
                    
                    night_username = get_employee_username_from_shift(detail.night_shift)
                    night_name = extract_employee_name_from_shift(detail.night_shift)
                    
                    # ====== الحصول على كود الوظيفة من قاعدة البيانات ======
                    # أولاً: البحث في قاموس job_codes_dict
                    job_code = job_codes_dict.get(detail.job_title, '')
                    
                    # إذا لم يتم العثور، استخدم الدالة القديمة كاحتياط
                    
                    # ====== FORMAT DATE AS dd/mm/yyyy ======
                    # Format: day/month/year (e.g., 20/02/2026)
                    formatted_date = detail.day_date.strftime('%d/%m/%Y')
                    
                    # Add data to Excel
                    worksheet.cell(row=row, column=1, value=detail.day_name)  # اليوم
                    worksheet.cell(row=row, column=2, value=formatted_date)  # التاريخ (dd/mm/yyyy)
                    
                    worksheet.cell(row=row, column=3, value=morning_username)  # كود موظف شفت A (username)
                    worksheet.cell(row=row, column=4, value=morning_name)  # الشيفت الصباحي (اسم الموظف)
                    
                    worksheet.cell(row=row, column=5, value=evening_username)  # كود موظف شفتB (username)
                    worksheet.cell(row=row, column=6, value=evening_name)  # الشيفت المسائي (اسم الموظف)
                    
                    worksheet.cell(row=row, column=7, value=night_username)  # كود موظف شفتC (username)
                    worksheet.cell(row=row, column=8, value=night_name)  # شيفت السهر (اسم الموظف)
                    
                    worksheet.cell(row=row, column=9, value=job_code)  # كود الوظيفة (من قاعدة البيانات)
                    worksheet.cell(row=row, column=10, value=detail.job_title)  # الوظيفة
                    worksheet.cell(row=row, column=11, value=department.name)  # القسم
                    worksheet.cell(row=row, column=12, value='')  # البيان (يمكن إضافته لاحقاً)
                    
                    # لا يحسب شفتA, شفتB, شفتC (كلها 0)
                    worksheet.cell(row=row, column=13, value=0)
                    worksheet.cell(row=row, column=14, value=0)
                    worksheet.cell(row=row, column=15, value=0)
                    
                    row += 1
            
            # ====== FORMATTING ======
            # Set column widths
            column_widths = {
                'A': 15,  # اليوم
                'B': 12,  # التاريخ
                'C': 15,  # كود موظف شفت A (username)
                'D': 20,  # الشيفت الصباحي (اسم الموظف)
                'E': 15,  # كود موظف شفتB (username)
                'F': 20,  # الشيفت المسائي (اسم الموظف)
                'G': 15,  # كود موظف شفتC (username)
                'H': 20,  # شيفت السهر (اسم الموظف)
                'I': 15,  # كود الوظيفة
                'J': 30,  # الوظيفة
                'K': 20,  # القسم
                'L': 20,  # البيان
                'M': 12,  # لا يحسب شفتA
                'N': 12,  # لا يحسب شفتB
                'O': 12   # لا يحسب شفتC
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Format date column as text to preserve the format
            for row_num in range(5, row):
                date_cell = worksheet.cell(row=row_num, column=2)
                date_cell.number_format = '@'  # Text format to keep as is
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            for row_num in range(4, row):
                for col_num in range(1, 16):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
            
            # Center align headers
            for col_num in range(1, 16):
                cell = worksheet.cell(row=4, column=col_num)
                cell.alignment = Alignment(horizontal='center')
        
        # Save workbook
        workbook.save(output)
        output.seek(0)
        
        # Generate filename
        dept_names = '_'.join([dept.name[:10] for dept in departments[:3]])
        if len(departments) > 3:
            dept_names += f'_و{len(departments)-3}أقسام'
        
        filename = f"جدول_العمل_الأسبوعي_{dept_names}_{start_date}_إلى_{end_date}.xlsx"
        
        # Check if it's an API request
        if request.method == 'GET' and request.args.get('format') == 'json':
            return jsonify({
                'success': True,
                'message': f'تم إنشاء ملف Excel بنجاح لـ {len(departments)} قسم',
                'filename': filename
            })
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error exporting schedule: {str(e)}")
        import traceback
        traceback.print_exc()
        
        if request.method == 'POST':
            flash(f'حدث خطأ أثناء تصدير الجدول: {str(e)}', 'error')
            return redirect(request.referrer or url_for('admin_weekly_schedules'))
        else:
            return jsonify({
                'success': False,
                'message': f'حدث خطأ: {str(e)}'
            })

def get_employee_username_from_shift(shift_text):
    """Get employee username from shift text by looking up in database"""
    if not shift_text or shift_text == '':
        return ''
    
    # Convert to string and strip
    text = str(shift_text).strip()
    
    # Try to find user by name in database
    # First, try to extract name if there's a code prefix
    import re
    match = re.match(r'^\d+\s+(.+)$', text)
    if match:
        employee_name = match.group(1).strip()
    else:
        employee_name = text
    
    # Search for user by name in the database
    user = db_session.query(User).filter(
        (User.name == employee_name) | 
        (User.name.ilike(f'%{employee_name}%'))
    ).first()
    
    if user:
        return user.username
    
    return ''


def extract_employee_name_from_shift(shift_text):
    """Extract employee name from shift text"""
    if not shift_text or shift_text == '':
        return ''
    
    # Convert to string and strip
    text = str(shift_text).strip()
    
    # Remove numeric code if present
    import re
    match = re.match(r'^\d+\s+(.+)$', text)
    if match:
        return match.group(1)
    
    return text




def extract_employee_code(shift_text):
    """Extract employee code from shift text (e.g., "463" from "463 اسم الموظف")"""
    if not shift_text or shift_text == '':
        return ''
    
    # Try to extract numeric code at the beginning
    import re
    match = re.match(r'^(\d+)', str(shift_text).strip())
    if match:
        return match.group(1)
    
    return ''


def extract_employee_name(shift_text):
    """Extract employee name from shift text (e.g., "اسم الموظف" from "463 اسم الموظف")"""
    if not shift_text or shift_text == '':
        return ''
    
    # Remove numeric code if present
    import re
    text = str(shift_text).strip()
    match = re.match(r'^\d+\s+(.+)$', text)
    if match:
        return match.group(1)
    
    return text


def find_similar_job_titles(job_title):
    """
    Find similar job titles in the JOB_CODES dictionary
    Useful for debugging
    """
    if not job_title:
        return []
    
    job_title = str(job_title).strip().lower()
    similar = []
    
    for key in JOB_CODES.keys():
        key_lower = key.lower()
        
        # Calculate similarity score
        score = 0
        
        # Check if one contains the other
        if key_lower in job_title:
            score += 10
        if job_title in key_lower:
            score += 10
        
        # Check word overlap
        key_words = set(key_lower.split())
        job_words = set(job_title.split())
        overlap = key_words.intersection(job_words)
        score += len(overlap) * 5
        
        # Check for common prefixes/suffixes
        common_prefix = 0
        for i in range(min(len(key_lower), len(job_title))):
            if key_lower[i] == job_title[i]:
                common_prefix += 1
            else:
                break
        score += common_prefix
        
        if score > 5:  # Threshold for similarity
            similar.append({
                'job_title': key,
                'code': JOB_CODES[key],
                'score': score
            })
    
    # Sort by score (highest first)
    similar.sort(key=lambda x: x['score'], reverse=True)
    return similar[:5]  # Return top 5 matches

# Add these imports at the top of your file with other openpyxl imports
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import unquote
import re

@app.route('/export/weekly_schedule_form')
@login_required
def export_weekly_schedule_form():
    """Display form for exporting weekly schedules"""
    
    # Get departments based on user role
    if current_user.is_admin:
        departments = db_session.query(Department).all()
    elif current_user.is_manager:
        managed_departments = db_session.query(Department).filter(
            (Department.primary_manager_id == current_user.id) | 
            (Department.id.in_(db_session.query(DepartmentManager.department_id).filter(
                DepartmentManager.user_id == current_user.id
            )))
        ).all()
        departments = managed_departments
    else:
        # Regular users can only see their own department
        departments = [db_session.query(Department).get(current_user.department_id)] if current_user.department_id else []
    
    # Get current date for default values
    today = date.today()
    
    # Calculate default start date (first day of current month)
    default_start_date = date(today.year, today.month, 1)
    
    # Calculate default end date (last day of current month)
    if today.month == 12:
        default_end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        default_end_date = date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('export_weekly_schedule_form.html',
                         departments=departments,
                         default_start_date=default_start_date.isoformat(),
                         default_end_date=default_end_date.isoformat(),
                         today=today.isoformat(),
                         notifications=notifications)

@app.route('/manager/copy_schedule/<int:source_schedule_id>', methods=['POST'])
@login_required
def copy_schedule(source_schedule_id):
    """Copy a schedule to a new week"""
    
    # Get source schedule
    source_schedule = db_session.query(WeeklySchedule).get(source_schedule_id)
    if not source_schedule:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'الجدول المصدر غير موجود'})
        flash('الجدول المصدر غير موجود', 'error')
        return redirect(url_for('manager_schedules'))
    
    # Check if manager manages this department
    department = db_session.query(Department).get(source_schedule.department_id)
    if not department:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
        flash('القسم غير موجود', 'error')
        return redirect(url_for('manager_schedules'))
    
    # Verify manager has access to this department
    if not current_user.is_admin:
        managed_dept_ids = get_managed_department_ids(current_user.id)
        if source_schedule.department_id not in managed_dept_ids:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'ليس لديك صلاحية للوصول إلى هذا الجدول'})
            flash('ليس لديك صلاحية للوصول إلى هذا الجدول', 'error')
            return redirect(url_for('manager_schedules'))
    
    try:
        # Get target week start date from form
        target_date_str = request.form.get('target_date')
        if not target_date_str:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'يرجى تحديد تاريخ بداية الأسبوع المستهدف'})
            flash('يرجى تحديد تاريخ بداية الأسبوع المستهدف', 'error')
            return redirect(url_for('manager_schedules'))
        
        target_start_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        
        # Ensure target date is a Saturday
        if target_start_date.weekday() != 5:  # Saturday = 5
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'يجب أن يكون تاريخ البداية يوم سبت'})
            flash('يجب أن يكون تاريخ البداية يوم سبت', 'error')
            return redirect(url_for('manager_schedules'))
        
        target_end_date = target_start_date + timedelta(days=6)
        
        # Check if schedule already exists for target week
        existing_schedule = db_session.query(WeeklySchedule).filter_by(
            department_id=source_schedule.department_id,
            week_start_date=target_start_date
        ).first()
        
        if existing_schedule:
            # If schedule exists, we'll replace it
            # Delete existing schedule details
            db_session.query(ScheduleDetail).filter_by(
                weekly_schedule_id=existing_schedule.id
            ).delete()
            
            # Update existing schedule
            target_schedule = existing_schedule
            target_schedule.updated_at = datetime.now()
            target_schedule.updated_by = current_user.id
            target_schedule.notes = f'تم النسخ من جدول الأسبوع {source_schedule.week_start_date.strftime("%Y-%m-%d")}'
        else:
            # Create new schedule
            week_number = target_start_date.isocalendar()[1]
            month = target_start_date.month
            year = target_start_date.year
            
            target_schedule = WeeklySchedule(
                department_id=source_schedule.department_id,
                week_start_date=target_start_date,
                week_end_date=target_end_date,
                week_number=week_number,
                month=month,
                year=year,
                created_by=current_user.id,
                is_approved=False,
                is_locked=False,
                is_template=False,
                notes=f'تم النسخ من جدول الأسبوع {source_schedule.week_start_date.strftime("%Y-%m-%d")}'
            )
            
            db_session.add(target_schedule)
            db_session.flush()  # Get ID for new schedule
        
        # Get source schedule details
        source_details = db_session.query(ScheduleDetail).filter_by(
            weekly_schedule_id=source_schedule_id
        ).order_by('day_date', 'row_order').all()
        
        # Calculate date difference between source and target weeks
        date_diff = target_start_date - source_schedule.week_start_date
        
        # Copy each detail with adjusted dates
        copied_count = 0
        for source_detail in source_details:
            # Calculate new date for this detail
            new_date = source_detail.day_date + date_diff
            
            # Get day name for the new date
            new_day_name = get_arabic_day_name(new_date)
            
            # Create new detail
            new_detail = ScheduleDetail(
                weekly_schedule_id=target_schedule.id,
                day_date=new_date,
                day_name=new_day_name,
                job_title=source_detail.job_title,
                morning_shift=source_detail.morning_shift,
                evening_shift=source_detail.evening_shift,
                night_shift=source_detail.night_shift,
                row_order=source_detail.row_order,
                is_custom=source_detail.is_custom,
                notes=source_detail.notes
            )
            
            db_session.add(new_detail)
            copied_count += 1
        
        db_session.commit()
        
        # Send notification
        create_notification(
            current_user.id,
            'تم نسخ الجدول بنجاح',
            f'تم نسخ {copied_count} تفصيلة من جدول {source_schedule.week_start_date.strftime("%Y-%m-%d")} إلى {target_start_date.strftime("%Y-%m-%d")}',
            'schedule_copied',
            related_id=target_schedule.id,
            action_url=url_for('manager_edit_schedule', schedule_id=target_schedule.id)
        )
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': f'تم نسخ الجدول بنجاح مع {copied_count} تفصيلة',
                'redirect_url': url_for('manager_edit_schedule', schedule_id=target_schedule.id)
            })
        
        flash(f'تم نسخ الجدول بنجاح مع {copied_count} تفصيلة', 'success')
        return redirect(url_for('manager_edit_schedule', schedule_id=target_schedule.id))
        
    except Exception as e:
        db_session.rollback()
        print(f"Error copying schedule: {str(e)}")
        import traceback
        traceback.print_exc()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})
        
        flash(f'حدث خطأ أثناء نسخ الجدول: {str(e)}', 'error')
        return redirect(url_for('manager_schedules'))


# Add to your app.py

@app.route('/api/send_telegram_notification', methods=['POST'])
@login_required
def send_telegram_notification():
    """Send notification to Telegram bot"""
    try:
        data = request.json
        notification_id = data.get('notification_id')
        
        if not notification_id:
            return jsonify({'success': False, 'message': 'notification_id required'})
        
        # Send to Telegram bot webhook
        response = requests.post(
            'http://192.168.2.70:5002/webhook/notification',  # Your bot webhook URL
            json={'notification_id': notification_id},
            timeout=5
        )
        
        if response.status_code == 200:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': response.text})
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Add these imports at the top of your app.py if not already present
import requests
import threading
from flask import jsonify, request

# Add this configuration near your other app.config settings
app.config['TELEGRAM_BOT_URL'] = 'http://localhost:5002'  # Your Telegram bot webhook URL

def send_to_telegram(notification_id):
    """Send notification to Telegram bot (runs in background)"""
    try:
        response = requests.post(
            f"{app.config['TELEGRAM_BOT_URL']}/webhook/notification",
            json={'notification_id': notification_id},
            timeout=3
        )
        if response.status_code == 200:
            app.logger.info(f"Notification {notification_id} sent to Telegram")
        else:
            app.logger.warning(f"Failed to send notification {notification_id} to Telegram: {response.text}")
    except requests.exceptions.ConnectionError:
        app.logger.warning(f"Telegram bot not reachable at {app.config['TELEGRAM_BOT_URL']}")
    except Exception as e:
        app.logger.error(f"Error sending to Telegram: {str(e)}")

# REPLACE your existing create_notification function with this enhanced version
def create_notification(user_id, title, message, notification_type='general', related_id=None, action_url=None):
    """
    Create notification and send to both database and Telegram
    """
    # Create notification in database
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        notification_type=notification_type,
        related_id=related_id,
        action_url=action_url
    )
    db_session.add(notification)
    db_session.commit()
    
    # Send to Telegram in background thread (non-blocking)
    try:
        thread = threading.Thread(target=send_to_telegram, args=(notification.id,))
        thread.daemon = True  # Thread will exit when main program exits
        thread.start()
        app.logger.debug(f"Started Telegram notification thread for notification {notification.id}")
    except Exception as e:
        app.logger.error(f"Failed to start Telegram thread: {str(e)}")
    
    return notification

# Add this new endpoint to manually trigger Telegram notifications (for testing)
@app.route('/api/test_telegram/<int:notification_id>')
@login_required
def test_telegram(notification_id):
    """Test sending a specific notification to Telegram"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    notification = db_session.query(Notification).get(notification_id)
    if not notification:
        return jsonify({'success': False, 'message': 'Notification not found'})
    
    # Send to Telegram
    send_to_telegram(notification_id)
    
    return jsonify({
        'success': True,
        'message': f'Notification {notification_id} sent to Telegram',
        'notification': {
            'id': notification.id,
            'title': notification.title,
            'user_id': notification.user_id
        }
    })

# Add this endpoint to broadcast a message to all Telegram users
@app.route('/api/broadcast_telegram', methods=['POST'])
@login_required
def broadcast_telegram():
    """Broadcast a message to all connected Telegram users"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    data = request.json
    message = data.get('message', '')
    
    if not message:
        return jsonify({'success': False, 'message': 'Message is required'})
    
    try:
        response = requests.post(
            f"{app.config['TELEGRAM_BOT_URL']}/webhook/broadcast",
            json={'message': message, 'type': 'broadcast'},
            timeout=10
        )
        
        if response.status_code == 200:
            result = response.json()
            return jsonify({
                'success': True,
                'message': f'Broadcast sent to {result.get("sent_count", 0)} users',
                'details': result
            })
        else:
            return jsonify({'success': False, 'message': 'Telegram bot error'})
            
    except requests.exceptions.ConnectionError:
        return jsonify({'success': False, 'message': 'Telegram bot not reachable'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Add this endpoint to check Telegram bot status
@app.route('/api/telegram_status')
@login_required
def telegram_status():
    """Check if Telegram bot is running"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        response = requests.get(f"{app.config['TELEGRAM_BOT_URL']}/health", timeout=3)
        if response.status_code == 200:
            return jsonify({'success': True, 'status': 'online', 'details': response.json()})
        else:
            return jsonify({'success': False, 'status': 'error'})
    except requests.exceptions.ConnectionError:
        return jsonify({'success': False, 'status': 'offline'})
    except Exception as e:
        return jsonify({'success': False, 'status': 'error', 'message': str(e)})



@app.route('/admin/telegram')
@login_required
def admin_telegram():
    """Telegram integration management page with employee data"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    # Get all users with their phone numbers from employee_data
    users_with_phones = db_session.query(
        User.id,
        User.name,
        User.username,
        EmployeeData.phone,
        EmployeeData.whatsapp
    ).outerjoin(
        EmployeeData, User.id == EmployeeData.user_id
    ).filter(
        User.is_admin == False  # Exclude admins
    ).all()
    
    # Get Telegram mappings
    try:
        # Using SQLAlchemy to query telegram_mapping if table exists
        from sqlalchemy import text
        result = db_session.execute(text("""
            SELECT tm.*, u.name as user_name, u.username 
            FROM telegram_mapping tm
            JOIN users u ON tm.user_id = u.id
            ORDER BY tm.verified_at DESC
        """))
        mappings = result.fetchall()
        
        # Convert to list of dicts
        mappings_list = []
        for row in mappings:
            mappings_list.append({
                'id': row[0],
                'user_id': row[1],
                'telegram_id': row[2],
                'phone_number': row[3],
                'verified_at': row[4],
                'is_active': row[5],
                'user_name': row[7],
                'username': row[8]
            })
    except Exception as e:
        print(f"Error getting mappings: {e}")
        mappings_list = []
    
    # Get statistics
    try:
        result = db_session.execute(text("""
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN is_active = 1 THEN 1 ELSE 0 END) as active
            FROM telegram_mapping
        """))
        stats = result.fetchone()
        active_users = stats[1] if stats and stats[1] else 0
        total_users = stats[0] if stats else 0
    except:
        active_users = 0
        total_users = 0
    
    # Prepare users data for display
    users_data = []
    for user in users_with_phones:
        # Check if this user is already mapped
        is_mapped = any(m.get('user_id') == user.id for m in mappings_list)
        phone = user.phone or user.whatsapp or 'لم يضف رقم'
        
        users_data.append({
            'id': user.id,
            'name': user.name,
            'username': user.username,
            'phone': phone,
            'has_phone': bool(user.phone or user.whatsapp),
            'is_mapped': is_mapped
        })
    
    # Check bot status
    bot_status = 'unknown'
    try:
        response = requests.get(f"{app.config['TELEGRAM_BOT_URL']}/health", timeout=2)
        bot_status = 'online' if response.status_code == 200 else 'offline'
    except:
        bot_status = 'offline'
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_telegram.html',
                         users=users_data,
                         mappings=mappings_list,
                         active_users=active_users,
                         total_users=total_users,
                         bot_status=bot_status,
                         bot_url=app.config['TELEGRAM_BOT_URL'],
                         notifications=notifications)


# ======== Delete Operations Based on Date ========

def can_delete_based_on_date(item_date):
    """
    التحقق مما إذا كان يمكن حذف عنصر بناءً على تاريخه
    القاعدة: يمكن الحذف فقط إذا كان التاريخ <= 25 من نفس الشهر
    أي قبل يوم 26
    """
    if not item_date:
        return False
    
    # تحويل التاريخ إذا كان نصياً
    if isinstance(item_date, str):
        try:
            item_date = datetime.strptime(item_date, '%Y-%m-%d').date()
        except:
            return False
    
    today = date.today()
    
    # التحقق من أن التاريخ في نفس الشهر
    if item_date.month != today.month or item_date.year != today.year:
        return False
    
    # يمكن الحذف فقط إذا كان اليوم <= 25
    return today.day <= 25

def get_deletion_period_message():
    """الحصول على رسالة توضح فترة الحذف المسموح بها"""
    today = date.today()
    if today.day <= 25:
        return f"يمكنك حذف المعاملات حتى يوم 25 من الشهر الحالي (اليوم: {today.day})"
    else:
        return f"لا يمكن حذف المعاملات بعد يوم 25 من الشهر (اليوم: {today.day})، انتظر حتى الشهر القادم"

# ======== Delete Reward/Penalty ========

@app.route('/manager/delete_reward/<int:reward_id>', methods=['POST'])
@login_required
def manager_delete_reward(reward_id):
    """حذف مكافأة أو خصم - فقط إذا كان ضمن الفترة المسموح بها"""
    
    try:
        # الحصول على المكافأة/الخصم
        reward_penalty = db_session.query(RewardPenalty).get(reward_id)
        if not reward_penalty:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'العنصر غير موجود'})
            flash('العنصر غير موجود', 'error')
            return redirect(url_for('manager_rewards_penalties'))
        
        # التحقق من أن المدير مسؤول عن القسم
        department = db_session.query(Department).get(reward_penalty.department_id)
        if not department:
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
        
        # التحقق من صلاحية المستخدم
        if not current_user.is_admin:
            managed_dept_ids = get_managed_department_ids(current_user.id)
            if reward_penalty.department_id not in managed_dept_ids:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'message': 'ليس لديك صلاحية لحذف هذا العنصر'})
                flash('ليس لديك صلاحية لحذف هذا العنصر', 'error')
                return redirect(url_for('manager_rewards_penalties'))
        
        # التحقق من إمكانية الحذف بناءً على التاريخ
        effective_date = reward_penalty.effective_date
        if not can_delete_based_on_date(effective_date):
            message = f'لا يمكن حذف هذا العنصر لأنه مضاف قبل يوم 26 من الشهر (تاريخ التطبيق: {effective_date.strftime("%Y-%m-%d")})'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': message})
            flash(message, 'error')
            return redirect(url_for('manager_rewards_penalties'))
        
        # حفظ معلومات للإشعار
        user_id = reward_penalty.user_id
        user = db_session.query(User).get(user_id)
        type_arabic = 'مكافأة' if reward_penalty.type == 'reward' else 'خصم'
        amount = reward_penalty.amount
        
        # حذف العنصر
        db_session.delete(reward_penalty)
        db_session.commit()
        
        # إرسال إشعار للموظف
        create_notification(
            user_id,
            f'تم حذف {type_arabic}',
            f'تم حذف {type_arabic} بقيمة {amount} جنيه (تاريخ التطبيق: {effective_date.strftime("%Y-%m-%d")})',
            'reward_deleted',
            action_url=url_for('user_rewards_penalties')
        )
        
        # إرسال إشعار للمدير
        create_notification(
            current_user.id,
            f'تم حذف {type_arabic}',
            f'تم حذف {type_arabic} للموظف {user.name} بقيمة {amount} جنيه',
            'reward_deleted',
            action_url=url_for('manager_rewards_penalties')
        )
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': f'تم حذف {type_arabic} بنجاح'
            })
        
        flash(f'تم حذف {type_arabic} بنجاح', 'success')
        return redirect(url_for('manager_rewards_penalties'))
        
    except Exception as e:
        db_session.rollback()
        print(f"Error deleting reward/penalty: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})
        flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('manager_rewards_penalties'))

# ======== Delete Leave Request ========

@app.route('/manager/delete_leave/<int:leave_id>', methods=['POST'])
@login_required
def manager_delete_leave(leave_id):
    """حذف طلب إجازة - فقط إذا كان ضمن الفترة المسموح بها"""
    
    try:
        # الحصول على طلب الإجازة
        leave_request = db_session.query(LeaveRequest).get(leave_id)
        if not leave_request:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'طلب الإجازة غير موجود'})
            flash('طلب الإجازة غير موجود', 'error')
            return redirect(url_for('manager_leave_requests'))
        
        # التحقق من أن المدير مسؤول عن القسم
        department = db_session.query(Department).get(leave_request.department_id)
        if not department:
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
        
        # التحقق من صلاحية المستخدم
        if not current_user.is_admin:
            managed_dept_ids = get_managed_department_ids(current_user.id)
            if leave_request.department_id not in managed_dept_ids:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'message': 'ليس لديك صلاحية لحذف هذا الطلب'})
                flash('ليس لديك صلاحية لحذف هذا الطلب', 'error')
                return redirect(url_for('manager_leave_requests'))
        
        # تحديد تاريخ الإجازة (استخدم leave_date إذا كان موجوداً، وإلا استخدم start_date)
        leave_date = leave_request.leave_date if hasattr(leave_request, 'leave_date') and leave_request.leave_date else leave_request.start_date
        
        # التحقق من إمكانية الحذف بناءً على التاريخ
        if not can_delete_based_on_date(leave_date):
            message = f'لا يمكن حذف هذا الطلب لأنه مضاف قبل يوم 26 من الشهر (تاريخ الإجازة: {leave_date.strftime("%Y-%m-%d")})'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': message})
            flash(message, 'error')
            return redirect(url_for('manager_leave_requests'))
        
        # حفظ معلومات للإشعار
        user_id = leave_request.user_id
        user = db_session.query(User).get(user_id)
        leave_type = leave_request.leave_type
        shift_name = getattr(leave_request, 'shift_name', '')
        
        # إذا كانت الإجازة من رصيد الإجازات، أعد الرصيد للموظف
        balance_updated = False
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(user_id=user_id).first()
            if balance:
                days_to_return = leave_request.total_days if leave_request.total_days else 1
                balance.leave_balance += days_to_return
                balance.last_updated = datetime.now()
                balance_updated = True
        
        # حذف الطلب
        db_session.delete(leave_request)
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'تم حذف طلب الإجازة ({leave_type})'
        if shift_name:
            notification_message += f' للشيفت {shift_name}'
        if balance_updated:
            notification_message += f' وتم إعادة {days_to_return} يوم إلى رصيدك'
        
        create_notification(
            user_id,
            'تم حذف طلب إجازة',
            notification_message,
            'leave_deleted',
            action_url=url_for('user_leave_requests')
        )
        
        # إرسال إشعار للمدير
        create_notification(
            current_user.id,
            'تم حذف طلب إجازة',
            f'تم حذف طلب إجازة للموظف {user.name} ({leave_type})',
            'leave_deleted',
            action_url=url_for('manager_leave_requests')
        )
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': f'تم حذف طلب الإجازة بنجاح{" وإعادة الرصيد للموظف" if balance_updated else ""}'
            })
        
        flash(f'تم حذف طلب الإجازة بنجاح{" وإعادة الرصيد للموظف" if balance_updated else ""}', 'success')
        return redirect(url_for('manager_leave_requests'))
        
    except Exception as e:
        db_session.rollback()
        print(f"Error deleting leave request: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})
        flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('manager_leave_requests'))

# ======== Delete Permission Request ========

@app.route('/manager/delete_permission/<int:permission_id>', methods=['POST'])
@login_required
def manager_delete_permission(permission_id):
    """حذف طلب إذن - فقط إذا كان ضمن الفترة المسموح بها"""
    
    try:
        # الحصول على طلب الإذن
        permission_request = db_session.query(PermissionRequest).get(permission_id)
        if not permission_request:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'طلب الإذن غير موجود'})
            flash('طلب الإذن غير موجود', 'error')
            return redirect(url_for('manager_permission_requests'))
        
        # التحقق من أن المدير مسؤول عن القسم
        department = db_session.query(Department).get(permission_request.department_id)
        if not department:
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
        
        # التحقق من صلاحية المستخدم
        if not current_user.is_admin:
            managed_dept_ids = get_managed_department_ids(current_user.id)
            if permission_request.department_id not in managed_dept_ids:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'message': 'ليس لديك صلاحية لحذف هذا الطلب'})
                flash('ليس لديك صلاحية لحذف هذا الطلب', 'error')
                return redirect(url_for('manager_permission_requests'))
        
        # التحقق من إمكانية الحذف بناءً على التاريخ
        permission_date = permission_request.date
        if not can_delete_based_on_date(permission_date):
            message = f'لا يمكن حذف هذا الطلب لأنه مضاف قبل يوم 26 من الشهر (تاريخ الإذن: {permission_date.strftime("%Y-%m-%d")})'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': message})
            flash(message, 'error')
            return redirect(url_for('manager_permission_requests'))
        
        # حفظ معلومات للإشعار
        user_id = permission_request.user_id
        user = db_session.query(User).get(user_id)
        permission_type = permission_request.permission_type
        
        # إذا كان الإذن من الرصيد العادي (غير الساعات الإضافية وتبديل الورديات)، أعد الرصيد
        balance_updated = False
        if 'ساعات اضافي' not in permission_type and 'تبديل' not in permission_type:
            balance = db_session.query(EmployeeBalance).filter_by(user_id=user_id).first()
            if balance:
                balance.permission_balance += 1
                balance.last_updated = datetime.now()
                balance_updated = True
        
        # حذف الطلب
        db_session.delete(permission_request)
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'تم حذف طلب الإذن ({permission_type})'
        if balance_updated:
            notification_message += f' وتم إعادة رصيد الإذن'
        
        create_notification(
            user_id,
            'تم حذف طلب إذن',
            notification_message,
            'permission_deleted',
            action_url=url_for('user_leave_requests')
        )
        
        # إرسال إشعار للمدير
        create_notification(
            current_user.id,
            'تم حذف طلب إذن',
            f'تم حذف طلب إذن للموظف {user.name} ({permission_type})',
            'permission_deleted',
            action_url=url_for('manager_permission_requests')
        )
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': f'تم حذف طلب الإذن بنجاح{" وإعادة الرصيد للموظف" if balance_updated else ""}'
            })
        
        flash(f'تم حذف طلب الإذن بنجاح{" وإعادة الرصيد للموظف" if balance_updated else ""}', 'success')
        return redirect(url_for('manager_permission_requests'))
        
    except Exception as e:
        db_session.rollback()
        print(f"Error deleting permission request: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})
        flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('manager_permission_requests'))

# ======== Delete Operations for Employees (Self-Service) ========

@app.route('/user/delete_own_leave/<int:leave_id>', methods=['POST'])
@login_required
def user_delete_own_leave(leave_id):
    """حذف طلب إجازة خاص بالموظف - فقط إذا كان ضمن الفترة المسموح بها"""
    
    try:
        # الحصول على طلب الإجازة
        leave_request = db_session.query(LeaveRequest).get(leave_id)
        if not leave_request:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'طلب الإجازة غير موجود'})
            flash('طلب الإجازة غير موجود', 'error')
            return redirect(url_for('user_leave_requests'))
        
        # التحقق من أن الطلب يخص الموظف الحالي
        if leave_request.user_id != current_user.id:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'ليس لديك صلاحية لحذف هذا الطلب'})
            flash('ليس لديك صلاحية لحذف هذا الطلب', 'error')
            return redirect(url_for('user_leave_requests'))
        
        # تحديد تاريخ الإجازة
        leave_date = leave_request.leave_date if hasattr(leave_request, 'leave_date') and leave_request.leave_date else leave_request.start_date
        
        # التحقق من إمكانية الحذف بناءً على التاريخ
        if not can_delete_based_on_date(leave_date):
            message = f'لا يمكن حذف هذا الطلب لأنه مضاف قبل يوم 26 من الشهر (تاريخ الإجازة: {leave_date.strftime("%Y-%m-%d")})'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': message})
            flash(message, 'error')
            return redirect(url_for('user_leave_requests'))
        
        # إذا كانت الإجازة من رصيد الإجازات، أعد الرصيد
        balance_updated = False
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(user_id=current_user.id).first()
            if balance:
                days_to_return = leave_request.total_days if leave_request.total_days else 1
                balance.leave_balance += days_to_return
                balance.last_updated = datetime.now()
                balance_updated = True
        
        # حذف الطلب
        db_session.delete(leave_request)
        db_session.commit()
        
        # إرسال إشعار للمدير
        if leave_request.department_id:
            department = db_session.query(Department).get(leave_request.department_id)
            if department and department.primary_manager_id:
                create_notification(
                    department.primary_manager_id,
                    'موظف حذف طلب إجازة',
                    f'قام الموظف {current_user.name} بحذف طلب إجازة ({leave_request.leave_type})',
                    'leave_deleted_by_employee',
                    action_url=url_for('manager_leave_requests')
                )
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': f'تم حذف طلب الإجازة بنجاح{" وإعادة الرصيد" if balance_updated else ""}'
            })
        
        flash(f'تم حذف طلب الإجازة بنجاح{" وإعادة الرصيد" if balance_updated else ""}', 'success')
        return redirect(url_for('user_leave_requests'))
        
    except Exception as e:
        db_session.rollback()
        print(f"Error deleting own leave: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})
        flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('user_leave_requests'))

@app.route('/user/delete_own_permission/<int:permission_id>', methods=['POST'])
@login_required
def user_delete_own_permission(permission_id):
    """حذف طلب إذن خاص بالموظف - فقط إذا كان ضمن الفترة المسموح بها"""
    
    try:
        # الحصول على طلب الإذن
        permission_request = db_session.query(PermissionRequest).get(permission_id)
        if not permission_request:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'طلب الإذن غير موجود'})
            flash('طلب الإذن غير موجود', 'error')
            return redirect(url_for('user_leave_requests'))
        
        # التحقق من أن الطلب يخص الموظف الحالي
        if permission_request.user_id != current_user.id:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'ليس لديك صلاحية لحذف هذا الطلب'})
            flash('ليس لديك صلاحية لحذف هذا الطلب', 'error')
            return redirect(url_for('user_leave_requests'))
        
        # التحقق من إمكانية الحذف بناءً على التاريخ
        permission_date = permission_request.date
        if not can_delete_based_on_date(permission_date):
            message = f'لا يمكن حذف هذا الطلب لأنه مضاف قبل يوم 26 من الشهر (تاريخ الإذن: {permission_date.strftime("%Y-%m-%d")})'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': message})
            flash(message, 'error')
            return redirect(url_for('user_leave_requests'))
        
        # إذا كان الإذن من الرصيد العادي، أعد الرصيد
        balance_updated = False
        if 'ساعات اضافي' not in permission_request.permission_type and 'تبديل' not in permission_request.permission_type:
            balance = db_session.query(EmployeeBalance).filter_by(user_id=current_user.id).first()
            if balance:
                balance.permission_balance += 1
                balance.last_updated = datetime.now()
                balance_updated = True
        
        # حذف الطلب
        db_session.delete(permission_request)
        db_session.commit()
        
        # إرسال إشعار للمدير
        if permission_request.department_id:
            department = db_session.query(Department).get(permission_request.department_id)
            if department and department.primary_manager_id:
                create_notification(
                    department.primary_manager_id,
                    'موظف حذف طلب إذن',
                    f'قام الموظف {current_user.name} بحذف طلب إذن ({permission_request.permission_type})',
                    'permission_deleted_by_employee',
                    action_url=url_for('manager_permission_requests')
                )
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': f'تم حذف طلب الإذن بنجاح{" وإعادة الرصيد" if balance_updated else ""}'
            })
        
        flash(f'تم حذف طلب الإذن بنجاح{" وإعادة الرصيد" if balance_updated else ""}', 'success')
        return redirect(url_for('user_leave_requests'))
        
    except Exception as e:
        db_session.rollback()
        print(f"Error deleting own permission: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})
        flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('user_leave_requests'))

# ======== Bulk Delete Operations ========

@app.route('/manager/bulk_delete_items', methods=['POST'])
@login_required
def manager_bulk_delete_items():
    """حذف مجموعة من العناصر دفعة واحدة (مع التحقق من التاريخ)"""
    
    try:
        data = request.get_json()
        item_type = data.get('type')  # 'rewards', 'leaves', 'permissions'
        item_ids = data.get('ids', [])
        
        if not item_type or not item_ids:
            return jsonify({'success': False, 'message': 'بيانات غير كاملة'})
        
        # التحقق من صلاحية المستخدم
        if not current_user.is_admin and not current_user.is_manager:
            return jsonify({'success': False, 'message': 'ليس لديك صلاحية'})
        
        deleted_count = 0
        failed_count = 0
        results = []
        
        for item_id in item_ids:
            try:
                if item_type == 'rewards':
                    # حذف مكافأة/خصم
                    reward = db_session.query(RewardPenalty).get(item_id)
                    if reward and can_delete_based_on_date(reward.effective_date):
                        db_session.delete(reward)
                        deleted_count += 1
                        results.append({'id': item_id, 'success': True})
                    else:
                        failed_count += 1
                        results.append({'id': item_id, 'success': False, 'reason': 'تاريخ غير مسموح'})
                
                elif item_type == 'leaves':
                    # حذف إجازة
                    leave = db_session.query(LeaveRequest).get(item_id)
                    if leave:
                        leave_date = leave.leave_date if hasattr(leave, 'leave_date') and leave.leave_date else leave.start_date
                        if can_delete_based_on_date(leave_date):
                            # إذا كانت من الرصيد، أعد الرصيد
                            if leave.leave_type == 'من رصيد الإجازات':
                                balance = db_session.query(EmployeeBalance).filter_by(user_id=leave.user_id).first()
                                if balance:
                                    balance.leave_balance += leave.total_days if leave.total_days else 1
                            
                            db_session.delete(leave)
                            deleted_count += 1
                            results.append({'id': item_id, 'success': True})
                        else:
                            failed_count += 1
                            results.append({'id': item_id, 'success': False, 'reason': 'تاريخ غير مسموح'})
                    else:
                        failed_count += 1
                        results.append({'id': item_id, 'success': False, 'reason': 'غير موجود'})
                
                elif item_type == 'permissions':
                    # حذف إذن
                    permission = db_session.query(PermissionRequest).get(item_id)
                    if permission and can_delete_based_on_date(permission.date):
                        # إذا كان من الرصيد العادي، أعد الرصيد
                        if 'ساعات اضافي' not in permission.permission_type and 'تبديل' not in permission.permission_type:
                            balance = db_session.query(EmployeeBalance).filter_by(user_id=permission.user_id).first()
                            if balance:
                                balance.permission_balance += 1
                        
                        db_session.delete(permission)
                        deleted_count += 1
                        results.append({'id': item_id, 'success': True})
                    else:
                        failed_count += 1
                        results.append({'id': item_id, 'success': False, 'reason': 'تاريخ غير مسموح' if permission else 'غير موجود'})
                
            except Exception as e:
                failed_count += 1
                results.append({'id': item_id, 'success': False, 'reason': str(e)})
        
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم حذف {deleted_count} عنصر، فشل {failed_count} عنصر',
            'deleted': deleted_count,
            'failed': failed_count,
            'details': results
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

# ======== Helper Route to Check Deletion Status ========

@app.route('/api/check_deletion_status')
@login_required
def api_check_deletion_status():
    """API للتحقق من حالة الحذف (هل اليوم <= 25)"""
    today = date.today()
    can_delete = today.day <= 25
    
    # حساب الأيام المتبقية
    days_left = 25 - today.day if can_delete else 0
    
    # حساب تاريخ بداية الفترة القادمة
    next_month = today.month + 1 if today.month < 12 else 1
    next_year = today.year if today.month < 12 else today.year + 1
    next_period_start = date(next_year, next_month, 26)
    
    return jsonify({
        'success': True,
        'can_delete': can_delete,
        'today': today.strftime('%Y-%m-%d'),
        'day_of_month': today.day,
        'days_left': days_left,
        'message': get_deletion_period_message(),
        'next_period_start': next_period_start.strftime('%Y-%m-%d') if not can_delete else None
    })

# ======== Update Context Processor to Include Deletion Info ========

@app.context_processor
def inject_deletion_info():
    """حقن معلومات الحذف في جميع القوالب"""
    today = date.today()
    can_delete = today.day <= 25
    
    return {
        'can_delete_today': can_delete,
        'deletion_message': get_deletion_period_message(),
        'deletion_deadline': 25,
        'today_day': today.day
    }

SPECIAL_USERS = ['98', '510', '100']


# ======== Export Leave Report ========

@app.route('/export/leave_report_form', methods=['GET'])
@login_required
def export_leave_report_form():
    """Display form for exporting leave report"""
    # Determine departments the user can see
    if current_user.is_admin:
        departments = db_session.query(Department).all()
    elif current_user.is_manager:
        managed_dept_ids = get_managed_department_ids(current_user.id)
        departments = db_session.query(Department).filter(Department.id.in_(managed_dept_ids)).all()
    else:
        departments = [db_session.query(Department).get(current_user.department_id)] if current_user.department_id else []
    
    # Get employees for filter (optional)
    employees = db_session.query(User).filter(User.is_admin == False).order_by(User.name).all()
    
    # Leave types for filter
    leave_types = db_session.query(LeaveRequest.leave_type).distinct().all()
    leave_types = [lt[0] for lt in leave_types if lt[0]]
    
    # Statuses
    statuses = ['pending', 'approved', 'rejected']
    
    # Default dates: current month
    today = date.today()
    default_start = date(today.year, today.month, 1)
    if today.month == 12:
        default_end = date(today.year, 12, 31)
    else:
        default_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('export_leave_report_form.html',
                         departments=departments,
                         employees=employees,
                         leave_types=leave_types,
                         statuses=statuses,
                         default_start=default_start.isoformat(),
                         default_end=default_end.isoformat(),
                         notifications=notifications)


@app.route('/export/leave_report', methods=['GET', 'POST'])
@login_required
def export_leave_report():
    """Export leave report to Excel"""
    try:
        # Get filter parameters
        if request.method == 'POST':
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            department_id = request.form.get('department_id')
            employee_id = request.form.get('employee_id')
            leave_type = request.form.get('leave_type')
            status = request.form.get('status')
        else:
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            department_id = request.args.get('department_id')
            employee_id = request.args.get('employee_id')
            leave_type = request.args.get('leave_type')
            status = request.args.get('status')
        
        # Build base query
        query = db_session.query(LeaveRequest).options(
            joinedload(LeaveRequest.user),
            joinedload(LeaveRequest.department),
            joinedload(LeaveRequest.approver)
        )
        
        # Apply filters
        if start_date:
            query = query.filter(LeaveRequest.start_date >= start_date)
        if end_date:
            query = query.filter(LeaveRequest.start_date <= end_date)
        if department_id and department_id != 'all':
            query = query.filter(LeaveRequest.department_id == department_id)
        if employee_id and employee_id != 'all':
            query = query.filter(LeaveRequest.user_id == employee_id)
        if leave_type and leave_type != 'all':
            query = query.filter(LeaveRequest.leave_type == leave_type)
        if status and status != 'all':
            query = query.filter(LeaveRequest.status == status)
        
        # Restrict to user's accessible departments
        if not current_user.is_admin:
            accessible_dept_ids = get_managed_department_ids(current_user.id)
            if current_user.department_id:
                accessible_dept_ids.append(current_user.department_id)
            query = query.filter(LeaveRequest.department_id.in_(accessible_dept_ids))
        
        leave_requests = query.order_by(LeaveRequest.start_date.desc()).all()
        
        # Create Excel file
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "تقرير الإجازات"
        
        # Title
        title = f"تقرير الإجازات - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        if start_date and end_date:
            title += f" (من {start_date} إلى {end_date})"
        worksheet.merge_cells('A1:N1')
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = [
            'كود الموظف', 'اسم الموظف', 'كود الوظيفة', 'القسم',
            'كود الاجازة', 'نوع الإجازة', 'التاريخ', 'نصف يوم',
            'عدد الأيام', 'شفتA', 'شفتB', 'شفتC', 'الحالة', 'ملاحظات'
        ]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = 4
        arabic_status = {'approved': 'مقبولة', 'pending': 'معلقة', 'rejected': 'مرفوضة'}
        
        for req in leave_requests:
            # Employee info
            user = req.user
            emp_data = db_session.query(EmployeeData).filter_by(user_id=user.id).first() if user else None
            
            worksheet.cell(row=row, column=1, value=user.username if user else '')
            worksheet.cell(row=row, column=2, value=user.name if user else '')
            worksheet.cell(row=row, column=3, value=emp_data.job_title if emp_data and emp_data.job_title else '')
            worksheet.cell(row=row, column=4, value=req.department.name if req.department else '')
            
            # Leave code – maybe use a fixed mapping or leave blank
            worksheet.cell(row=row, column=5, value='')  # كود الاجازة (optional)
            worksheet.cell(row=row, column=6, value=req.leave_type or '')
            
            # التاريخ – use start_date or leave_date if shift-based
            leave_date = req.leave_date if req.leave_date else req.start_date
            worksheet.cell(row=row, column=7, value=leave_date.strftime('%Y-%m-%d') if leave_date else '')
            
            # نصف يوم – determine if half day (total_days < 1 or shift_name exists)
            is_half = req.total_days and req.total_days < 1
            worksheet.cell(row=row, column=8, value='نعم' if is_half else '')
            
            # عدد الأيام
            worksheet.cell(row=row, column=9, value=req.total_days if req.total_days else 1)
            
            # Shift columns – fill based on shift_name if available
            shift_col = 10  # شفتA
            if req.shift_name:
                # Assume shift_name is something like "صباحي", "مسائي", "سهر"
                # We'll place it in the appropriate column, but here just put in شفتA
                worksheet.cell(row=row, column=shift_col, value=req.shift_name)
                if req.shift_job:
                    worksheet.cell(row=row, column=shift_col+1, value=req.shift_job)  # maybe شفتB gets job? unclear
            else:
                # If multiple shifts, they would be separate requests with same parent_request_id
                # We'll leave columns blank
                pass
            
            # الحالة
            worksheet.cell(row=row, column=13, value=arabic_status.get(req.status, req.status))
            
            # ملاحظات
            worksheet.cell(row=row, column=14, value=req.reason or '')
            
            row += 1
        
        # Adjust column widths
        for col_num, col_letter in enumerate(['A','B','C','D','E','F','G','H','I','J','K','L','M','N'], 1):
            worksheet.column_dimensions[col_letter].width = 15
        
        workbook.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"تقرير_الإجازات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        print(f"Error exporting leave report: {str(e)}")
        import traceback
        traceback.print_exc()
        if request.method == 'POST':
            flash(f'حدث خطأ أثناء تصدير التقرير: {str(e)}', 'error')
            return redirect(url_for('export_leave_report_form'))
        else:
            return jsonify({'success': False, 'message': str(e)})


# ======== Export Rewards & Penalties Report ========

@app.route('/export/reward_penalty_form', methods=['GET'])
@login_required
def export_reward_penalty_form():
    """Display form for exporting rewards & penalties report"""
    # Determine accessible departments
    if current_user.is_admin:
        departments = db_session.query(Department).all()
    elif current_user.is_manager:
        managed_dept_ids = get_managed_department_ids(current_user.id)
        departments = db_session.query(Department).filter(Department.id.in_(managed_dept_ids)).all()
    else:
        departments = [db_session.query(Department).get(current_user.department_id)] if current_user.department_id else []
    
    # Employees for filter
    employees = db_session.query(User).filter(User.is_admin == False).order_by(User.name).all()
    
    # Default dates: current month
    today = date.today()
    default_start = date(today.year, today.month, 1)
    if today.month == 12:
        default_end = date(today.year, 12, 31)
    else:
        default_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('export_reward_penalty_form.html',
                         departments=departments,
                         employees=employees,
                         default_start=default_start.isoformat(),
                         default_end=default_end.isoformat(),
                         notifications=notifications)


@app.route('/export/reward_penalty', methods=['GET', 'POST'])
@login_required
def export_reward_penalty():
    """Export rewards and penalties report to Excel"""
    try:
        # Get parameters
        if request.method == 'POST':
            department_ids = request.form.getlist('department_ids[]')
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            employee_id = request.form.get('employee_id')
            rp_type = request.form.get('type')  # reward, penalty, or all
        else:
            department_ids = request.args.getlist('department_ids[]')
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            employee_id = request.args.get('employee_id')
            rp_type = request.args.get('type')
        
        # Validate: at least one department must be selected
        if not department_ids:
            if request.method == 'POST':
                flash('يرجى اختيار قسم واحد على الأقل', 'error')
                return redirect(url_for('export_reward_penalty_form'))
            else:
                return jsonify({'success': False, 'message': 'يرجى اختيار قسم واحد على الأقل'})
        
        # Convert department_ids to integers
        try:
            department_ids = [int(did) for did in department_ids if did]
        except ValueError:
            if request.method == 'POST':
                flash('معرف القسم غير صالح', 'error')
                return redirect(url_for('export_reward_penalty_form'))
            else:
                return jsonify({'success': False, 'message': 'معرف القسم غير صالح'})
        
        # Build query
        query = db_session.query(RewardPenalty).filter(
            RewardPenalty.department_id.in_(department_ids)
        )
        
        if start_date:
            query = query.filter(RewardPenalty.effective_date >= start_date)
        if end_date:
            query = query.filter(RewardPenalty.effective_date <= end_date)
        if employee_id and employee_id != 'all':
            query = query.filter(RewardPenalty.user_id == employee_id)
        if rp_type and rp_type != 'all':
            query = query.filter(RewardPenalty.type == rp_type)
        
        # Restrict to user's accessible departments (if not admin)
        if not current_user.is_admin:
            accessible_dept_ids = get_managed_department_ids(current_user.id)
            if current_user.department_id:
                accessible_dept_ids.append(current_user.department_id)
            # Only keep department_ids that are in accessible_dept_ids
            query = query.filter(RewardPenalty.department_id.in_(accessible_dept_ids))
        
        items = query.order_by(RewardPenalty.effective_date.desc()).all()
        
        # Create Excel file
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "تقرير المكافآت والجزاءات"
        
        # Title
        title = f"تقرير المكافآت والجزاءات - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        if start_date and end_date:
            title += f" (من {start_date} إلى {end_date})"
        worksheet.merge_cells('A1:F1')
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers (including type column as requested)
        headers = ['كود الموظف', 'اسم الموظف', 'تاريخ المكافأة', 'قيمة المكافأة', 'ملاحظات', 'النوع']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = 4
        for item in items:
            # Get user info manually (since no relationship defined)
            user = db_session.query(User).get(item.user_id)
            username = user.username if user else ''
            user_name = user.name if user else 'غير معروف'
            
            worksheet.cell(row=row, column=1, value=username)
            worksheet.cell(row=row, column=2, value=user_name)
            worksheet.cell(row=row, column=3, value=item.effective_date.strftime('%Y-%m-%d') if item.effective_date else '')
            worksheet.cell(row=row, column=4, value=item.amount)
            worksheet.cell(row=row, column=5, value=item.reason or '')
            # نوع (reward/penalty) in Arabic
            type_arabic = 'مكافأة' if item.type == 'reward' else 'جزاء' if item.type == 'penalty' else item.type
            worksheet.cell(row=row, column=6, value=type_arabic)
            row += 1
        
        # Adjust column widths
        col_widths = [15, 20, 15, 15, 30, 15]
        for i, width in enumerate(col_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width
        
        workbook.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"تقرير_المكافآت_والجزاءات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        print(f"Error exporting rewards/penalties: {str(e)}")
        import traceback
        traceback.print_exc()
        if request.method == 'POST':
            flash(f'حدث خطأ أثناء تصدير التقرير: {str(e)}', 'error')
            return redirect(url_for('export_reward_penalty_form'))
        else:
            return jsonify({'success': False, 'message': str(e)})
        
if __name__ == '__main__':
    init_db()
    create_default_admin()
    reset_monthly_permission_balances()
    app.run(host='0.0.0.0', port=5551, debug=True)