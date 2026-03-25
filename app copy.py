import os
import json
import re
import warnings
import traceback
import smtplib
from io import BytesIO
from datetime import datetime, date, timedelta
from decimal import Decimal
from functools import wraps
from typing import Any, Dict, List, Union, Optional
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import pandas as pd
from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import extract, or_
from sqlalchemy.exc import SAWarning
from sqlalchemy.orm import joinedload
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from database import init_db, db_session
from models import *
from schedules_sync import sync_all_schedules_with_department_structures



app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here-change-this-in-production'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['EXPORT_FOLDER'] = 'exports'
app.config['SALARY_FOLDER'] = 'salary_slips'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'pdf'}
load_dotenv()





# Create folders if not exist
for folder in [app.config['UPLOAD_FOLDER'], app.config['EXPORT_FOLDER'], app.config['SALARY_FOLDER']]:
    os.makedirs(folder, exist_ok=True)

app.config.update(
    MAIL_SERVER='smtp.gmail.com',
    MAIL_PORT=587,
    MAIL_USE_TLS=True,
    MAIL_USERNAME=os.getenv('GMAIL_USERNAME', 'haronywaled@gmail.com'),
    MAIL_PASSWORD=os.getenv('GMAIL_PASSWORD', 'H@r0n011**'),
    MAIL_DEFAULT_SENDER=os.getenv('GMAIL_USERNAME', 'haronywaled@gmail.com')
)
from flask_mail import Mail
mail = Mail(app)
# Setup Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return db_session.get(User, int(user_id))  # Changed from query().get()

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def calculate_age(birth_date):
    today = date.today()
    return today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

def send_notification(user_id, title, message, notification_type='general', related_id=None):
    """إرسال إشعار للمستخدم"""
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        notification_type=notification_type,
        related_id=related_id
    )
    db_session.add(notification)
    db_session.commit()
    return notification

sync_all_schedules_with_department_structures()



def send_email(to_email, subject, html_content, text_content=None):
    """
    Send email using Gmail SMTP
    """
    try:
        # Get email credentials from config
        sender_email = app.config['MAIL_USERNAME']
        sender_password = app.config['MAIL_PASSWORD']
        
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = sender_email
        msg['To'] = to_email
        
        # Create the plain-text and HTML version of your message
        if text_content:
            part1 = MIMEText(text_content, 'plain')
            msg.attach(part1)
        
        part2 = MIMEText(html_content, 'html')
        msg.attach(part2)
        
        # Send email using Gmail SMTP
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
        
        print(f"✅ Email sent to {to_email}")
        return True
        
    except Exception as e:
        print(f"❌ Failed to send email to {to_email}: {str(e)}")
        return False


def send_notification_email(user_email, notification_title, notification_message, notification_type='general'):
    """
    Send formatted notification email to user
    """
    try:
        # Create HTML email content
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background-color: #4CAF50; color: white; padding: 20px; text-align: center; border-radius: 5px 5px 0 0; }}
                .content {{ background-color: #f9f9f9; padding: 30px; border-radius: 0 0 5px 5px; }}
                .notification-type {{ display: inline-block; padding: 5px 10px; background-color: #e7f3fe; color: #2196F3; border-radius: 3px; font-size: 12px; margin-bottom: 10px; }}
                .footer {{ text-align: center; margin-top: 30px; font-size: 12px; color: #666; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2>🔔 إشعار جديد</h2>
                </div>
                <div class="content">
                    <div class="notification-type">
                        {notification_type}
                    </div>
                    <h3>{notification_title}</h3>
                    <p>{notification_message}</p>
                    <p>📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
                </div>
                <div class="footer">
                    <p>هذا الإشعار تم إرساله تلقائياً من نظام إدارة الموظفين</p>
                    <p>© {datetime.now().year} جميع الحقوق محفوظة</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Create plain text version
        text_content = f"""
        إشعار جديد
        =========
        
        العنوان: {notification_title}
        النوع: {notification_type}
        الرسالة: {notification_message}
        
        التاريخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}
        
        هذا الإشعار تم إرساله تلقائياً من نظام إدارة الموظفين
        """
        
        # Send email
        subject = f"🔔 {notification_title}"
        return send_email(user_email, subject, html_content, text_content)
        
    except Exception as e:
        print(f"❌ Error creating notification email: {str(e)}")
        return False


def create_schedule_from_template_structure(template_data, department, week_start_date):
    """إنشاء جدول من هيكل القالب المحدد في القسم"""
    try:
        week_end_date = week_start_date + timedelta(days=6)
        
        # أيام الأسبوع بالعربية
        days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        
        # إنشاء الهيكل الأساسي
        new_schedule_data = {
            'department': department.name,
            'week_start_date': week_start_date.strftime('%Y-%m-%d'),
            'week_end_date': week_end_date.strftime('%Y-%m-%d'),
            'source': 'department_template',
            'schedule': []
        }
        
        # إذا كان القالب يحتوي على بيانات
        if isinstance(template_data, dict) and 'schedule' in template_data:
            # نسخ الهيكل الموجود
            for i, day_name in enumerate(days_of_week):
                current_date = week_start_date + timedelta(days=i)
                
                # البحث عن اليوم المناسب في القالب
                day_template = None
                for template_day in template_data['schedule']:
                    if isinstance(template_day, dict) and template_day.get('day') == day_name:
                        day_template = template_day
                        break
                
                if day_template:
                    # نسخ اليوم من القالب مع تحديث التاريخ
                    day_entry = day_template.copy()
                    day_entry['date'] = current_date.strftime('%Y-%m-%d')
                    day_entry['day'] = day_name
                else:
                    # إنشاء يوم افتراضي
                    day_entry = {
                        'day': day_name,
                        'date': current_date.strftime('%Y-%m-%d'),
                        'department': department.name,
                        'morning_shift': '',
                        'evening_shift': '',
                        'night_shift': '',
                        'job': 'موظف'
                    }
                
                new_schedule_data['schedule'].append(day_entry)
        else:
            # إذا لم يكن هناك قالب محدد، إنشاء هيكل افتراضي
            for i, day_name in enumerate(days_of_week):
                current_date = week_start_date + timedelta(days=i)
                
                day_entry = {
                    'day': day_name,
                    'date': current_date.strftime('%Y-%m-%d'),
                    'department': department.name,
                    'morning_shift': '',
                    'evening_shift': '',
                    'night_shift': '',
                    'job': 'موظف'
                }
                
                new_schedule_data['schedule'].append(day_entry)
        
        return new_schedule_data
        
    except Exception as e:
        print(f"خطأ في إنشاء الجدول من القالب: {str(e)}")
        return None

def generate_schedule_from_template(department, employees, week_start_date):
    """إنشاء بيانات الجدول من القالب المحدد"""
    try:
        if not department.schedule_structure:
            print(f"لا يوجد هيكل جدول محدد للقسم: {department.name}")
            return create_default_schedule_structure(department, employees, week_start_date)
        
        # تحليل هيكل الجدول المحدد
        template = json.loads(department.schedule_structure)
        
        # إنشاء الجدول بناءً على نوع المصدر
        if template.get('source') == 'excel':
            return create_excel_based_schedule(template, department, employees, week_start_date)
        else:
            return create_custom_schedule(template, department, employees, week_start_date)
            
    except Exception as e:
        print(f"خطأ في إنشاء الجدول من القالب: {str(e)}")
        return create_default_schedule_structure(department, employees, week_start_date)

def create_excel_based_schedule(template, department, employees, week_start_date):
    """إنشاء جدول بناءً على هيكل Excel"""
    schedule_data = {
        'department': department.name,
        'week_start_date': week_start_date.strftime('%Y-%m-%d'),
        'week_end_date': (week_start_date + timedelta(days=6)).strftime('%Y-%m-%d'),
        'source': 'excel_template',
        'schedule': []
    }
    
    # استخدام الهيكل الموجود في القالب إذا كان موجوداً
    if 'structure' in template and template['structure']:
        print(f"استخدام الهيكل الموجود مع {len(template['structure'])} إدخال")
        schedule_data['schedule'] = template['structure']
        
        # تحديث التواريخ لتتناسب مع الأسبوع الحالي
        days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        for i, day_entry in enumerate(schedule_data['schedule']):
            if i < len(days_of_week):
                current_date = week_start_date + timedelta(days=i)
                day_entry['date'] = current_date.strftime('%Y-%m-%d')
                day_entry['day'] = days_of_week[i]
                day_entry['department'] = department.name
    else:
        # إذا لم يكن هناك هيكل، إنشاء هيكل افتراضي
        print("إنشاء هيكل افتراضي")
        schedule_data['schedule'] = create_default_schedule_days(week_start_date, department.name)
    
    return schedule_data


def create_excel_based_schedule(template, department, employees, week_start_date):
    """إنشاء جدول بناءً على هيكل Excel"""
    schedule_data = {
        'department': department.name,
        'week_start_date': week_start_date.strftime('%Y-%m-%d'),
        'week_end_date': (week_start_date + timedelta(days=6)).strftime('%Y-%m-%d'),
        'source': 'excel_template',
        'schedule': []
    }
    
    # إنشاء أيام الأسبوع
    days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
    
    for i, day_name in enumerate(days_of_week):
        current_date = week_start_date + timedelta(days=i)
        
        day_schedule = {
            'day': day_name,
            'date': current_date.strftime('%Y-%m-%d'),
            'department': department.name,
            'morning_shift': '',
            'evening_shift': '',
            'night_shift': '',
            'job': ''
        }
        
        # إذا كان هناك هيكل محدد في القالب، استخدمه
        if 'structure' in template and template['structure']:
            # حاول مطابقة اليوم مع الهيكل المحدد
            for template_day in template['structure']:
                if template_day.get('day') == day_name:
                    day_schedule.update({
                        'morning_shift': template_day.get('morning_shift', ''),
                        'evening_shift': template_day.get('evening_shift', ''),
                        'night_shift': template_day.get('night_shift', ''),
                        'job': template_day.get('job', 'موظف')
                    })
                    break
        
        schedule_data['schedule'].append(day_schedule)
    
    return schedule_data

def create_custom_schedule(template, department, employees, week_start_date):
    """إنشاء جدول مخصص بناءً على الهيكل"""
    schedule_data = {
        'department': department.name,
        'week_start_date': week_start_date.strftime('%Y-%m-%d'),
        'week_end_date': (week_start_date + timedelta(days=6)).strftime('%Y-%m-%d'),
        'source': 'custom_template',
        'schedule': []
    }
    
    # استخدام الهيكل المخصص الموجود في القالب
    if 'structure' in template:
        schedule_data['schedule'] = template['structure']
    else:
        # إذا لم يكن هناك هيكل، إنشاء هيكل افتراضي
        schedule_data['schedule'] = create_default_schedule_days(week_start_date, department.name)
    
    return schedule_data

def create_default_schedule_structure(department, employees, week_start_date):
    """إنشاء هيكل جدول افتراضي"""
    return {
        'department': department.name,
        'week_start_date': week_start_date.strftime('%Y-%m-%d'),
        'week_end_date': (week_start_date + timedelta(days=6)).strftime('%Y-%m-%d'),
        'source': 'default',
        'schedule': create_default_schedule_days(week_start_date, department.name)
    }


def create_notification(user_id, title, message, notification_type='general', related_id=None, action_url=None):
    """
    Create notification and send email
    """
    # Create notification in database
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        notification_type=notification_type,
        related_id=related_id,
        action_url=action_url
    )
    db_session.add(notification)
    db_session.commit()
    
    # Send email notification
    try:
        # Get user information including email
        user = db_session.query(User).get(user_id)
        if user and hasattr(user, 'email') and user.email:
            # Send email in background (non-blocking)
            send_notification_email(user.email, title, message, notification_type)
        else:
            print(f"⚠️ User {user_id} has no email address or email field doesn't exist")
    except Exception as e:
        print(f"❌ Error sending email for notification: {str(e)}")
    
    return notification


def get_notification_email_template(notification_type, title, message):
    """
    Get different email templates based on notification type
    """
    templates = {
        'leave_approved': {
            'subject': '✅ موافقة على طلب الإجازة',
            'icon': '🎉',
            'color': '#4CAF50'
        },
        'leave_rejected': {
            'subject': '❌ رفض طلب الإجازة',
            'icon': '⚠️',
            'color': '#f44336'
        },
        'salary': {
            'subject': '💰 شيت مرتب جديد',
            'icon': '💰',
            'color': '#FF9800'
        },
        'permission_approved': {
            'subject': '✅ موافقة على طلب الإذن',
            'icon': '👍',
            'color': '#2196F3'
        },
        'schedule': {
            'subject': '📅 جدول عمل جديد',
            'icon': '📅',
            'color': '#9C27B0'
        },
        'default': {
            'subject': '🔔 إشعار جديد',
            'icon': '🔔',
            'color': '#607D8B'
        }
    }
    
    template = templates.get(notification_type, templates['default'])
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: 'Arial', sans-serif; direction: rtl; line-height: 1.6; }}
            .container {{ max-width: 600px; margin: 20px auto; background: #fff; border-radius: 10px; overflow: hidden; box-shadow: 0 0 20px rgba(0,0,0,0.1); }}
            .header {{ background: {template['color']}; padding: 30px; text-align: center; color: white; }}
            .header h1 {{ margin: 0; font-size: 24px; }}
            .icon {{ font-size: 40px; margin-bottom: 15px; }}
            .content {{ padding: 30px; }}
            .title {{ color: {template['color']}; font-size: 20px; margin-bottom: 20px; border-bottom: 2px solid {template['color']}; padding-bottom: 10px; }}
            .message {{ font-size: 16px; line-height: 1.8; color: #444; margin-bottom: 30px; }}
            .footer {{ background: #f5f5f5; padding: 20px; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #ddd; }}
            .button {{ display: inline-block; background: {template['color']}; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; margin-top: 20px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <div class="icon">{template['icon']}</div>
                <h1>{template['subject']}</h1>
            </div>
            <div class="content">
                <div class="title">{title}</div>
                <div class="message">{message}</div>
                <div style="text-align: center;">
                    <a href="{app.config.get('APP_URL', 'http://localhost:5551')}" class="button">الدخول إلى النظام</a>
                </div>
            </div>
            <div class="footer">
                <p>© {datetime.now().year} نظام إدارة الموظفين - جميع الحقوق محفوظة</p>
                <p>هذا إشعار تلقائي، يرجى عدم الرد على هذا البريد الإلكتروني</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return template['subject'], html_content


def send_notification_email(user_email, notification_title, notification_message, notification_type='general'):
    """
    Send formatted notification email using templates
    """
    try:
        # Get email template based on notification type
        subject, html_content = get_notification_email_template(
            notification_type, 
            notification_title, 
            notification_message
        )
        
        # Create plain text version
        text_content = f"""
        {notification_title}
        {'=' * len(notification_title)}
        
        {notification_message}
        
        نوع الإشعار: {notification_type}
        التاريخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}
        
        للدخول إلى النظام: {app.config.get('APP_URL', 'http://localhost:5551')}
        
        ---
        نظام إدارة الموظفين
        """
        
        # Send email
        return send_email(user_email, subject, html_content, text_content)
        
    except Exception as e:
        print(f"❌ Error sending notification email: {str(e)}")
        return False

def create_notification(user_id, title, message, notification_type='general', related_id=None, action_url=None):
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        notification_type=notification_type,
        related_id=related_id,
        action_url=action_url
    )
    db_session.add(notification)
    db_session.commit()
    return notification

def get_user_notifications(user_id):
    return db_session.query(Notification).filter_by(
        user_id=user_id, is_read=False
    ).order_by(Notification.created_at.desc()).all()

def export_to_excel(user_id):
    user = db_session.query(User).get(user_id)
    employee_data = db_session.query(EmployeeData).filter_by(user_id=user_id).first()
    
    if not employee_data:
        return None
    
    data_dict = {
        'الاسم بالعربية': [employee_data.arabic_name or ''],
        'الاسم بالإنجليزية': [employee_data.english_name or ''],
        'الرقم القومي': [employee_data.national_id or ''],
        'تاريخ إصدار البطاقة': [employee_data.id_issue_date or ''],
        'تاريخ الميلاد': [employee_data.birth_date or ''],
        'السن': [employee_data.age or ''],
        'واتساب': [employee_data.whatsapp or ''],
        'الهاتف': [employee_data.phone or ''],
        'العنوان': [employee_data.address or ''],
        'الموقف من التجنيد': [employee_data.military_status or ''],
        'الحالة الاجتماعية': [employee_data.marital_status or ''],
        'المؤهل الدراسي': [employee_data.qualification or ''],
        'سنة التخرج': [employee_data.graduation_year or ''],
        'التقدير': [employee_data.grade or ''],
        'يعمل حالياً': ['نعم' if employee_data.has_work else 'لا'],
        'جهة العمل': [employee_data.workplace or ''],
        'الوظيفة': [employee_data.job_title or ''],
        'الرقم التأميني': [employee_data.insurance_number or ''],
        'الرقم الضريبي': [employee_data.tax_number or ''],
        'ترخيص المهنة': [employee_data.profession_license or ''],
        'كارنية النقابة': [employee_data.union_card or ''],
        'جهة الاتصال 1 - الاسم': [employee_data.emergency1_name or ''],
        'جهة الاتصال 1 - التليفون': [employee_data.emergency1_phone or ''],
        'جهة الاتصال 1 - العنوان': [employee_data.emergency1_address or ''],
        'جهة الاتصال 1 - صلة القرابة': [employee_data.emergency1_relation or ''],
        'جهة الاتصال 2 - الاسم': [employee_data.emergency2_name or ''],
        'جهة الاتصال 2 - التليفون': [employee_data.emergency2_phone or ''],
        'جهة الاتصال 2 - العنوان': [employee_data.emergency2_address or ''],
        'جهة الاتصال 2 - صلة القرابة': [employee_data.emergency2_relation or ''],
        'نسبة الإكتمال': [f'{employee_data.completion_percentage}%'],
        'آخر تحديث': [employee_data.last_updated or ''],
        'تم التحديث بواسطة': [employee_data.updated_by or '']
    }
    
    df = pd.DataFrame(data_dict)
    
    user_export_folder = os.path.join(app.config['EXPORT_FOLDER'], str(user_id))
    os.makedirs(user_export_folder, exist_ok=True)
    
    filename = f"employee_data_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(user_export_folder, filename)
    df.to_excel(filepath, index=False, engine='openpyxl')
    
    return filepath



def update_database_schema():
    """تحديث هيكل قاعدة البيانات بإضافة الحقول المفقودة"""
    try:
        # استيراد سكريبت الإصلاح وتشغيله
        from database_fix import fix_database, check_specific_columns
        
        print("جاري تحديث هيكل قاعدة البيانات...")
        if fix_database():
            print("✓ تم تحديث هيكل قاعدة البيانات بنجاح")
            if check_specific_columns():
                print("✓ جميع الحقول المطلوبة موجودة")
                return True
            else:
                print("⚠️ بعض الحقول لا تزال مفقودة")
                return False
        else:
            print("❌ فشل تحديث قاعدة البيانات")
            return False
            
    except Exception as e:
        print(f"❌ خطأ في تحديث قاعدة البيانات: {e}")
        return False



def remove_weak_tables():
    """إزالة الجداول الضعيفة التي لم تعد مستخدمة"""
    try:
        from database_fix import remove_weak_tables as remove_tables
        print("جاري إزالة الجداول الضعيفة...")
        return remove_tables()
    except Exception as e:
        print(f"❌ خطأ في إزالة الجداول الضعيفة: {e}")
        return False

def update_database_schema():
    """تحديث هيكل قاعدة البيانات بإضافة الحقول المفقودة"""
    try:
        # استيراد سكريبت الإصلاح وتشغيله
        from database_fix import fix_database, check_specific_columns
        
        print("جاري تحديث هيكل قاعدة البيانات...")
        if fix_database():
            print("✓ تم تحديث هيكل قاعدة البيانات بنجاح")
            if check_specific_columns():
                print("✓ جميع الحقول المطلوبة موجودة")
                return True
            else:
                print("⚠️ بعض الحقول لا تزال مفقودة")
                return False
        else:
            print("❌ فشل تحديث قاعدة البيانات")
            return False
            
    except Exception as e:
        print(f"❌ خطأ في تحديث قاعدة البيانات: {e}")
        return False


def create_default_admin():
    admin = db_session.query(User).filter_by(username='admin').first()
    if not admin:
        admin_user = User(
            username='admin',
            name='المسؤول العام',
            password_hash=generate_password_hash('admin123'),
            is_admin=True,
            is_manager=True
        )
        db_session.add(admin_user)
        
        # Create default department
        default_dept = Department(
            name='الإدارة العامة',
            created_by=1
        )
        db_session.add(default_dept)
        db_session.commit()
        
        # Update admin department
        admin_user.department_id = default_dept.id
        default_dept.primary_manager_id = admin_user.id
        db_session.commit()
        
        print('تم إنشاء المستخدم المسؤول الافتراضي: admin / admin123')

# ======== Basic Routes ========

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = db_session.query(User).filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            user.last_login = datetime.now()
            db_session.commit()
            
            if user.is_admin:
                return redirect(url_for('admin_dashboard'))
            elif user.is_manager:
                return redirect(url_for('manager_dashboard'))
            else:
                return redirect(url_for('user_dashboard'))
        else:
            flash('اسم المستخدم أو كلمة المرور غير صحيحة')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ======== Admin Routes ========

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    users_count = db_session.query(User).count()
    departments_count = db_session.query(Department).count()
    pending_leaves = db_session.query(LeaveRequest).filter_by(status='pending').count()
    pending_permissions = db_session.query(PermissionRequest).filter_by(status='pending').count()
    pending_schedules = db_session.query(WeeklySchedule).filter_by(is_approved=False).count()
    
    notifications = get_user_notifications(current_user.id)
    notifications_count = len(notifications)  # Add this line
    
    return render_template('admin_dashboard.html',
                         users_count=users_count,
                         departments_count=departments_count,
                         pending_leaves=pending_leaves,
                         pending_permissions=pending_permissions,
                         pending_schedules=pending_schedules,
                         notifications=notifications,
                         notifications_count=notifications_count)  # Add this
@app.route('/admin/users')
@login_required
def admin_users():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    users = db_session.query(User).all()
    departments = db_session.query(Department).all()
    
    for user in users:
        user.employee_data = db_session.query(EmployeeData).filter_by(user_id=user.id).first()
        if user.employee_data:
            user.employee_data.calculate_completion()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_users.html', 
                         users=users, 
                         departments=departments,
                         notifications=notifications)


@app.route('/admin/export_all_users')
@login_required
def export_all_users():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        # Get all non-admin users with department information
        users = db_session.query(User).filter_by(is_admin=False).all()
        
        if not users:
            flash('لا يوجد مستخدمين للتصدير')
            return redirect(url_for('admin_users'))
        
        # Preload departments for better performance
        departments = {dept.id: dept for dept in db_session.query(Department).all()}
        
        # Create a list to hold all user data
        all_users_data = []
        
        for user in users:
            employee_data = db_session.query(EmployeeData).filter_by(user_id=user.id).first()
            
            # Get department name safely
            department_name = 'غير معين'
            if user.department_id and user.department_id in departments:
                department_name = departments[user.department_id].name
            
            if employee_data:
                user_dict = {
                    'اسم المستخدم': user.username,
                    'الاسم الكامل': user.name,
                    'الصفة': 'مدير' if user.is_manager else 'موظف',
                    'القسم': department_name,
                    'الاسم بالعربية': employee_data.arabic_name or '',
                    'الاسم بالإنجليزية': employee_data.english_name or '',
                    'الرقم القومي': employee_data.national_id or '',
                    'تاريخ إصدار البطاقة': employee_data.id_issue_date.strftime('%Y-%m-%d') if employee_data.id_issue_date else '',
                    'تاريخ الميلاد': employee_data.birth_date.strftime('%Y-%m-%d') if employee_data.birth_date else '',
                    'السن': employee_data.age or '',
                    'واتساب': employee_data.whatsapp or '',
                    'الهاتف': employee_data.phone or '',
                    'العنوان': employee_data.address or '',
                    'الموقف من التجنيد': employee_data.military_status or '',
                    'الحالة الاجتماعية': employee_data.marital_status or '',
                    'المؤهل الدراسي': employee_data.qualification or '',
                    'سنة التخرج': employee_data.graduation_year or '',
                    'التقدير': employee_data.grade or '',
                    'يعمل حالياً': 'نعم' if employee_data.has_work else 'لا',
                    'جهة العمل': employee_data.workplace or '',
                    'الوظيفة': employee_data.job_title or '',
                    'الرقم التأميني': employee_data.insurance_number or '',
                    'الرقم الضريبي': employee_data.tax_number or '',
                    'ترخيص المهنة': employee_data.profession_license or '',
                    'كارنية النقابة': employee_data.union_card or '',
                    'جهة الاتصال 1 - الاسم': employee_data.emergency1_name or '',
                    'جهة الاتصال 1 - التليفون': employee_data.emergency1_phone or '',
                    'جهة الاتصال 1 - العنوان': employee_data.emergency1_address or '',
                    'جهة الاتصال 1 - صلة القرابة': employee_data.emergency1_relation or '',
                    'جهة الاتصال 2 - الاسم': employee_data.emergency2_name or '',
                    'جهة الاتصال 2 - التليفون': employee_data.emergency2_phone or '',
                    'جهة الاتصال 2 - العنوان': employee_data.emergency2_address or '',
                    'جهة الاتصال 2 - صلة القرابة': employee_data.emergency2_relation or '',
                    'نسبة الإكتمال': f'{employee_data.completion_percentage}%',
                    'آخر تحديث': employee_data.last_updated.strftime('%Y-%m-%d %H:%M') if employee_data.last_updated else '',
                    'تم التحديث بواسطة': employee_data.updated_by or ''
                }
            else:
                user_dict = {
                    'اسم المستخدم': user.username,
                    'الاسم الكامل': user.name,
                    'الصفة': 'مدير' if user.is_manager else 'موظف',
                    'القسم': department_name,
                    'الاسم بالعربية': '',
                    'الاسم بالإنجليزية': '',
                    'الرقم القومي': '',
                    'تاريخ إصدار البطاقة': '',
                    'تاريخ الميلاد': '',
                    'السن': '',
                    'واتساب': '',
                    'الهاتف': '',
                    'العنوان': '',
                    'الموقف من التجنيد': '',
                    'الحالة الاجتماعية': '',
                    'المؤهل الدراسي': '',
                    'سنة التخرج': '',
                    'التقدير': '',
                    'يعمل حالياً': 'لا',
                    'جهة العمل': '',
                    'الوظيفة': '',
                    'الرقم التأميني': '',
                    'الرقم الضريبي': '',
                    'ترخيص المهنة': '',
                    'كارنية النقابة': '',
                    'جهة الاتصال 1 - الاسم': '',
                    'جهة الاتصال 1 - التليفون': '',
                    'جهة الاتصال 1 - العنوان': '',
                    'جهة الاتصال 1 - صلة القرابة': '',
                    'جهة الاتصال 2 - الاسم': '',
                    'جهة الاتصال 2 - التليفون': '',
                    'جهة الاتصال 2 - العنوان': '',
                    'جهة الاتصال 2 - صلة القرابة': '',
                    'نسبة الإكتمال': '0%',
                    'آخر تحديث': '',
                    'تم التحديث بواسطة': ''
                }
            
            all_users_data.append(user_dict)
        
        # Create DataFrame
        df = pd.DataFrame(all_users_data)
        
        # Create export folder if not exists
        export_folder = os.path.join(app.config['EXPORT_FOLDER'], 'all_users')
        os.makedirs(export_folder, exist_ok=True)
        
        # Generate filename
        filename = f"all_users_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(export_folder, filename)
        
        # Save to Excel
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        # Log the action
        print(f'تم تصدير بيانات {len(users)} مستخدم بواسطة {current_user.name}')
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        flash(f'حدث خطأ أثناء التصدير: {str(e)}')
        return redirect(url_for('admin_users'))
    
@app.route('/admin/create_user', methods=['GET', 'POST'])
@login_required
def create_user():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    departments = db_session.query(Department).all()
    
    if request.method == 'POST':
        username = request.form['username']
        name = request.form['name']
        password = request.form['password']
        is_manager = 'is_manager' in request.form
        department_id = request.form.get('department_id')
        
        if db_session.query(User).filter_by(username=username).first():
            flash('اسم المستخدم موجود بالفعل')
            return redirect(url_for('create_user'))
        
        user = User(
            username=username,
            name=name,
            password_hash=generate_password_hash(password),
            is_admin=False,
            is_manager=is_manager,
            department_id=department_id
        )
        
        db_session.add(user)
        db_session.commit()
        
        # Create employee balance
        balance = EmployeeBalance(user_id=user.id)
        db_session.add(balance)
        db_session.commit()
        
        create_notification(
            current_user.id,
            'تم إنشاء مستخدم جديد',
            f'تم إنشاء المستخدم {name} بنجاح',
            'user_created'
        )
        
        flash('تم إنشاء المستخدم بنجاح')
        return redirect(url_for('admin_users'))
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_create_user.html', 
                         departments=departments,
                         notifications=notifications)






@app.route('/admin/users/<int:user_id>/edit_complete', methods=['GET', 'POST'])
@login_required
def admin_edit_user_complete(user_id):
    """
    Complete user data editing - for admin only
    Allows editing all user fields including personal information
    """
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    # Get the user to edit
    user = db_session.query(User).filter_by(id=user_id).first()
    if not user:
        flash('المستخدم غير موجود', 'error')
        return redirect(url_for('admin_users'))
    
    # Get employee data or create new if doesn't exist
    employee_data = db_session.query(EmployeeData).filter_by(user_id=user_id).first()
    if not employee_data:
        employee_data = EmployeeData(user_id=user_id)
        db_session.add(employee_data)
        db_session.commit()
    
    # Get departments for dropdown
    departments = db_session.query(Department).all()
    
    if request.method == 'POST':
        try:
            # Update basic user information
            user.username = request.form.get('username', user.username)
            user.name = request.form.get('name', user.name)
            user.department_id = request.form.get('department_id') or None
            user.is_manager = 'is_manager' in request.form
            user.is_active = 'is_active' in request.form
            
            # Update password only if provided
            new_password = request.form.get('password', '').strip()
            if new_password:
                if len(new_password) < 6:
                    flash('كلمة المرور يجب أن تكون至少 6 أحرف', 'error')
                    return redirect(url_for('admin_edit_user_complete', user_id=user_id))
                user.password_hash = generate_password_hash(new_password)
            
            # Update employee data
            employee_data.arabic_name = request.form.get('arabic_name') or None
            employee_data.english_name = request.form.get('english_name') or None
            employee_data.national_id = request.form.get('national_id') or None
            
            # Handle dates
            if request.form.get('id_issue_date'):
                employee_data.id_issue_date = datetime.strptime(request.form['id_issue_date'], '%Y-%m-%d').date()
            else:
                employee_data.id_issue_date = None
            
            if request.form.get('birth_date'):
                employee_data.birth_date = datetime.strptime(request.form['birth_date'], '%Y-%m-%d').date()
                employee_data.age = calculate_age(employee_data.birth_date)
            else:
                employee_data.birth_date = None
                employee_data.age = None
            
            # Contact information
            employee_data.whatsapp = request.form.get('whatsapp') or None
            employee_data.phone = request.form.get('phone') or None
            employee_data.address = request.form.get('address') or None
            
            # Personal information
            employee_data.military_status = request.form.get('military_status') or None
            employee_data.marital_status = request.form.get('marital_status') or None
            employee_data.qualification = request.form.get('qualification') or None
            
            if request.form.get('graduation_year'):
                employee_data.graduation_year = int(request.form['graduation_year'])
            else:
                employee_data.graduation_year = None
            
            employee_data.grade = request.form.get('grade') or None
            employee_data.has_work = 'has_work' in request.form
            employee_data.workplace = request.form.get('workplace') or None
            employee_data.job_title = request.form.get('job_title') or None
            
            # Professional information
            employee_data.insurance_number = request.form.get('insurance_number') or None
            employee_data.tax_number = request.form.get('tax_number') or None
            employee_data.profession_license = request.form.get('profession_license') or None
            employee_data.union_card = request.form.get('union_card') or None
            
            # Emergency contacts
            employee_data.emergency1_name = request.form.get('emergency1_name') or None
            employee_data.emergency1_phone = request.form.get('emergency1_phone') or None
            employee_data.emergency1_address = request.form.get('emergency1_address') or None
            employee_data.emergency1_relation = request.form.get('emergency1_relation') or None
            employee_data.emergency2_name = request.form.get('emergency2_name') or None
            employee_data.emergency2_phone = request.form.get('emergency2_phone') or None
            employee_data.emergency2_address = request.form.get('emergency2_address') or None
            employee_data.emergency2_relation = request.form.get('emergency2_relation') or None
            
            # Handle file uploads
            user_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], str(user_id))
            os.makedirs(user_upload_folder, exist_ok=True)
            
            file_fields = [
                ('national_id_image', 'national_id_image'),
                ('military_status_image', 'military_status_image'),
                ('qualification_image', 'qualification_image'),
                ('salary_details', 'salary_details'),
                ('employment_status', 'employment_status')
            ]
            
            for form_field, db_field in file_fields:
                file = request.files.get(form_field)
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    file_path = os.path.join(user_upload_folder, filename)
                    file.save(file_path)
                    setattr(employee_data, db_field, filename)
            
            # Update timestamps
            employee_data.last_updated = datetime.now()
            employee_data.updated_by = current_user.name
            
            # Calculate completion percentage
            employee_data.calculate_completion()
            
            db_session.commit()
            
            # Create notification
            create_notification(
                current_user.id,
                'تم تحديث بيانات المستخدم',
                f'تم تحديث البيانات الكاملة للمستخدم {user.name}',
                'user_data_updated'
            )
            
            flash('تم تحديث بيانات المستخدم بنجاح', 'success')
            return redirect(url_for('admin_users'))
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ أثناء تحديث البيانات: {str(e)}', 'error')
            print(f"Error updating user data: {str(e)}")
            return redirect(url_for('admin_edit_user_complete', user_id=user_id))
    
    # GET request - display the form
    notifications = get_user_notifications(current_user.id)
    
    # Calculate completion percentage for display
    completion_percentage = employee_data.calculate_completion()
    missing_fields = employee_data.get_missing_fields()
    
    return render_template('admin/admin_edit_user_complete.html', 
                         user=user,
                         employee_data=employee_data,
                         departments=departments,
                         completion_percentage=completion_percentage,
                         missing_fields=missing_fields,
                         notifications=notifications)






@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    """
    Edit user information - only allows password updates
    All other fields are read-only as per the HTML template
    """
    
    # Get the user to edit
    user = db_session.query(User).filter_by(id=user_id).first()
    if not user:
        flash('المستخدم غير موجود', 'error')
        return redirect(url_for('admin_users'))
    
    # Get departments for display (read-only in template)
    departments = db_session.query(Department).all()
    
    if request.method == 'POST':
        try:
            # Only process password update if provided
            new_password = request.form.get('password', '').strip()
            
            # Update password only if not empty
            if new_password:
                # Validate password strength (optional)
                if len(new_password) < 6:
                    flash('كلمة المرور يجب أن تكون至少 6 أحرف', 'error')
                    return redirect(url_for('edit_user', user_id=user_id))
                
                user.password_hash = generate_password_hash(new_password)
                db_session.commit()
                
                # Create notification for password change
                create_notification(
                    current_user.id,
                    'تم تحديث كلمة المرور',
                    f'تم تحديث كلمة مرور المستخدم {user.name} بنجاح',
                    'password_updated'
                )
                
                flash('تم تحديث كلمة المرور بنجاح', 'success')
            else:
                flash('لم يتم إجراء أي تغييرات', 'info')
            
            return redirect(url_for('admin_users'))
            
        except Exception as e:
            db_session.rollback()
            flash('حدث خطأ أثناء تحديث البيانات', 'error')
            # Log the error for debugging
            print(f"Error updating user: {str(e)}")
            return redirect(url_for('edit_user', user_id=user_id))
    
    # GET request - display the form
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_edit_user.html', 
                         user=user,
                         departments=departments,
                         notifications=notifications)
@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    user = db_session.query(User).filter_by(id=user_id).first()
    if not user:
        flash('المستخدم غير موجود')
        return redirect(url_for('admin_users'))
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        flash('لا يمكنك حذف حسابك الخاص')
        return redirect(url_for('admin_users'))
    
    # Prevent deleting other admins (optional security measure)
    if user.is_admin and user.id != current_user.id:
        flash('لا يمكنك حذف مسؤول آخر')
        return redirect(url_for('admin_users'))
    
    # Store user info for notification before deletion
    user_name = user.name
    
    try:
        # Delete related records first (adjust based on your database relationships)
        # Delete employee data
        employee_data = db_session.query(EmployeeData).filter_by(user_id=user_id).first()
        if employee_data:
            db_session.delete(employee_data)
        
        # Delete employee balance
        employee_balance = db_session.query(EmployeeBalance).filter_by(user_id=user_id).first()
        if employee_balance:
            db_session.delete(employee_balance)
        
        # Delete notifications
        user_notifications = db_session.query(Notification).filter_by(user_id=user_id).all()
        for notification in user_notifications:
            db_session.delete(notification)
        
        # Delete the user
        db_session.delete(user)
        db_session.commit()
        
        create_notification(
            current_user.id,
            'تم حذف المستخدم',
            f'تم حذف المستخدم {user_name} بنجاح',
            'user_deleted'
        )
        
        flash('تم حذف المستخدم بنجاح')
        
    except Exception as e:
        db_session.rollback()
        flash('حدث خطأ أثناء حذف المستخدم')
        print(f"Error deleting user: {e}")
    
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/toggle_status', methods=['POST'])
@login_required
def toggle_user_status(user_id):
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    user = db_session.query(User).filter_by(id=user_id).first()
    if not user:
        flash('المستخدم غير موجود')
        return redirect(url_for('admin_users'))
    
    # Prevent deactivating yourself
    if user.id == current_user.id:
        flash('لا يمكنك تعطيل حسابك الخاص')
        return redirect(url_for('admin_users'))
    
    user.is_active = not user.is_active
    db_session.commit()
    
    status_text = "مفعل" if user.is_active else "معطل"
    create_notification(
        current_user.id,
        'تم تغيير حالة المستخدم',
        f'تم تغيير حالة المستخدم {user.name} إلى {status_text}',
        'user_status_changed'
    )
    
    flash(f'تم تغيير حالة المستخدم إلى {status_text}')
    return redirect(url_for('admin_users'))

@app.route('/admin/assign_manager', methods=['POST'])
@login_required
def assign_department_manager():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    department_id = request.form['department_id']
    manager_id = request.form['manager_id']
    
    department = db_session.query(Department).get(department_id)
    new_manager = db_session.query(User).get(manager_id)
    
    if department and new_manager:
        # Remove manager role from previous manager if exists
        if department.manager_id:
            old_manager = db_session.query(User).get(department.manager_id)
            if old_manager:
                # Check if old manager is manager of any other department
                other_departments = db_session.query(Department).filter(
                    Department.manager_id == old_manager.id,
                    Department.id != department.id
                ).count()
                if other_departments == 0:
                    old_manager.is_manager = False
        
        # Assign new manager
        department.manager_id = manager_id
        new_manager.is_manager = True
        new_manager.department_id = department_id
        
        db_session.commit()
        
        create_notification(
            new_manager.id,
            'تم تعيينك كمدير قسم',
            f'تم تعيينك كمدير للقسم {department.name}',
            'manager_assigned'
        )
        
        flash(f'تم تعيين {new_manager.name} كمدير للقسم {department.name} بنجاح')
    else:
        flash('حدث خطأ في تعيين المدير')
    
    return redirect(url_for('admin_departments'))

@app.route('/admin/migrate_schedules')
@login_required
def admin_migrate_schedules():
    """واجهة إدارة هجرة الجداول"""
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    # Get all schedules to analyze current state
    all_schedules = db_session.query(WeeklySchedule).all()
    
    # Analyze current structures
    structure_types = {
        'new_format': 0,
        'old_dict_format': 0,
        'list_format': 0,
        'empty': 0,
        'invalid': 0
    }
    
    for schedule in all_schedules:
        if not schedule.schedule_data:
            structure_types['empty'] += 1
            continue
        
        try:
            data = json.loads(schedule.schedule_data)
            
            if isinstance(data, dict):
                if 'schedule' in data:
                    structure_types['new_format'] += 1
                else:
                    structure_types['old_dict_format'] += 1
            elif isinstance(data, list):
                structure_types['list_format'] += 1
            else:
                structure_types['invalid'] += 1
                
        except:
            structure_types['invalid'] += 1
    
    # Get total counts
    total_schedules = len(all_schedules)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_migrate_schedules.html',
                         total_schedules=total_schedules,
                         structure_types=structure_types,
                         notifications=notifications)



def normalize_all_schedules_to_new_structure():
    """تحويل جميع الجداول إلى نفس الهيكل الجديد الموحد"""
    try:
        print("=== بدء تحويل جميع الجداول إلى الهيكل الموحد ===")
        
        # الحصول على جميع الجداول الحالية
        all_schedules = db_session.query(WeeklySchedule).all()
        
        normalized_count = 0
        failed_count = 0
        
        for schedule in all_schedules:
            try:
                # الحصول على القسم
                department = db_session.query(Department).get(schedule.department_id)
                if not department:
                    print(f"⚠️ القسم غير موجود للجدول {schedule.id}")
                    failed_count += 1
                    continue
                
                # تعريف الهيكل الجديد الموحد
                new_structure = {
                    'department': department.name,
                    'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
                    'week_end_date': schedule.week_end_date.strftime('%Y-%m-%d'),
                    'source': 'normalized_migration',
                    'schedule': []
                }
                
                # أيام الأسبوع
                days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
                
                # إنشاء الهيكل الموحد للأيام
                for i, day_name in enumerate(days_of_week):
                    current_date = schedule.week_start_date + timedelta(days=i)
                    
                    day_entry = {
                        'day': day_name,
                        'date': current_date.strftime('%Y-%m-%d'),
                        'department': department.name,
                        'morning_shift': '',
                        'evening_shift': '',
                        'night_shift': '',
                        'job': 'موظف'
                    }
                    
                    # محاولة استخراج البيانات من الهيكل القديم
                    if schedule.schedule_data:
                        try:
                            old_data = json.loads(schedule.schedule_data)
                            extracted = extract_shifts_from_old_structure(old_data, day_name, i, department.id)
                            if extracted:
                                day_entry.update(extracted)
                        except:
                            pass
                    
                    new_structure['schedule'].append(day_entry)
                
                # تحديث الجدول بالهيكل الموحد الجديد
                schedule.schedule_data = json.dumps(new_structure, ensure_ascii=False)
                normalized_count += 1
                
                print(f"✓ تم تحويل الجدول {schedule.id} للقسم {department.name} إلى الهيكل الموحد")
                
            except Exception as e:
                print(f"❌ خطأ في تحويل الجدول {schedule.id}: {str(e)}")
                failed_count += 1
                continue
        
        db_session.commit()
        
        print(f"=== تم تحويل {normalized_count} جدول بنجاح، فشل {failed_count} جدول ===")
        return normalized_count, failed_count
        
    except Exception as e:
        db_session.rollback()
        print(f"❌ خطأ عام في تحويل الجداول: {str(e)}")
        return 0, len(all_schedules)

@app.route('/admin/execute_normalize_migration', methods=['POST'])
@login_required
def admin_execute_normalize_migration():
    """تنفيذ عملية توحيد هياكل جميع الجداول"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        normalized_count, failed_count = normalize_all_schedules_to_new_structure()
        
        return jsonify({
            'success': True,
            'message': f'تم توحيد هياكل {normalized_count} جدول بنجاح',
            'normalized_count': normalized_count,
            'failed_count': failed_count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في توحيد الهياكل: {str(e)}'
        })


def extract_shifts_from_old_structure(old_data, day_name, day_index, department_id):
    """استخراج الشيفتات من الهيكل القديم"""
    try:
        extracted = {}
        arabic_days = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        
        if isinstance(old_data, dict):
            # الهيكل القديم: {employee_id: {day: shift, ...}}
            morning_shifts = []
            evening_shifts = []
            night_shifts = []
            
            for emp_id, emp_schedule in old_data.items():
                if isinstance(emp_schedule, dict):
                    arabic_day = arabic_days[day_index]
                    shift = emp_schedule.get(arabic_day, '')
                    
                    if shift:
                        # الحصول على اسم الموظف
                        try:
                            emp = db_session.query(User).get(int(emp_id))
                            if emp and emp.department_id == department_id:
                                emp_name = emp.name
                            else:
                                emp_name = f"مستخدم {emp_id}"
                        except:
                            emp_name = f"مستخدم {emp_id}"
                        
                        if 'صباحي' in str(shift):
                            morning_shifts.append(emp_name)
                        elif 'مسائي' in str(shift):
                            evening_shifts.append(emp_name)
                        elif 'ليلي' in str(shift) or 'سهر' in str(shift):
                            night_shifts.append(emp_name)
            
            if morning_shifts:
                extracted['morning_shift'] = ', '.join(morning_shifts)
            if evening_shifts:
                extracted['evening_shift'] = ', '.join(evening_shifts)
            if night_shifts:
                extracted['night_shift'] = ', '.join(night_shifts)
                
        elif isinstance(old_data, list):
            # الهيكل القديم: [{day: ..., shift: ...}]
            for item in old_data:
                if isinstance(item, dict):
                    item_day = item.get('day') or item.get('اليوم')
                    if item_day == day_name:
                        if item.get('morning_shift'):
                            extracted['morning_shift'] = item['morning_shift']
                        if item.get('evening_shift'):
                            extracted['evening_shift'] = item['evening_shift']
                        if item.get('night_shift'):
                            extracted['night_shift'] = item['night_shift']
                        if item.get('job'):
                            extracted['job'] = item['job']
                        break
        
        return extracted
        
    except Exception as e:
        print(f"⚠️ خطأ في استخراج البيانات: {str(e)}")
        return {}

def migrate_existing_schedules_to_new_structure():
    """هجرة الجداول الحالية إلى الهيكل الجديد مع الحفاظ على البيانات"""
    try:
        print("=== بدء هجرة الجداول إلى الهيكل الجديد ===")
        
        # الحصول على جميع الجداول الحالية
        existing_schedules = db_session.query(WeeklySchedule).all()
        
        migrated_count = 0
        
        for schedule in existing_schedules:
            try:
                # إذا كان الجدول يحتوي على بيانات
                if schedule.schedule_data:
                    # تحليل البيانات القديمة
                    old_data = json.loads(schedule.schedule_data)
                    
                    # إنشاء هيكل جديد بناءً على نوع البيانات القديمة
                    new_data = convert_to_new_structure(schedule, old_data)
                    
                    # تحديث بيانات الجدول بالهيكل الجديد
                    schedule.schedule_data = json.dumps(new_data, ensure_ascii=False)
                    
                    # تحديث حالة الجدول ليصبح draft (مسودة) حتى يتم مراجعته
                    schedule.status = 'draft'
                    schedule.is_approved = False
                    schedule.is_locked = False
                    
                    migrated_count += 1
                    print(f"✓ تم هجرة الجدول ID: {schedule.id} للقسم: {schedule.department_id}")
                    
            except Exception as e:
                print(f"❌ خطأ في هجرة الجدول ID: {schedule.id}: {str(e)}")
                continue
        
        db_session.commit()
        print(f"=== تم هجرة {migrated_count} جدول من أصل {len(existing_schedules)} ===")
        return migrated_count
        
    except Exception as e:
        print(f"❌ خطأ عام في عملية الهجرة: {str(e)}")
        db_session.rollback()
        return 0

def convert_to_new_structure(schedule, old_data):
    """تحويل البيانات القديمة إلى الهيكل الجديد"""
    try:
        # الحصول على معلومات القسم
        department = db_session.query(Department).get(schedule.department_id)
        
        # إنشاء الهيكل الجديد
        new_structure = {
            'department': department.name if department else f"القسم {schedule.department_id}",
            'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
            'week_end_date': schedule.week_end_date.strftime('%Y-%m-%d'),
            'source': 'migrated_from_old',
            'schedule': []
        }
        
        # تحويل البيانات القديمة حسب النوع
        if isinstance(old_data, dict):
            # النوع القديم: {employee_id: {day: shift, ...}}
            schedule_list = convert_old_dict_structure(old_data, schedule.week_start_date)
            new_structure['schedule'] = schedule_list
            
        elif isinstance(old_data, list):
            # النوع القديم: [{day: ..., shift: ...}]
            schedule_list = convert_old_list_structure(old_data, schedule.week_start_date)
            new_structure['schedule'] = schedule_list
            
        else:
            # إنشاء هيكل افتراضي
            new_structure['schedule'] = create_default_schedule_days(
                schedule.week_start_date, 
                department.name if department else "غير معين"
            )
        
        return new_structure
        
    except Exception as e:
        print(f"خطأ في تحويل الهيكل: {str(e)}")
        # الرجوع إلى الهيكل الافتراضي في حالة الخطأ
        return create_default_schedule_structure(
            department, 
            [], 
            schedule.week_start_date
        )

def convert_old_dict_structure(old_data, week_start_date):
    """تحويل الهيكل القديم (قاموس) إلى الهيكل الجديد"""
    days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
    schedule_list = []
    
    # إنشاء جدول لأيام الأسبوع
    for day_index, day_name in enumerate(days_of_week):
        current_date = week_start_date + timedelta(days=day_index)
        
        day_entry = {
            'day': day_name,
            'date': current_date.strftime('%Y-%m-%d'),
            'department': '',
            'morning_shift': '',
            'evening_shift': '',
            'night_shift': '',
            'job': 'موظف'
        }
        
        # محاولة تعبئة البيانات من الهيكل القديم
        morning_shifts = []
        evening_shifts = []
        night_shifts = []
        
        for employee_id, employee_schedule in old_data.items():
            if isinstance(employee_schedule, dict):
                # الحصول على اسم اليوم العربي
                arabic_day = get_arabic_day_name(day_index)
                
                # الحصول على الشيفت لهذا اليوم
                shift = employee_schedule.get(arabic_day, '')
                
                if 'صباحي' in str(shift):
                    # محاولة الحصول على اسم الموظف
                    employee = db_session.query(User).get(int(employee_id)) if employee_id.isdigit() else None
                    employee_name = employee.name if employee else f"مستخدم {employee_id}"
                    morning_shifts.append(employee_name)
                elif 'مسائي' in str(shift):
                    employee = db_session.query(User).get(int(employee_id)) if employee_id.isdigit() else None
                    employee_name = employee.name if employee else f"مستخدم {employee_id}"
                    evening_shifts.append(employee_name)
                elif 'ليلي' in str(shift):
                    employee = db_session.query(User).get(int(employee_id)) if employee_id.isdigit() else None
                    employee_name = employee.name if employee else f"مستخدم {employee_id}"
                    night_shifts.append(employee_name)
        
        # تعبئة البيانات في الهيكل الجديد
        if morning_shifts:
            day_entry['morning_shift'] = ', '.join(morning_shifts)
        if evening_shifts:
            day_entry['evening_shift'] = ', '.join(evening_shifts)
        if night_shifts:
            day_entry['night_shift'] = ', '.join(night_shifts)
        
        schedule_list.append(day_entry)
    
    return schedule_list

def convert_old_list_structure(old_data, week_start_date):
    """تحويل الهيكل القديم (قائمة) إلى الهيكل الجديد"""
    days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
    schedule_list = []
    
    for day_index, day_name in enumerate(days_of_week):
        current_date = week_start_date + timedelta(days=day_index)
        
        day_entry = {
            'day': day_name,
            'date': current_date.strftime('%Y-%m-%d'),
            'department': '',
            'morning_shift': '',
            'evening_shift': '',
            'night_shift': '',
            'job': 'موظف'
        }
        
        # البحث عن إدخال لهذا اليوم في البيانات القديمة
        for old_entry in old_data:
            if isinstance(old_entry, dict):
                if old_entry.get('day') == day_name:
                    # نسخ البيانات المطابقة
                    day_entry.update({
                        'morning_shift': old_entry.get('morning_shift', ''),
                        'evening_shift': old_entry.get('evening_shift', ''),
                        'night_shift': old_entry.get('night_shift', ''),
                        'job': old_entry.get('job', 'موظف'),
                        'department': old_entry.get('department', '')
                    })
                    break
        
        schedule_list.append(day_entry)
    
    return schedule_list

def get_arabic_day_name(day_index):
    """الحصول على اسم اليوم بالعربية"""
    arabic_days = {
        0: 'السبت',
        1: 'الأحد', 
        2: 'الاثنين',
        3: 'الثلاثاء',
        4: 'الأربعاء',
        5: 'الخميس',
        6: 'الجمعة'
    }
    return arabic_days.get(day_index, '')

def migrate_all_schedules_to_template_structure():
    """هجرة جميع الجداول لتبنى على هيكل القالب المحدد في القسم"""
    try:
        print("=== بدء هجرة جميع الجداول إلى هيكل القالب ===")
        
        # الحصول على جميع الجداول الحالية
        existing_schedules = db_session.query(WeeklySchedule).all()
        
        migrated_count = 0
        failed_count = 0
        
        for schedule in existing_schedules:
            try:
                # الحصول على القسم الخاص بالجدول
                department = db_session.query(Department).get(schedule.department_id)
                if not department:
                    print(f"❌ القسم غير موجود للجدول ID: {schedule.id}")
                    failed_count += 1
                    continue
                
                # الحصول على هيكل الجدول المحدد في القسم
                if department.schedule_structure:
                    # استخدام الهيكل المحدد في القسم
                    template_data = json.loads(department.schedule_structure)
                    new_structure = create_schedule_from_template(template_data, department, schedule.week_start_date)
                else:
                    # إذا لم يكن هناك هيكل، إنشاء هيكل افتراضي
                    new_structure = create_default_schedule_structure(
                        department, 
                        [], 
                        schedule.week_start_date
                    )
                
                # تحديث بيانات الجدول بالهيكل الجديد مع الاحتفاظ بتواريخ الورديات القديمة إذا أمكن
                updated_structure = merge_old_shifts_to_new_structure(schedule, new_structure)
                
                # تحديث بيانات الجدول
                schedule.schedule_data = json.dumps(updated_structure, ensure_ascii=False)
                
                # تحديث حالة الجدول ليصبح draft (مسودة) حتى يتم مراجعته
                schedule.status = 'draft'
                schedule.is_approved = False
                schedule.is_locked = False
                
                migrated_count += 1
                print(f"✓ تم هجرة الجدول ID: {schedule.id} للقسم: {department.name}")
                
            except Exception as e:
                print(f"❌ خطأ في هجرة الجدول ID: {schedule.id}: {str(e)}")
                failed_count += 1
                continue
        
        db_session.commit()
        print(f"=== تم هجرة {migrated_count} جدول، فشل {failed_count} جدول ===")
        return migrated_count
        
    except Exception as e:
        print(f"❌ خطأ عام في عملية الهجرة: {str(e)}")
        db_session.rollback()
        return 0

def create_schedule_from_template(template_data, department, week_start_date):
    """إنشاء جدول من القالب المحدد في القسم"""
    try:
        if isinstance(template_data, dict) and 'schedule' in template_data:
            # إذا كان القالب يحتوي على هيكل محدد
            schedule_list = template_data['schedule']
            
            # إنشاء هيكل جديد مع تحديث التواريخ
            new_structure = {
                'department': department.name,
                'week_start_date': week_start_date.strftime('%Y-%m-%d'),
                'week_end_date': (week_start_date + timedelta(days=6)).strftime('%Y-%m-%d'),
                'source': 'template_migration',
                'schedule': []
            }
            
            # تحديث أيام الجدول لتتناسب مع الأسبوع الجديد
            days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
            
            for i, day_name in enumerate(days_of_week):
                current_date = week_start_date + timedelta(days=i)
                
                # البحث عن اليوم المناسب في القالب
                day_template = None
                for template_day in schedule_list:
                    if isinstance(template_day, dict):
                        # مطابقة اليوم بعدة طرق
                        template_day_name = template_day.get('day', '')
                        if (template_day_name == day_name or 
                            template_day_name == get_arabic_day_name(i) or
                            (i < len(schedule_list) and isinstance(schedule_list[i], dict))):
                            day_template = template_day
                            break
                
                # إنشاء إدخال اليوم
                day_entry = {
                    'day': day_name,
                    'date': current_date.strftime('%Y-%m-%d'),
                    'department': department.name,
                    'morning_shift': '',
                    'evening_shift': '',
                    'night_shift': '',
                    'job': 'موظف'
                }
                
                # تعبئة البيانات من القالب إذا وجد
                if day_template:
                    day_entry.update({
                        'morning_shift': day_template.get('morning_shift', ''),
                        'evening_shift': day_template.get('evening_shift', ''),
                        'night_shift': day_template.get('night_shift', ''),
                        'job': day_template.get('job', 'موظف')
                    })
                
                new_structure['schedule'].append(day_entry)
            
            return new_structure
        else:
            # إنشاء هيكل افتراضي إذا لم يكن هناك قالب مناسب
            return create_default_schedule_structure(department, [], week_start_date)
            
    except Exception as e:
        print(f"خطأ في إنشاء الجدول من القالب: {str(e)}")
        return create_default_schedule_structure(department, [], week_start_date)

def merge_old_shifts_to_new_structure(schedule, new_structure):
    """دمج بيانات الشيفتات القديمة مع الهيكل الجديد"""
    try:
        if not schedule.schedule_data:
            return new_structure
        
        # تحليل البيانات القديمة
        old_data = json.loads(schedule.schedule_data)
        
        # إذا كانت البيانات القديمة فارغة، إرجاع الهيكل الجديد كما هو
        if not old_data:
            return new_structure
        
        # الحصول على قسم الجدول
        department = db_session.query(Department).get(schedule.department_id)
        if not department:
            return new_structure
        
        # الحصول على جميع الموظفين في القسم
        employees = db_session.query(User).filter_by(department_id=department.id).all()
        employee_map = {str(emp.id): emp.name for emp in employees}
        
        # دمج البيانات حسب نوع الهيكل القديم
        if isinstance(old_data, dict):
            # الهيكل القديم: {employee_id: {day: shift, ...}}
            return merge_dict_structure(old_data, new_structure, employee_map, schedule.week_start_date)
        elif isinstance(old_data, list):
            # الهيكل القديم: [{day: ..., shift: ...}]
            return merge_list_structure(old_data, new_structure, schedule.week_start_date)
        else:
            return new_structure
            
    except Exception as e:
        print(f"خطأ في دمج البيانات القديمة: {str(e)}")
        return new_structure

def merge_dict_structure(old_data, new_structure, employee_map, week_start_date):
    """دمج الهيكل القديم (قاموس) مع الهيكل الجديد"""
    days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
    
    for i, day_name in enumerate(days_of_week):
        if i >= len(new_structure['schedule']):
            break
            
        # تجميع الشيفتات القدية لهذا اليوم
        morning_shifts = []
        evening_shifts = []
        night_shifts = []
        
        arabic_day = get_arabic_day_name(i)
        
        for employee_id, employee_schedule in old_data.items():
            if isinstance(employee_schedule, dict):
                shift = employee_schedule.get(arabic_day, '')
                if shift:
                    employee_name = employee_map.get(employee_id, f"مستخدم {employee_id}")
                    
                    if 'صباحي' in str(shift):
                        morning_shifts.append(employee_name)
                    elif 'مسائي' in str(shift):
                        evening_shifts.append(employee_name)
                    elif 'ليلي' in str(shift):
                        night_shifts.append(employee_name)
        
        # إذا كان هناك شيفتات قديمة، استبدال البيانات
        if morning_shifts or evening_shifts or night_shifts:
            new_structure['schedule'][i]['morning_shift'] = ', '.join(morning_shifts) if morning_shifts else new_structure['schedule'][i].get('morning_shift', '')
            new_structure['schedule'][i]['evening_shift'] = ', '.join(evening_shifts) if evening_shifts else new_structure['schedule'][i].get('evening_shift', '')
            new_structure['schedule'][i]['night_shift'] = ', '.join(night_shifts) if night_shifts else new_structure['schedule'][i].get('night_shift', '')
    
    return new_structure

def merge_list_structure(old_data, new_structure, week_start_date):
    """دمج الهيكل القديم (قائمة) مع الهيكل الجديد"""
    days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
    
    for i, day_name in enumerate(days_of_week):
        if i >= len(new_structure['schedule']):
            break
            
        # البحث عن اليوم في البيانات القديمة
        for old_entry in old_data:
            if isinstance(old_entry, dict):
                old_day = old_entry.get('day', '')
                # مطابقة اليوم
                if (old_day == day_name or 
                    old_day == get_arabic_day_name(i)):
                    # استبدال البيانات من القديم إلى الجديد
                    if old_entry.get('morning_shift'):
                        new_structure['schedule'][i]['morning_shift'] = old_entry['morning_shift']
                    if old_entry.get('evening_shift'):
                        new_structure['schedule'][i]['evening_shift'] = old_entry['evening_shift']
                    if old_entry.get('night_shift'):
                        new_structure['schedule'][i]['night_shift'] = old_entry['night_shift']
                    if old_entry.get('job'):
                        new_structure['schedule'][i]['job'] = old_entry['job']
                    break
    
    return new_structure

def force_migrate_all_to_template():
    """هجرة قسرية لجميع الجداول باستخدام قالب القسم بدون الحفاظ على البيانات القديمة"""
    try:
        print("=== بدء الهجرة القسرية لجميع الجداول ===")
        
        # الحصول على جميع الجداول الحالية
        existing_schedules = db_session.query(WeeklySchedule).all()
        
        migrated_count = 0
        failed_count = 0
        
        for schedule in existing_schedules:
            try:
                # الحصول على القسم الخاص بالجدول
                department = db_session.query(Department).get(schedule.department_id)
                if not department:
                    print(f"❌ القسم غير موجود للجدول ID: {schedule.id}")
                    failed_count += 1
                    continue
                
                # الحصول على هيكل الجدول المحدد في القسم
                if department.schedule_structure:
                    # استخدام الهيكل المحدد في القسم كما هو
                    template_data = json.loads(department.schedule_structure)
                    new_structure = create_schedule_from_template_exact(template_data, department, schedule.week_start_date)
                else:
                    # إذا لم يكن هناك هيكل، إنشاء هيكل افتراضي
                    new_structure = create_default_schedule_structure(
                        department, 
                        [], 
                        schedule.week_start_date
                    )
                
                # تحديث بيانات الجدول بالهيكل الجديد (بدون دمج)
                schedule.schedule_data = json.dumps(new_structure, ensure_ascii=False)
                
                # تحديث حالة الجدول ليصبح draft (مسودة)
                schedule.status = 'draft'
                schedule.is_approved = False
                schedule.is_locked = False
                
                migrated_count += 1
                print(f"✓ تم هجرة الجدول ID: {schedule.id} للقسم: {department.name}")
                
            except Exception as e:
                print(f"❌ خطأ في هجرة الجدول ID: {schedule.id}: {str(e)}")
                failed_count += 1
                continue
        
        db_session.commit()
        print(f"=== تم هجرة {migrated_count} جدول بنفس هيكل القالب ===")
        return migrated_count
        
    except Exception as e:
        print(f"❌ خطأ عام في عملية الهجرة: {str(e)}")
        db_session.rollback()
        return 0


def migrate_all_schedules_to_new_structure():
    """هجرة جميع الجداول إلى الهيكل الجديد مع الحفاظ على البيانات"""
    try:
        print("=== بدء هجرة جميع الجداول ===")
        
        # الحصول على جميع الجداول الحالية
        existing_schedules = db_session.query(WeeklySchedule).all()
        migrated_count = 0
        preserved_data_count = 0
        
        for schedule in existing_schedules:
            try:
                department = db_session.query(Department).get(schedule.department_id)
                if not department:
                    print(f"⚠️ القسم غير موجود للجدول {schedule.id}")
                    continue
                
                # إنشاء الهيكل الجديد
                new_structure = {
                    'department': department.name,
                    'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
                    'week_end_date': schedule.week_end_date.strftime('%Y-%m-%d'),
                    'source': 'migration_' + datetime.now().strftime('%Y%m%d'),
                    'schedule': []
                }
                
                # إنشاء أيام الأسبوع
                days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
                
                for i, day_name in enumerate(days_of_week):
                    current_date = schedule.week_start_date + timedelta(days=i)
                    
                    day_entry = {
                        'day': day_name,
                        'date': current_date.strftime('%Y-%m-%d'),
                        'department': department.name,
                        'morning_shift': '',
                        'evening_shift': '',
                        'night_shift': '',
                        'job': 'موظف'
                    }
                    
                    # إذا كانت هناك بيانات قديمة، حاول استخراجها
                    if schedule.schedule_data:
                        try:
                            old_data = json.loads(schedule.schedule_data)
                            preserved_data_count += 1
                            
                            # استخراج البيانات من الهيكل القديم
                            extracted_data = extract_data_from_old_structure(old_data, day_name, i, department.id)
                            if extracted_data:
                                day_entry.update(extracted_data)
                        except:
                            pass
                    
                    new_structure['schedule'].append(day_entry)
                
                # تحديث الجدول بالهيكل الجديد
                schedule.schedule_data = json.dumps(new_structure, ensure_ascii=False)
                schedule.status = 'draft'
                schedule.is_approved = False
                schedule.is_locked = False
                
                migrated_count += 1
                print(f"✅ تم هجرة الجدول {schedule.id} للقسم {department.name}")
                
            except Exception as e:
                print(f"❌ خطأ في هجرة الجدول {schedule.id}: {str(e)}")
                continue
        
        db_session.commit()
        
        print(f"=== تم هجرة {migrated_count} جدول ===")
        print(f"=== تم الحفاظ على بيانات من {preserved_data_count} جدول ===")
        
        return migrated_count, preserved_data_count
        
    except Exception as e:
        db_session.rollback()
        print(f"❌ خطأ عام في الهجرة: {str(e)}")
        return 0, 0

def extract_data_from_old_structure(old_data, day_name, day_index, department_id):
    """استخراج البيانات من الهيكل القديم"""
    try:
        extracted = {}
        arabic_days = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        
        if isinstance(old_data, dict):
            # الهيكل القديم: {employee_id: {day: shift, ...}}
            morning_shifts = []
            evening_shifts = []
            night_shifts = []
            
            for emp_id, emp_schedule in old_data.items():
                if isinstance(emp_schedule, dict):
                    # الحصول على اسم اليوم العربي
                    arabic_day = arabic_days[day_index]
                    shift = emp_schedule.get(arabic_day, '')
                    
                    if shift:
                        # الحصول على اسم الموظف
                        try:
                            emp = db_session.query(User).get(int(emp_id))
                            emp_name = emp.name if emp else f"مستخدم {emp_id}"
                        except:
                            emp_name = f"مستخدم {emp_id}"
                        
                        if 'صباحي' in str(shift):
                            morning_shifts.append(emp_name)
                        elif 'مسائي' in str(shift):
                            evening_shifts.append(emp_name)
                        elif 'ليلي' in str(shift) or 'سهر' in str(shift):
                            night_shifts.append(emp_name)
            
            if morning_shifts:
                extracted['morning_shift'] = ', '.join(morning_shifts)
            if evening_shifts:
                extracted['evening_shift'] = ', '.join(evening_shifts)
            if night_shifts:
                extracted['night_shift'] = ', '.join(night_shifts)
                
        elif isinstance(old_data, list):
            # الهيكل القديم: [{day: ..., shift: ...}]
            for item in old_data:
                if isinstance(item, dict):
                    item_day = item.get('day') or item.get('اليوم')
                    if item_day == day_name:
                        # استخراج بيانات الشيفتات
                        shift_fields = {
                            'morning_shift': ['morning_shift', 'الشيفت الصباحي'],
                            'evening_shift': ['evening_shift', 'الشيفت المسائي'],
                            'night_shift': ['night_shift', 'شيفت السهر']
                        }
                        
                        for new_field, old_fields in shift_fields.items():
                            for old_field in old_fields:
                                if old_field in item and item[old_field]:
                                    extracted[new_field] = item[old_field]
                                    break
                        
                        if 'job' in item and item['job']:
                            extracted['job'] = item['job']
                        break
        
        return extracted
        
    except Exception as e:
        print(f"⚠️ خطأ في استخراج البيانات: {str(e)}")
        return {}

def force_migrate_all_to_template():
    """الهجرة القسرية باستخدام قوالب القسم"""
    try:
        print("=== بدء الهجرة القسرية ===")
        
        existing_schedules = db_session.query(WeeklySchedule).all()
        migrated_count = 0
        
        for schedule in existing_schedules:
            try:
                department = db_session.query(Department).get(schedule.department_id)
                if not department:
                    continue
                
                # إنشاء هيكل جديد من القالب
                new_structure = create_schedule_from_template(department, schedule.week_start_date)
                
                # تحديث الجدول
                schedule.schedule_data = json.dumps(new_structure, ensure_ascii=False)
                schedule.status = 'draft'
                schedule.is_approved = False
                schedule.is_locked = False
                
                migrated_count += 1
                print(f"✅ تم هجرة الجدول {schedule.id} للقسم {department.name}")
                
            except Exception as e:
                print(f"❌ خطأ في هجرة الجدول {schedule.id}: {str(e)}")
                continue
        
        db_session.commit()
        print(f"=== تم هجرة {migrated_count} جدول قسرياً ===")
        return migrated_count
        
    except Exception as e:
        db_session.rollback()
        print(f"❌ خطأ عام في الهجرة القسرية: {str(e)}")
        return 0

def create_schedule_from_template(department, week_start_date):
    """إنشاء جدول من قالب القسم"""
    week_end_date = week_start_date + timedelta(days=6)
    
    # أيام الأسبوع
    days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
    
    schedule_list = []
    for i, day_name in enumerate(days_of_week):
        current_date = week_start_date + timedelta(days=i)
        
        day_entry = {
            'day': day_name,
            'date': current_date.strftime('%Y-%m-%d'),
            'department': department.name,
            'morning_shift': '',
            'evening_shift': '',
            'night_shift': '',
            'job': 'موظف'
        }
        
        schedule_list.append(day_entry)
    
    return {
        'department': department.name,
        'week_start_date': week_start_date.strftime('%Y-%m-%d'),
        'week_end_date': week_end_date.strftime('%Y-%m-%d'),
        'source': 'template_migration',
        'schedule': schedule_list
    }

def analyze_schedule_structures():
    """تحليل هياكل الجداول الحالية"""
    try:
        schedules = db_session.query(WeeklySchedule).all()
        structure_types = {}
        
        for schedule in schedules:
            if not schedule.schedule_data:
                structure_types['empty'] = structure_types.get('empty', 0) + 1
                continue
            
            try:
                data = json.loads(schedule.schedule_data)
                
                if isinstance(data, dict):
                    if 'schedule' in data:
                        key = 'new_structure'
                    else:
                        key = 'old_dict_structure'
                elif isinstance(data, list):
                    key = 'list_structure'
                else:
                    key = 'other_structure'
                    
            except:
                key = 'invalid_structure'
            
            structure_types[key] = structure_types.get(key, 0) + 1
        
        return {
            'success': True,
            'structure_types': structure_types,
            'total_schedules': len(schedules)
        }
        
    except Exception as e:
        return {
            'success': False,
            'message': str(e)
        }

def create_schedule_from_template_exact(template_data, department, week_start_date):
    """إنشاء جدول مطابق تماماً للقالب المحدد في القسم"""
    try:
        # إذا كان القالب يحتوي على بيانات
        if isinstance(template_data, dict) and 'schedule' in template_data:
            # نسخ الهيكل كما هو مع تحديث التواريخ فقط
            new_structure = {
                'department': department.name,
                'week_start_date': week_start_date.strftime('%Y-%m-%d'),
                'week_end_date': (week_start_date + timedelta(days=6)).strftime('%Y-%m-%d'),
                'source': 'exact_template_copy',
                'schedule': []
            }
            
            # نسخ أيام القالب مع تحديث التواريخ فقط
            days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
            
            for i, day_name in enumerate(days_of_week):
                current_date = week_start_date + timedelta(days=i)
                
                # البحث عن القالب المناسب لهذا اليوم
                day_template = None
                for template_day in template_data['schedule']:
                    if isinstance(template_day, dict):
                        template_day_name = template_day.get('day', '')
                        if template_day_name == day_name:
                            day_template = template_day
                            break
                
                if day_template:
                    # نسخ اليوم كما هو مع تحديث التاريخ فقط
                    day_entry = day_template.copy()
                    day_entry['date'] = current_date.strftime('%Y-%m-%d')
                    day_entry['department'] = department.name
                    new_structure['schedule'].append(day_entry)
                else:
                    # إذا لم يكن هناك قالب لهذا اليوم، إنشاء يوم افتراضي
                    day_entry = {
                        'day': day_name,
                        'date': current_date.strftime('%Y-%m-%d'),
                        'department': department.name,
                        'morning_shift': '',
                        'evening_shift': '',
                        'night_shift': '',
                        'job': 'موظف'
                    }
                    new_structure['schedule'].append(day_entry)
            
            return new_structure
        else:
            # إنشاء هيكل افتراضي
            return create_default_schedule_structure(department, [], week_start_date)
            
    except Exception as e:
        print(f"خطأ في نسخ القالب: {str(e)}")
        return create_default_schedule_structure(department, [], week_start_date)





def check_department_templates():
    """فحص حالة القوالب في الأقسام"""
    try:
        departments = db_session.query(Department).all()
        template_status = {}
        
        for department in departments:
            has_template = bool(department.schedule_structure)
            template_info = {
                'has_template': has_template,
                'department_name': department.name,
                'template_data': None
            }
            
            if has_template:
                try:
                    template_data = json.loads(department.schedule_structure)
                    if isinstance(template_data, dict) and 'schedule' in template_data:
                        template_info['template_data'] = {
                            'days_count': len(template_data['schedule']),
                            'sample_day': template_data['schedule'][0] if template_data['schedule'] else None
                        }
                except:
                    template_info['has_template'] = False
            
            template_status[department.id] = template_info
        
        return template_status
    except Exception as e:
        print(f"خطأ في فحص القوالب: {str(e)}")
        return {}


@app.route('/admin/execute_comprehensive_migration', methods=['POST'])
@login_required
def admin_execute_comprehensive_migration():
    """Execute comprehensive migration that reads ALL table data"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        # Run comprehensive migration
        from migrate_schedules import (
            analyze_schedule_data_structure,
            migrate_all_schedules_with_data,
            create_future_schedules,
            generate_migration_report
        )
        
        # Step 1: Analyze
        structure_types = analyze_schedule_data_structure()
        
        # Step 2: Migrate with data preservation
        migrated_count, preserved_count = migrate_all_schedules_with_data()
        
        # Step 3: Create future schedules
        future_count = create_future_schedules()
        
        # Step 4: Generate report
        report = generate_migration_report(migrated_count, preserved_count, future_count)
        
        return jsonify({
            'success': True,
            'message': f'تم الهجرة الشاملة لـ {migrated_count} جدول مع الحفاظ على بيانات {preserved_count} جدول',
            'migrated_count': migrated_count,
            'preserved_count': preserved_count,
            'future_count': future_count,
            'report': report
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في الهجرة الشاملة: {str(e)}'
        })


@app.route('/admin/execute_migration', methods=['POST'])
@login_required
def admin_execute_migration():
    """تنفيذ عملية هجرة الجداول"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        migration_type = request.json.get('type', 'smart')
        
        if migration_type == 'force':
            # الهجرة القسرية
            migrated_count = force_migrate_all_to_template()
            message = f'تم الهجرة القسرية لـ {migrated_count} جدول'
        else:
            # الهجرة الذكية مع الحفاظ على البيانات
            migrated_count, preserved_count = migrate_all_schedules_to_new_structure()
            message = f'تم هجرة {migrated_count} جدول مع الحفاظ على بيانات من {preserved_count} جدول'
        
        # إنشاء الجداول المستقبلية
        future_count = generate_future_schedules_after_migration()
        
        return jsonify({
            'success': True,
            'message': message + f' وتم إنشاء {future_count} جدول مستقبلي',
            'migrated_count': migrated_count,
            'future_count': future_count,
            'type': migration_type
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في الهجرة: {str(e)}'
        })

@app.route('/admin/analyze_schedules')
@login_required
def admin_analyze_schedules():
    """تحليل هياكل الجداول"""
    if not current_user.is_admin:
        return jsonify({'success': False})
    
    result = analyze_schedule_structures()
    return jsonify(result)

def generate_future_schedules_after_migration():
    """إنشاء الجداول المستقبلية بعد الهجرة"""
    try:
        print("=== بدء إنشاء الجداول المستقبلية ===")
        
        departments = db_session.query(Department).filter_by(auto_generate_schedule=True).all()
        created_count = 0
        
        for department in departments:
            # الحصول على أحدث جدول للقسم
            latest_schedule = db_session.query(WeeklySchedule).filter_by(
                department_id=department.id
            ).order_by(WeeklySchedule.week_start_date.desc()).first()
            
            if not latest_schedule:
                continue
            
            # حساب الأسبوع القادم
            today = date.today()
            days_since_saturday = (today.weekday() - 5) % 7
            current_week_start = today - timedelta(days=days_since_saturday)
            
            # إنشاء جداول للأسابيع القادمة
            for week_offset in range(1, 5):
                week_start = current_week_start + timedelta(days=7 * week_offset)
                week_end = week_start + timedelta(days=6)
                
                # التحقق من عدم وجود جدول لهذا الأسبوع
                existing = db_session.query(WeeklySchedule).filter_by(
                    department_id=department.id,
                    week_start_date=week_start
                ).first()
                
                if not existing:
                    # إنشاء جدول جديد بنفس هيكل أحدث جدول
                    new_schedule = WeeklySchedule(
                        department_id=department.id,
                        week_start_date=week_start,
                        week_end_date=week_end,
                        schedule_data=latest_schedule.schedule_data,
                        created_by=1,  # النظام
                        status='draft',
                        is_approved=False,
                        is_locked=False
                    )
                    
                    db_session.add(new_schedule)
                    created_count += 1
        
        db_session.commit()
        print(f"=== تم إنشاء {created_count} جدول مستقبلي ===")
        return created_count
        
    except Exception as e:
        db_session.rollback()
        print(f"❌ خطأ في إنشاء الجداول المستقبلية: {str(e)}")
        return 0
    
@app.route('/admin/preview_migration')
@login_required
def admin_preview_migration():
    """Preview what will happen during migration"""
    if not current_user.is_admin:
        return jsonify({'success': False})
    
    try:
        schedules = db_session.query(WeeklySchedule).limit(5).all()
        preview_data = []
        
        for schedule in schedules:
            schedule_info = {
                'id': schedule.id,
                'department_id': schedule.department_id,
                'week_start': schedule.week_start_date.strftime('%Y-%m-%d') if schedule.week_start_date else '',
                'has_data': bool(schedule.schedule_data),
                'status': schedule.status,
                'is_approved': schedule.is_approved,
                'data_preview': ''
            }
            
            if schedule.schedule_data:
                try:
                    data = json.loads(schedule.schedule_data)
                    # Summarize data
                    if isinstance(data, dict):
                        if 'schedule' in data:
                            schedule_info['structure'] = 'new_format'
                            schedule_info['days_count'] = len(data.get('schedule', []))
                        else:
                            schedule_info['structure'] = 'old_dict_format'
                            schedule_info['employee_count'] = len(data)
                    elif isinstance(data, list):
                        schedule_info['structure'] = 'list_format'
                        schedule_info['items_count'] = len(data)
                        
                    # Show a preview
                    preview = str(data)[:100] + '...' if len(str(data)) > 100 else str(data)
                    schedule_info['data_preview'] = preview
                    
                except:
                    schedule_info['data_preview'] = 'Invalid JSON'
            
            preview_data.append(schedule_info)
        
        return jsonify({
            'success': True,
            'preview': preview_data,
            'total_schedules': db_session.query(WeeklySchedule).count()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في المعاينة: {str(e)}'
        })

def check_and_create_future_schedules():
    """Create future schedules if they don't exist"""
    try:
        departments = db_session.query(Department).filter_by(auto_generate_schedule=True).all()
        created_count = 0
        
        for department in departments:
            # Calculate next week start (Saturday)
            today = date.today()
            days_since_saturday = (today.weekday() - 5) % 7
            current_week_start = today - timedelta(days=days_since_saturday)
            
            # Create schedules for the next 4 weeks
            for week_offset in range(1, 5):
                week_start_date = current_week_start + timedelta(days=7 * week_offset)
                
                # Check if schedule exists
                existing = db_session.query(WeeklySchedule).filter_by(
                    department_id=department.id,
                    week_start_date=week_start_date
                ).first()
                
                if not existing:
                    # Create new schedule from template
                    new_schedule = create_weekly_schedule_from_structure(
                        department.id, week_start_date
                    )
                    if new_schedule:
                        created_count += 1
                        print(f"Created schedule for {department.name} week starting {week_start_date}")
        
        if created_count > 0:
            print(f"Created {created_count} future schedules")
        return created_count
        
    except Exception as e:
        print(f"Error creating future schedules: {e}")
        return 0

def verify_schedule_structure(schedule_id):
    """Verify and fix schedule structure"""
    schedule = db_session.query(WeeklySchedule).get(schedule_id)
    if not schedule or not schedule.schedule_data:
        return False
    
    try:
        data = json.loads(schedule.schedule_data)
        
        # Check if data has the new structure
        if isinstance(data, dict) and 'schedule' in data:
            # New structure is correct
            return True
        elif isinstance(data, dict) and 'schedule' not in data:
            # Old structure needs migration
            new_data = migrate_single_schedule(schedule, data)
            schedule.schedule_data = json.dumps(new_data, ensure_ascii=False)
            db_session.commit()
            print(f"Migrated schedule {schedule_id} to new structure")
            return True
        else:
            # Unknown structure, create default
            department = db_session.query(Department).get(schedule.department_id)
            employees = db_session.query(User).filter_by(
                department_id=department.id,
                is_admin=False
            ).all() if department else []
            
            default_data = create_default_schedule_structure(
                department, employees, schedule.week_start_date
            )
            schedule.schedule_data = json.dumps(default_data, ensure_ascii=False)
            db_session.commit()
            print(f"Reset schedule {schedule_id} to default structure")
            return True
            
    except Exception as e:
        print(f"Error verifying schedule {schedule_id}: {e}")
        return False

def migrate_single_schedule(schedule, old_data):
    """Migrate single schedule from old to new structure"""
    try:
        department = db_session.query(Department).get(schedule.department_id)
        
        # Default new structure
        new_structure = {
            'department': department.name if department else f"القسم {schedule.department_id}",
            'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
            'week_end_date': schedule.week_end_date.strftime('%Y-%m-%d'),
            'source': 'migrated',
            'schedule': []
        }
        
        # Days of week
        days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        
        for i, day_name in enumerate(days_of_week):
            current_date = schedule.week_start_date + timedelta(days=i)
            
            day_entry = {
                'day': day_name,
                'date': current_date.strftime('%Y-%m-%d'),
                'department': department.name if department else '',
                'morning_shift': '',
                'evening_shift': '',
                'night_shift': '',
                'job': 'موظف'
            }
            
            # Try to extract data from old structure
            if isinstance(old_data, dict):
                # Try employee-based structure
                morning_shifts = []
                evening_shifts = []
                night_shifts = []
                
                for emp_id, emp_data in old_data.items():
                    if isinstance(emp_data, dict):
                        # Get day in Arabic
                        arabic_days = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
                        if i < len(arabic_days):
                            arabic_day = arabic_days[i]
                            shift = emp_data.get(arabic_day, '')
                            
                            # Get employee name
                            try:
                                emp = db_session.query(User).get(int(emp_id))
                                emp_name = emp.name if emp else f"مستخدم {emp_id}"
                            except:
                                emp_name = f"مستخدم {emp_id}"
                            
                            if 'صباحي' in str(shift):
                                morning_shifts.append(emp_name)
                            elif 'مسائي' in str(shift):
                                evening_shifts.append(emp_name)
                            elif 'ليلي' in str(shift):
                                night_shifts.append(emp_name)
                
                # Join the shifts
                if morning_shifts:
                    day_entry['morning_shift'] = ', '.join(morning_shifts)
                if evening_shifts:
                    day_entry['evening_shift'] = ', '.join(evening_shifts)
                if night_shifts:
                    day_entry['night_shift'] = ', '.join(night_shifts)
            
            new_structure['schedule'].append(day_entry)
        
        return new_structure
        
    except Exception as e:
        print(f"Error migrating single schedule: {e}")
        # Return default structure on error
        return create_default_schedule_structure(
            db_session.query(Department).get(schedule.department_id),
            [],
            schedule.week_start_date
        )

@app.route('/admin/check_templates')
@login_required
def admin_check_templates():
    """فحص حالة القوالب"""
    if not current_user.is_admin:
        return jsonify({'success': False})
    
    template_status = check_department_templates()
    
    # إحصائيات
    departments_with_template = sum(1 for info in template_status.values() if info['has_template'])
    total_departments = len(template_status)
    
    return jsonify({
        'success': True,
        'template_status': template_status,
        'stats': {
            'total_departments': total_departments,
            'departments_with_template': departments_with_template,
            'departments_without_template': total_departments - departments_with_template
        }
    })

def generate_future_after_migration():
    """إنشاء الجداول المستقبلية بعد الهجرة"""
    try:
        print("=== بدء إنشاء الجداول المستقبلية بعد الهجرة ===")
        
        # تفعيل التوليد التلقائي لجميع الأقسام
        departments = db_session.query(Department).all()
        for department in departments:
            if not department.auto_generate_schedule:
                department.auto_generate_schedule = True
        
        db_session.commit()
        
        # إنشاء الجداول المستقبلية
        generated_count = auto_generate_weekly_schedules()
        
        print(f"=== تم إنشاء {generated_count} جدول مستقبلي ===")
        return generated_count
        
    except Exception as e:
        print(f"❌ خطأ في إنشاء الجداول المستقبلية: {str(e)}")
        return 0



@app.route('/admin/departments', methods=['GET', 'POST'])
@login_required
def admin_departments():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    # Handle POST request for creating department
    if request.method == 'POST':
        try:
            name = request.form['name']
            primary_manager_id = request.form.get('primary_manager_id') or None
            
            # Handle empty values for numeric fields
            max_advance_amount_str = request.form.get('max_advance_amount', '0')
            max_installments_str = request.form.get('max_installments', '1')
            
            max_advance_amount = float(max_advance_amount_str) if max_advance_amount_str else 0.0
            max_installments = int(max_installments_str) if max_installments_str else 1
            
            department = Department(
                name=name,
                primary_manager_id=primary_manager_id,
                advance_policy_max_amount=max_advance_amount,
                advance_policy_max_installments=max_installments,
                created_by=current_user.id
            )
            
            db_session.add(department)
            db_session.commit()
            
            flash('تم إنشاء القسم بنجاح', 'success')
            return redirect(url_for('admin_departments'))
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'error')
            return redirect(url_for('admin_departments'))
    
    # For GET requests, show the page
    try:
        # Get all departments
        departments = db_session.query(Department).all()
        
        # Get all users
        users = db_session.query(User).all()
        
        # Get all managers (users with is_manager=True)
        managers = db_session.query(User).filter_by(is_manager=True).all()
        
        # Prepare department data with managers
        departments_data = []
        for department in departments:
            # Create a dictionary with department info
            dept_data = {
                'id': department.id,
                'name': department.name,
                'primary_manager_id': department.primary_manager_id,
                'employee_count': db_session.query(User).filter_by(
                    department_id=department.id,
                    is_admin=False
                ).count(),
                'managers': []
            }
            
            # Get managers for this department
            dept_managers = db_session.query(DepartmentManager).filter_by(
                department_id=department.id
            ).all()
            
            # Add user info for each manager
            for dept_manager in dept_managers:
                user = db_session.get(User, dept_manager.user_id)  # Use get() instead of query().get()
                if user:
                    dept_data['managers'].append({
                        'manager_id': dept_manager.id,
                        'user_id': dept_manager.user_id,
                        'user_name': user.name,
                        'user_username': user.username,
                        'can_manage_schedules': dept_manager.can_manage_schedules,
                        'can_manage_leaves': dept_manager.can_manage_leaves,
                        'can_manage_permissions': dept_manager.can_manage_permissions,
                        'can_manage_advances': dept_manager.can_manage_advances,
                        'can_manage_rewards': dept_manager.can_manage_rewards,
                        'can_view_reports': dept_manager.can_view_reports,
                    })
            
            departments_data.append(dept_data)
        
        notifications = get_user_notifications(current_user.id)
        
        return render_template('admin/admin_departments.html',
                             departments=departments_data,  # Pass prepared data instead of ORM objects
                             users=users,
                             managers=managers,
                             notifications=notifications)
                             
    except Exception as e:
        flash(f'حدث خطأ في تحميل البيانات: {str(e)}', 'error')
        import traceback
        print(f"Error details: {traceback.format_exc()}")
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/add_manager_to_department', methods=['POST'])
@login_required
def admin_add_manager_to_department():
    """Add a manager to a department with specific permissions"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        department_id = request.form['department_id']
        user_id = request.form['user_id']
        
        # Check if user is already a manager for this department
        existing_manager = db_session.query(DepartmentManager).filter_by(
            department_id=department_id,
            user_id=user_id
        ).first()
        
        if existing_manager:
            flash('هذا المستخدم مدير بالفعل لهذا القسم', 'warning')
            return redirect(url_for('admin_departments'))
        
        # Get permissions from form
        can_manage_schedules = 'can_manage_schedules' in request.form
        can_manage_leaves = 'can_manage_leaves' in request.form
        can_manage_permissions = 'can_manage_permissions' in request.form
        can_manage_advances = 'can_manage_advances' in request.form
        can_manage_rewards = 'can_manage_rewards' in request.form
        can_view_reports = 'can_view_reports' in request.form
        
        # Check if should be primary manager
        if 'is_primary_manager' in request.form:
            department = db_session.get(Department, department_id)  # Use get() instead of query().get()
            if department:
                department.primary_manager_id = user_id
        
        # Create department manager record
        department_manager = DepartmentManager(
            department_id=department_id,
            user_id=user_id,
            can_manage_schedules=can_manage_schedules,
            can_manage_leaves=can_manage_leaves,
            can_manage_permissions=can_manage_permissions,
            can_manage_advances=can_manage_advances,
            can_manage_rewards=can_manage_rewards,
            can_view_reports=can_view_reports,
            created_by=current_user.id
        )
        
        db_session.add(department_manager)
        
        # Update user to be a manager if not already
        user = db_session.get(User, user_id)  # Use get() instead of query().get()
        if user and not user.is_manager:
            user.is_manager = True
        
        db_session.commit()
        
        flash('تم إضافة المدير بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'error')
        import traceback
        print(f"Error in add_manager_to_department: {traceback.format_exc()}")
    
    return redirect(url_for('admin_departments'))

@app.route('/admin/remove_department_manager/<int:manager_id>')
@login_required
def admin_remove_department_manager(manager_id):
    """Remove a manager from a department"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        manager = db_session.get(DepartmentManager, manager_id)  # Use get() instead of query().get()
        if not manager:
            flash('المدير غير موجود', 'error')
            return redirect(url_for('admin_departments'))
        
        # Check if this manager is the primary manager
        department = db_session.get(Department, manager.department_id)
        if department and department.primary_manager_id == manager.user_id:
            department.primary_manager_id = None
        
        # Remove the manager record
        db_session.delete(manager)
        
        # Check if user is manager of any other department
        other_managements = db_session.query(DepartmentManager).filter_by(
            user_id=manager.user_id
        ).count()
        
        # If not managing any other department, remove manager role
        if other_managements == 0:
            user = db_session.get(User, manager.user_id)
            if user:
                user.is_manager = False
        
        db_session.commit()
        flash('تم حذف المدير بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('admin_departments'))

@app.route('/admin/set_primary_manager/<int:department_id>', methods=['POST'])
@login_required
def admin_set_primary_manager(department_id):
    """Set or remove primary manager for a department"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        manager_id = request.form.get('manager_id')
        department = db_session.get(Department, department_id)  # Use get() instead of query().get()
        
        if not department:
            flash('القسم غير موجود', 'error')
            return redirect(url_for('admin_departments'))
        
        if manager_id:
            # Verify manager exists in this department
            manager = db_session.query(DepartmentManager).filter_by(
                department_id=department_id,
                user_id=manager_id
            ).first()
            
            if not manager:
                flash('المدير غير موجود في هذا القسم', 'error')
                return redirect(url_for('admin_departments'))
            
            department.primary_manager_id = manager_id
        else:
            department.primary_manager_id = None
        
        db_session.commit()
        flash('تم تحديث المدير الرئيسي بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('admin_departments'))

def get_department_managers_with_users(department_id):
    """Get all managers for a department with their user info"""
    return db_session.query(
        DepartmentManager, User
    ).join(
        User, DepartmentManager.user_id == User.id
    ).filter(
        DepartmentManager.department_id == department_id
    ).all()


@app.route('/admin/schedule_structure/<int:dept_id>', methods=['GET', 'POST'])
@login_required
def admin_schedule_structure(dept_id):
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    department = db_session.query(Department).get(dept_id)
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('admin_departments'))
    
    if request.method == 'POST':
        # Handle Excel file upload
        if 'excel_file' in request.files:
            excel_file = request.files['excel_file']
            if excel_file and excel_file.filename != '' and excel_file.filename.endswith(('.xlsx', '.xls')):
                try:
                    # Read Excel file
                    df = pd.read_excel(excel_file)
                    
                    # Convert DataFrame to dictionary structure
                    schedule_data = []
                    for _, row in df.iterrows():
                        schedule_row = {
                            'night_shift': str(row.get('شيفت السهر', '')),
                            'evening_shift': str(row.get('الشيفت المسائي', '')),
                            'morning_shift': str(row.get('الشيفت الصباحي', '')),
                            'job': str(row.get('الوظيفة', 'موظف')),
                            'department': str(row.get('القسم', department.name)),
                            'date': str(row.get('التاريخ', '')),
                            'day': str(row.get('اليوم', ''))
                        }
                        schedule_data.append(schedule_row)
                    
                    # Update department with Excel structure
                    department.schedule_structure = json.dumps({
                        'source': 'excel',
                        'columns': ['شيفت السهر', 'الشيفت المسائي', 'الشيفت الصباحي', 'الوظيفة', 'القسم', 'التاريخ', 'اليوم'],
                        'structure': schedule_data
                    })
                    department.auto_generate_schedule = 'auto_generate' in request.form
                    
                    # Update all existing schedules for this department
                    if 'update_existing' in request.form and request.form['update_existing'] == 'yes':
                        update_all_schedules_to_match_structure(dept_id, schedule_data)
                    
                    db_session.commit()
                    
                    flash('تم حفظ هيكل الجدول من ملف Excel بنجاح')
                    return redirect(url_for('admin_departments'))
                    
                except Exception as e:
                    flash(f'خطأ في قراءة ملف Excel: {str(e)}', 'error')
                    return redirect(url_for('admin_schedule_structure', dept_id=dept_id))
        
        # Handle manual structure data
        elif 'schedule_structure' in request.form:
            try:
                schedule_structure = request.form['schedule_structure']
                structure_data = json.loads(schedule_structure)
                
                # Ensure proper structure format
                if not isinstance(structure_data, dict):
                    structure_data = {
                        'source': 'manual',
                        'structure': structure_data if isinstance(structure_data, list) else []
                    }
                
                department.schedule_structure = json.dumps(structure_data)
                department.auto_generate_schedule = 'auto_generate' in request.form
                
                # Update all existing schedules for this department
                if 'update_existing' in request.form and request.form['update_existing'] == 'yes':
                    update_all_schedules_to_match_structure(dept_id, structure_data.get('structure', []))
                
                db_session.commit()
                
                flash('تم حفظ هيكل الجدول بنجاح')
                return redirect(url_for('admin_departments'))
                
            except Exception as e:
                flash(f'خطأ في حفظ البيانات: {str(e)}', 'error')
                return redirect(url_for('admin_schedule_structure', dept_id=dept_id))
        
        flash('يرجى تحميل ملف Excel صحيح أو إدخال البيانات يدوياً', 'error')
        return redirect(url_for('admin_schedule_structure', dept_id=dept_id))
    
    # Get employees for this department
    employees = db_session.query(User).filter_by(department_id=dept_id, is_admin=False).all()
    employee_data = [{'id': emp.id, 'name': emp.name} for emp in employees] if employees else []
    
    # Get existing structure or create default
    if department.schedule_structure:
        try:
            schedule_data = json.loads(department.schedule_structure)
            # Ensure all values are serializable
            if 'structure' in schedule_data:
                for item in schedule_data['structure']:
                    for key, value in item.items():
                        if value is None:
                            item[key] = ''
        except:
            schedule_data = {"source": "manual", "structure": []}
    else:
        schedule_data = {
            "source": "manual",
            "structure": create_default_schedule_days(date.today(), department.name)
        }
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_schedule_structure.html',
                         department=department,
                         schedule_data=schedule_data,
                         employees=employee_data,
                         notifications=notifications)


def update_all_schedules_to_match_structure(dept_id, structure_data):
    """
    تحديث جميع الجداول الأسبوعية الموجودة لتطابق الهيكل الجديد
    """
    try:
        # الحصول على جميع الجداول الأسبوعية للقسم
        from models import WeeklySchedule  # تأكد من استيراد النموذج
        
        weekly_schedules = db_session.query(WeeklySchedule).filter_by(
            department_id=dept_id
        ).all()
        
        updated_count = 0
        for schedule in weekly_schedules:
            if schedule.schedule_data:
                try:
                    existing_data = json.loads(schedule.schedule_data)
                    
                    # إنشاء هيكل جديد بناءً على النموذج المحدد
                    new_structure = []
                    
                    for structure_item in structure_data:
                        # البحث عن سطر مطابق في الجدول الحالي
                        matched_row = None
                        if isinstance(existing_data, list):
                            for existing_item in existing_data:
                                # المطابقة بناءً على التاريخ أو اليوم
                                if 'date' in structure_item and 'date' in existing_item:
                                    if structure_item.get('date') == existing_item.get('date'):
                                        matched_row = existing_item
                                        break
                                elif 'day' in structure_item and 'day' in existing_item:
                                    if structure_item.get('day') == existing_item.get('day'):
                                        matched_row = existing_item
                                        break
                        
                        # إذا وجدنا صف مطابق، نستخدم البيانات الموجودة
                        if matched_row:
                            new_row = {}
                            for key in ['night_shift', 'evening_shift', 'morning_shift', 'job', 'department', 'date', 'day']:
                                new_row[key] = matched_row.get(key, structure_item.get(key, ''))
                        else:
                            # استخدام البيانات من الهيكل الجديد
                            new_row = structure_item.copy()
                        
                        new_structure.append(new_row)
                    
                    # تحديث البيانات
                    schedule.schedule_data = json.dumps(new_structure)
                    updated_count += 1
                    
                except Exception as e:
                    print(f"خطأ في تحديث الجدول {schedule.id}: {str(e)}")
                    continue
        
        return updated_count
    
    except Exception as e:
        print(f"خطأ في تحديث الجداول: {str(e)}")
        return 0

def update_single_schedule_to_match_structure(schedule_id, structure_data):
    """
    تحديث جدول أسبوعي محدد لتطابق الهيكل
    """
    try:
        from models import WeeklySchedule
        
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return False
        
        # إنشاء هيكل جديد بناءً على النموذج
        new_structure = []
        
        for structure_item in structure_data:
            new_row = structure_item.copy()
            new_structure.append(new_row)
        
        # تحديث البيانات
        schedule.schedule_data = json.dumps(new_structure)
        db_session.commit()
        
        return True
    
    except Exception as e:
        print(f"خطأ في تحديث الجدول: {str(e)}")
        return False


@app.route('/admin/update_all_schedules/<int:dept_id>', methods=['POST'])
@login_required
def update_all_schedules(dept_id):
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    department = db_session.query(Department).get(dept_id)
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('admin_departments'))
    
    # الحصول على هيكل الجدول
    if not department.schedule_structure:
        flash('لم يتم تعيين هيكل للجدول بعد', 'error')
        return redirect(url_for('admin_departments'))
    
    try:
        structure_data = json.loads(department.schedule_structure)
        schedule_structure = structure_data.get('structure', [])
        
        # تحديث جميع الجداول
        updated_count = update_all_schedules_to_match_structure(dept_id, schedule_structure)
        db_session.commit()
        
        flash(f'تم تحديث {updated_count} جدول أسبوعي لتطابق الهيكل الجديد', 'success')
        
    except Exception as e:
        flash(f'خطأ في تحديث الجداول: {str(e)}', 'error')
    
    return redirect(url_for('admin_departments'))

@app.route('/admin/employee_progress')
@login_required
def admin_employee_progress():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    users = db_session.query(User).filter_by(is_admin=False).all()
    employee_data_list = []
    
    for user in users:
        data = db_session.query(EmployeeData).filter_by(user_id=user.id).first()
        if data:
            # استخدم الدوال من النموذج مباشرة
            data.calculate_completion()
            employee_data_list.append({
                'user': user,
                'data': data,
                'missing_fields': data.get_missing_fields()  # استدعاء الدالة من النموذج
            })
        else:
            # إذا لم يكن هناك بيانات للموظف، أضفه كـ "لم يبدأ"
            employee_data_list.append({
                'user': user,
                'data': None,
                'missing_fields': []
            })
    
    db_session.commit()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_employee_progress.html',
                         employee_data_list=employee_data_list,
                         notifications=notifications)

@app.route('/admin/request_data_update/<int:user_id>')
@login_required
def request_data_update(user_id):
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    employee_data = db_session.query(EmployeeData).filter_by(user_id=user_id).first()
    if employee_data:
        employee_data.needs_update = True
        db_session.commit()
        
        create_notification(
            user_id,
            'مطلوب تحديث البيانات',
            'يرجى تحديث بياناتك الشخصية',
            'data_update_requested',
            action_url=url_for('user_form')
        )
        
        flash('تم إرسال طلب تحديث البيانات')
    
    return redirect(url_for('admin_employee_progress'))



def save_schedule_data(schedule_data):
    # تأكد من أن البيانات في التنسيق الصحيح
    if isinstance(schedule_data, (dict, list)):
        # حول إلى JSON string للتخزين
        return json.dumps(schedule_data, ensure_ascii=False)
    else:
        raise ValueError("بيانات الجدول يجب أن تكون dictionary أو list")
    
@app.route('/export_schedule/<int:schedule_id>')
@login_required
def export_schedule(schedule_id):
    """تصدير الجدول - للمسؤولين فقط"""
    if not current_user.is_admin:
        flash('غير مصرح لك بتنزيل الجداول', 'error')
        return redirect(url_for('user_dashboard'))
    
    schedule = db_session.get(WeeklySchedule, schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('admin_schedules'))
    
    try:
        # إنشاء ملف Excel
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"جدول {schedule.department.name}"
        
        # إضافة البيانات إلى Excel
        if schedule.schedule_data:
            schedule_data = json.loads(schedule.schedule_data)
            
            # كتابة العناوين
            headers = ['الموظف', 'السبت', 'الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # كتابة بيانات الموظفين
            row = 2
            for employee_id, days in schedule_data.items():
                employee = db_session.query(User).get(int(employee_id))
                worksheet.cell(row=row, column=1, value=employee.name if employee else f"مستخدم {employee_id}")
                
                days_order = ['السبت', 'الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس']
                for col, day in enumerate(days_order, 2):
                    worksheet.cell(row=row, column=col, value=days.get(day, ''))
                row += 1
        
        # حفظ الملف
        workbook.save(output)
        output.seek(0)
        
        # تحديث حالة الجدول إلى مغلق
        schedule.is_locked = True
        schedule.exported_at = datetime.now()
        schedule.exported_by = current_user.id
        db_session.commit()
        
        # إرسال الإشعارات
        send_notification(
            user_id=schedule.created_by,
            title="تم تصدير الجدول الأسبوعي",
            message=f"تم تصدير جدول قسم {schedule.department.name} للفترة {schedule.week_start_date} إلى {schedule.week_end_date} ولا يمكن التعديل عليه الآن",
            notification_type="schedule_exported",
            related_id=schedule.id
        )
        
        filename = f"جدول_{schedule.department.name}_{schedule.week_start_date}_to_{schedule.week_end_date}.xlsx"
        return send_file(output, 
                        download_name=filename,
                        as_attachment=True,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        flash(f'خطأ في تصدير الملف: {str(e)}', 'error')
        return redirect(url_for('admin_schedules'))
    


@app.route('/view_schedule/<int:schedule_id>')
@login_required
def view_schedule(schedule_id):
    """عرض الجدول للقراءة فقط"""
    schedule = db_session.get(WeeklySchedule, schedule_id)
    if not schedule:
        flash('الجدول غير موجود')
        return redirect(url_for('user_dashboard'))
    
    department = db_session.query(Department).get(schedule.department_id)
    
    # التحقق من الصلاحيات
    can_edit = False
    if current_user.is_admin:
        can_edit = False  # Admin can only view
    elif current_user.is_manager and department and department.manager_id == current_user.id:
        can_edit = not (schedule.is_locked or schedule.is_approved)
    
    # الحصول على موظفي القسم
    employees = db_session.query(User).filter_by(
        department_id=department.id,
        is_admin=False
    ).all()
    
    # تحضير بيانات الموظفين للقالب
    employees_data = []
    for emp in employees:
        employee_data = {
            'id': emp.id,
            'name': emp.name,
            'username': emp.username
        }
        if hasattr(emp, 'job'):
            employee_data['job'] = emp.job or 'موظف'
        else:
            employee_data['job'] = 'موظف'
        employees_data.append(employee_data)
    
    # تحميل بيانات الجدول الحالية
    current_schedule = {}
    if schedule.schedule_data:
        try:
            current_schedule = json.loads(schedule.schedule_data)
        except:
            current_schedule = {}
    
    notifications = get_user_notifications(current_user.id)
    
    template = 'view_schedule.html'
    if current_user.is_admin:
        template = 'admin/view_schedule.html'
    
    return render_template(template,
                         schedule=schedule,
                         department=department,
                         current_schedule=current_schedule,
                         employees=employees_data,
                         can_edit=can_edit,
                         notifications=notifications)


@app.route('/approve_schedule/<int:schedule_id>')
@login_required
def approve_schedule(schedule_id):
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    schedule = db_session.get(WeeklySchedule, schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('admin_schedules'))
    
    try:
        # اعتماد الجدول وتفعيل القفل
        schedule.is_approved = True
        schedule.approved_by = current_user.id
        schedule.approved_at = datetime.now()
        schedule.is_locked = True  # قفل الجدول بعد الاعتماد
        
        db_session.commit()
        
        # إرسال إشعار للمدير
        send_notification(
            user_id=schedule.created_by,
            title="تم اعتماد الجدول الأسبوعي",
            message=f"تم اعتماد جدول قسم {schedule.department.name} للفترة {schedule.week_start_date} إلى {schedule.week_end_date} ولا يمكن التعديل عليه",
            notification_type="schedule_approved",
            related_id=schedule.id
        )
        
        flash('تم اعتماد الجدول بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'خطأ في اعتماد الجدول: {str(e)}', 'error')
    
    return redirect(url_for('admin_schedules'))


from sqlalchemy.orm import joinedload

@app.route('/admin/schedules')
@login_required
def admin_schedules():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    # إنشاء الجداول تلقائياً إذا لم تكن موجودة
    #auto_generate_weekly_schedules()
    
    # بناء الاستعلام مع التصفية والعلاقات
    query = db_session.query(WeeklySchedule).options(
        joinedload(WeeklySchedule.department),
        joinedload(WeeklySchedule.creator),
        joinedload(WeeklySchedule.approver)
    ).filter(
        WeeklySchedule.department_id.isnot(None)  # إضافة فلتر للتأكد من وجود القسم
    )
    
    # تصفية حسب القسم
    department_id = request.args.get('department_id')
    if department_id and department_id.isdigit():
        query = query.filter(WeeklySchedule.department_id == int(department_id))
    
    # تصفية حسب الحالة
    status = request.args.get('status')
    if status == 'approved':
        query = query.filter(WeeklySchedule.is_approved == True)
    elif status == 'pending':
        query = query.filter(WeeklySchedule.is_approved == False, WeeklySchedule.status == 'pending')
    elif status == 'draft':
        query = query.filter(WeeklySchedule.is_approved == False, WeeklySchedule.status == 'draft')
    
    # تصفية حسب الأسبوع
    week = request.args.get('week')
    if week:
        try:
            # تحويل week input إلى تاريخ بداية الأسبوع
            year, week_num = map(int, week.split('-W'))
            first_day = datetime.strptime(f'{year}-W{week_num}-1', '%G-W%V-%u').date()
            query = query.filter(WeeklySchedule.week_start_date == first_day)
        except ValueError:
            pass
    
    # تصفية حسب النطاق الزمني
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(WeeklySchedule.week_start_date >= date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(WeeklySchedule.week_end_date <= date_to_obj)
        except ValueError:
            pass
    
    schedules = query.order_by(WeeklySchedule.week_start_date.desc()).all()
    departments = db_session.query(Department).all()
    
    notifications = get_user_notifications(current_user.id)
    
    # حساب التواريخ الحالية
    today = date.today()
    days_since_saturday = (today.weekday() - 5) % 7
    current_week_start = today - timedelta(days=days_since_saturday)
    current_week_end = current_week_start + timedelta(days=6)
    
    return render_template('admin/admin_schedules.html',
                         schedules=schedules,
                         departments=departments,
                         today=today,
                         current_week_start=current_week_start,
                         current_week_end=current_week_end,
                         date=date,  # Pass date to template
                         timedelta=timedelta,  # Pass timedelta to template
                         notifications=notifications)

@app.route('/admin/generate_schedules')
@login_required
def admin_generate_schedules():
    """إنشاء الجداول الأسبوعية يدوياً"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        generated_count = auto_generate_weekly_schedules()
        return jsonify({
            'success': True, 
            'message': f'تم إنشاء {generated_count} جدول أسبوعي بنجاح',
            'count': generated_count
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في إنشاء الجداول: {str(e)}'
        })


def process_schedule_data(schedule_data):
    employees_added = 0
    
    if isinstance(schedule_data, dict):
        # معالجة كـ dictionary
        for key, value in schedule_data.items():
            print(f"معالجة المفتاح: {key}, النوع: {type(value)}")
            
            if key == 'schedule' and isinstance(value, list):
                # معالجة قائمة الموظفين
                for employee_data in value:
                    if add_employee_from_schedule(employee_data):
                        employees_added += 1
            elif isinstance(value, (dict, list)):
                # معالجة الهياكل المتداخلة
                employees_added += process_schedule_data(value)
    
    elif isinstance(schedule_data, list):
        # معالجة كـ list مباشرة
        for employee_data in schedule_data:
            if add_employee_from_schedule(employee_data):
                employees_added += 1
    
    return employees_added


def add_employee_from_schedule(employee_data):
    try:
        if not isinstance(employee_data, dict):
            print(f"بيانات الموظف ليست dictionary: {type(employee_data)}")
            return False
        
        # استخراج البيانات مع قيم افتراضية
        name = employee_data.get('morning_shift') or employee_data.get('evening_shift') or employee_data.get('night_shift')
        if not name or name.strip() == '':
            return False
            
        job = employee_data.get('job', '')
        department = employee_data.get('department', '')
        date = employee_data.get('date', '')
        
        # هنا أضف منطق حفظ الموظف في قاعدة البيانات
        print(f"إضافة موظف: {name}, وظيفة: {job}, قسم: {department}, تاريخ: {date}")
        return True
        
    except Exception as e:
        print(f"خطأ في إضافة الموظف: {str(e)}")
        return False
    

@app.route('/admin/schedule/export/<int:schedule_id>')
@login_required
def admin_export_schedule(schedule_id):
    """تصدير الجدول بنفس هيكل العرض"""
    if not current_user.is_admin:
        flash('غير مصرح لك بتنزيل الجداول', 'error')
        return redirect(url_for('user_dashboard'))
    
    schedule = db_session.get(WeeklySchedule, schedule_id)
    if not schedule:
        flash('الجدول غير موجود', 'error')
        return redirect(url_for('admin_schedules'))
    
    try:
        # الحصول على اسم القسم
        department_name = "غير معين"
        department = db_session.query(Department).get(schedule.department_id)
        if department:
            department_name = department.name
        
        print(f"تصدير جدول: {schedule_id}")
        print(f"بيانات الجدول: {schedule.schedule_data}")
        
        # إنشاء ملف Excel
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "جدول العمل"
        
        # إضافة العنوان
        worksheet.merge_cells('A1:G1')
        worksheet['A1'] = f'جدول العمل الأسبوعي - {department_name}'
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        worksheet.merge_cells('A2:G2')
        worksheet['A2'] = f'الفترة: {schedule.week_start_date} إلى {schedule.week_end_date}'
        worksheet['A2'].font = Font(size=12, bold=True)
        worksheet['A2'].alignment = Alignment(horizontal='center')
        
        # إضافة رؤوس الأعمدة
        headers = ['اليوم', 'التاريخ', 'الشيفت الصباحي', 'الشيفت المسائي', 'شيفت السهر', 'الوظيفة', 'القسم']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # معالجة البيانات بناءً على الهيكل الجديد
        row = 5
        entry_count = 0
        
        if schedule.schedule_data:
            try:
                schedule_data = json.loads(schedule.schedule_data)
                print(f"نوع البيانات: {type(schedule_data)}")
                print(f"مفاتيح البيانات: {schedule_data.keys() if isinstance(schedule_data, dict) else 'Not a dict'}")
                
                # الهيكل الجديد: يحتوي على schedule كمفتاح رئيسي
                if isinstance(schedule_data, dict) and 'schedule' in schedule_data:
                    schedule_entries = schedule_data['schedule']
                    print(f"عدد الإدخالات في الجدول: {len(schedule_entries)}")
                    
                    for entry in schedule_entries:
                        if isinstance(entry, dict):
                            # إضافة بيانات كل صف
                            worksheet.cell(row=row, column=1, value=entry.get('day', ''))
                            worksheet.cell(row=row, column=2, value=entry.get('date', ''))
                            worksheet.cell(row=row, column=3, value=entry.get('morning_shift', ''))
                            worksheet.cell(row=row, column=4, value=entry.get('evening_shift', ''))
                            worksheet.cell(row=row, column=5, value=entry.get('night_shift', ''))
                            worksheet.cell(row=row, column=6, value=entry.get('job', ''))
                            worksheet.cell(row=row, column=7, value=entry.get('department', department_name))
                            
                            row += 1
                            entry_count += 1
                            
                    print(f"تم إضافة {entry_count} إدخال إلى Excel")
                    
                else:
                    # محاولة الهيكل القديم
                    print("الهيكل غير متوقع، محاولة الهيكل القديم")
                    for key, value in schedule_data.items():
                        if isinstance(value, dict):
                            employee = db_session.query(User).get(int(key)) if key.isdigit() else None
                            if employee:
                                worksheet.cell(row=row, column=1, value=employee.name)
                                worksheet.cell(row=row, column=2, value=value.get('saturday', ''))
                                worksheet.cell(row=row, column=3, value=value.get('sunday', ''))
                                worksheet.cell(row=row, column=4, value=value.get('monday', ''))
                                worksheet.cell(row=row, column=5, value=value.get('tuesday', ''))
                                worksheet.cell(row=row, column=6, value=value.get('wednesday', ''))
                                worksheet.cell(row=row, column=7, value=value.get('thursday', ''))
                                row += 1
                                entry_count += 1
                                
            except Exception as e:
                print(f"خطأ في معالجة بيانات الجدول: {str(e)}")
                import traceback
                print(f"Traceback: {traceback.format_exc()}")
        
        # إذا لم توجد بيانات، إضافة رسالة
        if entry_count == 0:
            worksheet.merge_cells('A5:G5')
            worksheet.cell(row=5, column=1, value="لا توجد بيانات في الجدول")
            worksheet.cell(row=5, column=1).alignment = Alignment(horizontal='center')
            worksheet.cell(row=5, column=1).font = Font(color="FF0000", bold=True)
            print("لم يتم العثور على بيانات للتصدير")
        
        # ضبط عرض الأعمدة
        column_widths = {
            'A': 15,  # اليوم
            'B': 12,  # التاريخ
            'C': 20,  # الشيفت الصباحي
            'D': 20,  # الشيفت المسائي
            'E': 20,  # شيفت السهر
            'F': 30,  # الوظيفة
            'G': 20   # القسم
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # إضافة حدود للجدول
        from openpyxl.styles import Border, Side
        
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        for row_num in range(4, worksheet.max_row + 1):
            for col in range(1, 8):
                worksheet.cell(row=row_num, column=col).border = thin_border
        
        workbook.save(output)
        output.seek(0)
        
        filename = f"جدول_{department_name}_{schedule.week_start_date}_to_{schedule.week_end_date}.xlsx".replace(" ", "_")
        
        # تحديث حالة الجدول إذا كان معتمداً
        if schedule.is_approved and not schedule.is_locked:
            schedule.is_locked = True
            schedule.exported_at = datetime.now()
            schedule.exported_by = current_user.id
            db_session.commit()
        
        flash(f'تم تصدير الجدول بنجاح ({entry_count} إدخال)', 'success')
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"خطأ تفصيلي في التصدير: {e}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        flash(f'خطأ في تصدير الملف: {str(e)}', 'error')
        return redirect(url_for('admin_schedules'))



@app.teardown_appcontext
def shutdown_session(exception=None):
    db_session.remove()

@app.before_request
def create_session():
    # التأكد من وجود جلسة قاعدة بيانات
    pass

@app.after_request
def close_session(response):
    try:
        db_session.remove()
    except Exception as e:
        print(f"Error closing session: {e}")
    return response

from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from io import BytesIO

@app.route('/admin/schedule/approve/<int:schedule_id>')
@login_required
def admin_approve_schedule(schedule_id):
    """اعتماد الجدول من قبل المدير العام دون حذف أي بيانات"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    schedule = db_session.get(WeeklySchedule, schedule_id)
    if not schedule:
        return jsonify({'success': False, 'message': 'الجدول غير موجود'})
    
    try:
        schedule.is_approved = True
        schedule.approved_by = current_user.id
        schedule.approved_at = datetime.now()
        schedule.status = 'approved'
        
        db_session.commit()
        
        # إنشاء الجداول المستقبلية تلقائياً (بدون حذف القديمة)
        generate_future_schedules(schedule.department_id)
        
        # إرسال إشعار للمدير
        if schedule.department and schedule.department.manager_id:
            create_notification(
                schedule.department.manager_id,
                "تم اعتماد الجدول الأسبوعي",
                f"تم اعتماد جدول قسم {schedule.department.name} للفترة {schedule.week_start_date} إلى {schedule.week_end_date}",
                'schedule_approved',
                related_id=schedule.id
            )
        
        return jsonify({'success': True, 'message': 'تم اعتماد الجدول بنجاح'})
        
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'خطأ في اعتماد الجدول: {str(e)}'})

def generate_future_schedules(department_id):
    """إنشاء الجداول المستقبلية تلقائياً مع الحفاظ على البيانات القديمة"""
    try:
        department = db_session.query(Department).get(department_id)
        if not department or not department.auto_generate_schedule:
            return 0
        
        # تاريخ بداية الأسبوع الحالي
        today = date.today()
        days_since_saturday = (today.weekday() - 5) % 7
        current_week_start = today - timedelta(days=days_since_saturday)
        
        # إنشاء جداول للأسابيع القادمة (بدون حذف القديمة)
        weeks_to_generate = 3  # 3 أسابيع قادمة
        generated_count = 0
        
        for week_offset in range(1, weeks_to_generate + 1):  # ابدأ من الأسبوع القادم
            week_start_date = current_week_start + timedelta(days=7 * week_offset)
            
            # التحقق من عدم وجود جدول لهذا الأسبوع
            existing_schedule = db_session.query(WeeklySchedule).filter_by(
                department_id=department_id,
                week_start_date=week_start_date
            ).first()
            
            if not existing_schedule:
                new_schedule = create_weekly_schedule_from_structure(department_id, week_start_date)
                if new_schedule:
                    generated_count += 1
                    print(f"تم إنشاء جدول مستقبلي للقسم {department.name} للفترة {week_start_date}")
        
        return generated_count
        
    except Exception as e:
        print(f"خطأ في إنشاء الجداول المستقبلية: {str(e)}")
        return 0

def generate_next_week_schedule(department_id):
    """إنشاء جدول للأسبوع القادم تلقائياً"""
    try:
        department = db_session.query(Department).get(department_id)
        if not department or not department.auto_generate_schedule:
            return None
        
        # تاريخ بداية الأسبوع القادم
        next_week_start = date.today() + timedelta(days=(7 - date.today().weekday() + 5) % 7)
        
        # التحقق من عدم وجود جدول للأسبوع القادم
        existing_schedule = db_session.query(WeeklySchedule).filter_by(
            department_id=department_id,
            week_start_date=next_week_start
        ).first()
        
        if existing_schedule:
            print(f"جدول الأسبوع القادم موجود مسبقاً للقسم {department.name}")
            return existing_schedule
        
        print(f"إنشاء جدول الأسبوع القادم للقسم {department.name}")
        return create_weekly_schedule_from_structure(department_id, next_week_start)
        
    except Exception as e:
        print(f"خطأ في إنشاء جدول الأسبوع القادم: {str(e)}")
        return None


@app.route('/admin/schedule/lock/<int:schedule_id>')
@login_required
def admin_lock_schedule(schedule_id):
    """قفل الجدول بعد الاعتماد"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    schedule = db_session.get(WeeklySchedule, schedule_id)
    if not schedule:
        return jsonify({'success': False, 'message': 'الجدول غير موجود'})
    
    try:
        schedule.is_locked = True
        schedule.exported_at = datetime.now()
        schedule.exported_by = current_user.id
        db_session.commit()
        
        return jsonify({'success': True, 'message': 'تم قفل الجدول بنجاح'})
        
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'خطأ في قفل الجدول: {str(e)}'})
    

@app.route('/user/dashboard')
@login_required
def user_dashboard():
    if current_user.is_admin:
        return redirect(url_for('admin_dashboard'))
    elif current_user.is_manager:
        return redirect(url_for('manager_dashboard'))
    
    employee_data = db_session.query(EmployeeData).filter_by(user_id=current_user.id).first()
    
    if employee_data:
        completion_percentage = employee_data.calculate_completion()
        missing_fields = employee_data.get_missing_fields()
        db_session.commit()
    else:
        completion_percentage = 0
        missing_fields = []
    
    # Check for new salary slips
    new_salary_slips = db_session.query(SalarySlip).filter_by(
        user_id=current_user.id, is_viewed=False
    ).count()
    
    # Check schedule notifications
    current_week_start = date.today() - timedelta(days=date.today().weekday())
    has_current_schedule = db_session.query(WeeklySchedule).filter_by(
        department_id=current_user.department_id,
        week_start_date=current_week_start,
        is_approved=True
    ).first()
    
    # Get employee balance
    employee_balance = db_session.query(EmployeeBalance).filter_by(user_id=current_user.id).first()
    if not employee_balance:
        # Create balance if it doesn't exist
        employee_balance = EmployeeBalance(user_id=current_user.id)
        db_session.add(employee_balance)
        db_session.commit()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_dashboard.html', 
                         employee_data=employee_data,
                         completion_percentage=completion_percentage,
                         missing_fields=missing_fields,
                         new_salary_slips=new_salary_slips,
                         has_current_schedule=has_current_schedule,
                         employee_balance=employee_balance,
                         notifications=notifications)

@app.route('/user/form', methods=['GET', 'POST'])
@login_required
def user_form():

    employee_data = db_session.query(EmployeeData).filter_by(user_id=current_user.id).first()
    
    if request.method == 'POST':
        if not current_user.is_admin or not current_user.is_manager:
            if not employee_data:
                employee_data = EmployeeData(user_id=current_user.id)
                db_session.add(employee_data)
        
        # Update all fields from form - none are required
        employee_data.arabic_name = request.form.get('arabic_name') or None
        employee_data.english_name = request.form.get('english_name') or None
        employee_data.national_id = request.form.get('national_id') or None
        
        if request.form.get('id_issue_date'):
            employee_data.id_issue_date = datetime.strptime(request.form['id_issue_date'], '%Y-%m-%d').date()
        else:
            employee_data.id_issue_date = None
        
        if request.form.get('birth_date'):
            employee_data.birth_date = datetime.strptime(request.form['birth_date'], '%Y-%m-%d').date()
            employee_data.age = calculate_age(employee_data.birth_date)
        else:
            employee_data.birth_date = None
            employee_data.age = None
        
        employee_data.whatsapp = request.form.get('whatsapp') or None
        employee_data.phone = request.form.get('phone') or None
        employee_data.address = request.form.get('address') or None
        employee_data.military_status = request.form.get('military_status') or None
        employee_data.marital_status = request.form.get('marital_status') or None
        employee_data.qualification = request.form.get('qualification') or None
        
        if request.form.get('graduation_year'):
            employee_data.graduation_year = int(request.form['graduation_year'])
        else:
            employee_data.graduation_year = None
        
        employee_data.grade = request.form.get('grade') or None
        employee_data.has_work = 'has_work' in request.form
        employee_data.workplace = request.form.get('workplace') or None
        employee_data.job_title = request.form.get('job_title') or None
        employee_data.insurance_number = request.form.get('insurance_number') or None
        employee_data.tax_number = request.form.get('tax_number') or None
        
        # Emergency contacts
        employee_data.emergency1_name = request.form.get('emergency1_name') or None
        employee_data.emergency1_phone = request.form.get('emergency1_phone') or None
        employee_data.emergency1_address = request.form.get('emergency1_address') or None
        employee_data.emergency1_relation = request.form.get('emergency1_relation') or None
        employee_data.emergency2_name = request.form.get('emergency2_name') or None
        employee_data.emergency2_phone = request.form.get('emergency2_phone') or None
        employee_data.emergency2_address = request.form.get('emergency2_address') or None
        employee_data.emergency2_relation = request.form.get('emergency2_relation') or None
        
        employee_data.profession_license = request.form.get('profession_license') or None
        employee_data.union_card = request.form.get('union_card') or None
        
        # Handle file uploads
        user_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], str(current_user.name))
        os.makedirs(user_upload_folder, exist_ok=True)
        
        file_fields = [
            ('national_id_image', 'national_id_image'),
            ('military_status_image', 'military_status_image'),
            ('qualification_image', 'qualification_image'),
            ('salary_details', 'salary_details'),
            ('employment_status', 'employment_status')
        ]
        
        for form_field, db_field in file_fields:
            file = request.files.get(form_field)
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file_path = os.path.join(user_upload_folder, filename)
                file.save(file_path)
                setattr(employee_data, db_field, filename)
        
        employee_data.last_updated = datetime.now()
        employee_data.updated_by = current_user.name
        
        # حساب نسبة الإكتمال باستخدام الدالة المساعدة
        completion_percentage = employee_data.calculate_completion()

        # Check if this is a final update or temporary save
        if 'save_complete' in request.form:
            employee_data.needs_update = False
            flash('تم التحديث النهائي للبيانات بنجاح')
            
            # Export to Excel
            export_to_excel(current_user.id)
            
            # Notify admin
            admins = db_session.query(User).filter_by(is_admin=True).all()
            for admin in admins:
                create_notification(
                    admin.id,
                    'تم تحديث البيانات نهائياً',
                    f'قام {current_user.name} بتحديث بياناته نهائياً',
                    'data_updated'
                )
        else:
            # Temporary save
            flash('تم الحفظ المؤقت للبيانات بنجاح')
        
        db_session.commit()
        
        return redirect(url_for('user_dashboard'))
    
    missing_fields = []
    completion_percentage = 0
    
    if employee_data:
        completion_percentage = employee_data.calculate_completion()
        missing_fields = employee_data.get_missing_fields()
        db_session.commit()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_form.html', 
                         employee_data=employee_data,
                         completion_percentage=completion_percentage,
                         missing_fields=missing_fields,
                         notifications=notifications)


@app.route('/user/leave', methods=['GET', 'POST'])
@login_required
def user_leave():
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    balance = db_session.query(EmployeeBalance).filter_by(user_id=current_user.id).first()
    
    # الحصول على نطاق الشهر المالي
    financial_month_start, financial_month_end = get_financial_month_range()
    
    # الحصول على قائمة الوظائف من هيكل القسم
    department_jobs = get_department_jobs(current_user.department_id)
    
    if request.method == 'POST':
        # التحقق من وجود الحقول المطلوبة
        required_fields = ['leave_type', 'leave_date', 'reason']
        for field in required_fields:
            if field not in request.form:
                flash(f'حقل {field} مطلوب', 'error')
                return redirect(url_for('user_leave'))
        
        leave_type = request.form['leave_type']
        leave_date = datetime.strptime(request.form['leave_date'], '%Y-%m-%d').date()
        reason = request.form['reason']
        
        # جلب معلومات الشيفتات المختارة
        shifts = request.form.getlist('shifts[]')
        
        if not shifts:
            flash('يجب اختيار شيفت واحد على الأقل', 'error')
            return redirect(url_for('user_leave'))
        
        # التحقق من أن التاريخ في نطاق الشهر المالي
        if leave_date < financial_month_start or leave_date > financial_month_end:
            flash(f'يمكن اختيار التواريخ فقط من {financial_month_start.strftime("%Y-%m-%d")} إلى {financial_month_end.strftime("%Y-%m-%d")}', 'error')
            return redirect(url_for('user_leave'))
        
        # التحقق من أن التاريخ ليس في الماضي
        if leave_date < date.today():
            flash('لا يمكن طلب إجازة لتاريخ ماضي', 'error')
            return redirect(url_for('user_leave'))
        
        # إنشاء معرف فريد لهذه المجموعة من الشيفتات
        parent_request_id = int(datetime.now().timestamp())
        created_requests = 0
        
        try:
            for i, shift in enumerate(shifts):
                # الحصول على الوظيفة المختارة لهذا الشيفت
                job = request.form.get(f'job_{shift}', '')
                
                # التحقق من اختيار وظيفة لكل شيفت (خاص بالإجازات من الرصيد)
                if leave_type == 'من رصيد الإجازات' and not job:
                    flash(f'يجب اختيار وظيفة للشيفت {shift}', 'error')
                    return redirect(url_for('user_leave'))
                
                # التحقق من الرصيد فقط للشيفت الأول من النوع "من رصيد الإجازات"
                if leave_type == 'من رصيد الإجازات' and i == 0:
                    if balance.leave_balance < len(shifts):
                        flash(f'رصيد الإجازات غير كافي. تحتاج {len(shifts)} يوم، رصيدك الحالي: {balance.leave_balance} يوم', 'error')
                        return redirect(url_for('user_leave'))
                
                # إنشاء طلب إجازة منفصل لكل شيفت
                leave_request = LeaveRequest(
                    user_id=current_user.id,
                    department_id=current_user.department_id,
                    leave_type=leave_type,
                    start_date=leave_date,  # استخدام leave_date كـ start_date
                    end_date=leave_date,    # إذا كان نفس اليوم لبداية ونهاية الإجازة
                    leave_date=leave_date,  # أيضًا حفظه كـ leave_date إذا كان الحقل موجودًا
                    shift_name=shift,
                    shift_job=job,
                    total_days=1,  # كل شيفت = يوم واحد
                    reason=reason,
                    parent_request_id=parent_request_id,
                    shift_order=i+1,
                    status='pending'
                )
                
                db_session.add(leave_request)
                created_requests += 1
            
            # خصم من الرصيد إذا كانت الإجازة من رصيد الإجازات
            if leave_type == 'من رصيد الإجازات':
                balance.leave_balance -= len(shifts)
                balance.last_updated = datetime.now()
            
            db_session.commit()
            
            # إرسال إشعار للمدير
            department = db_session.query(Department).get(current_user.department_id)
            if department and department.primary_manager_id:
                shift_names = "، ".join(shifts)
                
                notification_message = f'طلب {current_user.name} إجازة من نوع {leave_type} '
                notification_message += f'بعدد {len(shifts)} شيفت ({shift_names}) '
                notification_message += f'بتاريخ {leave_date.strftime("%Y-%m-%d")}'
                
                create_notification(
                    department.primary_manager_id,
                    'طلب إجازة جديد',
                    notification_message,
                    'leave_request',
                    related_id=parent_request_id,
                    action_url=url_for('manager_leave_requests')
                )
            
            flash(f'تم تقديم {created_requests} طلب إجازة منفصل بنجاح', 'success')
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ أثناء تقديم الطلب: {str(e)}', 'error')
        
        return redirect(url_for('user_leave_requests'))
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_leave.html', 
                         balance=balance,
                         financial_month_start=financial_month_start,
                         financial_month_end=financial_month_end,
                         department_jobs=department_jobs,
                         notifications=notifications)

@app.template_filter('from_json')
def from_json_filter(value):
    """تحويل نص JSON إلى كائن Python في القوالب"""
    if not value or not isinstance(value, str):
        return {}
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return {}
    
# تأكد من أن Jinja2 يعرف الفلتر الجديد
app.jinja_env.filters['from_json'] = from_json_filter

def is_date_in_financial_month(check_date):
    """Check if date is within current financial month (26th to 25th)"""
    today = date.today()
    
    # Determine financial month range
    if today.day >= 26:
        # From 26th of current month to 25th of next month
        financial_start = date(today.year, today.month, 26)
        if today.month == 12:
            financial_end = date(today.year + 1, 1, 25)
        else:
            financial_end = date(today.year, today.month + 1, 25)
    else:
        # From 26th of previous month to 25th of current month
        if today.month == 1:
            financial_start = date(today.year - 1, 12, 26)
            financial_end = date(today.year, 1, 25)
        else:
            financial_start = date(today.year, today.month - 1, 26)
            financial_end = date(today.year, today.month, 25)
    
    return financial_start <= check_date <= financial_end


from datetime import date  # تأكد من وجود هذا الاستيراد في الأعلى
def get_financial_month_range():
    """الحصول على نطاق الشهر المالي (26 إلى 25 من الشهر التالي)"""
    today = date.today()
    
    if today.day >= 26:
        # الفترة: 26 من الشهر الحالي إلى 25 من الشهر التالي
        start_date = date(today.year, today.month, 26)
        if today.month == 12:
            end_date = date(today.year + 1, 1, 25)
        else:
            end_date = date(today.year, today.month + 1, 25)
    else:
        # الفترة: 26 من الشهر السابق إلى 25 من الشهر الحالي
        if today.month == 1:
            start_date = date(today.year - 1, 12, 26)
            end_date = date(today.year, 1, 25)
        else:
            start_date = date(today.year, today.month - 1, 26)
            end_date = date(today.year, today.month, 25)
    
    return start_date, end_date


def get_department_jobs(department_id):
    """الحصول على قائمة الوظائف من هيكل الجدول الخاص بالقسم"""
    department = db_session.query(Department).get(department_id)
    
    if not department:
        return []
    
    jobs = []
    
    # محاولة استخراج الوظائف من هيكل الجدول
    if department.schedule_structure:
        try:
            structure_data = json.loads(department.schedule_structure)
            
            # استخراج الوظائف من الهيكل
            if isinstance(structure_data, dict):
                # النوع الجديد: يحتوي على schedule كمفتاح
                if 'schedule' in structure_data:
                    schedule_list = structure_data['schedule']
                    for day_schedule in schedule_list:
                        if isinstance(day_schedule, dict):
                            job = day_schedule.get('job')
                            if job and job not in jobs:
                                jobs.append(job)
                
                # محاولة استخراج الوظائف من المفاتيح الأخرى
                elif 'jobs' in structure_data:
                    jobs = structure_data['jobs']
                
                # النوع القديم: قد يحتوي على حقل الوظائف مباشرة
                elif 'structure' in structure_data:
                    for item in structure_data['structure']:
                        if isinstance(item, dict):
                            job = item.get('job') or item.get('الوظيفة')
                            if job and job not in jobs:
                                jobs.append(job)
            
            elif isinstance(structure_data, list):
                # إذا كانت البيانات مباشرة كقائمة
                for item in structure_data:
                    if isinstance(item, dict):
                        job = item.get('job') or item.get('الوظيفة')
                        if job and job not in jobs:
                            jobs.append(job)
        
        except Exception as e:
            print(f"خطأ في استخراج الوظائف من هيكل القسم: {str(e)}")
    
    # إذا لم توجد وظائف في الهيكل، استخدم القائمة الافتراضية
    if not jobs:
        jobs = [
            'طبيب',
            'ممرض',
            'فني مختبر',
            'سكرتير طبي',
            'موظف استقبال',
            'أخصائي أشعة',
            'صيدلي',
            'إداري',
            'موظف خدمة'
        ]
    
    # إزالة القيم الفارغة وتصفية القيم
    jobs = [job for job in jobs if job and job.strip()]
    
    # إضافة "موظف" كخيار افتراضي إذا لم يكن موجوداً
    if 'موظف' not in jobs:
        jobs.insert(0, 'موظف')
    
    return jobs


@app.route('/approve_permission/<int:request_id>')
@login_required
def approve_permission(request_id):
    """Alias for manager_approve_permission to maintain template compatibility"""
    return redirect(url_for('manager_approve_permission', request_id=request_id))

@app.route('/user/leave_requests')
@login_required
def user_leave_requests():
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    # Get period filter parameter
    period_filter = request.args.get('period', 'all')
    
    # Get all leave requests (individual shifts)
    query = db_session.query(LeaveRequest)\
        .filter_by(user_id=current_user.id)\
        .order_by(LeaveRequest.created_at.desc())
    
    # Apply period filter
    if period_filter != 'all':
        today = date.today()
        
        if period_filter == 'week':
            week_ago = today - timedelta(days=7)
            query = query.filter(LeaveRequest.created_at >= week_ago)
        elif period_filter == 'month':
            month_ago = today - timedelta(days=30)
            query = query.filter(LeaveRequest.created_at >= month_ago)
        elif period_filter == '3months':
            three_months_ago = today - timedelta(days=90)
            query = query.filter(LeaveRequest.created_at >= three_months_ago)
    
    # Get all individual leave requests
    leave_requests = query.all()
    
    # Get permission requests with proper handling
    permission_query = db_session.query(PermissionRequest).filter_by(
        user_id=current_user.id
    ).order_by(PermissionRequest.created_at.desc())
    
    if period_filter != 'all':
        today = date.today()
        if period_filter == 'week':
            week_ago = today - timedelta(days=7)
            permission_query = permission_query.filter(PermissionRequest.created_at >= week_ago)
        elif period_filter == 'month':
            month_ago = today - timedelta(days=30)
            permission_query = permission_query.filter(PermissionRequest.created_at >= month_ago)
        elif period_filter == '3months':
            three_months_ago = today - timedelta(days=90)
            permission_query = permission_query.filter(PermissionRequest.created_at >= three_months_ago)
    
    permission_requests = permission_query.all()
    
    # Debug: Print permission request data
    print(f"Found {len(permission_requests)} permission requests")
    for i, req in enumerate(permission_requests[:5]):  # Show first 5
        print(f"Permission {i+1}: ID={req.id}, Type={req.permission_type}, Extra Data={req.extra_data}")
    
    # Calculate completion percentage
    employee_data = db_session.query(EmployeeData).filter_by(user_id=current_user.id).first()
    new_salary_slips = 0
    if employee_data:
        employee_data.calculate_completion()
        new_salary_slips = db_session.query(SalarySlip).filter_by(
            user_id=current_user.id, is_viewed=False
        ).count()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_leave_requests.html',
                         leave_requests=leave_requests,
                         permission_requests=permission_requests,
                         period_filter=period_filter,
                         new_salary_slips=new_salary_slips,
                         notifications=notifications)



@app.route('/debug_permissions')
@login_required
def debug_permissions():
    """Debug route to check permission data structure"""
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    permission_requests = db_session.query(PermissionRequest).filter_by(
        user_id=current_user.id
    ).order_by(PermissionRequest.created_at.desc()).all()
    
    debug_data = []
    for req in permission_requests:
        debug_data.append({
            'id': req.id,
            'permission_type': req.permission_type,
            'date': req.date.strftime('%Y-%m-%d') if req.date else None,
            'time': req.time,
            'reason': req.reason,
            'extra_data': req.extra_data,
            'extra_data_dict': req.extra_data_dict if hasattr(req, 'extra_data_dict') else {},
            'shift_name': req.shift_name if hasattr(req, 'shift_name') else 'N/A',
            'shift_job': req.shift_job if hasattr(req, 'shift_job') else 'N/A',
            'exchange_employee_name': req.exchange_employee_name if hasattr(req, 'exchange_employee_name') else 'N/A',
            'attendance_type': req.attendance_type if hasattr(req, 'attendance_type') else 'N/A',
            'overtime_hours': req.overtime_hours if hasattr(req, 'overtime_hours') else 'N/A',
            'status': req.status
        })
    
    return jsonify({
        'user_id': current_user.id,
        'total_requests': len(permission_requests),
        'requests': debug_data
    })

# In your app.py, add these template filters:

@app.template_filter('get_extra_data')
def get_extra_data_filter(extra_data_json):
    """Parse extra_data JSON in template"""
    if not extra_data_json:
        return {}
    try:
        return json.loads(extra_data_json)
    except:
        return {}

@app.template_filter('get_shift_name')
def get_shift_name_filter(extra_data_json):
    """Get shift name from extra_data in template"""
    data = get_extra_data_filter(extra_data_json)
    return data.get('shift', '')

@app.template_filter('get_shift_job')
def get_shift_job_filter(extra_data_json):
    """Get job from extra_data in template"""
    data = get_extra_data_filter(extra_data_json)
    return data.get('job', '')

@app.template_filter('get_employee_name')
def get_employee_name_filter(extra_data_json):
    """Get employee name from extra_data in template"""
    data = get_extra_data_filter(extra_data_json)
    return data.get('employee_name', '')

@app.template_filter('get_attendance_type')
def get_attendance_type_filter(extra_data_json):
    """Get attendance type from extra_data in template"""
    data = get_extra_data_filter(extra_data_json)
    return data.get('attendance_type', '')

@app.template_filter('get_overtime_hours')
def get_overtime_hours_filter(extra_data_json):
    """Get overtime hours from extra_data in template"""
    data = get_extra_data_filter(extra_data_json)
    return data.get('hours') or data.get('overtime_hours') or 0


def reset_monthly_permission_balances():
    """Reset permission balance to 2 for all employees at the beginning of each month"""
    try:
        today = date.today()
        
        # Check if it's the first day of the month
        if today.day == 1:
            print(f"=== Resetting permission balances for {today.strftime('%Y-%m')} ===")
            
            # Get all employee balances
            employee_balances = db_session.query(EmployeeBalance).all()
            reset_count = 0
            
            for balance in employee_balances:
                # Only reset if the balance is different from default (2)
                if balance.permission_balance != 2:
                    old_balance = balance.permission_balance
                    balance.permission_balance = 2
                    balance.last_updated = datetime.now()
                    reset_count += 1
                    
                    print(f"Reset permission balance for user {balance.user_id}: {old_balance} -> 2")
                    
                    # Notify employee about balance reset
                    create_notification(
                        balance.user_id,
                        'تم تجديد رصيد الإذونات',
                        f'تم تجديد رصيد الإذونات الخاص بك إلى 2 إذن لشهر {today.strftime("%Y-%m")}',
                        'balance_reset',
                        action_url=url_for('user_dashboard')
                    )
            
            if reset_count > 0:
                db_session.commit()
                print(f"Successfully reset permission balances for {reset_count} employees")
            else:
                print("No permission balances needed resetting")
                
            return reset_count
        else:
            print(f"Not the first day of month ({today.day}), skipping permission balance reset")
            return 0
            
    except Exception as e:
        print(f"Error resetting permission balances: {str(e)}")
        db_session.rollback()
        return 0
    

def check_and_reset_balances():
    """Check if balances need to be reset (runs daily)"""
    try:
        # Check if we've already reset balances this month
        today = date.today()
        current_month = today.strftime('%Y-%m')
        
        if not hasattr(app, 'last_balance_reset_month'):
            app.last_balance_reset_month = None
        
        # Reset if it's a new month and we haven't reset yet
        if app.last_balance_reset_month != current_month and today.day == 1:
            reset_count = reset_monthly_permission_balances()
            app.last_balance_reset_month = current_month
            return reset_count
        
        return 0
        
    except Exception as e:
        print(f"Error in balance reset check: {str(e)}")
        return 0    

@app.before_request
def check_weekly_schedules_and_balances():
    """التحقق من وجود الجداول الأسبوعية وضبط أرصدة الإذونات قبل كل طلب"""
    try:
        # تشغيل هذا فقط مرة في اليوم لتجنب التحميل الزائد
        if not hasattr(app, 'last_schedule_check'):
            app.last_schedule_check = date.today() - timedelta(days=1)
        
        if date.today() > app.last_schedule_check:
            #auto_generate_weekly_schedules()
            check_and_reset_balances()  # Add this line
            app.last_schedule_check = date.today()
    except Exception as e:
        print(f"خطأ في التحقق من الجداول والأرصدة: {e}")



@app.route('/user/permission', methods=['GET', 'POST'])
@login_required
def user_permission():
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    balance = db_session.query(EmployeeBalance).filter_by(user_id=current_user.id).first()
    department_jobs = get_department_jobs(current_user.department_id)
    
    # Get department employees for auto-complete
    department_employees = db_session.query(User).filter(
        User.department_id == current_user.department_id,
        User.is_active == True,
        User.is_admin == False
    ).all()
    
    if request.method == 'POST':
        permission_type = request.form['permission_type']
        permission_date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        reason = request.form['reason']
        
        # التحقق من أن التاريخ ليس في الماضي
        if permission_date < date.today():
            flash('لا يمكن طلب إذن لتاريخ ماضي')
            return redirect(url_for('user_permission'))
        
        if balance.permission_balance <= 0:
            flash('رصيد الإذونات غير كافي')
            return redirect(url_for('user_permission'))
        
        created_count = 0
        
        try:
            # معالجة أنواع الإذن المختلفة
            if permission_type == 'اذن تبديل وردية':
                # جمع بيانات تبديل الشيفتات
                exchange_shifts = request.form.getlist('exchange_shifts[]')
                
                if not exchange_shifts:
                    flash('يرجى اختيار شيفت واحد على الأقل للتبديل')
                    return redirect(url_for('user_permission'))
                
                # التحقق من أن جميع الشيفتات المختارة لها وظائف محددة
                for shift in exchange_shifts:
                    job_key = f'exchange_job_{shift}'
                    job = request.form.get(job_key)
                    
                    if not job:
                        flash(f'يرجى اختيار وظيفة للشيفت {shift}')
                        return redirect(url_for('user_permission'))
                
                # إنشاء سجل منفصل لكل شيفت
                for shift in exchange_shifts:
                    job_key = f'exchange_job_{shift}'
                    employee_key = f'exchange_employee_{shift}'
                    employee_id_key = f'exchange_employee_{shift}_id'
                    
                    job = request.form.get(job_key)
                    employee_name = request.form.get(employee_key)
                    employee_id = request.form.get(employee_id_key)
                    
                    # بيانات إضافية لهذا الشيفت المحدد
                    extra_data = {
                        'type': 'shift_exchange',
                        'shift': shift,
                        'job': job,
                        'employee_name': employee_name,
                        'employee_id': employee_id
                    }
                    
                    # حساب الوقت
                    time_minutes = 0
                    time_input = request.form.get('time')
                    if time_input:
                        try:
                            time_obj = datetime.strptime(time_input, '%H:%M').time()
                            time_minutes = time_obj.hour * 60 + time_obj.minute
                        except:
                            pass
                    
                    # إنشاء طلب إذن منفصل لهذا الشيفت
                    permission_request = PermissionRequest(
                        user_id=current_user.id,
                        department_id=current_user.department_id,
                        permission_type=f'{permission_type} - {shift}',
                        date=permission_date,
                        time=time_minutes,
                        reason=reason,
                        extra_data=json.dumps(extra_data, ensure_ascii=False),
                        status='pending'
                    )
                    
                    db_session.add(permission_request)
                    created_count += 1
                
                # خصم من رصيد الإذونات (شيفت واحد = إذن واحد)
                balance.permission_balance -= len(exchange_shifts)
                
            elif permission_type == 'اذن طلب ساعات اضافي':
                # جمع بيانات الساعات الإضافية
                overtime_shifts = request.form.getlist('overtime_shifts[]')
                overtime_hours = int(request.form.get('overtime_hours', 1))
                
                if not overtime_shifts:
                    flash('يرجى اختيار شيفت واحد على الأقل للعمل الإضافي')
                    return redirect(url_for('user_permission'))
                
                # التحقق من عدد الساعات
                if overtime_hours < 1 or overtime_hours > 12:
                    flash('عدد الساعات الإضافية يجب أن يكون بين 1 و 12 ساعة')
                    return redirect(url_for('user_permission'))
                
                # التحقق من أن جميع الشيفتات المختارة لها وظائف محددة
                for shift in overtime_shifts:
                    job_key = f'overtime_job_{shift}'
                    job = request.form.get(job_key)
                    
                    if not job:
                        flash(f'يرجى اختيار وظيفة للشيفت {shift}')
                        return redirect(url_for('user_permission'))
                    
                    # التحقق من اختيار توقيت
                    if shift == 'صباحي':
                        attendance_type = request.form.get('morning_attendance_type')
                        if not attendance_type:
                            flash(f'يرجى اختيار توقيت للشيفت الصباحي')
                            return redirect(url_for('user_permission'))
                    else:
                        attendance_type = request.form.get('evening_attendance_type')
                        if not attendance_type:
                            flash(f'يرجى اختيار توقيت للشيفت المسائي')
                            return redirect(url_for('user_permission'))
                
                # إنشاء سجل منفصل لكل شيفت
                for shift in overtime_shifts:
                    job_key = f'overtime_job_{shift}'
                    job = request.form.get(job_key)
                    
                    # تحديد نوع الحضور/الانصراف
                    if shift == 'صباحي':
                        attendance_type = request.form.get('morning_attendance_type')
                    else:
                        attendance_type = request.form.get('evening_attendance_type')
                    
                    # بيانات إضافية لهذا الشيفت المحدد
                    extra_data = {
                        'type': 'overtime',
                        'shift': shift,
                        'job': job,
                        'hours': overtime_hours,
                        'attendance_type': attendance_type
                    }
                    
                    # حساب الوقت
                    time_minutes = 0
                    time_input = request.form.get('time')
                    if time_input:
                        try:
                            time_obj = datetime.strptime(time_input, '%H:%M').time()
                            time_minutes = time_obj.hour * 60 + time_obj.minute
                        except:
                            pass
                    
                    # إنشاء طلب إذن منفصل لهذا الشيفت
                    permission_request = PermissionRequest(
                        user_id=current_user.id,
                        department_id=current_user.department_id,
                        permission_type=f'{permission_type} - {shift}',
                        date=permission_date,
                        time=time_minutes,
                        reason=reason,
                        extra_data=json.dumps(extra_data, ensure_ascii=False),
                        status='pending'
                    )
                    
                    db_session.add(permission_request)
                    created_count += 1
                
                # خصم من رصيد الإذونات (شيفت واحد = إذن واحد)
                balance.permission_balance -= len(overtime_shifts)
            
            else:
                # الإذونات العادية (حضور/انصراف/نسيان توقيع)
                extra_data = {}
                time_minutes = 0
                time_input = request.form.get('time')
                
                if time_input:
                    try:
                        time_obj = datetime.strptime(time_input, '%H:%M').time()
                        time_minutes = time_obj.hour * 60 + time_obj.minute
                    except:
                        pass
                
                # إنشاء طلب إذن واحد
                permission_request = PermissionRequest(
                    user_id=current_user.id,
                    department_id=current_user.department_id,
                    permission_type=permission_type,
                    date=permission_date,
                    time=time_minutes,
                    reason=reason,
                    extra_data=json.dumps(extra_data) if extra_data else None,
                    status='pending'
                )
                
                db_session.add(permission_request)
                created_count += 1
                
                # خصم من رصيد الإذونات
                balance.permission_balance -= 1
            
            # تحديث رصيد الموظف
            balance.last_updated = datetime.now()
            
            # حفظ جميع التغييرات
            db_session.commit()
            
            # إرسال إشعار للمدير
            department = db_session.query(Department).get(current_user.department_id)
            if department and department.primary_manager_id:
                notification_message = f'طلب {current_user.name} {created_count} إذن من نوع {permission_type}'
                
                create_notification(
                    department.primary_manager_id,
                    'طلب إذن جديد',
                    notification_message,
                    'permission_request',
                    related_id=permission_request.id if created_count == 1 else None,
                    action_url=url_for('manager_permission_requests')
                )
            
            if created_count > 1:
                flash(f'تم تقديم {created_count} طلب إذن منفصل بنجاح')
            else:
                flash('تم تقديم طلب الإذن بنجاح وفي انتظار الموافقة')
            
            return redirect(url_for('user_leave_requests'))
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ أثناء تقديم الطلب: {str(e)}', 'error')
            print(f"Error in permission submission: {str(e)}")
            import traceback
            traceback.print_exc()
            return redirect(url_for('user_permission'))
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_permission.html',
                         balance=balance,
                         department_jobs=department_jobs,
                         department_employees=department_employees,
                         notifications=notifications)


@app.route('/admin/create_department', methods=['POST'])
@login_required
def admin_create_department():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        name = request.form['name']
        primary_manager_id = request.form.get('primary_manager_id') or None
        max_advance_amount = float(request.form.get('max_advance_amount', 0))
        max_installments = int(request.form.get('max_installments', 1))
        auto_generate = 'auto_generate_schedule' in request.form
        
        department = Department(
            name=name,
            primary_manager_id=primary_manager_id,
            advance_policy_max_amount=max_advance_amount,
            advance_policy_max_installments=max_installments,
            auto_generate_schedule=auto_generate,
            created_by=current_user.id
        )
        
        db_session.add(department)
        db_session.commit()
        
        flash('تم إنشاء القسم بنجاح', 'success')
        
    except Exception as e:
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('admin_departments'))

@app.route('/admin/add_department_manager', methods=['POST'])
@login_required
def admin_add_department_manager():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        department_id = request.form['department_id']
        user_id = request.form['user_id']
        
        # Get permissions from form
        permissions = {
            'can_manage_schedules': 'can_manage_schedules' in request.form,
            'can_manage_leaves': 'can_manage_leaves' in request.form,
            'can_manage_permissions': 'can_manage_permissions' in request.form,
            'can_manage_advances': 'can_manage_advances' in request.form,
            'can_manage_rewards': 'can_manage_rewards' in request.form,
            'can_view_reports': 'can_view_reports' in request.form
        }
        
        # Create department manager record
        department_manager = DepartmentManager(
            department_id=department_id,
            user_id=user_id,
            created_by=current_user.id,
            **permissions
        )
        
        db_session.add(department_manager)
        
        # Update user to be a manager if not already
        user = db_session.query(User).get(user_id)
        if not user.is_manager:
            user.is_manager = True
        
        db_session.commit()
        
        flash('تم إضافة المدير بنجاح', 'success')
        
    except Exception as e:
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('admin_departments'))

@app.route('/admin/update_manager_permissions/<int:manager_id>', methods=['POST'])
@login_required
def admin_update_manager_permissions(manager_id):
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        manager = db_session.query(DepartmentManager).get(manager_id)
        if not manager:
            flash('المدير غير موجود', 'error')
            return redirect(url_for('admin_departments'))
        
        # Update permissions
        manager.can_manage_schedules = 'can_manage_schedules' in request.form
        manager.can_manage_leaves = 'can_manage_leaves' in request.form
        manager.can_manage_permissions = 'can_manage_permissions' in request.form
        manager.can_manage_advances = 'can_manage_advances' in request.form
        manager.can_manage_rewards = 'can_manage_rewards' in request.form
        manager.can_view_reports = 'can_view_reports' in request.form
        
        db_session.commit()
        flash('تم تحديث الصلاحيات بنجاح', 'success')
        
    except Exception as e:
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(url_for('admin_departments'))


# Add this helper function
def get_manager_permissions_display(manager):
    """Return a string representation of manager permissions"""
    if not manager:
        return "لا توجد صلاحيات"
    
    permissions = []
    if manager.can_manage_schedules:
        permissions.append('إدارة الجداول')
    if manager.can_manage_leaves:
        permissions.append('إدارة الإجازات')
    if manager.can_manage_permissions:
        permissions.append('إدارة الإذونات')
    if manager.can_manage_advances:
        permissions.append('إدارة السلف')
    if manager.can_manage_rewards:
        permissions.append('إدارة المكافآت')
    if manager.can_view_reports:
        permissions.append('عرض التقارير')
    
    return '، '.join(permissions) if permissions else 'لا توجد صلاحيات'

# Update the context processor
@app.context_processor
def utility_processor():
    def get_manager_permissions(manager):
        return get_manager_permissions_display(manager)
    
    def get_approver_name(approver_id):
        if not approver_id:
            return "غير معين"
        approver = db_session.query(User).get(approver_id)
        return approver.name if approver else "غير معين"
    
    return dict(
        get_manager_permissions_display=get_manager_permissions,
        get_approver_name=get_approver_name
    )


@app.route('/user/cancel_leave/<int:request_id>')
@login_required
def cancel_leave_request(request_id):
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    leave_request = db_session.query(LeaveRequest).get(request_id)
    
    if leave_request and leave_request.user_id == current_user.id and leave_request.status == 'pending':
        # Check if this is a leave from balance and restore it
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=current_user.id
            ).first()
            if balance:
                balance.leave_balance += 1  # Restore 1 day for each shift
                balance.last_updated = datetime.now()
        
        db_session.delete(leave_request)
        db_session.commit()
        flash('تم إلغاء الشيفت بنجاح')
    
    return redirect(url_for('user_leave_requests'))


@app.route('/user/cancel_permission/<int:request_id>')
@login_required
def cancel_permission_request(request_id):
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    permission_request = db_session.query(PermissionRequest).get(request_id)
    
    if permission_request and permission_request.user_id == current_user.id and permission_request.status == 'pending':
        db_session.delete(permission_request)
        db_session.commit()
        flash('تم إلغاء طلب الإذن بنجاح')
    
    return redirect(url_for('user_leave_requests'))

@app.route('/user/permission_requests')
@login_required
def user_permission_requests():
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    permission_requests = db_session.query(PermissionRequest).filter_by(
        user_id=current_user.id
    ).order_by(PermissionRequest.created_at.desc()).all()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_permission_requests.html',
                         permission_requests=permission_requests,
                         notifications=notifications)


import glob
import re
def detect_salary_slips():
    """اكتشاف شيتات المرتب الجديدة تلقائياً"""
    try:
        salary_folder = app.config['SALARY_FOLDER']
        if not os.path.exists(salary_folder):
            os.makedirs(salary_folder)
        
        detected_count = 0
        for filename in os.listdir(salary_folder):
            if filename.lower().endswith('.pdf'):
                file_path = os.path.join(salary_folder, filename)
                
                # استخراج معلومات الملف
                file_info = extract_file_info_by_username(filename)
                
                if file_info:
                    username = file_info['username']
                    month = file_info['month']
                    arabic_month = file_info['arabic_month']
                    
                    # البحث عن المستخدم باستخدام username فقط
                    user = db_session.query(User).filter(
                        User.username == username
                    ).first()
                    
                    if user:
                        # التحقق من عدم وجود شيت مكرر
                        existing_slip = db_session.query(SalarySlip).filter_by(
                            user_id=user.id,
                            month=month
                        ).first()
                        
                        if not existing_slip:
                            # إنشاء سجل جديد
                            salary_slip = SalarySlip(
                                user_id=user.id,
                                month=month,
                                arabic_month=arabic_month,
                                file_path=file_path,
                                file_name=filename,
                                uploaded_by=1,  # النظام
                                is_auto_detected=True,
                                is_viewed=False
                            )
                            db_session.add(salary_slip)
                            
                            # إنشاء إشعار للمستخدم
                            create_notification(
                                user.id,
                                'شيت مرتب جديد',
                                f'تم اكتشاف شيت مرتب جديد لشهر {arabic_month}',
                                'salary',
                                salary_slip.id
                            )
                            
                            detected_count += 1
                            print(f"✅ تم اكتشاف شيت مرتب جديد: {filename} للمستخدم: {username} لشهر: {arabic_month}")
                    else:
                        print(f"⚠️  لم يتم العثور على مستخدم باسم: {username}")
        
        db_session.commit()
        return detected_count
        
    except Exception as e:
        print(f'❌ خطأ في اكتشاف شيتات المرتب: {str(e)}')
        db_session.rollback()
        return 0

import os
import re
from datetime import datetime

def extract_month_from_filename(filename):
    """استخراج الشهر من اسم الملف مع دعم العربية والإنجليزية"""
    try:
        name_without_ext = os.path.splitext(filename)[0]
        
        # البحث عن أنماط التاريخ المختلفة
        # النمط الجديد: 2025_9---510.PDF
        year_month_match = re.search(r'(\d{4})_(\d{1,2})', name_without_ext)
        if year_month_match:
            year = int(year_month_match.group(1))
            month = int(year_month_match.group(2))
            return f"{year}-{month:02d}"
        
        # الأنماط القديمة (للتوافق مع الإصدارات السابقة)
        month_year_match = re.search(r'(\d{4})-(\d{2})', name_without_ext)  # 2024-01
        if month_year_match:
            return month_year_match.group(0)
        
        # الأشهر العربية والإنجليزية
        month_names = {
            # العربية الكاملة
            'يناير': '01', 'فبراير': '02', 'مارس': '03', 'أبريل': '04',
            'مايو': '05', 'يونيو': '06', 'يوليو': '07', 'أغسطس': '08',
            'سبتمبر': '09', 'أكتوبر': '10', 'نوفمبر': '11', 'ديسمبر': '12',
            # العربية المختصرة
            'ينا': '01', 'فبر': '02', 'مار': '03', 'أبر': '04',
            'ماي': '05', 'يون': '06', 'يول': '07', 'أغس': '08',
            'سبت': '09', 'أكت': '10', 'نوف': '11', 'ديس': '12',
            # الإنجليزية الكاملة
            'january': '01', 'february': '02', 'march': '03', 'april': '04',
            'may': '05', 'june': '06', 'july': '07', 'august': '08',
            'september': '09', 'october': '10', 'november': '11', 'december': '12',
            # الإنجليزية المختصرة
            'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04', 'may': '05', 'jun': '06',
            'jul': '07', 'aug': '08', 'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'
        }
        
        for month_name, month_num in month_names.items():
            if re.search(month_name, name_without_ext, re.IGNORECASE):
                current_year = datetime.now().year
                return f"{current_year}-{month_num}"
        
        # إذا لم يتم العثور على شهر، نستخدم الشهر الحالي
        return datetime.now().strftime('%Y-%m')
        
    except Exception as e:
        print(f'خطأ في استخراج الشهر من {filename}: {str(e)}')
        return datetime.now().strftime('%Y-%m')

def get_arabic_month_name(month_num):
    """الحصول على اسم الشهر بالعربية"""
    arabic_months = {
        '01': 'يناير', '02': 'فبراير', '03': 'مارس', '04': 'أبريل',
        '05': 'مايو', '06': 'يونيو', '07': 'يوليو', '08': 'أغسطس',
        '09': 'سبتمبر', '10': 'أكتوبر', '11': 'نوفمبر', '12': 'ديسمبر'
    }
    return arabic_months.get(month_num, 'غير معروف')

def extract_file_info_by_username(filename):
    """استخراج معلومات المستخدم والشهر من اسم الملف مع دعم العربية"""
    try:
        # إزالة الامتداد
        name_without_ext = os.path.splitext(filename)[0]
        
        print(f"معالجة الملف: {filename}")
        print(f"الاسم بدون امتداد: {name_without_ext}")
        
        # محاولة استخراج الشهر من اسم الملف
        month = extract_month_from_filename(filename)
        print(f"الشهر المستخرج: {month}")
        
        # استخراج اسم الشهر بالعربية
        month_num = month.split('-')[1] if '-' in month else datetime.now().strftime('%m')
        arabic_month = get_arabic_month_name(month_num)
        print(f"اسم الشهر بالعربية: {arabic_month}")
        
        # استخراج اسم المستخدم
        username = None
        
        # النمط 1: تاريخ_شهر---اسم مستخدم (2025_9---510)
        pattern1 = r'\d{4}_\d{1,2}---(\w+)'
        match1 = re.search(pattern1, name_without_ext)
        if match1:
            username = match1.group(1)
            print(f"تم التعرف على النمط 1 - اسم المستخدم: {username}")
        
        # النمط 2: اسم مستخدم_تاريخ (username_2024-01)
        if not username:
            pattern2 = r'^([^_]+)_\d'
            match2 = re.search(pattern2, name_without_ext)
            if match2:
                username = match2.group(1)
                print(f"تم التعرف على النمط 2 - اسم المستخدم: {username}")
        
        # النمط 3: اسم مستخدم فقط (username.pdf)
        if not username:
            # إزالة أنماط التاريخ المعروفة
            temp_username = re.sub(r'_\d{4}-\d{2}', '', name_without_ext)
            temp_username = re.sub(r'_\d{4}_\d{1,2}', '', temp_username)
            temp_username = re.sub(r'---\d+$', '', temp_username)
            temp_username = re.sub(r'_\d{6}', '', temp_username)
            
            # إزالة أسماء الأشهر العربية والإنجليزية
            month_patterns = [
                # العربية الكاملة
                r'_يناير', r'_فبراير', r'_مارس', r'_أبريل', r'_مايو', r'_يونيو',
                r'_يوليو', r'_أغسطس', r'_سبتمبر', r'_أكتوبر', r'_نوفمبر', r'_ديسمبر',
                # العربية المختصرة
                r'_ينا', r'_فبر', r'_مار', r'_أبر', r'_ماي', r'_يون',
                r'_يول', r'_أغس', r'_سبت', r'_أكت', r'_نوف', r'_ديس',
                # الإنجليزية
                r'_january', r'_february', r'_march', r'_april', r'_may', r'_june',
                r'_july', r'_august', r'_september', r'_october', r'_november', r'_december',
                r'_jan', r'_feb', r'_mar', r'_apr', r'_may', r'_jun',
                r'_jul', r'_aug', r'_sep', r'_oct', r'_nov', r'_dec'
            ]
            
            for pattern in month_patterns:
                temp_username = re.sub(pattern, '', temp_username, flags=re.IGNORECASE)
            
            username = temp_username.strip(' _-')
            print(f"تم التعرف على النمط 3 - اسم المستخدم: {username}")
        
        if username:
            return {
                'username': username, 
                'month': month,  # التنسيق الرقمي: 2025-09
                'arabic_month': arabic_month,  # اسم الشهر العربي: سبتمبر
                'display_month': arabic_month  # اسم العرض العربي
            }
        else:
            print(f"لم يتم العثور على اسم مستخدم في: {filename}")
            return None
            
    except Exception as e:
        print(f'خطأ في تحليل اسم الملف {filename}: {str(e)}')
        return None

@app.route('/user/salary_slips')
@login_required
def user_salary_slips():
    if current_user.is_admin:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    # اكتشاف شيتات المرتب الجديدة تلقائياً قبل العرض
    detect_salary_slips()
    
    # الحصول على جميع شيتات المرتب للمستخدم
    all_salary_slips = db_session.query(SalarySlip).filter_by(
        user_id=current_user.id
    ).order_by(SalarySlip.month.desc()).all()
    
    # تصفية الشيتات التي لا يوجد لها ملف فعلي
    valid_salary_slips = []
    for slip in all_salary_slips:
        if slip.file_path and os.path.exists(slip.file_path):
            valid_salary_slips.append(slip)
        else:
            # حذف السجل من قاعدة البيانات إذا لم يكن الملف موجوداً
            print(f"تحذير: ملف شيت المرتب غير موجود - {slip.file_path}")
            db_session.delete(slip)
    
    # تحديث حالة المشاهدة للشيتات الصالحة فقط
    for slip in valid_salary_slips:
        if not slip.is_viewed:
            slip.is_viewed = True
    
    db_session.commit()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_salary_slips.html',
                         salary_slips=valid_salary_slips,  # إرسال القائمة المصفاة فقط
                         notifications=notifications)

@app.route('/admin/upload_salary_slip', methods=['GET', 'POST'])
@login_required
def admin_upload_salary_slip():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    if request.method == 'POST':
        user_id = request.form['user_id']
        month = request.form['month']  # 2025-09
        file = request.files['salary_file']
        
        if file and allowed_file(file.filename):
            user = db_session.query(User).get(user_id)
            if not user:
                flash('المستخدم غير موجود')
                return redirect(url_for('admin_upload_salary_slip'))
            
            # استخراج اسم الشهر العربي
            month_num = month.split('-')[1] if '-' in month else datetime.now().strftime('%m')
            arabic_month = get_arabic_month_name(month_num)
            
            existing_slip = db_session.query(SalarySlip).filter_by(
                user_id=user_id,
                month=month
            ).first()
            
            if existing_slip:
                flash('يوجد already شيت مرتب لنفس الشهر')
                return redirect(url_for('admin_upload_salary_slip'))
            
            filename = secure_filename(f"{user_id}_{month}_{file.filename}")
            file_path = os.path.join(app.config['SALARY_FOLDER'], filename)
            file.save(file_path)
            
            salary_slip = SalarySlip(
                user_id=user_id,
                month=month,
                arabic_month=arabic_month,  # حفظ اسم الشهر العربي
                file_path=file_path,
                file_name=filename,
                uploaded_by=current_user.id
            )
            db_session.add(salary_slip)
            
            create_notification(
                user_id,
                'شيت مرتب جديد',
                f'تم رفع شيت المرتب لشهر {arabic_month}',  # استخدام الاسم العربي
                'salary',
                salary_slip.id
            )
            
            db_session.commit()
            flash('تم رفع شيت المرتب بنجاح')
            return redirect(url_for('admin_salary_slips'))
        else:
            flash('الملف غير مسموح به')
    
    users = db_session.query(User).filter_by(is_admin=False).all()
    notifications_list = db_session.query(Notification).filter_by(
        user_id=current_user.id, is_read=False
    ).order_by(Notification.created_at.desc()).limit(5).all()
    
    return render_template('admin_upload_salary_slip.html', 
                         users=users, 
                         notifications=notifications_list)



def apply_department_structure_to_schedule(schedule_id):
    """تطبيق هيكل القسم على جدول محدد"""
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return {'success': False, 'message': 'الجدول غير موجود'}
        
        department = db_session.query(Department).get(schedule.department_id)
        if not department or not department.schedule_structure:
            return {'success': False, 'message': 'القسم ليس لديه هيكل'}
        
        # تحليل هيكل القسم
        department_structure = json.loads(department.schedule_structure)
        
        # تحليل بيانات الجدول الحالية
        current_schedule_data = {}
        if schedule.schedule_data:
            try:
                current_schedule_data = json.loads(schedule.schedule_data)
            except:
                current_schedule_data = {}
        
        # إنشاء هيكل جديد مطابق لهيكل القسم
        new_schedule_data = {
            'department': department.name,
            'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
            'week_end_date': schedule.week_end_date.strftime('%Y-%m-%d'),
            'source': 'department_structure_forced',
            'schedule': []
        }
        
        # أيام الأسبوع
        days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        
        # تعبئة الأيام من هيكل القسم
        for i, day_name in enumerate(days_of_week):
            current_date = schedule.week_start_date + timedelta(days=i)
            
            # البحث عن اليوم في هيكل القسم
            day_structure = None
            if isinstance(department_structure, dict) and 'schedule' in department_structure:
                for day in department_structure['schedule']:
                    if isinstance(day, dict) and day.get('day') == day_name:
                        day_structure = day
                        break
            
            # إنشاء إدخال اليوم
            if day_structure:
                # نسخ الهيكل تماماً من القسم
                day_entry = day_structure.copy()
                day_entry['date'] = current_date.strftime('%Y-%m-%d')
                day_entry['department'] = department.name
                
                # التأكد من وجود جميع الحقول المطلوبة
                required_fields = ['morning_shift', 'evening_shift', 'night_shift', 'job']
                for field in required_fields:
                    if field not in day_entry:
                        day_entry[field] = ''
                
                # إذا كان job فارغاً، استخدم القيمة الافتراضية
                if not day_entry.get('job'):
                    day_entry['job'] = 'موظف'
            else:
                # إذا لم يكن اليوم موجوداً في الهيكل
                day_entry = {
                    'day': day_name,
                    'date': current_date.strftime('%Y-%m-%d'),
                    'department': department.name,
                    'morning_shift': '',
                    'evening_shift': '',
                    'night_shift': '',
                    'job': 'موظف'
                }
            
            # المحاولة للحفاظ على بيانات الموظفين من الجدول الحالي
            if current_schedule_data and 'schedule' in current_schedule_data:
                for current_day in current_schedule_data['schedule']:
                    if isinstance(current_day, dict) and current_day.get('day') == day_name:
                        # نسخ بيانات الموظفين فقط
                        shift_fields = ['morning_shift', 'evening_shift', 'night_shift']
                        for field in shift_fields:
                            if current_day.get(field):
                                day_entry[field] = current_day[field]
                        break
            
            new_schedule_data['schedule'].append(day_entry)
        
        # تحديث الجدول بالهيكل الجديد
        schedule.schedule_data = json.dumps(new_schedule_data, ensure_ascii=False)
        schedule.structure_version = department.schedule_structure_version
        schedule.structure_hash = hashlib.sha256(
            json.dumps(department_structure, ensure_ascii=False).encode('utf-8')
        ).hexdigest()
        schedule.is_generated_from_structure = True
        schedule.sync_status = 'synced'
        schedule.last_sync_check = datetime.now()
        
        db_session.commit()
        
        return {
            'success': True,
            'message': f'تم تطبيق هيكل القسم على الجدول',
            'schedule_id': schedule.id,
            'department_id': department.id,
            'days_updated': len(new_schedule_data['schedule'])
        }
        
    except Exception as e:
        db_session.rollback()
        return {'success': False, 'message': f'خطأ في تطبيق الهيكل: {str(e)}'}
    


@app.route('/admin/apply_structure/<int:schedule_id>')
@login_required
def admin_apply_structure(schedule_id):
    """تطبيق هيكل القسم على جدول محدد"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        result = apply_department_structure_to_schedule(schedule_id)
        
        if result['success']:
            flash(result['message'], 'success')
        else:
            flash(result['message'], 'error')
        
    except Exception as e:
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(request.referrer or url_for('admin_schedules'))


def apply_structure_to_all_department_schedules(department_id):
    """تطبيق الهيكل على جميع جداول القسم"""
    try:
        department = db_session.query(Department).get(department_id)
        if not department or not department.schedule_structure:
            return {'success': False, 'message': 'القسم ليس لديه هيكل'}
        
        schedules = db_session.query(WeeklySchedule).filter_by(
            department_id=department_id
        ).all()
        
        applied_count = 0
        failed_count = 0
        
        for schedule in schedules:
            if schedule.is_approved and schedule.is_locked:
                continue  # تخطي الجداول المعتمدة المقفلة
            
            result = apply_department_structure_to_schedule(schedule.id)
            if result['success']:
                applied_count += 1
                print(f"✓ تم تطبيق الهيكل على الجدول {schedule.id}")
            else:
                failed_count += 1
                print(f"❌ فشل في تطبيق الهيكل على الجدول {schedule.id}: {result['message']}")
        
        db_session.commit()
        
        return {
            'success': True,
            'message': f'تم تطبيق الهيكل على {applied_count} جدول، فشل {failed_count}',
            'applied_count': applied_count,
            'failed_count': failed_count
        }
        
    except Exception as e:
        db_session.rollback()
        return {'success': False, 'message': f'خطأ عام: {str(e)}'}


@app.route('/admin/view_structure/<int:dept_id>')
@login_required
def admin_view_structure(dept_id):
    """عرض هيكل القسم"""
    if not current_user.is_admin and not current_user.is_manager:
        flash('غير مصرح لك بالوصول', 'error')
        return redirect(url_for('user_dashboard'))
    
    department = db_session.query(Department).get(dept_id)
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('admin_departments'))
    
    structure_data = {}
    if department.schedule_structure:
        try:
            structure_data = json.loads(department.schedule_structure)
        except json.JSONDecodeError:
            flash('خطأ في تحليل بيانات الهيكل', 'error')
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_view_structure.html',
                         department=department,
                         structure_data=structure_data,
                         notifications=notifications)


@app.route('/admin/apply_structure_to_all/<int:department_id>')
@login_required
def admin_apply_structure_to_all(department_id):
    """تطبيق هيكل القسم على جميع جداوله"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        result = apply_structure_to_all_department_schedules(department_id)
        
        if result['success']:
            flash(result['message'], 'success')
        else:
            flash(result['message'], 'error')
        
    except Exception as e:
        flash(f'حدث خطأ: {str(e)}', 'error')
    
    return redirect(request.referrer or url_for('admin_schedules'))


def sync_schedule_with_department_structure(schedule_id):
    """مزامنة جدول مع هيكل القسم مع فرض المطابقة"""
    try:
        schedule = db_session.query(WeeklySchedule).get(schedule_id)
        if not schedule:
            return False
        
        department = db_session.query(Department).get(schedule.department_id)
        if not department or not department.schedule_structure:
            return False
        
        # تطبيق الهيكل بالقوة
        result = apply_department_structure_to_schedule(schedule.id)
        
        # إرسال إشعار إذا نجحت المزامنة
        if result['success']:
            create_notification(
                schedule.created_by,
                'تم تحديث هيكل الجدول',
                f'تم تحديث جدول الأسبوع {schedule.week_start_date} لمطابقة هيكل القسم {department.name}',
                'schedule_synced',
                related_id=schedule.id
            )
        
        return result['success']
        
    except Exception as e:
        print(f"خطأ في مزامنة الجدول: {str(e)}")
        return False

@app.route('/admin/salary_slips', methods=['GET', 'POST'])
@login_required
def admin_salary_slips():
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    if request.method == 'POST':
        # Manual upload functionality
        user_id = request.form['user_id']
        month = request.form['month']
        file = request.files['file']
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['SALARY_FOLDER'], filename)
            file.save(file_path)
            
            salary_slip = SalarySlip(
                user_id=user_id,
                month=month,
                file_path=file_path,
                file_name=filename,
                uploaded_by=current_user.id,
                is_auto_detected=False
            )
            db_session.add(salary_slip)
            
            # Notify user
            user = db_session.query(User).get(user_id)
            create_notification(
                user_id,
                'مستحق راتب جديد',
                f'تم إضافة مستحق راتب جديد لشهر {month}',
                'salary_slip',
                action_url=url_for('user_salary_slips')
            )
            
            db_session.commit()
            flash('تم رفع مستحق الراتب بنجاح')
    
    # Auto-detect on page load
    detected_count = detect_salary_slips()
    if detected_count > 0:
        flash(f'تم اكتشاف {detected_count} مستحق راتب جديد تلقائياً')
    
    users = db_session.query(User).filter_by(is_admin=False).all()
    salary_slips = db_session.query(SalarySlip).order_by(SalarySlip.month.desc()).all()
    
    # Add user info to salary slips for display
    for slip in salary_slips:
        slip.user = db_session.query(User).get(slip.user_id)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_salary_slips.html',
                         users=users,
                         salary_slips=salary_slips,
                         notifications=notifications)

@app.route('/download_salary_slip/<int:slip_id>')
@login_required
def download_salary_slip(slip_id):
    salary_slip = db_session.query(SalarySlip).get(slip_id)
    
    # Check permissions
    if not current_user.is_admin and not current_user.is_manager and salary_slip.user_id != current_user.id:
        flash('غير مصرح لك بالوصول إلى هذا الملف')
        return redirect(url_for('user_dashboard'))
    
    if salary_slip and os.path.exists(salary_slip.file_path):
        return send_file(salary_slip.file_path, as_attachment=True)
    else:
        flash('الملف غير موجود')
        return redirect(request.referrer or url_for('user_dashboard'))




@app.route('/user/rewards_penalties')
@login_required
def user_rewards_penalties():
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    rewards_penalties = db_session.query(RewardPenalty).filter_by(
        user_id=current_user.id
    ).order_by(RewardPenalty.effective_date.desc()).all()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_rewards_penalties.html',
                         rewards_penalties=rewards_penalties,
                         notifications=notifications)




@app.route('/admin/all_leaves')
@login_required
def admin_all_leaves():
    """عرض جميع الإجازات في كل قسم حسب الشهر المالي"""
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    # الحصول على معاملات التصفية
    month_filter = request.args.get('month', '')
    department_filter = request.args.get('department', '')
    status_filter = request.args.get('status', '')
    
    # بناء الاستعلام الأساسي بدون joinedload
    query = db_session.query(LeaveRequest).options(joinedload(LeaveRequest.approver))

    
    # تطبيق الفلاتر
    if department_filter and department_filter.isdigit():
        query = query.filter(LeaveRequest.department_id == int(department_filter))
    
    if status_filter:
        query = query.filter(LeaveRequest.status == status_filter)
    
    # الحصول على جميع الإجازات
    all_leaves = query.order_by(LeaveRequest.created_at.desc()).all()
    
    # تحميل البيانات المرتبطة يدوياً
    for leave in all_leaves:
        leave.user = db_session.query(User).get(leave.user_id)
        leave.department = db_session.query(Department).get(leave.department_id)
    
    # باقي الكود يبقى كما هو...
    # تصفية حسب الشهر المالي إذا تم تحديده
    filtered_leaves = []
    if month_filter:
        try:
            # تحويل month_filter إلى تاريخ (تنسيق: YYYY-MM)
            year, month = map(int, month_filter.split('-'))
            
            # حساب بداية ونهاية الشهر المالي (26 إلى 25)
            if month == 12:
                financial_month_start = date(year, 12, 26)
                financial_month_end = date(year + 1, 1, 25)
            else:
                financial_month_start = date(year, month, 26)
                financial_month_end = date(year, month + 1, 25)
            
            # تصفية الإجازات التي تتداخل مع الشهر المالي
            for leave in all_leaves:
                if (leave.start_date <= financial_month_end and 
                    leave.end_date >= financial_month_start):
                    filtered_leaves.append(leave)
        except ValueError:
            filtered_leaves = all_leaves
    else:
        filtered_leaves = all_leaves
    
    # إحصائيات
    stats = {
        'total': len(filtered_leaves),
        'approved': len([l for l in filtered_leaves if l.status == 'approved']),
        'pending': len([l for l in filtered_leaves if l.status == 'pending']),
        'rejected': len([l for l in filtered_leaves if l.status == 'rejected']),
        'current_month': len([l for l in filtered_leaves if is_in_current_financial_month(l.start_date)])
    }
    
    # تجميع البيانات للعرض
    leaves_by_department = {}
    for leave in filtered_leaves:
        dept_name = leave.department.name if leave.department else 'غير معين'
        if dept_name not in leaves_by_department:
            leaves_by_department[dept_name] = []
        leaves_by_department[dept_name].append(leave)
    
    # الحصول على الأقسام للفلتر
    departments = db_session.query(Department).all()
    
    # توليد قائمة الأشهر المالية للفلتر
    financial_months = generate_financial_months()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_all_leaves.html',
                         leaves=filtered_leaves,
                         leaves_by_department=leaves_by_department,
                         departments=departments,
                         financial_months=financial_months,
                         selected_month=month_filter,
                         selected_department=department_filter,
                         selected_status=status_filter,
                         stats=stats,
                         notifications=notifications)


def get_approver_name(approver_id):
    """الحصول على اسم الموافق على الإجازة"""
    if not approver_id:
        return "غير معين"
    
    approver = db_session.query(User).get(approver_id)
    return approver.name if approver else "غير معين"

@app.context_processor
def utility_processor():
    def get_approver_name(approver_id):
        if not approver_id:
            return "غير معين"
        approver = db_session.query(User).get(approver_id)
        return approver.name if approver else "غير معين"
    
    return dict(get_approver_name=get_approver_name)

def generate_financial_months():
    """توليد قائمة بالأشهر المالية للسنوات السابقة والحالية والمستقبلية"""
    months = []
    current_date = date.today()
    current_year = current_date.year
    
    # إضافة الأشهر للعام الحالي
    for month_num in range(1, 13):
        # تنسيقين: YYYY-MM للفلتر و MM للعرض البسيط
        year_month = f"{current_year}-{month_num:02d}"
        month_only = str(month_num)
        
        month_name = get_month_name_arabic(month_num)
        prev_month_name = get_month_name_arabic(month_num-1 if month_num>1 else 12)
        prev_year = current_year if month_num > 1 else current_year - 1
        
        display_name = f"{month_name} {current_year} (26 {prev_month_name} {prev_year} - 25 {month_name} {current_year})"
        
        months.append({
            'value': year_month,  # للفلتر
            'simple_value': month_only,  # للتوافق مع القيم القديمة
            'display': display_name
        })
    
    # تحديد الشهر الحالي
    if current_date.day >= 26:
        current_financial_month = current_date.month
    else:
        current_financial_month = current_date.month - 1 if current_date.month > 1 else 12
    
    # وضع علامة على الشهر الحالي
    for month in months:
        if month['simple_value'] == str(current_financial_month):
            month['display'] += " - الحالي"
            break
    
    # إضافة أشهر من العام الماضي
    for month_num in range(1, 13):
        year_month = f"{current_year-1}-{month_num:02d}"
        month_only = str(month_num)
        
        month_name = get_month_name_arabic(month_num)
        prev_month_name = get_month_name_arabic(month_num-1 if month_num>1 else 12)
        prev_year = current_year-1 if month_num > 1 else current_year - 2
        
        display_name = f"{month_name} {current_year-1} (26 {prev_month_name} {prev_year} - 25 {month_name} {current_year-1})"
        
        months.insert(0, {
            'value': year_month,
            'simple_value': month_only,
            'display': display_name
        })
    
    return months

def get_month_name_arabic(month):
    """إرجاع اسم الشهر بالعربية"""
    months = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    return months.get(month, '')

def is_in_current_financial_month(leave_date):
    """التحقق مما إذا كان تاريخ الإجارة يقع في الشهر المالي الحالي"""
    today = date.today()
    
    # حساب بداية ونهاية الشهر المالي الحالي
    if today.day >= 26:
        # الشهر المالي يبدأ من 26 من الشهر الحالي إلى 25 من الشهر القادم
        financial_month_start = date(today.year, today.month, 26)
        if today.month == 12:
            financial_month_end = date(today.year + 1, 1, 25)
        else:
            financial_month_end = date(today.year, today.month + 1, 25)
    else:
        # الشهر المالي يبدأ من 26 من الشهر الماضي إلى 25 من الشهر الحالي
        if today.month == 1:
            financial_month_start = date(today.year - 1, 12, 26)
            financial_month_end = date(today.year, 1, 25)
        else:
            financial_month_start = date(today.year, today.month - 1, 26)
            financial_month_end = date(today.year, today.month, 25)
    
    return financial_month_start <= leave_date <= financial_month_end


@app.route('/admin/export_leaves')
@login_required
def admin_export_leaves():
    """تصدير بيانات الإجازات إلى Excel - بدون استخدام العلاقات"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        # الحصول على معاملات التصفية
        month_filter = request.args.get('month', '')
        department_filter = request.args.get('department', '')
        status_filter = request.args.get('status', '')
        
        # بناء الاستعلام الأساسي
        query = db_session.query(LeaveRequest)
        
        if department_filter and department_filter.isdigit():
            query = query.filter(LeaveRequest.department_id == int(department_filter))
        
        if status_filter:
            query = query.filter(LeaveRequest.status == status_filter)
        
        all_leaves = query.order_by(LeaveRequest.created_at.desc()).all()
        
        # الحصول على بيانات المستخدمين والأقسام مسبقاً
        user_ids = [leave.user_id for leave in all_leaves]
        department_ids = [leave.department_id for leave in all_leaves]
        
        users = {user.id: user for user in db_session.query(User).filter(User.id.in_(user_ids)).all()}
        departments = {dept.id: dept for dept in db_session.query(Department).filter(Department.id.in_(department_ids)).all()}
        
        # تطبيق فلتر الشهر المالي
        filtered_leaves = []
        if month_filter:
            try:
                year, month = map(int, month_filter.split('-'))
                if month == 12:
                    financial_month_start = date(year, 12, 26)
                    financial_month_end = date(year + 1, 1, 25)
                else:
                    financial_month_start = date(year, month, 26)
                    financial_month_end = date(year, month + 1, 25)
                
                for leave in all_leaves:
                    if (leave.start_date <= financial_month_end and 
                        leave.end_date >= financial_month_start):
                        filtered_leaves.append(leave)
            except ValueError:
                filtered_leaves = all_leaves
        else:
            filtered_leaves = all_leaves
        
        # باقي كود التصدير يبقى كما هو...
        # إنشاء ملف Excel
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "تقرير الإجازات"
        
        # إضافة العنوان
        title = "تقرير الإجازات"
        if month_filter:
            month_name = get_month_name_arabic(int(month_filter.split('-')[1]))
            year = month_filter.split('-')[0]
            title += f" - {month_name} {year}"
        
        if department_filter:
            department = db_session.query(Department).get(int(department_filter))
            if department:
                title += f" - قسم {department.name}"
        
        if status_filter:
            status_names = {'approved': 'المقبولة', 'pending': 'المعلقة', 'rejected': 'المرفوضة'}
            title += f" - {status_names.get(status_filter, status_filter)}"
        
        worksheet.merge_cells('A1:H1')
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # رؤوس الأعمدة
        headers = [
            'اسم الموظف', 'القسم', 'نوع الإجازة', 'تاريخ البداية', 
            'تاريخ النهاية', 'عدد الأيام', 'الحالة', 'ملاحظات'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # بيانات الإجازات
        row = 4
        for leave in filtered_leaves:
            user = users.get(leave.user_id)
            department = departments.get(leave.department_id)
            
            worksheet.cell(row=row, column=1, value=user.name if user else 'غير معين')
            worksheet.cell(row=row, column=2, value=department.name if department else 'غير معين')
            worksheet.cell(row=row, column=3, value=leave.leave_type)
            worksheet.cell(row=row, column=4, value=leave.start_date.strftime('%Y-%m-%d'))
            worksheet.cell(row=row, column=5, value=leave.end_date.strftime('%Y-%m-%d'))
            worksheet.cell(row=row, column=6, value=leave.total_days)
            
            # تحويل الحالة للعربية
            status_arabic = {
                'approved': 'مقبولة',
                'pending': 'معلقة', 
                'rejected': 'مرفوضة'
            }.get(leave.status, leave.status)
            
            worksheet.cell(row=row, column=7, value=status_arabic)
            worksheet.cell(row=row, column=8, value=leave.reason or '')
            
            row += 1
        
        # ضبط عرض الأعمدة
        column_widths = [25, 20, 15, 15, 15, 12, 12, 30]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width
        
        workbook.save(output)
        output.seek(0)
        
        filename = f"تقرير_الإجازات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'خطأ في تصدير الملف: {str(e)}', 'error')
        return redirect(url_for('admin_all_leaves'))

# في قسم inject_manager_stats (البحث عن هذا القسم في app.py):

@app.context_processor
def inject_manager_stats():
    """Inject manager stats into all templates"""
    try:
        if current_user.is_authenticated and hasattr(current_user, 'is_manager') and current_user.is_manager:
            # استخدام primary_manager_id بدلاً من manager_id
            managed_departments = db_session.query(Department).filter_by(primary_manager_id=current_user.id).all()
            department_ids = [dept.id for dept in managed_departments]
            
            stats = {
                'pending_leaves': db_session.query(LeaveRequest).filter(
                    LeaveRequest.department_id.in_(department_ids),
                    LeaveRequest.status == 'pending'
                ).count(),
                'pending_permissions': db_session.query(PermissionRequest).filter(
                    PermissionRequest.department_id.in_(department_ids),
                    PermissionRequest.status == 'pending'
                ).count(),
                'pending_advances': db_session.query(AdvanceRequest).filter(
                    AdvanceRequest.department_id.in_(department_ids),
                    AdvanceRequest.status == 'pending'
                ).count(),
                'department_employees': db_session.query(User).filter(
                    User.department_id.in_(department_ids),
                    User.is_admin == False
                ).count()
            }
            return {'stats': stats}
    except Exception as e:
        print(f"Error in inject_manager_stats: {e}")
    
    return {'stats': {}}


@app.route('/admin/all_permissions')
@login_required
def admin_all_permissions():
    """عرض جميع طلبات الإذن في كل قسم حسب الشهر المالي"""
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    # الحصول على معاملات التصفية
    month_filter = request.args.get('month', '')
    department_filter = request.args.get('department', '')
    status_filter = request.args.get('status', '')
    
    # بناء الاستعلام الأساسي
    query = db_session.query(PermissionRequest)
    
    # تطبيق الفلاتر
    if department_filter and department_filter.isdigit():
        query = query.filter(PermissionRequest.department_id == int(department_filter))
    
    if status_filter:
        query = query.filter(PermissionRequest.status == status_filter)
    
    # الحصول على جميع طلبات الإذن
    all_permissions = query.order_by(PermissionRequest.created_at.desc()).all()
    
    # تحميل البيانات المرتبطة يدوياً
    for permission in all_permissions:
        permission.user = db_session.query(User).get(permission.user_id)
        permission.department = db_session.query(Department).get(permission.department_id)
    
    # تصفية حسب الشهر المالي إذا تم تحديده
    filtered_permissions = []
    if month_filter:
        try:
            # تحويل month_filter إلى تاريخ (تنسيق: YYYY-MM)
            year, month = map(int, month_filter.split('-'))
            
            # حساب بداية ونهاية الشهر المالي (26 إلى 25)
            if month == 12:
                financial_month_start = date(year, 12, 26)
                financial_month_end = date(year + 1, 1, 25)
            else:
                financial_month_start = date(year, month, 26)
                financial_month_end = date(year, month + 1, 25)
            
            # تصفية طلبات الإذن التي تتداخل مع الشهر المالي
            for permission in all_permissions:
                if (permission.date <= financial_month_end and 
                    permission.date >= financial_month_start):
                    filtered_permissions.append(permission)
        except ValueError:
            filtered_permissions = all_permissions
    else:
        filtered_permissions = all_permissions
    
    # إحصائيات
    stats = {
        'total': len(filtered_permissions),
        'approved': len([p for p in filtered_permissions if p.status == 'approved']),
        'pending': len([p for p in filtered_permissions if p.status == 'pending']),
        'rejected': len([p for p in filtered_permissions if p.status == 'rejected']),
        'current_month': len([p for p in filtered_permissions if is_in_current_financial_month(p.date)])
    }
    
    # تجميع البيانات للعرض
    permissions_by_department = {}
    for permission in filtered_permissions:
        dept_name = permission.department.name if permission.department else 'غير معين'
        if dept_name not in permissions_by_department:
            permissions_by_department[dept_name] = []
        permissions_by_department[dept_name].append(permission)
    
    # الحصول على الأقسام للفلتر
    departments = db_session.query(Department).all()
    
    # توليد قائمة الأشهر المالية للفلتر
    financial_months = generate_financial_months()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_all_permissions.html',
                         permissions=filtered_permissions,
                         permissions_by_department=permissions_by_department,
                         departments=departments,
                         financial_months=financial_months,
                         selected_month=month_filter,
                         selected_department=department_filter,
                         selected_status=status_filter,
                         stats=stats,
                         notifications=notifications)

@app.route('/admin/export_permissions')
@login_required
def admin_export_permissions():
    """تصدير بيانات طلبات الإذن إلى Excel"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        # الحصول على معاملات التصفية
        month_filter = request.args.get('month', '')
        department_filter = request.args.get('department', '')
        status_filter = request.args.get('status', '')
        
        # بناء الاستعلام الأساسي
        query = db_session.query(PermissionRequest)
        
        if department_filter and department_filter.isdigit():
            query = query.filter(PermissionRequest.department_id == int(department_filter))
        
        if status_filter:
            query = query.filter(PermissionRequest.status == status_filter)
        
        all_permissions = query.order_by(PermissionRequest.created_at.desc()).all()
        
        # الحصول على بيانات المستخدمين والأقسام مسبقاً
        user_ids = [permission.user_id for permission in all_permissions]
        department_ids = [permission.department_id for permission in all_permissions]
        
        users = {user.id: user for user in db_session.query(User).filter(User.id.in_(user_ids)).all()}
        departments = {dept.id: dept for dept in db_session.query(Department).filter(Department.id.in_(department_ids)).all()}
        
        # تطبيق فلتر الشهر المالي
        filtered_permissions = []
        if month_filter:
            try:
                year, month = map(int, month_filter.split('-'))
                if month == 12:
                    financial_month_start = date(year, 12, 26)
                    financial_month_end = date(year + 1, 1, 25)
                else:
                    financial_month_start = date(year, month, 26)
                    financial_month_end = date(year, month + 1, 25)
                
                for permission in all_permissions:
                    if (permission.date <= financial_month_end and 
                        permission.date >= financial_month_start):
                        filtered_permissions.append(permission)
            except ValueError:
                filtered_permissions = all_permissions
        else:
            filtered_permissions = all_permissions
        
        # إنشاء ملف Excel
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "تقرير الإذونات"
        
        # إضافة العنوان
        title = "تقرير طلبات الإذن"
        if month_filter:
            month_name = get_month_name_arabic(int(month_filter.split('-')[1]))
            year = month_filter.split('-')[0]
            title += f" - {month_name} {year}"
        
        if department_filter:
            department = db_session.query(Department).get(int(department_filter))
            if department:
                title += f" - قسم {department.name}"
        
        if status_filter:
            status_names = {'approved': 'المقبولة', 'pending': 'المعلقة', 'rejected': 'المرفوضة'}
            title += f" - {status_names.get(status_filter, status_filter)}"
        
        worksheet.merge_cells('A1:H1')
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # رؤوس الأعمدة
        headers = [
            'اسم الموظف', 'القسم', 'نوع الإذن', 'التاريخ', 
            'المدة (دقيقة)', 'الحالة', 'السبب'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # بيانات طلبات الإذن
        row = 4
        for permission in filtered_permissions:
            user = users.get(permission.user_id)
            department = departments.get(permission.department_id)
            
            worksheet.cell(row=row, column=1, value=user.name if user else 'غير معين')
            worksheet.cell(row=row, column=2, value=department.name if department else 'غير معين')
            worksheet.cell(row=row, column=3, value=permission.permission_type)
            worksheet.cell(row=row, column=4, value=permission.date.strftime('%Y-%m-%d'))
            worksheet.cell(row=row, column=5, value=permission.time or '-')
            
            # تحويل الحالة للعربية
            status_arabic = {
                'approved': 'مقبولة',
                'pending': 'معلقة', 
                'rejected': 'مرفوضة'
            }.get(permission.status, permission.status)
            
            worksheet.cell(row=row, column=6, value=status_arabic)
            worksheet.cell(row=row, column=7, value=permission.reason or '')
            
            row += 1
        
        # ضبط عرض الأعمدة
        column_widths = {
            'A': 25,  # اسم الموظف
            'B': 20,  # القسم
            'C': 15,  # نوع الإذن
            'D': 12,  # التاريخ
            'E': 10,  # من
            'F': 10,  # إلى
            'G': 15,  # المدة
            'H': 12,  # الحالة
            'I': 25   # السبب
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        workbook.save(output)
        output.seek(0)
        
        filename = f"تقرير_الإذونات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'خطأ في تصدير الملف: {str(e)}', 'error')
        return redirect(url_for('admin_all_permissions'))


@app.route('/manager/rewards_penalties', methods=['GET', 'POST'])
@login_required
def manager_rewards_penalties():
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    managed_departments = db_session.query(Department).filter_by(primary_manager_id=current_user.id).all()
    department_ids = [dept.id for dept in managed_departments]
    
    if request.method == 'POST':
        user_id = request.form['user_id']
        type = request.form['type']
        amount = float(request.form['amount'])
        reason = request.form['reason']
        effective_date = datetime.strptime(request.form['effective_date'], '%Y-%m-%d').date()
        
        user = db_session.query(User).get(user_id)
        
        reward_penalty = RewardPenalty(
            user_id=user_id,
            department_id=user.department_id,
            type=type,
            amount=amount,
            reason=reason,
            effective_date=effective_date,
            created_by=current_user.id
        )
        
        db_session.add(reward_penalty)
        db_session.commit()
        
        # Notify employee
        type_arabic = 'مكافأة' if type == 'reward' else 'خصم'
        create_notification(
            user_id,
            f'{type_arabic} جديد',
            f'تم إضافة {type_arabic} لك بقيمة {amount} جنيه',
            'reward_penalty',
            action_url=url_for('user_rewards_penalties')
        )
        
        flash(f'تم إضافة {type_arabic} بنجاح')
        return redirect(url_for('manager_rewards_penalties'))
    
    department_employees = db_session.query(User).filter(
        User.department_id.in_(department_ids)
    ).all()
    
    rewards_penalties = db_session.query(RewardPenalty).filter(
        RewardPenalty.department_id.in_(department_ids)
    ).order_by(RewardPenalty.created_at.desc()).all()
    
    for rp in rewards_penalties:
        rp.user = db_session.query(User).get(rp.user_id)
    
    # حساب الإحصائيات بشكل صحيح
    stats = {
        'pending_leaves': db_session.query(LeaveRequest).filter(
            LeaveRequest.department_id.in_(department_ids),
            LeaveRequest.status == 'pending'
        ).count(),
        'pending_permissions': db_session.query(PermissionRequest).filter(
            PermissionRequest.department_id.in_(department_ids),
            PermissionRequest.status == 'pending'
        ).count(),
        'pending_advances': db_session.query(AdvanceRequest).filter(
            AdvanceRequest.department_id.in_(department_ids),
            AdvanceRequest.status == 'pending'
        ).count(),
        'pending_advances': 0  # إضافة قيمة افتراضية لتجنب الخطأ
    }
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager/manager_rewards_penalties.html',
                         department_employees=department_employees,
                         rewards_penalties=rewards_penalties,
                         stats=stats,
                         notifications=notifications)

def parse_shift_info(shift_info_json):
    """تحليل معلومات الشيفتات من JSON"""
    if not shift_info_json:
        return []
    
    try:
        shift_data = json.loads(shift_info_json)
        return shift_data.get('shifts', [])
    except:
        return []

def get_shift_jobs(shift_info_json):
    """الحصول على الوظائف المحددة لكل شيفت"""
    if not shift_info_json:
        return {}
    
    try:
        shift_data = json.loads(shift_info_json)
        return shift_data.get('jobs', {})
    except:
        return {}



@app.route('/manager/approve_leave_partial/<int:request_id>', methods=['POST'])
@login_required
def approve_leave_partial(request_id):
    """موافقة جزئية على طلب إجازة (قبول شيفتات معينة)"""
    if not current_user.is_manager and not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    leave_request = db_session.query(LeaveRequest).get(request_id)
    if not leave_request:
        flash('طلب الإجازة غير موجود', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_request.department_id)
    if department.primary_manager_id != current_user.id and not current_user.is_admin:
        flash('غير مصرح لك بالموافقة على هذا الطلب', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # الحصول على الشيفتات المقبولة والمرفوضة
    approved_shifts_json = request.form.get('approved_shifts', '[]')
    rejected_shifts_json = request.form.get('rejected_shifts', '[]')
    approval_notes = request.form.get('approval_notes', '')
    
    try:
        approved_shifts = json.loads(approved_shifts_json)
        rejected_shifts = json.loads(rejected_shifts_json)
        
        if not approved_shifts:
            flash('لم يتم تحديد أي شيفتات للموافقة', 'error')
            return redirect(url_for('manager_leave_requests'))
        
        # حساب عدد الشيفتات المقبولة والمرفوضة
        approved_count = len(approved_shifts)
        rejected_count = len(rejected_shifts)
        
        # إذا تم رفض جميع الشيفتات، اعتبر الطلب مرفوضاً بالكامل
        if approved_count == 0:
            leave_request.status = 'rejected'
            leave_request.rejection_reason = f'تم رفض جميع الشيفتات. {approval_notes}'
        # إذا تم قبول جميع الشيفتات، اعتبر الطلب مقبولاً بالكامل
        elif rejected_count == 0:
            leave_request.status = 'approved'
        # إذا كان هناك قبول ورفض جزئي، اعتبر الطلب مقبولاً جزئياً
        else:
            leave_request.status = 'partially_approved'
        
        # حفظ بيانات القبول/الرفض الجزئي
        leave_request.partial_leave_data = json.dumps({
            'approved_shifts': approved_shifts,
            'rejected_shifts': rejected_shifts,
            'approval_notes': approval_notes,
            'approved_count': approved_count,
            'rejected_count': rejected_count
        })
        
        leave_request.approved_shifts = json.dumps(approved_shifts)
        leave_request.rejected_shifts = json.dumps(rejected_shifts)
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        
        # إذا كانت الإجازة من رصيد الإجازات، خصم فقط الشيفتات المقبولة
        if leave_request.leave_type == 'من رصيد الإجازات' and approved_count > 0:
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=leave_request.user_id
            ).first()
            
            if balance:
                # خصم فقط الشيفتات المقبولة
                balance.leave_balance -= approved_count
                balance.last_updated = datetime.now()
        
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'تمت معالجة طلب إجازتك جزئياً. '
        notification_message += f'تم قبول {approved_count} شيفت ورفض {rejected_count} شيفت.'
        if approval_notes:
            notification_message += f' ملاحظات: {approval_notes}'
        
        create_notification(
            leave_request.user_id,
            'معالجة جزئية لطلب الإجازة',
            notification_message,
            'leave_partially_processed',
            related_id=leave_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        flash(f'تمت الموافقة على {approved_count} شيفت ورفض {rejected_count} شيفت بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء معالجة الطلب: {str(e)}', 'error')
    
    return redirect(url_for('manager_leave_requests'))



@app.route('/manager/reject_leave_partial/<int:request_id>', methods=['POST'])
@login_required
def reject_leave_partial(request_id):
    """رفض جزئي لطلب إجازة مع زيادة رصيد الموظف إذا كانت من رصيد الإجازات"""
    if not current_user.is_manager and not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    leave_request = db_session.query(LeaveRequest).get(request_id)
    if not leave_request:
        flash('طلب الإجازة غير موجود', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.get(Department, leave_request.department_id)
    if department.primary_manager_id != current_user.id and not current_user.is_admin:
        flash('غير مصرح لك برفض هذا الطلب', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من التاريخ (لا يمكن الرفض إلا في نفس الفترة الزمنية)
    leave_date = leave_request.leave_date if hasattr(leave_request, 'leave_date') and leave_request.leave_date else leave_request.start_date
    
    if not can_reject_leave_by_date(leave_date):
        # الحصول على معلومات الفترة الزمنية للعرض
        period_start, period_end = get_rejection_period_dates(leave_date)
        flash(
            f'لا يمكن رفض الإجازة إلا خلال الفترة من {period_start.strftime("%Y-%m-%d")} '
            f'إلى {period_end.strftime("%Y-%m-%d")}',
            'error'
        )
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_request.department_id)
    if department.primary_manager_id != current_user.id and not current_user.is_admin:
        flash('غير مصرح لك برفض هذا الطلب', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من التاريخ (لا يمكن الرفض بعد أو في يوم 26)
    if not can_reject_leave_by_date(leave_request.leave_date or leave_request.start_date):
        flash('لا يمكن رفض الإجازة بعد أو في يوم 26 من الشهر الحالي', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # الحصول على بيانات الرفض
    rejected_shifts_json = request.form.get('rejected_shifts', '[]')
    approved_shifts_json = request.form.get('approved_shifts', '[]')
    rejection_reason = request.form.get('rejection_reason', '').strip()
    
    if not rejection_reason:
        flash('يرجى كتابة سبب الرفض', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    try:
        rejected_shifts = json.loads(rejected_shifts_json)
        approved_shifts = json.loads(approved_shifts_json)
        
        if not rejected_shifts:
            flash('لم يتم تحديد أي شيفتات للرفض', 'error')
            return redirect(url_for('manager_leave_requests'))
        
        rejected_count = len(rejected_shifts)
        approved_count = len(approved_shifts)
        
        # تحديد حالة الطلب النهائية
        if approved_count == 0:
            leave_request.status = 'rejected'
        elif rejected_count == 0:
            leave_request.status = 'approved'
        else:
            leave_request.status = 'partially_rejected'
        
        # حفظ بيانات الرفض الجزئي
        leave_request.partial_leave_data = json.dumps({
            'rejected_shifts': rejected_shifts,
            'approved_shifts': approved_shifts,
            'rejection_reason': rejection_reason,
            'rejected_count': rejected_count,
            'approved_count': approved_count
        })
        
        leave_request.rejected_shifts = json.dumps(rejected_shifts)
        leave_request.approved_shifts = json.dumps(approved_shifts)
        leave_request.rejection_reason = rejection_reason
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        
        # زيادة رصيد الموظف إذا كانت الإجازة من رصيد الإجازات
        balance_increased = False
        new_balance = None
        
        if leave_request.leave_type == 'من رصيد الإجازات' and rejected_count > 0:
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=leave_request.user_id
            ).first()
            
            if balance:
                old_balance = balance.leave_balance
                balance.leave_balance += rejected_count
                balance.last_updated = datetime.now()
                new_balance = balance.leave_balance
                balance_increased = True
        
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'تم رفض {rejected_count} شيفت من طلب إجازتك. السبب: {rejection_reason}'
        if balance_increased:
            notification_message += f'\nتم زيادة رصيدك بمقدار {rejected_count} يوم (الرصيد الجديد: {new_balance} يوم)'
        
        create_notification(
            leave_request.user_id,
            'رفض جزئي لطلب الإجازة',
            notification_message,
            'leave_partially_rejected',
            related_id=leave_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        success_message = f'تم رفض {rejected_count} شيفت بنجاح'
        if balance_increased:
            success_message += f' وتم زيادة رصيد الموظف بمقدار {rejected_count} يوم'
        
        flash(success_message, 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء رفض الطلب: {str(e)}', 'error')
    
    return redirect(url_for('manager_leave_requests'))



def update_employee_balance_on_rejection(leave_request, rejected_shifts_count):
    """زيادة رصيد الموظف عند رفض إجازة من رصيد الإجازات"""
    if leave_request.leave_type == 'من رصيد الإجازات' and rejected_shifts_count > 0:
        balance = db_session.query(EmployeeBalance).filter_by(
            user_id=leave_request.user_id
        ).first()
        
        if balance:
            # زيادة الرصيد بعدد الشيفتات المرفوضة
            balance.leave_balance += rejected_shifts_count
            balance.last_updated = datetime.now()
            db_session.commit()
            
            return True, balance.leave_balance
    
    return False, None


@app.route('/manager/approve_leave/<int:request_id>')
@login_required
def approve_leave(request_id):
    """موافقة على طلب إجازة فردي"""
    if not current_user.is_manager and not current_user.is_admin:
        flash('غير مصرح لك بالموافقة على هذا الطلب', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    leave_request = db_session.get(LeaveRequest, request_id)
    
    if not leave_request:
        flash('طلب الإجازة غير موجود', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_request.department_id)
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    if department.primary_manager_id != current_user.id and not current_user.is_admin:
        flash('غير مصرح لك بالموافقة على هذا الطلب', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    if leave_request.status != 'pending':
        flash('تم معالجة هذا الطلب مسبقاً', 'warning')
        return redirect(url_for('manager_leave_requests'))
    
    try:
        # موافقة على الشيفت الفردي
        leave_request.status = 'approved'
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        
        # إذا كانت الإجازة من رصيد الإجازات، خصم من الرصيد
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=leave_request.user_id
            ).first()
            if balance:
                balance.leave_balance -= 1  # خصم يوم واحد لكل شيفت
                balance.last_updated = datetime.now()
        
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'تم الموافقة على شيفت إجازتك ({leave_request.shift_name})'
        if leave_request.shift_job:
            notification_message += f' للوظيفة {leave_request.shift_job}'
        
        create_notification(
            leave_request.user_id,
            'موافقة على طلب الإجازة',
            notification_message,
            'leave_approved',
            related_id=leave_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        flash('تم الموافقة على الشيفت بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء الموافقة على الطلب: {str(e)}', 'error')
    
    return redirect(url_for('manager_leave_requests'))




@app.context_processor
def utility_processor():
    def can_reject_leave(leave_request):
        """دالة مساعدة للقالب للتحقق من إمكانية رفض الإجازة"""
        leave_date = getattr(leave_request, 'leave_date', None) or leave_request.start_date
        return can_reject_leave_by_date(leave_date)
    
    def get_month_name_arabic(month):
        """إرجاع اسم الشهر بالعربية"""
        months = {
            1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
            5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
            9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
        }
        return months.get(month, '')
    
    return dict(
        can_reject_leave=can_reject_leave,
        get_month_name_arabic=get_month_name_arabic
    )


@app.route('/api/reject_leave/<int:request_id>', methods=['POST'])
@login_required
def api_reject_leave(request_id):
    """API لرفض طلب إجازة (للاستخدام مع AJAX)"""
    if not current_user.is_manager and not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    leave_request = db_session.query(LeaveRequest).get(request_id)
    if not leave_request:
        return jsonify({'success': False, 'message': 'طلب الإجازة غير موجود'})
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_request.department_id)
    if department.manager_id != current_user.id and not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    # التحقق من أن الطلب معلق
    if leave_request.status != 'pending':
        return jsonify({'success': False, 'message': 'تم معالجة هذا الطلب مسبقاً'})
    
    # الحصول على البيانات من طلب AJAX
    data = request.get_json()
    if not data or 'rejection_reason' not in data:
        return jsonify({'success': False, 'message': 'بيانات غير كاملة'})
    
    rejection_reason = data['rejection_reason'].strip()
    if len(rejection_reason) < 10:
        return jsonify({'success': False, 'message': 'سبب الرفض يجب أن يكون على الأقل 10 أحرف'})
    
    try:
        # تحديث حالة الطلب
        leave_request.status = 'rejected'
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        leave_request.rejection_reason = rejection_reason
        
        # زيادة الرصيد إذا كانت الإجازة من الرصيد
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=leave_request.user_id
            ).first()
            
            if balance:
                # زيادة الرصيد
                balance.leave_balance += leave_request.total_days
                balance.last_updated = datetime.now()
                
                # يمكنك إضافة سجل للعملية
                print(f"تم زيادة رصيد الموظف {leave_request.user_id} بمقدار {leave_request.total_days} يوم")
        
        db_session.commit()
        
        # إرسال إشعار
        create_notification(
            leave_request.user_id,
            'تم رفض طلب الإجازة',
            f'تم رفض طلب إجازتك. السبب: {rejection_reason[:50]}...',
            'leave_rejected',
            related_id=leave_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        return jsonify({
            'success': True,
            'message': f'تم رفض طلب الإجازة بنجاح{" وتم زيادة رصيد الموظف" if leave_request.leave_type == "من رصيد الإجازات" else ""}',
            'new_balance': balance.leave_balance if balance and leave_request.leave_type == 'من رصيد الإجازات' else None
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})
    


@app.route('/manager/approve_permission/<int:request_id>')
@login_required
def manager_approve_permission(request_id):
    """موافقة المدير على طلب إذن مع معالجة الأنواع الجديدة"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    permission_request = db_session.query(PermissionRequest).get(request_id)
    
    # Check if manager manages this department
    department = db_session.query(Department).get(permission_request.department_id)
    if department.primary_manager_id != current_user.id and not current_user.is_admin:
        flash('غير مصرح لك بالموافقة على هذا الطلب')
        return redirect(url_for('manager_permission_requests'))
    
    if permission_request.status == 'pending':
        permission_request.status = 'approved'
        permission_request.approved_by = current_user.id
        permission_request.approved_at = datetime.now()
        
        # معالجة خاصة لأنواع الإذن الجديدة
        extra_data = {}
        if permission_request.extra_data:
            try:
                extra_data = json.loads(permission_request.extra_data)
            except:
                pass
        
        # إرسال إشعار خاص للموظف بناءً على نوع الإذن
        notification_message = 'تم الموافقة على طلب الإذن الخاص بك'
        
        if permission_request.permission_type == 'اذن طلب ساعات اضافي':
            hours = extra_data.get('hours', 0)
            shifts = extra_data.get('shifts', [])
            notification_message = f'تم الموافقة على طلب ساعات إضافية لمدة {hours} ساعات للشيفتات: {", ".join(shifts)}'
        
        elif permission_request.permission_type == 'اذن تبديل وردية':
            shifts = extra_data.get('shifts', [])
            notification_message = f'تم الموافقة على طلب تبديل وردية للشيفتات: {", ".join(shifts)}'
            
            # يمكنك هنا إضافة منطق إضافي لتحديث جداول العمل
        
        # Notify employee
        create_notification(
            permission_request.user_id,
            'موافقة على طلب الإذن',
            notification_message,
            'permission_approved',
            action_url=url_for('user_leave_requests')
        )
        
        db_session.commit()
        flash('تم الموافقة على طلب الإذن')
    
    return redirect(url_for('manager_permission_requests'))


@app.template_filter('parse_permission_extra_data')
def parse_permission_extra_data_filter(extra_data_json):
    """فلتر لعرض البيانات الإضافية لطلب الإذن"""
    if not extra_data_json:
        return {}
    try:
        return json.loads(extra_data_json)
    except:
        return {}


@app.context_processor
def utility_processor():
    def parse_permission_extra_data(extra_data_json):
        if not extra_data_json:
            return {}
        try:
            return json.loads(extra_data_json)
        except:
            return {}
    
    return dict(
        parse_permission_extra_data=parse_permission_extra_data,
        get_approver_name=get_approver_name
    )

@app.route('/manager/reject_permission/<int:request_id>', methods=['POST'])
@login_required
def reject_permission(request_id):
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    permission_request = db_session.query(PermissionRequest).get(request_id)
    
    # Check if manager manages this department
    department = db_session.query(Department).get(permission_request.department_id)
    if department.primary_manager_id != current_user.id and not current_user.is_admin:  # CHANGED HERE
        flash('غير مصرح لك برفض هذا الطلب')
        return redirect(url_for('manager_permission_requests'))
    
    if permission_request.status == 'pending':
        permission_request.status = 'rejected'
        permission_request.approved_by = current_user.id
        permission_request.approved_at = datetime.now()
        permission_request.rejection_reason = request.form.get('rejection_reason')
        
        # Notify employee
        create_notification(
            permission_request.user_id,
            'رفض طلب الإذن',
            'تم رفض طلب الإذن الخاص بك',
            'permission_rejected',
            action_url=url_for('user_leave_requests')
        )
        
        db_session.commit()
        flash('تم رفض طلب الإذن')
    
    return redirect(url_for('manager_permission_requests'))

@app.route('/manager/approve_advance/<int:request_id>')
@login_required
def approve_advance(request_id):
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    advance_request = db_session.query(AdvanceRequest).get(request_id)
    
    # Check if manager manages this department
    department = db_session.query(Department).get(advance_request.department_id)
    if department.manager_id != current_user.id and not current_user.is_admin:
        flash('غير مصرح لك بالموافقة على هذا الطلب')
        return redirect(url_for('manager_advance_requests'))
    
    if advance_request.status == 'pending':
        advance_request.status = 'approved'
        advance_request.approved_by = current_user.id
        advance_request.approved_at = datetime.now()
        
        # Update balance
        balance = db_session.query(EmployeeBalance).filter_by(user_id=advance_request.user_id).first()
        if balance:
            balance.advance_balance += advance_request.amount
            balance.last_updated = datetime.now()
        
        # Notify employee
        create_notification(
            advance_request.user_id,
            'موافقة على طلب السلفة',
            f'تم الموافقة على طلب سلفتك بقيمة {advance_request.amount} جنيه',
            'advance_approved',
            action_url=url_for('user_advance_requests')
        )
        
        db_session.commit()
        flash('تم الموافقة على طلب السلفة')
    
    return redirect(url_for('manager_advance_requests'))

@app.route('/manager/reject_advance/<int:request_id>', methods=['POST'])
@login_required
def reject_advance(request_id):
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    advance_request = db_session.query(AdvanceRequest).get(request_id)
    
    # Check if manager manages this department
    department = db_session.query(Department).get(advance_request.department_id)
    if department.manager_id != current_user.id and not current_user.is_admin:
        flash('غير مصرح لك برفض هذا الطلب')
        return redirect(url_for('manager_advance_requests'))
    
    if advance_request.status == 'pending':
        advance_request.status = 'rejected'
        advance_request.approved_by = current_user.id
        advance_request.approved_at = datetime.now()
        advance_request.rejection_reason = request.form.get('rejection_reason')
        
        # Notify employee
        create_notification(
            advance_request.user_id,
            'رفض طلب السلفة',
            'تم رفض طلب السلفة الخاص بك',
            'advance_rejected',
            action_url=url_for('user_advance_requests')
        )
        
        db_session.commit()
        flash('تم رفض طلب السلفة')
    
    return redirect(url_for('manager_advance_requests'))

# ======== Notification Routes ========

@app.route('/notifications')
@login_required
def notifications():
    notifications_list = db_session.query(Notification).filter_by(
        user_id=current_user.id
    ).order_by(Notification.created_at.desc()).all()
    
    # Get only unread notifications
    unread_notifications = [n for n in notifications_list if not n.is_read]
    unread_count = len(unread_notifications)
    
    return render_template('notifications.html', 
                         notifications=notifications_list,
                         unread_notifications=unread_notifications,
                         unread_count=unread_count)

@app.route('/mark_notification_read/<int:notification_id>')
@login_required
def mark_notification_read(notification_id):
    notification = db_session.query(Notification).get(notification_id)
    if notification and notification.user_id == current_user.id:
        notification.is_read = True
        db_session.commit()
    
    return redirect(request.referrer or url_for('notifications'))

@app.route('/mark_all_read')
@login_required
def mark_all_read():
    notifications_list = db_session.query(Notification).filter_by(
        user_id=current_user.id, is_read=False
    ).all()
    
    for notification in notifications_list:
        notification.is_read = True
    
    db_session.commit()
    flash('تم تعيين جميع الإشعارات كمقروءة')
    return redirect(url_for('notifications'))

# ======== API Routes ========

@app.route('/api/employee_progress')
@login_required
def api_employee_progress():
    if current_user.is_admin:
        return jsonify({'success': False})
    
    employee_data = db_session.query(EmployeeData).filter_by(user_id=current_user.id).first()
    
    if employee_data:
        completion_percentage = employee_data.calculate_completion()
        missing_fields = employee_data.get_missing_fields()
        db_session.commit()
    else:
        completion_percentage = 0
        missing_fields = []
    
    return jsonify({
        'completion_percentage': completion_percentage,
        'missing_fields': missing_fields
    })


@app.route('/api/department_employeess')
@login_required
def api_department_employeess():
    """الحصول على موظفي القسم الحالي"""
    try:
        dept_id = request.args.get('dept_id')
        if not dept_id:
            # إذا لم يتم تمرير dept_id، استخدم قسم المستخدم الحالي
            dept_id = current_user.department_id
        
        if not dept_id:
            return jsonify({
                'success': False,
                'message': 'لا يوجد قسم محدد'
            })
        
        # جلب جميع موظفي القسم (غير المسؤولين)
        employees = db_session.query(User).filter(
            User.department_id == dept_id,
            User.is_admin == False,
            User.is_active == True
        ).all()
        
        # تحضير البيانات
        employees_data = []
        for emp in employees:
            # الحصول على بيانات الموظف
            employee_data = db_session.query(EmployeeData).filter_by(user_id=emp.id).first()
            
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'username': emp.username,
                'position': employee_data.job_title if employee_data else 'موظف',
                'full_name': employee_data.arabic_name if employee_data else emp.name
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        print(f"خطأ في جلب موظفي القسم: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'خطأ في جلب البيانات: {str(e)}'
        })

@app.route('/department/employees')
@login_required
def department_employees():
    """عرض جميع موظفي القسم"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    # الحصول على معرف القسم من المعاملات
    dept_id = request.args.get('dept_id', current_user.department_id)
    
    # جلب القسم
    department = db_session.query(Department).get(dept_id)
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('manager_dashboard' if current_user.is_manager else 'admin_dashboard'))
    
    # جلب موظفي القسم
    employees = db_session.query(User).filter(
        User.department_id == dept_id,
        User.is_admin == False,
        User.is_active == True
    ).order_by(User.name).all()
    
    # جلب بيانات الموظفين الإضافية
    for emp in employees:
        emp.employee_data = db_session.query(EmployeeData).filter_by(user_id=emp.id).first()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('department_employees.html',
                         department=department,
                         employees=employees,
                         notifications=notifications)


@app.route('/api/department_employees/<int:department_id>')
@login_required
def api_department_employees(department_id):
    if not current_user.is_manager and not current_user.is_admin:
        return jsonify({'success': False})
    
    employees = db_session.query(User).filter_by(department_id=department_id).all()
    employees_data = []
    
    for employee in employees:
        employees_data.append({
            'id': employee.id,
            'name': employee.name,
            'username': employee.username
        })
    
    return jsonify({'employees': employees_data})



# ======== Admin User Data View ========
@app.route('/admin/view_user/<int:user_id>')
@login_required
def view_user_data(user_id):
    if not current_user.is_admin and not current_user.is_manager:
        return redirect(url_for('user_dashboard'))
    
    user = db_session.query(User).get(user_id)
    employee_data = db_session.query(EmployeeData).filter_by(user_id=user_id).first()
    
    if employee_data:
        # Use the method from EmployeeData model if it exists
        missing_fields = employee_data.get_missing_fields()
    else:
        missing_fields = []
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin_view_user.html', 
                         user=user, 
                         employee_data=employee_data,
                         missing_fields=missing_fields,
                         notifications=notifications)

@app.route('/admin/export_user/<int:user_id>')
@login_required
def export_user_data(user_id):
    if not current_user.is_admin and not current_user.is_manager:
        return redirect(url_for('user_dashboard'))
    
    filepath = export_to_excel(user_id)
    if filepath:
        return send_file(filepath, as_attachment=True)
    else:
        flash('لا توجد بيانات للتصدير')
        return redirect(url_for('admin_users'))


@login_manager.user_loader
def load_user(user_id):
    try:
        return db_session.query(User).get(int(user_id))
    except Exception as e:
        print(f"Error loading user {user_id}: {e}")
        return None

@app.route('/admin/enable_auto_generation')
@login_required
def admin_enable_auto_generation():
    """تفعيل التوليد التلقائي لجميع الأقسام"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        # تفعيل التوليد التلقائي لجميع الأقسام
        departments = db_session.query(Department).all()
        enabled_count = 0
        
        for department in departments:
            if not department.auto_generate_schedule:
                department.auto_generate_schedule = True
                enabled_count += 1
                print(f"✅ تم تفعيل التوليد التلقائي للقسم: {department.name}")
        
        db_session.commit()
        
        flash(f'تم تفعيل التوليد التلقائي لـ {enabled_count} قسم', 'success')
        
        # إنشاء الجداول تلقائياً بعد التفعيل
        auto_generate_weekly_schedules()
        
    except Exception as e:
        db_session.rollback()
        flash(f'خطأ في تفعيل التوليد التلقائي: {str(e)}', 'error')
    
    return redirect(url_for('admin_schedules'))


def create_weekly_schedule_from_structure(department_id, start_date):
    """
    إنشاء جدول أسبوعي جديد بناءً على هيكل القسم
    """
    department = db_session.query(Department).get(department_id)
    if not department or not department.schedule_structure:
        return None
    
    try:
        structure_data = json.loads(department.schedule_structure)
        schedule_structure = structure_data.get('structure', [])
        
        # إنشاء نسخة جديدة من الهيكل مع تحديث التواريخ إذا لزم الأمر
        new_schedule = []
        for item in schedule_structure:
            new_item = item.copy()
            # هنا يمكنك إضافة منطق لتحديث التواريخ بناءً على start_date
            # إذا كان الهيكل يحتوي على أيام فقط وليس تواريخ
            new_schedule.append(new_item)
        
        return new_schedule
    
    except Exception as e:
        print(f"خطأ في إنشاء الجدول من الهيكل: {str(e)}")
        return None



# Add this to your Flask app
@app.template_filter('safe_tojson')
def safe_tojson(value):
    if value is None or isinstance(value, Undefined):
        return '[]'
    return json.dumps(value)


def create_default_schedule_days(week_start_date, department_name):
    """إنشاء أيام الجدول الافتراضية"""
    days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
    schedule_days = []
    
    for i, day_name in enumerate(days_of_week):
        current_date = week_start_date + timedelta(days=i)
        schedule_days.append({
            'day': day_name,
            'date': current_date.strftime('%Y-%m-%d'),
            'department': department_name,
            'morning_shift': 'موظف',
            'evening_shift': 'موظف',
            'night_shift': 'موظف',
            'job': 'موظف'
        })
    
    return schedule_days


@app.route('/admin/generate_test_schedules')
@login_required
def generate_test_schedules():
    """إنشاء جداول اختبارية (للتطوير فقط)"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    try:
        # إنشاء جدول للأسبوع الحالي
        today = date.today()
        days_since_saturday = (today.weekday() - 5) % 7
        week_start_date = today - timedelta(days=days_since_saturday)
        
        generated_count = auto_generate_weekly_schedules()
        
        flash(f'تم إنشاء {generated_count} جدول أسبوعي تلقائياً')
        
    except Exception as e:
        flash(f'خطأ في إنشاء الجداول: {str(e)}')
    
    return redirect(url_for('admin_schedules'))

def get_current_week_schedule(user):
    """الحصول على جدول الأسبوع الحالي للمستخدم"""
    today = date.today()
    week_start_date = today - timedelta(days=today.weekday() + 2)  # السبت الماضي
    
    if user.is_admin:
        # للمدير العام: جميع الجداول
        return db_session.query(WeeklySchedule).filter_by(
            week_start_date=week_start_date
        ).all()
    elif user.is_manager:
        # لمدير القسم: جداول قسمه
        department = db_session.query(Department).filter_by(manager_id=user.id).first()
        if department:
            return db_session.query(WeeklySchedule).filter_by(
                department_id=department.id,
                week_start_date=week_start_date
            ).first()
    else:
        # للموظف العادي: جدول قسمه المعتمد
        return db_session.query(WeeklySchedule).filter_by(
            department_id=user.department_id,
            week_start_date=week_start_date,
            is_approved=True
        ).first()
    
    return None

@app.route('/user/schedule')
@login_required
def user_schedule():
    if current_user.is_admin or current_user.is_manager:
        return redirect(url_for('admin_dashboard' if current_user.is_admin else 'manager_dashboard'))
    
    # Calculate current financial month (26th to 25th)
    today = date.today()
    
    # Determine current financial month
    if today.day >= 26:
        # Financial month starts from 26th of current month to 25th of next month
        financial_month_start = date(today.year, today.month, 26)
        if today.month == 12:
            financial_month_end = date(today.year + 1, 1, 25)
        else:
            financial_month_end = date(today.year, today.month + 1, 25)
    else:
        # Financial month starts from 26th of previous month to 25th of current month
        if today.month == 1:
            financial_month_start = date(today.year - 1, 12, 26)
            financial_month_end = date(today.year, 1, 25)
        else:
            financial_month_start = date(today.year, today.month - 1, 26)
            financial_month_end = date(today.year, today.month, 25)
    
    # Get all approved schedules for the user's department in the current financial month
    schedules = db_session.query(WeeklySchedule).filter(
        WeeklySchedule.department_id == current_user.department_id,
        WeeklySchedule.week_start_date >= financial_month_start,
        WeeklySchedule.week_end_date <= financial_month_end
    ).order_by(WeeklySchedule.week_start_date).all()
    
    # Prepare schedule data for template
    schedule_data_list = []
    for schedule in schedules:
        schedule_data = {}
        if schedule.schedule_data:
            try:
                schedule_data = json.loads(schedule.schedule_data)
            except:
                schedule_data = {}
        
        schedule_data_list.append({
            'schedule': schedule,
            'data': schedule_data,
            'is_current_week': (
                schedule.week_start_date <= today <= schedule.week_end_date
            )
        })
    
    # Get department info
    department = db_session.query(Department).get(current_user.department_id)
    
    # Generate month name for display
    month_name = get_month_name_arabic(financial_month_start.month)
    year = financial_month_start.year
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('user_schedule.html',
                         schedule_data_list=schedule_data_list,
                         financial_month_start=financial_month_start,
                         financial_month_end=financial_month_end,
                         month_name=month_name,
                         year=year,
                         department=department,
                         today=today,
                         notifications=notifications)

@app.route('/admin/regenerate_schedule/<int:dept_id>')
@login_required
def admin_regenerate_schedule(dept_id):
    """إنشاء الجداول المستقبلية لقسم معين دون حذف البيانات القديمة"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        department = db_session.query(Department).get(dept_id)
        if not department:
            flash('القسم غير موجود', 'error')
            return redirect(url_for('admin_schedules'))
        
        generated_count = generate_future_schedules(dept_id)
        
        if generated_count > 0:
            flash(f'تم إنشاء {generated_count} جدول مستقبلي جديد للقسم {department.name}')
        else:
            flash(f'جميع الجداول المستقبلية موجودة مسبقاً للقسم {department.name}', 'info')
            
    except Exception as e:
        flash(f'خطأ في إنشاء الجداول: {str(e)}', 'error')
    
    return redirect(url_for('admin_schedules'))

@app.route('/admin/regenerate_all_future')
@login_required
def admin_regenerate_all_future():
    """إنشاء الجداول المستقبلية لجميع الأقسام"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        departments = db_session.query(Department).filter_by(auto_generate_schedule=True).all()
        total_generated = 0
        
        for department in departments:
            generated_count = generate_future_schedules(department.id)
            total_generated += generated_count
            if generated_count > 0:
                print(f"تم إنشاء {generated_count} جدول للقسم {department.name}")
        
        flash(f'تم إنشاء {total_generated} جدول مستقبلي جديد لجميع الأقسام')
        
    except Exception as e:
        flash(f'خطأ في إنشاء الجداول: {str(e)}', 'error')
    
    return redirect(url_for('admin_schedules'))

@app.route('/admin/schedule_history/<int:dept_id>')
@login_required
def admin_schedule_history(dept_id):
    """عرض التاريخ الكامل للجداول لقسم معين"""
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    department = db_session.query(Department).get(dept_id)
    if not department:
        flash('القسم غير موجود', 'error')
        return redirect(url_for('admin_schedules'))
    
    # الحصول على جميع الجداول للقسم مرتبة من الأحدث إلى الأقدم
    schedules = db_session.query(WeeklySchedule).filter_by(
        department_id=dept_id
    ).order_by(WeeklySchedule.week_start_date.desc()).all()
    
    # تجميع الجداول حسب الشهر والسنة
    schedules_by_period = {}
    for schedule in schedules:
        period_key = schedule.week_start_date.strftime('%Y-%m')  # السنة-الشهر
        if period_key not in schedules_by_period:
            schedules_by_period[period_key] = {
                'display_name': schedule.week_start_date.strftime('%B %Y'),  # اسم الشهر والسنة
                'schedules': []
            }
        schedules_by_period[period_key]['schedules'].append(schedule)
    
    # إحصائيات
    stats = {
        'total_schedules': len(schedules),
        'approved_schedules': len([s for s in schedules if s.is_approved]),
        'pending_schedules': len([s for s in schedules if s.status == 'pending']),
        'draft_schedules': len([s for s in schedules if s.status == 'draft']),
        'locked_schedules': len([s for s in schedules if s.is_locked]),
        'oldest_schedule': min(schedules, key=lambda x: x.week_start_date).week_start_date if schedules else None,
        'newest_schedule': max(schedules, key=lambda x: x.week_start_date).week_start_date if schedules else None
    }
    
    # حساب التواريخ الحالية
    today = date.today()
    days_since_saturday = (today.weekday() - 5) % 7
    current_week_start = today - timedelta(days=days_since_saturday)
    current_week_end = current_week_start + timedelta(days=6)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_schedule_history.html',
                         department=department,
                         schedules_by_period=schedules_by_period,
                         stats=stats,
                         today=today,
                         current_week_start=current_week_start,
                         current_week_end=current_week_end,
                         notifications=notifications)

@app.before_request
def check_weekly_schedules():
    """التحقق من وجود الجداول الأسبوعية قبل كل طلب (بحدود معقولة)"""
    try:
        # تشغيل هذا فقط مرة في اليوم لتجنب التحميل الزائد
        if not hasattr(app, 'last_schedule_check'):
            app.last_schedule_check = date.today() - timedelta(days=1)
        
        if date.today() > app.last_schedule_check:
            auto_generate_weekly_schedules()
            app.last_schedule_check = date.today()
    except Exception as e:
        print(f"خطأ في التحقق من الجداول: {e}")
        

@app.context_processor
def inject_timedelta():
    return dict(timedelta=timedelta)




def handle_schedule_submission(request, schedule, department):
    """معالجة طلبات حفظ الجدول"""
    try:
        action = request.form.get('action', 'save')
        schedule_data_json = request.form.get('schedule_data', '{}')
        
        # التحقق من صحة البيانات
        if not schedule_data_json:
            flash('لم يتم استلام بيانات الجدول', 'error')
            return redirect(url_for('manager_edit_schedule', schedule_id=schedule.id))
        
        # تحليل بيانات JSON
        try:
            schedule_data = json.loads(schedule_data_json)
        except json.JSONDecodeError:
            flash('بيانات الجدول غير صالحة', 'error')
            return redirect(url_for('manager_edit_schedule', schedule_id=schedule.id))
        
        # تحديث حالة الجدول بناءً على الإجراء
        if action == 'submit':
            schedule.status = 'pending'
            schedule.submitted_at = datetime.utcnow()
            schedule.is_locked = True
            flash('تم إرسال الجدول للاعتماد بنجاح', 'success')
        else:  # save
            schedule.status = 'draft'
            flash('تم حفظ الجدول بنجاح', 'success')
        
        # حفظ بيانات الجدول
        schedule.schedule_data = json.dumps(schedule_data, ensure_ascii=False)
        schedule.updated_at = datetime.utcnow()
        
        db_session.commit()
        
        # إرسال إشعارات إذا كان الإرسال للاعتماد
        if action == 'submit':
            send_approval_notification(schedule, department)
        
        return redirect(url_for('manager_schedules'))
    
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء حفظ الجدول: {str(e)}', 'error')
        return redirect(url_for('manager_edit_schedule', schedule_id=schedule.id))



def send_approval_notification(schedule, department):
    """إرسال إشعارات عند تقديم الجدول للاعتماد"""
    try:
        # إرسال إشعار للمدير العام أو المسؤولين
        admin_users = db_session.query(User).filter(
            User.role.in_(['admin', 'general_manager']),
            User.is_active == True
        ).all()
        
        for admin in admin_users:
            notification = Notification(
                user_id=admin.id,
                title=f'جدول جديد يحتاج للاعتماد - {department.name}',
                message=f'تم تقديم جدول الأسبوع {schedule.week_start_date} من قسم {department.name} للاعتماد',
                type='schedule_approval',
                related_entity_type='schedule',
                related_entity_id=schedule.id
            )
            db_session.add(notification)
        
        db_session.commit()
        
    except Exception as e:
        print(f"Error sending approval notifications: {e}")
        db_session.rollback()


 
@app.before_request
def check_weekly_schedules():
    """التحقق من وجود الجداول الأسبوعية قبل كل طلب (بحدود معقولة)"""
    try:
        # تشغيل هذا فقط مرة في اليوم لتجنب التحميل الزائد
        if not hasattr(app, 'last_schedule_check'):
            app.last_schedule_check = date.today() - timedelta(days=1)
        
        if date.today() > app.last_schedule_check:
            auto_generate_weekly_schedules()
            app.last_schedule_check = date.today()
    except Exception as e:
        print(f"خطأ في التحقق من الجداول: {e}")


@app.route('/debug/schedule/<int:schedule_id>')
@login_required
def debug_schedule(schedule_id):
    """Route for debugging schedule data"""
    schedule = db_session.get(WeeklySchedule, schedule_id)
    if schedule:
        return jsonify({
            'schedule_id': schedule.id,
            'department_id': schedule.department_id,
            'has_schedule_data': bool(schedule.schedule_data),
            'schedule_data_length': len(schedule.schedule_data) if schedule.schedule_data else 0,
            'is_locked': schedule.is_locked,
            'is_approved': schedule.is_approved
        })
    return jsonify({'error': 'Schedule not found'})


def get_default_days_structure():
    """إرجاع هيكل الأيام الافتراضي"""
    return [
        {'name': 'saturday', 'title': 'السبت', 'type': 'select', 'options': ['صباحي', 'مسائي', 'ليلي', 'إجازة'], 'default': 'صباحي'},
        {'name': 'sunday', 'title': 'الأحد', 'type': 'select', 'options': ['صباحي', 'مسائي', 'ليلي', 'إجازة'], 'default': 'صباحي'},
        {'name': 'monday', 'title': 'الإثنين', 'type': 'select', 'options': ['صباحي', 'مسائي', 'ليلي', 'إجازة'], 'default': 'صباحي'},
        {'name': 'tuesday', 'title': 'الثلاثاء', 'type': 'select', 'options': ['صباحي', 'مسائي', 'ليلي', 'إجازة'], 'default': 'صباحي'},
        {'name': 'wednesday', 'title': 'الأربعاء', 'type': 'select', 'options': ['صباحي', 'مسائي', 'ليلي', 'إجازة'], 'default': 'صباحي'},
        {'name': 'thursday', 'title': 'الخميس', 'type': 'select', 'options': ['صباحي', 'مسائي', 'ليلي', 'إجازة'], 'default': 'صباحي'}
    ]

def create_default_schedule(employees, days_structure):
    """إنشاء جدول افتراضي"""
    schedule_data = {}
    for employee in employees:
        employee_schedule = {"employee": employee.name}
        for day in days_structure:
            employee_schedule[day['name']] = day.get('default', '')
        schedule_data[str(employee.id)] = employee_schedule
    return schedule_data

@app.route('/manager/schedule/submit/<int:schedule_id>')
@login_required
def manager_submit_schedule(schedule_id):
    """إرسال الجدول للاعتماد (منفصل عن حفظ التعديلات)"""
    if not current_user.is_manager:
        flash('غير مصرح لك بهذا الإجراء')
        return redirect(url_for('user_dashboard'))
    
    schedule = db_session.get(WeeklySchedule, schedule_id)
    if not schedule:
        flash('الجدول غير موجود')
        return redirect(url_for('manager_schedules'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(schedule.department_id)
    if not department or department.manager_id != current_user.id:
        flash('غير مصرح لك بهذا الإجراء')
        return redirect(url_for('manager_schedules'))
    
    try:
        # تحديث حالة الجدول
        schedule.status = 'pending'
        schedule.is_approved = False
        schedule.is_locked = False
        schedule.updated_at = datetime.now()
        
        db_session.commit()
        
        # إرسال إشعار للمدير العام
        admins = db_session.query(User).filter_by(is_admin=True).all()
        for admin in admins:
            create_notification(
                admin.id,
                'جدول جديد يحتاج اعتماد',
                f'تم إرسال جدول قسم {department.name} للاعتماد للفترة {schedule.week_start_date} إلى {schedule.week_end_date}',
                'schedule_pending',
                related_id=schedule.id,
                action_url=url_for('admin_schedules')
            )
        
        flash('تم إرسال الجدول للاعتماد بنجاح')
        
    except Exception as e:
        flash(f'حدث خطأ أثناء إرسال الجدول: {str(e)}')
        db_session.rollback()
    
    return redirect(url_for('manager_schedules'))



# ======== Manager Routes ========

@app.route('/manager/dashboard')
@login_required
def manager_dashboard():
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('manager_dashboard'))
    
    managed_departments = db_session.query(Department).filter_by(primary_manager_id=current_user.id).all()
    department_ids = [dept.id for dept in managed_departments]
    
    # Get all users for department counts
    users = db_session.query(User).filter(User.department_id.in_(department_ids)).all()
    
    stats = {
        'pending_leaves': db_session.query(LeaveRequest).filter(
            LeaveRequest.department_id.in_(department_ids),
            LeaveRequest.status == 'pending'
        ).count(),
        'pending_permissions': db_session.query(PermissionRequest).filter(
            PermissionRequest.department_id.in_(department_ids),
            PermissionRequest.status == 'pending'
        ).count(),
        'pending_advances': db_session.query(AdvanceRequest).filter(
            AdvanceRequest.department_id.in_(department_ids),
            AdvanceRequest.status == 'pending'
        ).count(),
        'department_employees': db_session.query(User).filter(
            User.department_id.in_(department_ids),
            User.is_admin == False
        ).count()
    }
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager_dashboard.html',
                         managed_departments=managed_departments,
                         stats=stats,
                         users=users,
                         notifications=notifications)

@app.route('/manager/employees')
@login_required
def manager_employees():
    """إدارة موظفي القسم"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    managed_departments = db_session.query(Department).filter_by(primary_manager_id=current_user.id).all()
    department_ids = [dept.id for dept in managed_departments]
    
    employees = db_session.query(User).filter(
        User.department_id.in_(department_ids),
        User.is_admin == False
    ).all()
    
    # Load employee data
    for employee in employees:
        employee.employee_data = db_session.query(EmployeeData).filter_by(user_id=employee.id).first()
        if employee.employee_data:
            employee.employee_data.calculate_completion()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager/manager_employees.html',
                         employees=employees,
                         notifications=notifications)


@app.route('/admin/schedule_sync_dashboard')
@login_required
def admin_schedule_sync_dashboard():
    """لوحة تحكم مزامنة الجداول"""
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    from database import db_session
    from models import Department, WeeklySchedule
    
    try:
        # إحصائيات عامة
        total_departments = db_session.query(Department).count()
        departments_with_structure = db_session.query(Department).filter(
            Department.schedule_structure.isnot(None)
        ).count()
        
        total_schedules = db_session.query(WeeklySchedule).count()
        synced_schedules = db_session.query(WeeklySchedule).filter(
            WeeklySchedule.is_generated_from_structure == True
        ).count()
        
        # تحليل هياكل الجداول
        from schedules_sync import analyze_schedule_structure
        
        schedules = db_session.query(WeeklySchedule).limit(100).all()
        structure_types = {
            'new_format': 0,
            'old_format': 0,
            'list_format': 0,
            'empty': 0,
            'synced': 0,
            'not_synced': 0
        }
        
        for schedule in schedules:
            structure_type, _ = analyze_schedule_structure(schedule)
            structure_types[structure_type] = structure_types.get(structure_type, 0) + 1
            
            if hasattr(schedule, 'is_generated_from_structure') and schedule.is_generated_from_structure:
                structure_types['synced'] += 1
            else:
                structure_types['not_synced'] += 1
        
        # الأقسام التي تحتاج اهتماماً
        departments_needing_attention = []
        departments = db_session.query(Department).all()
        
        for dept in departments:
            # عدد الجداول غير المزامنة
            unsynced_schedules = db_session.query(WeeklySchedule).filter(
                WeeklySchedule.department_id == dept.id,
                or_(
                    WeeklySchedule.is_generated_from_structure == False,
                    WeeklySchedule.is_generated_from_structure == None
                )
            ).count()
            
            if unsynced_schedules > 0 and dept.schedule_structure:
                departments_needing_attention.append({
                    'department': dept,
                    'unsynced_count': unsynced_schedules,
                    'total_schedules': db_session.query(WeeklySchedule).filter_by(department_id=dept.id).count()
                })
        
        notifications = get_user_notifications(current_user.id)
        
        return render_template('admin/admin_schedule_sync_dashboard.html',
                             total_departments=total_departments,
                             departments_with_structure=departments_with_structure,
                             total_schedules=total_schedules,
                             synced_schedules=synced_schedules,
                             structure_types=structure_types,
                             departments_needing_attention=departments_needing_attention,
                             notifications=notifications)
        
    except Exception as e:
        flash(f'حدث خطأ في تحميل البيانات: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route('/admin/execute_full_sync', methods=['POST'])
@login_required
def admin_execute_full_sync():
    """تنفيذ مزامنة كاملة لجميع الجداول"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        from schedules_sync import force_sync_all_schedules
        
        synced_count = force_sync_all_schedules()
        
        return jsonify({
            'success': True,
            'message': f'تم مزامنة {synced_count} جدول مع هياكل الأقسام',
            'synced_count': synced_count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في المزامنة: {str(e)}'
        })
    

@app.route('/admin/create_default_structures', methods=['POST'])
@login_required
def admin_create_default_structures():
    """إنشاء هياكل افتراضية للأقسام التي ليس لديها هياكل"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        from database import db_session
        from models import Department
        
        departments_without_structure = db_session.query(Department).filter(
            Department.schedule_structure.is_(None)
        ).all()
        
        created_count = 0
        
        for department in departments_without_structure:
            # إنشاء هيكل افتراضي
            default_structure = create_default_schedule_structure_for_department(department)
            
            if default_structure:
                department.schedule_structure = json.dumps(default_structure, ensure_ascii=False)
                
                if hasattr(department, 'schedule_structure_version'):
                    department.schedule_structure_version = 1
                
                created_count += 1
                print(f"✓ تم إنشاء هيكل افتراضي للقسم {department.name}")
        
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء هياكل افتراضية لـ {created_count} قسم',
            'created_count': created_count
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في إنشاء الهياكل: {str(e)}'
        })

def create_default_schedule_structure_for_department(department):
    """إنشاء هيكل جدول افتراضي للقسم"""
    try:
        days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        
        schedule_structure = {
            'source': 'default_auto_created',
            'department': department.name,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'schedule': []
        }
        
        for day_name in days_of_week:
            day_entry = {
                'day': day_name,
                'morning_shift': 'موظف',
                'evening_shift': 'موظف',
                'night_shift': 'موظف',
                'job': 'موظف'
            }
            schedule_structure['schedule'].append(day_entry)
        
        return schedule_structure
        
    except Exception as e:
        print(f"Error creating default structure: {str(e)}")
        return None


@app.route('/admin/departments_without_structures')
@login_required
def admin_departments_without_structures():
    """عرض الأقسام التي ليس لديها هياكل"""
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        from database import db_session
        from models import Department
        
        departments_without_structure = db_session.query(Department).filter(
            Department.schedule_structure.is_(None)
        ).order_by(Department.name).all()
        
        notifications = get_user_notifications(current_user.id)
        
        return render_template('admin/admin_departments_without_structures.html',
                             departments=departments_without_structure,
                             notifications=notifications)
        
    except Exception as e:
        flash(f'حدث خطأ في تحميل البيانات: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


def normalize_schedule_with_cached_data(schedule, department, cached_employee_data):
    """
    تطبيع الجدول مع البيانات المخزنة مؤقتاً
    """
    try:
        if not schedule.schedule_data:
            return None
        
        # تحليل بيانات الجدول الحالي
        schedule_data = json.loads(schedule.schedule_data)
        
        # تحديد نوع الهيكل وتحويله إلى الهيكل الموحد
        normalized_structure = {
            'department': department.name,
            'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
            'week_end_date': schedule.week_end_date.strftime('%Y-%m-%d'),
            'source': 'normalized_from_mixed',
            'schedule': []
        }
        
        # أيام الأسبوع
        days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        
        for i, day_name in enumerate(days_of_week):
            current_date = schedule.week_start_date + timedelta(days=i)
            
            day_entry = {
                'day': day_name,
                'date': current_date.strftime('%Y-%m-%d'),
                'department': department.name,
                'morning_shift': '',
                'evening_shift': '',
                'night_shift': '',
                'job': 'موظف'
            }
            
            # استخراج البيانات من الهيكل القديم
            if isinstance(schedule_data, dict):
                # الهيكل القديم: {employee_id: {day: shift, ...}}
                morning_shifts = []
                evening_shifts = []
                night_shifts = []
                
                for emp_id, emp_schedule in schedule_data.items():
                    if isinstance(emp_schedule, dict):
                        # الحصول على اسم اليوم العربي
                        arabic_days = ['السبت', 'الأحد', 'الأثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
                        if i < len(arabic_days):
                            arabic_day = arabic_days[i]
                            shift = emp_schedule.get(arabic_day, '')
                            
                            if shift:
                                # الحصول على اسم الموظف من البيانات المخزنة مؤقتاً
                                emp_name = cached_employee_data.get(str(emp_id), f"مستخدم {emp_id}")
                                
                                if 'صباحي' in str(shift):
                                    morning_shifts.append(emp_name)
                                elif 'مسائي' in str(shift):
                                    evening_shifts.append(emp_name)
                                elif 'ليلي' in str(shift) or 'سهر' in str(shift):
                                    night_shifts.append(emp_name)
                
                if morning_shifts:
                    day_entry['morning_shift'] = ', '.join(morning_shifts)
                if evening_shifts:
                    day_entry['evening_shift'] = ', '.join(evening_shifts)
                if night_shifts:
                    day_entry['night_shift'] = ', '.join(night_shifts)
                    
            elif isinstance(schedule_data, list):
                # الهيكل القديم: [{day: ..., shift: ...}]
                for item in schedule_data:
                    if isinstance(item, dict):
                        item_day = item.get('day') or item.get('اليوم')
                        if item_day == day_name:
                            # استخراج الشيفتات
                            if item.get('morning_shift'):
                                day_entry['morning_shift'] = item['morning_shift']
                            if item.get('evening_shift'):
                                day_entry['evening_shift'] = item['evening_shift']
                            if item.get('night_shift'):
                                day_entry['night_shift'] = item['night_shift']
                            if item.get('job'):
                                day_entry['job'] = item['job']
                            break
            
            normalized_structure['schedule'].append(day_entry)
        
        return normalized_structure
        
    except Exception as e:
        print(f"خطأ في تطبيع الجدول {schedule.id}: {str(e)}")
        return None


def sync_all_schedules_with_department_structures():
    """
    مزامنة جميع الجداول مع هياكل الأقسام - نسخة مبسطة
    """
    try:
        print("=== بدء مزامنة جميع الجداول مع هياكل الأقسام ===")
        
        # الحصول على جميع الأقسام
        departments = db_session.query(Department).all()
        total_synced = 0
        
        for department in departments:
            if not department.schedule_structure:
                print(f"⚠️ القسم {department.name} ليس لديه هيكل")
                continue
            
            try:
                print(f"مزامنة جداول القسم: {department.name}")
                
                # الحصول على جميع جداول القسم
                schedules = db_session.query(WeeklySchedule).filter_by(
                    department_id=department.id
                ).all()
                
                department_synced = 0
                
                for schedule in schedules:
                    try:
                        # إذا كان الجدول معتمداً ومقفولاً، تخطيه
                        if schedule.is_approved and schedule.is_locked:
                            continue
                        
                        # إنشاء هيكل جديد من هيكل القسم
                        new_schedule_data = create_schedule_from_department_structure(
                            department, schedule.week_start_date
                        )
                        
                        if new_schedule_data:
                            # تحديث الجدول بالهيكل الجديد
                            schedule.schedule_data = json.dumps(new_schedule_data, ensure_ascii=False)
                            if hasattr(schedule, 'is_generated_from_structure'):
                                schedule.is_generated_from_structure = True
                            if hasattr(schedule, 'structure_version'):
                                schedule.structure_version = department.schedule_structure_version
                            department_synced += 1
                            total_synced += 1
                            
                            print(f"✓ تمت مزامنة الجدول {schedule.id}")
                        else:
                            print(f"⚠️ لم يتمكن من إنشاء جدول للجدول {schedule.id}")
                            
                    except Exception as e:
                        print(f"❌ خطأ في مزامنة الجدول {schedule.id}: {str(e)}")
                        continue
                
                print(f"تمت مزامنة {department_synced} جدول للقسم {department.name}")
                
            except Exception as e:
                print(f"❌ خطأ في مزامنة قسم {department.name}: {str(e)}")
                continue
        
        db_session.commit()
        print(f"=== تمت مزامنة {total_synced} جدول ===")
        return total_synced
        
    except Exception as e:
        db_session.rollback()
        print(f"❌ خطأ عام في المزامنة: {str(e)}")
        return 0

def create_unified_schedule_structure():
    """
    إنشاء هيكل موحد للجداول يشبه الجدول 54
    """
    unified_structure = {
        'department': 'اسم القسم',
        'week_start_date': 'تاريخ بداية الأسبوع',
        'week_end_date': 'تاريخ نهاية الأسبوع',
        'source': 'unified_structure',
        'columns': [
            'شيفت السهر',
            'الشيفت المسائي', 
            'الشيفت الصباحي',
            'الوظيفة',
            'القسم',
            'التاريخ',
            'اليوم'
        ],
        'schedule': []  # سيحتوي على بيانات الأيام
    }
    
    return unified_structure


def create_new_unified_schedule(department_id, week_start_date):
    """
    إنشاء جدول جديد بالهيكل الموحد
    """
    try:
        department = db_session.query(Department).get(department_id)
        if not department:
            print(f"❌ القسم غير موجود: {department_id}")
            return None
        
        week_end_date = week_start_date + timedelta(days=6)
        
        # أيام الأسبوع بالعربية
        days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        
        # إنشاء الهيكل الموحد
        unified_structure = {
            'department': department.name,
            'week_start_date': week_start_date.strftime('%Y-%m-%d'),
            'week_end_date': week_end_date.strftime('%Y-%m-%d'),
            'source': 'unified_structure_created',
            'columns': [
                'شيفت السهر',
                'الشيفت المسائي', 
                'الشيفت الصباحي',
                'الوظيفة',
                'القسم',
                'التاريخ',
                'اليوم'
            ],
            'schedule': []
        }
        
        # إضافة أيام الأسبوع
        for i, day_name in enumerate(days_of_week):
            current_date = week_start_date + timedelta(days=i)
            
            day_entry = {
                'day': day_name,
                'date': current_date.strftime('%Y-%m-%d'),
                'department': department.name,
                'morning_shift': 'موظف',
                'evening_shift': 'موظف',
                'night_shift': 'موظف',
                'job': 'موظف'
            }
            
            unified_structure['schedule'].append(day_entry)
        
        return unified_structure
        
    except Exception as e:
        print(f"❌ خطأ في إنشاء الجدول الموحد: {str(e)}")
        return None

def convert_dict_structure(old_data, schedule):
    """
    تحويل الهيكل القديم (قاموس) إلى الهيكل الموحد
    """
    try:
        department = db_session.query(Department).get(schedule.department_id)
        if not department:
            return create_new_unified_schedule(schedule.department_id, schedule.week_start_date)
        
        # استخدام الدالة التي أنشأناها
        return create_new_unified_schedule(department.id, schedule.week_start_date)
        
    except Exception as e:
        print(f"❌ خطأ في تحويل الهيكل القاموسي: {str(e)}")
        return create_new_unified_schedule(schedule.department_id, schedule.week_start_date)

def convert_list_structure(old_data, schedule):
    """
    تحويل الهيكل القديم (قائمة) إلى الهيكل الموحد
    """
    try:
        department = db_session.query(Department).get(schedule.department_id)
        if not department:
            return create_new_unified_schedule(schedule.department_id, schedule.week_start_date)
        
        week_end_date = schedule.week_start_date + timedelta(days=6)
        
        # إنشاء الهيكل الموحد
        unified_structure = {
            'department': department.name,
            'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
            'week_end_date': week_end_date.strftime('%Y-%m-%d'),
            'source': 'converted_from_list',
            'schedule': []
        }
        
        # أيام الأسبوع
        days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        
        for i, day_name in enumerate(days_of_week):
            current_date = schedule.week_start_date + timedelta(days=i)
            
            # البحث عن اليوم في البيانات القديمة
            day_data = None
            for item in old_data:
                if isinstance(item, dict):
                    item_day = item.get('day') or item.get('اليوم')
                    if item_day == day_name:
                        day_data = item
                        break
            
            if day_data:
                day_entry = {
                    'day': day_name,
                    'date': current_date.strftime('%Y-%m-%d'),
                    'department': department.name,
                    'morning_shift': day_data.get('morning_shift') or day_data.get('الشيفت الصباحي') or '',
                    'evening_shift': day_data.get('evening_shift') or day_data.get('الشيفت المسائي') or '',
                    'night_shift': day_data.get('night_shift') or day_data.get('شيفت السهر') or '',
                    'job': day_data.get('job') or day_data.get('الوظيفة') or 'موظف'
                }
            else:
                day_entry = {
                    'day': day_name,
                    'date': current_date.strftime('%Y-%m-%d'),
                    'department': department.name,
                    'morning_shift': 'موظف',
                    'evening_shift': 'موظف',
                    'night_shift': 'موظف',
                    'job': 'موظف'
                }
            
            unified_structure['schedule'].append(day_entry)
        
        return unified_structure
        
    except Exception as e:
        print(f"❌ خطأ في تحويل الهيكل القائمة: {str(e)}")
        return create_new_unified_schedule(schedule.department_id, schedule.week_start_date)



def convert_old_schedule_to_unified(schedule_id):
    """
    تحويل جدول قديم إلى الهيكل الموحد
    """
    schedule = db_session.get(WeeklySchedule, schedule_id)
    
    if not schedule:
        print(f"❌ الجدول غير موجود: {schedule_id}")
        return None
    
    if not schedule.schedule_data:
        # إنشاء جدول جديد بالهيكل الموحد
        return create_new_unified_schedule(schedule.department_id, schedule.week_start_date)
    
    try:
        old_data = json.loads(schedule.schedule_data)
        
        # تحديد نوع الهيكل القديم وتحويله
        if isinstance(old_data, dict):
            # الهيكل القديم: {employee_id: {day: shift}}
            return convert_dict_structure(old_data, schedule)
        elif isinstance(old_data, list):
            # الهيكل القديم: [{day: ..., shift: ...}]
            return convert_list_structure(old_data, schedule)
        else:
            # هيكل غير معروف، إنشاء هيكل جديد
            return create_new_unified_schedule(schedule.department_id, schedule.week_start_date)
            
    except Exception as e:
        print(f"❌ خطأ في تحويل الجدول {schedule_id}: {str(e)}")
        return create_new_unified_schedule(schedule.department_id, schedule.week_start_date)

@app.route('/admin/sync_all_schedules')
@login_required
def admin_sync_all_schedules():
    """واجهة مزامنة جميع الجداول"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    # تنفيذ المزامنة
    synced_count = sync_all_schedules_with_department_structures()
    
    flash(f'تم مزامنة {synced_count} جدول مع هياكل الأقسام', 'success')
    return redirect(url_for('admin_schedules'))

@app.route('/admin/generate_future_schedules', methods=['POST'])
@login_required
def admin_generate_future_schedules():
    """إنشاء جداول مستقبلية من هياكل الأقسام"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        from schedules_sync import create_future_schedules_from_structure
        
        created_count = create_future_schedules_from_structure()
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء {created_count} جدول مستقبلي',
            'created_count': created_count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في إنشاء الجداول: {str(e)}'
        })

@app.route('/admin/schedule_sync_report')
@login_required
def admin_schedule_sync_report():
    """تقرير مفصل عن مزامنة الجداول"""
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    from database import db_session
    from models import Department, WeeklySchedule
    from schedules_sync import analyze_schedule_structure
    
    try:
        # الحصول على جميع الأقسام مع هياكل
        departments = db_session.query(Department).filter(
            Department.schedule_structure.isnot(None)
        ).order_by(Department.name).all()
        
        department_reports = []
        
        for dept in departments:
            # الحصول على جميع جداول القسم
            schedules = db_session.query(WeeklySchedule).filter_by(
                department_id=dept.id
            ).order_by(WeeklySchedule.week_start_date.desc()).all()
            
            # تحليل كل جدول
            schedule_details = []
            synced_count = 0
            unsynced_count = 0
            
            for schedule in schedules:
                structure_type, _ = analyze_schedule_structure(schedule)
                is_synced = hasattr(schedule, 'is_generated_from_structure') and schedule.is_generated_from_structure
                
                schedule_details.append({
                    'id': schedule.id,
                    'week_start': schedule.week_start_date,
                    'week_end': schedule.week_end_date,
                    'status': schedule.status,
                    'is_approved': schedule.is_approved,
                    'is_locked': schedule.is_locked,
                    'structure_type': structure_type,
                    'is_synced': is_synced,
                    'is_generated_from_structure': schedule.is_generated_from_structure if hasattr(schedule, 'is_generated_from_structure') else False,
                    'structure_hash_match': check_structure_hash_match(schedule, dept) if hasattr(schedule, 'structure_hash') else False
                })
                
                if is_synced:
                    synced_count += 1
                else:
                    unsynced_count += 1
            
            department_reports.append({
                'department': dept,
                'total_schedules': len(schedules),
                'synced_count': synced_count,
                'unsynced_count': unsynced_count,
                'sync_percentage': (synced_count / len(schedules) * 100) if schedules else 0,
                'has_structure': bool(dept.schedule_structure),
                'structure_version': dept.schedule_structure_version if hasattr(dept, 'schedule_structure_version') else None,
                'schedule_details': schedule_details
            })
        
        # إحصائيات عامة
        total_schedules = sum(r['total_schedules'] for r in department_reports)
        total_synced = sum(r['synced_count'] for r in department_reports)
        overall_sync_percentage = (total_synced / total_schedules * 100) if total_schedules > 0 else 0
        
        notifications = get_user_notifications(current_user.id)
        
        return render_template('admin/admin_schedule_sync_report.html',
                             department_reports=department_reports,
                             total_schedules=total_schedules,
                             total_synced=total_synced,
                             overall_sync_percentage=overall_sync_percentage,
                             notifications=notifications)
        
    except Exception as e:
        flash(f'حدث خطأ في تحميل التقرير: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

def check_structure_hash_match(schedule, department):
    """التحقق من تطابق بصمة الهيكل"""
    from schedules_sync import calculate_structure_hash
    
    if not hasattr(schedule, 'structure_hash') or not schedule.structure_hash:
        return False
    
    if not department.schedule_structure:
        return False
    
    current_hash = calculate_structure_hash(department.schedule_structure)
    return schedule.structure_hash == current_hash

@app.route('/admin/sync_department_schedules/<int:dept_id>', methods=['POST'])
@login_required
def admin_sync_department_schedules(dept_id):
    """مزامنة جميع جداول قسم معين"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        from database import db_session
        from models import Department, WeeklySchedule
        from schedules_sync import normalize_schedule_to_structure, calculate_structure_hash
        
        department = db_session.query(Department).get(dept_id)
        if not department:
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
        
        if not department.schedule_structure:
            return jsonify({'success': False, 'message': 'القسم ليس لديه هيكل'})
        
        # الحصول على جميع جداول القسم
        schedules = db_session.query(WeeklySchedule).filter_by(
            department_id=dept_id
        ).all()
        
        synced_count = 0
        
        for schedule in schedules:
            # توحيد الجدول
            new_structure = normalize_schedule_to_structure(schedule, department)
            if new_structure:
                schedule.schedule_data = json.dumps(new_structure, ensure_ascii=False)
                
                # تحديث معلومات الهيكل
                if hasattr(schedule, 'structure_hash'):
                    schedule.structure_hash = calculate_structure_hash(department.schedule_structure)
                
                schedule.is_generated_from_structure = True
                synced_count += 1
        
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم مزامنة {synced_count} جدول للقسم {department.name}',
            'synced_count': synced_count
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في المزامنة: {str(e)}'
        })

@app.route('/admin/force_sync_schedule/<int:schedule_id>', methods=['POST'])
@login_required
def admin_force_sync_schedule(schedule_id):
    """إجبار مزامنة جدول معين"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        from schedules_sync import force_sync_single_schedule
        
        success, message = force_sync_single_schedule(schedule_id)
        
        return jsonify({
            'success': success,
            'message': message
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ في المزامنة: {str(e)}'
        })

@app.route('/admin/auto_sync_settings', methods=['GET', 'POST'])
@login_required
def admin_auto_sync_settings():
    """إعدادات المزامنة التلقائية"""
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    from database import db_session
    from models import Department
    
    if request.method == 'POST':
        try:
            # تحديث إعدادات المزامنة للأقسام
            for dept in db_session.query(Department).all():
                auto_sync_key = f'auto_sync_{dept.id}'
                structure_key = f'structure_version_{dept.id}'
                
                if auto_sync_key in request.form:
                    dept.auto_sync_schedules = True
                    
                    # تحديث نسخة الهيكل إذا تم التغيير
                    if structure_key in request.form:
                        new_version = int(request.form[structure_key])
                        if hasattr(dept, 'schedule_structure_version'):
                            dept.schedule_structure_version = new_version
                else:
                    dept.auto_sync_schedules = False
            
            db_session.commit()
            flash('تم حفظ إعدادات المزامنة بنجاح', 'success')
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ في حفظ الإعدادات: {str(e)}', 'error')
    
    # تحميل الإعدادات الحالية
    departments = db_session.query(Department).order_by(Department.name).all()
    
    # إعدادات النظام
    system_settings = {
        'last_auto_sync': getattr(app, 'last_auto_sync', None),
        'auto_sync_enabled': getattr(app, 'auto_sync_enabled', False),
        'sync_frequency': getattr(app, 'sync_frequency', 'daily')
    }
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_auto_sync_settings.html',
                         departments=departments,
                         system_settings=system_settings,
                         notifications=notifications)






@app.before_request
def auto_sync_check():
    """التحقق من المزامنة التلقائية قبل كل طلب (مرة يومياً)"""
    try:
        # تفعيل المزامنة التلقائية
        if not hasattr(app, 'auto_sync_enabled'):
            app.auto_sync_enabled = True
        
        if not app.auto_sync_enabled:
            return
        
        # تشغيل مرة واحدة يومياً
        if not hasattr(app, 'last_auto_sync'):
            app.last_auto_sync = datetime.now() - timedelta(days=1)
        
        # إصلاح المقارنة: تحويل datetime.datetime إلى datetime.date
        last_sync_date = app.last_auto_sync.date() if isinstance(app.last_auto_sync, datetime) else app.last_auto_sync
        today_date = date.today()
        
        # التحقق إذا مر يوم كامل منذ آخر مزامنة
        if isinstance(last_sync_date, date) and (today_date > last_sync_date):
            from schedules_sync import (
                sync_all_schedules_with_department_structures,
                create_future_schedules_from_structure
            )
            
            print("=== بدء المزامنة التلقائية ===")
            
            # 1. مزامنة الجداول الحالية
            synced_count = sync_all_schedules_with_department_structures()
            print(f"تم مزامنة {synced_count} جدول")
            
            # 2. إنشاء جداول مستقبلية
            created_count = create_future_schedules_from_structure()
            print(f"تم إنشاء {created_count} جدول مستقبلي")
            
            app.last_auto_sync = datetime.now()
            print("=== انتهت المزامنة التلقائية ===")
            
    except Exception as e:
        print(f"خطأ في المزامنة التلقائية: {e}")
        import traceback
        traceback.print_exc()


@app.before_request
def check_and_sync_schedules():
    """التحقق من المزامنة التلقائية"""
    try:
        if not hasattr(app, 'last_sync_check'):
            app.last_sync_check = datetime.now() - timedelta(hours=1)
        
        # تشغيل كل ساعة
        if datetime.now() > app.last_sync_check + timedelta(hours=1):
            from schedules_sync import (
                sync_all_schedules_with_department_structures,
                create_future_schedules_from_structure
            )
            
            # مزامنة الجداول الحالية
            sync_all_schedules_with_department_structures()
            
            # إنشاء جداول مستقبلية
            create_future_schedules_from_structure()
            
            app.last_sync_check = datetime.now()
            
    except Exception as e:
        print(f"خطأ في التحقق التلقائي: {e}")

def reject_leave(request_id, rejection_reason):
    # 1. التحقق من تاريخ الطلب
    if leave_request.start_date > date(اليوم.year, اليوم.month, 26):
        # 2. زيادة رصيد الموظف
        balance = EmployeeBalance.query.filter_by(user_id=request.user_id).first()
        balance.leave_balance += leave_request.total_days  # يزيد بالمثل
        # إذا كان total_days = 3، يزيد الرصيد بـ 3 أيام
        
        # 3. تحديث حالة الطلب
        leave_request.status = 'rejected'
        leave_request.rejection_reason = rejection_reason
        leave_request.approved_at = datetime.now()
        
        # 4. إرسال إشعار للموظف
        send_notification(request.user_id, "تم رفض طلب الإجازة", "زيادة في الرصيد")

@app.route('/manager/schedules')
@login_required
def manager_schedules():
    """صفحة مدير القسم للجداول مع الفلاتر"""
    if not current_user.is_manager:
        return redirect(url_for('user_dashboard'))
    
    # الحصول على قسم المدير
    department = db_session.query(Department).filter_by(primary_manager_id=current_user.id).first()
    if not department:
        flash('لم يتم تعيينك كمدير لأي قسم', 'error')
        return redirect(url_for('user_dashboard'))
    
    # الحصول على معاملات التصفية
    status_filter = request.args.get('status', '')
    month_filter = request.args.get('month', '')
    year_filter = request.args.get('year', '')
    
    # بناء الاستعلام الأساسي
    query = db_session.query(WeeklySchedule).filter_by(
        department_id=department.id
    )
    
    # تطبيق الفلاتر
    if status_filter:
        if status_filter == 'approved':
            query = query.filter(WeeklySchedule.is_approved == True)
        elif status_filter == 'pending':
            query = query.filter(WeeklySchedule.is_approved == False, WeeklySchedule.status == 'pending')
        elif status_filter == 'draft':
            query = query.filter(WeeklySchedule.is_approved == False, WeeklySchedule.status == 'draft')
        elif status_filter == 'locked':
            query = query.filter(WeeklySchedule.is_locked == True)
    
    # تصفية حسب الشهر والسنة
    if month_filter and month_filter.isdigit():
        query = query.filter(extract('month', WeeklySchedule.week_start_date) == int(month_filter))

    if year_filter and year_filter.isdigit():
        query = query.filter(extract('year', WeeklySchedule.week_start_date) == int(year_filter))
    # الحصول على جداول القسم
    schedules = query.order_by(WeeklySchedule.week_start_date.desc()).all()
    
    # تحميل البيانات المرتبطة
    for schedule in schedules:
        schedule.department = department
        schedule.creator = db_session.query(User).get(schedule.created_by) if schedule.created_by else None
        schedule.approver = db_session.query(User).get(schedule.approved_by) if schedule.approved_by else None
    
    # إحصائيات
    stats = {
        'total_schedules': len(schedules),
        'approved_schedules': len([s for s in schedules if s.is_approved]),
        'pending_schedules': len([s for s in schedules if s.status == 'pending']),
        'draft_schedules': len([s for s in schedules if s.status == 'draft']),
        'locked_schedules': len([s for s in schedules if s.is_locked]),
        'current_schedules': len([s for s in schedules if s.week_start_date <= date.today() <= s.week_end_date])
    }
    
    # تجميع الجداول حسب الشهر والسنة
    schedules_by_period = {}
    for schedule in schedules:
        period_key = schedule.week_start_date.strftime('%Y-%m')
        month_name = get_month_name_arabic(schedule.week_start_date.month)
        year = schedule.week_start_date.year
        
        if period_key not in schedules_by_period:
            schedules_by_period[period_key] = {
                'display_name': f"{month_name} {year}",
                'schedules': []
            }
        schedules_by_period[period_key]['schedules'].append(schedule)
    
    # توليد قائمة السنوات المتاحة
    available_years = sorted(set([s.week_start_date.year for s in schedules]), reverse=True)
    
    # حساب التواريخ الحالية
    today = date.today()
    days_since_saturday = (today.weekday() - 5) % 7
    current_week_start = today - timedelta(days=days_since_saturday)
    current_week_end = current_week_start + timedelta(days=6)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager/manager_schedules.html',
                         schedules=schedules,
                         schedules_by_period=schedules_by_period,
                         department=department,
                         stats=stats,
                         today=today,
                         current_week_start=current_week_start,
                         current_week_end=current_week_end,
                         available_years=available_years,
                         selected_status=status_filter,
                         selected_month=month_filter,
                         selected_year=year_filter,
                         notifications=notifications)

@app.route('/manager/schedule')
@login_required
def manager_schedule():
    """صفحة مدير القسم للجداول مع الفلاتر"""
    if not current_user.is_manager:
        return redirect(url_for('user_dashboard'))
    
    # الحصول على قسم المدير
    department = db_session.query(Department).filter_by(primary_manager_id=current_user.id).first()
    if not department:
        flash('لم يتم تعيينك كمدير لأي قسم', 'error')
        return redirect(url_for('user_dashboard'))
    
    # الحصول على معاملات التصفية
    status_filter = request.args.get('status', '')
    month_filter = request.args.get('month', '')
    year_filter = request.args.get('year', '')
    
    # بناء الاستعلام الأساسي
    query = db_session.query(WeeklySchedule).filter_by(
        department_id=department.id
    )
    
    # تطبيق الفلاتر
    if status_filter:
        if status_filter == 'approved':
            query = query.filter(WeeklySchedule.is_approved == True)
        elif status_filter == 'pending':
            query = query.filter(WeeklySchedule.is_approved == False, WeeklySchedule.status == 'pending')
        elif status_filter == 'draft':
            query = query.filter(WeeklySchedule.is_approved == False, WeeklySchedule.status == 'draft')
        elif status_filter == 'locked':
            query = query.filter(WeeklySchedule.is_locked == True)
    
    # تصفية حسب الشهر والسنة
    if month_filter and month_filter.isdigit():
        query = query.filter(extract('month', WeeklySchedule.week_start_date) == int(month_filter))

    if year_filter and year_filter.isdigit():
        query = query.filter(extract('year', WeeklySchedule.week_start_date) == int(year_filter))
    # الحصول على جداول القسم
    schedules = query.order_by(WeeklySchedule.week_start_date.desc()).all()
    
    # تحميل البيانات المرتبطة
    for schedule in schedules:
        schedule.department = department
        schedule.creator = db_session.query(User).get(schedule.created_by) if schedule.created_by else None
        schedule.approver = db_session.query(User).get(schedule.approved_by) if schedule.approved_by else None
    
    # إحصائيات
    stats = {
        'total_schedules': len(schedules),
        'approved_schedules': len([s for s in schedules if s.is_approved]),
        'pending_schedules': len([s for s in schedules if s.status == 'pending']),
        'draft_schedules': len([s for s in schedules if s.status == 'draft']),
        'locked_schedules': len([s for s in schedules if s.is_locked]),
        'current_schedules': len([s for s in schedules if s.week_start_date <= date.today() <= s.week_end_date])
    }
    
    # تجميع الجداول حسب الشهر والسنة
    schedules_by_period = {}
    for schedule in schedules:
        period_key = schedule.week_start_date.strftime('%Y-%m')
        month_name = get_month_name_arabic(schedule.week_start_date.month)
        year = schedule.week_start_date.year
        
        if period_key not in schedules_by_period:
            schedules_by_period[period_key] = {
                'display_name': f"{month_name} {year}",
                'schedules': []
            }
        schedules_by_period[period_key]['schedules'].append(schedule)
    
    # توليد قائمة السنوات المتاحة
    available_years = sorted(set([s.week_start_date.year for s in schedules]), reverse=True)
    
    # حساب التواريخ الحالية
    today = date.today()
    days_since_saturday = (today.weekday() - 5) % 7
    current_week_start = today - timedelta(days=days_since_saturday)
    current_week_end = current_week_start + timedelta(days=6)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager/manager_schedules.html',
                         schedules=schedules,
                         schedules_by_period=schedules_by_period,
                         department=department,
                         stats=stats,
                         today=today,
                         current_week_start=current_week_start,
                         current_week_end=current_week_end,
                         available_years=available_years,
                         selected_status=status_filter,
                         selected_month=month_filter,
                         selected_year=year_filter,
                         notifications=notifications)


def sync_all_schedules_with_structure():
    """مزامنة جميع الجداول مع أحدث هيكل للقسم"""
    try:
        print("=== بدء مزامنة جميع الجداول مع هياكل الأقسام ===")
        
        departments = db_session.query(Department).all()
        updated_count = 0
        
        for department in departments:
            if not department.schedule_structure:
                continue
            
            # التحقق من وجود السمات المطلوبة
            if not hasattr(department, 'schedule_structure_version'):
                continue
            
            # تحديث جميع الجداول غير المعتمدة لهذا القسم
            schedules = db_session.query(WeeklySchedule).filter_by(
                department_id=department.id,
                is_approved=False
            ).all()
            
            for schedule in schedules:
                try:
                    # إنشاء جدول جديد من الهيكل
                    new_schedule_data = create_schedule_from_department_structure(
                        department, schedule.week_start_date
                    )
                    
                    if new_schedule_data:
                        # استبدال بيانات الشيفتات فقط (الموظفين) مع الحفاظ على الأيام والوظائف
                        schedule.schedule_data = json.dumps(new_schedule_data, ensure_ascii=False)
                        schedule.structure_version = department.schedule_structure_version
                        
                        # تحديث بصمة الهيكل
                        structure_hash = calculate_structure_hash(department.schedule_structure)
                        schedule.structure_hash = structure_hash
                        
                        updated_count += 1
                        print(f"✓ تم تحديث الجدول {schedule.id} للقسم {department.name}")
                        
                except Exception as e:
                    print(f"❌ خطأ في تحديث الجدول {schedule.id}: {str(e)}")
                    continue
        
        db_session.commit()
        print(f"=== تم تحديث {updated_count} جدول ===")
        return updated_count
        
    except Exception as e:
        db_session.rollback()
        print(f"❌ خطأ في المزامنة: {str(e)}")
        import traceback
        traceback.print_exc()
        return 0



def auto_sync_all_templates():
    """مزامنة جميع القوالب تلقائياً مع هياكل الأقسام"""
    try:
        print("=== بدء مزامنة القوالب مع هياكل الأقسام ===")
        
        templates = db_session.query(DepartmentScheduleTemplate).filter_by(
            is_active=True,
            is_auto_synced=True
        ).all()
        
        synced_count = 0
        for template in templates:
            try:
                if template.sync_with_department_structure():
                    db_session.add(template)
                    synced_count += 1
                    print(f"✓ تمت مزامنة القالب: {template.template_name}")
            except Exception as e:
                print(f"❌ خطأ في مزامنة القالب {template.id}: {str(e)}")
                continue
        
        db_session.commit()
        print(f"=== تمت مزامنة {synced_count} قالب ===")
        return synced_count
        
    except Exception as e:
        db_session.rollback()
        print(f"❌ خطأ في المزامنة الجماعية: {str(e)}")
        return 0


def check_and_sync_templates():
    """التحقق من الحاجة للمزامنة وتنفيذها"""
    try:
        if not hasattr(app, 'last_template_sync'):
            app.last_template_sync = datetime.now() - timedelta(days=1)
        
        # التحقق حسب التردد
        now = datetime.now()
        
        # جلب القوالب التي تحتاج مزامنة
        templates_to_sync = db_session.query(DepartmentScheduleTemplate).filter(
            DepartmentScheduleTemplate.is_active == True,
            DepartmentScheduleTemplate.is_auto_synced == True,
            or_(
                DepartmentScheduleTemplate.next_sync_at == None,
                DepartmentScheduleTemplate.next_sync_at <= now
            )
        ).all()
        
        synced_count = 0
        for template in templates_to_sync:
            try:
                # تحديث تاريخ المزامنة التالي
                if template.sync_frequency == 'daily':
                    next_sync = now + timedelta(days=1)
                elif template.sync_frequency == 'weekly':
                    next_sync = now + timedelta(days=7)
                elif template.sync_frequency == 'monthly':
                    next_sync = now + timedelta(days=30)
                else:
                    next_sync = now + timedelta(days=7)
                
                # تنفيذ المزامنة
                if template.sync_with_department_structure():
                    template.last_synced_at = now
                    template.next_sync_at = next_sync
                    synced_count += 1
                    
            except Exception as e:
                print(f"❌ خطأ في مزامنة القالب {template.id}: {str(e)}")
                continue
        
        if synced_count > 0:
            db_session.commit()
            print(f"تمت مزامنة {synced_count} قالب تلقائياً")
        
        app.last_template_sync = now
        return synced_count
        
    except Exception as e:
        print(f"خطأ في التحقق من المزامنة: {str(e)}")
        return 0


@app.route('/admin/check_structures')
@login_required
def admin_check_structures():
    """فحص هياكل الجداول في جميع الأقسام"""
    if not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    departments = db_session.query(Department).all()
    structure_info = []
    
    for dept in departments:
        has_structure = bool(dept.schedule_structure)
        structure_info.append({
            'id': dept.id,
            'name': dept.name,
            'has_structure': has_structure,
            'structure_length': len(dept.schedule_structure) if dept.schedule_structure else 0,
            'auto_generate': dept.auto_generate_schedule
        })
    
    return render_template('admin/admin_check_structures.html',
                         departments=structure_info,
                         notifications=get_user_notifications(current_user.id))



@app.route('/admin/create_schedule_template', methods=['GET', 'POST'])
@login_required
def admin_create_schedule_template():
    """إنشاء قالب جدول جديد"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    if request.method == 'POST':
        try:
            department_id = request.form['department_id']
            template_name = request.form['template_name']
            description = request.form.get('description', '')
            sync_frequency = request.form.get('sync_frequency', 'weekly')
            is_auto_synced = 'is_auto_synced' in request.form
            
            department = db_session.query(Department).get(department_id)
            if not department:
                flash('القسم غير موجود', 'error')
                return redirect(url_for('admin_schedule_templates'))
            
            # إنشاء القالب الجديد
            template = DepartmentScheduleTemplate(
                department_id=department_id,
                template_name=template_name,
                description=description,
                sync_frequency=sync_frequency,
                is_auto_synced=is_auto_synced,
                created_by=current_user.id,
                updated_by=current_user.id
            )
            
            # مزامنة مع هيكل القسم
            if is_auto_synced and department.schedule_structure:
                template.schedule_structure = department.schedule_structure
                template.structure_version = department.schedule_structure_version
                template.structure_hash = template.calculate_structure_hash(department.schedule_structure)
            
            # تعيين تاريخ المزامنة التالي
            if is_auto_synced:
                now = datetime.now()
                template.last_synced_at = now
                if sync_frequency == 'daily':
                    template.next_sync_at = now + timedelta(days=1)
                elif sync_frequency == 'weekly':
                    template.next_sync_at = now + timedelta(days=7)
                elif sync_frequency == 'monthly':
                    template.next_sync_at = now + timedelta(days=30)
            
            db_session.add(template)
            db_session.commit()
            
            flash(f'تم إنشاء القالب "{template_name}" بنجاح', 'success')
            return redirect(url_for('admin_schedule_templates'))
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ أثناء إنشاء القالب: {str(e)}', 'error')
            return redirect(url_for('admin_create_schedule_template'))
    
    departments = db_session.query(Department).all()
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_create_schedule_template.html',
                         departments=departments,
                         notifications=notifications)


@app.route('/admin/sync_template/<int:template_id>')
@login_required
def admin_sync_template(template_id):
    """مزامنة قالب يدوياً"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        template = db_session.query(DepartmentScheduleTemplate).get(template_id)
        if not template:
            return jsonify({'success': False, 'message': 'القالب غير موجود'})
        
        # مزامنة القالب
        if template.sync_with_department_structure(force=True):
            template.updated_by = current_user.id
            template.updated_at = datetime.now()
            db_session.commit()
            
            return jsonify({
                'success': True,
                'message': f'تمت مزامنة القالب "{template.template_name}" بنجاح',
                'new_version': template.structure_version
            })
        else:
            return jsonify({
                'success': True,
                'message': 'القالب محدث بالفعل مع أحدث إصدار',
                'current_version': template.structure_version
            })
            
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'خطأ في المزامنة: {str(e)}'})


@app.route('/admin/generate_schedule_from_template/<int:template_id>', methods=['POST'])
@login_required
def admin_generate_schedule_from_template(template_id):
    """إنشاء جدول أسبوعي من قالب"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        template = db_session.query(DepartmentScheduleTemplate).get(template_id)
        if not template:
            return jsonify({'success': False, 'message': 'القالب غير موجود'})
        
        # الحصول على تاريخ بداية الأسبوع
        week_start_str = request.json.get('week_start_date')
        if week_start_str:
            week_start_date = datetime.strptime(week_start_str, '%Y-%m-%d').date()
        else:
            # استخدام الأسبوع القادم إذا لم يتم تحديد تاريخ
            today = date.today()
            days_since_saturday = (today.weekday() - 5) % 7
            week_start_date = today - timedelta(days=days_since_saturday) + timedelta(days=7)
        
        week_end_date = week_start_date + timedelta(days=6)
        
        # التحقق من عدم وجود جدول لهذا الأسبوع
        existing_schedule = db_session.query(WeeklySchedule).filter_by(
            department_id=template.department_id,
            week_start_date=week_start_date
        ).first()
        
        if existing_schedule:
            return jsonify({
                'success': False,
                'message': f'يوجد جدول بالفعل للفترة {week_start_date} إلى {week_end_date}'
            })
        
        # إنشاء الجدول من القالب
        schedule_data = template.generate_schedule_from_template(week_start_date)
        if not schedule_data:
            return jsonify({'success': False, 'message': 'لا يمكن إنشاء جدول من القالب'})
        
        # حفظ الجدول الجديد
        new_schedule = WeeklySchedule(
            department_id=template.department_id,
            week_start_date=week_start_date,
            week_end_date=week_end_date,
            schedule_data=json.dumps(schedule_data, ensure_ascii=False),
            created_by=current_user.id,
            status='draft',
            is_approved=False,
            is_locked=False,
            is_generated_from_structure=True,
            structure_version=template.structure_version,
            structure_hash=template.structure_hash
        )
        
        db_session.add(new_schedule)
        db_session.commit()
        
        # ربط القالب بالجدول
        template.weekly_schedule_id = new_schedule.id
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء جدول جديد من القالب "{template.template_name}"',
            'schedule_id': new_schedule.id,
            'week_start_date': week_start_date.strftime('%Y-%m-%d'),
            'week_end_date': week_end_date.strftime('%Y-%m-%d')
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({'success': False, 'message': f'خطأ في إنشاء الجدول: {str(e)}'})


@app.route('/admin/edit_template/<int:template_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_template(template_id):
    """تعديل قالب جدول"""
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    template = db_session.query(DepartmentScheduleTemplate).get(template_id)
    if not template:
        flash('القالب غير موجود', 'error')
        return redirect(url_for('admin_schedule_templates'))
    
    if request.method == 'POST':
        try:
            template.template_name = request.form['template_name']
            template.description = request.form.get('description', '')
            template.sync_frequency = request.form.get('sync_frequency', 'weekly')
            template.is_auto_synced = 'is_auto_synced' in request.form
            template.is_active = 'is_active' in request.form
            template.is_locked = 'is_locked' in request.form
            template.updated_by = current_user.id
            template.updated_at = datetime.now()
            
            # معالجة التواريخ
            valid_from = request.form.get('valid_from')
            valid_until = request.form.get('valid_until')
            
            if valid_from:
                template.valid_from = datetime.strptime(valid_from, '%Y-%m-%d').date()
            else:
                template.valid_from = None
            
            if valid_until:
                template.valid_until = datetime.strptime(valid_until, '%Y-%m-%d').date()
            else:
                template.valid_until = None
            
            # إذا تم تفعيل المزامنة التلقائية، تحديث الهيكل
            if template.is_auto_synced and template.department.schedule_structure:
                template.sync_with_department_structure(force=True)
            
            db_session.commit()
            flash('تم تحديث القالب بنجاح', 'success')
            return redirect(url_for('admin_schedule_templates'))
            
        except Exception as e:
            db_session.rollback()
            flash(f'حدث خطأ أثناء تحديث القالب: {str(e)}', 'error')
            return redirect(url_for('admin_edit_template', template_id=template_id))
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_edit_template.html',
                         template=template,
                         notifications=notifications)


@app.route('/admin/delete_template/<int:template_id>')
@login_required
def admin_delete_template(template_id):
    """حذف قالب"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        template = db_session.query(DepartmentScheduleTemplate).get(template_id)
        if not template:
            flash('القالب غير موجود', 'error')
            return redirect(url_for('admin_schedule_templates'))
        
        db_session.delete(template)
        db_session.commit()
        
        flash('تم حذف القالب بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء حذف القالب: {str(e)}', 'error')
    
    return redirect(url_for('admin_schedule_templates'))


@app.route('/admin/preview_template/<int:template_id>')
@login_required
def admin_preview_template(template_id):
    """معاينة القالب"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        template = db_session.query(DepartmentScheduleTemplate).get(template_id)
        if not template:
            return jsonify({'success': False, 'message': 'القالب غير موجود'})
        
        # إنشاء معاينة للأسبوع الحالي
        today = date.today()
        days_since_saturday = (today.weekday() - 5) % 7
        week_start_date = today - timedelta(days=days_since_saturday)
        
        schedule_data = template.generate_schedule_from_template(week_start_date)
        
        if not schedule_data:
            return jsonify({'success': False, 'message': 'لا يمكن إنشاء المعاينة'})
        
        return jsonify({
            'success': True,
            'template': template.to_dict(),
            'preview': schedule_data,
            'week_info': {
                'start_date': week_start_date.strftime('%Y-%m-%d'),
                'end_date': (week_start_date + timedelta(days=6)).strftime('%Y-%m-%d')
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'خطأ في المعاينة: {str(e)}'})



@app.route('/admin/template_sync_report')
@login_required
def admin_template_sync_report():
    """تقرير مزامنة القوالب"""
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    # إحصائيات
    total_templates = db_session.query(DepartmentScheduleTemplate).count()
    active_templates = db_session.query(DepartmentScheduleTemplate).filter_by(is_active=True).count()
    auto_sync_templates = db_session.query(DepartmentScheduleTemplate).filter_by(is_auto_synced=True).count()
    
    # آخر المزامنات
    last_synced_templates = db_session.query(DepartmentScheduleTemplate).filter(
        DepartmentScheduleTemplate.last_synced_at != None
    ).order_by(DepartmentScheduleTemplate.last_synced_at.desc()).limit(10).all()
    
    # القوالب التي تحتاج مزامنة
    now = datetime.now()
    templates_needing_sync = db_session.query(DepartmentScheduleTemplate).filter(
        DepartmentScheduleTemplate.is_active == True,
        DepartmentScheduleTemplate.is_auto_synced == True,
        or_(
            DepartmentScheduleTemplate.next_sync_at == None,
            DepartmentScheduleTemplate.next_sync_at <= now
        )
    ).all()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_template_sync_report.html',
                         total_templates=total_templates,
                         active_templates=active_templates,
                         auto_sync_templates=auto_sync_templates,
                         last_synced_templates=last_synced_templates,
                         templates_needing_sync=templates_needing_sync,
                         now=now,
                         notifications=notifications)

@app.before_request
def auto_check_templates():
    """التحقق من مزامنة القوالب قبل كل طلب (بحدود معقولة)"""
    try:
        if not hasattr(app, 'last_template_check'):
            app.last_template_check = datetime.now() - timedelta(hours=1)
        
        if datetime.now() > app.last_template_check + timedelta(hours=1):
            check_and_sync_templates()
            app.last_template_check = datetime.now()
    except Exception as e:
        print(f"خطأ في التحقق من القوالب: {e}")



def create_default_templates_for_departments():
    """إنشاء قوالب افتراضية لجميع الأقسام"""
    try:
        departments = db_session.query(Department).all()
        created_count = 0
        
        for department in departments:
            # التحقق من عدم وجود قالب بالفعل
            existing_template = db_session.query(DepartmentScheduleTemplate).filter_by(
                department_id=department.id,
                template_name=f"قالب {department.name} الافتراضي"
            ).first()
            
            if not existing_template:
                template = DepartmentScheduleTemplate(
                    department_id=department.id,
                    template_name=f"قالب {department.name} الافتراضي",
                    description=f"قالب افتراضي للقسم {department.name} - تم إنشاؤه تلقائياً",
                    is_auto_synced=True,
                    sync_frequency='weekly',
                    created_by=1,  # النظام
                    updated_by=1,
                    is_active=True
                )
                
                db_session.add(template)
                created_count += 1
                print(f"✓ تم إنشاء قالب افتراضي للقسم {department.name}")
        
        if created_count > 0:
            db_session.commit()
            print(f"=== تم إنشاء {created_count} قالب افتراضي ===")
        
        return created_count
        
    except Exception as e:
        db_session.rollback()
        print(f"❌ خطأ في إنشاء القوالب الافتراضية: {str(e)}")
        return 0


@app.before_request
def auto_check_templates_and_generate():
    """التحقق من مزامنة القوالب وإنشاء الجداول"""
    try:
        # تشغيل مرة واحدة يومياً
        if not hasattr(app, 'last_full_check'):
            app.last_full_check = datetime.now() - timedelta(days=1)
        
        if datetime.now() > app.last_full_check + timedelta(days=1):
            # 1. التحقق من مزامنة القوالب
            check_and_sync_templates()
            
            # 2. إنشاء الجداول التلقائية
            auto_generate_weekly_schedules()
            
            # 3. التحقق من القوالب الافتراضية
            create_default_templates_for_departments()
            
            app.last_full_check = datetime.now()
            
    except Exception as e:
        print(f"خطأ في التحقق التلقائي: {e}")


@app.route('/manager/edit_schedule/<int:schedule_id>', methods=['GET', 'POST'])
@login_required
def manager_edit_schedule(schedule_id):
    """عرض وتعديل الجدول باستخدام الهيكل المحدد"""
    try:
        # التحقق من صلاحية المستخدم
        if not current_user.is_manager:
            flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
            return redirect(url_for('user_dashboard'))
        
        # جلب الجدول من قاعدة البيانات
        schedule = db_session.get(WeeklySchedule, schedule_id)
        if not schedule:
            flash('الجدول غير موجود', 'error')
            return redirect(url_for('manager_schedules'))
        
        # التحقق من أن المدير مسؤول عن قسم هذا الجدول
        department = db_session.query(Department).get(schedule.department_id)
        if not department:
            flash('القسم غير موجود', 'error')
            return redirect(url_for('manager_schedules'))
            
        if department.primary_manager_id != current_user.id:
            flash('غير مصرح لك بتعديل هذا الجدول', 'error')
            return redirect(url_for('manager_schedules'))
        
        # جلب الموظفين في القسم
        employees = db_session.query(User).filter(
            User.department_id == department.id,
            User.is_active == True
        ).all()
        
        # جلب قوالب القسم
        templates = db_session.query(DepartmentScheduleTemplate).filter(
            DepartmentScheduleTemplate.department_id == department.id,
            DepartmentScheduleTemplate.is_active == True
        ).all()
        
        # جلب أحدث قالب نشط للقسم (اختياري)
        latest_template = None
        if templates:
            latest_template = templates[0]  # يمكنك استخدام منطق آخر لاختيار القالب
        
        # معالجة طلبات POST (حفظ البيانات)
        if request.method == 'POST':
            return handle_schedule_submission(request, schedule, department)
        
        # تحضير بيانات الجدول للعرض
        schedule_data = prepare_schedule_data(schedule, department, employees)
        
        # جلب الإشعارات
        notifications = get_user_notifications(current_user.id)
        
        return render_template('manager_edit_schedule.html',
                             schedule=schedule,
                             department=department,
                             employees=employees,
                             templates=templates,
                             latest_template=latest_template,
                             current_schedule=schedule_data,
                             notifications=notifications)
    
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ غير متوقع: {str(e)}', 'error')
        return redirect(url_for('manager_schedules'))


def prepare_schedule_data(schedule, department, employees):
    """تحضير بيانات الجدول للعرض مع دعم القوالب"""
    try:
        # جلب أحدث قالب نشط للقسم
        template = db_session.query(DepartmentScheduleTemplate).filter(
            DepartmentScheduleTemplate.department_id == department.id,
            DepartmentScheduleTemplate.is_active == True,
            DepartmentScheduleTemplate.valid_from <= schedule.week_start_date,
            (DepartmentScheduleTemplate.valid_until >= schedule.week_start_date) | 
            (DepartmentScheduleTemplate.valid_until == None)
        ).order_by(DepartmentScheduleTemplate.created_at.desc()).first()
        
        # إذا كان هناك قالب وكان الجدول ليس لديه بيانات، استخدم القالب
        if template and (not schedule.schedule_data or schedule.schedule_data == '{}'):
            print(f"⚡ إنشاء جدول من القالب: {template.template_name}")
            generated_schedule = template.generate_schedule_from_template(schedule.week_start_date)
            
            if generated_schedule:
                # حفظ الجدول المولد في قاعدة البيانات
                schedule.schedule_data = json.dumps(generated_schedule, ensure_ascii=False)
                schedule.source_template_id = template.id
                schedule.is_generated_from_structure = True
                schedule.structure_version = template.structure_version
                schedule.structure_hash = template.structure_hash
                db_session.commit()
                
                return generated_schedule
        
        # إذا كان هناك بيانات في الجدول، استخدمها
        if schedule.schedule_data:
            try:
                schedule_data = json.loads(schedule.schedule_data)
                
                # التحقق من صحة البيانات وهيكلها
                if isinstance(schedule_data, dict) and 'schedule' in schedule_data:
                    return schedule_data
                elif isinstance(schedule_data, list):
                    # تحويل البيانات القديمة إلى الهيكل الجديد
                    return convert_old_format(schedule_data, department, employees)
                else:
                    print(f"⚠️ هيكل بيانات غير معروف للجدول {schedule.id}")
            except json.JSONDecodeError as e:
                print(f"❌ خطأ في تحليل JSON للجدول {schedule.id}: {str(e)}")
        
        # إذا لم يكن هناك بيانات، إنشاء جدول فارغ من هيكل القسم
        return create_empty_schedule(schedule, department, employees)
        
    except Exception as e:
        print(f"❌ خطأ في تحضير بيانات الجدول: {str(e)}")
        return create_empty_schedule(schedule, department, employees)


def create_empty_schedule(schedule, department, employees):
    """إنشاء جدول فارغ من هيكل القسم"""
    try:
        # استخدام هيكل القسم إذا كان موجوداً
        if department.schedule_structure:
            structure_data = json.loads(department.schedule_structure)
            
            # إنشاء هيكل الجدول
            schedule_data = {
                'department': department.name,
                'department_id': department.id,
                'week_start_date': schedule.week_start_date.strftime('%Y-%m-%d'),
                'week_end_date': schedule.week_end_date.strftime('%Y-%m-%d'),
                'source': 'department_structure',
                'structure_version': department.schedule_structure_version,
                'schedule': []
            }
            
            # أيام الأسبوع
            days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
            
            # تعبئة الأيام من هيكل القسم
            for i, day_name in enumerate(days_of_week):
                current_date = schedule.week_start_date + timedelta(days=i)
                
                # البحث عن يوم في هيكل القسم
                day_structure = find_day_in_structure(structure_data, day_name)
                
                if day_structure:
                    day_entry = day_structure.copy()
                else:
                    day_entry = {
                        'day': day_name,
                        'date': current_date.strftime('%Y-%m-%d'),
                        'department': department.name,
                        'morning_shift': '',
                        'evening_shift': '',
                        'night_shift': '',
                        'job': 'موظف'
                    }
                
                schedule_data['schedule'].append(day_entry)
            
            return schedule_data
            
    except Exception as e:
        print(f"❌ خطأ في إنشاء الجدول الفارغ: {str(e)}")
    
    # إذا فشل كل شيء، إرجاع جدول فارغ بسيط
    return {
        'department': department.name,
        'schedule': []
    }


def find_day_in_structure(structure_data, day_name):
    """البحث عن يوم في بيانات الهيكل"""
    try:
        if isinstance(structure_data, dict) and 'schedule' in structure_data:
            for day in structure_data['schedule']:
                if isinstance(day, dict) and day.get('day') == day_name:
                    return day
        
        elif isinstance(structure_data, list):
            for day in structure_data:
                if isinstance(day, dict) and day.get('day') == day_name:
                    return day
        
        return None
    except Exception:
        return None


def convert_old_format(old_data, department, employees):
    """تحويل البيانات القديمة إلى الهيكل الجديد"""
    schedule_data = {
        'department': department.name,
        'department_id': department.id,
        'schedule': []
    }
    
    days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
    
    for entry in old_data:
        if isinstance(entry, dict) and 'day' in entry:
            schedule_data['schedule'].append(entry)
        elif len(entry) >= 7:
            # تحويل من صف array
            schedule_data['schedule'].append({
                'night_shift': entry[0] or '',
                'evening_shift': entry[1] or '',
                'morning_shift': entry[2] or '',
                'job': entry[3] or 'موظف',
                'department': entry[4] or department.name,
                'date': entry[5] or '',
                'day': entry[6] or ''
            })
    
    return schedule_data

@app.route('/api/templates/<int:template_id>/generate')
@login_required
def generate_schedule_from_template(template_id):
    """إنشاء جدول من قالب معين"""
    try:
        template = db_session.get(DepartmentScheduleTemplate, template_id)
        if not template:
            return jsonify({'success': False, 'message': 'القالب غير موجود'})
        
        week_start = request.args.get('week_start')
        if not week_start:
            return jsonify({'success': False, 'message': 'يرجى تحديد تاريخ بداية الأسبوع'})
        
        week_start_date = datetime.strptime(week_start, '%Y-%m-%d').date()
        
        # إنشاء الجدول من القالب
        schedule_data = template.generate_schedule_from_template(week_start_date)
        
        if schedule_data:
            return jsonify({
                'success': True,
                'schedule': schedule_data,
                'template_name': template.template_name,
                'structure_version': template.structure_version
            })
        else:
            return jsonify({'success': False, 'message': 'فشل في إنشاء الجدول من القالب'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


def create_schedule_from_department_structure(department, week_start_date):
    """
    إنشاء جدول جديد من هيكل القسم
    """
    try:
        if not department.schedule_structure:
            print(f"❌ القسم {department.name} ليس لديه هيكل")
            return None
        
        week_end_date = week_start_date + timedelta(days=6)
        
        # إنشاء هيكل جديد من هيكل القسم
        new_schedule_data = {
            'department': department.name,
            'week_start_date': week_start_date.strftime('%Y-%m-%d'),
            'week_end_date': week_end_date.strftime('%Y-%m-%d'),
            'source': 'department_structure',
            'schedule': []
        }
        
        # أيام الأسبوع
        days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
        
        for i, day_name in enumerate(days_of_week):
            current_date = week_start_date + timedelta(days=i)
            
            day_entry = {
                'day': day_name,
                'date': current_date.strftime('%Y-%m-%d'),
                'department': department.name,
                'morning_shift': 'موظف',
                'evening_shift': 'موظف',
                'night_shift': 'موظف',
                'job': 'موظف'
            }
            
            new_schedule_data['schedule'].append(day_entry)
        
        return new_schedule_data
        
    except Exception as e:
        print(f"❌ خطأ في إنشاء الجدول من هيكل القسم: {str(e)}")
        return None

def calculate_structure_hash(structure_json):
    """حساب بصمة الهيكل"""
    import hashlib
    if structure_json:
        return hashlib.sha256(structure_json.encode('utf-8')).hexdigest()
    return None



@app.route('/admin/schedule_templates')
@login_required
def admin_schedule_templates():
    """إدارة قوالب الجداول"""
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    templates = db_session.query(DepartmentScheduleTemplate).options(
        joinedload(DepartmentScheduleTemplate.department),
        joinedload(DepartmentScheduleTemplate.creator),
        joinedload(DepartmentScheduleTemplate.updater)
    ).order_by(DepartmentScheduleTemplate.updated_at.desc()).all()
    
    departments = db_session.query(Department).all()
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('admin/admin_schedule_templates.html',
                         templates=templates,
                         departments=departments,
                         notifications=notifications)

def update_department_structure(dept_id, new_structure):
    """تحديث هيكل القسم وتحديث جميع الجداول المرتبطة"""
    try:
        department = db_session.query(Department).get(dept_id)
        if not department:
            return False
        
        # تحديث هيكل القسم
        department.update_structure(new_structure)
        
        # تحديث جميع الجداول غير المعتمدة
        schedules = db_session.query(WeeklySchedule).filter_by(
            department_id=dept_id,
            is_approved=False
        ).all()
        
        updated_count = 0
        for schedule in schedules:
            new_schedule_data = create_schedule_from_department_structure(
                department, schedule.week_start_date
            )
            
            if new_schedule_data:
                schedule.schedule_data = json.dumps(new_schedule_data, ensure_ascii=False)
                schedule.structure_version = department.schedule_structure_version
                schedule.structure_hash = calculate_structure_hash(new_structure)
                updated_count += 1
        
        db_session.commit()
        
        # إرسال إشعارات
        if department.primary_manager_id:
            create_notification(
                department.primary_manager_id,
                'تم تحديث هيكل الجدول',
                f'تم تحديث هيكل جدول القسم {department.name} وتحديث {updated_count} جدول',
                'structure_updated'
            )
        
        print(f"تم تحديث هيكل القسم {department.name} و {updated_count} جدول")
        return True
        
    except Exception as e:
        db_session.rollback()
        print(f"❌ خطأ في تحديث هيكل القسم: {str(e)}")
        return False

def regenerate_all_schedules_from_structure():
    """إعادة توليد جميع الجداول من هياكل الأقسام (حتى المعتمدة)"""
    try:
        print("=== إعادة توليد جميع الجداول من هياكل الأقسام ===")
        
        departments = db_session.query(Department).all()
        regenerated_count = 0
        
        for department in departments:
            if not department.schedule_structure:
                continue
            
            # الحصول على جميع الجداول للقسم
            schedules = db_session.query(WeeklySchedule).filter_by(
                department_id=department.id
            ).all()
            
            for schedule in schedules:
                try:
                    # إنشاء جدول جديد من الهيكل
                    new_schedule_data = create_schedule_from_department_structure(
                        department, schedule.week_start_date
                    )
                    
                    if new_schedule_data:
                        # استبدال بيانات الشيفتات مع الحفاظ على الأيام والوظائف
                        schedule.schedule_data = json.dumps(new_schedule_data, ensure_ascii=False)
                        schedule.structure_version = department.schedule_structure_version
                        schedule.structure_hash = calculate_structure_hash(department.schedule_structure)
                        schedule.is_generated_from_structure = True
                        
                        regenerated_count += 1
                        print(f"✓ تم إعادة توليد الجدول {schedule.id} للقسم {department.name}")
                        
                except Exception as e:
                    print(f"❌ خطأ في إعادة توليد الجدول {schedule.id}: {str(e)}")
                    continue
        
        db_session.commit()
        print(f"=== تم إعادة توليد {regenerated_count} جدول ===")
        return regenerated_count
        
    except Exception as e:
        db_session.rollback()
        print(f"❌ خطأ في إعادة التوليد: {str(e)}")
        return 0

# إضافة route جديد في قسم المسؤول
@app.route('/admin/sync_schedules_with_structure')
@login_required
def admin_sync_schedules_with_structure():
    """مزامنة جميع الجداول مع هياكل الأقسام"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        updated_count = sync_all_schedules_with_structure()
        flash(f'تم مزامنة {updated_count} جدول مع هياكل الأقسام', 'success')
        
    except Exception as e:
        flash(f'خطأ في المزامنة: {str(e)}', 'error')
    
    return redirect(url_for('admin_schedules'))

@app.route('/admin/regenerate_all_from_structure')
@login_required
def admin_regenerate_all_from_structure():
    """إعادة توليد جميع الجداول من هياكل الأقسام"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        regenerated_count = regenerate_all_schedules_from_structure()
        flash(f'تم إعادة توليد {regenerated_count} جدول من هياكل الأقسام', 'success')
        
    except Exception as e:
        flash(f'خطأ في إعادة التوليد: {str(e)}', 'error')
    
    return redirect(url_for('admin_schedules'))


@app.route('/admin/update_structure/<int:dept_id>', methods=['POST'])
@login_required
def admin_update_structure(dept_id):
    """تحديث هيكل القسم مع مزامنة الجداول والقوالب"""
    if not current_user.is_admin:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        # الحصول على البيانات الجديدة
        schedule_structure = request.form.get('schedule_structure')
        if not schedule_structure:
            flash('بيانات الهيكل مطلوبة', 'error')
            return redirect(url_for('admin_schedule_structure', dept_id=dept_id))
        
        # التحقق من صحة JSON
        try:
            json.loads(schedule_structure)
        except json.JSONDecodeError:
            flash('بيانات الهيكل غير صالحة', 'error')
            return redirect(url_for('admin_schedule_structure', dept_id=dept_id))
        
        # تحديث هيكل القسم
        department = db_session.query(Department).get(dept_id)
        if not department:
            flash('القسم غير موجود', 'error')
            return redirect(url_for('admin_schedule_structure', dept_id=dept_id))
        
        # تحديث الهيكل
        department.schedule_structure = schedule_structure
        department.schedule_structure_version = (department.schedule_structure_version or 0) + 1
        department.structure_last_modified = datetime.now()
        
        # 1. مزامنة الجداول غير المعتمدة
        schedule_count = sync_all_schedules_with_structure()
        
        # 2. مزامنة القوالب المرتبطة
        template_count = 0
        templates = db_session.query(DepartmentScheduleTemplate).filter_by(
            department_id=dept_id,
            is_auto_synced=True
        ).all()
        
        for template in templates:
            if template.sync_with_department_structure(force=True):
                template_count += 1
        
        db_session.commit()
        
        flash(f'تم تحديث هيكل القسم ومزامنة {schedule_count} جدول و {template_count} قالب بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'خطأ في تحديث الهيكل: {str(e)}', 'error')
    
    return redirect(url_for('admin_schedule_structure', dept_id=dept_id))


@app.route('/debug/sync')
@login_required
def debug_sync():
    """طريق للتصحيح الفوري"""
    if not current_user.is_admin:
        return "غير مصرح"
    
    from database import db_session
    from models import Department, WeeklySchedule
    
    # 1. إعادة تفعيل جميع الأقسام
    departments = db_session.query(Department).all()
    for dept in departments:
        dept.auto_generate_schedule = True
    
    # 2. حذف جميع الجداول المستقبلية (اختياري)
    # today = date.today()
    # future_schedules = db_session.query(WeeklySchedule).filter(
    #     WeeklySchedule.week_start_date >= today
    # ).all()
    # for schedule in future_schedules:
    #     db_session.delete(schedule)
    
    db_session.commit()
    
    # 3. إنشاء جداول جديدة
    created = auto_generate_weekly_schedules()
    
    return f"تم تفعيل {len(departments)} قسم وإنشاء {created} جدول"

def auto_generate_weekly_schedules():
    """إنشاء الجداول الأسبوعية تلقائياً بنفس هيكل القسم"""
    try:
        # تاريخ بداية الأسبوع الحالي (السبت)
        today = date.today()
        days_since_saturday = (today.weekday() - 5) % 7
        current_week_start = today - timedelta(days=days_since_saturday)
        
        # إنشاء جداول للأسابيع القادمة
        weeks_to_generate = 4  # 4 أسابيع قادمة
        generated_count = 0
        
        print(f"=== بدء إنشاء الجداول التلقائية بنفس هيكل القسم ===")
        
        # الحصول على جميع الأقسام المفعلة للتوليد التلقائي
        auto_departments = db_session.query(Department).filter_by(auto_generate_schedule=True).all()
        
        for department in auto_departments:
            print(f"معالجة القسم: {department.name} (ID: {department.id})")
            
            for week_offset in range(weeks_to_generate):
                week_start_date = current_week_start + timedelta(days=7 * week_offset)
                week_end_date = week_start_date + timedelta(days=6)
                
                # التحقق من وجود جدول لهذا الأسبوع
                existing_schedule = db_session.query(WeeklySchedule).filter_by(
                    department_id=department.id,
                    week_start_date=week_start_date
                ).first()
                
                if not existing_schedule:
                    print(f"إنشاء جدول جديد بنفس هيكل القسم للفترة {week_start_date}")
                    
                    # إنشاء جدول جديد بنفس هيكل القسم
                    schedule_data = create_weekly_schedule_from_structure(department.id, week_start_date)
                    
                    if schedule_data:
                        new_schedule = WeeklySchedule(
                            department_id=department.id,
                            week_start_date=week_start_date,
                            week_end_date=week_end_date,
                            schedule_data=json.dumps(schedule_data, ensure_ascii=False),
                            created_by=1,  # النظام
                            status='draft',
                            is_locked=False,
                            is_approved=False,
                            is_generated_from_structure=True
                        )
                        
                        # حفظ الجدول في قاعدة البيانات
                        db_session.add(new_schedule)
                        generated_count += 1
                        print(f"✓ تم إنشاء جدول للقسم {department.name} للفترة {week_start_date}")
                    else:
                        print(f"❌ فشل في إنشاء بيانات الجدول للقسم {department.name}")
        
        db_session.commit()
        print(f"=== تم إنشاء {generated_count} جدول أسبوعي جديد بنفس هيكل القسم ===")
        return generated_count
        
    except Exception as e:
        print(f'❌ خطأ في الإنشاء التلقائي للجداول: {str(e)}')
        import traceback
        traceback.print_exc()
        db_session.rollback()
        return 0

def copy_schedule_structure_from_existing(source_schedule_id, target_department_id, week_start_date):
    """نسخ هيكل جدول موجود لجدول جديد"""
    try:
        source_schedule = db_session.query(WeeklySchedule).get(source_schedule_id)
        if not source_schedule:
            return None
        
        department = db_session.query(Department).get(target_department_id)
        if not department:
            return None
        
        # تحليل بيانات الجدول المصدر
        if source_schedule.schedule_data:
            source_data = json.loads(source_schedule.schedule_data)
            
            # إنشاء نسخة من الهيكل مع تحديث التواريخ
            week_end_date = week_start_date + timedelta(days=6)
            
            new_schedule_data = {
                'department': department.name,
                'week_start_date': week_start_date.strftime('%Y-%m-%d'),
                'week_end_date': week_end_date.strftime('%Y-%m-%d'),
                'source': 'copied_from_existing',
                'original_schedule_id': source_schedule_id,
                'schedule': []
            }
            
            # إذا كان المصدر يحتوي على بيانات جدول
            if 'schedule' in source_data:
                days_of_week = ['السبت', 'الأحد', 'الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
                
                for i, day_name in enumerate(days_of_week):
                    current_date = week_start_date + timedelta(days=i)
                    
                    # البحث عن اليوم المناسب في المصدر
                    day_template = None
                    for source_day in source_data['schedule']:
                        if isinstance(source_day, dict) and source_day.get('day') == day_name:
                            day_template = source_day
                            break
                    
                    if day_template:
                        # نسخ اليوم مع تحديث التاريخ
                        day_entry = day_template.copy()
                        day_entry['date'] = current_date.strftime('%Y-%m-%d')
                        day_entry['department'] = department.name
                    else:
                        # إنشاء يوم افتراضي
                        day_entry = {
                            'day': day_name,
                            'date': current_date.strftime('%Y-%m-%d'),
                            'department': department.name,
                            'morning_shift': '',
                            'evening_shift': '',
                            'night_shift': '',
                            'job': 'موظف'
                        }
                    
                    new_schedule_data['schedule'].append(day_entry)
                
                return new_schedule_data
        
        return None
        
    except Exception as e:
        print(f"خطأ في نسخ هيكل الجدول: {str(e)}")
        return None

@app.route('/admin/create_schedules_from_structure', methods=['POST'])
@login_required
def admin_create_schedules_from_structure():
    """إنشاء جداول جديدة بنفس هيكل القسم"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        department_id = request.json.get('department_id')
        weeks_count = int(request.json.get('weeks_count', 4))
        
        department = db_session.query(Department).get(department_id)
        if not department:
            return jsonify({'success': False, 'message': 'القسم غير موجود'})
        
        # تاريخ بداية الأسبوع الحالي
        today = date.today()
        days_since_saturday = (today.weekday() - 5) % 7
        current_week_start = today - timedelta(days=days_since_saturday)
        
        generated_count = 0
        
        # إنشاء جداول للأسبوع القادم فصاعداً
        for week_offset in range(1, weeks_count + 1):
            week_start_date = current_week_start + timedelta(days=7 * week_offset)
            
            # التحقق من عدم وجود جدول لهذا الأسبوع
            existing_schedule = db_session.query(WeeklySchedule).filter_by(
                department_id=department_id,
                week_start_date=week_start_date
            ).first()
            
            if not existing_schedule:
                # إنشاء جدول جديد بنفس هيكل القسم
                schedule_data = create_weekly_schedule_from_structure(department_id, week_start_date)
                
                if schedule_data:
                    week_end_date = week_start_date + timedelta(days=6)
                    
                    new_schedule = WeeklySchedule(
                        department_id=department_id,
                        week_start_date=week_start_date,
                        week_end_date=week_end_date,
                        schedule_data=json.dumps(schedule_data, ensure_ascii=False),
                        created_by=current_user.id,
                        status='draft',
                        is_locked=False,
                        is_approved=False,
                        is_generated_from_structure=True
                    )
                    
                    db_session.add(new_schedule)
                    generated_count += 1
                    print(f"تم إنشاء جدول جديد للقسم {department.name} للفترة {week_start_date}")
        
        db_session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء {generated_count} جدول جديد بنفس هيكل القسم',
            'generated_count': generated_count
        })
        
    except Exception as e:
        db_session.rollback()
        return jsonify({
            'success': False,
            'message': f'خطأ في إنشاء الجداول: {str(e)}'
        })


@app.before_request
def auto_check_and_sync_schedules():
    """التحقق من الجداول والمزامنة تلقائياً"""
    try:
        # تشغيل هذا فقط مرة في اليوم
        if not hasattr(app, 'last_schedule_check'):
            app.last_schedule_check = date.today() - timedelta(days=1)
        
        if date.today() > app.last_schedule_check:
            # 1. إنشاء جداول جديدة بنفس هيكل القسم
            auto_generate_weekly_schedules()
            
            # 2. التحقق من أرصدة الإذونات
            check_and_reset_balances()
            
            # 3. تحديث آخر تاريخ تحقق
            app.last_schedule_check = date.today()
            
    except Exception as e:
        print(f"خطأ في التحقق التلقائي: {e}")


def debug_sync_process():
    """دالة تصحيح لعملية المزامنة"""
    print("=== بدء تصحيح المزامنة ===")
    
    # 1. فحص الأقسام
    departments = db_session.query(Department).all()
    print(f"عدد الأقسام: {len(departments)}")
    
    for dept in departments:
        print(f"القسم: {dept.name} (ID: {dept.id}) - Auto Generate: {dept.auto_generate_schedule}")
    
    # 2. فحص جداول موجودة
    schedules = db_session.query(WeeklySchedule).all()
    print(f"عدد الجداول الموجودة: {len(schedules)}")
    
    # 3. محاولة إنشاء جداول
    created_count = auto_generate_weekly_schedules()
    print(f"تم إنشاء {created_count} جدول جديد")
    
    return created_count


# إضافة وظيفة تلقائية لمزامنة الجداول
@app.before_request
def auto_sync_schedules():
    """مزامنة الجداول تلقائياً مع هياكل الأقسام"""
    try:
        # تشغيل هذا مرة واحدة يومياً
        if not hasattr(app, 'last_sync_check'):
            app.last_sync_check = date.today() - timedelta(days=1)
        
        if date.today() > app.last_sync_check.date():
            # مزامنة الجداول غير المعتمدة
            sync_all_schedules_with_structure()
            
            # إنشاء الجداول المستقبلية
            auto_generate_weekly_schedules()
            
            app.last_sync_check = date.today()
    except Exception as e:
        print(f"خطأ في المزامنة التلقائية: {e}")

# تحديث دالة manager_edit_schedule للحفاظ على الهيكل
def handle_schedule_submission(request, schedule, department):
    """معالجة حفظ الجدول مع الحفاظ على الهيكل"""
    try:
        action = request.form.get('action', 'save')
        schedule_data_json = request.form.get('schedule_data', '{}')
        
        # التحقق من صحة البيانات
        if not schedule_data_json:
            flash('لم يتم استلام بيانات الجدول', 'error')
            return redirect(url_for('manager_edit_schedule', schedule_id=schedule.id))
        
        try:
            schedule_data = json.loads(schedule_data_json)
        except json.JSONDecodeError:
            flash('بيانات الجدول غير صالحة', 'error')
            return redirect(url_for('manager_edit_schedule', schedule_id=schedule.id))
        
        # الحفاظ على الهيكل الأساسي من القسم
        structure_data = json.loads(department.schedule_structure) if department.schedule_structure else None
        
        if structure_data:
            # تحديث بيانات الموظفين فقط مع الحفاظ على الأيام والوظائف
            for day_entry in schedule_data.get('schedule', []):
                # إيجاد اليوم المقابل في الهيكل
                day_structure = find_day_in_structure(structure_data, day_entry.get('day'))
                if day_structure:
                    # الحفاظ على الوظيفة من الهيكل إذا لم يتم التعديل
                    if not day_entry.get('job') or day_entry.get('job') == 'موظف':
                        day_entry['job'] = day_structure.get('job', 'موظف')
        
        # تحديث حالة الجدول بناءً على الإجراء
        if action == 'submit':
            schedule.status = 'pending'
            schedule.is_locked = True
            flash('تم إرسال الجدول للاعتماد بنجاح', 'success')
        else:  # save
            schedule.status = 'draft'
            flash('تم حفظ الجدول بنجاح', 'success')
        
        # تحديث بيانات الجدول
        schedule.schedule_data = json.dumps(schedule_data, ensure_ascii=False)
        schedule.updated_at = datetime.now()
        
        # إزالة علامة "منشأ من الهيكل" لأن المدير قام بالتعديل
        schedule.is_generated_from_structure = False
        
        db_session.commit()
        
        return redirect(url_for('manager_schedules'))
    
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء حفظ الجدول: {str(e)}', 'error')
        return redirect(url_for('manager_edit_schedule', schedule_id=schedule.id))



@app.context_processor
def inject_schedule_info():
    """حقن معلومات الجداول في جميع القوالب"""
    def is_schedule_synced(schedule, department):
        """التحقق من أن الجدول مزامن مع أحدث هيكل"""
        if not schedule or not department:
            return True
        
        if not schedule.is_generated_from_structure:
            return True  # تم تعديله يدوياً
        
        if not department.schedule_structure_version:
            return True
        
        if schedule.structure_version != department.schedule_structure_version:
            return False
        
        if schedule.structure_hash != calculate_structure_hash(department.schedule_structure):
            return False
        
        return True
    
    return dict(
        is_schedule_synced=is_schedule_synced,
        calculate_structure_hash=calculate_structure_hash
    )


@app.route('/manager/create_future_schedules', methods=['POST'])
@login_required
def manager_create_future_schedules():
    """إنشاء 4 جداول مستقبلية جديدة للقسم بناءً على آخر جدول"""
    if not current_user.is_manager:
        flash('غير مصرح لك بهذا الإجراء', 'error')
        return redirect(url_for('user_dashboard'))
    
    # الحصول على قسم المدير
    department = db_session.query(Department).filter_by(primary_manager_id=current_user.id).first()
    if not department:
        flash('لم يتم تعيينك كمدير لأي قسم', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        # الحصول على آخر جدول للقسم
        last_schedule = db_session.query(WeeklySchedule).filter_by(
            department_id=department.id
        ).order_by(WeeklySchedule.week_start_date.desc()).first()
        
        if not last_schedule:
            flash('لا توجد جداول سابقة لنسخ الهيكل منها', 'error')
            return redirect(url_for('manager_schedules'))
        
        # حساب تاريخ بداية الأسبوع القادم
        today = date.today()
        days_since_saturday = (today.weekday() - 5) % 7
        current_week_start = today - timedelta(days=days_since_saturday)
        
        # إنشاء 4 جداول مستقبلية
        created_count = 0
        for week_offset in range(1, 49):  # 4 أسابيع قادمة
            week_start_date = current_week_start + timedelta(days=7 * week_offset)
            week_end_date = week_start_date + timedelta(days=6)
            
            # التحقق من عدم وجود جدول لهذا الأسبوع
            existing_schedule = db_session.query(WeeklySchedule).filter_by(
                department_id=department.id,
                week_start_date=week_start_date
            ).first()
            
            if not existing_schedule:
                # إنشاء جدول جديد بنفس هيكل آخر جدول
                new_schedule = WeeklySchedule(
                    department_id=department.id,
                    week_start_date=week_start_date,
                    week_end_date=week_end_date,
                    schedule_data=last_schedule.schedule_data,  # نسخ نفس البيانات
                    created_by=current_user.id,
                    status='draft',
                    is_locked=False,
                    is_approved=False
                )
                
                db_session.add(new_schedule)
                created_count += 1
                
                print(f"تم إنشاء جدول جديد للقسم {department.name} للفترة {week_start_date}")
        
        db_session.commit()
        
        if created_count > 0:
            flash(f'تم إنشاء {created_count} جدول مستقبلي جديد للقسم بناءً على آخر جدول', 'success')
            
            # إرسال إشعار للمدير
            create_notification(
                current_user.id,
                'تم إنشاء جداول مستقبلية',
                f'تم إنشاء {created_count} جدول مستقبلي جديد للقسم {department.name}',
                'schedules_created',
                action_url=url_for('manager_schedules')
            )
        else:
            flash('جميع الجداول المستقبلية موجودة مسبقاً', 'info')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء إنشاء الجداول: {str(e)}', 'error')
        print(f"Error creating future schedules: {e}")
    
    return redirect(url_for('manager_schedules'))

# Add this context processor to make the function available in all templates
@app.context_processor
def utility_processor():
    def get_month_name_arabic(month):
        """إرجاع اسم الشهر بالعربية"""
        months = {
            1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
            5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
            9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
        }
        return months.get(month, '')
    
    return dict(get_month_name_arabic=get_month_name_arabic)


@app.route('/manager/export_leave_requests')
@login_required
def export_leave_requests():
    """تصدير طلبات الإجازات إلى Excel"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    managed_departments = db_session.query(Department).filter_by(primary_manager_id=current_user.id).all()
    department_ids = [dept.id for dept in managed_departments]
    
    # الحصول على معاملات التصفية
    filter_status = request.args.get('status', 'all')
    filter_type = request.args.get('type', 'all')
    filter_month = request.args.get('month', 'all')
    
    # بناء الاستعلام الأساسي (نفس منطق manager_leave_requests)
    query = db_session.query(LeaveRequest).filter(
        LeaveRequest.department_id.in_(department_ids)
    )
    
    # تطبيق الفلاتر (نفس منطق manager_leave_requests)
    if filter_status != 'all':
        query = query.filter(LeaveRequest.status == filter_status)
    
    if filter_type != 'all':
        query = query.filter(LeaveRequest.leave_type == filter_type)
    
    if filter_month != 'all':
        try:
            if '-' not in filter_month:
                current_year = date.today().year
                month_num = int(filter_month)
                year_month = f"{current_year}-{month_num:02d}"
            else:
                year_month = filter_month
            
            year, month = map(int, year_month.split('-'))
            
            if month == 12:
                financial_month_start = date(year, 12, 26)
                financial_month_end = date(year + 1, 1, 25)
            else:
                financial_month_start = date(year, month, 26)
                financial_month_end = date(year, month + 1, 25)
            
            query = query.filter(
                LeaveRequest.start_date <= financial_month_end,
                LeaveRequest.end_date >= financial_month_start
            )
        except (ValueError, IndexError):
            pass
    
    leave_requests = query.order_by(LeaveRequest.created_at.desc()).all()
    
    # تحضير البيانات للتصدير
    data = []
    for req in leave_requests:
        user = db_session.get(User, req.user_id)
        department = db_session.get(Department, req.department_id)
        
        data.append({
            'اسم الموظف': user.name if user else 'غير معين',
            'القسم': department.name if department else 'غير معين',
            'نوع الإجازة': req.leave_type,
            'تاريخ البداية': req.start_date.strftime('%Y-%m-%d'),
            'تاريخ النهاية': req.end_date.strftime('%Y-%m-%d'),
            'عدد الأيام': req.total_days,
            'السبب': req.reason or '',
            'حالة الطلب': get_status_arabic(req.status),
            'تاريخ الطلب': req.created_at.strftime('%Y-%m-%d %H:%M'),
            'تاريخ القرار': req.approved_at.strftime('%Y-%m-%d %H:%M') if req.approved_at else '',
            'سبب الرفض': req.rejection_reason or ''
        })
    
    # إنشاء DataFrame
    df = pd.DataFrame(data)
    
    # إنشاء ملف Excel في الذاكرة
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='طلبات الإجازات', index=False)
        
        # تحسين تنسيق الأعمدة
        worksheet = writer.sheets['طلبات الإجازات']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # إنشاء اسم الملف
    filename = f"طلبات_الإجازات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def get_status_arabic(status):
    """تحويل حالة الطلب إلى العربية"""
    status_map = {
        'pending': 'معلقة',
        'approved': 'مقبولة',
        'rejected': 'مرفوضة'
    }
    return status_map.get(status, status)


from sqlalchemy.orm import joinedload, contains_eager

@app.route('/manager/leave_requests')
@login_required
def manager_leave_requests():
    """طلبات الإجازات المعلقة مع تجميع الشيفتات"""
    if not current_user.is_manager and not current_user.is_admin:
        flash('غير مصرح لك بالوصول', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter_by(primary_manager_id=current_user.id).all()
        
        if not managed_departments:
            flash('لم يتم تعيينك كمدير لأي قسم', 'warning')
            return render_template('manager/manager_leave_requests.html',
                                grouped_requests={},
                                pending_requests=[],
                                leave_requests=[],
                                departments=[],
                                processed_requests=[],
                                notifications=[])
        
        department_ids = [dept.id for dept in managed_departments]
        
        # Get ALL leave requests with eager loading
        all_requests = db_session.query(LeaveRequest)\
            .options(
                joinedload(LeaveRequest.user),
                joinedload(LeaveRequest.department),
                joinedload(LeaveRequest.approver)
            )\
            .filter(LeaveRequest.department_id.in_(department_ids))\
            .order_by(LeaveRequest.created_at.desc())\
            .all()
        
        # للتصحيح: تحقق من تحميل البيانات
        print("=" * 50)
        print(f"عدد الطلبات الكلي: {len(all_requests)}")
        print("=" * 50)
        
        for i, req in enumerate(all_requests[:10]):  # أول 10 طلبات فقط للتصحيح
            has_user = req.user is not None
            has_dept = req.department is not None
            user_name = req.user.name if has_user else "لا يوجد"
            dept_name = req.department.name if has_dept else "لا يوجد"
            
            print(f"طلب {i+1}: ID={req.id}, User ID={req.user_id}, Has User={has_user}, User Name='{user_name}'")
        
        # Separate pending and processed requests
        pending_requests = [req for req in all_requests if req.status == 'pending']
        processed_requests = [req for req in all_requests if req.status != 'pending']
        
        # تجميع الطلبات المعلقة حسب parent_request_id
        grouped_requests = {}
        
        for req in pending_requests:
            if req.parent_request_id not in grouped_requests:
                grouped_requests[req.parent_request_id] = {
                    'requests': [],
                    'user_id': req.user_id,
                    'department_id': req.department_id,
                    'leave_type': req.leave_type,
                    'date': req.leave_date,
                    'reason': req.reason,
                    'created_at': req.created_at
                }
            grouped_requests[req.parent_request_id]['requests'].append(req)
        
        # تجميع بيانات المستخدمين والأقسام من الطلبات المعلقة
        for parent_id, group in grouped_requests.items():
            # البحث عن الطلب الأول في المجموعة
            first_request = next((req for req in pending_requests if req.parent_request_id == parent_id), None)
            if first_request:
                group['user'] = first_request.user
                group['department'] = first_request.department
        
        notifications = get_user_notifications(current_user.id)

        current_period_start, current_period_end = get_rejection_period_dates()

        can_reject_requests = []
        for req in pending_requests:
            req_date = req.leave_date if hasattr(req, 'leave_date') and req.leave_date else req.start_date
            if can_reject_leave_by_date(req_date):
                can_reject_requests.append(req)
        
        return render_template('manager/manager_leave_requests.html',
                            grouped_requests=grouped_requests,
                            pending_requests=pending_requests,
                            leave_requests=all_requests,
                            departments=managed_departments,
                            processed_requests=processed_requests,
                            notifications=notifications,
                            current_user=current_user,
                            can_reject_requests=can_reject_requests,
                            current_period_start=current_period_start,
                            current_period_end=current_period_end)
            
    except Exception as e:
        print(f"Error in manager_leave_requests: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'حدث خطأ في تحميل البيانات: {str(e)}', 'error')
        return render_template('manager/manager_leave_requests.html',
                            grouped_requests={},
                            pending_requests=[],
                            leave_requests=[],
                            departments=[],
                            processed_requests=[],
                            notifications=[])


def can_reject_leave_by_date(leave_date):
    print(leave_date)
    """
    التحقق مما إذا كان يمكن رفض الإجازة بناءً على التاريخ
    القاعدة: يمكن رفض الإجازة فقط في نفس الفترة الزمنية (الشهر المالي)
    
    أمثلة:
    - إجازة 15 يناير 2025 ← تتبع لنطاق يناير (26 ديسمبر 2024 إلى 25 يناير 2025)
    - اليوم 20 يناير 2025 ← يقع في نفس النطاق ← يمكن الرفض
    - اليوم 26 يناير 2025 ← يقع في نطاق فبراير ← لا يمكن الرفض
    """
    if not leave_date:
        return False
    
    # تحويل التاريخ إذا كان نصياً
    if isinstance(leave_date, str):
        try:
            leave_date = datetime.strptime(leave_date, '%Y-%m-%d').date()
        except:
            return False
    
    today = date.today()
    
    # تحديد الفترة الزمنية (الشهر المالي) لتاريخ الإجازة
    leave_period_start, leave_period_end = get_financial_period_for_date(leave_date)
    
    # تحديد الفترة الزمنية (الشهر المالي) لليوم الحالي
    today_period_start, today_period_end = get_financial_period_for_date(today)
    
    # يمكن الرفض فقط إذا كان اليوم في نفس الفترة الزمنية لتاريخ الإجازة
    return (leave_period_start == today_period_start and 
            leave_period_end == today_period_end)



def get_manager_department_stats(manager_id):
    """الحصول على إحصائيات قسم المدير"""
    # الحصول على الأقسام التي يديرها المدير
    managed_departments = db_session.query(Department).filter_by(
        primary_manager_id=manager_id
    ).all()
    
    if not managed_departments:
        return {}
    
    department_ids = [dept.id for dept in managed_departments]
    
    # حساب عدد الموظفين في القسم
    department_employees = db_session.query(User).filter(
        User.department_id.in_(department_ids),
        User.is_admin == False,
        User.is_active == True
    ).all()
    
    # حساب عدد الإذونات اليوم
    today = date.today()
    today_permissions = db_session.query(PermissionRequest).filter(
        PermissionRequest.department_id.in_(department_ids),
        PermissionRequest.date == today,
        PermissionRequest.status == 'approved'
    ).count()
    
    return {
        'department_employees': department_employees,
        'today_permissions': today_permissions,
        'managed_departments': managed_departments
    }

def get_department_employees_without_managers(department_id):
    """الحصول على موظفي القسم (غير المديرين)"""
    employees = db_session.query(User).filter(
        User.department_id == department_id,
        User.is_admin == False,
        User.is_active == True
    ).all()
    
    return employees



@app.route('/manager/create_permission', methods=['GET', 'POST'])
@login_required
def manager_create_permission():
    """صفحة إنشاء إذن للموظفين بواسطة المدير"""
    if not current_user.is_manager and not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('user_dashboard'))
    
    try:
        # الحصول على إحصائيات قسم المدير
        stats = get_manager_department_stats(current_user.id)
        
        if not stats or not stats.get('managed_departments'):
            flash('لم يتم تعيينك كمدير لأي قسم', 'error')
            return redirect(url_for('manager_dashboard'))
        
        # الحصول على موظفي القسم الأول الذي يديره (لتبسيط الأمثلة)
        department = stats['managed_departments'][0]
        department_employees = get_department_employees_without_managers(department.id)
        
        # معالجة طلب POST (إنشاء إذن)
        if request.method == 'POST':
            return handle_manager_permission_creation(request, current_user, department)
        
        # تحضير البيانات للعرض
        notifications = get_user_notifications(current_user.id)
        
        return render_template('manager_create_permission.html',
                             department_employees=department_employees,
                             today_permissions=stats['today_permissions'],
                             current_user=current_user,
                             notifications=notifications)
    
    except Exception as e:
        print(f"Error in manager_create_permission: {str(e)}")
        flash(f'حدث خطأ: {str(e)}', 'error')
        return redirect(url_for('manager_dashboard'))


def handle_manager_permission_creation(request, manager, department):
    """معالجة إنشاء إذن من قبل المدير"""
    try:
        # التحقق من البيانات المطلوبة
        required_fields = ['employee_id', 'permission_type', 'date', 'time', 'reason']
        for field in required_fields:
            if field not in request.form or not request.form[field].strip():
                flash(f'حقل {field} مطلوب', 'error')
                return redirect(url_for('manager_create_permission'))
        
        employee_id = request.form['employee_id']
        permission_type = request.form['permission_type']
        permission_date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        time_str = request.form['time']
        reason = request.form['reason']
        
        # التحقق من أن الموظف ينتمي لنفس قسم المدير
        employee = db_session.query(User).get(employee_id)
        if not employee or employee.department_id != department.id:
            flash('الموظف غير موجود أو لا ينتمي لقسمك', 'error')
            return redirect(url_for('manager_create_permission'))
        
        # التحقق من أن التاريخ ليس في الماضي
        if permission_date < date.today():
            flash('لا يمكن إنشاء إذن لتاريخ ماضي', 'error')
            return redirect(url_for('manager_create_permission'))
        
        # تحويل الوقت من تنسيق HH:MM إلى دقائق
        try:
            time_obj = datetime.strptime(time_str, '%H:%M').time()
            time_minutes = time_obj.hour * 60 + time_obj.minute
        except ValueError:
            flash('تنسيق الوقت غير صحيح', 'error')
            return redirect(url_for('manager_create_permission'))
        
        # إنشاء طلب الإذن مع الموافقة المباشرة
        permission_request = PermissionRequest(
            user_id=employee_id,
            department_id=department.id,
            permission_type=permission_type,
            date=permission_date,
            time=time_minutes,
            reason=reason,
            extra_data=json.dumps({
                'created_by_manager': True,
                'manager_id': manager.id,
                'manager_name': manager.name,
                'source': 'manager_direct'
            }, ensure_ascii=False),
            status='approved',  # موافقة مباشرة من المدير
            approved_by=manager.id,
            approved_at=datetime.now(),
            created_at=datetime.now()
        )
        
        db_session.add(permission_request)
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'قام مديرك {manager.name} بمنحك إذن من نوع {permission_type}'
        notification_message += f' بتاريخ {permission_date.strftime("%Y-%m-%d")}'
        notification_message += f' بسبب: {reason}'
        
        create_notification(
            employee_id,
            'تم منحك إذن جديد',
            notification_message,
            'permission_granted',
            related_id=permission_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        # إرسال إشعار للمدير نفسه
        create_notification(
            manager.id,
            'تم إنشاء إذن للموظف',
            f'تم بنجاح منح إذن للموظف {employee.name}',
            'manager_permission_created',
            related_id=permission_request.id,
            action_url=url_for('manager_permission_requests')
        )
        
        flash(f'تم منح الإذن للموظف {employee.name} بنجاح', 'success')
        return redirect(url_for('manager_create_permission'))
    
    except Exception as e:
        db_session.rollback()
        print(f"Error handling manager permission creation: {str(e)}")
        flash(f'حدث خطأ أثناء إنشاء الإذن: {str(e)}', 'error')
        return redirect(url_for('manager_create_permission'))



@app.route('/api/manager/employee_suggestions')
@login_required
def api_manager_employee_suggestions():
    """API للحصول على اقتراحات الموظفين للبحث"""
    if not current_user.is_manager and not current_user.is_admin:
        return jsonify({'success': False, 'message': 'غير مصرح'})
    
    try:
        query = request.args.get('q', '')
        
        # الحصول على الأقسام التي يديرها المستخدم
        managed_departments = db_session.query(Department).filter_by(
            primary_manager_id=current_user.id
        ).all()
        
        if not managed_departments:
            return jsonify({'success': True, 'employees': []})
        
        department_ids = [dept.id for dept in managed_departments]
        
        # البحث عن الموظفين
        employees_query = db_session.query(User).filter(
            User.department_id.in_(department_ids),
            User.is_admin == False,
            User.is_active == True
        )
        
        if query:
            employees_query = employees_query.filter(
                User.name.ilike(f'%{query}%') |
                User.username.ilike(f'%{query}%')
            )
        
        employees = employees_query.order_by(User.name).limit(10).all()
        
        # تحضير البيانات
        employees_data = []
        for emp in employees:
            employee_data = db_session.query(EmployeeData).filter_by(
                user_id=emp.id
            ).first()
            
            employees_data.append({
                'id': emp.id,
                'name': emp.name,
                'username': emp.username,
                'department_id': emp.department_id,
                'job_title': employee_data.job_title if employee_data else 'موظف',
                'display_text': f"{emp.name} ({emp.username})"
            })
        
        return jsonify({
            'success': True,
            'employees': employees_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'خطأ: {str(e)}'
        })



@app.context_processor
def inject_manager_context():
    """حقن سياق المدير في جميع القوالب"""
    if current_user.is_authenticated and current_user.is_manager:
        # الحصول على قسم المدير
        managed_departments = db_session.query(Department).filter_by(
            primary_manager_id=current_user.id
        ).all()
        
        if managed_departments:
            # الحصول على إحصائيات
            department = managed_departments[0]
            employee_count = db_session.query(User).filter(
                User.department_id == department.id,
                User.is_admin == False
            ).count()
            
            today = date.today()
            today_permissions = db_session.query(PermissionRequest).filter(
                PermissionRequest.department_id == department.id,
                PermissionRequest.date == today,
                PermissionRequest.status == 'approved'
            ).count()
            
            return {
                'managed_department': department,
                'department_employee_count': employee_count,
                'today_permissions': today_permissions
            }
    
    return {}

def get_financial_period_for_date(check_date):
    """
    الحصول على بداية ونهاية الفترة المالية (الشهر المالي) لتاريخ معين
    
    القاعدة:
    - يناير: 26 ديسمبر السابق إلى 25 يناير الحالي
    - فبراير: 26 يناير إلى 25 فبراير
    - مارس: 26 فبراير إلى 25 مارس
    - ... وهكذا
    """
    # حساب الشهر المالي بناءً على تاريخ الإجازة
    if check_date.day >= 26:
        # يقع في الشهر المالي الذي يبدأ من يوم 26 من الشهر الحالي
        period_start = date(check_date.year, check_date.month, 26)
        if check_date.month == 12:
            period_end = date(check_date.year + 1, 1, 25)
        else:
            period_end = date(check_date.year, check_date.month + 1, 25)
    else:
        # يقع في الشهر المالي الذي يبدأ من يوم 26 من الشهر السابق
        if check_date.month == 1:
            period_start = date(check_date.year - 1, 12, 26)
            period_end = date(check_date.year, 1, 25)
        else:
            period_start = date(check_date.year, check_date.month - 1, 26)
            period_end = date(check_date.year, check_date.month, 25)
    
    return period_start, period_end


@app.route('/manager/reject_leave/<int:request_id>', methods=['POST'])
@login_required
def reject_leave(request_id):
    if not current_user.is_manager and not current_user.is_admin:
        flash('غير مصرح لك برفض هذا الطلب', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    leave_request = db_session.query(LeaveRequest).get(request_id)
    if not leave_request:
        flash('طلب الإجازة غير موجود', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # الحصول على تاريخ الإجازة الصحيح
    # استخدم leave_date إذا كان موجوداً، وإلا استخدم start_date
    leave_date = leave_request.leave_date if hasattr(leave_request, 'leave_date') and leave_request.leave_date else leave_request.start_date
    
    # التحقق من التاريخ (لا يمكن الرفض إلا في نفس الفترة الزمنية)
    if not can_reject_leave_by_date(leave_date):
        # الحصول على معلومات الفترة الزمنية للعرض
        leave_period_start, leave_period_end = get_financial_period_for_date(leave_date)
        today_period_start, today_period_end = get_financial_period_for_date(date.today())
        
        # جلب اسم الشهر بالعربية
        leave_month_name = get_month_name_arabic(leave_date.month)
        today_month_name = get_month_name_arabic(date.today().month)
        
        flash(
            f'لا يمكن رفض الإجازة لأنها في نطاق {leave_month_name} '
            f'({leave_period_start.strftime("%Y-%m-%d")} إلى {leave_period_end.strftime("%Y-%m-%d")}) '
            f'واليوم في نطاق {today_month_name} '
            f'({today_period_start.strftime("%Y-%m-%d")} إلى {today_period_end.strftime("%Y-%m-%d")})',
            'error'
        )
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.get(Department, leave_request.department_id)
    if department.primary_manager_id != current_user.id and not current_user.is_admin:
        flash('غير مصرح لك برفض هذا الطلب', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن الطلب معلق
    if leave_request.status != 'pending':
        flash('تم معالجة هذا الطلب مسبقاً', 'warning')
        return redirect(url_for('manager_leave_requests'))
    
    rejection_reason = request.form.get('rejection_reason', '').strip()
    
    if not rejection_reason:
        flash('يرجى كتابة سبب الرفض', 'error')
        return redirect(url_for('manager_leave_requests'))
    
    try:
        # رفض الشيفت الفردي
        leave_request.status = 'rejected'
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        leave_request.rejection_reason = rejection_reason
        
        # زيادة الرصيد إذا كانت الإجازة من رصيد الإجازات
        balance_increased = False
        new_balance = None
        
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(
                user_id=leave_request.user_id
            ).first()
            
            if balance:
                old_balance = balance.leave_balance
                balance.leave_balance += 1  # إرجاع يوم واحد
                balance.last_updated = datetime.now()
                new_balance = balance.leave_balance
                balance_increased = True
        
        db_session.commit()
        
        # إرسال إشعار للموظف
        notification_message = f'تم رفض شيفت إجازتك ({leave_request.shift_name}). السبب: {rejection_reason}'
        if balance_increased:
            notification_message += f'\nتم زيادة رصيدك بمقدار 1 يوم (الرصيد الجديد: {new_balance} يوم)'
        
        create_notification(
            leave_request.user_id,
            'رفض طلب الإجازة',
            notification_message,
            'leave_rejected',
            related_id=leave_request.id,
            action_url=url_for('user_leave_requests')
        )
        
        success_message = 'تم رفض الشيفت بنجاح'
        if balance_increased:
            success_message += ' وتم زيادة رصيد الموظف'
        
        flash(success_message, 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء رفض الطلب: {str(e)}', 'error')
    
    return redirect(url_for('manager_leave_requests'))


@app.context_processor
def inject_rejection_period():
    """تمرير معلومات الفترة الزمنية الحالية للرفض لجميع القوالب"""
    if current_user.is_authenticated and (current_user.is_manager or current_user.is_admin):
        period_start, period_end = get_financial_period_for_date(date.today())
        return {
            'rejection_period_start': period_start,
            'rejection_period_end': period_end,
            'can_reject_leave_by_date': can_reject_leave_by_date,
            'get_financial_period_for_date': get_financial_period_for_date
        }
    return {}



def get_financial_month_range(check_date=None):
    """الحصول على نطاق الشهر المالي (26 الشهر السابق إلى 25 الشهر الحالي)"""
    if not check_date:
        check_date = date.today()
    
    if check_date.day >= 26:
        # الفترة: 26 من الشهر الحالي إلى 25 من الشهر التالي
        start_date = date(check_date.year, check_date.month, 26)
        if check_date.month == 12:
            end_date = date(check_date.year + 1, 1, 25)
        else:
            end_date = date(check_date.year, check_date.month + 1, 25)
    else:
        # الفترة: 26 من الشهر السابق إلى 25 من الشهر الحالي
        if check_date.month == 1:
            start_date = date(check_date.year - 1, 12, 26)
            end_date = date(check_date.year, 1, 25)
        else:
            start_date = date(check_date.year, check_date.month - 1, 26)
            end_date = date(check_date.year, check_date.month, 25)
    
    return start_date, end_date

@app.context_processor
def utility_processor():
    def get_month_name_arabic_filter(month_num):
        months = {
            1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
            5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
            9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
        }
        return months.get(month_num, f'شهر {month_num}')
    
    return dict(get_month_name_arabic=get_month_name_arabic_filter)


def get_month_name_arabic(month_num):
    """Get Arabic month name - make sure this function exists"""
    months = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    return months.get(month_num, '')

@app.template_filter('get_month_name_arabic')
def get_month_name_arabic_filter(month_num):
    """فلتر للقالب للحصول على اسم الشهر بالعربية"""
    months = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    return months.get(month_num, f'شهر {month_num}')

@app.route('/manager/approve_leave_group/<int:parent_request_id>')
@login_required
def approve_leave_group(parent_request_id):
    """موافقة على مجموعة كاملة من الشيفتات"""
    if not current_user.is_manager and not current_user.is_admin:
        flash('غير مصرح لك بالموافقة على هذا الطلب')
        return redirect(url_for('manager_leave_requests'))
    
    # الحصول على جميع الشيفتات المرتبطة بهذا الطلب
    leave_requests = db_session.query(LeaveRequest).filter_by(
        parent_request_id=parent_request_id,
        status='pending'
    ).all()
    
    if not leave_requests:
        flash('لم يتم العثور على طلبات إجازة معلقة', 'warning')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_requests[0].department_id)
    if department.primary_manager_id != current_user.id and not current_user.is_admin:
        flash('غير مصرح لك بالموافقة على هذا الطلب')
        return redirect(url_for('manager_leave_requests'))
    
    try:
        approved_count = 0
        user_id = leave_requests[0].user_id
        
        for request in leave_requests:
            if request.status == 'pending':
                request.status = 'approved'
                request.approved_by = current_user.id
                request.approved_at = datetime.now()
                approved_count += 1
        
        # خصم من الرصيد فقط إذا كانت من رصيد الإجازات
        if leave_requests[0].leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(user_id=user_id).first()
            if balance:
                balance.leave_balance -= approved_count
                balance.last_updated = datetime.now()
        
        db_session.commit()
        
        # إرسال إشعار للموظف
        create_notification(
            user_id,
            'موافقة على طلب الإجازة',
            f'تم الموافقة على جميع شيفتات طلب إجازتك البالغ عددها {approved_count} شيفت',
            'leave_approved',
            related_id=parent_request_id,
            action_url=url_for('user_leave_requests')
        )
        
        flash(f'تم الموافقة على {approved_count} شيفت بنجاح', 'success')
        
    except Exception as e:
        db_session.rollback()
        flash(f'حدث خطأ أثناء الموافقة على الطلب: {str(e)}', 'error')
    
    return redirect(url_for('manager_leave_requests'))


@app.route('/manager/approve_single_shift/<int:request_id>')
@login_required
def approve_single_shift(request_id):
    """موافقة على شيفت فردي"""
    if not current_user.is_manager and not current_user.is_admin:
        flash('غير مصرح لك بالموافقة على هذا الطلب')
        return redirect(url_for('manager_leave_requests'))
    
    leave_request = db_session.query(LeaveRequest).get(request_id)
    
    if not leave_request:
        flash('طلب الإجازة غير موجود')
        return redirect(url_for('manager_leave_requests'))
    
    # التحقق من أن المدير مسؤول عن القسم
    department = db_session.query(Department).get(leave_request.department_id)
    if department.primary_manager_id != current_user.id and not current_user.is_admin:
        flash('غير مصرح لك بالموافقة على هذا الطلب')
        return redirect(url_for('manager_leave_requests'))
    
    if leave_request.status == 'pending':
        leave_request.status = 'approved'
        leave_request.approved_by = current_user.id
        leave_request.approved_at = datetime.now()
        
        # خصم من الرصيد فقط إذا كانت من رصيد الإجازات
        if leave_request.leave_type == 'من رصيد الإجازات':
            balance = db_session.query(EmployeeBalance).filter_by(user_id=leave_request.user_id).first()
            if balance:
                balance.leave_balance -= 1  # شيفت واحد = يوم واحد
                balance.last_updated = datetime.now()
        
        db_session.commit()
        
        flash('تم الموافقة على الشيفت بنجاح')
    
    return redirect(url_for('manager_leave_requests'))


@app.template_filter('parse_shift_info')
def parse_shift_info_filter(shift_info_json):
    """فلتر للقالب لتحليل معلومات الشيفتات"""
    if not shift_info_json:
        return []
    try:
        shift_data = json.loads(shift_info_json)
        return shift_data.get('shifts', [])
    except:
        return []

@app.template_filter('get_shift_jobs')
def get_shift_jobs_filter(shift_info_json):
    """فلتر للقالب للحصول على الوظائف المحددة لكل شيفت"""
    if not shift_info_json:
        return {}
    try:
        shift_data = json.loads(shift_info_json)
        return shift_data.get('jobs', {})
    except:
        return {}



@app.route('/manager/permission_requests')
@login_required
def manager_permission_requests():
    """طلبات الإذونات مع الفلاتر"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    managed_departments = db_session.query(Department).filter_by(primary_manager_id=current_user.id).all()
    department_ids = [dept.id for dept in managed_departments]
    
    # الحصول على معاملات التصفية - استخدام request من Flask
    filter_status = request.args.get('status', 'all')
    filter_type = request.args.get('type', 'all')
    filter_month = request.args.get('month', 'all')
    filter_employee = request.args.get('employee', 'all')  # فلتر الموظف الجديد
    
    # بناء الاستعلام الأساسي
    query = db_session.query(PermissionRequest).filter(
        PermissionRequest.department_id.in_(department_ids)
    )
    
    # تطبيق الفلاتر
    if filter_status != 'all':
        query = query.filter(PermissionRequest.status == filter_status)
    
    if filter_type != 'all':
        query = query.filter(PermissionRequest.permission_type == filter_type)
    
    # تطبيق فلتر الشهر فقط (بدون السنة)
    if filter_month != 'all' and filter_month.isdigit():
        month_num = int(filter_month)
        # تصفية حسب الشهر فقط (أي سنة)
        query = query.filter(extract('month', PermissionRequest.date) == month_num)
    
    # تطبيق فلتر الموظف
    if filter_employee != 'all' and filter_employee.isdigit():
        employee_id = int(filter_employee)
        query = query.filter(PermissionRequest.user_id == employee_id)
    
    permission_requests = query.order_by(PermissionRequest.created_at.desc()).all()
    
    # الحصول على جميع المستخدمين للفلتر
    all_users = db_session.query(User).filter(
        User.id.in_(
            db_session.query(PermissionRequest.user_id).filter(
                PermissionRequest.department_id.in_(department_ids)
            ).distinct()
        )
    ).order_by(User.name).all()
    
    # فصل الطلبات إلى معلقة وسابقة
    pending_requests = []
    approved_requests = []
    rejected_requests = []
    history_requests = []
    
    for req in permission_requests:
        # تحميل البيانات ذات الصلة
        req.user = db_session.query(User).get(req.user_id)
        req.department = db_session.query(Department).get(req.department_id)
        
        # تحميل بيانات المعالجة
        if req.approved_by:
            req.approved_by = db_session.query(User).get(req.approved_by)
        
        # تحليل extra_data يدوياً لتخزينه في متغير مؤقت
        # هذا لأن extra_data_dict هو property للقراءة فقط
        if req.extra_data:
            try:
                # تحميل JSON إلى متغير مؤقت
                req._parsed_extra_data = json.loads(req.extra_data)
            except Exception as e:
                req._parsed_extra_data = {}
        else:
            req._parsed_extra_data = {}
        
        # فصل الطلبات
        if req.status == 'pending':
            pending_requests.append(req)
        elif req.status == 'approved':
            approved_requests.append(req)
            history_requests.append(req)
        elif req.status == 'rejected':
            rejected_requests.append(req)
            history_requests.append(req)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager/manager_permission_requests.html',
                         permission_requests=permission_requests,
                         pending_requests=pending_requests,
                         approved_requests=approved_requests,
                         rejected_requests=rejected_requests,
                         history_requests=history_requests,
                         all_users=all_users,  # جميع المستخدمين للفلتر
                         notifications=notifications,
                         current_status=filter_status,
                         current_type=filter_type,
                         current_month=filter_month,
                         current_user_id=filter_employee,  # المستخدم المحدد
                         stats={  # الإحصائيات القديمة لا تزال متاحة
                             'total': len(permission_requests),
                             'pending': len(pending_requests),
                             'approved': len(approved_requests),
                             'rejected': len(rejected_requests),
                         })






@app.route('/manager/advance_requests')
@login_required
def manager_advance_requests():
    """طلبات السلف"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    managed_departments = db_session.query(Department).filter_by(primary_manager_id=current_user.id).all()
    department_ids = [dept.id for dept in managed_departments]
    
    advance_requests = db_session.query(AdvanceRequest).filter(
        AdvanceRequest.department_id.in_(department_ids)
    ).order_by(AdvanceRequest.created_at.desc()).all()
    
    # Load related data
    for request in advance_requests:
        request.user = db_session.query(User).get(request.user_id)
        request.department = db_session.query(Department).get(request.department_id)
    
    notifications = get_user_notifications(current_user.id)
    
    return render_template('manager/manager_advance_requests.html',
                         advance_requests=advance_requests,
                         notifications=notifications)



def is_within_rejection_period(submit_date, reject_date=None):
    """
    التحقق مما إذا كان تاريخ الرفض يقع في نفس الفترة الزمنية لتقديم الطلب
    الفترات: من 26 من الشهر إلى 25 من الشهر التالي
    """
    if not reject_date:
        reject_date = date.today()
    
    # تحويل التواريخ إذا كانت نصوصاً
    if isinstance(submit_date, str):
        submit_date = datetime.strptime(submit_date, '%Y-%m-%d').date()
    if isinstance(reject_date, str):
        reject_date = datetime.strptime(reject_date, '%Y-%m-%d').date()
    
    # تحديد الفترة الزمنية لتاريخ تقديم الطلب
    if submit_date.day >= 26:
        # الفترة: من 26 من الشهر الحالي إلى 25 من الشهر التالي
        period_start = date(submit_date.year, submit_date.month, 26)
        if submit_date.month == 12:
            period_end = date(submit_date.year + 1, 1, 25)
        else:
            period_end = date(submit_date.year, submit_date.month + 1, 25)
    else:
        # الفترة: من 26 من الشهر السابق إلى 25 من الشهر الحالي
        if submit_date.month == 1:
            period_start = date(submit_date.year - 1, 12, 26)
            period_end = date(submit_date.year, 1, 25)
        else:
            period_start = date(submit_date.year, submit_date.month - 1, 26)
            period_end = date(submit_date.year, submit_date.month, 25)
    
    # التحقق مما إذا كان تاريخ الرفض يقع ضمن هذه الفترة
    return period_start <= reject_date <= period_end


def get_rejection_period_dates(check_date=None):
    """
    الحصول على تواريخ بداية ونهاية الفترة الزمنية للرفض
    """
    if not check_date:
        check_date = date.today()
    
    if check_date.day >= 26:
        # الفترة: من 26 من الشهر الحالي إلى 25 من الشهر التالي
        period_start = date(check_date.year, check_date.month, 26)
        if check_date.month == 12:
            period_end = date(check_date.year + 1, 1, 25)
        else:
            period_end = date(check_date.year, check_date.month + 1, 25)
    else:
        # الفترة: من 26 من الشهر السابق إلى 25 من الشهر الحالي
        if check_date.month == 1:
            period_start = date(check_date.year - 1, 12, 26)
            period_end = date(check_date.year, 1, 25)
        else:
            period_start = date(check_date.year, check_date.month - 1, 26)
            period_end = date(check_date.year, check_date.month, 25)
    
    return period_start, period_end


def can_reject_leave_by_financial_month(leave_date):
    """
    التحقق من رفض الإجازة بناءً على نطاق الشهر المالي
    القاعدة: لا يمكن رفض الإجازة بعد الوصول إلى يوم 26 من نفس الشهر المالي
    """
    if not leave_date:
        return True
    
    # تحويل التاريخ إذا كان نصياً
    if isinstance(leave_date, str):
        try:
            leave_date = datetime.strptime(leave_date, '%Y-%m-%d').date()
        except:
            return True
    
    today = date.today()
    
    # إذا كان تاريخ الإجازة في الماضي، لا يمكن رفضه
    if leave_date < today:
        return False
    
    # تحديد الشهر المالي لتاريخ الإجازة
    # الشهر المالي يبدأ من 26 الشهر السابق إلى 25 الشهر الحالي
    if leave_date.day >= 26:
        # يقع في الشهر المالي التالي
        financial_month_26th = date(leave_date.year, leave_date.month, 26)
    else:
        # يقع في الشهر المالي الحالي
        if leave_date.month == 1:
            financial_month_26th = date(leave_date.year - 1, 12, 26)
        else:
            financial_month_26th = date(leave_date.year, leave_date.month - 1, 26)
    
    # إذا وصلنا إلى يوم 26 من الشهر المالي، لا يمكن الرفض
    if today >= financial_month_26th:
        return False
    
    return True



@app.route('/manager/approve_advance/<int:request_id>')
@login_required
def manager_approve_advance(request_id):
    """موافقة المدير على طلب سلفة"""
    if not current_user.is_manager and not current_user.is_admin:
        return redirect(url_for('user_dashboard'))
    
    advance_request = db_session.query(AdvanceRequest).get(request_id)
    
    # Check if manager manages this department
    department = db_session.query(Department).get(advance_request.department_id)
    if department.manager_id != current_user.id and not current_user.is_admin:
        flash('غير مصرح لك بالموافقة على هذا الطلب')
        return redirect(url_for('manager_advance_requests'))
    
    if advance_request.status == 'pending':
        advance_request.status = 'approved'
        advance_request.approved_by = current_user.id
        advance_request.approved_at = datetime.now()
        
        # Update balance
        balance = db_session.query(EmployeeBalance).filter_by(user_id=advance_request.user_id).first()
        if balance:
            balance.advance_balance += advance_request.amount
            balance.last_updated = datetime.now()
        
        # Notify employee
        create_notification(
            advance_request.user_id,
            'موافقة على طلب السلفة',
            f'تم الموافقة على طلب سلفتك بقيمة {advance_request.amount} جنيه',
            'advance_approved',
            action_url=url_for('user_advance_requests')
        )
        
        db_session.commit()
        flash('تم الموافقة على طلب السلفة')
    
    return redirect(url_for('manager_advance_requests'))

@app.route('/admin/reset_future_schedules')
@login_required
def admin_reset_future_schedules():
    """حذف جميع الجداول المستقبلية وإعادة إنشائها بنفس هيكل القسم"""

    try:
        from database import db_session
        from models import WeeklySchedule, Department
        
        today = date.today()
        
        # 1. حساب تاريخ بداية الأسبوع الحالي (السبت)
        days_since_saturday = (today.weekday() - 5) % 7
        current_week_start = today - timedelta(days=days_since_saturday)
        
        print(f"=== بدء إعادة تعيين الجداول المستقبلية ===")
        print(f"تاريخ اليوم: {today}")
        print(f"بداية الأسبوع الحالي: {current_week_start}")
        
        # 2. حذف جميع الجداول من الأسبوع الحالي فصاعداً
        future_schedules = db_session.query(WeeklySchedule).filter(
            WeeklySchedule.week_start_date >= current_week_start
        ).all()
        
        deleted_count = len(future_schedules)
        
        for schedule in future_schedules:
            db_session.delete(schedule)
            print(f"تم حذف الجدول {schedule.id} للقسم {schedule.department_id}")
        
        db_session.commit()
        
        # 3. إنشاء جداول جديدة بنفس هيكل كل قسم
        created_count = 0
        
        # الحصول على جميع الأقسام المفعلة للتوليد التلقائي
        auto_departments = db_session.query(Department).filter_by(auto_generate_schedule=True).all()
        
        for department in auto_departments:
            print(f"معالجة القسم: {department.name} (ID: {department.id})")
            
            # إنشاء جداول للأسبوع الحالي والأسابيع الثلاثة القادمة
            for week_offset in range(4):  # 4 أسابيع (0 للأسبوع الحالي، 1-3 للأسابيع القادمة)
                week_start_date = current_week_start + timedelta(days=7 * week_offset)
                week_end_date = week_start_date + timedelta(days=6)
                
                # إنشاء جدول جديد بنفس هيكل القسم
                schedule_data = create_weekly_schedule_from_structure(department.id, week_start_date)
                
                if schedule_data:
                    new_schedule = WeeklySchedule(
                        department_id=department.id,
                        week_start_date=week_start_date,
                        week_end_date=week_end_date,
                        schedule_data=json.dumps(schedule_data, ensure_ascii=False),
                        created_by=current_user.id,
                        status='draft',
                        is_locked=False,
                        is_approved=False,
                        is_generated_from_structure=True
                    )
                    
                    db_session.add(new_schedule)
                    created_count += 1
                    print(f"✓ تم إنشاء جدول جديد للقسم {department.name} للفترة {week_start_date}")
        
        db_session.commit()
        
        result_html = f"""
        <!DOCTYPE html>
        <html dir="rtl">
        <head>
            <meta charset="UTF-8">
            <title>نتيجة إعادة التعيين</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; background-color: #f5f5f5; }}
                .container {{ max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .success {{ color: #28a745; font-weight: bold; }}
                .error {{ color: #dc3545; font-weight: bold; }}
                .btn {{ display: inline-block; padding: 10px 20px; background: #007bff; color: white; text-decoration: none; border-radius: 5px; margin-top: 20px; }}
                .btn:hover {{ background: #0056b3; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>نتيجة إعادة تعيين الجداول المستقبلية</h2>
                <hr>
                <h3 class="success">✓ تمت العملية بنجاح</h3>
                <p><strong>عدد الجداول المحذوفة:</strong> {deleted_count} جدول</p>
                <p><strong>عدد الجداول المنشأة:</strong> {created_count} جدول جديد</p>
                <p><strong>عدد الأقسام المعالجة:</strong> {len(auto_departments)} قسم</p>
                
                <h4>تفاصيل العملية:</h4>
                <ul>
                    <li>تم حذف جميع الجداول من تاريخ {current_week_start} فصاعداً</li>
                    <li>تم إنشاء جداول جديدة بنفس هيكل كل قسم</li>
                    <li>تم إنشاء جداول للأسبوع الحالي و 3 أسابيع قادمة</li>
                    <li>تم تعيين جميع الجداول الجديدة كـ "مسودة" (draft)</li>
                </ul>
                
                <a href="/admin/schedules" class="btn">العودة إلى صفحة الجداول</a>
                <a href="/admin/dashboard" class="btn" style="background: #6c757d; margin-right: 10px;">العودة للوحة التحكم</a>
            </div>
        </body>
        </html>
        """
        
        return result_html
        
    except Exception as e:
        db_session.rollback()
        error_html = f"""
        <!DOCTYPE html>
        <html dir="rtl">
        <head>
            <meta charset="UTF-8">
            <title>خطأ في إعادة التعيين</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; background-color: #f5f5f5; }}
                .container {{ max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .error {{ color: #dc3545; font-weight: bold; }}
                .btn {{ display: inline-block; padding: 10px 20px; background: #6c757d; color: white; text-decoration: none; border-radius: 5px; margin-top: 20px; }}
                .btn:hover {{ background: #5a6268; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2 class="error">❌ حدث خطأ في إعادة تعيين الجداول</h2>
                <hr>
                <p><strong>تفاصيل الخطأ:</strong> {str(e)}</p>
                <p>تم التراجع عن جميع التغييرات</p>
                <a href="/admin/schedules" class="btn">العودة إلى صفحة الجداول</a>
            </div>
        </body>
        </html>
        """
        
        return error_html


if __name__ == '__main__':
    init_db()
    create_default_admin()
    debug_sync_process()
    app.run(host='0.0.0.0', port=5551, debug=True)